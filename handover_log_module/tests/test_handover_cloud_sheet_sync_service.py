from __future__ import annotations

from copy import deepcopy
from pathlib import Path

import openpyxl

from handover_log_module.service.handover_cloud_sheet_sync_service import HandoverCloudSheetSyncService


def _merge(start_row: int, end_row: int, start_col: int, end_col: int) -> dict[str, int]:
    return {
        "start_row_index": start_row,
        "end_row_index": end_row,
        "start_column_index": start_col,
        "end_column_index": end_col,
    }


class FakeSheetsClient:
    def __init__(self) -> None:
        self.sheets = [
            {
                "sheet_id": "sheet_a_keep",
                "title": "A楼",
                "index": 0,
                "row_count": 30,
                "column_count": 12,
                "merges": [
                    _merge(2, 3, 0, 2),
                    _merge(22, 23, 6, 8),
                ],
            },
            {
                "sheet_id": "sheet_a_dup",
                "title": "A楼",
                "index": 3,
                "row_count": 30,
                "column_count": 12,
                "merges": [],
            },
            {
                "sheet_id": "sheet_b_keep",
                "title": "B楼",
                "index": 1,
                "row_count": 30,
                "column_count": 12,
                "merges": [],
            },
        ]
        self.sheet_counter = 10
        self.deleted: list[dict] = []
        self.added: list[dict] = []
        self.copied: list[dict] = []
        self.renamed: list[dict] = []
        self.clear_updates: list[dict] = []
        self.dimension_adds: list[dict] = []
        self.dimension_deletes: list[dict] = []
        self.dimension_updates: list[dict] = []
        self.value_updates: list[dict] = []
        self.merge_updates: list[dict] = []
        self.unmerge_updates: list[dict] = []
        self.query_calls: list[dict] = []
        self.fail_rename_titles: set[str] = set()

    def find_or_create_date_spreadsheet(self, **kwargs):
        return {
            "spreadsheet_token": "sheet_token_1",
            "title": kwargs.get("spreadsheet_title", ""),
            "url": "https://vnet.feishu.cn/wiki/wiki_token_1",
        }

    def query_sheets(self, spreadsheet_token: str, *, sheet_cache=None, force_refresh: bool = False):  # noqa: ARG002, ANN001
        self.query_calls.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "has_cache": isinstance(sheet_cache, dict),
                "force_refresh": force_refresh,
            }
        )
        return [deepcopy(item) for item in self.sheets]

    def _sort_sheets(self) -> None:
        self.sheets.sort(key=lambda item: (int(item.get("index", 0) or 0), str(item.get("title", ""))))

    def _next_sheet_id(self) -> str:
        self.sheet_counter += 1
        return f"sheet_{self.sheet_counter}"

    def _find_sheet(self, sheet_id: str) -> dict:
        for item in self.sheets:
            if item["sheet_id"] == sheet_id:
                return item
        raise RuntimeError(f"sheet not found: {sheet_id}")

    def delete_sheet(self, spreadsheet_token: str, sheet_id: str, *, sheet_cache=None) -> None:  # noqa: ARG002, ANN001
        self.deleted.append({"spreadsheet_token": spreadsheet_token, "sheet_id": sheet_id})
        self.sheets = [item for item in self.sheets if item["sheet_id"] != sheet_id]

    def add_sheet(self, spreadsheet_token: str, title: str, index: int = 0, *, sheet_cache=None):  # noqa: ARG002, ANN001
        new_sheet = {
            "sheet_id": self._next_sheet_id(),
            "title": title,
            "index": index,
            "row_count": 20,
            "column_count": 20,
            "merges": [],
        }
        self.added.append({"spreadsheet_token": spreadsheet_token, "title": title, "index": index})
        self.sheets.append(new_sheet)
        self._sort_sheets()
        return deepcopy(new_sheet)

    def copy_sheet(self, spreadsheet_token: str, *, source_sheet_id: str, title: str, sheet_cache=None):  # noqa: ARG002, ANN001
        source = deepcopy(self._find_sheet(source_sheet_id))
        copied = {
            **source,
            "sheet_id": self._next_sheet_id(),
            "title": title,
            "index": int(source.get("index", 0) or 0) + 1,
            "merges": [dict(item) for item in source.get("merges", [])],
        }
        self.copied.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "source_sheet_id": source_sheet_id,
                "title": title,
            }
        )
        self.sheets.append(copied)
        self._sort_sheets()
        return deepcopy(copied)

    def rename_sheet(self, spreadsheet_token: str, *, sheet_id: str, title: str, sheet_cache=None):  # noqa: ARG002, ANN001
        target = self._find_sheet(sheet_id)
        target["title"] = title
        self.renamed.append({"spreadsheet_token": spreadsheet_token, "sheet_id": sheet_id, "title": title})
        self._sort_sheets()
        return deepcopy(target)

    def rename_and_move_sheet(  # noqa: ANN001
        self,
        spreadsheet_token: str,
        *,
        sheet_id: str,
        title=None,
        index=None,
        sheet_cache=None,
    ):
        if title in self.fail_rename_titles:
            raise RuntimeError(f"rename_failed: {title}")
        target = self._find_sheet(sheet_id)
        if title is not None:
            target["title"] = title
        if index is not None:
            target["index"] = int(index or 0)
        self.renamed.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "title": target["title"],
                "index": target["index"],
            }
        )
        self._sort_sheets()
        return deepcopy(target)

    def get_or_create_named_sheet(self, spreadsheet_token: str, title: str, index: int = 0, *, sheet_cache=None):  # noqa: ANN001
        matched = [item for item in self.sheets if item["title"] == title]
        if matched:
            matched.sort(key=lambda item: int(item.get("index", 0) or 0))
            return deepcopy(matched[0])
        return self.add_sheet(spreadsheet_token, title, index=index, sheet_cache=sheet_cache)

    def dedupe_named_sheets(self, spreadsheet_token: str, title: str, *, sheet_cache=None):  # noqa: ANN001
        matched = [item for item in self.sheets if item["title"] == title]
        if not matched:
            return {}
        matched.sort(key=lambda item: int(item.get("index", 0) or 0))
        keep = matched[0]
        for duplicate in matched[1:]:
            self.delete_sheet(spreadsheet_token, duplicate["sheet_id"], sheet_cache=sheet_cache)
        return deepcopy(keep)

    def batch_clear_values(self, spreadsheet_token: str, range_name: str, rows: int, cols: int):
        self.clear_updates.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "range_name": range_name,
                "rows": rows,
                "cols": cols,
            }
        )
        return {}

    def batch_update_values(self, spreadsheet_token: str, value_ranges: list[dict]):
        self.value_updates.append({"spreadsheet_token": spreadsheet_token, "value_ranges": value_ranges})
        return {}

    def add_dimension(self, spreadsheet_token: str, *, sheet_id: str, major_dimension: str, length: int):
        self.dimension_adds.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "major_dimension": major_dimension,
                "length": length,
            }
        )
        target = self._find_sheet(sheet_id)
        if str(major_dimension).upper() == "ROWS":
            target["row_count"] = int(target.get("row_count", 0) or 0) + int(length or 0)
        else:
            target["column_count"] = int(target.get("column_count", 0) or 0) + int(length or 0)
        return {"addCount": length, "majorDimension": major_dimension}

    def delete_dimension(self, spreadsheet_token: str, *, sheet_id: str, major_dimension: str, start_index: int, end_index: int):
        self.dimension_deletes.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "major_dimension": major_dimension,
                "start_index": start_index,
                "end_index": end_index,
            }
        )
        target = self._find_sheet(sheet_id)
        length = max(0, int(end_index or 0) - int(start_index or 0))
        if str(major_dimension).upper() == "ROWS":
            target["row_count"] = max(1, int(target.get("row_count", 0) or 0) - length)
        else:
            target["column_count"] = max(1, int(target.get("column_count", 0) or 0) - length)
        return {"delCount": length, "majorDimension": major_dimension}

    def update_dimension_range(
        self,
        spreadsheet_token: str,
        *,
        sheet_id: str,
        major_dimension: str,
        start_index: int,
        end_index: int,
        pixel_size: int,
    ):
        self.dimension_updates.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "major_dimension": major_dimension,
                "start_index": start_index,
                "end_index": end_index,
                "pixel_size": pixel_size,
            }
        )
        return {}

    def batch_merge_cells(self, spreadsheet_token: str, sheet_id: str, merges: list[dict]):
        self.merge_updates.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "merges": deepcopy(merges),
            }
        )
        target = self._find_sheet(sheet_id)
        existing = {
            (
                int(item["start_row_index"]),
                int(item["end_row_index"]),
                int(item["start_column_index"]),
                int(item["end_column_index"]),
            ): dict(item)
            for item in target.get("merges", [])
        }
        for item in merges:
            key = (
                int(item["start_row_index"]),
                int(item["end_row_index"]),
                int(item["start_column_index"]),
                int(item["end_column_index"]),
            )
            existing[key] = dict(item)
        target["merges"] = list(existing.values())
        return {}

    def batch_unmerge_cells(self, spreadsheet_token: str, sheet_id: str, merges: list[dict]):
        self.unmerge_updates.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "sheet_id": sheet_id,
                "merges": deepcopy(merges),
            }
        )
        remove_keys = {
            (
                int(item["start_row_index"]),
                int(item["end_row_index"]),
                int(item["start_column_index"]),
                int(item["end_column_index"]),
            )
            for item in merges
        }
        target = self._find_sheet(sheet_id)
        target["merges"] = [
            dict(item)
            for item in target.get("merges", [])
            if (
                int(item["start_row_index"]),
                int(item["end_row_index"]),
                int(item["start_column_index"]),
                int(item["end_column_index"]),
            )
            not in remove_keys
        ]
        return {}


def create_workbook(path: Path, *, add_cross_boundary_merge: bool = False) -> None:
    workbook = openpyxl.Workbook()
    cover = workbook.active
    cover.title = "封面"
    cover["A1"] = "cover"

    sheet = workbook.create_sheet("交接班日志")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 24
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[23].height = 30
    sheet["A1"] = "标题"
    sheet["F2"] = "白班"
    sheet["C3"] = "张三/李四"
    sheet["A3"] = "固定合并标题"
    sheet.merge_cells("A3:B3")
    sheet["B18"] = '=IF(F2="白班","9:00","21:00")'
    sheet["B19"] = '=IF(F2="白班","16:00","4:00")'
    sheet["H18"] = "=C3"
    sheet["H19"] = "=C3"
    sheet["A23"] = "动态合并标题"
    sheet["C23"] = "动态右侧值"
    sheet.merge_cells("A23:B23")
    if add_cross_boundary_merge:
        sheet.merge_cells("A21:B22")
    workbook.save(path)
    workbook.close()


def build_service(monkeypatch, fake_client: FakeSheetsClient) -> HandoverCloudSheetSyncService:
    service = HandoverCloudSheetSyncService(
        {
            "_global_feishu": {"app_id": "app_id", "app_secret": "app_secret"},
            "cloud_sheet_sync": {
                "enabled": True,
                "source_sheet_name": "交接班日志",
                "sheet_names": {
                    "A楼": "A楼",
                    "B楼": "B楼",
                    "C楼": "C楼",
                    "D楼": "D楼",
                    "E楼": "E楼",
                },
            },
        }
    )
    monkeypatch.setattr(service, "_build_client", lambda: fake_client)  # noqa: SLF001
    return service


def test_select_source_sheet_requires_exact_handover_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "handover.xlsx"
    create_workbook(file_path)
    workbook = openpyxl.load_workbook(file_path)
    try:
        service = HandoverCloudSheetSyncService({"cloud_sheet_sync": {"source_sheet_name": "交接班日志"}})
        worksheet = service._select_source_sheet(workbook)  # noqa: SLF001
    finally:
        workbook.close()

    assert worksheet.title == "交接班日志"


def test_prepare_batch_spreadsheet_creates_missing_target_sheets_and_dedupes_titles(monkeypatch) -> None:
    fake_client = FakeSheetsClient()
    service = build_service(monkeypatch, fake_client)

    result = service.prepare_batch_spreadsheet(
        duty_date="2026-03-22",
        duty_date_text="3月22日",
        shift_text="夜班",
        emit_log=lambda *_args: None,
    )

    assert result["success"] is True
    assert result["status"] == "prepared"
    assert {item["title"] for item in fake_client.sheets} >= {"A楼", "B楼", "C楼", "D楼", "E楼"}
    assert fake_client.deleted == [{"spreadsheet_token": "sheet_token_1", "sheet_id": "sheet_a_dup"}]
    assert fake_client.value_updates == []


def test_collect_sheet_snapshot_splits_fixed_and_dynamic_merges_and_collects_dimensions(tmp_path: Path) -> None:
    file_path = tmp_path / "handover.xlsx"
    create_workbook(file_path)

    service = HandoverCloudSheetSyncService({"cloud_sheet_sync": {"source_sheet_name": "交接班日志"}})
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    try:
        worksheet = workbook["交接班日志"]
        snapshot = service._collect_sheet_snapshot(worksheet)  # noqa: SLF001
    finally:
        workbook.close()

    assert snapshot["values"][17][1] == "9:00"
    assert snapshot["values"][18][1] == "16:00"
    assert snapshot["values"][17][7] == "张三/李四"
    assert snapshot["values"][18][7] == "张三/李四"
    assert snapshot["fixed_header_merges"] == [_merge(2, 3, 0, 2)]
    assert snapshot["dynamic_merges"] == [_merge(22, 23, 0, 2)]
    assert snapshot["dynamic_merge_signature"]
    assert snapshot["row_heights"]
    assert snapshot["column_widths"]


def test_collect_sheet_snapshot_rejects_cross_boundary_merge(tmp_path: Path) -> None:
    file_path = tmp_path / "handover.xlsx"
    create_workbook(file_path, add_cross_boundary_merge=True)

    service = HandoverCloudSheetSyncService({"cloud_sheet_sync": {"source_sheet_name": "交接班日志"}})
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    try:
        worksheet = workbook["交接班日志"]
        try:
            service._collect_sheet_snapshot(worksheet)  # noqa: SLF001
        except RuntimeError as exc:
            error = str(exc)
        else:  # pragma: no cover - defensive
            error = ""
    finally:
        workbook.close()

    assert "unsupported_cross_boundary_merge" in error


def test_sync_confirmed_buildings_cleans_shadow_sheet_and_rebuilds_original_sheet(tmp_path: Path, monkeypatch) -> None:
    file_path = tmp_path / "handover.xlsx"
    create_workbook(file_path)
    fake_client = FakeSheetsClient()
    fake_client.sheets.append(
        {
            "sheet_id": "sheet_a_tmp_old",
            "title": "A楼__tmp__20260412151311",
            "index": 4,
            "row_count": 20,
            "column_count": 20,
            "merges": [],
        }
    )
    service = build_service(monkeypatch, fake_client)

    result = service.sync_confirmed_buildings(
        batch_meta={
            "spreadsheet_token": "sheet_token_1",
            "spreadsheet_url": "https://vnet.feishu.cn/wiki/wiki_token_1",
            "spreadsheet_title": "南通园区交接班日志-3月22日夜班",
        },
        building_items=[
            {
                "building": "A楼",
                "output_file": str(file_path),
                "revision": 3,
                "cloud_sheet_sync": {
                    "synced_row_count": 25,
                    "synced_column_count": 8,
                    "synced_merges": [_merge(22, 23, 6, 8)],
                    "dynamic_merge_signature": "legacy-signature",
                },
            }
        ],
        emit_log=lambda *_args: None,
    )

    assert result["status"] == "ok"
    assert fake_client.copied == []
    assert fake_client.renamed == []
    assert {"spreadsheet_token": "sheet_token_1", "sheet_id": "sheet_a_tmp_old"} in fake_client.deleted
    assert fake_client.unmerge_updates[0]["merges"] == [_merge(2, 3, 0, 2), _merge(22, 23, 6, 8)]
    assert fake_client.clear_updates
    assert fake_client.dimension_updates == []
    assert fake_client.merge_updates[0]["merges"] == [_merge(2, 3, 0, 2), _merge(22, 23, 0, 2)]
    assert fake_client.deleted[-1] != {"spreadsheet_token": "sheet_token_1", "sheet_id": "sheet_a_keep"}

    final_a = [item for item in fake_client.sheets if item["title"] == "A楼"]
    assert len(final_a) == 1
    assert final_a[0]["sheet_id"] == "sheet_a_keep"
    assert final_a[0]["merges"] == [_merge(2, 3, 0, 2), _merge(22, 23, 0, 2)]
    assert result["details"]["A楼"]["dynamic_merge_signature"]
    assert result["details"]["A楼"]["sheet_title"] == "A楼"


def test_sync_confirmed_buildings_overwrites_existing_sheet_in_place_without_copy_or_swap(tmp_path: Path, monkeypatch) -> None:
    file_path = tmp_path / "handover.xlsx"
    create_workbook(file_path)
    fake_client = FakeSheetsClient()
    service = build_service(monkeypatch, fake_client)

    result = service.sync_confirmed_buildings(
        batch_meta={
            "spreadsheet_token": "sheet_token_1",
            "spreadsheet_url": "https://vnet.feishu.cn/wiki/wiki_token_1",
            "spreadsheet_title": "南通园区交接班日志-3月22日夜班",
        },
        building_items=[
            {
                "building": "A楼",
                "output_file": str(file_path),
                "revision": 1,
                "cloud_sheet_sync": {},
            }
        ],
        emit_log=lambda *_args: None,
    )

    assert result["status"] == "ok"
    final_a = [item for item in fake_client.sheets if item["title"] == "A楼"]
    assert len(final_a) == 1
    assert final_a[0]["sheet_id"] == "sheet_a_keep"
    assert fake_client.copied == []
    assert fake_client.renamed == []
    assert result["details"]["A楼"]["status"] == "success"


def test_sync_confirmed_buildings_returns_failed_when_exact_handover_sheet_missing(tmp_path: Path, monkeypatch) -> None:
    file_path = tmp_path / "handover.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "封面"
    workbook.create_sheet("其他页面")
    workbook.save(file_path)
    workbook.close()

    service = build_service(monkeypatch, FakeSheetsClient())

    result = service.sync_confirmed_buildings(
        batch_meta={
            "spreadsheet_token": "sheet_token_1",
            "spreadsheet_url": "https://vnet.feishu.cn/wiki/wiki_token_1",
            "spreadsheet_title": "南通园区交接班日志-3月22日夜班",
        },
        building_items=[{"building": "A楼", "output_file": str(file_path), "revision": 1, "cloud_sheet_sync": {}}],
        emit_log=lambda *_args: None,
    )

    assert result["status"] == "failed"
    assert result["failed_buildings"][0]["building"] == "A楼"
    assert "missing_source_sheet: 交接班日志" in result["failed_buildings"][0]["error"]


def test_sync_confirmed_buildings_runs_serial_and_retries_rate_limit(monkeypatch) -> None:
    service = HandoverCloudSheetSyncService({"cloud_sheet_sync": {"enabled": True}})
    monkeypatch.setattr(service, "_build_base_sheet_cache", lambda **_kwargs: {})  # noqa: SLF001
    monkeypatch.setattr("handover_log_module.service.handover_cloud_sheet_sync_service.time.sleep", lambda _sec: None)
    call_order = []
    attempts_by_building = {}

    def _fake_sync_one_building(*, spreadsheet_token, item, emit_log, base_sheet_cache=None):  # noqa: ARG001
        building = str(item.get("building", "")).strip()
        call_order.append(building)
        attempts_by_building[building] = attempts_by_building.get(building, 0) + 1
        if building == "B" and attempts_by_building[building] == 1:
            return {
                "building": building,
                "success": False,
                "error": "飞书接口调用失败: {'code': 90217, 'msg': 'too many request'}",
                "detail": {
                    "status": "failed",
                    "sheet_title": building,
                    "synced_revision": 0,
                    "rows": 0,
                    "cols": 0,
                    "merged": 0,
                    "synced_row_count": 0,
                    "synced_column_count": 0,
                    "synced_merges": [],
                    "dynamic_merge_signature": "",
                    "error": "飞书接口调用失败: {'code': 90217, 'msg': 'too many request'}",
                },
            }
        return {
            "building": building,
            "success": True,
            "error": "",
            "detail": {
                "status": "success",
                "sheet_title": building,
                "synced_revision": int(item.get("revision", 0) or 0),
                "rows": 57,
                "cols": 9,
                "merged": 10,
                "synced_row_count": 57,
                "synced_column_count": 9,
                "synced_merges": [],
                "dynamic_merge_signature": f"sig-{building}",
                "error": "",
            },
        }

    monkeypatch.setattr(service, "_sync_one_building", _fake_sync_one_building)  # noqa: SLF001

    result = service.sync_confirmed_buildings(
        batch_meta={
            "batch_key": "2026-03-23|night",
            "spreadsheet_token": "sheet_token_1",
            "spreadsheet_url": "https://example.test/wiki",
            "spreadsheet_title": "demo",
        },
        building_items=[
            {"building": "A", "output_file": "A.xlsx", "revision": 1, "cloud_sheet_sync": {}},
            {"building": "B", "output_file": "B.xlsx", "revision": 2, "cloud_sheet_sync": {}},
            {"building": "C", "output_file": "C.xlsx", "revision": 3, "cloud_sheet_sync": {}},
        ],
        emit_log=lambda *_args: None,
    )

    assert result["status"] == "ok"
    assert result["uploaded_buildings"] == ["A", "B", "C"]
    assert list(result["details"].keys()) == ["A", "B", "C"]
    assert call_order == ["A", "B", "B", "C"]
    assert attempts_by_building == {"A": 1, "B": 2, "C": 1}
