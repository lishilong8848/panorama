from __future__ import annotations

import openpyxl

from handover_log_module.core.footer_layout import (
    FOOTER_GROUP_TITLE_TEXT,
    FOOTER_INVENTORY_COLUMNS,
    FOOTER_SIGNOFF_MARKER,
    FOOTER_TITLE_TEXT,
    find_footer_inventory_layout,
)
from handover_log_module.repository.footer_inventory_writer import write_footer_inventory_table
from handover_log_module.service.review_document_parser import ReviewDocumentParser


SHEET_NAME = 'handover'
AUDIT_LABEL = 'audit:'


def _build_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws['A50'] = FOOTER_TITLE_TEXT
    ws.merge_cells('A50:H50')

    ws['A51'] = FOOTER_GROUP_TITLE_TEXT
    ws['B51'] = str(FOOTER_INVENTORY_COLUMNS[0]['label'])
    ws['C51'] = str(FOOTER_INVENTORY_COLUMNS[1]['label'])
    ws['E51'] = str(FOOTER_INVENTORY_COLUMNS[2]['label'])
    ws['F51'] = str(FOOTER_INVENTORY_COLUMNS[3]['label'])
    ws['G51'] = str(FOOTER_INVENTORY_COLUMNS[4]['label'])
    ws['H51'] = str(FOOTER_INVENTORY_COLUMNS[5]['label'])
    ws.merge_cells('A51:A55')
    ws.merge_cells('C51:D51')

    data_rows = [
        ('tool-1', 'room-1', '1', 'no', 'none', 'user-1'),
        ('tool-2', 'room-1', '5', 'no', 'none', 'user-2'),
        ('tool-3', 'room-2', '4', 'no', 'none', 'user-3'),
        ('tool-4', 'room-1', '1', 'no', 'none', 'user-4'),
    ]
    for offset, values in enumerate(data_rows, start=52):
        ws[f'B{offset}'] = values[0]
        ws[f'C{offset}'] = values[1]
        ws[f'E{offset}'] = values[2]
        ws[f'F{offset}'] = values[3]
        ws[f'G{offset}'] = values[4]
        ws[f'H{offset}'] = values[5]
        ws.merge_cells(f'C{offset}:D{offset}')

    ws['A56'] = FOOTER_SIGNOFF_MARKER
    ws['C56'] = 'handover-sign'
    ws['F56'] = 'receive-sign'
    ws['F57'] = AUDIT_LABEL
    ws.merge_cells('A56:B56')
    ws.merge_cells('D56:E56')
    ws.merge_cells('G56:H56')
    return wb, ws


def test_review_document_parser_splits_footer_into_inventory_and_signoff_blocks() -> None:
    wb, ws = _build_ws()
    try:
        parser = ReviewDocumentParser({'template': {'sheet_name': SHEET_NAME}, 'review_ui': {'fixed_cells': {}}})
        document = parser.parse_from_worksheet(ws) if hasattr(parser, 'parse_from_worksheet') else {
            'footer_blocks': parser._footer_blocks(ws),  # noqa: SLF001
        }
        footer_blocks = document['footer_blocks']
        assert footer_blocks[0]['type'] == 'inventory_table'
        assert footer_blocks[0]['columns'][1]['key'] == 'C'
        assert footer_blocks[0]['columns'][1]['source_cols'] == ['C', 'D']
        assert footer_blocks[1]['type'] == 'readonly_grid'
    finally:
        wb.close()


def test_footer_inventory_writer_inserts_row_and_preserves_signoff_block() -> None:
    wb, ws = _build_ws()
    try:
        layout = find_footer_inventory_layout(ws)
        assert layout is not None
        assert layout.data_end_row == 55

        write_footer_inventory_table(
            ws=ws,
            inventory_block={
                'type': 'inventory_table',
                'columns': [
                    {'key': 'B', 'source_cols': ['B']},
                    {'key': 'C', 'source_cols': ['C', 'D']},
                    {'key': 'E', 'source_cols': ['E']},
                    {'key': 'F', 'source_cols': ['F']},
                    {'key': 'G', 'source_cols': ['G']},
                    {'key': 'H', 'source_cols': ['H']},
                ],
                'rows': [
                    {'cells': {'B': 'tool-1', 'C': 'room-1', 'E': '1', 'F': 'no', 'G': 'none', 'H': 'user-1'}},
                    {'cells': {'B': 'tool-2', 'C': 'room-1', 'E': '5', 'F': 'no', 'G': 'none', 'H': 'user-2'}},
                    {'cells': {'B': 'tool-3', 'C': 'room-2', 'E': '4', 'F': 'no', 'G': 'none', 'H': 'user-3'}},
                    {'cells': {'B': 'tool-4', 'C': 'room-1', 'E': '1', 'F': 'no', 'G': 'none', 'H': 'user-4'}},
                    {'cells': {'B': 'tool-5', 'C': 'room-5', 'E': '2', 'F': 'yes', 'G': 'note', 'H': 'user-5'}},
                ],
            },
            emit_log=lambda *_: None,
        )

        layout_after = find_footer_inventory_layout(ws)
        assert layout_after is not None
        assert layout_after.data_end_row == 56
        assert ws.max_row == layout_after.last_row
        assert ws['B56'].value == 'tool-5'
        assert ws['C56'].value == 'room-5'
        assert ws['A57'].value == FOOTER_SIGNOFF_MARKER
        assert 'A51:A56' in {str(item) for item in ws.merged_cells.ranges}
        assert 'C56:D56' in {str(item) for item in ws.merged_cells.ranges}
    finally:
        wb.close()


def test_footer_inventory_writer_shrinks_rows_without_merged_cell_crash() -> None:
    wb, ws = _build_ws()
    try:
        write_footer_inventory_table(
            ws=ws,
            inventory_block={
                'type': 'inventory_table',
                'columns': [
                    {'key': 'B', 'source_cols': ['B']},
                    {'key': 'C', 'source_cols': ['C', 'D']},
                    {'key': 'E', 'source_cols': ['E']},
                    {'key': 'F', 'source_cols': ['F']},
                    {'key': 'G', 'source_cols': ['G']},
                    {'key': 'H', 'source_cols': ['H']},
                ],
                'rows': [
                    {'cells': {'B': 'tool-1', 'C': 'room-1', 'E': '1', 'F': 'no', 'G': 'none', 'H': 'user-1'}},
                ],
            },
            emit_log=lambda *_: None,
        )

        layout_after = find_footer_inventory_layout(ws)
        assert layout_after is not None
        assert layout_after.data_end_row == 52
        assert ws.max_row == layout_after.last_row
        assert ws['B52'].value == 'tool-1'
        assert ws['F54'].value == AUDIT_LABEL
        assert 'A51:A52' in {str(item) for item in ws.merged_cells.ranges}
    finally:
        wb.close()
