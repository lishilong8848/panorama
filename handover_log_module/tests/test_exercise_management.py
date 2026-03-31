from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from handover_log_module.repository.exercise_management_repository import (
    ExerciseManagementRepository,
    ExerciseManagementRow,
)
from handover_log_module.service.exercise_management_payload_builder import (
    ExerciseManagementPayloadBuilder,
)


def _build_exercise_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A29"] = "演练管理"
    ws["A30"] = "序号"
    ws["B30"] = "演练类型"
    ws["C30"] = "演练项目"
    ws["D30"] = "演练完成情况"
    ws["H30"] = "执行人"
    ws["A31"] = 1
    ws["B31"] = "/"
    ws["C31"] = "/"
    ws["D31"] = "/"
    ws["H31"] = "/"
    ws.merge_cells("D31:G31")
    ws["A32"] = "交接确认"
    wb.save(path)
    wb.close()


class _FakeExerciseRepo:
    def __init__(self, cfg, rows):
        self._cfg = cfg
        self._rows = rows

    def get_config(self):
        return self._cfg

    def list_current_shift_rows(self, **_kwargs):
        return list(self._rows), self._cfg


class _GroupedExerciseRepo(ExerciseManagementRepository):
    def __init__(self, rows):
        super().__init__({})
        self._rows = rows

    def _load_rows_for_shift(self, *, duty_date, duty_shift, emit_log):  # noqa: ARG002
        cfg = self.get_config()
        return list(self._rows), cfg, {"total": len(self._rows), "in_shift": len(self._rows), "parse_fail": 0}


def test_exercise_management_builder_maps_template_columns(tmp_path: Path) -> None:
    template_path = tmp_path / "exercise_template.xlsx"
    _build_exercise_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"exercise_management": "演练管理"},
        "fixed_values": {"exercise_type": "计划性演练", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "exercise_type": ["演练类型"],
                "exercise_item": ["演练项目"],
                "completion": ["演练完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "exercise_type": "B",
                "exercise_item": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        ExerciseManagementRow(
            record_id="rec1",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            project_text="UPS切换演练",
            raw_fields={},
        )
    ]
    builder = ExerciseManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeExerciseRepo(cfg, rows),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        executor_text="张三、李四",
        emit_log=lambda *_args: None,
    )

    assert list(payload.keys()) == ["演练管理"]
    assert payload["演练管理"] == [
        {"cells": {"B": "计划性演练", "C": "UPS切换演练", "D": "已完成", "H": "张三、李四"}}
    ]


def test_exercise_management_grouped_rows_reuse_multi_building_records() -> None:
    rows = [
        ExerciseManagementRow(
            record_id="rec-a",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            project_text="A项目",
            raw_fields={},
        ),
        ExerciseManagementRow(
            record_id="rec-ac",
            building_values=["A楼", "C楼"],
            start_time=datetime(2026, 3, 14, 9, 30, 0),
            project_text="AC项目",
            raw_fields={},
        ),
        ExerciseManagementRow(
            record_id="rec-b",
            building_values=["B楼"],
            start_time=datetime(2026, 3, 14, 10, 0, 0),
            project_text="B项目",
            raw_fields={},
        ),
    ]
    repo = _GroupedExerciseRepo(rows)

    grouped, _ = repo.list_current_shift_rows_grouped(
        buildings=["A楼", "C楼"],
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert [row.record_id for row in grouped["A楼"]] == ["rec-a", "rec-ac"]
    assert [row.record_id for row in grouped["C楼"]] == ["rec-ac"]


def test_exercise_management_builder_skips_blank_project_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "exercise_template_blank.xlsx"
    _build_exercise_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"exercise_management": "演练管理"},
        "fixed_values": {"exercise_type": "计划性演练", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "exercise_type": ["演练类型"],
                "exercise_item": ["演练项目"],
                "completion": ["演练完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "exercise_type": "B",
                "exercise_item": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        ExerciseManagementRow(
            record_id="rec-blank",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 8, 0, 0),
            project_text="",
            raw_fields={},
        ),
        ExerciseManagementRow(
            record_id="rec-keep",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            project_text="UPS切换演练",
            raw_fields={},
        ),
    ]
    builder = ExerciseManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeExerciseRepo(cfg, rows),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        executor_text="张三、李四",
        emit_log=lambda *_args: None,
    )

    assert payload["演练管理"] == [
        {"cells": {"B": "计划性演练", "C": "UPS切换演练", "D": "已完成", "H": "张三、李四"}}
    ]
