from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from app.config.config_adapter import ensure_v3_config
from app.modules.handover_review.api.routes import _build_review_document_state_service, _persist_review_defaults
from handover_log_module.service.cabinet_power_defaults_service import CabinetPowerDefaultsService


SHEET_NAME = "交接班日志"
TEMPLATE_CONFIG = Path(__file__).resolve().parents[2] / "config" / "表格计算配置.template.json"


def _build_handover_file(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A12"] = "机柜上下电信息"
    ws["B13"] = ""
    ws["D13"] = ""
    ws["F13"] = ""
    ws["H13"] = ""
    wb.save(path)
    wb.close()


class _DummyContainer:
    def __init__(self, config_path: Path) -> None:
        raw = json.loads(TEMPLATE_CONFIG.read_text(encoding="utf-8-sig"))
        self.config = ensure_v3_config(raw)
        self.config.setdefault("features", {}).setdefault("alarm_export", {})["window_days"] = 1
        self.config.setdefault("common", {}).setdefault("paths", {})["runtime_state_root"] = str(
            config_path.parent / ".runtime"
        )
        self.runtime_config = self.config
        self.config_path = config_path
        self.config_path.write_text(json.dumps(self.config, ensure_ascii=False, indent=2), encoding="utf-8-sig")
        self.reload_calls = 0

    def reload_config(self, settings):
        self.reload_calls += 1
        self.config = settings


def test_cabinet_power_defaults_roundtrip_v3_config() -> None:
    service = CabinetPowerDefaultsService()
    document = {
        "fixed_blocks": [
            {
                "id": "cabinet_power_info",
                "title": "机柜上下电信息",
                "fields": [
                    {"cell": "B13", "value": "1272"},
                    {"cell": "D13", "value": "2"},
                    {"cell": "F13", "value": "/"},
                    {"cell": "H13", "value": "/"},
                ],
            }
        ]
    }

    updated = service.set_building_defaults(
        ensure_v3_config({}),
        "A楼",
        service.extract_cells_from_document(document),
    )

    assert service.get_building_defaults(updated, "A楼") == {
        "B13": "1272",
        "D13": "2",
        "F13": "/",
        "H13": "/",
    }
    assert (
        updated["features"]["handover_log"]["review_ui"]["cabinet_power_defaults_by_building"]["A楼"]["cells"]["B13"]
        == "1272"
    )


def test_apply_building_defaults_to_output_overwrites_cabinet_cells(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_handover_file(output_file)
    service = CabinetPowerDefaultsService()

    applied_fields = service.apply_building_defaults_to_output(
        config={
            "review_ui": {
                "cabinet_power_defaults_by_building": {
                    "A楼": {
                        "cells": {
                            "B13": "1272",
                            "D13": "2",
                            "F13": "1",
                            "H13": "0",
                        }
                    }
                }
            }
        },
        building="A楼",
        output_file=output_file,
        sheet_name=SHEET_NAME,
        emit_log=lambda *_: None,
    )

    assert applied_fields == 4

    wb = openpyxl.load_workbook(output_file)
    try:
        ws = wb[SHEET_NAME]
        assert ws["B13"].value == "1272"
        assert ws["D13"].value == "2"
        assert ws["F13"].value == "1"
        assert ws["H13"].value == "0"
    finally:
        wb.close()


def test_persist_review_defaults_updates_cabinet_and_footer_defaults(tmp_path: Path) -> None:
    config_path = tmp_path / "config.json"
    container = _DummyContainer(config_path)
    document = {
        "fixed_blocks": [
            {
                "id": "cabinet_power_info",
                "fields": [
                    {"cell": "B13", "value": "1272"},
                    {"cell": "D13", "value": "2"},
                    {"cell": "F13", "value": "1"},
                    {"cell": "H13", "value": "0"},
                ],
            }
        ],
        "footer_blocks": [
            {
                "id": "handover_inventory_table",
                "type": "inventory_table",
                "rows": [
                    {"cells": {"B": "对讲机", "C": "A楼值班室", "E": "5", "F": "否", "G": "无", "H": "李四"}}
                ],
            }
        ],
    }

    persisted = _persist_review_defaults(container, building="A楼", document=document)

    assert persisted["footer_inventory_rows"] == 1
    assert persisted["cabinet_power_fields"] == 4
    assert persisted["config_updated"] is False
    assert persisted["defaults_updated"] is True
    assert persisted["config_sync_required"] is True
    assert persisted["config_building_code"] == "A"
    assert isinstance(persisted["config_data"], dict)
    assert container.reload_calls == 0
    state_service = _build_review_document_state_service(container)
    store = state_service._store("A楼")
    assert store.get_default("cabinet_power") == {"B13": "1272", "D13": "2", "F13": "1", "H13": "0"}
    assert store.get_default("footer_inventory")[0]["cells"] == {
        "B": "对讲机",
        "C": "A楼值班室",
        "E": "5",
        "F": "否",
        "G": "无",
    }
    saved = json.loads(config_path.read_text(encoding="utf-8-sig"))
    assert saved == container.config


def test_persist_review_defaults_skips_config_write_when_defaults_unchanged(tmp_path: Path) -> None:
    config_path = tmp_path / "config.json"
    container = _DummyContainer(config_path)
    document = {
        "fixed_blocks": [
            {
                "id": "cabinet_power_info",
                "fields": [
                    {"cell": "B13", "value": "1272"},
                    {"cell": "D13", "value": "2"},
                    {"cell": "F13", "value": "1"},
                    {"cell": "H13", "value": "0"},
                ],
            }
        ],
        "footer_blocks": [
            {
                "id": "handover_inventory_table",
                "type": "inventory_table",
                "rows": [
                    {"cells": {"B": "对讲机", "C": "A楼值班室", "E": "5", "F": "否", "G": "无", "H": "李四"}}
                ],
            }
        ],
    }

    first = _persist_review_defaults(container, building="A楼", document=document)
    first_config = config_path.read_text(encoding="utf-8-sig")

    second = _persist_review_defaults(container, building="A楼", document=document)
    second_config = config_path.read_text(encoding="utf-8-sig")

    assert first["config_updated"] is False
    assert first["defaults_updated"] is True
    assert second["footer_inventory_rows"] == 1
    assert second["cabinet_power_fields"] == 4
    assert second["config_updated"] is False
    assert second["defaults_updated"] is False
    assert second["config_sync_required"] is True
    assert second["config_building_code"] == "A"
    assert isinstance(second["config_data"], dict)
    assert container.reload_calls == 0
    assert second_config == first_config
