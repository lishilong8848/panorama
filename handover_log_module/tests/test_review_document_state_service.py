from __future__ import annotations

from pathlib import Path

import openpyxl

from handover_log_module.service.review_document_state_service import ReviewDocumentStateService


def _config(tmp_path: Path) -> dict:
    return {
        "_global_paths": {"runtime_state_root": str(tmp_path / ".runtime")},
        "template": {"sheet_name": "交接班日志"},
        "review_ui": {
            "fixed_cells": {
                "header_basic": ["A1"],
                "cabinet_power_info": ["B13", "D13", "F13", "H13"],
            }
        },
    }


def _build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A1"] = "原标题"
    ws["B13"] = "旧规划"
    ws["D13"] = "旧上电"
    ws["F13"] = "旧本班上电"
    ws["H13"] = "旧本班下电"
    wb.save(path)
    wb.close()


def _session(path: Path, *, revision: int = 1) -> dict:
    return {
        "session_id": "A楼|2026-04-15|day",
        "building": "A楼",
        "duty_date": "2026-04-15",
        "duty_shift": "day",
        "batch_key": "2026-04-15|day",
        "revision": revision,
        "output_file": str(path),
    }


def _fixed_value(document: dict, cell: str) -> str:
    for block in document.get("fixed_blocks", []):
        for field in block.get("fields", []):
            if field.get("cell") == cell:
                return field.get("value", "")
    return ""


def _set_fixed_value(document: dict, cell: str, value: str) -> None:
    for block in document.get("fixed_blocks", []):
        for field in block.get("fields", []):
            if field.get("cell") == cell:
                field["value"] = value
                return


def test_review_document_state_imports_saves_and_force_syncs_excel(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, loaded_session = service.load_document(session)
    assert loaded_session["revision"] == 1
    assert _fixed_value(document, "B13") == "旧规划"

    _set_fixed_value(document, "B13", "新规划")
    state, _previous = service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    assert state["revision"] == 2

    wb = openpyxl.load_workbook(output_file)
    try:
        assert wb["交接班日志"]["B13"].value == "旧规划"
    finally:
        wb.close()

    service.force_sync_session(building="A楼", session_id=session["session_id"], target_revision=2, reason="test")

    wb = openpyxl.load_workbook(output_file)
    try:
        assert wb["交接班日志"]["B13"].value == "新规划"
    finally:
        wb.close()


def test_review_document_state_stores_building_defaults_without_config_file(tmp_path: Path) -> None:
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    document = {
        "fixed_blocks": [
            {
                "id": "cabinet_power_info",
                "fields": [
                    {"cell": "B13", "value": "100"},
                    {"cell": "D13", "value": "20"},
                    {"cell": "F13", "value": "1"},
                    {"cell": "H13", "value": "0"},
                ],
            }
        ],
        "footer_blocks": [
            {
                "type": "inventory_table",
                "rows": [
                    {"cells": {"B": "对讲机", "C": "值班室", "E": "5", "F": "否", "G": "无", "H": "不保存"}}
                ],
            }
        ],
    }

    result = service.persist_defaults_from_document(
        building="A楼",
        document=document,
        dirty_regions={"fixed_blocks": True, "footer_inventory": True},
    )

    assert result["defaults_updated"] is True
    store = service._store("A楼")
    assert store.get_default("cabinet_power") == {"B13": "100", "D13": "20", "F13": "1", "H13": "0"}
    assert store.get_default("footer_inventory")[0]["cells"] == {
        "B": "对讲机",
        "C": "值班室",
        "E": "5",
        "F": "否",
        "G": "无",
    }


def test_review_document_state_save_enqueues_sync_job_atomically(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, _loaded_session = service.load_document(session)
    _set_fixed_value(document, "B13", "新规划")
    state, _previous = service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )

    store = service._store("A楼")
    job = store.claim_next_job()
    sync = store.get_sync_state(session["session_id"])

    assert state["revision"] == 2
    assert job == {"session_id": session["session_id"], "target_revision": 2, "attempts": 1}
    assert sync["status"] == "pending"
    assert sync["pending_revision"] == 2


def test_force_sync_session_uses_saved_dirty_regions(tmp_path: Path, monkeypatch) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, _loaded_session = service.load_document(session)
    _set_fixed_value(document, "B13", "新规划")
    service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True, "sections": False, "footer_inventory": False},
    )

    captured: list[dict] = []

    def _fake_write(*, output_file: str, document: dict, dirty_regions: dict | None = None) -> None:
        captured.append(
            {
                "output_file": output_file,
                "dirty_regions": dict(dirty_regions or {}),
                "title": document.get("title", ""),
            }
        )

    monkeypatch.setattr(service.writer, "write", _fake_write)

    sync = service.force_sync_session(
        building="A楼",
        session_id=session["session_id"],
        target_revision=2,
        reason="test",
    )

    assert sync["status"] == "synced"
    assert captured == [
        {
            "output_file": str(output_file),
            "dirty_regions": {"fixed_blocks": True, "sections": False, "footer_inventory": False},
            "title": "原标题",
        }
    ]


def test_ensure_document_reimports_when_excel_fingerprint_changes_without_path_change(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, _loaded_session = service.load_document(session)
    assert _fixed_value(document, "B13") == "旧规划"

    wb = openpyxl.load_workbook(output_file)
    try:
        ws = wb["交接班日志"]
        ws["B13"] = "覆盖后规划"
        wb.save(output_file)
    finally:
        wb.close()

    reloaded_document, _reloaded_session = service.load_document(session)
    assert _fixed_value(reloaded_document, "B13") == "覆盖后规划"


def test_worker_loop_recovers_after_claim_exception(tmp_path: Path) -> None:
    logs: list[str] = []
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=logs.append)

    class _Store:
        def __init__(self):
            self.calls = 0

        def claim_next_job(self):
            self.calls += 1
            if self.calls == 1:
                raise RuntimeError("temporary sqlite fault")
            return None

    store = _Store()
    service._store = lambda _building: store  # type: ignore[method-assign]

    service._worker_loop("A楼")

    assert any("后台Excel同步线程异常，已自动恢复" in message for message in logs)


def test_finish_job_success_keeps_newer_pending_revision(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, _loaded_session = service.load_document(session)
    _set_fixed_value(document, "B13", "rev2")
    state, _previous = service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    store = service._store("A楼")
    first_job = store.claim_next_job()
    assert first_job is not None

    _set_fixed_value(document, "B13", "rev3")
    state, _previous = service.save_document(
        session={**session, "revision": 2},
        document=document,
        base_revision=2,
        dirty_regions={"fixed_blocks": True},
    )
    sync = store.finish_job(
        session_id=session["session_id"],
        success=True,
        claimed_target_revision=2,
        synced_revision=2,
    )
    next_job = store.claim_next_job()

    assert state["revision"] == 3
    assert sync["status"] == "pending"
    assert sync["synced_revision"] == 2
    assert sync["pending_revision"] == 3
    assert next_job == {"session_id": session["session_id"], "target_revision": 3, "attempts": 1}


def test_finish_job_failure_keeps_newer_pending_revision(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)

    document, _loaded_session = service.load_document(session)
    _set_fixed_value(document, "B13", "rev2")
    service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    store = service._store("A楼")
    first_job = store.claim_next_job()
    assert first_job is not None

    _set_fixed_value(document, "B13", "rev3")
    service.save_document(
        session={**session, "revision": 2},
        document=document,
        base_revision=2,
        dirty_regions={"fixed_blocks": True},
    )
    sync = store.finish_job(
        session_id=session["session_id"],
        success=False,
        claimed_target_revision=2,
        error="boom",
    )
    next_job = store.claim_next_job()

    assert sync["status"] == "pending"
    assert sync["pending_revision"] == 3
    assert sync["error"] == ""
    assert next_job == {"session_id": session["session_id"], "target_revision": 3, "attempts": 1}


def test_review_document_state_mirrors_and_clears_defaults_from_config(tmp_path: Path) -> None:
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    config = {
        "features": {
            "handover_log": {
                "review_ui": {
                    "cabinet_power_defaults_by_building": {
                        "A楼": {"cells": {"B13": "10", "D13": "11", "F13": "12", "H13": "13"}}
                    },
                    "footer_inventory_defaults_by_building": {
                        "A楼": {
                            "rows": [
                                {"cells": {"B": "对讲机", "C": "值班室", "E": "5", "F": "否", "G": "无", "H": "忽略"}}
                            ]
                        }
                    },
                }
            }
        }
    }

    mirrored = service.persist_defaults_from_config(building="A楼", config=config)
    store = service._store("A楼")

    assert mirrored["defaults_updated"] is True
    assert store.get_default("cabinet_power") == {"B13": "10", "D13": "11", "F13": "12", "H13": "13"}
    assert store.get_default("footer_inventory")[0]["cells"] == {
        "B": "对讲机",
        "C": "值班室",
        "E": "5",
        "F": "否",
        "G": "无",
    }

    cleared = service.persist_defaults_from_config(
        building="A楼",
        config={"features": {"handover_log": {"review_ui": {}}}},
    )

    assert cleared["defaults_updated"] is True
    assert store.get_default("cabinet_power") is None
    assert store.get_default("footer_inventory") is None


def test_review_document_state_reimports_when_session_output_file_changes(tmp_path: Path) -> None:
    first_output = tmp_path / "handover_v1.xlsx"
    second_output = tmp_path / "handover_v2.xlsx"
    _build_workbook(first_output)
    _build_workbook(second_output)
    wb = openpyxl.load_workbook(second_output)
    try:
        ws = wb["交接班日志"]
        ws["B13"] = "新版本规划"
        ws["D13"] = "新版本上电"
        wb.save(second_output)
    finally:
        wb.close()

    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session_v1 = _session(first_output)
    document_v1, loaded_session_v1 = service.load_document(session_v1)
    assert loaded_session_v1["revision"] == 1
    assert _fixed_value(document_v1, "B13") == "旧规划"

    document_v1["title"] = "本地旧版本编辑"
    _set_fixed_value(document_v1, "B13", "旧版本本地修改")
    service.save_document(
        session=session_v1,
        document=document_v1,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )

    session_v2 = _session(second_output, revision=2)
    document_v2, loaded_session_v2 = service.load_document(session_v2)

    assert loaded_session_v2["revision"] == 2
    assert document_v2["title"] == "原标题"
    assert _fixed_value(document_v2, "B13") == "新版本规划"
    assert _fixed_value(document_v2, "D13") == "新版本上电"

    store = service._store("A楼")
    state = store.get_document(session_v2["session_id"])
    assert state is not None
    assert state["source_excel_path"] == str(second_output)
