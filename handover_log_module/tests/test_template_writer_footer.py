from __future__ import annotations

from pathlib import Path

import openpyxl

from handover_log_module.core.footer_layout import find_footer_inventory_layout
from handover_log_module.core.section_layout import parse_category_sections
from handover_log_module.repository.template_writer import _write_workbook


def _template_path() -> Path:
    return Path(__file__).resolve().parents[2] / "交接班日志空模板.xlsx"


def test_write_workbook_preserves_footer_after_large_category_expansion(tmp_path: Path) -> None:
    out_path = tmp_path / "handover_footer.xlsx"
    _write_workbook(
        source_path=_template_path(),
        out_path=out_path,
        sheet_name="交接班日志",
        cell_values={
            "B2": "2026年3月15日",
            "F2": "白班",
            "C3": "当前班组",
            "G3": "下一班组",
            "H52": "徐志远",
            "H53": "徐志远",
            "H54": "徐志远",
            "H55": "徐志远",
        },
        category_payloads={
            "新事件处理": [{"cells": {"B": "紧急", "F": "未完成", "H": "甲"}} for _ in range(7)],
            "历史事件跟进": [{"cells": {"B": "重要", "F": "未完成", "H": "乙"}} for _ in range(5)],
            "其他重要工作记录": [{"cells": {"B": "巡检", "F": "已完成", "H": "丙"}} for _ in range(6)],
        },
        emit_log=lambda *_: None,
    )

    wb = openpyxl.load_workbook(out_path)
    try:
        ws = wb["交接班日志"]
        layout = find_footer_inventory_layout(ws)
        assert layout is not None
        assert ws.max_row == layout.last_row

        assert ws.cell(layout.title_row, 1).value == "交接确认"
        assert ws.cell(layout.header_row, 1).value == "工具及物品交接清点"
        assert ws.cell(layout.header_row, 2).value == "交接工具名称"
        assert ws.cell(layout.header_row, 5).value == "数量"
        assert ws.cell(layout.header_row, 8).value == "清点确认人（接班）"

        actual_rows = [
            (
                ws.cell(row_idx, 2).value,
                ws.cell(row_idx, 3).value,
                ws.cell(row_idx, 5).value,
                ws.cell(row_idx, 6).value,
                ws.cell(row_idx, 7).value,
                ws.cell(row_idx, 8).value,
            )
            for row_idx in range(layout.data_start_row, layout.data_end_row + 1)
        ]
        assert actual_rows == [
            ("值班手机", "E-112值班室", 1, "否", "无", "徐志远"),
            ("对讲机", "E-112值班室", 5, "否", "无", "徐志远"),
            ("钥匙", "E-112值班室钥匙箱", 4, "否", "无", "徐志远"),
            ("应急工具", "E-112值班室", 1, "否", "无", "徐志远"),
        ]

        merged_refs = {str(item) for item in ws.merged_cells.ranges}
        assert f"A{layout.title_row}:H{layout.title_row}" in merged_refs
        assert f"A{layout.header_row}:A{layout.data_end_row}" in merged_refs
        for row_idx in range(layout.header_row, layout.data_end_row + 1):
            assert f"C{row_idx}:D{row_idx}" in merged_refs
        assert layout.signoff_start_row is not None
        assert f"A{layout.signoff_start_row}:B{layout.signoff_start_row}" in merged_refs
        assert f"D{layout.signoff_start_row}:E{layout.signoff_start_row}" in merged_refs
        assert f"G{layout.signoff_start_row}:H{layout.signoff_start_row}" in merged_refs
    finally:
        wb.close()


def test_write_workbook_trims_blank_rows_below_footer_without_category_changes(tmp_path: Path) -> None:
    out_path = tmp_path / "handover_footer_trim.xlsx"
    _write_workbook(
        source_path=_template_path(),
        out_path=out_path,
        sheet_name="交接班日志",
        cell_values={
            "B2": "2026年3月15日",
            "F2": "白班",
        },
        category_payloads=None,
        emit_log=lambda *_: None,
    )

    wb = openpyxl.load_workbook(out_path)
    try:
        ws = wb["交接班日志"]
        layout = find_footer_inventory_layout(ws)
        assert layout is not None
        assert ws.max_row == layout.last_row
        assert ws.cell(layout.last_row, 6).value == "审核人："
    finally:
        wb.close()


def test_write_workbook_clears_empty_sections_and_keeps_footer_contiguous(tmp_path: Path) -> None:
    out_path = tmp_path / "handover_empty_sections.xlsx"
    _write_workbook(
        source_path=_template_path(),
        out_path=out_path,
        sheet_name="交接班日志",
        cell_values={
            "B2": "2026年3月15日",
            "F2": "白班",
        },
        category_payloads={
            "新事件处理": [{"cells": {"B": "紧急", "F": "未完成", "H": "甲"}} for _ in range(6)],
            "历史事件跟进": [],
            "其他重要工作记录": [{"cells": {"B": "巡检", "F": "已完成", "H": "丙"}} for _ in range(2)],
        },
        emit_log=lambda *_: None,
    )

    wb = openpyxl.load_workbook(out_path)
    try:
        ws = wb["交接班日志"]
        sections = parse_category_sections(ws)
        hist_section = next(section for section in sections if section.name == "历史事件跟进")
        later_section = next(section for section in sections if section.name == "其他重要工作记录")
        layout = find_footer_inventory_layout(ws)

        assert hist_section.end_row - hist_section.template_data_row + 1 == 1
        assert ws[f"A{hist_section.template_data_row}"].value in ("", None)
        assert ws[f"B{hist_section.template_data_row}"].value in ("", None)
        assert ws[f"D{hist_section.template_data_row}"].value in ("", None)
        assert ws[f"F{hist_section.template_data_row}"].value in ("", None)
        assert ws.cell(later_section.title_row, 1).value == "其他重要工作记录"
        assert layout is not None
        assert ws.max_row == layout.last_row
        assert layout.title_row > later_section.end_row
        assert ws.cell(layout.title_row, 1).value == "交接确认"
    finally:
        wb.close()
