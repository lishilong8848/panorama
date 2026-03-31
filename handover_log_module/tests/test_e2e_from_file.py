from __future__ import annotations

from pathlib import Path

import openpyxl

from handover_log_module.api.facade import run_from_existing_file


SHEET_NAME = "handover"


def _build_data_file(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"
    # Minimal rows: parser reads B/C/D/E from row >= 4.
    ws.cell(4, 2, "zone/A-301")
    ws.cell(4, 3, "C3-1")
    ws.cell(4, 4, "冷通道温度")
    ws.cell(4, 5, 25.1)
    ws.cell(5, 2, "zone/A-302")
    ws.cell(5, 3, "C3-2")
    ws.cell(5, 4, "冷通道湿度")
    ws.cell(5, 5, 45.2)
    wb.save(path)
    wb.close()


def _build_template_file(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    wb.save(path)
    wb.close()


def test_e2e_from_existing_file_success(tmp_path: Path) -> None:
    data_file = tmp_path / "data.xlsx"
    tpl_file = tmp_path / "template.xlsx"
    _build_data_file(data_file)
    _build_template_file(tpl_file)

    cfg = {
        "handover_log": {
            "template": {
                "source_path": str(tpl_file),
                "sheet_name": SHEET_NAME,
                "output_dir": str(tmp_path / "output"),
                "file_name_pattern": "{building}_{date}_handover.xlsx",
                "date_format": "%Y%m%d",
            }
        }
    }

    summary = run_from_existing_file(
        config=cfg,
        building="C楼",
        data_file=str(data_file),
        end_time="2026-03-01 10:00:00",
        emit_log=lambda _: None,
    )
    results = summary["results"]
    assert len(results) == 1
    assert results[0]["success"] is True
    out_path = Path(results[0]["output_file"])
    assert out_path.exists()


def test_from_existing_file_with_duty_fills_fixed_cells(tmp_path: Path) -> None:
    data_file = tmp_path / "data_with_duty.xlsx"
    tpl_file = tmp_path / "template_with_duty.xlsx"
    _build_data_file(data_file)
    _build_template_file(tpl_file)

    cfg = {
        "handover_log": {
            "template": {
                "source_path": str(tpl_file),
                "sheet_name": SHEET_NAME,
                "output_dir": str(tmp_path / "output"),
                "file_name_pattern": "{building}_{date}_handover.xlsx",
                "date_format": "%Y%m%d",
            },
            "template_fixed_fill": {
                "date_cell": "B2",
                "shift_cell": "F2",
                "alarm_total_cell": "B15",
                "alarm_unrecovered_cell": "D15",
                "alarm_accept_desc_cell": "F15",
                "date_text_format": "{year}-{month}-{day}",
                "shift_text": {"day": "DAY", "night": "NIGHT"},
                "on_alarm_query_fail": {"total": "0", "unrecovered": "0", "accept_desc": "/"},
            },
        }
    }

    summary = run_from_existing_file(
        config=cfg,
        building="X楼",  # No site host -> alarm query falls back by design.
        data_file=str(data_file),
        duty_date="2026-03-01",
        duty_shift="day",
        emit_log=lambda _: None,
    )
    results = summary["results"]
    assert len(results) == 1
    assert results[0]["success"] is True

    out_path = Path(results[0]["output_file"])
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb[SHEET_NAME]
    assert ws["B2"].value == "2026-3-1"
    assert ws["F2"].value == "DAY"
    assert ws["B15"].value == "0"
    assert ws["D15"].value == "0"
    assert ws["F15"].value == "/"
    wb.close()
