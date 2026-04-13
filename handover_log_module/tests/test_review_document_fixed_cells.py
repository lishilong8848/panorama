from __future__ import annotations

from pathlib import Path

import openpyxl

from app.config.config_adapter import ensure_v3_config
import handover_log_module.service.review_document_writer as review_document_writer_module
from handover_log_module.service.review_document_parser import ReviewDocumentParser
from handover_log_module.service.review_document_writer import ReviewDocumentWriter


SHEET_NAME = "handover"


def _build_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet["A1"] = "title"
    worksheet["A6"] = "PUE"
    worksheet["B6"] = "1.31"
    worksheet["C6"] = "Total Load"
    worksheet["D6"] = "1200"
    worksheet["E6"] = "IT Load"
    worksheet["F6"] = "900"
    worksheet["G6"] = "Oil Backup"
    worksheet["H6"] = "18"
    worksheet["B15"] = "10"
    workbook.save(path)
    workbook.close()


def test_ensure_v3_config_upgrades_metrics_summary_strings_to_pairs() -> None:
    cfg = ensure_v3_config(
        {
            "common": {},
            "features": {
                "handover_log": {
                    "review_ui": {
                        "fixed_cells": {
                            "metrics_summary": ["D6", "F6", "H6", "B15"],
                        }
                    }
                }
            },
        }
    )

    metrics_summary = cfg["features"]["handover_log"]["review_ui"]["fixed_cells"]["metrics_summary"]

    assert {"label_cell": "A6", "value_cell": "B6"} in metrics_summary
    assert {"label_cell": "C6", "value_cell": "D6"} in metrics_summary
    assert {"label_cell": "E6", "value_cell": "F6"} in metrics_summary
    assert {"label_cell": "G6", "value_cell": "H6"} in metrics_summary
    assert "B15" in metrics_summary
    assert "D6" not in metrics_summary
    assert "F6" not in metrics_summary
    assert "H6" not in metrics_summary


def test_review_document_parser_and_writer_support_paired_fixed_cells(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    config = {
        "template": {"sheet_name": SHEET_NAME},
        "review_ui": {
            "fixed_cells": {
                "metrics_summary": [
                    {"label_cell": "A6", "value_cell": "B6"},
                    {"label_cell": "C6", "value_cell": "D6"},
                    {"label_cell": "E6", "value_cell": "F6"},
                    {"label_cell": "G6", "value_cell": "H6"},
                    "B15",
                ]
            }
        },
    }

    parser = ReviewDocumentParser(config)
    workbook = openpyxl.load_workbook(output_file)
    try:
        worksheet = workbook[SHEET_NAME]
        metrics_block = next(block for block in parser._fixed_blocks(worksheet) if block["id"] == "metrics_summary")  # noqa: SLF001
    finally:
        workbook.close()

    by_cell = {field["cell"]: field for field in metrics_block["fields"]}

    assert by_cell["B6"]["label"] == "PUE"
    assert by_cell["B6"]["value"] == "1.31"
    assert by_cell["D6"]["label"] == "Total Load"
    assert by_cell["F6"]["label"] == "IT Load"
    assert by_cell["H6"]["label"] == "Oil Backup"

    by_cell["B6"]["value"] = "1.42"
    writer = ReviewDocumentWriter(config)
    writer.write(
        output_file=str(output_file),
        document={"title": "title", "fixed_blocks": [metrics_block], "sections": [], "footer_blocks": []},
    )

    workbook = openpyxl.load_workbook(output_file)
    try:
        worksheet = workbook[SHEET_NAME]
        assert worksheet["A6"].value == "PUE"
        assert worksheet["B6"].value == "1.42"
    finally:
        workbook.close()


def test_review_document_writer_uses_dirty_regions_to_skip_section_and_footer_writes(tmp_path: Path, monkeypatch) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    config = {
        "template": {"sheet_name": SHEET_NAME},
        "review_ui": {"fixed_cells": {"metrics_summary": [{"label_cell": "A6", "value_cell": "B6"}]}},
    }
    calls = {"sections": 0, "footer": 0}

    monkeypatch.setattr(
        review_document_writer_module,
        "write_category_sections",
        lambda **_kwargs: calls.__setitem__("sections", calls["sections"] + 1),
    )
    monkeypatch.setattr(
        review_document_writer_module,
        "write_footer_inventory_table",
        lambda **_kwargs: calls.__setitem__("footer", calls["footer"] + 1),
    )

    writer = ReviewDocumentWriter(config)
    writer.write(
        output_file=str(output_file),
        document={
            "title": "title",
            "fixed_blocks": [{"id": "metrics_summary", "fields": [{"cell": "B6", "value": "1.88"}]}],
            "sections": [{"name": "新事件处理", "rows": [{"cells": {"B": "x"}}]}],
            "footer_blocks": [{"id": "handover_inventory_table", "type": "inventory_table", "rows": []}],
        },
        dirty_regions={"fixed_blocks": True, "sections": False, "footer_inventory": False},
    )

    workbook = openpyxl.load_workbook(output_file)
    try:
        worksheet = workbook[SHEET_NAME]
        assert worksheet["B6"].value == "1.88"
    finally:
        workbook.close()
    assert calls == {"sections": 0, "footer": 0}
