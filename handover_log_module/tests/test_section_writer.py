from __future__ import annotations

from openpyxl.styles import PatternFill
import openpyxl

from handover_log_module.core.section_layout import capture_section_snapshots, parse_category_sections
from handover_log_module.repository.section_writer import write_category_sections


def _build_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"

    ws["A20"] = "新事件处理"
    ws["A21"] = "序号"
    ws["B21"] = "事件等级"
    ws["D21"] = "描述"
    ws["F21"] = "处理进展"
    ws["A22"] = 1
    ws["B22"] = "/"
    ws["D22"] = "/"
    ws["F22"] = "/"
    ws.merge_cells("D22:E22")
    ws.merge_cells("F22:G22")

    ws["A23"] = "历史事件跟进"
    ws["A24"] = "序号"
    ws["B24"] = "事件等级"
    ws["D24"] = "描述"
    ws["A25"] = 1
    ws["B25"] = "/"
    ws["D25"] = "/"
    ws.merge_cells("D25:E25")

    ws["A26"] = "交接确认"
    return wb, ws


def test_parse_category_sections_detects_all() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        assert [s.name for s in sections] == ["新事件处理", "历史事件跟进"]
        assert sections[0].title_row == 20
        assert sections[0].template_data_row == 22
        assert sections[0].end_row == 22
        assert sections[0].baseline_rows == 1
        assert sections[1].title_row == 23
        assert sections[1].template_data_row == 25
        assert sections[1].end_row == 25
        assert sections[1].baseline_rows == 1
    finally:
        wb.close()


def test_write_category_sections_expand_and_number_rows() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1", "F": "处理中", "I": "甲"}},
                {"cells": {"B": "严重", "D": "事件2", "F": "处理中", "I": "乙"}},
                {"cells": {"B": "重要", "D": "事件3", "F": "已完成", "I": "丙"}},
            ],
            "历史事件跟进": [
                {"cells": {"B": "次要", "D": "跟进1", "I": "丁"}},
                {"cells": {"B": "重要", "D": "跟进2", "I": "戊"}},
            ],
        }
        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        sections_after = parse_category_sections(ws)
        sec_new = next(s for s in sections_after if s.name == "新事件处理")
        sec_hist = next(s for s in sections_after if s.name == "历史事件跟进")

        assert sec_new.end_row - sec_new.template_data_row + 1 == 3
        assert ws[f"A{sec_new.template_data_row}"].value == 1
        assert ws[f"A{sec_new.template_data_row + 1}"].value == 2
        assert ws[f"A{sec_new.template_data_row + 2}"].value == 3
        assert ws[f"B{sec_new.template_data_row}"].value == "紧急"
        assert ws[f"D{sec_new.template_data_row + 2}"].value == "事件3"

        assert sec_hist.end_row - sec_hist.template_data_row + 1 == 2
        assert ws[f"A{sec_hist.template_data_row}"].value == 1
        assert ws[f"A{sec_hist.template_data_row + 1}"].value == 2
        assert ws[f"D{sec_hist.template_data_row}"].value == "跟进1"
        assert ws[f"D{sec_hist.template_data_row + 1}"].value == "跟进2"
    finally:
        wb.close()


def test_style_snapshot_protects_lower_section_title_and_header() -> None:
    wb, ws = _build_ws()
    try:
        red = PatternFill(fill_type="solid", fgColor="FFFF0000")
        green = PatternFill(fill_type="solid", fgColor="FF00FF00")

        ws["A20"].fill = red
        ws["A23"].fill = green

        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1"}},
                {"cells": {"B": "严重", "D": "事件2"}},
                {"cells": {"B": "重要", "D": "事件3"}},
            ],
            "历史事件跟进": [
                {"cells": {"B": "次要", "D": "跟进1"}},
            ],
        }

        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        sections_after = parse_category_sections(ws)
        sec_new = next(s for s in sections_after if s.name == "新事件处理")
        sec_hist = next(s for s in sections_after if s.name == "历史事件跟进")

        # 下方分类被上方扩行后，标题样式和标题文本应保持
        assert ws.cell(row=sec_new.title_row, column=1).fill.fgColor.rgb == "FFFF0000"
        assert ws.cell(row=sec_hist.title_row, column=1).fill.fgColor.rgb == "FF00FF00"
        assert ws.cell(row=sec_hist.title_row, column=1).value == "历史事件跟进"

        # 下方分类结构保持：表头文本仍在，数据区合并仍存在
        merged_refs = {str(x) for x in ws.merged_cells.ranges}
        assert ws[f"D{sec_hist.header_row}"].value == "描述"
        assert f"D{sec_hist.template_data_row}:E{sec_hist.template_data_row}" in merged_refs
    finally:
        wb.close()


def test_insert_upper_section_clears_empty_lower_section_template_values() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1", "F": "处理中"}},
                {"cells": {"B": "严重", "D": "事件2", "F": "处理中"}},
                {"cells": {"B": "重要", "D": "事件3", "F": "已完成"}},
            ],
            # 历史事件跟进不传数据，要求只保留 1 行空白可编辑行，不保留模板示例值。
            "历史事件跟进": [],
        }

        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        sections_after = parse_category_sections(ws)
        sec_hist = next(s for s in sections_after if s.name == "历史事件跟进")

        assert sec_hist.end_row - sec_hist.template_data_row + 1 == 1
        assert ws[f"A{sec_hist.template_data_row}"].value in ("", None)
        assert ws[f"B{sec_hist.template_data_row}"].value in ("", None)
        assert ws[f"D{sec_hist.template_data_row}"].value in ("", None)
        assert ws[f"F{sec_hist.template_data_row}"].value in ("", None)
    finally:
        wb.close()


def test_insert_preserves_handover_confirmation_block() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1"}},
                {"cells": {"B": "严重", "D": "事件2"}},
                {"cells": {"B": "重要", "D": "事件3"}},
            ],
            "历史事件跟进": [],
        }

        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        # 原 A26=交接确认，上方新增2行后应下移到 A28
        assert ws["A28"].value == "交接确认"
        assert ws["A26"].value != "交接确认"
    finally:
        wb.close()


def test_merge_pattern_cloned_for_inserted_rows() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1"}},
                {"cells": {"B": "严重", "D": "事件2"}},
                {"cells": {"B": "重要", "D": "事件3"}},
            ],
            "历史事件跟进": [],
        }

        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        merged_refs = {str(x) for x in ws.merged_cells.ranges}
        assert "D22:E22" in merged_refs
        assert "D23:E23" in merged_refs
        assert "D24:E24" in merged_refs
        assert "F22:G22" in merged_refs
        assert "F23:G23" in merged_refs
        assert "F24:G24" in merged_refs
    finally:
        wb.close()


def test_no_overlapping_merge_after_insert() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1"}},
                {"cells": {"B": "严重", "D": "事件2"}},
                {"cells": {"B": "重要", "D": "事件3"}},
            ],
            "历史事件跟进": [],
        }

        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        ranges = list(ws.merged_cells.ranges)
        for i, left in enumerate(ranges):
            for right in ranges[i + 1 :]:
                row_overlap = not (left.max_row < right.min_row or right.max_row < left.min_row)
                col_overlap = not (left.max_col < right.min_col or right.max_col < left.min_col)
                assert not (row_overlap and col_overlap), f"overlap found: {left} vs {right}"
    finally:
        wb.close()


def test_write_category_sections_shrinks_when_rows_deleted() -> None:
    wb, ws = _build_ws()
    try:
        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        expanded_payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1", "F": "处理中", "I": "甲"}},
                {"cells": {"B": "严重", "D": "事件2", "F": "处理中", "I": "乙"}},
                {"cells": {"B": "重要", "D": "事件3", "F": "已完成", "I": "丙"}},
            ],
            "历史事件跟进": [
                {"cells": {"B": "次要", "D": "跟进1", "I": "丁"}},
            ],
        }
        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=expanded_payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        sections = parse_category_sections(ws)
        snapshots = capture_section_snapshots(ws, sections)
        shrink_payloads = {
            "新事件处理": [
                {"cells": {"B": "紧急", "D": "事件1", "F": "处理中", "I": "甲"}},
            ],
            "历史事件跟进": [
                {"cells": {"B": "次要", "D": "跟进1", "I": "丁"}},
            ],
        }
        write_category_sections(
            ws=ws,
            sections=sections,
            category_payloads=shrink_payloads,
            snapshots=snapshots,
            emit_log=lambda *_: None,
        )

        sections_after = parse_category_sections(ws)
        sec_new = next(s for s in sections_after if s.name == "新事件处理")
        sec_hist = next(s for s in sections_after if s.name == "历史事件跟进")

        assert sec_new.end_row - sec_new.template_data_row + 1 == 1
        assert ws[f"A{sec_new.template_data_row}"].value == 1
        assert ws[f"B{sec_new.template_data_row}"].value == "紧急"
        assert ws[f"D{sec_new.template_data_row}"].value == "事件1"
        assert ws[f"A{sec_hist.title_row}"].value == "历史事件跟进"
        assert ws[f"D{sec_hist.template_data_row}"].value == "跟进1"
        assert ws[f"A{sec_hist.title_row + 3}"].value == "交接确认"
    finally:
        wb.close()
