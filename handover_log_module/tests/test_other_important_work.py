from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from handover_log_module.repository.other_important_work_repository import (
    OtherImportantWorkRepository,
    OtherImportantWorkRow,
)
from handover_log_module.service.other_important_work_payload_builder import (
    OtherImportantWorkPayloadBuilder,
)


def _build_other_work_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A44"] = "其他重要工作记录"
    ws["A45"] = "序号"
    ws["B45"] = "描述"
    ws["F45"] = "完成情况"
    ws["H45"] = "执行人"
    ws["A46"] = 1
    ws["B46"] = "/"
    ws["F46"] = "/"
    ws["H46"] = "/"
    wb.save(path)
    wb.close()


class _FakeOtherWorkRepo:
    def __init__(self, cfg, rows):
        self._cfg = cfg
        self._rows = rows

    def get_config(self):
        return self._cfg

    def list_current_shift_rows(self, **_kwargs):
        return list(self._rows), self._cfg


class _FakeShiftRosterRepo:
    def __init__(self, rows):
        self._rows = rows
        self.calls = []

    def list_engineer_directory(self, **_kwargs):
        self.calls.append(dict(_kwargs))
        return list(self._rows)


class _GroupedOtherWorkRepo(OtherImportantWorkRepository):
    def __init__(self, rows):
        super().__init__({})
        self._rows = rows

    def _load_rows_for_shift(self, *, duty_date, duty_shift, emit_log):  # noqa: ARG002
        cfg = self.get_config()
        return list(self._rows), cfg, {"total": len(self._rows), "in_scope": len(self._rows), "parse_fail": 0}


def test_other_important_work_builder_maps_template_columns_and_preserves_source_order(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="power_notice",
            source_label="上电通告",
            record_id="rec-1",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="上电事项",
            completion_text="进行中",
            specialty_text="电气",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_adjustment",
            source_label="设备调整",
            record_id="rec-2",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 30, 0),
            actual_end_time=datetime(2026, 3, 14, 10, 0, 0),
            description_text="调整事项",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_patrol",
            source_label="设备轮巡",
            record_id="rec-3",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 40, 0),
            actual_end_time=datetime(2026, 3, 14, 10, 5, 0),
            description_text="轮巡事项",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_repair",
            source_label="设备检修",
            record_id="rec-4",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 50, 0),
            actual_end_time=datetime(2026, 3, 14, 10, 10, 0),
            description_text="检修事项",
            completion_text="未完成",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert list(payload.keys()) == ["其他重要工作记录"]
    assert payload["其他重要工作记录"] == [
        {"cells": {"B": "上电事项", "F": "进行中", "H": "汪根尚"}},
        {"cells": {"B": "调整事项", "F": "已完成", "H": "汪根尚"}},
        {"cells": {"B": "轮巡事项", "F": "已完成", "H": "汪根尚"}},
        {"cells": {"B": "检修事项", "F": "未完成", "H": "汪根尚"}},
    ]


def test_other_important_work_grouped_rows_reuse_multi_building_records() -> None:
    rows = [
        OtherImportantWorkRow(
            source_key="power_notice",
            source_label="上电通告",
            record_id="rec-a",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="A事项",
            completion_text="已完成",
            specialty_text="暖通",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_adjustment",
            source_label="设备调整",
            record_id="rec-ac",
            building_values=["A楼", "C楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 10, 0),
            actual_end_time=datetime(2026, 3, 14, 9, 30, 0),
            description_text="AC事项",
            completion_text="进行中",
            specialty_text="暖通",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_repair",
            source_label="设备检修",
            record_id="rec-b",
            building_values=["B楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 40, 0),
            actual_end_time=datetime(2026, 3, 14, 10, 0, 0),
            description_text="B事项",
            completion_text="已完成",
            specialty_text="暖通",
            raw_fields={},
        ),
    ]
    repo = _GroupedOtherWorkRepo(rows)

    grouped, _ = repo.list_current_shift_rows_grouped(
        buildings=["A楼", "C楼"],
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert [row.record_id for row in grouped["A楼"]] == ["rec-a", "rec-ac"]
    assert [row.record_id for row in grouped["C楼"]] == ["rec-ac"]


def test_other_important_work_builder_skips_blank_description_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_blank_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="power_notice",
            source_label="上电通告",
            record_id="rec-keep",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 8, 0, 0),
            actual_end_time=None,
            description_text="保留事项",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_adjustment",
            source_label="设备调整",
            record_id="rec-blank",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 8, 10, 0),
            actual_end_time=None,
            description_text="",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["其他重要工作记录"] == [
        {"cells": {"B": "保留事项", "F": "已完成", "H": "汪根尚"}},
    ]


def test_other_important_work_builder_normalizes_specialty_for_executor_match(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_specialty_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="device_patrol",
            source_label="设备轮巡",
            record_id="rec-normalized",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="暖通轮巡事项",
            completion_text="已完成",
            specialty_text="暖通巡检",
            raw_fields={},
        ),
    ]
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "土建、暖通", "supervisor": "郭克成"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["其他重要工作记录"] == [
        {"cells": {"B": "暖通轮巡事项", "F": "已完成", "H": "郭克成"}},
    ]


def test_other_important_work_builder_uses_shared_fire_supervisor_across_buildings(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_fire_shared_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="device_patrol",
            source_label="设备轮巡",
            record_id="rec-fire-shared",
            building_values=["C楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="C楼消防专项巡检",
            completion_text="已完成",
            specialty_text="消防专项",
            raw_fields={},
        ),
    ]
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "消防、安全", "specialty": "消防、安全", "supervisor": "明志勇"}]
        ),
    )

    payload = builder.build(
        building="C楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["其他重要工作记录"] == [
        {"cells": {"B": "C楼消防专项巡检", "F": "已完成", "H": "明志勇"}},
    ]


def test_other_important_work_builder_dedupes_duplicate_descriptions(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_dedupe_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="device_patrol",
            source_label="设备轮巡",
            record_id="rec-1",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="重复事项",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
        OtherImportantWorkRow(
            source_key="device_repair",
            source_label="设备检修",
            record_id="rec-2",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 10, 0),
            actual_end_time=None,
            description_text=" 重复事项 ",
            completion_text="未完成",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["其他重要工作记录"] == [
        {"cells": {"B": "重复事项", "F": "已完成", "H": "汪根尚"}},
    ]


def test_other_important_work_builder_uses_target_duty_for_engineer_directory(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_context_template.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "description": "B",
                "completion": "F",
                "executor": "H",
            },
        },
    }
    rows = [
        OtherImportantWorkRow(
            source_key="device_patrol",
            source_label="设备轮巡",
            record_id="rec-duty-context",
            building_values=["A楼"],
            actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
            actual_end_time=None,
            description_text="轮巡事项",
            completion_text="已完成",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    fake_shift_repo = _FakeShiftRosterRepo(
        [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
    )
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(cfg, rows),
        shift_roster_repo=fake_shift_repo,
    )

    builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert fake_shift_repo.calls
    assert fake_shift_repo.calls[-1]["duty_date"] == "2026-03-14"
    assert fake_shift_repo.calls[-1]["duty_shift"] == "day"


def test_other_important_work_builder_reuses_preloaded_engineers(tmp_path: Path) -> None:
    template_path = tmp_path / "other_work_preloaded_engineers.xlsx"
    _build_other_work_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"other_important_work": "其他重要工作记录"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "description": ["描述"],
                "completion": ["完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {"description": "B", "completion": "F", "executor": "H"},
        },
    }
    fake_shift_repo = _FakeShiftRosterRepo([])
    builder = OtherImportantWorkPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeOtherWorkRepo(
            cfg,
            [
                OtherImportantWorkRow(
                    source_key="device_patrol",
                    source_label="设备轮巡",
                    record_id="rec-preloaded",
                    building_values=["A楼"],
                    actual_start_time=datetime(2026, 3, 14, 9, 0, 0),
                    actual_end_time=None,
                    description_text="轮巡事项",
                    completion_text="已完成",
                    specialty_text="电气",
                    raw_fields={},
                )
            ],
        ),
        shift_roster_repo=fake_shift_repo,
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        preloaded_engineers=[{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}],
        emit_log=lambda *_args: None,
    )

    assert fake_shift_repo.calls == []
    assert payload["其他重要工作记录"][0]["cells"]["H"] == "汪根尚"
