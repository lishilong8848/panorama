from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from app.config.config_adapter import ensure_v3_config
from app.modules.handover_review.api.routes import _persist_footer_inventory_defaults
from handover_log_module.core.footer_layout import (
    FOOTER_GROUP_TITLE_TEXT,
    FOOTER_SIGNOFF_MARKER,
    FOOTER_TITLE_TEXT,
    find_footer_inventory_layout,
)
from handover_log_module.service.footer_inventory_defaults_service import FooterInventoryDefaultsService


SHEET_NAME = "\u4ea4\u63a5\u73ed\u65e5\u5fd7"
TEMPLATE_CONFIG = Path(__file__).resolve().parents[2] / "config" / "\u8868\u683c\u8ba1\u7b97\u914d\u7f6e.template.json"


def _build_footer_file(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws["A50"] = FOOTER_TITLE_TEXT
    ws.merge_cells("A50:H50")

    ws["A51"] = FOOTER_GROUP_TITLE_TEXT
    ws["B51"] = "\u4ea4\u63a5\u5de5\u5177\u540d\u79f0"
    ws["C51"] = "\u5b58\u653e\u4f4d\u7f6e"
    ws["E51"] = "\u6570\u91cf"
    ws["F51"] = "\u662f\u5426\u5b58\u5728\u635f\u574f"
    ws["G51"] = "\u5176\u4ed6\u8865\u5145\u8bf4\u660e"
    ws["H51"] = "\u6e05\u70b9\u786e\u8ba4\u4eba\uff08\u63a5\u73ed\uff09"
    ws.merge_cells("A51:A55")
    ws.merge_cells("C51:D51")

    data_rows = [
        ("\u503c\u73ed\u624b\u673a", "E-112\u503c\u73ed\u5ba4", "1", "\u5426", "\u65e0", "\u66fe\u5c0f\u5e86"),
        ("\u5bf9\u8bb2\u673a", "E-112\u503c\u73ed\u5ba4", "5", "\u5426", "\u65e0", "\u66fe\u5c0f\u5e86"),
        ("\u94a5\u5319", "E-112\u503c\u73ed\u5ba4\u94a5\u5319\u7bb1", "4", "\u5426", "\u65e0", "\u66fe\u5c0f\u5e86"),
        ("\u5e94\u6025\u5de5\u5177", "E-112\u503c\u73ed\u5ba4", "1", "\u5426", "\u65e0", "\u66fe\u5c0f\u5e86"),
    ]
    for row_no, values in enumerate(data_rows, start=52):
        ws[f"B{row_no}"] = values[0]
        ws[f"C{row_no}"] = values[1]
        ws[f"E{row_no}"] = values[2]
        ws[f"F{row_no}"] = values[3]
        ws[f"G{row_no}"] = values[4]
        ws[f"H{row_no}"] = values[5]
        ws.merge_cells(f"C{row_no}:D{row_no}")

    ws["A56"] = FOOTER_SIGNOFF_MARKER
    ws["C56"] = "\u4ea4\u73ed\u503c\u73ed\u957f\u7b7e\u5b57"
    ws["F56"] = "\u63a5\u73ed\u503c\u73ed\u957f\u7b7e\u5b57"
    ws["F57"] = "\u5ba1\u6838\u4eba\uff1a"
    ws.merge_cells("A56:B56")
    ws.merge_cells("D56:E56")
    ws.merge_cells("G56:H56")

    wb.save(path)
    wb.close()


class _DummyContainer:
    def __init__(self, config_path: Path) -> None:
        raw = json.loads(TEMPLATE_CONFIG.read_text(encoding="utf-8-sig"))
        self.config = ensure_v3_config(raw)
        self.config.setdefault("features", {}).setdefault("alarm_export", {})["window_days"] = 1
        self.config_path = config_path
        self.config_path.write_text(json.dumps(self.config, ensure_ascii=False, indent=2), encoding="utf-8-sig")

    def reload_config(self, settings):
        self.config = settings


def test_footer_inventory_defaults_roundtrip_v3_config() -> None:
    service = FooterInventoryDefaultsService()
    document = {
        "footer_blocks": [
            {
                "type": "inventory_table",
                "rows": [
                    {
                        "row_id": "inventory:1",
                        "cells": {
                            "B": "\u5bf9\u8bb2\u673a",
                            "C": "E-112\u503c\u73ed\u5ba4",
                            "D": "ignored",
                            "E": "5",
                            "F": "\u5426",
                            "G": "\u65e0",
                            "H": "\u5f20\u4e09",
                            "I": "ignored",
                        },
                    }
                ],
            }
        ]
    }

    updated = service.set_building_defaults(
        ensure_v3_config({}),
        "A\u697c",
        service.extract_rows_from_document(document),
    )
    rows = service.get_building_defaults(updated, "A\u697c")

    expected = [
        {
            "cells": {
                "B": "\u5bf9\u8bb2\u673a",
                "C": "E-112\u503c\u73ed\u5ba4",
                "E": "5",
                "F": "\u5426",
                "G": "\u65e0",
                "H": "\u5f20\u4e09",
            }
        }
    ]

    assert rows == expected
    assert (
        updated["features"]["handover_log"]["review_ui"]["footer_inventory_defaults_by_building"]["A\u697c"]["rows"]
        == expected
    )


def test_apply_building_defaults_to_output_overwrites_template_footer(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_footer_file(output_file)

    service = FooterInventoryDefaultsService()
    applied_rows = service.apply_building_defaults_to_output(
        config={
            "review_ui": {
                "footer_inventory_defaults_by_building": {
                    "A\u697c": {
                        "rows": [
                            {"cells": {"B": "\u503c\u73ed\u624b\u673a", "C": "A\u697c\u503c\u73ed\u5ba4", "E": "1", "F": "\u5426", "G": "\u65e0", "H": "\u7532"}},
                            {"cells": {"B": "\u5bf9\u8bb2\u673a", "C": "A\u697c\u503c\u73ed\u5ba4", "E": "3", "F": "\u5426", "G": "\u5907\u7528", "H": "\u4e59"}},
                            {"cells": {"B": "\u95e8\u7981\u5361", "C": "A\u697c\u95e8\u5c97", "E": "2", "F": "\u5426", "G": "", "H": "\u4e19"}},
                            {"cells": {"B": "\u94a5\u5319", "C": "A\u697c\u914d\u7535\u95f4", "E": "4", "F": "\u5426", "G": "", "H": "\u4e01"}},
                            {"cells": {"B": "\u5e94\u6025\u706f", "C": "A\u697c\u4ed3\u5e93", "E": "2", "F": "\u662f", "G": "\u9700\u8865", "H": "\u620a"}},
                        ]
                    }
                }
            }
        },
        building="A\u697c",
        output_file=output_file,
        sheet_name=SHEET_NAME,
        emit_log=lambda *_: None,
    )

    assert applied_rows == 5

    wb = openpyxl.load_workbook(output_file)
    try:
        ws = wb[SHEET_NAME]
        layout = find_footer_inventory_layout(ws)
        assert layout is not None
        assert layout.data_end_row == 56
        assert ws.max_row == layout.last_row
        assert ws["B56"].value == "\u5e94\u6025\u706f"
        assert ws["C56"].value == "A\u697c\u4ed3\u5e93"
        assert ws["F56"].value == "\u662f"
        assert ws["A57"].value == FOOTER_SIGNOFF_MARKER
        merged = {str(item) for item in ws.merged_cells.ranges}
        assert "A51:A56" in merged
        assert "C56:D56" in merged
    finally:
        wb.close()


def test_persist_footer_inventory_defaults_updates_main_config_file(tmp_path: Path) -> None:
    config_path = tmp_path / "config.json"
    container = _DummyContainer(config_path)
    document = {
        "footer_blocks": [
            {
                "id": "handover_inventory_table",
                "type": "inventory_table",
                "rows": [
                    {"cells": {"B": "\u5bf9\u8bb2\u673a", "C": "B\u697c\u503c\u73ed\u5ba4", "E": "5", "F": "\u5426", "G": "\u65e0", "H": "\u674e\u56db"}}
                ],
            },
            {
                "id": "handover_signoff_block",
                "type": "readonly_grid",
                "rows": [{"cells": [{"column": "A", "value": FOOTER_SIGNOFF_MARKER, "colspan": 2}]}],
            },
        ]
    }

    persisted_rows = _persist_footer_inventory_defaults(container, building="B\u697c", document=document)

    expected_rows = [
        {
            "cells": {
                "B": "\u5bf9\u8bb2\u673a",
                "C": "B\u697c\u503c\u73ed\u5ba4",
                "E": "5",
                "F": "\u5426",
                "G": "\u65e0",
                "H": "\u674e\u56db",
            }
        }
    ]

    assert persisted_rows == 1
    saved = json.loads(config_path.read_text(encoding="utf-8-sig"))
    assert (
        saved["features"]["handover_log"]["review_ui"]["footer_inventory_defaults_by_building"]["B\u697c"]["rows"]
        == expected_rows
    )
    assert (
        container.config["features"]["handover_log"]["review_ui"]["footer_inventory_defaults_by_building"]["B\u697c"]["rows"]
        == expected_rows
    )
