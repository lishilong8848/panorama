from __future__ import annotations

import copy
import threading
import time
from pathlib import Path

import openpyxl

from handover_log_module.service.review_document_state_service import (
    ReviewDocumentStateConflictError,
    ReviewDocumentStateService,
)


def _config(tmp_path: Path) -> dict:
    return {
        "_global_paths": {"runtime_state_root": str(tmp_path / ".runtime")},
        "template": {"sheet_name": "交接班日志"},
        "review_ui": {
            "fixed_cells": {
                "header_basic": ["A1"],
                "cabinet_power_info": ["B13", "D13", "F13", "H13"],
            }
        },
    }


def _build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A1"] = "原标题"
    ws["B13"] = "旧规划"
    ws["D13"] = "旧上电"
    ws["F13"] = "旧本班上电"
    ws["H13"] = "旧本班下电"
    wb.save(path)
    wb.close()


def _session(path: Path, *, revision: int = 1) -> dict:
    return {
        "session_id": "A楼|2026-04-15|day",
        "building": "A楼",
        "duty_date": "2026-04-15",
        "duty_shift": "day",
        "batch_key": "2026-04-15|day",
        "revision": revision,
        "output_file": str(path),
    }


def _set_fixed_value(document: dict, cell: str, value: str) -> None:
    for block in document.get("fixed_blocks", []):
        for field in block.get("fields", []):
            if field.get("cell") == cell:
                field["value"] = value
                return


class _SlowWorkbookWriter:
    def __init__(self, delay_sec: float = 0.03) -> None:
        self.delay_sec = delay_sec
        self.active = 0
        self.max_active = 0
        self._lock = threading.Lock()

    def write(self, *, output_file, document, dirty_regions=None):  # noqa: ANN001
        del dirty_regions
        with self._lock:
            self.active += 1
            self.max_active = max(self.max_active, self.active)
        try:
            time.sleep(self.delay_sec)
            value = ""
            for block in document.get("fixed_blocks", []):
                for field in block.get("fields", []):
                    if field.get("cell") == "B13":
                        value = str(field.get("value", "") or "")
                        break
            wb = openpyxl.load_workbook(output_file)
            try:
                ws = wb["交接班日志"]
                ws["B13"] = value
                wb.save(output_file)
            finally:
                wb.close()
        finally:
            with self._lock:
                self.active -= 1


def test_concurrent_same_revision_save_allows_one_success_and_one_conflict(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    service = ReviewDocumentStateService(_config(tmp_path), emit_log=lambda *_: None)
    session = _session(output_file)
    document, _loaded_session = service.load_document(session)

    barrier = threading.Barrier(3)
    results: list[tuple[str, int | str]] = []
    result_lock = threading.Lock()

    def _worker(value: str) -> None:
        local_document = copy.deepcopy(document)
        _set_fixed_value(local_document, "B13", value)
        barrier.wait()
        try:
            state, _previous = service.save_document(
                session=session,
                document=local_document,
                base_revision=1,
                dirty_regions={"fixed_blocks": True},
            )
            outcome: tuple[str, int | str] = ("ok", int(state["revision"]))
        except ReviewDocumentStateConflictError as exc:
            outcome = ("conflict", str(exc))
        with result_lock:
            results.append(outcome)

    threads = [
        threading.Thread(target=_worker, args=("并发-A",)),
        threading.Thread(target=_worker, args=("并发-B",)),
    ]
    for thread in threads:
        thread.start()
    barrier.wait()
    for thread in threads:
        thread.join(timeout=5)

    statuses = sorted(item[0] for item in results)
    assert statuses == ["conflict", "ok"]
    store_state = service._store("A楼").get_document(session["session_id"])
    assert store_state is not None
    assert store_state["revision"] == 2


def test_background_sync_worker_eventually_reaches_latest_revision_under_save_burst(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    writer = _SlowWorkbookWriter(delay_sec=0.02)
    service = ReviewDocumentStateService(_config(tmp_path), writer=writer, emit_log=lambda *_: None)
    session = _session(output_file)
    document, _loaded_session = service.load_document(session)

    latest_revision = 1
    for revision in range(2, 12):
        latest_revision = revision
        local_document = copy.deepcopy(document)
        _set_fixed_value(local_document, "B13", f"rev-{revision}")
        state, _previous = service.save_document(
            session={**session, "revision": revision - 1},
            document=local_document,
            base_revision=revision - 1,
            dirty_regions={"fixed_blocks": True},
        )
        document = copy.deepcopy(local_document)
        sync = service.enqueue_excel_sync(
            {**session, "revision": state["revision"]},
            target_revision=state["revision"],
        )
        assert sync["status"] in {"pending", "syncing", "synced"}

    deadline = time.time() + 8
    store = service._store("A楼")
    last_sync = store.get_sync_state(session["session_id"])
    while time.time() < deadline:
        last_sync = store.get_sync_state(session["session_id"])
        if last_sync["status"] == "synced" and last_sync["synced_revision"] == latest_revision:
            break
        time.sleep(0.05)

    assert last_sync["status"] == "synced"
    assert last_sync["synced_revision"] == latest_revision
    wb = openpyxl.load_workbook(output_file)
    try:
        assert wb["交接班日志"]["B13"].value == f"rev-{latest_revision}"
    finally:
        wb.close()


def test_force_sync_serializes_writer_access_for_same_output_file(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    _build_workbook(output_file)
    writer = _SlowWorkbookWriter(delay_sec=0.05)
    service = ReviewDocumentStateService(_config(tmp_path), writer=writer, emit_log=lambda *_: None)
    session = _session(output_file)
    document, _loaded_session = service.load_document(session)
    _set_fixed_value(document, "B13", "最新值")
    state, _previous = service.save_document(
        session=session,
        document=document,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )

    sync_errors: list[Exception] = []
    barrier = threading.Barrier(3)

    def _run_force_sync() -> None:
        barrier.wait()
        try:
            service.force_sync_session(
                building="A楼",
                session_id=session["session_id"],
                target_revision=state["revision"],
                reason="concurrency-test",
                reconcile_sync_job=False,
            )
        except Exception as exc:  # noqa: BLE001
            sync_errors.append(exc)

    threads = [threading.Thread(target=_run_force_sync), threading.Thread(target=_run_force_sync)]
    for thread in threads:
        thread.start()
    barrier.wait()
    for thread in threads:
        thread.join(timeout=5)

    assert sync_errors == []
    assert writer.max_active == 1


def test_force_sync_allows_parallel_writer_access_for_distinct_buildings(tmp_path: Path) -> None:
    output_a = tmp_path / "handover-a.xlsx"
    output_b = tmp_path / "handover-b.xlsx"
    _build_workbook(output_a)
    _build_workbook(output_b)
    writer = _SlowWorkbookWriter(delay_sec=0.05)
    service = ReviewDocumentStateService(_config(tmp_path), writer=writer, emit_log=lambda *_: None)

    session_a = _session(output_a)
    session_b = {
        **_session(output_b),
        "session_id": "B楼|2026-04-15|day",
        "building": "B楼",
    }

    document_a, _loaded_session_a = service.load_document(session_a)
    document_b, _loaded_session_b = service.load_document(session_b)
    _set_fixed_value(document_a, "B13", "A最新值")
    _set_fixed_value(document_b, "B13", "B最新值")

    state_a, _previous_a = service.save_document(
        session=session_a,
        document=document_a,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    state_b, _previous_b = service.save_document(
        session=session_b,
        document=document_b,
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )

    sync_errors: list[Exception] = []
    barrier = threading.Barrier(3)

    def _run_force_sync(building: str, session_id: str, target_revision: int) -> None:
        barrier.wait()
        try:
            service.force_sync_session(
                building=building,
                session_id=session_id,
                target_revision=target_revision,
                reason="parallel-buildings-test",
                reconcile_sync_job=False,
            )
        except Exception as exc:  # noqa: BLE001
            sync_errors.append(exc)

    threads = [
        threading.Thread(target=_run_force_sync, args=("A楼", session_a["session_id"], state_a["revision"])),
        threading.Thread(target=_run_force_sync, args=("B楼", session_b["session_id"], state_b["revision"])),
    ]
    for thread in threads:
        thread.start()
    barrier.wait()
    for thread in threads:
        thread.join(timeout=5)

    assert sync_errors == []
    assert writer.max_active >= 2
