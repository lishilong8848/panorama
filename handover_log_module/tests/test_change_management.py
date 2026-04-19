from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from handover_log_module.repository.change_management_repository import ChangeManagementRow
from handover_log_module.service.change_management_payload_builder import ChangeManagementPayloadBuilder


def _build_change_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A35"] = "变更管理"
    ws["A36"] = "序号"
    ws["B36"] = "变更等级"
    ws["D36"] = "描述"
    ws["E36"] = "作业时间段"
    ws["H36"] = "执行人"
    ws["A37"] = 1
    ws["B37"] = "/"
    ws["D37"] = "/"
    ws["E37"] = "/"
    ws["H37"] = "/"
    wb.save(path)
    wb.close()


class _FakeChangeRepo:
    def __init__(self, cfg, rows):
        self._cfg = cfg
        self._rows = rows

    def list_current_shift_rows(self, **_kwargs):
        return list(self._rows), self._cfg


class _FakeShiftRosterRepo:
    def __init__(self, rows):
        self._rows = rows
        self.calls = []

    def list_engineer_directory(self, **_kwargs):
        self.calls.append(dict(_kwargs))
        return list(self._rows)


def test_change_management_builder_skips_blank_description_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "change_level": "B",
                "work_window": "E",
                "description": "D",
                "executor": "H",
            },
        },
        "work_window_text": {
            "day_anchor": "08:00:00",
            "day_default_end": "18:30:00",
            "night_anchor": "18:00:00",
            "night_default_end_next_day": "08:00:00",
        },
    }
    rows = [
        ChangeManagementRow(
            record_id="rec-blank",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 09:10:00",
            description="",
            specialty_text="电气",
            raw_fields={},
        ),
        ChangeManagementRow(
            record_id="rec-keep",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 10, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 10:10:00",
            description="变更描述",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["变更管理"] == [
        {"cells": {"B": "低", "D": "变更描述", "E": "10:10-18:30", "H": "汪根尚"}}
    ]


def test_change_management_builder_normalizes_specialty_for_executor_match(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template_specialty.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "change_level": "B",
                "work_window": "E",
                "description": "D",
                "executor": "H",
            },
        },
        "work_window_text": {
            "day_anchor": "08:00:00",
            "day_default_end": "18:30:00",
            "night_anchor": "18:00:00",
            "night_default_end_next_day": "08:00:00",
        },
    }
    rows = [
        ChangeManagementRow(
            record_id="rec-normalized",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 09:10:00",
            description="消防系统变更",
            specialty_text="消防值守",
            raw_fields={},
        ),
    ]
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "消防", "supervisor": "高荣"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["变更管理"] == [
        {"cells": {"B": "低", "D": "消防系统变更", "E": "09:10-18:30", "H": "高荣"}}
    ]


def test_change_management_builder_uses_shared_fire_supervisor_across_buildings(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template_fire_shared.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "change_level": "B",
                "work_window": "E",
                "description": "D",
                "executor": "H",
            },
        },
        "work_window_text": {
            "day_anchor": "08:00:00",
            "day_default_end": "18:30:00",
            "night_anchor": "18:00:00",
            "night_default_end_next_day": "08:00:00",
        },
    }
    rows = [
        ChangeManagementRow(
            record_id="rec-fire-shared",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 09:10:00",
            description="消防系统联动测试",
            specialty_text="消防",
            raw_fields={},
        ),
    ]
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "消防、安全", "specialty": "消防、安全", "supervisor": "明志勇"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["变更管理"] == [
        {"cells": {"B": "低", "D": "消防系统联动测试", "E": "09:10-18:30", "H": "明志勇"}}
    ]


def test_change_management_builder_dedupes_duplicate_descriptions(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template_dedupe.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "change_level": "B",
                "work_window": "E",
                "description": "D",
                "executor": "H",
            },
        },
        "work_window_text": {
            "day_anchor": "08:00:00",
            "day_default_end": "18:30:00",
            "night_anchor": "18:00:00",
            "night_default_end_next_day": "08:00:00",
        },
    }
    rows = [
        ChangeManagementRow(
            record_id="rec-1",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 09:10:00",
            description="重复变更",
            specialty_text="电气",
            raw_fields={},
        ),
        ChangeManagementRow(
            record_id="rec-2",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 10, 0, 0),
            end_time=None,
            change_level="高",
            process_updates_text="2026-03-14 10:10:00",
            description=" 重复变更 ",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["变更管理"] == [
        {"cells": {"B": "低", "D": "重复变更", "E": "09:10-18:30", "H": "汪根尚"}}
    ]


def test_change_management_builder_uses_target_duty_for_engineer_directory(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template_context.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "change_level": "B",
                "work_window": "E",
                "description": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        ChangeManagementRow(
            record_id="rec-duty-context",
            building_values=["A楼"],
            start_time=datetime(2026, 3, 14, 9, 0, 0),
            end_time=None,
            change_level="低",
            process_updates_text="2026-03-14 09:10:00",
            description="变更描述",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    fake_shift_repo = _FakeShiftRosterRepo(
        [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
    )
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(cfg, rows),
        shift_roster_repo=fake_shift_repo,
    )

    builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert fake_shift_repo.calls
    assert fake_shift_repo.calls[-1]["duty_date"] == "2026-03-14"
    assert fake_shift_repo.calls[-1]["duty_shift"] == "day"


def test_change_management_builder_reuses_preloaded_engineers(tmp_path: Path) -> None:
    template_path = tmp_path / "change_template_preloaded_engineers.xlsx"
    _build_change_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"change_management": "变更管理"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "change_level": ["变更等级"],
                "work_window": ["作业时间段"],
                "description": ["描述"],
                "executor": ["执行人"],
            },
            "fallback_cols": {"change_level": "B", "work_window": "E", "description": "D", "executor": "H"},
        },
    }
    fake_shift_repo = _FakeShiftRosterRepo([])
    builder = ChangeManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeChangeRepo(
            cfg,
            [
                ChangeManagementRow(
                    record_id="rec-preloaded",
                    building_values=["A楼"],
                    start_time=datetime(2026, 3, 14, 9, 0, 0),
                    end_time=None,
                    change_level="低",
                    process_updates_text="2026-03-14 09:10:00",
                    description="预取执行人",
                    specialty_text="电气",
                    raw_fields={},
                )
            ],
        ),
        shift_roster_repo=fake_shift_repo,
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        preloaded_engineers=[{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}],
        emit_log=lambda *_args: None,
    )

    assert fake_shift_repo.calls == []
    assert payload["变更管理"][0]["cells"]["H"] == "汪根尚"
