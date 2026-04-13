from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from handover_log_module.repository.maintenance_management_repository import (
    MaintenanceManagementRepository,
    MaintenanceManagementRow,
)
from handover_log_module.service.maintenance_management_payload_builder import (
    MaintenanceManagementPayloadBuilder,
)


def _build_maintenance_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交接班日志"
    ws["A32"] = "维护管理"
    ws["A33"] = "序号"
    ws["B33"] = "维护总项"
    ws["C33"] = "维护执行方"
    ws["D33"] = "维护完成情况"
    ws["H33"] = "执行人"
    ws["A34"] = 1
    ws["B34"] = "/"
    ws["C34"] = "/"
    ws["D34"] = "/"
    ws["H34"] = "/"
    ws.merge_cells("D34:G34")
    wb.save(path)
    wb.close()


class _FakeMaintenanceRepo:
    def __init__(self, cfg, rows):
        self._cfg = cfg
        self._rows = rows

    def get_config(self):
        return self._cfg

    def list_current_shift_rows(self, **_kwargs):
        return list(self._rows), self._cfg


class _FakeShiftRosterRepo:
    def __init__(self, rows):
        self._rows = rows
        self.calls = []

    def list_engineer_directory(self, **_kwargs):
        self.calls.append(dict(_kwargs))
        return list(self._rows)


class _GroupedMaintenanceRepo(MaintenanceManagementRepository):
    def __init__(self, rows):
        super().__init__({})
        self._rows = rows

    def _load_rows_for_shift(self, *, duty_date, duty_shift, emit_log):  # noqa: ARG002
        cfg = self.get_config()
        return list(self._rows), cfg, {"total": len(self._rows), "in_shift": len(self._rows), "parse_fail": 0}


def test_maintenance_management_builder_maps_template_columns_and_vendor_type(tmp_path: Path) -> None:
    template_path = tmp_path / "maintenance_template.xlsx"
    _build_maintenance_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"maintenance_management": "维护管理"},
        "fixed_values": {"vendor_internal": "自维", "vendor_external": "厂维", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "maintenance_item": ["维护总项"],
                "maintenance_party": ["维护执行方"],
                "completion": ["维护完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "maintenance_item": "B",
                "maintenance_party": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        MaintenanceManagementRow(
            record_id="rec-factory",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A楼厂家巡检维护",
            specialty_text="电气",
            raw_fields={},
        ),
        MaintenanceManagementRow(
            record_id="rec-self",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 10, 0, 0),
            item_text="A楼日常维护",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = MaintenanceManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeMaintenanceRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert list(payload.keys()) == ["维护管理"]
    assert payload["维护管理"] == [
        {"cells": {"B": "A楼厂家巡检维护", "C": "厂维", "D": "已完成", "H": "汪根尚"}},
        {"cells": {"B": "A楼日常维护", "C": "自维", "D": "已完成", "H": "汪根尚"}},
    ]


def test_maintenance_management_grouped_rows_reuse_multi_building_records() -> None:
    rows = [
        MaintenanceManagementRow(
            record_id="rec-a",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A项目",
            specialty_text="暖通",
            raw_fields={},
        ),
        MaintenanceManagementRow(
            record_id="rec-ac",
            building_values=["A楼", "C楼"],
            updated_time=datetime(2026, 3, 14, 9, 30, 0),
            item_text="AC项目",
            specialty_text="暖通",
            raw_fields={},
        ),
        MaintenanceManagementRow(
            record_id="rec-b",
            building_values=["B楼"],
            updated_time=datetime(2026, 3, 14, 10, 0, 0),
            item_text="B项目",
            specialty_text="暖通",
            raw_fields={},
        ),
    ]
    repo = _GroupedMaintenanceRepo(rows)

    grouped, _ = repo.list_current_shift_rows_grouped(
        buildings=["A楼", "C楼"],
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert [row.record_id for row in grouped["A楼"]] == ["rec-a", "rec-ac"]
    assert [row.record_id for row in grouped["C楼"]] == ["rec-ac"]


def test_maintenance_management_builder_skips_blank_item_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "maintenance_template_blank.xlsx"
    _build_maintenance_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"maintenance_management": "维护管理"},
        "fixed_values": {"vendor_internal": "自维", "vendor_external": "厂维", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "maintenance_item": ["维护总项"],
                "maintenance_party": ["维护执行方"],
                "completion": ["维护完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "maintenance_item": "B",
                "maintenance_party": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        MaintenanceManagementRow(
            record_id="rec-blank",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 8, 0, 0),
            item_text="",
            specialty_text="电气",
            raw_fields={},
        ),
        MaintenanceManagementRow(
            record_id="rec-keep",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A楼日常维护",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = MaintenanceManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeMaintenanceRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["维护管理"] == [
        {"cells": {"B": "A楼日常维护", "C": "自维", "D": "已完成", "H": "汪根尚"}}
    ]


def test_maintenance_management_builder_normalizes_specialty_for_executor_match(tmp_path: Path) -> None:
    template_path = tmp_path / "maintenance_template_specialty.xlsx"
    _build_maintenance_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"maintenance_management": "维护管理"},
        "fixed_values": {"vendor_internal": "自维", "vendor_external": "厂维", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "maintenance_item": ["维护总项"],
                "maintenance_party": ["维护执行方"],
                "completion": ["维护完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "maintenance_item": "B",
                "maintenance_party": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        MaintenanceManagementRow(
            record_id="rec-normalized",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A楼配电维护",
            specialty_text="配电",
            raw_fields={},
        ),
    ]
    builder = MaintenanceManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeMaintenanceRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["维护管理"] == [
        {"cells": {"B": "A楼配电维护", "C": "自维", "D": "已完成", "H": "汪根尚"}}
    ]


def test_maintenance_management_builder_uses_target_duty_for_engineer_directory(tmp_path: Path) -> None:
    template_path = tmp_path / "maintenance_template_context.xlsx"
    _build_maintenance_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"maintenance_management": "维护管理"},
        "fixed_values": {"vendor_internal": "自维", "vendor_external": "厂维", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "maintenance_item": ["维护总项"],
                "maintenance_party": ["维护执行方"],
                "completion": ["维护完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "maintenance_item": "B",
                "maintenance_party": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        MaintenanceManagementRow(
            record_id="rec-duty-context",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A楼日常维护",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    fake_shift_repo = _FakeShiftRosterRepo(
        [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
    )
    builder = MaintenanceManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeMaintenanceRepo(cfg, rows),
        shift_roster_repo=fake_shift_repo,
    )

    builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert fake_shift_repo.calls
    assert fake_shift_repo.calls[-1]["duty_date"] == "2026-03-14"
    assert fake_shift_repo.calls[-1]["duty_shift"] == "day"


def test_maintenance_management_builder_dedupes_duplicate_items(tmp_path: Path) -> None:
    template_path = tmp_path / "maintenance_template_dedupe.xlsx"
    _build_maintenance_template(template_path)
    cfg = {
        "enabled": True,
        "sections": {"maintenance_management": "维护管理"},
        "fixed_values": {"vendor_internal": "自维", "vendor_external": "厂维", "completion": "已完成"},
        "column_mapping": {
            "resolve_by_header": True,
            "header_alias": {
                "maintenance_item": ["维护总项"],
                "maintenance_party": ["维护执行方"],
                "completion": ["维护完成情况"],
                "executor": ["执行人"],
            },
            "fallback_cols": {
                "maintenance_item": "B",
                "maintenance_party": "C",
                "completion": "D",
                "executor": "H",
            },
        },
    }
    rows = [
        MaintenanceManagementRow(
            record_id="rec-1",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 9, 0, 0),
            item_text="A楼日常维护",
            specialty_text="电气",
            raw_fields={},
        ),
        MaintenanceManagementRow(
            record_id="rec-2",
            building_values=["A楼"],
            updated_time=datetime(2026, 3, 14, 10, 0, 0),
            item_text=" A楼日常维护 ",
            specialty_text="电气",
            raw_fields={},
        ),
    ]
    builder = MaintenanceManagementPayloadBuilder(
        {"template": {"source_path": str(template_path), "sheet_name": "交接班日志"}},
        repository=_FakeMaintenanceRepo(cfg, rows),
        shift_roster_repo=_FakeShiftRosterRepo(
            [{"building": "A楼", "specialty": "电气", "supervisor": "汪根尚"}]
        ),
    )

    payload = builder.build(
        building="A楼",
        duty_date="2026-03-14",
        duty_shift="day",
        emit_log=lambda *_args: None,
    )

    assert payload["维护管理"] == [
        {"cells": {"B": "A楼日常维护", "C": "自维", "D": "已完成", "H": "汪根尚"}}
    ]
