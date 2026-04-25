from __future__ import annotations

import hashlib
import math
import os
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from PIL import Image, ImageDraw, ImageFont

from app.shared.utils.atomic_file import atomic_write_bytes, validate_image_file
from app.shared.utils.runtime_temp_workspace import resolve_runtime_state_root
from handover_log_module.service.review_access_snapshot_service import materialize_review_access_snapshot
from handover_log_module.service.review_link_delivery_service import ReviewLinkDeliveryService
from handover_log_module.service.review_session_service import ReviewSessionService


_CAPACITY_IMAGE_DELIVERY_LOCK = threading.RLock()
_CAPACITY_IMAGE_EXCEL_COPY_LOCK = threading.RLock()
_CAPACITY_IMAGE_DELIVERY_RUNNING_STATUSES = {"sending"}
_CAPACITY_IMAGE_EXCEL_LOCK_TIMEOUT_SEC = 120.0
_CAPACITY_IMAGE_LOCK_TIMEOUT_ERROR = "容量图片截图繁忙，请稍后重试"
_EXCEL_COM_PROGID = "Excel.Application"
_WPS_EXCEL_COM_PROGIDS = ("KET.Application", "KET.Application.9")


def _now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_name(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", text, flags=re.UNICODE)
    return text.strip("._") or "unknown"


class CapacityReportImageRenderer:
    def __init__(self, handover_cfg: Dict[str, Any]) -> None:
        self.handover_cfg = handover_cfg if isinstance(handover_cfg, dict) else {}

    def _runtime_root(self) -> Path:
        return resolve_runtime_state_root(
            runtime_config={"paths": self.handover_cfg.get("_global_paths", {})},
            app_dir=Path(__file__).resolve().parents[2],
        )

    def _configured_sheet_name(self) -> str:
        capacity_cfg = self.handover_cfg.get("capacity_report", {})
        if not isinstance(capacity_cfg, dict):
            return ""
        template_cfg = capacity_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return ""
        return _text(template_cfg.get("sheet_name"))

    @staticmethod
    def _registry_default_value(path: str) -> str:
        if os.name != "nt":
            return ""
        try:
            import winreg

            with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, path) as key:
                value, _ = winreg.QueryValueEx(key, "")
        except Exception:
            return ""
        return str(value or "").strip()

    @classmethod
    def _registered_com_clsid(cls, progid: str) -> str:
        return cls._registry_default_value(f"{str(progid or '').strip()}\\CLSID")

    @classmethod
    def _registered_com_local_server(cls, clsid: str) -> str:
        normalized = str(clsid or "").strip()
        if not normalized:
            return ""
        return cls._registry_default_value(f"CLSID\\{normalized}\\LocalServer32")

    @classmethod
    def _excel_dispatch_progids(cls) -> List[str]:
        candidates = [_EXCEL_COM_PROGID, *_WPS_EXCEL_COM_PROGIDS]
        excel_clsid = cls._registered_com_clsid(_EXCEL_COM_PROGID)
        wps_registered = any(cls._registered_com_clsid(progid) for progid in _WPS_EXCEL_COM_PROGIDS)
        if wps_registered and excel_clsid and not cls._registered_com_local_server(excel_clsid):
            candidates = [*_WPS_EXCEL_COM_PROGIDS, _EXCEL_COM_PROGID]
        unique: List[str] = []
        for progid in candidates:
            text = str(progid or "").strip()
            if text and text not in unique:
                unique.append(text)
        return unique

    @classmethod
    def _dispatch_excel_application(cls, win32com_client: Any, *, emit_log: Callable[[str], None] | None = None) -> Tuple[Any, str]:
        errors: List[str] = []
        for progid in cls._excel_dispatch_progids():
            try:
                excel = win32com_client.DispatchEx(progid)
                if emit_log and progid != _EXCEL_COM_PROGID:
                    emit_log(f"[交接班][容量表图片发送] 已使用兼容表格COM启动 progid={progid}")
                return excel, progid
            except Exception as exc:  # noqa: BLE001
                detail = f"{progid}: {exc}"
                errors.append(detail)
                if emit_log:
                    emit_log(
                        "[交接班][容量表图片发送] 表格COM启动失败，尝试下一个兼容ProgID "
                        f"progid={progid}, error={exc}"
                    )
        raise RuntimeError("Excel/WPS COM启动失败: " + "; ".join(errors))

    def _cache_path(
        self,
        *,
        source_path: Path,
        building: str,
        duty_date: str,
        duty_shift: str,
        session_id: str,
        cache_signature: str = "",
    ) -> Path:
        stat = source_path.stat()
        signature_text = str(cache_signature or "").strip()
        if signature_text:
            source_sig = hashlib.sha1(
                f"v7|{signature_text}".encode("utf-8", errors="ignore")
            ).hexdigest()[:16]
        else:
            source_sig = hashlib.sha1(
                f"v6|{source_path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}".encode("utf-8", errors="ignore")
            ).hexdigest()[:12]
        session_sig = hashlib.sha1(str(session_id or "").encode("utf-8", errors="ignore")).hexdigest()[:10]
        batch_dir = self._runtime_root() / "handover" / "capacity_report_images" / _safe_name(f"{duty_date}--{duty_shift}")
        return batch_dir / f"{_safe_name(building)}_{session_sig}_{source_sig}.png"

    def render_to_image(
        self,
        *,
        source_file: str,
        building: str,
        duty_date: str,
        duty_shift: str,
        session_id: str,
        force_refresh: bool = False,
        allow_fallback: bool = True,
        cache_signature: str = "",
        render_metadata: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] | None = None,
    ) -> Path:
        source_path = Path(str(source_file or "").strip())
        if not source_path.exists() or not source_path.is_file():
            raise FileNotFoundError(f"交接班容量报表文件不存在: {source_path}")
        metadata = render_metadata if isinstance(render_metadata, dict) else None
        if metadata is not None:
            metadata.update(
                {
                    "cache_hit": False,
                    "cache_signature": str(cache_signature or "").strip(),
                    "generated_at": "",
                    "image_size": 0,
                }
            )
        output_path = self._cache_path(
            source_path=source_path,
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            session_id=session_id,
            cache_signature=cache_signature,
        )
        if metadata is not None:
            metadata["image_path"] = str(output_path)
        if force_refresh:
            attempt_sig = datetime.now().strftime("%Y%m%d%H%M%S%f")
            output_path = output_path.with_name(f"{output_path.stem}_{attempt_sig}{output_path.suffix}")
            if metadata is not None:
                metadata["image_path"] = str(output_path)
        if output_path.exists() and output_path.is_file():
            if not force_refresh:
                try:
                    validate_image_file(output_path)
                except Exception as exc:
                    if emit_log:
                        emit_log(
                            "[交接班][容量表图片发送] 容量表图片缓存校验失败，准备重新截图 "
                            f"building={building}, session_id={session_id}, image={output_path}, error={exc}"
                        )
                    try:
                        output_path.unlink()
                    except FileNotFoundError:
                        pass
                else:
                    if metadata is not None:
                        metadata.update(
                            {
                                "cache_hit": True,
                                "image_path": str(output_path),
                                "image_size": output_path.stat().st_size,
                            }
                        )
                    if emit_log:
                        emit_log(
                            "[交接班][容量表图片发送] 命中容量表图片缓存 "
                            f"building={building}, session_id={session_id}, image={output_path}, "
                            f"size={output_path.stat().st_size}, cache_signature={str(cache_signature or '').strip() or '-'}"
                        )
                    return output_path
            else:
                try:
                    output_path.unlink()
                    if emit_log:
                        emit_log(
                            "[交接班][容量表图片发送] 已清理旧容量表图片缓存，准备重新截图 "
                            f"building={building}, session_id={session_id}, image={output_path}"
                        )
                except FileNotFoundError:
                    pass
                except Exception as exc:
                    if emit_log:
                        emit_log(
                            "[交接班][容量表图片发送] 旧容量表图片缓存清理失败，将覆盖写入 "
                            f"building={building}, session_id={session_id}, image={output_path}, error={exc}"
                        )
        if output_path.exists() and output_path.is_file() and force_refresh:
            try:
                output_path.unlink()
            except FileNotFoundError:
                pass
        if output_path.exists() and output_path.is_file():
            # Cache file was invalid and could not be removed; continue and overwrite.
            pass
        if emit_log:
            emit_log(
                "[交接班][容量表图片发送] 开始生成容量表图片 "
                f"building={building}, session_id={session_id}, source={source_path}, target={output_path}, "
                f"cache_signature={str(cache_signature or '').strip() or '-'}"
            )
        if self._render_with_excel_copy_picture(source_path=source_path, output_path=output_path, emit_log=emit_log):
            if metadata is not None:
                metadata.update(
                    {
                        "cache_hit": False,
                        "image_path": str(output_path),
                        "generated_at": _now_text(),
                        "image_size": output_path.stat().st_size if output_path.exists() else 0,
                    }
                )
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图生成成功 "
                    f"building={building}, session_id={session_id}, image={output_path}, "
                    f"size={output_path.stat().st_size if output_path.exists() else 0}"
                )
            return output_path
        if emit_log:
            if allow_fallback:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图不可用，已切换到内置渲染兜底 "
                    f"building={building}, session_id={session_id}, source={source_path}"
                )
            else:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图不可用，已停止发送，避免发送旧格式图片 "
                    f"building={building}, session_id={session_id}, source={source_path}"
                )
        if not allow_fallback:
            raise RuntimeError(
                "Excel截图失败，未生成容量表图片；为避免发送错误格式图片，本次不使用内置渲染兜底。"
                "如果日志提示缺少 pythoncom/win32com，请等待启动依赖自动安装 pywin32 后重启；"
                "如果日志提示“服务器运行失败”，说明 pywin32 已导入但 Excel COM 未能启动，"
                "请确认当前账号可正常打开 Excel，且没有首次启动、激活或弹窗卡住。"
            )

        workbook = load_workbook(source_path, data_only=False)
        try:
            configured_sheet = self._configured_sheet_name()
            sheet = workbook[configured_sheet] if configured_sheet and configured_sheet in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            image = self._render_sheet(sheet, display_values=self._formula_display_values(sheet))
            output_path.parent.mkdir(parents=True, exist_ok=True)
            buffer = self._image_bytes(image)
            atomic_write_bytes(output_path, buffer, validator=validate_image_file, temp_suffix=".tmp")
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] 内置渲染生成成功 "
                    f"building={building}, session_id={session_id}, image={output_path}"
                )
            if metadata is not None:
                metadata.update(
                    {
                        "cache_hit": False,
                        "image_path": str(output_path),
                        "generated_at": _now_text(),
                        "image_size": output_path.stat().st_size if output_path.exists() else 0,
                    }
                )
            return output_path
        finally:
            workbook.close()

    def _render_with_excel_copy_picture(
        self,
        *,
        source_path: Path,
        output_path: Path,
        emit_log: Callable[[str], None] | None = None,
    ) -> bool:
        # Excel CopyPicture and the clipboard are process-global resources; serialize
        # this path so concurrent review-page sends cannot steal each other's image.
        wait_started = time.perf_counter()
        if emit_log:
            emit_log(
                "[交接班][容量表图片发送] 等待Excel截图锁 "
                f"source={source_path}, timeout_sec={_CAPACITY_IMAGE_EXCEL_LOCK_TIMEOUT_SEC:g}"
            )
        acquired = _CAPACITY_IMAGE_EXCEL_COPY_LOCK.acquire(timeout=_CAPACITY_IMAGE_EXCEL_LOCK_TIMEOUT_SEC)
        wait_ms = int((time.perf_counter() - wait_started) * 1000)
        if not acquired:
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图锁等待超时 "
                    f"source={source_path}, wait_ms={wait_ms}"
                )
            raise TimeoutError(_CAPACITY_IMAGE_LOCK_TIMEOUT_ERROR)
        try:
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] 已获取Excel截图锁 "
                    f"source={source_path}, wait_ms={wait_ms}"
                )
            return self._render_with_excel_copy_picture_locked(source_path=source_path, output_path=output_path, emit_log=emit_log)
        finally:
            _CAPACITY_IMAGE_EXCEL_COPY_LOCK.release()

    def _render_with_excel_copy_picture_locked(
        self,
        *,
        source_path: Path,
        output_path: Path,
        emit_log: Callable[[str], None] | None = None,
    ) -> bool:
        excel = None
        workbook = None
        co_initialized = False
        excel_pid = 0
        excel_progid = _EXCEL_COM_PROGID
        quit_ok = False
        try:
            import pythoncom
            import win32com.client
            from PIL import ImageGrab

            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图开始 "
                    f"source={source_path}, target={output_path}"
                )
            pythoncom.CoInitialize()
            co_initialized = True
            excel, excel_progid = self._dispatch_excel_application(win32com.client, emit_log=emit_log)
            try:
                import win32process

                _, excel_pid = win32process.GetWindowThreadProcessId(int(excel.Hwnd))
            except Exception:
                excel_pid = 0
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel已启动 "
                    f"source={source_path}, progid={excel_progid}, excel_pid={excel_pid or '-'}"
                )
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            try:
                excel.EnableEvents = False
            except Exception:
                pass
            workbook = excel.Workbooks.Open(os.path.abspath(str(source_path)), UpdateLinks=0, ReadOnly=True, AddToMru=False)
            configured_sheet = self._configured_sheet_name()
            if configured_sheet:
                try:
                    sheet = workbook.Sheets(configured_sheet)
                except Exception:
                    sheet = workbook.Sheets(1)
                    if emit_log:
                        emit_log(
                            "[交接班][容量表图片发送] 配置Sheet不存在，已使用首个Sheet "
                            f"configured_sheet={configured_sheet}"
                        )
            else:
                sheet = workbook.Sheets(1)
            try:
                sheet_name = str(sheet.Name)
            except Exception:
                sheet_name = configured_sheet or "1"
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图Sheet已选中 "
                    f"sheet={sheet_name}, progid={excel_progid}, excel_pid={excel_pid or '-'}"
                )
            try:
                workbook.Activate()
                sheet.Activate()
                sheet.DisplayPageBreaks = False
                excel.ActiveWindow.View = 1
                excel.ActiveWindow.DisplayGridlines = False
            except Exception:
                pass
            try:
                sheet.PageSetup.LeftHeader = ""
                sheet.PageSetup.CenterHeader = ""
                sheet.PageSetup.RightHeader = ""
                sheet.PageSetup.LeftFooter = ""
                sheet.PageSetup.CenterFooter = ""
                sheet.PageSetup.RightFooter = ""
            except Exception:
                pass
            try:
                excel.CalculateFullRebuild()
            except Exception:
                try:
                    excel.Calculate()
                except Exception:
                    pass

            used_range = sheet.UsedRange
            try:
                address = str(used_range.Address)
            except Exception:
                address = "UsedRange"
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel复制区域到剪贴板 "
                    f"sheet={sheet_name}, range={address}"
                )
            used_range.CopyPicture(Format=2)
            time.sleep(1)
            image = None
            for _ in range(10):
                image = ImageGrab.grabclipboard()
                if isinstance(image, Image.Image):
                    break
                time.sleep(0.2)
            if not isinstance(image, Image.Image):
                if emit_log:
                    emit_log("[交接班][容量表图片发送] Excel截图失败: 剪贴板未返回图片")
                return False
            output_path.parent.mkdir(parents=True, exist_ok=True)
            atomic_write_bytes(output_path, self._image_bytes(image.convert("RGB")), validator=validate_image_file, temp_suffix=".tmp")
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图图片保存成功 "
                    f"image={output_path}, size={output_path.stat().st_size if output_path.exists() else 0}"
                )
            return True
        except ModuleNotFoundError as exc:
            if emit_log:
                emit_log(
                    "[交接班][容量表图片发送] Excel截图依赖缺失 "
                    f"missing={exc.name or '-'}, error={exc}; 请在服务运行的Python环境安装 pywin32"
                )
            return False
        except Exception as exc:
            if emit_log:
                detail = str(exc)
                hint = ""
                if "-2146959355" in detail or "服务器运行失败" in detail:
                    hint = (
                        "；pywin32已导入，但 Excel/WPS COM 启动失败，请用当前运行账号手动打开一次 Excel 或 WPS表格，"
                        "处理首次启动/激活/弹窗后再重试；如使用WPS，请确认 KET.Application 可用"
                    )
                emit_log(f"[交接班][容量表图片发送] Excel截图异常: error={exc}{hint}")
            return False
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                    quit_ok = True
                except Exception:
                    pass
            if excel_pid and not quit_ok:
                try:
                    import psutil

                    process = psutil.Process(excel_pid)
                    if process.is_running():
                        process.terminate()
                        if emit_log:
                            emit_log(f"[交接班][容量表图片发送] 已终止本次创建的Excel进程 pid={excel_pid}")
                except Exception:
                    pass
            if co_initialized:
                try:
                    import pythoncom

                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    @staticmethod
    def _number_value(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                return 0.0
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text:
            return 0.0
        percent = text.endswith("%")
        if percent:
            text = text[:-1].strip()
        try:
            number = float(text)
        except Exception:
            return 0.0
        return number / 100 if percent else number

    @classmethod
    def _formula_display_values(cls, sheet) -> Dict[str, str]:
        min_col, min_row, max_col, max_row = cls._render_bounds(sheet)
        cache: Dict[str, Any] = {}
        visiting: set[str] = set()

        def eval_cell(coord: str) -> Any:
            normalized = coord.replace("$", "").upper()
            if normalized in cache:
                return cache[normalized]
            if normalized in visiting:
                return ""
            visiting.add(normalized)
            try:
                cell = sheet[normalized]
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    result = eval_formula(value)
                else:
                    result = value
                cache[normalized] = result
                return result
            finally:
                visiting.discard(normalized)

        def range_values(ref: str) -> List[Any]:
            try:
                min_c, min_r, max_c, max_r = range_boundaries(ref.replace("$", ""))
            except Exception:
                return []
            values: List[Any] = []
            for row_index in range(min_r, max_r + 1):
                for col_index in range(min_c, max_c + 1):
                    values.append(eval_cell(f"{get_column_letter(col_index)}{row_index}"))
            return values

        def replace_sum(match: re.Match[str]) -> str:
            args = str(match.group(1) or "")
            total = 0.0
            for part in [item.strip() for item in args.split(",") if item.strip()]:
                if ":" in part:
                    total += sum(cls._number_value(item) for item in range_values(part))
                else:
                    total += cls._number_value(eval_expr(part))
            return str(total)

        cell_ref_pattern = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?([0-9]{1,7})(?![A-Za-z0-9_])")

        def eval_expr(expr: str) -> Any:
            text = str(expr or "").strip()
            if not text:
                return 0.0
            if "#REF!" in text.upper():
                return ""
            while re.search(r"\bSUM\s*\(([^()]*)\)", text, flags=re.IGNORECASE):
                text = re.sub(r"\bSUM\s*\(([^()]*)\)", replace_sum, text, flags=re.IGNORECASE)
            if ":" in text:
                return ""
            text = re.sub(r"(\d+(?:\.\d+)?)\s*%", r"(\1/100)", text)

            def repl_ref(match: re.Match[str]) -> str:
                value = eval_cell(f"{match.group(1)}{match.group(2)}")
                return str(cls._number_value(value))

            text = cell_ref_pattern.sub(repl_ref, text)
            if not re.fullmatch(r"[0-9eE+\-*/().\s]+", text):
                return ""
            try:
                result = eval(text, {"__builtins__": {}}, {})  # noqa: S307 - sanitized arithmetic only
            except Exception:
                return ""
            if isinstance(result, (int, float)) and not (isinstance(result, float) and (math.isnan(result) or math.isinf(result))):
                return result
            return ""

        def eval_formula(formula: str) -> Any:
            return eval_expr(str(formula or "")[1:])

        display_values: Dict[str, str] = {}
        for row_index in range(min_row, max_row + 1):
            for col_index in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col_index)}{row_index}"
                cell = sheet.cell(row=row_index, column=col_index)
                value = eval_cell(coord)
                display_values[coord] = cls._format_display_value(value, cell)
        return display_values

    @staticmethod
    def _format_display_value(value: Any, cell) -> str:
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                return ""
            fmt = str(getattr(cell, "number_format", "") or "")
            if "%" in fmt:
                decimals = 0
                match = re.search(r"\.([0#]+)%", fmt)
                if match:
                    decimals = len(match.group(1))
                return f"{float(value) * 100:.{decimals}f}%"
            match = re.search(r"\.([0#]+)", fmt)
            if match and fmt != "@":
                decimals = len(match.group(1))
                return f"{float(value):.{decimals}f}"
            if float(value).is_integer():
                return str(int(value))
            return f"{float(value):.6f}".rstrip("0").rstrip(".")
        return _text(value)

    @staticmethod
    def _image_bytes(image: Image.Image) -> bytes:
        from io import BytesIO

        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()

    @staticmethod
    def _print_area_bounds(sheet) -> Tuple[int, int, int, int] | None:
        raw = getattr(sheet, "print_area", None)
        if not raw:
            return None
        areas = raw if isinstance(raw, (list, tuple)) else str(raw).split(",")
        for area in areas:
            ref = str(area or "").strip()
            if not ref:
                continue
            if "!" in ref:
                ref = ref.split("!", 1)[1]
            ref = ref.replace("'", "").replace("$", "")
            try:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
            except Exception:
                continue
            if min_col and min_row and max_col and max_row:
                return min_col, min_row, max_col, max_row
        return None

    @staticmethod
    def _content_bounds(sheet) -> Tuple[int, int, int, int]:
        min_row = sheet.max_row or 1
        min_col = sheet.max_column or 1
        max_row = 1
        max_col = 1
        found = False
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if _text(cell.value):
                    found = True
                    min_row = min(min_row, cell.row)
                    min_col = min(min_col, cell.column)
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)
        for merged in sheet.merged_cells.ranges:
            min_col = min(min_col, merged.min_col)
            min_row = min(min_row, merged.min_row)
            max_col = max(max_col, merged.max_col)
            max_row = max(max_row, merged.max_row)
            found = True
        if not found:
            return 1, 1, 1, 1
        return min_col, min_row, max_col, max_row

    @classmethod
    def _render_bounds(cls, sheet) -> Tuple[int, int, int, int]:
        return cls._print_area_bounds(sheet) or cls._content_bounds(sheet)

    @staticmethod
    def _column_width_px(sheet, column_index: int) -> int:
        letter = get_column_letter(column_index)
        dimension = sheet.column_dimensions.get(letter)
        if dimension is not None and bool(getattr(dimension, "hidden", False)):
            return 0
        width = getattr(dimension, "width", None) if dimension is not None else None
        try:
            value = float(width if width is not None else 8.43)
        except Exception:
            value = 8.43
        return max(28, min(220, int(round(value * 7 + 8))))

    @staticmethod
    def _row_height_px(sheet, row_index: int) -> int:
        dimension = sheet.row_dimensions.get(row_index)
        if dimension is not None and bool(getattr(dimension, "hidden", False)):
            return 0
        height = getattr(dimension, "height", None) if dimension is not None else None
        try:
            points = float(height if height is not None else 15)
        except Exception:
            points = 15
        return max(20, min(120, int(round(points * 96 / 72 + 6))))

    @staticmethod
    def _rgb(color: Any, default: Tuple[int, int, int] | None = None) -> Tuple[int, int, int] | None:
        if color is None:
            return default
        rgb = str(getattr(color, "rgb", "") or "").strip()
        if rgb and len(rgb) in {6, 8}:
            rgb = rgb[-6:]
            try:
                return int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            except Exception:
                return default
        return default

    @classmethod
    def _fill_color(cls, cell) -> Tuple[int, int, int]:
        fill = getattr(cell, "fill", None)
        if fill is None or str(getattr(fill, "fill_type", "") or "").lower() in {"", "none"}:
            return 255, 255, 255
        color = cls._rgb(getattr(fill, "fgColor", None))
        if color is None or color == (0, 0, 0):
            return 255, 255, 255
        return color

    @classmethod
    def _font_color(cls, cell) -> Tuple[int, int, int]:
        color = cls._rgb(getattr(getattr(cell, "font", None), "color", None))
        return color or (0, 0, 0)

    @staticmethod
    def _font_path(bold: bool) -> str:
        candidates = [
            Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("C:/Windows/Fonts/simsun.ttc"),
        ]
        for path in candidates:
            if path.exists():
                return str(path)
        return ""

    @classmethod
    def _font(cls, cell) -> ImageFont.ImageFont:
        raw_size = getattr(getattr(cell, "font", None), "sz", None)
        try:
            size = max(10, min(36, int(round(float(raw_size or 11) * 96 / 72))))
        except Exception:
            size = 15
        bold = bool(getattr(getattr(cell, "font", None), "bold", False))
        font_path = cls._font_path(bold)
        if font_path:
            try:
                return ImageFont.truetype(font_path, size=size)
            except Exception:
                pass
        return ImageFont.load_default()

    @staticmethod
    def _measure(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> Tuple[int, int]:
        if not text:
            return 0, 0
        box = draw.textbbox((0, 0), text, font=font)
        return max(0, box[2] - box[0]), max(0, box[3] - box[1])

    @classmethod
    def _wrap_text(cls, draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, width: int) -> List[str]:
        max_width = max(1, width)
        output: List[str] = []
        for raw_line in str(text or "").splitlines() or [""]:
            line = ""
            for char in raw_line:
                candidate = f"{line}{char}"
                if line and cls._measure(draw, candidate, font)[0] > max_width:
                    output.append(line)
                    line = char
                else:
                    line = candidate
            output.append(line)
        return output

    @staticmethod
    def _alignment(cell) -> Tuple[str, str, bool]:
        alignment = getattr(cell, "alignment", None)
        horizontal = str(getattr(alignment, "horizontal", "") or "center").lower()
        vertical = str(getattr(alignment, "vertical", "") or "center").lower()
        wrap = bool(getattr(alignment, "wrap_text", False))
        return horizontal, vertical, wrap

    @staticmethod
    def _merged_bounds(sheet) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
        mapping: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
        for merged in sheet.merged_cells.ranges:
            bounds = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    mapping[(row, col)] = bounds
        return mapping

    @staticmethod
    def _border_color(side) -> Tuple[int, int, int]:
        return CapacityReportImageRenderer._rgb(getattr(side, "color", None), (0, 0, 0)) or (0, 0, 0)

    @staticmethod
    def _has_border(side) -> bool:
        return bool(side is not None and str(getattr(side, "style", "") or "").strip())

    @classmethod
    def _draw_border(cls, draw: ImageDraw.ImageDraw, rect: Tuple[int, int, int, int], cell) -> None:
        border = getattr(cell, "border", None)
        if border is None:
            draw.rectangle(rect, outline=(210, 210, 210), width=1)
            return
        left, top, right, bottom = rect
        if cls._has_border(border.left):
            draw.line((left, top, left, bottom), fill=cls._border_color(border.left), width=1)
        if cls._has_border(border.right):
            draw.line((right, top, right, bottom), fill=cls._border_color(border.right), width=1)
        if cls._has_border(border.top):
            draw.line((left, top, right, top), fill=cls._border_color(border.top), width=1)
        if cls._has_border(border.bottom):
            draw.line((left, bottom, right, bottom), fill=cls._border_color(border.bottom), width=1)

    @classmethod
    def _render_sheet(cls, sheet, *, display_values: Dict[str, str] | None = None) -> Image.Image:
        display_map = display_values if isinstance(display_values, dict) else {}
        min_col, min_row, max_col, max_row = cls._render_bounds(sheet)
        col_widths = {col: cls._column_width_px(sheet, col) for col in range(min_col, max_col + 1)}
        row_heights = {row: cls._row_height_px(sheet, row) for row in range(min_row, max_row + 1)}
        x_offsets: Dict[int, int] = {}
        y_offsets: Dict[int, int] = {}
        x = 0
        for col in range(min_col, max_col + 1):
            x_offsets[col] = x
            x += col_widths[col]
        y = 0
        for row in range(min_row, max_row + 1):
            y_offsets[row] = y
            y += row_heights[row]
        width = max(1, x)
        height = max(1, y)
        image = Image.new("RGB", (width, height), (255, 255, 255))
        draw = ImageDraw.Draw(image)
        merged_map = cls._merged_bounds(sheet)
        rendered_merged: set[Tuple[int, int, int, int]] = set()

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                bounds = merged_map.get((row, col))
                if bounds:
                    if bounds in rendered_merged:
                        continue
                    rendered_merged.add(bounds)
                    start_col, start_row, end_col, end_row = bounds
                    if start_col < min_col or start_row < min_row:
                        continue
                    cell = sheet.cell(start_row, start_col)
                else:
                    start_col = end_col = col
                    start_row = end_row = row
                    cell = sheet.cell(row, col)
                left = x_offsets[start_col]
                top = y_offsets[start_row]
                right = x_offsets[end_col] + col_widths[end_col]
                bottom = y_offsets[end_row] + row_heights[end_row]
                rect = (left, top, max(left + 1, right), max(top + 1, bottom))
                draw.rectangle(rect, fill=cls._fill_color(cell))
                cls._draw_border(draw, rect, cell)
                value = _text(display_map.get(cell.coordinate, cls._format_display_value(cell.value, cell)))
                if not value:
                    continue
                font = cls._font(cell)
                font_color = cls._font_color(cell)
                horizontal, vertical, wrap = cls._alignment(cell)
                padding_x = 5
                padding_y = 3
                text_width = max(1, rect[2] - rect[0] - padding_x * 2)
                lines = cls._wrap_text(draw, value, font, text_width) if wrap or "\n" in value else [value]
                line_heights = [max(1, cls._measure(draw, line, font)[1]) for line in lines]
                line_gap = 3
                total_text_height = sum(line_heights) + max(0, len(lines) - 1) * line_gap
                available_height = max(1, rect[3] - rect[1] - padding_y * 2)
                if total_text_height > available_height and lines:
                    max_lines = max(1, math.floor((available_height + line_gap) / (max(line_heights) + line_gap)))
                    lines = lines[:max_lines]
                    if len(lines) == max_lines:
                        lines[-1] = lines[-1].rstrip()
                    line_heights = [max(1, cls._measure(draw, line, font)[1]) for line in lines]
                    total_text_height = sum(line_heights) + max(0, len(lines) - 1) * line_gap
                if vertical == "top":
                    text_y = rect[1] + padding_y
                elif vertical == "bottom":
                    text_y = rect[3] - padding_y - total_text_height
                else:
                    text_y = rect[1] + max(0, (rect[3] - rect[1] - total_text_height) // 2)
                for line, line_height in zip(lines, line_heights):
                    measured_width = cls._measure(draw, line, font)[0]
                    if horizontal in {"right", "distributed"}:
                        text_x = rect[2] - padding_x - measured_width
                    elif horizontal in {"left", "general"}:
                        text_x = rect[0] + padding_x
                    else:
                        text_x = rect[0] + max(0, (rect[2] - rect[0] - measured_width) // 2)
                    draw.text((text_x, text_y), line, font=font, fill=font_color)
                    text_y += line_height + line_gap
        return image


class CapacityReportImageDeliveryService:
    def __init__(self, handover_cfg: Dict[str, Any], *, config_path: str | Path | None = None) -> None:
        self.handover_cfg = handover_cfg if isinstance(handover_cfg, dict) else {}
        self.config_path = Path(config_path) if config_path else None
        self._review_service = ReviewSessionService(self.handover_cfg)
        self._link_service = ReviewLinkDeliveryService(self.handover_cfg, config_path=self.config_path)
        self._renderer = CapacityReportImageRenderer(self.handover_cfg)

    @staticmethod
    def _elapsed_ms(started_at: float) -> int:
        return int((time.perf_counter() - started_at) * 1000)

    def validate_preflight(self, session: Dict[str, Any], *, building: str) -> None:
        self._validate_session(session, building=building)
        self._validate_not_running(session)
        recipients = self._link_service._recipients_for_building(building)
        if not recipients:
            raise ValueError("当前楼未配置启用的审核链接接收人")

    def _build_review_text_message(
        self,
        session: Dict[str, Any],
        *,
        building: str,
        emit_log: Callable[[str], None],
    ) -> Tuple[str, str]:
        snapshot = materialize_review_access_snapshot(self.handover_cfg)
        url = self._link_service._review_url_for_building(snapshot, building) or self._link_service._manual_test_review_url_for_building(snapshot, building)
        try:
            message_text = str(self._link_service._summary_message_service.build_for_session(
                session,
                emit_log=emit_log,
            ) or "").strip()
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"审核文本生成失败: {exc}") from exc
        if not message_text:
            raise ValueError("审核文本为空，无法发送")
        return message_text, url

    @staticmethod
    def _capacity_image_cache_signature(
        session: Dict[str, Any],
        *,
        source_file: str,
    ) -> Tuple[str, bool, str]:
        source_path = Path(str(source_file or "").strip())
        try:
            stat = source_path.stat()
            resolved = str(source_path.resolve())
        except Exception as exc:  # noqa: BLE001
            return "", False, f"file_stat_failed:{exc}"
        capacity_sync = session.get("capacity_sync", {}) if isinstance(session.get("capacity_sync", {}), dict) else {}
        parts = {
            "source_path": resolved,
            "source_size": str(stat.st_size),
            "source_mtime_ns": str(stat.st_mtime_ns),
            "session_id": str(session.get("session_id", "") or "").strip(),
            "capacity_input_signature": str(capacity_sync.get("input_signature", "") or "").strip(),
            "capacity_sync_updated_at": str(capacity_sync.get("updated_at", "") or "").strip(),
        }
        missing = [key for key, value in parts.items() if not value]
        if missing:
            return "", False, "missing:" + ",".join(missing)
        raw = "|".join(f"{key}={parts[key]}" for key in sorted(parts))
        return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest(), True, ""

    def _save_review_link_delivery(
        self,
        *,
        session_id: str,
        status: str,
        attempt_at: str,
        url: str,
        successful_recipients: List[str],
        failed_recipients: List[Dict[str, str]],
        source: str,
        error: str = "",
    ) -> None:
        self._review_service.update_review_link_delivery(
            session_id=session_id,
            review_link_delivery={
                "status": status,
                "last_attempt_at": attempt_at,
                "last_sent_at": attempt_at if status == "success" else "",
                "error": str(error or "").strip(),
                "url": url,
                "successful_recipients": successful_recipients,
                "failed_recipients": failed_recipients,
                "source": str(source or "manual").strip().lower() or "manual",
                "auto_attempted": False,
                "auto_attempted_at": "",
            },
        )

    @staticmethod
    def _validate_not_running(session: Dict[str, Any]) -> None:
        delivery = session.get("capacity_image_delivery", {}) if isinstance(session, dict) else {}
        if not isinstance(delivery, dict):
            return
        status = str(delivery.get("status", "") or "").strip().lower()
        if status in _CAPACITY_IMAGE_DELIVERY_RUNNING_STATUSES:
            raise ValueError("容量表图片正在发送中，请等待发送完成")

    @staticmethod
    def _validate_session(session: Dict[str, Any], *, building: str) -> None:
        if not isinstance(session, dict) or not str(session.get("session_id", "") or "").strip():
            raise ValueError("未找到交接班审核会话")
        if str(session.get("building", "") or "").strip() != str(building or "").strip():
            raise ValueError("审核会话楼栋不匹配")
        output_file_text = str(session.get("capacity_output_file", "") or "").strip()
        if not output_file_text:
            raise ValueError("当前交接班容量报表尚未生成")
        output_file = Path(output_file_text)
        if not output_file.exists() or not output_file.is_file():
            raise FileNotFoundError("交接班容量报表文件不存在，请重新生成")
        capacity_sync = session.get("capacity_sync", {}) if isinstance(session.get("capacity_sync", {}), dict) else {}
        if str(capacity_sync.get("status", "") or "").strip().lower() != "ready":
            raise ValueError(str(capacity_sync.get("error", "") or "").strip() or "容量报表待补写完成后才能发送")

    def begin_delivery(self, session: Dict[str, Any], *, building: str, source: str = "manual") -> Dict[str, Any]:
        session_id = str(session.get("session_id", "") or "").strip() if isinstance(session, dict) else ""
        if not session_id:
            raise ValueError("未找到交接班审核会话")
        with _CAPACITY_IMAGE_DELIVERY_LOCK:
            latest_session = self._review_service.get_session_by_id(session_id)
            if not isinstance(latest_session, dict):
                latest_session = session
            self.validate_preflight(latest_session, building=building)
            delivery = {
                "status": "sending",
                "last_attempt_at": _now_text(),
                "last_sent_at": "",
                "error": "",
                "image_path": "",
                "image_key": "",
                "image_signature": "",
                "cache_hit": False,
                "generated_at": "",
                "successful_recipients": [],
                "failed_recipients": [],
                "source": str(source or "manual").strip().lower() or "manual",
            }
            updated = self._review_service.update_capacity_image_delivery(
                session_id=session_id,
                capacity_image_delivery=delivery,
            )
            return dict(updated.get("capacity_image_delivery", delivery))

    def mark_failed(self, *, session_id: str, error: str, source: str = "manual") -> Dict[str, Any]:
        delivery = {
            "status": "failed",
            "last_attempt_at": _now_text(),
            "last_sent_at": "",
            "error": str(error or "容量表图片发送失败").strip() or "容量表图片发送失败",
            "image_path": "",
            "image_key": "",
            "image_signature": "",
            "cache_hit": False,
            "generated_at": "",
            "successful_recipients": [],
            "failed_recipients": [],
            "source": str(source or "manual").strip().lower() or "manual",
        }
        updated = self._review_service.update_capacity_image_delivery(
            session_id=str(session_id or "").strip(),
            capacity_image_delivery=delivery,
        )
        return dict(updated.get("capacity_image_delivery", delivery))

    def send_for_session(
        self,
        session: Dict[str, Any],
        *,
        building: str,
        source: str = "manual",
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        normalized_session = dict(session) if isinstance(session, dict) else {}
        building_text = str(building or normalized_session.get("building", "") or "").strip()
        session_id = str(normalized_session.get("session_id", "") or "").strip()
        total_started = time.perf_counter()
        self._validate_session(normalized_session, building=building_text)
        recipient_started = time.perf_counter()
        recipient_snapshot = self._link_service._recipient_snapshot_for_building(building_text)
        recipients = list(recipient_snapshot.get("recipients", []))
        if not recipients:
            raise ValueError("当前楼未配置启用的审核链接接收人")

        attempt_at = _now_text()
        emit_log(
            "[交接班][容量表图片发送] 开始 "
            f"building={building_text}, session_id={session_id}, "
            f"source_file={normalized_session.get('capacity_output_file', '')}, recipients={len(recipients)}, "
            f"raw={int(recipient_snapshot.get('raw_count', 0) or 0)}, "
            f"enabled={int(recipient_snapshot.get('enabled_count', 0) or 0)}, "
            f"disabled={int(recipient_snapshot.get('disabled_count', 0) or 0)}, "
            f"invalid={int(recipient_snapshot.get('invalid_count', 0) or 0)}, "
            f"open_ids={recipient_snapshot.get('open_ids', [])}, "
            f"disabled_open_ids={recipient_snapshot.get('disabled_open_ids', [])}, "
            f"invalid_recipients={recipient_snapshot.get('invalid_recipients', [])}"
        )
        emit_log(
            "[交接班][容量表图片发送] 阶段耗时 "
            f"stage=recipient_snapshot, building={building_text}, session_id={session_id}, "
            f"elapsed_ms={self._elapsed_ms(recipient_started)}"
        )
        self._review_service.update_capacity_image_delivery(
            session_id=session_id,
            capacity_image_delivery={
                "status": "sending",
                "last_attempt_at": attempt_at,
                "last_sent_at": "",
                "error": "",
                "image_path": "",
                "image_key": "",
                "image_signature": "",
                "cache_hit": False,
                "generated_at": "",
                "successful_recipients": [],
                "failed_recipients": [],
                "source": str(source or "manual").strip().lower() or "manual",
            },
        )
        image_path: Path | None = None
        image_key = ""
        image_signature = ""
        image_cache_hit = False
        image_generated_at = ""
        render_metadata: Dict[str, Any] = {}
        review_url = ""
        successful_recipients: List[str] = []
        failed_recipients: List[Dict[str, str]] = []
        try:
            text_started = time.perf_counter()
            message_text, review_url = self._build_review_text_message(
                normalized_session,
                building=building_text,
                emit_log=emit_log,
            )
            emit_log(
                "[交接班][容量表图片发送] 审核文本已生成 "
                f"building={building_text}, session_id={session_id}, length={len(message_text)}, review_url={review_url or '-'}"
            )
            emit_log(
                "[交接班][容量表图片发送] 本次将发送审核文本内容如下:\n"
                f"{message_text}"
            )
            emit_log(
                "[交接班][容量表图片发送] 阶段耗时 "
                f"stage=build_review_text, building={building_text}, session_id={session_id}, "
                f"elapsed_ms={self._elapsed_ms(text_started)}"
            )
            image_signature, image_cache_reusable, image_signature_reason = self._capacity_image_cache_signature(
                normalized_session,
                source_file=str(normalized_session.get("capacity_output_file", "") or "").strip(),
            )
            if image_cache_reusable:
                emit_log(
                    "[交接班][容量表图片发送] 容量表图片缓存签名已生成 "
                    f"building={building_text}, session_id={session_id}, image_signature={image_signature}"
                )
            else:
                emit_log(
                    "[交接班][容量表图片发送] 容量表图片缓存不可复用，强制重新截图 "
                    f"building={building_text}, session_id={session_id}, reason={image_signature_reason or '-'}"
                )
            render_started = time.perf_counter()
            image_path = self._renderer.render_to_image(
                source_file=str(normalized_session.get("capacity_output_file", "") or "").strip(),
                building=building_text,
                duty_date=str(normalized_session.get("duty_date", "") or "").strip(),
                duty_shift=str(normalized_session.get("duty_shift", "") or "").strip(),
                session_id=session_id,
                force_refresh=not image_cache_reusable,
                allow_fallback=False,
                cache_signature=image_signature if image_cache_reusable else "",
                render_metadata=render_metadata,
                emit_log=emit_log,
            )
            image_cache_hit = bool(render_metadata.get("cache_hit", False))
            image_generated_at = str(render_metadata.get("generated_at", "") or "").strip()
            emit_log(
                "[交接班][容量表图片发送] 图片已生成 "
                f"building={building_text}, session_id={session_id}, image={image_path}, "
                f"cache_hit={image_cache_hit}, size={int(render_metadata.get('image_size', 0) or 0)}"
            )
            emit_log(
                "[交接班][容量表图片发送] 阶段耗时 "
                f"stage=render_capacity_image, building={building_text}, session_id={session_id}, "
                f"elapsed_ms={self._elapsed_ms(render_started)}, cache_hit={image_cache_hit}"
            )

            client = self._link_service._build_feishu_client()
            upload_started = time.perf_counter()
            emit_log(
                "[交接班][容量表图片发送] 开始上传飞书图片 "
                f"building={building_text}, session_id={session_id}, image={image_path}"
            )
            image_upload = client.upload_image(str(image_path))
            image_key = str(image_upload.get("image_key", "") or "").strip()
            if not image_key:
                raise RuntimeError(f"飞书图片上传未返回 image_key: {image_upload}")
            emit_log(
                "[交接班][容量表图片发送] 飞书图片上传成功 "
                f"building={building_text}, session_id={session_id}, image_key={image_key}"
            )
            emit_log(
                "[交接班][容量表图片发送] 阶段耗时 "
                f"stage=upload_image, building={building_text}, session_id={session_id}, "
                f"elapsed_ms={self._elapsed_ms(upload_started)}"
            )
            for recipient in recipients:
                open_id = str(recipient.get("open_id", "") or "").strip()
                note = str(recipient.get("note", "") or "").strip()
                receive_id_type = self._link_service._resolve_effective_receive_id_type(open_id, "open_id")
                text_sent = False
                try:
                    recipient_started = time.perf_counter()
                    emit_log(
                        "[交接班][容量表图片发送] 准备发送审核文本 "
                        f"building={building_text}, session_id={session_id}, open_id={open_id}, "
                        f"receive_id_type={receive_id_type}, note={note or '-'}"
                    )
                    client.send_text_message(
                        receive_id=open_id,
                        receive_id_type=receive_id_type,
                        text=message_text,
                    )
                    text_sent = True
                    emit_log(
                        "[交接班][容量表图片发送] 审核文本发送成功 "
                        f"building={building_text}, session_id={session_id}, open_id={open_id}, note={note or '-'}"
                    )
                    emit_log(
                        "[交接班][容量表图片发送] 阶段耗时 "
                        f"stage=send_text, building={building_text}, session_id={session_id}, "
                        f"open_id={open_id}, elapsed_ms={self._elapsed_ms(recipient_started)}"
                    )
                    image_send_started = time.perf_counter()
                    emit_log(
                        "[交接班][容量表图片发送] 准备发送容量图片 "
                        f"building={building_text}, session_id={session_id}, open_id={open_id}, "
                        f"receive_id_type={receive_id_type}, note={note or '-'}"
                    )
                    client.send_image_message(
                        receive_id=open_id,
                        receive_id_type=receive_id_type,
                        image_key=image_key,
                    )
                    successful_recipients.append(open_id)
                    emit_log(
                        "[交接班][容量表图片发送] 容量图片发送成功 "
                        f"building={building_text}, session_id={session_id}, open_id={open_id}, note={note or '-'}"
                    )
                    emit_log(
                        "[交接班][容量表图片发送] 阶段耗时 "
                        f"stage=send_image, building={building_text}, session_id={session_id}, "
                        f"open_id={open_id}, elapsed_ms={self._elapsed_ms(image_send_started)}"
                    )
                except Exception as exc:  # noqa: BLE001
                    step = "image" if text_sent else "text"
                    failed_recipients.append({"open_id": open_id, "note": note, "step": step, "error": str(exc)})
                    if step == "image":
                        emit_log(
                            "[交接班][容量表图片发送] 容量图片发送失败 "
                            f"building={building_text}, session_id={session_id}, open_id={open_id}, "
                            f"receive_id_type={receive_id_type}, note={note or '-'}, error={exc}"
                        )
                    else:
                        emit_log(
                            "[交接班][容量表图片发送] 审核文本发送失败 "
                            f"building={building_text}, session_id={session_id}, open_id={open_id}, "
                            f"receive_id_type={receive_id_type}, note={note or '-'}, error={exc}"
                        )
        except Exception as exc:  # noqa: BLE001
            delivery = {
                "status": "failed",
                "last_attempt_at": attempt_at,
                "last_sent_at": "",
                "error": str(exc),
                "image_path": str(image_path or ""),
                "image_key": image_key,
                "image_signature": image_signature,
                "cache_hit": image_cache_hit,
                "generated_at": image_generated_at,
                "successful_recipients": successful_recipients,
                "failed_recipients": failed_recipients,
                "source": str(source or "manual").strip().lower() or "manual",
            }
            try:
                self._review_service.update_capacity_image_delivery(
                    session_id=session_id,
                    capacity_image_delivery=delivery,
                )
            except Exception as save_exc:  # noqa: BLE001
                emit_log(
                    "[交接班][容量表图片发送] 失败状态保存失败 "
                    f"building={building_text}, session_id={session_id}, error={save_exc}"
                )
            emit_log(
                "[交接班][容量表图片发送] 失败 "
                f"building={building_text}, session_id={session_id}, error={exc}, "
                f"failed_recipients={failed_recipients}, elapsed_ms={self._elapsed_ms(total_started)}"
            )
            review_delivery = {
                "status": "failed",
                "last_attempt_at": attempt_at,
                "last_sent_at": "",
                "error": str(exc),
                "url": review_url,
                "successful_recipients": successful_recipients,
                "failed_recipients": failed_recipients,
                "source": str(source or "manual").strip().lower() or "manual",
            }
            try:
                self._save_review_link_delivery(
                    session_id=session_id,
                    status="failed",
                    attempt_at=attempt_at,
                    url=review_url,
                    successful_recipients=successful_recipients,
                    failed_recipients=failed_recipients,
                    source=source,
                    error=str(exc),
                )
            except Exception as save_exc:  # noqa: BLE001
                emit_log(
                    "[交接班][容量表图片发送] 审核文本失败状态保存失败 "
                    f"building={building_text}, session_id={session_id}, error={save_exc}"
                )
            return {
                "ok": False,
                "status": "failed",
                "error": str(exc),
                "building": building_text,
                "session_id": session_id,
                "successful_recipients": successful_recipients,
                "failed_recipients": failed_recipients,
                "capacity_image_delivery": delivery,
                "review_link_delivery": review_delivery,
            }

        if not failed_recipients and len(successful_recipients) == len(recipients):
            status = "success"
            error = ""
        else:
            status = "failed"
            error = "发送失败，详见收件人明细"
        delivery = {
            "status": status,
            "last_attempt_at": attempt_at,
            "last_sent_at": attempt_at if status == "success" else "",
            "error": error,
            "image_path": str(image_path or ""),
            "image_key": image_key,
            "image_signature": image_signature,
            "cache_hit": image_cache_hit,
            "generated_at": image_generated_at,
            "successful_recipients": successful_recipients,
            "failed_recipients": failed_recipients,
            "source": str(source or "manual").strip().lower() or "manual",
        }
        try:
            updated = self._review_service.update_capacity_image_delivery(
                session_id=session_id,
                capacity_image_delivery=delivery,
            )
            delivery = dict(updated.get("capacity_image_delivery", delivery))
        except Exception as exc:  # noqa: BLE001
            emit_log(
                "[交接班][容量表图片发送] 状态保存失败但不影响发送结果 "
                f"building={building_text}, session_id={session_id}, error={exc}"
            )
        try:
            self._save_review_link_delivery(
                session_id=session_id,
                status=status,
                attempt_at=attempt_at,
                url=review_url,
                successful_recipients=successful_recipients,
                failed_recipients=failed_recipients,
                source=source,
                error=error,
            )
        except Exception as exc:  # noqa: BLE001
            emit_log(
                "[交接班][容量表图片发送] 审核文本发送状态保存失败但不影响发送结果 "
                f"building={building_text}, session_id={session_id}, error={exc}"
            )
        emit_log(
            "[交接班][容量表图片发送] 完成 "
            f"building={building_text}, session_id={session_id}, status={status}, "
            f"successful={len(successful_recipients)}, failed={len(failed_recipients)}, "
            f"failed_recipients={failed_recipients}, elapsed_ms={self._elapsed_ms(total_started)}"
        )
        return {
            "ok": status == "success",
            "status": status,
            "error": error,
            "building": building_text,
            "session_id": session_id,
            "successful_recipients": successful_recipients,
            "failed_recipients": failed_recipients,
            "capacity_image_delivery": delivery,
            "review_link_delivery": {
                "status": status,
                "last_attempt_at": attempt_at,
                "last_sent_at": attempt_at if status == "success" else "",
                "error": error,
                "url": review_url,
                "successful_recipients": successful_recipients,
                "failed_recipients": failed_recipients,
                "source": str(source or "manual").strip().lower() or "manual",
            },
        }
