from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from pipeline_utils import get_app_dir
from handover_log_module.repository.excel_reader import load_workbook_quietly
from handover_log_module.core.change_work_window import resolve_work_window
from handover_log_module.core.section_layout import parse_category_sections
from handover_log_module.core.specialty_normalizer import normalize_specialty_text
from handover_log_module.repository.change_management_repository import (
    ChangeManagementRepository,
    ChangeManagementRow,
    ChangeRowsByBuilding,
)
from handover_log_module.repository.shift_roster_repository import ShiftRosterRepository


def _norm_header(value: Any) -> str:
    return str(value or "").replace(" ", "").strip().casefold()


def _dedupe_text(value: Any) -> str:
    return "".join(str(value or "").split()).casefold()


class ChangeManagementPayloadBuilder:
    def __init__(
        self,
        handover_cfg: Dict[str, Any],
        *,
        repository: ChangeManagementRepository | None = None,
        shift_roster_repo: ShiftRosterRepository | None = None,
    ) -> None:
        self.handover_cfg = handover_cfg
        self.repo = repository or ChangeManagementRepository(handover_cfg)
        self.shift_roster_repo = shift_roster_repo or ShiftRosterRepository(handover_cfg)

    def _resolve_template_path(self, source_path: str) -> Path:
        path = Path(str(source_path or "").strip())
        if path.is_absolute():
            return path
        return get_app_dir() / path

    def _resolve_section_column_mapping(
        self,
        *,
        cfg: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Dict[str, str]]:
        col_cfg = cfg.get("column_mapping", {}) if isinstance(cfg.get("column_mapping", {}), dict) else {}
        fallback_cols = col_cfg.get("fallback_cols", {}) if isinstance(col_cfg.get("fallback_cols", {}), dict) else {}
        header_alias = col_cfg.get("header_alias", {}) if isinstance(col_cfg.get("header_alias", {}), dict) else {}
        resolve_by_header = bool(col_cfg.get("resolve_by_header", True))
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("change_management", "变更管理")).strip() or "变更管理"
        semantic_keys = ["change_level", "work_window", "description", "executor"]
        output = {
            section_name: {
                key: str(fallback_cols.get(key, "")).strip().upper()
                for key in semantic_keys
            }
        }
        defaults = {
            "change_level": "B",
            "work_window": "E",
            "description": "D",
            "executor": "H",
        }
        for key, value in defaults.items():
            if not output[section_name].get(key):
                output[section_name][key] = value

        if not resolve_by_header:
            return output

        template_cfg = self.handover_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return output
        source_path = self._resolve_template_path(str(template_cfg.get("source_path", "")).strip())
        sheet_name = str(template_cfg.get("sheet_name", "交接班日志")).strip() or "交接班日志"
        if not source_path.exists():
            return output

        wb = load_workbook_quietly(source_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return output
            ws = wb[sheet_name]
            sections = parse_category_sections(ws)
            section = next((item for item in sections if item.name == section_name), None)
            if section is None:
                return output
            header_row = section.header_row
            normalized_header_map: Dict[str, str] = {}
            max_col = int(ws.max_column or 0)
            for col_idx in range(1, max_col + 1):
                text = _norm_header(ws.cell(row=header_row, column=col_idx).value)
                if not text:
                    continue
                normalized_header_map[text] = get_column_letter(col_idx)
            for semantic_key in semantic_keys:
                aliases = header_alias.get(semantic_key, [])
                if not isinstance(aliases, list):
                    aliases = []
                found_col = ""
                for alias in aliases:
                    normalized = _norm_header(alias)
                    if normalized and normalized in normalized_header_map:
                        found_col = normalized_header_map[normalized]
                        break
                if found_col:
                    output[section_name][semantic_key] = found_col
            emit_log(f"[交接班][变更管理] 列映射已解析: {output.get(section_name)}")
            return output
        finally:
            wb.close()

    @staticmethod
    def _to_cells(row_payload: Dict[str, str], col_map: Dict[str, str]) -> Dict[str, str]:
        cells: Dict[str, str] = {}
        for semantic_key, text in row_payload.items():
            col = str(col_map.get(semantic_key, "")).strip().upper()
            if not col:
                continue
            cells[col] = str(text or "")
        return cells

    @staticmethod
    def _pick_supervisor_for_single_building(
        engineers: List[Dict[str, str]],
        *,
        building: str,
        specialty_text: str,
    ) -> str:
        specialty = normalize_specialty_text(specialty_text)
        current_building = str(building or "").strip()
        if not specialty or not current_building:
            return ""
        for row in engineers:
            if str(row.get("building", "")).strip() != current_building:
                continue
            if normalize_specialty_text(row.get("specialty", "")) != specialty:
                continue
            supervisor = str(row.get("supervisor", "")).strip()
            if supervisor:
                return supervisor
        return ""

    def _resolve_executor(
        self,
        *,
        row: ChangeManagementRow,
        building: str,
        engineers: List[Dict[str, str]],
        emit_log: Callable[[str], None],
    ) -> str:
        supervisor = self._pick_supervisor_for_single_building(
            engineers,
            building=building,
            specialty_text=row.specialty_text,
        )
        if supervisor:
            emit_log(
                f"[交接班][变更管理] 执行人匹配: mode=single_building, "
                f"record_id={row.record_id}, building={building}, specialty={row.specialty_text}, supervisor={supervisor}"
            )
            return supervisor

        emit_log(
            f"[交接班][变更管理] 执行人未匹配: record_id={row.record_id}, "
            f"buildings={'/'.join(row.building_values) or '-'}, specialty={row.specialty_text or '-'}"
        )
        return "/"

    def build(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        preloaded_rows_by_building: ChangeRowsByBuilding | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        if preloaded_rows_by_building is None:
            try:
                rows, cfg = self.repo.list_current_shift_rows(
                    building=building,
                    duty_date=duty_date,
                    duty_shift=duty_shift,
                    emit_log=emit_log,
                )
            except Exception as exc:  # noqa: BLE001
                emit_log(f"[交接班][变更管理] 读取失败，按空分类继续: {exc}")
                return {}
        else:
            cfg = self.repo.get_config()
            rows = list(preloaded_rows_by_building.get(building, []))
            emit_log(f"[交接班][变更管理] 命中批量预取: building={building}, rows={len(rows)}")

        if not bool(cfg.get("enabled", True)):
            return {}

        try:
            engineers = self.shift_roster_repo.list_engineer_directory(emit_log=emit_log)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][变更管理] 工程师目录读取失败，执行人按 / 继续: {exc}")
            engineers = []

        work_window_cfg = cfg.get("work_window_text", {}) if isinstance(cfg.get("work_window_text", {}), dict) else {}
        payload_rows: List[Dict[str, str]] = []
        matched_supervisor = 0
        unmatched_supervisor = 0
        deduped_count = 0
        seen_descriptions: set[str] = set()

        for row in rows:
            description_text = str(row.description or "").strip()
            if not description_text:
                continue
            dedupe_key = _dedupe_text(description_text)
            if dedupe_key and dedupe_key in seen_descriptions:
                deduped_count += 1
                continue
            if dedupe_key:
                seen_descriptions.add(dedupe_key)
            work_window = resolve_work_window(
                process_updates_text=row.process_updates_text,
                duty_date=duty_date,
                duty_shift=duty_shift,
                day_anchor=str(work_window_cfg.get("day_anchor", "08:00:00")).strip() or "08:00:00",
                day_default_end=str(work_window_cfg.get("day_default_end", "18:30:00")).strip() or "18:30:00",
                night_anchor=str(work_window_cfg.get("night_anchor", "18:00:00")).strip() or "18:00:00",
                night_default_end_next_day=str(
                    work_window_cfg.get("night_default_end_next_day", "08:00:00")
                ).strip()
                or "08:00:00",
            )
            executor = self._resolve_executor(row=row, building=building, engineers=engineers, emit_log=emit_log)
            if executor == "/":
                unmatched_supervisor += 1
            else:
                matched_supervisor += 1
            payload_rows.append(
                {
                    "change_level": str(row.change_level or "").strip(),
                    "work_window": work_window.text,
                    "description": description_text,
                    "executor": executor,
                }
            )

        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("change_management", "变更管理")).strip() or "变更管理"
        col_map = self._resolve_section_column_mapping(cfg=cfg, emit_log=emit_log).get(section_name, {})
        output_rows = [{"cells": self._to_cells(row_payload, col_map)} for row_payload in payload_rows]
        emit_log(
            f"[交接班][变更管理] 构建完成: count={len(output_rows)}, "
            f"deduped={deduped_count}, matched_supervisor={matched_supervisor}, "
            f"unmatched_supervisor={unmatched_supervisor}"
        )
        return {section_name: output_rows}
