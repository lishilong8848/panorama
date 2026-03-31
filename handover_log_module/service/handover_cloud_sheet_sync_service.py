from __future__ import annotations

from copy import deepcopy
from datetime import datetime
import hashlib
import json
from pathlib import Path
import time
from typing import Any, Callable, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from app.modules.feishu.service.sheets_client_runtime import FeishuSheetsClientRuntime


class HandoverCloudSheetSyncService:
    FIXED_HEADER_ROW_COUNT = 21
    DYNAMIC_START_ROW = FIXED_HEADER_ROW_COUNT + 1
    FINAL_SHEET_SERIAL_RETRY_ATTEMPTS = 4
    FINAL_SHEET_SERIAL_RETRY_BACKOFF_SECONDS = 2.0

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    @staticmethod
    def _now_text() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "enabled": True,
            "root_wiki_url": "https://vnet.feishu.cn/wiki/WlpWwkhQGi46pEkYbMTcNnOzntb",
            "template_node_token": "QxeYwGTHbiyz9bk2gRAca4nonod",
            "spreadsheet_name_pattern": "南通园区交接班日志-{date_text}{shift_text}",
            "source_sheet_name": "交接班日志",
            "sheet_names": {
                "A楼": "A楼",
                "B楼": "B楼",
                "C楼": "C楼",
                "D楼": "D楼",
                "E楼": "E楼",
            },
            "sync_mode": "overwrite_named_sheet",
            "copy": {
                "values": True,
                "formulas": True,
                "styles": True,
                "merges": True,
                "row_heights": True,
                "column_widths": True,
            },
            "request": {
                "timeout_sec": 20,
                "max_retries": 3,
                "retry_backoff_sec": 2,
            },
        }

    def _sync_cfg(self) -> Dict[str, Any]:
        raw_root = self.config if isinstance(self.config, dict) else {}
        raw = raw_root.get("cloud_sheet_sync", {})
        raw = raw if isinstance(raw, dict) else {}
        merged = deepcopy(self._defaults())
        for key, value in raw.items():
            if isinstance(value, dict) and isinstance(merged.get(key), dict):
                merged[key].update(value)
            else:
                merged[key] = value
        sync_mode = str(merged.get("sync_mode", "overwrite_named_sheet") or "").strip().lower()
        if sync_mode == "rebuild_sheet":
            sync_mode = "overwrite_named_sheet"
        merged["sync_mode"] = sync_mode or "overwrite_named_sheet"
        return merged

    def _build_client(self) -> FeishuSheetsClientRuntime:
        global_feishu = self.config.get("_global_feishu", {})
        global_feishu = global_feishu if isinstance(global_feishu, dict) else {}
        sync_cfg = self._sync_cfg()
        request_cfg = sync_cfg.get("request", {})
        request_cfg = request_cfg if isinstance(request_cfg, dict) else {}
        return FeishuSheetsClientRuntime(
            app_id=str(global_feishu.get("app_id", "") or "").strip(),
            app_secret=str(global_feishu.get("app_secret", "") or "").strip(),
            timeout=int(request_cfg.get("timeout_sec", 20) or 20),
            request_retry_count=int(request_cfg.get("max_retries", 3) or 3),
            request_retry_interval_sec=float(request_cfg.get("retry_backoff_sec", 2) or 2),
        )

    @staticmethod
    def _normalize_merge_ranges(raw: Any) -> List[Dict[str, int]]:
        output: List[Dict[str, int]] = []
        if not isinstance(raw, list):
            return output
        for item in raw:
            if not isinstance(item, dict):
                continue
            try:
                start_row = int(item.get("start_row_index", 0))
                end_row = int(item.get("end_row_index", 0))
                start_col = int(item.get("start_column_index", 0))
                end_col = int(item.get("end_column_index", 0))
            except (TypeError, ValueError):
                continue
            if end_row <= start_row or end_col <= start_col:
                continue
            output.append(
                {
                    "start_row_index": start_row,
                    "end_row_index": end_row,
                    "start_column_index": start_col,
                    "end_column_index": end_col,
                }
            )
        return output

    def _status_payload(
        self,
        *,
        attempted: bool,
        success: bool,
        status: str,
        spreadsheet_token: str = "",
        spreadsheet_url: str = "",
        spreadsheet_title: str = "",
        sheet_title: str = "",
        synced_revision: int = 0,
        last_attempt_revision: int = 0,
        prepared_at: str = "",
        error: str = "",
        synced_row_count: int = 0,
        synced_column_count: int = 0,
        synced_merges: List[Dict[str, int]] | None = None,
        dynamic_merge_signature: str = "",
    ) -> Dict[str, Any]:
        return {
            "attempted": bool(attempted),
            "success": bool(success),
            "status": str(status or "").strip(),
            "spreadsheet_token": str(spreadsheet_token or "").strip(),
            "spreadsheet_url": str(spreadsheet_url or "").strip(),
            "spreadsheet_title": str(spreadsheet_title or "").strip(),
            "sheet_title": str(sheet_title or "").strip(),
            "synced_revision": int(synced_revision or 0),
            "last_attempt_revision": int(last_attempt_revision or 0),
            "prepared_at": str(prepared_at or "").strip(),
            "updated_at": self._now_text(),
            "error": str(error or "").strip(),
            "synced_row_count": int(synced_row_count or 0),
            "synced_column_count": int(synced_column_count or 0),
            "synced_merges": self._normalize_merge_ranges(synced_merges or []),
            "dynamic_merge_signature": str(dynamic_merge_signature or "").strip(),
        }

    def prepare_batch_spreadsheet(
        self,
        *,
        duty_date: str,
        duty_date_text: str,
        shift_text: str,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        sync_cfg = self._sync_cfg()
        if not bool(sync_cfg.get("enabled", True)):
            return self._status_payload(attempted=False, success=False, status="disabled")

        try:
            client = self._build_client()
            sheet_cache: Dict[str, List[Dict[str, Any]]] = {}
            spreadsheet_meta = self._find_or_create_target_spreadsheet(
                client=client,
                duty_date=duty_date,
                duty_date_text=duty_date_text,
                shift_text=shift_text,
            )
            self._ensure_target_sheets(
                client=client,
                spreadsheet_token=str(spreadsheet_meta.get("spreadsheet_token", "")),
                sheet_cache=sheet_cache,
            )
            prepared_at = self._now_text()
            emit_log(
                f"[交接班][云表预建] 完成 duty={duty_date}, title={spreadsheet_meta.get('title', '-')}"
            )
            return self._status_payload(
                attempted=True,
                success=True,
                status="prepared",
                spreadsheet_token=str(spreadsheet_meta.get("spreadsheet_token", "")),
                spreadsheet_url=str(spreadsheet_meta.get("url", "")),
                spreadsheet_title=str(spreadsheet_meta.get("title", "")),
                prepared_at=prepared_at,
            )
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][云表预建] 失败 duty={duty_date}, error={exc}")
            return self._status_payload(
                attempted=True,
                success=False,
                status="prepare_failed",
                error=str(exc),
            )

    def validate_batch_spreadsheet(
        self,
        *,
        batch_meta: Dict[str, Any],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        normalized = batch_meta if isinstance(batch_meta, dict) else {}
        spreadsheet_token = str(normalized.get("spreadsheet_token", "")).strip()
        if not spreadsheet_token:
            return {"valid": False, "error": "missing_spreadsheet_token"}
        try:
            client = self._build_client()
            client.query_sheets(spreadsheet_token)
            return {"valid": True, "error": ""}
        except Exception as exc:  # noqa: BLE001
            emit_log(
                f"[交接班][云表预建] 检测到已缓存云文档不可用，准备重建:"
                f" token={spreadsheet_token}, error={exc}"
            )
            return {"valid": False, "error": str(exc)}

    def sync_confirmed_buildings(
        self,
        *,
        batch_meta: Dict[str, Any],
        building_items: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        sync_cfg = self._sync_cfg()
        normalized_batch = batch_meta if isinstance(batch_meta, dict) else {}
        spreadsheet_token = str(normalized_batch.get("spreadsheet_token", "")).strip()
        spreadsheet_url = str(normalized_batch.get("spreadsheet_url", "")).strip()
        spreadsheet_title = str(normalized_batch.get("spreadsheet_title", "")).strip()
        if not bool(sync_cfg.get("enabled", True)):
            return {
                "status": "skipped",
                "spreadsheet_token": spreadsheet_token,
                "spreadsheet_url": spreadsheet_url,
                "spreadsheet_title": spreadsheet_title,
                "uploaded_buildings": [],
                "skipped_buildings": [
                    {"building": str(item.get("building", "")).strip(), "reason": "disabled"}
                    for item in building_items
                    if isinstance(item, dict)
                ],
                "failed_buildings": [],
                "details": {},
            }
        if not spreadsheet_token:
            raise RuntimeError("cloud batch 缺少 spreadsheet_token")

        normalized_items = [dict(item) for item in building_items if isinstance(item, dict)]
        building_order = {
            str(item.get("building", "")).strip(): index
            for index, item in enumerate(normalized_items)
            if str(item.get("building", "")).strip()
        }
        uploaded_buildings: List[str] = []
        skipped_buildings: List[Dict[str, str]] = []
        failed_buildings: List[Dict[str, str]] = []
        details: Dict[str, Dict[str, Any]] = {}
        if not normalized_items:
            return {
                "status": "skipped",
                "spreadsheet_token": spreadsheet_token,
                "spreadsheet_url": spreadsheet_url,
                "spreadsheet_title": spreadsheet_title,
                "uploaded_buildings": uploaded_buildings,
                "skipped_buildings": skipped_buildings,
                "failed_buildings": failed_buildings,
                "details": details,
            }

        batch_key = str(normalized_batch.get("batch_key", "")).strip()
        emit_log(
            f"[交接班][云表最终上传] 批量开始 batch={batch_key or '-'}, "
            f"mode=serial, buildings={','.join(building_order.keys())}"
        )
        batch_started = time.perf_counter()
        base_sheet_cache = self._build_base_sheet_cache(
            spreadsheet_token=spreadsheet_token,
            emit_log=emit_log,
        )
        for item in normalized_items:
            building = str(item.get("building", "")).strip()
            if not building:
                continue
            try:
                result = self._run_sync_one_building_with_retry(
                    spreadsheet_token=spreadsheet_token,
                    item=item,
                    emit_log=emit_log,
                    base_sheet_cache=base_sheet_cache,
                )
            except Exception as exc:  # noqa: BLE001
                target_sheet_title = self._target_sheet_title(building)
                failed_buildings.append({"building": building, "error": str(exc)})
                details[building] = {
                    "status": "failed",
                    "sheet_title": target_sheet_title,
                    "synced_revision": 0,
                    "rows": 0,
                    "cols": 0,
                    "merged": 0,
                    "synced_row_count": 0,
                    "synced_column_count": 0,
                    "synced_merges": [],
                    "dynamic_merge_signature": "",
                    "error": str(exc),
                }
                emit_log(f"[交接班][云表最终上传] 失败 building={building}, error={exc}")
                continue

            if bool(result.get("success", False)):
                uploaded_buildings.append(building)
            else:
                failed_buildings.append({"building": building, "error": str(result.get("error", "")).strip()})
            detail = result.get("detail", {})
            details[building] = detail if isinstance(detail, dict) else {}

        if failed_buildings and uploaded_buildings:
            status = "partial_failed"
        elif failed_buildings:
            status = "failed"
        elif uploaded_buildings:
            status = "ok"
        else:
            status = "skipped"

        uploaded_buildings.sort(key=lambda item: building_order.get(str(item or "").strip(), 10**6))
        failed_buildings.sort(key=lambda item: building_order.get(str(item.get("building", "")).strip(), 10**6))
        ordered_details = {}
        for building in sorted(details.keys(), key=lambda item: building_order.get(str(item or "").strip(), 10**6)):
            ordered_details[building] = details[building]
        total_ms = int((time.perf_counter() - batch_started) * 1000)
        emit_log(
            f"[交接班][云表最终上传][批量耗时] batch={batch_key or '-'}, "
            f"total_ms={total_ms}, mode=serial"
        )

        return {
            "status": status,
            "spreadsheet_token": spreadsheet_token,
            "spreadsheet_url": spreadsheet_url,
            "spreadsheet_title": spreadsheet_title,
            "uploaded_buildings": uploaded_buildings,
            "skipped_buildings": skipped_buildings,
            "failed_buildings": failed_buildings,
            "details": ordered_details,
        }

    def _build_base_sheet_cache(
        self,
        *,
        spreadsheet_token: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, List[Dict[str, Any]]]:
        cache: Dict[str, List[Dict[str, Any]]] = {}
        try:
            client = self._build_client()
            cache[str(spreadsheet_token or "").strip()] = client.query_sheets(spreadsheet_token)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][云表最终上传] 预取 sheet 列表失败，将在 worker 内独立查询: {exc}")
        return cache

    @staticmethod
    def _is_retryable_sheet_sync_error(error: Any) -> bool:
        text = str(error or "").strip().lower()
        if not text:
            return False
        return "90217" in text or "too many request" in text or "too many requests" in text

    def _run_sync_one_building_with_retry(
        self,
        *,
        spreadsheet_token: str,
        item: Dict[str, Any],
        emit_log: Callable[[str], None],
        base_sheet_cache: Dict[str, List[Dict[str, Any]]] | None = None,
    ) -> Dict[str, Any]:
        building = str(item.get("building", "")).strip()
        attempts = max(1, int(self.FINAL_SHEET_SERIAL_RETRY_ATTEMPTS or 1))
        for attempt in range(1, attempts + 1):
            result = self._sync_one_building(
                spreadsheet_token=spreadsheet_token,
                item=item,
                emit_log=emit_log,
                base_sheet_cache=base_sheet_cache,
            )
            if bool(result.get("success", False)):
                return result
            error_text = str(result.get("error", "")).strip()
            if attempt >= attempts or not self._is_retryable_sheet_sync_error(error_text):
                return result
            wait_seconds = float(self.FINAL_SHEET_SERIAL_RETRY_BACKOFF_SECONDS) * attempt
            emit_log(
                f"[交接班][云表最终上传] 限流重试 building={building}, "
                f"attempt={attempt}/{attempts}, wait_sec={wait_seconds:.1f}"
            )
            time.sleep(wait_seconds)
        return {
            "building": building,
            "success": False,
            "error": "unknown_retry_exit",
            "detail": {
                "status": "failed",
                "sheet_title": self._target_sheet_title(building),
                "synced_revision": 0,
                "rows": 0,
                "cols": 0,
                "merged": 0,
                "synced_row_count": 0,
                "synced_column_count": 0,
                "synced_merges": [],
                "dynamic_merge_signature": "",
                "error": "unknown_retry_exit",
            },
        }

    def _sync_one_building(
        self,
        *,
        spreadsheet_token: str,
        item: Dict[str, Any],
        emit_log: Callable[[str], None],
        base_sheet_cache: Dict[str, List[Dict[str, Any]]] | None = None,
    ) -> Dict[str, Any]:
        building = str(item.get("building", "")).strip()
        output_file = Path(str(item.get("output_file", "")).strip())
        revision = int(item.get("revision", 0) or 0)
        target_sheet_title = self._target_sheet_title(building)
        previous_cloud_sync = item.get("cloud_sheet_sync", {})
        timings = {
            "workbook_ms": 0,
            "snapshot_ms": 0,
            "ensure_sheet_ms": 0,
            "clear_ms": 0,
            "value_ms": 0,
            "dimension_ms": 0,
            "merge_ms": 0,
        }
        total_started = time.perf_counter()
        if not output_file.exists():
            error = f"output_file_not_found: {output_file}"
            detail = {
                "status": "failed",
                "sheet_title": target_sheet_title,
                "synced_revision": 0,
                "rows": 0,
                "cols": 0,
                "merged": 0,
                "synced_row_count": 0,
                "synced_column_count": 0,
                "synced_merges": [],
                "dynamic_merge_signature": "",
                "error": error,
            }
            return {"building": building, "success": False, "error": error, "detail": detail}

        client = self._build_client()
        sheet_cache = deepcopy(base_sheet_cache) if isinstance(base_sheet_cache, dict) else {}
        try:
            emit_log(f"[交接班][云表最终上传] 开始 building={building}, output={output_file.name}")
            workbook_started = time.perf_counter()
            workbook = openpyxl.load_workbook(output_file, data_only=False)
            self._add_elapsed_ms(timings, "workbook_ms", workbook_started)
            try:
                snapshot_started = time.perf_counter()
                source_ws = self._select_source_sheet(workbook)
                snapshot = self._collect_sheet_snapshot(source_ws)
                self._add_elapsed_ms(timings, "snapshot_ms", snapshot_started)
            finally:
                workbook.close()

            applied = self._overwrite_named_target_sheet(
                client=client,
                spreadsheet_token=spreadsheet_token,
                sheet_title=target_sheet_title,
                snapshot=snapshot,
                previous_cloud_sync=previous_cloud_sync if isinstance(previous_cloud_sync, dict) else {},
                emit_log=emit_log,
                sheet_cache=sheet_cache,
                timings=timings,
            )
            total_ms = int((time.perf_counter() - total_started) * 1000)
            detail = {
                "status": "success",
                "sheet_title": target_sheet_title,
                "synced_revision": revision,
                "rows": int(snapshot.get("max_row", 0) or 0),
                "cols": int(snapshot.get("max_column", 0) or 0),
                "merged": len(snapshot.get("fixed_header_merges", []) or [])
                + len(snapshot.get("dynamic_merges", []) or []),
                "synced_row_count": int(applied.get("synced_row_count", 0) or 0),
                "synced_column_count": int(applied.get("synced_column_count", 0) or 0),
                "synced_merges": self._normalize_merge_ranges(applied.get("synced_merges", [])),
                "dynamic_merge_signature": str(applied.get("dynamic_merge_signature", "")).strip(),
                "error": "",
            }
            emit_log(
                f"[交接班][云表最终上传][耗时] building={building}, total_ms={total_ms}, "
                f"workbook_ms={timings.get('workbook_ms', 0)}, "
                f"snapshot_ms={timings.get('snapshot_ms', 0)}, "
                f"ensure_sheet_ms={timings.get('ensure_sheet_ms', 0)}, "
                f"clear_ms={timings.get('clear_ms', 0)}, "
                f"value_ms={timings.get('value_ms', 0)}, "
                f"dimension_ms={timings.get('dimension_ms', 0)}, "
                f"merge_ms={timings.get('merge_ms', 0)}"
            )
            return {"building": building, "success": True, "error": "", "detail": detail}
        except Exception as exc:  # noqa: BLE001
            total_ms = int((time.perf_counter() - total_started) * 1000)
            detail = {
                "status": "failed",
                "sheet_title": target_sheet_title,
                "synced_revision": 0,
                "rows": 0,
                "cols": 0,
                "merged": 0,
                "synced_row_count": 0,
                "synced_column_count": 0,
                "synced_merges": [],
                "dynamic_merge_signature": "",
                "error": str(exc),
            }
            emit_log(f"[交接班][云表最终上传] 失败 building={building}, error={exc}")
            emit_log(
                f"[交接班][云表最终上传][耗时] building={building}, total_ms={total_ms}, "
                f"workbook_ms={timings.get('workbook_ms', 0)}, "
                f"snapshot_ms={timings.get('snapshot_ms', 0)}, "
                f"ensure_sheet_ms={timings.get('ensure_sheet_ms', 0)}, "
                f"clear_ms={timings.get('clear_ms', 0)}, "
                f"value_ms={timings.get('value_ms', 0)}, "
                f"dimension_ms={timings.get('dimension_ms', 0)}, "
                f"merge_ms={timings.get('merge_ms', 0)}"
            )
            return {"building": building, "success": False, "error": str(exc), "detail": detail}

    def sync_output(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        duty_date_text: str,
        shift_text: str,
        output_file: str,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        batch_meta = self.prepare_batch_spreadsheet(
            duty_date=duty_date,
            duty_date_text=duty_date_text,
            shift_text=shift_text,
            emit_log=emit_log,
        )
        if not bool(batch_meta.get("success", False)):
            return self._status_payload(
                attempted=True,
                success=False,
                status="failed",
                spreadsheet_token=str(batch_meta.get("spreadsheet_token", "")),
                spreadsheet_url=str(batch_meta.get("spreadsheet_url", "")),
                spreadsheet_title=str(batch_meta.get("spreadsheet_title", "")),
                sheet_title=self._target_sheet_title(building),
                error=str(batch_meta.get("error", "")).strip() or "prepare_failed",
            )
        result = self.sync_confirmed_buildings(
            batch_meta=batch_meta,
            building_items=[
                {
                    "building": building,
                    "output_file": output_file,
                    "revision": 1,
                    "duty_shift": duty_shift,
                    "cloud_sheet_sync": {},
                }
            ],
            emit_log=emit_log,
        )
        detail_map = result.get("details", {})
        detail = detail_map.get(building, {}) if isinstance(detail_map, dict) else {}
        success = str(detail.get("status", "")).strip().lower() == "success"
        return self._status_payload(
            attempted=True,
            success=success,
            status="success" if success else "failed",
            spreadsheet_token=str(result.get("spreadsheet_token", "")),
            spreadsheet_url=str(result.get("spreadsheet_url", "")),
            spreadsheet_title=str(result.get("spreadsheet_title", "")),
            sheet_title=self._target_sheet_title(building),
            synced_revision=int(detail.get("synced_revision", 0) or 0),
            last_attempt_revision=1,
            prepared_at=str(batch_meta.get("prepared_at", "")).strip(),
            error=str(detail.get("error", "")).strip(),
            synced_row_count=int(detail.get("synced_row_count", 0) or 0),
            synced_column_count=int(detail.get("synced_column_count", 0) or 0),
            synced_merges=detail.get("synced_merges", []),
            dynamic_merge_signature=str(detail.get("dynamic_merge_signature", "")).strip(),
        )

    def _find_or_create_target_spreadsheet(
        self,
        *,
        client: FeishuSheetsClientRuntime,
        duty_date: str,
        duty_date_text: str,
        shift_text: str,
    ) -> Dict[str, Any]:
        cfg = self._sync_cfg()
        pattern = str(
            cfg.get("spreadsheet_name_pattern", "") or "南通园区交接班日志-{date_text}{shift_text}"
        ).strip()
        title = pattern.format(
            date_text=str(duty_date_text or "").strip(),
            shift_text=str(shift_text or "").strip(),
            duty_date=str(duty_date or "").strip(),
        ).strip()
        if not title:
            raise ValueError("cloud_sheet_sync.spreadsheet_name_pattern 生成的标题为空")
        meta = client.find_or_create_date_spreadsheet(
            root_wiki_url=str(cfg.get("root_wiki_url", "") or "").strip(),
            template_node_token=str(cfg.get("template_node_token", "") or "").strip(),
            spreadsheet_title=title,
            duty_date=str(duty_date or "").strip(),
        )
        if not str(meta.get("spreadsheet_token", "") or "").strip():
            raise RuntimeError("云表定位失败: 缺少 spreadsheet_token")
        return meta

    def _target_sheet_title(self, building: str) -> str:
        cfg = self._sync_cfg()
        names = cfg.get("sheet_names", {})
        names = names if isinstance(names, dict) else {}
        title = str(names.get(str(building or "").strip(), "") or "").strip()
        return title or str(building or "").strip()

    def _select_source_sheet(self, workbook) -> Any:
        cfg = self._sync_cfg()
        source_sheet_name = str(cfg.get("source_sheet_name", "交接班日志") or "").strip() or "交接班日志"
        if source_sheet_name in workbook.sheetnames:
            return workbook[source_sheet_name]
        raise RuntimeError(f"missing_source_sheet: {source_sheet_name}")

    @staticmethod
    def _normalize_display_value(value: Any) -> Any:
        return value if value is not None else ""

    def _display_value_for_handover_cell(self, worksheet: Any, cell: Any) -> Any:
        value = cell.value
        if cell.data_type != "f":
            return self._normalize_display_value(value)

        coordinate = str(getattr(cell, "coordinate", "") or "").strip().upper()
        if coordinate == "B18":
            shift_text = str(worksheet["F2"].value or "").strip()
            return "9:00" if shift_text == "白班" else "21:00"
        if coordinate == "B19":
            shift_text = str(worksheet["F2"].value or "").strip()
            return "16:00" if shift_text == "白班" else "4:00"
        if coordinate in {"H18", "H19"}:
            return self._normalize_display_value(worksheet["C3"].value)
        raise RuntimeError(f"unsupported_formula_cell: {coordinate}")

    @staticmethod
    def _column_width_to_pixels(width: float | None) -> int:
        raw = float(width or 8.43)
        return max(20, int(raw * 7 + 5))

    @staticmethod
    def _row_height_to_pixels(height: float | None) -> int:
        raw = float(height or 15.0)
        return max(20, int(raw * 96 / 72))

    @staticmethod
    def _compress_dimension_ranges(items: List[Dict[str, int]]) -> List[Dict[str, int]]:
        normalized = [dict(item) for item in items if isinstance(item, dict)]
        if not normalized:
            return []
        normalized.sort(key=lambda item: (int(item.get("start_index", 0) or 0), int(item.get("end_index", 0) or 0)))
        output: List[Dict[str, int]] = []
        current = dict(normalized[0])
        for item in normalized[1:]:
            same_size = int(item.get("pixel_size", 0) or 0) == int(current.get("pixel_size", 0) or 0)
            contiguous = int(item.get("start_index", 0) or 0) == int(current.get("end_index", 0) or 0)
            if same_size and contiguous:
                current["end_index"] = int(item.get("end_index", current.get("end_index", 0)) or 0)
                continue
            output.append(current)
            current = dict(item)
        output.append(current)
        return output

    @staticmethod
    def _add_elapsed_ms(timings: Dict[str, int] | None, key: str, started_at: float) -> None:
        if not isinstance(timings, dict):
            return
        timings[key] = timings.get(key, 0) + int((time.perf_counter() - started_at) * 1000)

    def _collect_sheet_snapshot(self, worksheet) -> Dict[str, Any]:
        max_row = max(1, int(worksheet.max_row or 1))
        max_column = max(1, int(worksheet.max_column or 1))
        values: List[List[Any]] = []
        for row_idx in range(1, max_row + 1):
            value_row: List[Any] = []
            for col_idx in range(1, max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value_row.append(self._display_value_for_handover_cell(worksheet, cell))
            values.append(value_row)

        fixed_header_merges = []
        dynamic_merges = []
        for merged in worksheet.merged_cells.ranges:
            if int(merged.min_row) <= self.FIXED_HEADER_ROW_COUNT < int(merged.max_row):
                raise RuntimeError(f"unsupported_cross_boundary_merge: {merged}")
            normalized = {
                "start_row_index": int(merged.min_row) - 1,
                "end_row_index": int(merged.max_row),
                "start_column_index": int(merged.min_col) - 1,
                "end_column_index": int(merged.max_col),
            }
            if int(merged.max_row) <= self.FIXED_HEADER_ROW_COUNT:
                fixed_header_merges.append(normalized)
            elif int(merged.min_row) >= self.DYNAMIC_START_ROW:
                dynamic_merges.append(normalized)

        return {
            "max_row": max_row,
            "max_column": max_column,
            "values": values,
            "fixed_header_merges": self._normalize_merge_ranges(fixed_header_merges),
            "dynamic_merges": self._normalize_merge_ranges(dynamic_merges),
            "dynamic_merge_signature": self._build_dynamic_merge_signature(dynamic_merges),
        }

    def _ensure_target_sheets(
        self,
        *,
        client: FeishuSheetsClientRuntime,
        spreadsheet_token: str,
        sheet_cache: Dict[str, List[Dict[str, Any]]] | None = None,
    ) -> None:
        for index, building in enumerate(("A楼", "B楼", "C楼", "D楼", "E楼")):
            title = self._target_sheet_title(building)
            client.dedupe_named_sheets(spreadsheet_token, title, sheet_cache=sheet_cache)
            client.get_or_create_named_sheet(spreadsheet_token, title, index=index, sheet_cache=sheet_cache)

    def _ensure_named_target_sheet(
        self,
        *,
        client: FeishuSheetsClientRuntime,
        spreadsheet_token: str,
        sheet_title: str,
        index: int = 0,
        sheet_cache: Dict[str, List[Dict[str, Any]]] | None = None,
    ) -> Dict[str, Any]:
        deduped = client.dedupe_named_sheets(spreadsheet_token, sheet_title, sheet_cache=sheet_cache)
        if deduped:
            return deduped
        return client.get_or_create_named_sheet(
            spreadsheet_token,
            sheet_title,
            index=index,
            sheet_cache=sheet_cache,
        )

    @staticmethod
    def _range_name(sheet_title: str, row_count: int, col_count: int) -> str:
        safe_rows = max(1, int(row_count or 1))
        safe_cols = max(1, int(col_count or 1))
        end_cell = f"{get_column_letter(safe_cols)}{safe_rows}"
        return f"{sheet_title}!A1:{end_cell}"

    @staticmethod
    def _range_name_by_sheet_id(sheet_id: str, row_count: int, col_count: int) -> str:
        safe_rows = max(1, int(row_count or 1))
        safe_cols = max(1, int(col_count or 1))
        end_cell = f"{get_column_letter(safe_cols)}{safe_rows}"
        return f"{str(sheet_id or '').strip()}!A1:{end_cell}"

    @staticmethod
    def _merge_key(item: Dict[str, int]) -> tuple[int, int, int, int]:
        return (
            int(item.get("start_row_index", 0)),
            int(item.get("end_row_index", 0)),
            int(item.get("start_column_index", 0)),
            int(item.get("end_column_index", 0)),
        )

    def _merge_union(
        self,
        left: List[Dict[str, int]] | None,
        right: List[Dict[str, int]] | None,
    ) -> List[Dict[str, int]]:
        merged: Dict[tuple[int, int, int, int], Dict[str, int]] = {}
        for item in self._normalize_merge_ranges(left or []) + self._normalize_merge_ranges(right or []):
            merged[self._merge_key(item)] = item
        return list(merged.values())

    def _build_dynamic_merge_signature(self, merges: List[Dict[str, int]] | None) -> str:
        normalized = sorted(self._normalize_merge_ranges(merges or []), key=self._merge_key)
        return hashlib.sha1(
            json.dumps(normalized, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        ).hexdigest()

    def _build_writable_ranges(
        self,
        *,
        sheet_id: str,
        values: List[List[Any]],
        effective_merges: List[Dict[str, int]],
        clear_rows: int,
        clear_cols: int,
    ) -> List[Dict[str, Any]]:
        safe_rows = max(1, int(clear_rows or 1))
        safe_cols = max(1, int(clear_cols or 1))
        current_rows = len(values)
        current_cols = max((len(row) for row in values), default=0)
        merged_interior: set[tuple[int, int]] = set()
        for merge in self._normalize_merge_ranges(effective_merges):
            start_row = int(merge.get("start_row_index", 0) or 0)
            end_row = int(merge.get("end_row_index", 0) or 0)
            start_col = int(merge.get("start_column_index", 0) or 0)
            end_col = int(merge.get("end_column_index", 0) or 0)
            for row_idx in range(start_row, end_row):
                for col_idx in range(start_col, end_col):
                    if row_idx == start_row and col_idx == start_col:
                        continue
                    merged_interior.add((row_idx, col_idx))

        ranges: List[Dict[str, Any]] = []
        for row_idx in range(safe_rows):
            col_idx = 0
            while col_idx < safe_cols:
                if (row_idx, col_idx) in merged_interior:
                    col_idx += 1
                    continue
                start_col = col_idx
                row_values: List[Any] = []
                while col_idx < safe_cols and (row_idx, col_idx) not in merged_interior:
                    if row_idx < current_rows and col_idx < current_cols:
                        row_values.append(values[row_idx][col_idx])
                    else:
                        row_values.append("")
                    col_idx += 1
                if not row_values:
                    continue
                ranges.append(
                    {
                        "range": FeishuSheetsClientRuntime.build_sheet_id_range(
                            sheet_id=sheet_id,
                            start_row_index=row_idx,
                            end_row_index=row_idx + 1,
                            start_column_index=start_col,
                            end_column_index=start_col + len(row_values),
                        ),
                        "values": [row_values],
                    }
                )
        return ranges

    def _overwrite_named_target_sheet(
        self,
        *,
        client: FeishuSheetsClientRuntime,
        spreadsheet_token: str,
        sheet_title: str,
        snapshot: Dict[str, Any],
        previous_cloud_sync: Dict[str, Any],
        emit_log: Callable[[str], None],
        sheet_cache: Dict[str, List[Dict[str, Any]]] | None = None,
        timings: Dict[str, int] | None = None,
    ) -> Dict[str, Any]:
        ensure_started = time.perf_counter()
        target_sheet = self._ensure_named_target_sheet(
            client=client,
            spreadsheet_token=spreadsheet_token,
            sheet_title=sheet_title,
            sheet_cache=sheet_cache,
        )
        self._add_elapsed_ms(timings, "ensure_sheet_ms", ensure_started)
        sheet_id = str(target_sheet.get("sheet_id", "")).strip()
        if not sheet_id:
            raise RuntimeError(f"目标 sheet 缺少 sheet_id: {sheet_title}")

        current_rows = max(1, int(snapshot.get("max_row", 1) or 1))
        current_cols = max(1, int(snapshot.get("max_column", 1) or 1))
        previous_rows = max(0, int(previous_cloud_sync.get("synced_row_count", 0) or 0))
        previous_cols = max(0, int(previous_cloud_sync.get("synced_column_count", 0) or 0))
        clear_rows = max(previous_rows, current_rows, 1)
        clear_cols = max(previous_cols, current_cols, 1)
        current_capacity_rows = max(0, int(target_sheet.get("row_count", 0) or 0))
        current_capacity_cols = max(0, int(target_sheet.get("column_count", 0) or 0))
        if current_capacity_rows < clear_rows:
            dimension_started = time.perf_counter()
            client.add_dimension(
                spreadsheet_token,
                sheet_id=sheet_id,
                major_dimension="ROWS",
                length=clear_rows - current_capacity_rows,
            )
            self._add_elapsed_ms(timings, "dimension_ms", dimension_started)
            current_capacity_rows = clear_rows
        if current_capacity_cols < clear_cols:
            dimension_started = time.perf_counter()
            client.add_dimension(
                spreadsheet_token,
                sheet_id=sheet_id,
                major_dimension="COLUMNS",
                length=clear_cols - current_capacity_cols,
            )
            self._add_elapsed_ms(timings, "dimension_ms", dimension_started)
            current_capacity_cols = clear_cols
        if isinstance(sheet_cache, dict):
            cache_key = str(spreadsheet_token or "").strip()
            cached_items = sheet_cache.get(cache_key, [])
            refreshed_items = []
            for item in cached_items:
                current_item = dict(item)
                if str(current_item.get("sheet_id", "")).strip() == sheet_id:
                    current_item["row_count"] = max(current_capacity_rows, clear_rows)
                    current_item["column_count"] = max(current_capacity_cols, clear_cols)
                refreshed_items.append(current_item)
            if refreshed_items:
                sheet_cache[cache_key] = refreshed_items

        previous_dynamic_merges = self._normalize_merge_ranges(previous_cloud_sync.get("synced_merges", []))
        current_fixed_merges = self._normalize_merge_ranges(snapshot.get("fixed_header_merges", []))
        current_dynamic_merges = self._normalize_merge_ranges(snapshot.get("dynamic_merges", []))
        current_dynamic_signature = str(snapshot.get("dynamic_merge_signature", "")).strip()
        previous_dynamic_signature = str(previous_cloud_sync.get("dynamic_merge_signature", "")).strip()
        dynamic_merge_changed = current_dynamic_signature != previous_dynamic_signature

        emit_log(
            f"[交接班][云表覆写] building={sheet_title}, spreadsheet={spreadsheet_token}, "
            f"sheet={sheet_title}, clear_rows={clear_rows}, clear_cols={clear_cols}"
        )
        emit_log(f"[交接班][云表覆写] dynamic_merge_changed={str(dynamic_merge_changed).lower()}")

        if dynamic_merge_changed:
            merge_started = time.perf_counter()
            if previous_dynamic_merges:
                client.batch_unmerge_cells(spreadsheet_token, sheet_id, previous_dynamic_merges)
            if current_dynamic_merges:
                client.batch_merge_cells(spreadsheet_token, sheet_id, current_dynamic_merges)
            self._add_elapsed_ms(timings, "merge_ms", merge_started)

        value_started = time.perf_counter()
        value_ranges = self._build_writable_ranges(
            sheet_id=sheet_id,
            values=snapshot.get("values", []),
            effective_merges=self._merge_union(current_fixed_merges, current_dynamic_merges),
            clear_rows=clear_rows,
            clear_cols=clear_cols,
        )
        if value_ranges:
            client.batch_update_values(spreadsheet_token, value_ranges)
        self._add_elapsed_ms(timings, "value_ms", value_started)

        emit_log(
            f"[交接班][云表覆写] 覆写完成 building={sheet_title}, sheet={sheet_title}, rows={current_rows}, cols={current_cols}"
        )
        return {
            "sheet_id": sheet_id,
            "sheet_title": sheet_title,
            "synced_row_count": current_rows,
            "synced_column_count": current_cols,
            "synced_merges": current_dynamic_merges,
            "dynamic_merge_signature": current_dynamic_signature,
        }
