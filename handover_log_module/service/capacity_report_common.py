from __future__ import annotations

import re
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

from handover_log_module.core.models import RawRow
from handover_log_module.core.normalizers import format_number


_WEST_PUBLIC_CELLS = {"first": "G22", "second": "J22", "tank": "M22"}
_EAST_PUBLIC_CELLS = {"first": "T22", "second": "W22", "tank": "Z22"}
_BLOCK_COL_OFFSET = column_index_from_string("Q") - column_index_from_string("D")
_BLOCK_ROW_OFFSET = 10
_WEST_BLOCK_BASE_CELLS = {
    "title": "D23",
    "current_or_power": "D25",
    "chilled_out_temp": "D27",
    "cooling_in_temp": "D28",
    "condenser_pressure": "D29",
    "evaporator_delta": "D30",
    "condenser_delta": "D31",
    "fan_1": "F27",
    "fan_2": "G27",
    "cooling_pump_freq": "I26",
    "primary_pump_freq": "J26",
    "plate_cooling_in_temp": "L25",
    "plate_cooling_out_temp": "L26",
    "plate_chilled_in_temp": "L28",
    "plate_chilled_out_temp": "L29",
    "plate_cooling_in_pressure": "N25",
    "plate_cooling_in_pressure_dup": "J29",
    "plate_chilled_in_pressure": "N29",
    "plate_chilled_out_pressure": "N30",
    "cooling_out_temp": "F29",
    "cooling_tower_out_temp": "F30",
    "ph": "I30",
    "conductivity": "I31",
}
_TEMPLATE_FAMILY_OTHER_BUILDINGS = "other_buildings"
_TEMPLATE_FAMILY_E_BUILDING = "e_building"
_AIRCON_TARGET_CELLS_BY_TEMPLATE_FAMILY = {
    _TEMPLATE_FAMILY_OTHER_BUILDINGS: {
        (2, "west", "south"): "AE70",
        (2, "west", "north"): "AE75",
        (2, "east", "south"): "AE80",
        (2, "east", "north"): "AE85",
        (3, "west", "south"): "AE91",
        (3, "west", "north"): "AE98",
        (3, "east", "south"): "AE105",
        (3, "east", "north"): "AE112",
        (4, "west", "south"): "AE118",
        (4, "west", "north"): "AE123",
        (4, "east", "south"): "AE128",
        (4, "east", "north"): "AE133",
    },
    _TEMPLATE_FAMILY_E_BUILDING: {
        (2, "west", "south"): "AE72",
        (2, "west", "north"): "AE82",
        (2, "east", "south"): "AE92",
        (2, "east", "north"): "AE102",
        (3, "west", "south"): "AE112",
        (3, "west", "north"): "AE122",
        (3, "east", "south"): "AE132",
        (3, "east", "north"): "AE142",
        (4, "west", "south"): "AE152",
        (4, "west", "north"): "AE162",
        (4, "east", "south"): "AE172",
        (4, "east", "north"): "AE182",
    },
}
_AIRCON_ZONE_DIRECTION_BY_AREA = {
    "1": ("east", "south"),
    "2": ("east", "north"),
    "3": ("west", "south"),
    "4": ("west", "north"),
}
_DEFAULT_PRIMARY_PUMP_ALIASES = {
    "A楼": ["冷冻水一次泵变频反馈"],
    "B楼": ["冷冻水一次泵变频反馈"],
    "C楼": ["一次冷冻泵频率反馈"],
    "D楼": ["冷冻泵频率反馈"],
    "E楼": ["一次冷冻泵频率反馈"],
}
_DEFAULT_REGION_ALIAS_GROUPS = {
    "current_or_power": ["冷机_电流百分比", "电机电流百分比", "冷机电流百分比", "电机功率"],
    "chilled_out_temp": ["冷机_冷冻水出水温度", "冷冻单元-冷冻水出水温度", "冷机冷冻水出水温度"],
    "cooling_in_temp": ["冷机_冷却水进水温度", "冷冻单元-冷却水进水温度", "冷机冷却水进水温度"],
    "condenser_pressure": ["冷凝器压力"],
    "evaporator_delta": ["蒸发器小温差", "蒸发器换热温差", "蒸发器端温差"],
    "condenser_delta": ["冷凝器小温差", "冷凝器换热温差", "冷凝器端温差"],
    "cooling_pump_freq": ["冷却水泵变频反馈", "冷却泵频率反馈"],
    "plate_cooling_in_temp": ["板换冷却水进水温度", "冷却水侧板换进口温度"],
    "plate_cooling_out_temp": ["板换冷却水出水温度", "冷却水侧板换出口温度"],
    "plate_chilled_in_temp": ["板交冷冻回水温度", "冷冻水侧板换进口温度", "板换冷冻水进水温度"],
    "plate_chilled_out_temp": ["板交冷冻供水温度", "冷冻水侧板换出口温度", "板换冷冻水出水温度"],
    "plate_cooling_in_pressure": ["冷却水侧板换进口压力", "冷机冷却水进口压力"],
    "plate_chilled_in_pressure": ["冷冻水侧板换进口压力", "板换冷冻水进口压力"],
    "plate_chilled_out_pressure": ["冷冻水侧板换出口压力", "板换冷冻水出口压力"],
    "cooling_out_temp": ["冷机侧冷却水出水温度", "冷冻单元-冷却水出水温度", "冷机冷却水出水温度", "冷却水出水温度（相对于冷机）"],
    "cooling_tower_out_temp": ["冷却塔出口温度", "冷却塔出水温度"],
    "ph": ["ph"],
    "conductivity": ["加药装置_电导率信号", "加药装置_电导率", "冷却水电导率", "towercondtor"],
}
_FLOW_ALIASES = {
    "primary": ["一次总流量", "单元总流量", "一次侧总流量"],
    "secondary": ["二侧总流量", "二次总流量", "二次侧流量"],
    "tank": ["盈亏管1流量计流量", "蓄冷罐流量", "蓄冷罐总管道流量"],
}
_CAPACITY_SOURCE_DIRECT_ALIASES = {
    "storage_total": ["蓄水池总储水量"],
    "oil_amount": ["油量"],
}
_CAPACITY_SOURCE_DIRECT_EXCLUDES = {
    "oil_amount": ["油量后备时间", "燃油后备时间"],
}
_E_BUILDING_FLOW_ALIASES = {
    "return_1": ["冷冻水回水流量_1"],
    "return_2": ["冷冻水回水流量_2"],
    "tank": ["蓄冷罐总管道流量"],
}
_SECONDARY_PUMP_ALIASES = ["冷冻水二次泵变频反馈", "二次冷冻泵频率反馈", "二次泵频率反馈", "冷冻水二次泵频率反馈"]
_TANK_LEVEL_ALIASES = ["水池液位", "蓄水罐液位", "补水罐液位"]
_ACTIVE_CHILLER_MODE_TEXTS = {"制冷", "预冷"}
_E_BUILDING_CHILLER_SKIP_KEYS = {
    "current_or_power",
    "chilled_out_temp",
    "cooling_in_temp",
    "condenser_pressure",
    "evaporator_delta",
    "condenser_delta",
}
def _text(value: Any) -> str:
    return str(value or "").strip()


def _casefold(value: Any) -> str:
    return _text(value).casefold()


def _alarm_count(value: Any) -> str:
    try:
        return str(int(value or 0))
    except Exception:  # noqa: BLE001
        text = _text(value)
        return text or "0"


def _split_metric_pair(value: Any) -> tuple[str, str]:
    text = _text(value)
    if not text:
        return "", ""
    if "/" not in text:
        return "", text
    left, right = text.split("/", 1)
    return _text(left), _text(right)


def _parse_tank_backup_pair(value: Any) -> tuple[str, str]:
    text = _text(value)
    if not text:
        return "", ""
    west_match = re.search(r"西区\s*([+-]?\d+(?:\.\d+)?)", text)
    east_match = re.search(r"东区\s*([+-]?\d+(?:\.\d+)?)", text)
    if west_match and east_match:
        return _text(west_match.group(1)), _text(east_match.group(1))
    numbers = re.findall(r"[+-]?\d+(?:\.\d+)?", text)
    if len(numbers) >= 2:
        return _text(numbers[0]), _text(numbers[1])
    return "", ""


def _to_float_text(value: Any) -> float | None:
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _extract_building_code(building: Any) -> str:
    match = re.search(r"([A-Za-z])", _text(building))
    return match.group(1).upper() if match else ""


def _offset_cell(cell_name: str, *, row_offset: int = 0, col_offset: int = 0) -> str:
    column, row = coordinate_from_string(str(cell_name).strip().upper())
    return f"{get_column_letter(column_index_from_string(column) + col_offset)}{int(row) + row_offset}"


def _zone_of_row(row: RawRow) -> str:
    combined = f"{_text(getattr(row, 'b_text', ''))} {_text(getattr(row, 'c_text', ''))}"
    if "150" in combined or "西区" in combined:
        return "west"
    if "124" in combined or "东区" in combined:
        return "east"
    return ""


def _unit_from_row(row: RawRow) -> int | None:
    text = _text(getattr(row, "c_text", ""))
    lowered = text.casefold()
    if ("150" in text or "西区" in text) and ("101" in text or "一号冷机" in text or "1号冷机" in lowered):
        return 1
    if ("150" in text or "西区" in text) and ("102" in text or "二号冷机" in text or "2号冷机" in lowered):
        return 2
    if ("150" in text or "西区" in text) and ("103" in text or "三号冷机" in text or "3号冷机" in lowered):
        return 3
    if ("124" in text or "东区" in text) and ("101" in text or "四号冷机" in text or "4号冷机" in lowered):
        return 4
    if ("124" in text or "东区" in text) and ("102" in text or "五号冷机" in text or "5号冷机" in lowered):
        return 5
    if ("124" in text or "东区" in text) and ("103" in text or "六号冷机" in text or "6号冷机" in lowered):
        return 6
    if "一号冷机" in text or "1号冷机" in lowered:
        return 1
    if "二号冷机" in text or "2号冷机" in lowered:
        return 2
    if "三号冷机" in text or "3号冷机" in lowered:
        return 3
    if "四号冷机" in text or "4号冷机" in lowered:
        return 4
    if "五号冷机" in text or "5号冷机" in lowered:
        return 5
    if "六号冷机" in text or "6号冷机" in lowered:
        return 6
    return None


def _match_contains(text: str, needle: str) -> bool:
    return _casefold(needle) in _casefold(text)


def _compile_placeholder_regex(pattern: str) -> re.Pattern[str]:
    normalized = str(pattern or "").replace("/d", r"\d")
    return re.compile(normalized, flags=re.IGNORECASE)


class CapacitySourceQuery:
    def __init__(self, rows: Iterable[RawRow] | None) -> None:
        self.rows: List[RawRow] = [row for row in (rows or []) if isinstance(row, RawRow)]

    def _iter_rows(
        self,
        *,
        zone: str = "",
        unit: int | None = None,
        allow_global: bool = False,
        b_contains: Sequence[str] | None = None,
        c_contains: Sequence[str] | None = None,
        c_regexes: Sequence[str] | None = None,
    ) -> List[RawRow]:
        normalized_zone = _text(zone).lower()
        target_rows: List[RawRow] = []
        for row in self.rows:
            row_zone = _zone_of_row(row)
            row_unit = _unit_from_row(row)
            if normalized_zone and row_zone != normalized_zone:
                continue
            if unit is not None and row_unit != int(unit):
                continue
            if b_contains and not any(_match_contains(getattr(row, "b_text", ""), needle) for needle in b_contains):
                continue
            if c_contains and not any(_match_contains(getattr(row, "c_text", ""), needle) for needle in c_contains):
                continue
            if c_regexes and not any(_compile_placeholder_regex(pattern).search(_text(getattr(row, "c_text", ""))) for pattern in c_regexes):
                continue
            target_rows.append(row)
        if target_rows or not allow_global or (not normalized_zone and unit is None):
            return target_rows
        return self._iter_rows(zone="", unit=None, allow_global=False, b_contains=b_contains, c_contains=c_contains, c_regexes=c_regexes)

    def first_row_by_d_aliases(
        self,
        d_aliases: Sequence[str],
        *,
        zone: str = "",
        unit: int | None = None,
        allow_global: bool = False,
        b_contains: Sequence[str] | None = None,
        c_contains: Sequence[str] | None = None,
        c_regexes: Sequence[str] | None = None,
    ) -> RawRow | None:
        candidate_rows = self._iter_rows(zone=zone, unit=unit, allow_global=allow_global, b_contains=b_contains, c_contains=c_contains, c_regexes=c_regexes)
        if not candidate_rows:
            return None
        for alias in d_aliases:
            alias_cf = _casefold(alias)
            for row in candidate_rows:
                if alias_cf and alias_cf in _casefold(getattr(row, "d_name", "")):
                    return row
        return None

    def rows_by_d_regexes(
        self,
        d_regexes: Sequence[str],
        *,
        zone: str = "",
        unit: int | None = None,
        allow_global: bool = False,
        b_contains: Sequence[str] | None = None,
        c_contains: Sequence[str] | None = None,
        c_regexes: Sequence[str] | None = None,
    ) -> List[RawRow]:
        candidate_rows = self._iter_rows(zone=zone, unit=unit, allow_global=allow_global, b_contains=b_contains, c_contains=c_contains, c_regexes=c_regexes)
        if not candidate_rows:
            return []
        compiled = [_compile_placeholder_regex(pattern) for pattern in d_regexes if _text(pattern)]
        if not compiled:
            return []
        return [row for row in candidate_rows if any(pattern.search(_text(getattr(row, "d_name", ""))) for pattern in compiled)]

    def first_text_by_d_aliases(self, d_aliases: Sequence[str], **kwargs: Any) -> str:
        row = self.first_row_by_d_aliases(d_aliases, **kwargs)
        return _text(getattr(row, "e_raw", None)) if row is not None else ""

    def first_number_by_d_aliases(self, d_aliases: Sequence[str], **kwargs: Any) -> float | None:
        row = self.first_row_by_d_aliases(d_aliases, **kwargs)
        return getattr(row, "value", None) if row is not None else None

    def first_row_by_identifier(
        self,
        *,
        identifier_tokens: Sequence[str],
        search_column: str,
        d_aliases: Sequence[str] | None = None,
    ) -> RawRow | None:
        tokens = [_casefold(token) for token in identifier_tokens if _text(token)]
        if not tokens:
            return None
        target_attr = "c_text" if _text(search_column).lower() == "c" else "b_text"
        candidate_rows = [row for row in self.rows if any(token in _casefold(getattr(row, target_attr, "")) for token in tokens)]
        if not candidate_rows:
            return None
        if not d_aliases:
            return candidate_rows[0]
        for alias in d_aliases:
            alias_cf = _casefold(alias)
            for row in candidate_rows:
                if alias_cf and alias_cf in _casefold(getattr(row, "d_name", "")):
                    return row
        return None

    def first_text_by_identifier(self, *, identifier_tokens: Sequence[str], search_column: str, d_aliases: Sequence[str] | None = None) -> str:
        row = self.first_row_by_identifier(identifier_tokens=identifier_tokens, search_column=search_column, d_aliases=d_aliases)
        return _text(getattr(row, "e_raw", None)) if row is not None else ""


def _format_calc_value(value: float | None) -> str:
    return format_number(value) if value is not None else ""


def _build_redundancy_text(count: int) -> str:
    safe_count = max(0, int(count))
    return f"{safe_count}用（{max(0, 3 - safe_count)}）备"


def _building_aliases(context: Dict[str, Any], key: str) -> List[str]:
    building = _text(context.get("building"))
    if key == "primary_pump_aliases":
        return list(_DEFAULT_PRIMARY_PUMP_ALIASES.get(building, []))
    if key == "chilled_out_extra_aliases":
        return ["冷冻水出水温度"] if building == "E楼" else []
    if key == "cooling_in_extra_aliases":
        return ["冷却水进水温度"] if building == "E楼" else []
    return []


def _prefix_building_identifier(raw_text: str, building_code: str) -> str:
    text = _text(raw_text)
    code = _text(building_code).upper()
    if not text or not code:
        return text
    if text.startswith("-"):
        return f"{code}{text}"
    if re.match(r"^[A-Za-z]-", text):
        return re.sub(r"^[A-Za-z](?=-)", code, text, count=1)
    return f"{code}-{text}"


def _identifier_search_tokens(identifier: str, *, kind: str) -> List[str]:
    tokens = [_text(identifier).upper()]
    if kind == "tr":
        tokens.append(tokens[0].replace("-TR-", "-TR"))
        tokens.append(tokens[0].replace("-TR", "-TR-"))
    deduped: List[str] = []
    for token in tokens:
        if token and token not in deduped:
            deduped.append(token)
    return deduped


def _tr_replacement_search_tokens(identifier: str) -> List[str]:
    text = _text(identifier).upper()
    if not text:
        return []
    transformed = re.sub(r"-TR-?", "-TRB", text, count=1)
    match = re.search(r"^(.*-TRB)-?(101|102|201|202)$", transformed)
    if not match:
        return [transformed] if transformed else []
    prefix = match.group(1)
    suffix = match.group(2)
    mapped_suffix = {
        "101": "101",
        "201": "101",
        "102": "201",
        "202": "201",
    }.get(suffix, suffix)
    tokens: List[str] = [f"{prefix}{mapped_suffix}", f"{prefix}-{mapped_suffix}"]
    deduped: List[str] = []
    for token in tokens:
        if token and token not in deduped:
            deduped.append(token)
    return deduped


def _hvdc_search_tokens(identifier: str, *, template_family: str) -> List[str]:
    text = _text(identifier).upper()
    if not text:
        return []
    tokens: List[str] = []
    if template_family == _TEMPLATE_FAMILY_OTHER_BUILDINGS:
        shifted = re.sub(r"(\d)$", "2", text)
        if shifted and shifted != text:
            tokens.append(shifted)
    tokens.append(text)
    deduped: List[str] = []
    for token in tokens:
        if token and token not in deduped:
            deduped.append(token)
    return deduped


def build_capacity_template_snapshot(sheet: Worksheet, building: str) -> Dict[str, Any]:
    building_code = _extract_building_code(building)
    template_family = (
        _TEMPLATE_FAMILY_E_BUILDING if _text(building) == "E楼" else _TEMPLATE_FAMILY_OTHER_BUILDINGS
    )

    def _search_tokens_for(identifier: str, *, kind: str) -> List[str]:
        if kind == "hvdc":
            return _hvdc_search_tokens(identifier, template_family=template_family)
        return _identifier_search_tokens(identifier, kind=kind)

    def _collect(column_letter: str, start_row: int, end_row: int, *, kind: str) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        for row_idx in range(start_row, end_row + 1):
            raw_text = _text(sheet[f"{column_letter}{row_idx}"].value)
            if not raw_text:
                continue
            identifier = _prefix_building_identifier(raw_text, building_code)
            items.append({"row": row_idx, "identifier": identifier, "search_tokens": _search_tokens_for(identifier, kind=kind)})
        return items

    def _collect_rows(column_letter: str, rows: Sequence[int], *, kind: str) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        for row_idx in rows:
            raw_text = _text(sheet[f"{column_letter}{row_idx}"].value)
            if not raw_text:
                continue
            identifier = _prefix_building_identifier(raw_text, building_code)
            items.append({"row": row_idx, "identifier": identifier, "search_tokens": _search_tokens_for(identifier, kind=kind)})
        return items

    def _collect_merged_anchor_rows(column_letter: str, start_row: int, end_row: int, *, kind: str) -> List[Dict[str, Any]]:
        target_col = column_index_from_string(column_letter)
        anchor_rows: List[int] = []
        for merged_range in getattr(sheet.merged_cells, "ranges", []):
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col != target_col or max_col != target_col:
                continue
            if min_row < start_row or min_row > end_row:
                continue
            if max_row < start_row:
                continue
            raw_text = _text(sheet[f"{column_letter}{min_row}"].value)
            if not raw_text:
                continue
            if min_row not in anchor_rows:
                anchor_rows.append(min_row)
        anchor_rows.sort()
        if anchor_rows:
            return _collect_rows(column_letter, anchor_rows, kind=kind)
        return _collect_rows(column_letter, list(range(start_row, end_row + 1)), kind=kind)

    return {
        "building_code": building_code,
        "template_family": template_family,
        "tr_entries": _collect_merged_anchor_rows("B", 67, 202, kind="tr"),
        "ups_entries": _collect_merged_anchor_rows("I", 67, 202, kind="ups"),
        "hvdc_entries": _collect("O", 67, 186, kind="hvdc"),
        "rpp_entries": _collect("S", 67, 186, kind="rpp"),
    }


def build_common_capacity_cell_values(context: Dict[str, Any]) -> Dict[str, str]:
    duty_shift = _text(context.get("duty_shift")).lower()
    handover_cells = context.get("handover_cells", {}) if isinstance(context.get("handover_cells", {}), dict) else {}
    roster = context.get("roster", {}) if isinstance(context.get("roster", {}), dict) else {}
    previous_alarm = context.get("previous_alarm_summary", {}) if isinstance(context.get("previous_alarm_summary", {}), dict) else {}
    current_alarm = context.get("current_alarm_summary", {}) if isinstance(context.get("current_alarm_summary", {}), dict) else {}
    oil_previous = context.get("oil_previous", {}) if isinstance(context.get("oil_previous", {}), dict) else {}
    oil_current = context.get("oil_current", {}) if isinstance(context.get("oil_current", {}), dict) else {}
    weather_text = _text(context.get("weather_text"))
    weather_humidity = _text(context.get("weather_humidity"))
    water_summary = (
        context.get("capacity_water_summary", {})
        if isinstance(context.get("capacity_water_summary", {}), dict)
        else {}
    )
    if not water_summary and isinstance(context.get("night_water_summary", {}), dict):
        water_summary = context.get("night_water_summary", {})
    tank_west, tank_east = _parse_tank_backup_pair(handover_cells.get("F8"))
    h16_left, h16_right = _split_metric_pair(handover_cells.get("B10"))
    h17_left, h17_right = _split_metric_pair(context.get("hvdc_text"))
    h18_left, h18_right = _split_metric_pair(handover_cells.get("D10"))

    cell_values: Dict[str, str] = {
        "D2": "白" if duty_shift == "day" else "夜",
        "M6": _text(handover_cells.get("C3")),
        "R2": _text(handover_cells.get("B7")),
        "AB2": _text(handover_cells.get("D7")),
        "G6": _text(roster.get("current_team")),
        "P6": _text(roster.get("next_team")),
        "U6": _text(handover_cells.get("G3")),
        "S7": _text(handover_cells.get("B4" if duty_shift == "day" else "F4")),
        "G8": f"接班未恢复告警：{_alarm_count(previous_alarm.get('unrecovered_count'))}",
        "L8": _alarm_count(previous_alarm.get("total_count")),
        "S8": _text(previous_alarm.get("accept_description")) or "/",
        "G9": f"交班未恢复告警：{_alarm_count(current_alarm.get('unrecovered_count'))}",
        "L9": _alarm_count(current_alarm.get("total_count")),
        "S9": _text(current_alarm.get("accept_description")) or "/",
        "H16": h16_right,
        "L16": h16_left,
        "H17": h17_right,
        "L17": h17_left,
        "H18": h18_right,
        "L18": h18_left,
        "U12": _text(oil_previous.get("first")),
        "X12": _text(oil_previous.get("second")),
        "U13": _text(oil_current.get("first")),
        "X13": _text(oil_current.get("second")),
        "U15": _text(handover_cells.get("H6")),
        "AD22": tank_west,
        "AD23": tank_east,
        "V60": _text(handover_cells.get("B6")),
        "O60": _text(handover_cells.get("D6")),
        "S60": _text(handover_cells.get("F6")),
        "AB56": _text(handover_cells.get("B13")),
        "AC56": _text(handover_cells.get("D13")),
        "L2": weather_text,
        "X2": weather_humidity,
        "O57": _text(water_summary.get("latest_daily_total")),
        "AC25": _text(water_summary.get("month_total")),
        "R57": _text(water_summary.get("month_total")),
    }
    if duty_shift == "day":
        cell_values["G7"] = _text(handover_cells.get("B4")) or "/"
    return {cell: value for cell, value in cell_values.items() if value != ""}


def _build_meter_values(query: CapacitySourceQuery) -> Dict[str, str]:
    values: Dict[str, str] = {}
    for row in query._iter_rows(
        c_regexes=[
            ".*\\d#.*\\u5e02\\u7535\\u8fdb\\u7ebf_\\u7535\\u91cf\\u4eea.*",
            ".*\\d#\\u5e02\\u7535\\u8fdb\\u7ebf.*\\u7535\\u91cf\\u4eea.*",
        ]
    ):
        match = re.search(r"([1-4])\s*#", _text(getattr(row, "c_text", "")))
        if not match:
            continue
        target_cell = f"G{11 + int(match.group(1))}"
        if target_cell in values:
            continue
        values[target_cell] = _text(getattr(row, "e_raw", None))
    return {cell: value for cell, value in values.items() if value != ""}


def _first_text_by_d_contains(
    query: CapacitySourceQuery,
    needles: Sequence[str],
    *,
    excludes: Sequence[str] | None = None,
) -> str:
    normalized_needles = [_casefold(item) for item in needles if _text(item)]
    normalized_excludes = [_casefold(item) for item in (excludes or []) if _text(item)]
    if not normalized_needles:
        return ""
    for needle in normalized_needles:
        for row in query.rows:
            d_name = _casefold(getattr(row, "d_name", ""))
            if not d_name or needle not in d_name:
                continue
            if normalized_excludes and any(token in d_name for token in normalized_excludes):
                continue
            value_text = _text(getattr(row, "e_raw", None))
            if value_text:
                return value_text
    return ""


def _resolve_public_flow_values(query: CapacitySourceQuery, *, zone: str, context: Dict[str, Any]) -> Dict[str, str]:
    building = _text(context.get("building"))
    cells = _WEST_PUBLIC_CELLS if zone == "west" else _EAST_PUBLIC_CELLS
    if building == "E楼":
        return_1 = query.first_number_by_d_aliases(_E_BUILDING_FLOW_ALIASES["return_1"], zone=zone, allow_global=True)
        return_2 = query.first_number_by_d_aliases(_E_BUILDING_FLOW_ALIASES["return_2"], zone=zone, allow_global=True)
        tank_flow = query.first_number_by_d_aliases(_E_BUILDING_FLOW_ALIASES["tank"], zone=zone, allow_global=True)
        total_1 = None if return_1 is None and return_2 is None and tank_flow is None else float(return_1 or 0) + float(return_2 or 0) + float(tank_flow or 0)
        total_2 = None if return_1 is None and return_2 is None else float(return_1 or 0) + float(return_2 or 0)
        return {cells["first"]: _format_calc_value(total_1), cells["second"]: _format_calc_value(total_2), cells["tank"]: _format_calc_value(tank_flow)}
    first_text = query.first_text_by_d_aliases(_FLOW_ALIASES["primary"], zone=zone, allow_global=True)
    second_text = query.first_text_by_d_aliases(_FLOW_ALIASES["secondary"], zone=zone, allow_global=True)
    tank_text = query.first_text_by_d_aliases(_FLOW_ALIASES["tank"], zone=zone, allow_global=True)
    if building == "D楼":
        primary = query.first_number_by_d_aliases(["一次侧总流量"], zone=zone, allow_global=True)
        tank = query.first_number_by_d_aliases(["蓄冷罐总管道流量"], zone=zone, allow_global=True)
        if primary is not None and tank is not None:
            second_text = _format_calc_value(float(primary) - float(tank))
    return {cells["first"]: first_text, cells["second"]: second_text, cells["tank"]: tank_text}


def _resolve_primary_flow_number(query: CapacitySourceQuery, *, zone: str, context: Dict[str, Any]) -> float | None:
    cells = _WEST_PUBLIC_CELLS if zone == "west" else _EAST_PUBLIC_CELLS
    public_values = _resolve_public_flow_values(query, zone=zone, context=context)
    return _to_float_text(public_values.get(cells["first"]))


def _region_summary_cells(zone: str) -> Dict[str, str]:
    if zone == "west":
        return {"redundancy": "D42", "secondary_value": "D48", "secondary_redundancy": "D49", "tank_level": "AC27"}
    return {"redundancy": "Q42", "secondary_value": "Q48", "secondary_redundancy": "Q49", "tank_level": "AC28"}


def _build_zone_summary_values(query: CapacitySourceQuery, *, zone: str, running_units: Dict[str, List[Dict[str, Any]]]) -> Dict[str, str]:
    cells = _region_summary_cells(zone)
    results: Dict[str, str] = {cells["redundancy"]: _build_redundancy_text(len(running_units.get(zone, [])))}
    secondary_rows = query.rows_by_d_regexes([re.escape(alias) for alias in _SECONDARY_PUMP_ALIASES], zone=zone, allow_global=True)
    running_secondary = [row for row in secondary_rows if getattr(row, "value", None) is not None and float(row.value) > 10]
    if running_secondary:
        results[cells["secondary_value"]] = _text(getattr(running_secondary[0], "e_raw", None))
    results[cells["secondary_redundancy"]] = _build_redundancy_text(len(running_secondary))
    results[cells["tank_level"]] = query.first_text_by_d_aliases(_TANK_LEVEL_ALIASES, zone=zone, allow_global=True)
    return {cell: value for cell, value in results.items() if value != ""}


def _select_active_chiller_unit(
    query: CapacitySourceQuery,
    *,
    zone: str,
    running_units: Dict[str, List[Dict[str, Any]]],
    context: Dict[str, Any],
) -> Dict[str, Any] | None:
    alias_groups = _default_alias_groups(context)
    candidates: List[tuple[bool, float, int, Dict[str, Any]]] = []
    for unit_info in list(running_units.get(zone, [])):
        mode_text = _text(unit_info.get("mode_text"))
        if mode_text not in _ACTIVE_CHILLER_MODE_TEXTS:
            continue
        unit_number = int(unit_info.get("unit", 0) or 0)
        if unit_number <= 0:
            continue
        chilled_out_temp = query.first_number_by_d_aliases(
            alias_groups["chilled_out_temp"],
            zone=zone,
            unit=unit_number,
            allow_global=False,
        )
        sort_temp = float(chilled_out_temp) if chilled_out_temp is not None else float("inf")
        candidates.append(
            (
                chilled_out_temp is None,
                sort_temp,
                unit_number,
                {
                    "unit": unit_number,
                    "mode_text": mode_text,
                    "chilled_out_temp": chilled_out_temp,
                },
            )
        )
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    return dict(candidates[0][3])


def _build_capacity_source_direct_values(query: CapacitySourceQuery, *, context: Dict[str, Any]) -> Dict[str, str]:
    alias_groups = _default_alias_groups(context)
    running_units = context.get("running_units", {}) if isinstance(context.get("running_units", {}), dict) else {}
    values: Dict[str, str] = {}

    storage_total = _first_text_by_d_contains(
        query,
        _CAPACITY_SOURCE_DIRECT_ALIASES["storage_total"],
    )
    if storage_total:
        values["AC29"] = storage_total

    oil_amount = _first_text_by_d_contains(
        query,
        _CAPACITY_SOURCE_DIRECT_ALIASES["oil_amount"],
        excludes=_CAPACITY_SOURCE_DIRECT_EXCLUDES["oil_amount"],
    )
    if oil_amount:
        values["U16"] = oil_amount

    def _zone_capacity(zone: str) -> float | None:
        primary_flow = _resolve_primary_flow_number(query, zone=zone, context=context)
        plate_cooling_in_temp = query.first_number_by_d_aliases(
            alias_groups["plate_cooling_in_temp"],
            zone=zone,
            allow_global=True,
        )
        if primary_flow is None or plate_cooling_in_temp is None:
            return None
        selected_unit = _select_active_chiller_unit(query, zone=zone, running_units=running_units, context=context)
        if selected_unit is not None:
            source_temp = selected_unit.get("chilled_out_temp")
        else:
            source_temp = query.first_number_by_d_aliases(
                alias_groups["plate_chilled_out_temp"],
                zone=zone,
                allow_global=True,
            )
        if source_temp is None:
            return None
        return abs(float(source_temp) - float(plate_cooling_in_temp)) * 4.2 * 1000 / 3600 * float(primary_flow)

    west_capacity = _zone_capacity("west")
    if west_capacity is not None:
        values["D22"] = format_number(west_capacity)

    east_capacity = _zone_capacity("east")
    if east_capacity is not None:
        values["Q22"] = format_number(east_capacity)

    return values


def _block_cell_map(zone: str, position: int) -> Dict[str, str]:
    row_offset = _BLOCK_ROW_OFFSET * int(position)
    col_offset = _BLOCK_COL_OFFSET if zone == "east" else 0
    return {name: _offset_cell(base_cell, row_offset=row_offset, col_offset=col_offset) for name, base_cell in _WEST_BLOCK_BASE_CELLS.items()}


def _fan_values(query: CapacitySourceQuery, *, zone: str, unit: int) -> Dict[int, str]:
    matched_rows = query.rows_by_d_regexes(
        [r"冷却塔\d风机变频反馈", r"冷塔/\d#风机频率反馈", r"冷塔\d#风机频率反馈", r"冷却塔\d号风扇频率反馈"],
        zone=zone,
        unit=unit,
        allow_global=False,
    )
    values: Dict[int, str] = {}
    for row in matched_rows:
        match = re.search(r"(\d)", _text(getattr(row, "d_name", "")))
        if not match:
            continue
        digit = int(match.group(1))
        if digit not in {1, 2} or digit in values:
            continue
        values[digit] = _text(getattr(row, "e_raw", None))
    return values


def _default_alias_groups(context: Dict[str, Any]) -> Dict[str, List[str]]:
    groups = {key: list(value) for key, value in _DEFAULT_REGION_ALIAS_GROUPS.items()}
    groups["primary_pump_freq"] = _building_aliases(context, "primary_pump_aliases")
    groups["chilled_out_temp"].extend(_building_aliases(context, "chilled_out_extra_aliases"))
    groups["cooling_in_temp"].extend(_building_aliases(context, "cooling_in_extra_aliases"))
    return groups


def _build_zone_unit_values(query: CapacitySourceQuery, *, zone: str, running_units: Dict[str, List[Dict[str, Any]]], context: Dict[str, Any]) -> Dict[str, str]:
    alias_groups = _default_alias_groups(context)
    results: Dict[str, str] = {}
    building = _text(context.get("building"))
    active_units = list(running_units.get(zone, []))[:2]
    for position, unit_info in enumerate(active_units):
        unit_number = int(unit_info.get("unit", 0) or 0)
        if unit_number <= 0:
            continue
        block = _block_cell_map(zone, position)
        results[block["title"]] = f"{unit_number}号制冷单元→{_text(unit_info.get('mode_text'))}"
        skip_chiller_values = building == "E楼" and _text(unit_info.get("mode_text")) == "板换"
        for key, aliases in alias_groups.items():
            if key == "primary_pump_freq" and not aliases:
                continue
            if skip_chiller_values and key in _E_BUILDING_CHILLER_SKIP_KEYS:
                continue
            value_text = query.first_text_by_d_aliases(aliases, zone=zone, unit=unit_number, allow_global=False)
            if not value_text:
                continue
            target_cell = block.get(key)
            if target_cell:
                results[target_cell] = value_text
            if key == "plate_cooling_in_pressure":
                results[block["plate_cooling_in_pressure_dup"]] = value_text
        fan_values = _fan_values(query, zone=zone, unit=unit_number)
        if fan_values.get(1):
            results[block["fan_1"]] = fan_values[1]
        if fan_values.get(2):
            results[block["fan_2"]] = fan_values[2]
    return {cell: value for cell, value in results.items() if value != ""}


def _build_tr_values(query: CapacitySourceQuery, snapshot: Dict[str, Any]) -> Dict[str, str]:
    values: Dict[str, str] = {}
    for entry in [item for item in snapshot.get("tr_entries", []) if isinstance(item, dict)]:
        row_idx = int(entry.get("row", 0) or 0)
        identifier = _text(entry.get("identifier"))
        tokens = list(entry.get("search_tokens", []) or [])
        if row_idx > 0 and identifier:
            values[f"B{row_idx}"] = identifier
        matched_row = query.first_row_by_identifier(identifier_tokens=tokens, search_column="c")
        if row_idx > 0 and matched_row is not None:
            values[f"D{row_idx}"] = _text(getattr(matched_row, "e_raw", None))
        replacement_tokens = _tr_replacement_search_tokens(identifier)
        replacement_row = query.first_row_by_identifier(identifier_tokens=replacement_tokens, search_column="b")
        if row_idx > 0 and replacement_row is not None:
            values[f"E{row_idx}"] = _text(getattr(replacement_row, "e_raw", None))
    return {cell: value for cell, value in values.items() if value != ""}


def _build_ups_values(query: CapacitySourceQuery, snapshot: Dict[str, Any]) -> Dict[str, str]:
    values: Dict[str, str] = {}
    alias_map = {"J": ["输出总有功功率"], "M": ["正极电压"], "N": ["输出电压"]}
    for entry in [item for item in snapshot.get("ups_entries", []) if isinstance(item, dict)]:
        row_idx = int(entry.get("row", 0) or 0)
        identifier = _text(entry.get("identifier"))
        tokens = list(entry.get("search_tokens", []) or [])
        if row_idx <= 0:
            continue
        if identifier:
            values[f"I{row_idx}"] = identifier
        for column_letter, aliases in alias_map.items():
            cell_text = query.first_text_by_identifier(identifier_tokens=tokens, search_column="c", d_aliases=aliases)
            if cell_text:
                values[f"{column_letter}{row_idx}"] = cell_text
    return values


def _build_hvdc_values(query: CapacitySourceQuery, snapshot: Dict[str, Any]) -> Dict[str, str]:
    values: Dict[str, str] = {}
    alias_map = {"P": ["电池组电压"], "Q": ["直流电压"], "R": ["直流总功率"]}
    for entry in [item for item in snapshot.get("hvdc_entries", []) if isinstance(item, dict)]:
        row_idx = int(entry.get("row", 0) or 0)
        identifier = _text(entry.get("identifier"))
        tokens = list(entry.get("search_tokens", []) or [])
        if row_idx <= 0:
            continue
        if identifier:
            values[f"O{row_idx}"] = identifier
        for column_letter, aliases in alias_map.items():
            cell_text = query.first_text_by_identifier(identifier_tokens=tokens, search_column="c", d_aliases=aliases)
            if cell_text:
                values[f"{column_letter}{row_idx}"] = cell_text
            elif column_letter == "R":
                values[f"{column_letter}{row_idx}"] = "0"
    return values


def _build_rpp_values(snapshot: Dict[str, Any]) -> Dict[str, str]:
    values: Dict[str, str] = {}
    for entry in [item for item in snapshot.get("rpp_entries", []) if isinstance(item, dict)]:
        row_idx = int(entry.get("row", 0) or 0)
        identifier = _text(entry.get("identifier"))
        if row_idx > 0 and identifier:
            values[f"S{row_idx}"] = identifier
    return values


def _aircon_quadrant(row: RawRow, *, building_code: str) -> tuple[int, str, str] | None:
    _ = building_code
    b_text = _text(getattr(row, "b_text", ""))
    c_text = _text(getattr(row, "c_text", ""))
    combined = f"{b_text} {c_text}"
    if "空调" not in combined:
        return None
    floor_match = re.search(r"([234])层", combined)
    code_match = re.search(r"([234](?:11|12|40|41))", combined)
    area_match = re.search(r"空调区([1-4])", combined)
    floor_token = _text(floor_match.group(1)) if floor_match else ""
    if not floor_token and code_match:
        floor_token = _text(code_match.group(1))[:1]
    if not floor_token:
        return None
    if area_match:
        mapping = _AIRCON_ZONE_DIRECTION_BY_AREA.get(_text(area_match.group(1)))
        if mapping:
            zone, direction = mapping
            return int(floor_token), zone, direction
    if code_match:
        suffix = _text(code_match.group(1))[1:]
        mapping = {
            "12": ("east", "south"),
            "11": ("east", "north"),
            "41": ("west", "south"),
            "40": ("west", "north"),
        }.get(suffix)
        if mapping:
            zone, direction = mapping
            return int(floor_token), zone, direction
    return None


def _build_aircon_matrix_values(query: CapacitySourceQuery, snapshot: Dict[str, Any]) -> Dict[str, str]:
    building_code = _text(snapshot.get("building_code")).upper()
    template_family = _text(snapshot.get("template_family")) or _TEMPLATE_FAMILY_OTHER_BUILDINGS
    target_cells = _AIRCON_TARGET_CELLS_BY_TEMPLATE_FAMILY.get(
        template_family,
        _AIRCON_TARGET_CELLS_BY_TEMPLATE_FAMILY[_TEMPLATE_FAMILY_OTHER_BUILDINGS],
    )
    values: Dict[str, str] = {}
    for row in query.rows:
        key = _aircon_quadrant(row, building_code=building_code)
        if key is None:
            continue
        target_cell = target_cells.get(key)
        value_text = _text(getattr(row, "e_raw", None))
        if not target_cell or not value_text:
            continue
        if target_cell in values and _text(values.get(target_cell)):
            continue
        values[target_cell] = value_text
    return {cell: value for cell, value in values.items() if value != ""}


def build_capacity_cells_with_config(context: Dict[str, Any], config: Dict[str, Any] | None = None) -> Dict[str, str]:
    _ = config
    query = CapacitySourceQuery(context.get("capacity_rows", []) if isinstance(context.get("capacity_rows", []), list) else [])
    running_units = context.get("running_units", {}) if isinstance(context.get("running_units", {}), dict) else {}
    snapshot = context.get("template_snapshot", {}) if isinstance(context.get("template_snapshot", {}), dict) else {}

    values: Dict[str, str] = {}
    values.update(build_common_capacity_cell_values(context))
    values.update(_build_meter_values(query))
    values.update(_resolve_public_flow_values(query, zone="west", context=context))
    values.update(_resolve_public_flow_values(query, zone="east", context=context))
    values.update(_build_capacity_source_direct_values(query, context=context))
    values.update(_build_zone_unit_values(query, zone="west", running_units=running_units, context=context))
    values.update(_build_zone_unit_values(query, zone="east", running_units=running_units, context=context))
    values.update(_build_zone_summary_values(query, zone="west", running_units=running_units))
    values.update(_build_zone_summary_values(query, zone="east", running_units=running_units))
    values.update(_build_tr_values(query, snapshot))
    values.update(_build_ups_values(query, snapshot))
    values.update(_build_hvdc_values(query, snapshot))
    values.update(_build_rpp_values(snapshot))
    values.update(_build_aircon_matrix_values(query, snapshot))
    return {cell: value for cell, value in values.items() if value != ""}
