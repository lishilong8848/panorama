from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from pipeline_utils import get_app_dir
from handover_log_module.core.section_layout import parse_category_sections
from handover_log_module.core.specialty_normalizer import normalize_specialty_text
from handover_log_module.repository.maintenance_management_repository import (
    MaintenanceManagementRepository,
    MaintenanceManagementRow,
    MaintenanceRowsByBuilding,
)
from handover_log_module.repository.shift_roster_repository import ShiftRosterRepository


def _norm_header(value: Any) -> str:
    return str(value or "").replace(" ", "").strip().casefold()


def _looks_like_factory_vendor(item_text: str) -> bool:
    text = str(item_text or "").strip()
    if not text:
        return False
    return ("厂家" in text) or ("厂商" in text)


class MaintenanceManagementPayloadBuilder:
    def __init__(
        self,
        handover_cfg: Dict[str, Any],
        *,
        repository: MaintenanceManagementRepository | None = None,
        shift_roster_repo: ShiftRosterRepository | None = None,
    ) -> None:
        self.handover_cfg = handover_cfg
        self.repo = repository or MaintenanceManagementRepository(handover_cfg)
        self.shift_roster_repo = shift_roster_repo or ShiftRosterRepository(handover_cfg)

    def _resolve_template_path(self, source_path: str) -> Path:
        path = Path(str(source_path or "").strip())
        if path.is_absolute():
            return path
        return get_app_dir() / path

    def _resolve_section_column_mapping(
        self,
        *,
        cfg: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Dict[str, str]]:
        col_cfg = cfg.get("column_mapping", {}) if isinstance(cfg.get("column_mapping", {}), dict) else {}
        fallback_cols = col_cfg.get("fallback_cols", {}) if isinstance(col_cfg.get("fallback_cols", {}), dict) else {}
        header_alias = col_cfg.get("header_alias", {}) if isinstance(col_cfg.get("header_alias", {}), dict) else {}
        resolve_by_header = bool(col_cfg.get("resolve_by_header", True))
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("maintenance_management", "维护管理")).strip() or "维护管理"
        semantic_keys = ["maintenance_item", "maintenance_party", "completion", "executor"]
        defaults = {
            "maintenance_item": "B",
            "maintenance_party": "C",
            "completion": "D",
            "executor": "H",
        }
        output = {
            section_name: {
                key: str(fallback_cols.get(key, "")).strip().upper() or defaults[key]
                for key in semantic_keys
            }
        }
        if not resolve_by_header:
            return output

        template_cfg = self.handover_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return output
        source_path = self._resolve_template_path(str(template_cfg.get("source_path", "")).strip())
        sheet_name = str(template_cfg.get("sheet_name", "交接班日志")).strip() or "交接班日志"
        if not source_path.exists():
            return output

        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return output
            ws = wb[sheet_name]
            sections = parse_category_sections(ws)
            section = next((item for item in sections if item.name == section_name), None)
            if section is None:
                return output
            header_row = section.header_row
            normalized_header_map: Dict[str, str] = {}
            max_col = int(ws.max_column or 0)
            for col_idx in range(1, max_col + 1):
                text = _norm_header(ws.cell(row=header_row, column=col_idx).value)
                if not text:
                    continue
                normalized_header_map[text] = get_column_letter(col_idx)
            for semantic_key in semantic_keys:
                aliases_raw = header_alias.get(semantic_key, [])
                aliases = aliases_raw if isinstance(aliases_raw, list) else []
                found_col = ""
                for alias in aliases:
                    normalized = _norm_header(alias)
                    if normalized and normalized in normalized_header_map:
                        found_col = normalized_header_map[normalized]
                        break
                if found_col:
                    output[section_name][semantic_key] = found_col
            emit_log(f"[交接班][维护管理] 列映射已解析: {output.get(section_name)}")
            return output
        finally:
            wb.close()

    @staticmethod
    def _to_cells(row_payload: Dict[str, str], col_map: Dict[str, str]) -> Dict[str, str]:
        cells: Dict[str, str] = {}
        for semantic_key, text in row_payload.items():
            col = str(col_map.get(semantic_key, "")).strip().upper()
            if not col:
                continue
            cells[col] = str(text or "")
        return cells

    @staticmethod
    def _pick_supervisor(
        engineers: List[Dict[str, str]],
        *,
        building: str,
        specialty_text: str,
    ) -> str:
        specialty = normalize_specialty_text(specialty_text)
        current_building = str(building or "").strip()
        if not specialty or not current_building:
            return ""
        for row in engineers:
            if str(row.get("building", "")).strip() != current_building:
                continue
            if normalize_specialty_text(row.get("specialty", "")) != specialty:
                continue
            supervisor = str(row.get("supervisor", "")).strip()
            if supervisor:
                return supervisor
        return ""

    def _resolve_executor(
        self,
        *,
        row: MaintenanceManagementRow,
        building: str,
        engineers: List[Dict[str, str]],
        emit_log: Callable[[str], None],
    ) -> str:
        supervisor = self._pick_supervisor(
            engineers,
            building=building,
            specialty_text=row.specialty_text,
        )
        if supervisor:
            emit_log(
                "[交接班][维护管理] 执行人匹配: "
                f"record_id={row.record_id}, building={building}, specialty={row.specialty_text}, supervisor={supervisor}"
            )
            return supervisor

        emit_log(
            "[交接班][维护管理] 执行人未匹配: "
            f"record_id={row.record_id}, buildings={'/'.join(row.building_values) or '-'}, specialty={row.specialty_text or '-'}"
        )
        return "/"

    def build(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        preloaded_rows_by_building: MaintenanceRowsByBuilding | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self.repo.get_config()
        if not bool(cfg.get("enabled", True)):
            return {}

        if preloaded_rows_by_building is None:
            try:
                rows, cfg = self.repo.list_current_shift_rows(
                    building=building,
                    duty_date=duty_date,
                    duty_shift=duty_shift,
                    emit_log=emit_log,
                )
            except Exception as exc:  # noqa: BLE001
                emit_log(f"[交接班][维护管理] 读取失败，按空分类继续: {exc}")
                return {}
        else:
            rows = list(preloaded_rows_by_building.get(building, []))

        try:
            engineers = self.shift_roster_repo.list_engineer_directory(emit_log=emit_log)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][维护管理] 工程师目录读取失败，执行人按 / 继续: {exc}")
            engineers = []

        fixed_values = cfg.get("fixed_values", {}) if isinstance(cfg.get("fixed_values", {}), dict) else {}
        vendor_internal = str(fixed_values.get("vendor_internal", "自维")).strip() or "自维"
        vendor_external = str(fixed_values.get("vendor_external", "厂维")).strip() or "厂维"
        completion_text = str(fixed_values.get("completion", "已完成")).strip() or "已完成"
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("maintenance_management", "维护管理")).strip() or "维护管理"
        col_map = self._resolve_section_column_mapping(cfg=cfg, emit_log=emit_log).get(section_name, {})

        payload_rows: List[Dict[str, str]] = []
        matched_supervisor = 0
        unmatched_supervisor = 0

        for row in rows:
            maintenance_item = str(row.item_text or "").strip()
            if not maintenance_item:
                continue
            executor = self._resolve_executor(row=row, building=building, engineers=engineers, emit_log=emit_log)
            if executor == "/":
                unmatched_supervisor += 1
            else:
                matched_supervisor += 1
            maintenance_party = vendor_external if _looks_like_factory_vendor(maintenance_item) else vendor_internal
            payload_rows.append(
                {
                    "maintenance_item": maintenance_item,
                    "maintenance_party": maintenance_party,
                    "completion": completion_text,
                    "executor": executor,
                }
            )

        output_rows = [{"cells": self._to_cells(row_payload, col_map)} for row_payload in payload_rows]
        emit_log(
            f"[交接班][维护管理] 构建完成: building={building}, count={len(output_rows)}, "
            f"matched_supervisor={matched_supervisor}, unmatched_supervisor={unmatched_supervisor}"
        )
        return {section_name: output_rows}
