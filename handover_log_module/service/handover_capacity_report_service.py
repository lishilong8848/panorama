from __future__ import annotations

import base64
import copy
import hashlib
import hmac
import json
import re
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from app.modules.sheet_import.core.field_value_converter import parse_timestamp_ms
from app.shared.utils.atomic_file import atomic_save_workbook
from app.shared.utils.artifact_naming import OUTPUT_TYPE_HANDOVER_CAPACITY, build_output_base_path
from app.shared.utils.file_utils import fallback_missing_windows_drive_path
from handover_log_module.core.shift_window import parse_duty_date
from handover_log_module.core.normalizers import format_number
from handover_log_module.repository.excel_reader import load_rows, load_workbook_quietly
from handover_log_module.repository.review_building_document_store import ReviewBuildingDocumentStore
from handover_log_module.service import capacity_report_a, capacity_report_b, capacity_report_c, capacity_report_d, capacity_report_e
from handover_log_module.service.capacity_report_common import CapacitySourceQuery, build_capacity_template_snapshot
from handover_log_module.service.review_session_service import ReviewSessionService
from pipeline_utils import get_app_dir


_BUILDER_BY_BUILDING = {
    "A楼": capacity_report_a.build_capacity_cells,
    "B楼": capacity_report_b.build_capacity_cells,
    "C楼": capacity_report_c.build_capacity_cells,
    "D楼": capacity_report_d.build_capacity_cells,
    "E楼": capacity_report_e.build_capacity_cells,
}
_RUNNING_MODE_TEXTS = {"制冷", "预冷", "板换"}
_DEFAULT_CHILLER_MODE_VALUE_MAP = {"1": "制冷", "2": "预冷", "3": "板换", "4": "停机"}
_CAPACITY_WATER_SOURCE = {
    "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
    "table_id": "tblz4TkZqrUJB90y",
    "page_size": 500,
    "max_records": 20000,
    "fields": {
        "date": "执行日期",
        "building": "楼栋",
        "water_total": "当日耗水总量（修正）",
    },
}
_CAPACITY_TOTAL_ELECTRICITY_SOURCE = {
    "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
    "table_id": "tblqSskJvBnx9UJj",
    "page_size": 500,
    "max_records": 0,
    "cache_ttl_sec": 30,
    "fields": {
        "category": "汇总分类",
        "building": "楼栋",
        "date": "日期",
        "value": "数值（整数）",
    },
    "category_value": "总用电量",
}
_TOTAL_ELECTRICITY_CACHE_LOCK = threading.RLock()
_TOTAL_ELECTRICITY_CACHE: Dict[tuple[str, str, str], tuple[float, Dict[str, Any]]] = {}
_TOTAL_ELECTRICITY_INFLIGHT: Dict[Any, threading.Event] = {}
_WATER_SUMMARY_CACHE_LOCK = threading.RLock()
_WATER_SUMMARY_CACHE: Dict[Any, tuple[float, Dict[str, Any]]] = {}
_WATER_SUMMARY_INFLIGHT: Dict[Any, threading.Event] = {}
_WEATHER_CACHE_LOCK = threading.RLock()
_WEATHER_CACHE: Dict[Any, tuple[float, Dict[str, Any]]] = {}
_WEATHER_INFLIGHT: Dict[Any, threading.Event] = {}
_CAPACITY_BATCH_CACHE_TTL_SEC = 30
_CAPACITY_TRACKED_CELLS = ("H6", "F8", "B6", "D6", "F6", "D8", "B13", "D13")
_CAPACITY_SYNC_REQUIRED_CELLS = ("H6", "F8", "B6", "D6", "F6", "B13", "D13")
_SUBSTATION_110KV_TARGET_ROWS = {
    "阿开": 57,
    "阿家": 58,
    "1#主变": 60,
    "2#主变": 61,
    "3#主变": 62,
    "4#主变": 63,
}
_SUBSTATION_110KV_TARGET_COLUMNS = {
    "line_voltage": "C",
    "current": "E",
    "power_kw": "G",
    "power_factor": "I",
    "load_rate": "K",
}
_COOLING_PUMP_PRESSURE_TARGETS = {
    ("west", 1): ("I28", "I29"),
    ("west", 2): ("I38", "I39"),
    ("east", 1): ("V28", "V29"),
    ("east", 2): ("V38", "V39"),
}
_WEATHER_PHENOMENON_PRIORITY = (
    "暴雨",
    "大雨",
    "中雨",
    "小雨",
    "雷阵雨",
    "阵雨",
    "雨夹雪",
    "大雪",
    "中雪",
    "小雪",
    "多云",
    "阴",
    "晴",
    "雾",
    "霾",
)
_SENIVERSE_DAILY_WEATHER_ENDPOINT = "https://api.seniverse.com/v3/weather/daily.json"
_SENIVERSE_SIGN_TTL_SEC = 1800
_SENIVERSE_MAX_DAYS = 15
_COOLING_TOWER_LEVEL_ALIASES = ["冷却塔液位", "冷塔液位", "冷却塔水位", "冷塔水位"]
_COOLING_TANK_TEMP_ALIASES = ["蓄冷罐后备温度", "蓄冷罐温度", "蓄冷罐供水温度", "蓄冷罐回水温度"]
_COOLING_TANK_LEVEL_ALIASES = ["蓄冷罐液位", "蓄冷罐水位", "蓄冷罐后备液位"]
_COOLING_SECONDARY_PUMP_ALIASES = ["冷冻水二次泵变频反馈", "二次冷冻泵频率反馈", "二次泵频率反馈", "冷冻水二次泵频率反馈"]
_LEGACY_CAPACITY_TEMPLATE_NAME = "交接班容量报表空模板.xlsx"
_CAPACITY_TEMPLATE_BY_FAMILY = {
    "other_buildings": "其他楼交接班容量报表空模板.xlsx",
    "e_building": "E楼交接班容量报表空模板.xlsx",
}
_CAPACITY_TEMPLATE_FAMILY_BY_FILENAME = {
    "其他楼交接班容量报表空模板.xlsx": "other_buildings",
    "E楼交接班容量报表空模板.xlsx": "e_building",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _with_index(path: Path, idx: int) -> Path:
    if idx <= 1:
        return path
    return path.with_name(f"{path.stem}_{idx}{path.suffix}")


def _field_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in ("text", "name", "value", "label"):
            text = str(value.get(key, "") or "").strip()
            if text:
                return text
        return ""
    if isinstance(value, list):
        parts = [_field_text(item) for item in value]
        return "、".join([item for item in parts if item])
    return str(value).strip()


def _field_text_with_option_map(value: Any, option_map: Dict[str, str]) -> str:
    text = _field_text(value)
    if not text:
        return ""
    return str(option_map.get(text, text)).strip()


def _parse_datetime(value: Any) -> datetime | None:
    timestamp_ms = parse_timestamp_ms(value, tz_offset_hours=8)
    if timestamp_ms is None:
        return None
    return datetime.fromtimestamp(timestamp_ms / 1000)


def _to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if number != number or number in {float("inf"), float("-inf")}:
            return None
        return number
    text = _field_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _extract_building_code(value: Any) -> str:
    match = re.search(r"([A-Za-z])", _text(value))
    return match.group(1).upper() if match else ""


def _normalize_building_text(value: Any) -> str:
    return _text(value).replace(" ", "").casefold()


def _formula_literal(value: Any) -> str:
    if isinstance(value, bool):
        return "TRUE()" if value else "FALSE()"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if isinstance(value, int) else str(value)
    text = _text(value).replace("\\", "\\\\").replace('"', '\\"')
    return f'"{text}"'


def _date_only(value: Any) -> str:
    parsed = _parse_datetime(value)
    if parsed is None:
        text = _text(value)
        return text[:10] if len(text) >= 10 else text
    return parsed.strftime("%Y-%m-%d")


def _read_short_cache(
    cache: Dict[Any, tuple[float, Dict[str, Any]]],
    lock: threading.RLock,
    key: Any,
    *,
    ttl_sec: int,
) -> Dict[str, Any] | None:
    now = time.monotonic()
    with lock:
        cached = cache.get(key)
        if cached and now - cached[0] <= max(1, int(ttl_sec)):
            return copy.deepcopy(cached[1])
    return None


def _store_short_cache(
    cache: Dict[Any, tuple[float, Dict[str, Any]]],
    lock: threading.RLock,
    key: Any,
    payload: Dict[str, Any],
) -> None:
    with lock:
        cache[key] = (time.monotonic(), copy.deepcopy(payload))


def _claim_singleflight(
    inflight: Dict[Any, threading.Event],
    lock: threading.RLock,
    key: Any,
) -> tuple[bool, threading.Event]:
    with lock:
        event = inflight.get(key)
        if event is not None:
            return False, event
        event = threading.Event()
        inflight[key] = event
        return True, event


def _finish_singleflight(
    inflight: Dict[Any, threading.Event],
    lock: threading.RLock,
    key: Any,
    event: threading.Event,
) -> None:
    with lock:
        inflight.pop(key, None)
        event.set()


def _build_fixed_header_cells(building: Any) -> Dict[str, str]:
    building_text = _text(building)
    building_code = _extract_building_code(building_text)
    normalized_code = building_code or building_text.replace("楼", "").replace("栋", "")
    building_floor_text = f"{normalized_code}楼" if normalized_code else building_text
    building_block_text = f"{normalized_code}栋" if normalized_code else building_floor_text.replace("楼", "栋", 1)
    return {
        "A1": f"世纪互联南通数据中心{building_block_text}FM运维交接班重要事项",
        "E5": building_floor_text,
        "G16": building_floor_text,
        "G17": building_floor_text,
        "G18": building_floor_text,
        "S15": building_floor_text,
        "S16": building_floor_text,
        "S17": building_floor_text,
        "S18": building_floor_text,
        "A20": building_floor_text,
        "A65": f"{building_floor_text}容量一览表",
        "O55": f"{building_floor_text}能耗一览",
    }


def _build_merged_anchor_map(sheet: Any) -> Dict[str, str]:
    anchor_map: Dict[str, str] = {}
    for merged_range in getattr(sheet.merged_cells, "ranges", []):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor = f"{get_column_letter(min_col)}{min_row}"
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                anchor_map[coord] = anchor
    return anchor_map


def _write_cells_with_merged_support(sheet: Any, cell_values: Dict[str, Any]) -> None:
    merged_anchor_map = _build_merged_anchor_map(sheet)
    resolved_values: Dict[str, Any] = {}
    source_coords: Dict[str, str] = {}
    for cell_name, value in (cell_values or {}).items():
        normalized_cell = _text(cell_name).upper()
        if not normalized_cell:
            continue
        resolved_cell = merged_anchor_map.get(normalized_cell, normalized_cell)
        if resolved_cell in resolved_values and resolved_values[resolved_cell] != value:
            raise RuntimeError(
                "merged_target_conflict: "
                f"{normalized_cell}->{resolved_cell} conflicts with {source_coords.get(resolved_cell, resolved_cell)}"
            )
        resolved_values[resolved_cell] = value
        source_coords[resolved_cell] = normalized_cell
    for resolved_cell, value in resolved_values.items():
        sheet[resolved_cell] = value


class HandoverCapacityReportService:
    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}
        self._review_session_service = ReviewSessionService(self.config)
        self._weather_payload_cache: Dict[str, Dict[str, str]] = {}
        self._water_summary_cache: Dict[tuple[str, str], Dict[str, str]] = {}

    def _capacity_cfg(self) -> Dict[str, Any]:
        raw = self.config.get("capacity_report", {})
        return raw if isinstance(raw, dict) else {}

    def _template_cfg(self) -> Dict[str, Any]:
        raw = self._capacity_cfg().get("template", {})
        return raw if isinstance(raw, dict) else {}

    def _parsing_cfg(self) -> Dict[str, Any]:
        raw = self._capacity_cfg().get("parsing", {})
        return raw if isinstance(raw, dict) else {}

    def _chiller_mode_cfg(self) -> Dict[str, Any]:
        raw = self.config.get("chiller_mode", {})
        return raw if isinstance(raw, dict) else {}

    def _weather_cfg(self) -> Dict[str, Any]:
        raw = self._capacity_cfg().get("weather", {})
        return raw if isinstance(raw, dict) else {}

    def _today_local_date(self):
        return datetime.now().date()

    @staticmethod
    def _resolve_template_path(raw_path: str) -> Path:
        path = Path(str(raw_path or "").strip())
        if path.is_absolute():
            return path
        return Path(__file__).resolve().parents[2] / path

    @staticmethod
    def _default_template_family_for_building(building: str) -> str:
        return "e_building" if _text(building) == "E楼" else "other_buildings"

    def resolve_template_selection(self, *, building: str) -> Dict[str, Any]:
        source_path = _text(self._template_cfg().get("source_path"))
        default_family = self._default_template_family_for_building(building)
        if source_path and source_path != _LEGACY_CAPACITY_TEMPLATE_NAME:
            resolved_path = self._resolve_template_path(source_path)
            return {
                "path": resolved_path,
                "template_family": _CAPACITY_TEMPLATE_FAMILY_BY_FILENAME.get(resolved_path.name, default_family),
                "source": "config_override",
            }
        default_template_name = _CAPACITY_TEMPLATE_BY_FAMILY[default_family]
        return {
            "path": self._resolve_template_path(default_template_name),
            "template_family": default_family,
            "source": "dual_default",
        }

    def resolve_template_path(self, *, building: str) -> Path:
        return self.resolve_template_selection(building=building)["path"]

    def resolve_output_dir(self) -> Path:
        output_dir = Path(_text(self._template_cfg().get("output_dir")))
        if not str(output_dir):
            raise ValueError("handover_log.capacity_report.template.output_dir 未配置")
        if not output_dir.is_absolute():
            output_dir = get_app_dir() / output_dir
        output_dir = fallback_missing_windows_drive_path(
            output_dir,
            app_dir=get_app_dir(),
            label="交接班容量报表输出目录",
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    def _build_output_path(self, *, building: str, duty_date: str, duty_shift: str) -> Path:
        return build_output_base_path(
            output_root=self.resolve_output_dir(),
            output_type=OUTPUT_TYPE_HANDOVER_CAPACITY,
            building=building,
            suffix=".xlsx",
            duty_date=duty_date,
            duty_shift=duty_shift,
        )

    def _next_available_output_path(self, *, building: str, duty_date: str, duty_shift: str) -> Path:
        base_path = self._build_output_path(
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
        )
        for idx in range(1, 1000):
            candidate = _with_index(base_path, idx)
            if not candidate.exists():
                return candidate
        raise RuntimeError(f"交接班容量报表输出文件序号已用尽: {base_path}")

    @staticmethod
    def _sheet_name(workbook: openpyxl.Workbook, configured_name: str) -> str:
        sheet_name = _text(configured_name)
        if sheet_name and sheet_name in workbook.sheetnames:
            return sheet_name
        return workbook.sheetnames[0]

    @staticmethod
    def _read_handover_cells(output_file: str, cells: List[str], sheet_name: str = "") -> Dict[str, str]:
        workbook = load_workbook_quietly(output_file, data_only=True, read_only=True)
        try:
            ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            values: Dict[str, str] = {}
            for cell_name in cells:
                normalized = _text(cell_name).upper()
                if not normalized:
                    continue
                values[normalized] = _text(ws[normalized].value)
            return values
        finally:
            workbook.close()

    @staticmethod
    def _raw_value_text(value: Any) -> str:
        return _text(value)

    def _load_capacity_rows(self, source_file: str) -> List[Any]:
        return load_rows(
            source_file,
            self._parsing_cfg(),
            self.config.get("normalize", {}) if isinstance(self.config.get("normalize", {}), dict) else {},
        )

    def _extract_oil_values(self, rows: List[Any]) -> Dict[str, str]:
        first_candidates = ["1#油罐容积", "油罐1液位", "1#油罐体积"]
        second_candidates = ["2#油罐容积", "油罐2液位", "2#油罐体积"]

        def _find_value(candidates: List[str]) -> str:
            for candidate in candidates:
                for row in rows:
                    if _text(getattr(row, "c_text", "")) != "燃油自控系统":
                        continue
                    if _text(getattr(row, "d_name", "")) != candidate:
                        continue
                    value_text = self._raw_value_text(getattr(row, "e_raw", None))
                    if value_text:
                        return value_text
            return ""

        return {"first": _find_value(first_candidates), "second": _find_value(second_candidates)}

    @staticmethod
    def _previous_duty_context(*, duty_date: str, duty_shift: str) -> tuple[str, str]:
        duty_day = parse_duty_date(duty_date)
        shift_text = _text(duty_shift).lower()
        if shift_text == "day":
            return (duty_day - timedelta(days=1)).strftime("%Y-%m-%d"), "night"
        return duty_day.strftime("%Y-%m-%d"), "day"

    @staticmethod
    def _scale_ab_oil_value(raw_value: Any) -> str:
        number = _to_float(raw_value)
        if number is None:
            return _text(raw_value)
        return format_number(number * 1000 / 0.84)

    def _extract_current_oil_display_values(
        self,
        *,
        building: str,
        rows: List[Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        building_text = _text(building)
        raw_values = self._extract_oil_values(rows)
        if building_text == "D楼":
            raw_values = {
                "first": self._extract_specific_oil_value(rows, ["1#油罐体积"]),
                "second": self._extract_specific_oil_value(rows, ["2#油罐体积"]),
            }
        display_values = dict(raw_values)
        if building_text in {"A楼", "B楼"}:
            display_values = {
                "first": self._scale_ab_oil_value(raw_values.get("first")),
                "second": self._scale_ab_oil_value(raw_values.get("second")),
            }
        emit_log(
            "[交接班][容量报表][燃油] 当前班次取值 "
            f"building={building_text}, raw_first={_text(raw_values.get('first')) or '-'}, "
            f"raw_second={_text(raw_values.get('second')) or '-'}, "
            f"display_first={_text(display_values.get('first')) or '-'}, "
            f"display_second={_text(display_values.get('second')) or '-'}"
        )
        return {
            "first": _text(display_values.get("first")),
            "second": _text(display_values.get("second")),
        }

    def _extract_specific_oil_value(self, rows: List[Any], aliases: List[str]) -> str:
        for candidate in aliases:
            for row in rows:
                if _text(getattr(row, "c_text", "")) != "燃油自控系统":
                    continue
                if _text(getattr(row, "d_name", "")) != candidate:
                    continue
                value_text = self._raw_value_text(getattr(row, "e_raw", None))
                if value_text:
                    return value_text
        return ""

    def _load_previous_capacity_display_oil_values(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        current_display_values: Dict[str, str],
        emit_log: Callable[[str], None],
    ) -> tuple[Dict[str, str], str]:
        previous_date, previous_shift = self._previous_duty_context(
            duty_date=duty_date,
            duty_shift=duty_shift,
        )
        try:
            previous_session = self._review_session_service.get_latest_session_for_context(
                building=building,
                duty_date=previous_date,
                duty_shift=previous_shift,
            )
        except Exception as exc:  # noqa: BLE001
            warning = f"上一班容量文件查询失败，已回退当前班次值: {exc}"
            emit_log(
                "[交接班][容量报表][燃油] 上一班容量文件查询失败 "
                f"building={building}, duty={previous_date}/{previous_shift}, error={exc}"
            )
            return dict(current_display_values), warning

        previous_capacity_output = ""
        if isinstance(previous_session, dict):
            previous_capacity_output = _text(previous_session.get("capacity_output_file"))
        if not previous_capacity_output:
            warning = "上一班容量文件未命中，已回退当前班次值"
            emit_log(
                "[交接班][容量报表][燃油] 上一班容量文件未命中 "
                f"building={building}, duty={previous_date}/{previous_shift}"
            )
            return dict(current_display_values), warning

        emit_log(
            "[交接班][容量报表][燃油] 上一班容量文件命中 "
            f"building={building}, duty={previous_date}/{previous_shift}, output={previous_capacity_output}"
        )
        previous_capacity_path = Path(previous_capacity_output)
        if not previous_capacity_path.exists() or not previous_capacity_path.is_file():
            warning = "上一班容量文件不存在，已回退当前班次值"
            emit_log(
                "[交接班][容量报表][燃油] 上一班容量文件不存在 "
                f"building={building}, duty={previous_date}/{previous_shift}, output={previous_capacity_output}"
            )
            return dict(current_display_values), warning
        try:
            previous_values = self._read_handover_cells(
                previous_capacity_output,
                ["U13", "X13"],
                sheet_name=_text(self._template_cfg().get("sheet_name")),
            )
        except Exception as exc:  # noqa: BLE001
            warning = f"上一班容量文件读取失败，已回退当前班次值: {exc}"
            emit_log(
                "[交接班][容量报表][燃油] 上一班容量文件读取失败 "
                f"building={building}, duty={previous_date}/{previous_shift}, output={previous_capacity_output}, error={exc}"
            )
            return dict(current_display_values), warning
        previous_first = _text(previous_values.get("U13"))
        previous_second = _text(previous_values.get("X13"))
        if not previous_first and not previous_second:
            warning = "上一班容量文件U13/X13为空，已回退当前班次值"
            emit_log(
                "[交接班][容量报表][燃油] 上一班容量文件U13/X13为空 "
                f"building={building}, duty={previous_date}/{previous_shift}, output={previous_capacity_output}"
            )
            return dict(current_display_values), warning
        return {
            "first": previous_first or _text(current_display_values.get("first")),
            "second": previous_second or _text(current_display_values.get("second")),
        }, ""

    def _normalize_alarm_summary(self, payload: Dict[str, Any] | None) -> Dict[str, Any]:
        data = payload if isinstance(payload, dict) else {}
        return {
            "total_count": int(data.get("total_count", 0) or 0),
            "unrecovered_count": int(data.get("unrecovered_count", 0) or 0),
            "accept_description": _text(data.get("accept_description")) or "/",
        }

    @staticmethod
    def _extract_hvdc_position_code_from_text(value: Any) -> str:
        match = re.search(r"([A-E]-\d{3}-HVDC-\d{3})", _text(value), flags=re.IGNORECASE)
        return _text(match.group(1)).upper() if match else ""

    @classmethod
    def _extract_hvdc_position_code(cls, origin_payload: Dict[str, Any] | None) -> str:
        payload = origin_payload if isinstance(origin_payload, dict) else {}
        for field_name in ("d_name", "b_text", "c_text", "b_norm", "c_norm"):
            extracted = cls._extract_hvdc_position_code_from_text(payload.get(field_name))
            if extracted:
                return extracted
        return ""

    def _resolve_hvdc_text(
        self,
        *,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None,
        hvdc_source_d_name: Any = None,
    ) -> Dict[str, str]:
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        raw_value = _text(resolved.get("hvdc_load_max"))
        if raw_value and "%" not in raw_value:
            raw_value = f"{raw_value}%"
        source_d_name_text = _text(hvdc_source_d_name)
        position_code = self._extract_hvdc_position_code_from_text(hvdc_source_d_name)
        origin_context = metric_origin_context if isinstance(metric_origin_context, dict) else {"by_metric_id": {}}
        if not position_code:
            origin_payload = origin_context.get("by_metric_id", {}).get("hvdc_load_max", {})
            position_code = self._extract_hvdc_position_code(origin_payload)
            if not source_d_name_text:
                source_d_name_text = _text(origin_payload.get("d_name"))
        if raw_value and position_code:
            formatted = f"{raw_value}/{position_code}"
        elif raw_value:
            formatted = f"{raw_value}/"
        else:
            formatted = position_code
        return {
            "raw_value": raw_value,
            "source_d_name": source_d_name_text,
            "position_code": position_code,
            "formatted": formatted,
        }

    @staticmethod
    def _normalize_mode_code(value: Any, value_map: Dict[str, str]) -> str:
        raw = _text(value)
        if not raw:
            return ""
        if raw in value_map:
            return raw
        try:
            number = float(raw)
        except ValueError:
            number = None
        if number is not None and int(number) == number:
            key = str(int(number))
            if key in value_map:
                return key
        lowered = raw.casefold()
        for key, text in value_map.items():
            if _text(text).casefold() == lowered:
                return _text(key)
        return ""

    def _resolve_running_units(self, resolved_values_by_id: Dict[str, Any] | None) -> Dict[str, List[Dict[str, Any]]]:
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        chiller_cfg = self._chiller_mode_cfg()
        raw_value_map = chiller_cfg.get("value_map", _DEFAULT_CHILLER_MODE_VALUE_MAP)
        value_map = (
            {str(key).strip(): str(value).strip() for key, value in raw_value_map.items() if _text(key)}
            if isinstance(raw_value_map, dict)
            else dict(_DEFAULT_CHILLER_MODE_VALUE_MAP)
        )
        running: Dict[str, List[Dict[str, Any]]] = {"west": [], "east": []}
        for unit_number in range(1, 7):
            metric_key = f"chiller_mode_{unit_number}"
            mode_code = self._normalize_mode_code(resolved.get(metric_key), value_map)
            mode_text = _text(value_map.get(mode_code))
            if mode_text not in _RUNNING_MODE_TEXTS:
                continue
            zone = "west" if unit_number <= 3 else "east"
            running[zone].append(
                {
                    "unit": unit_number,
                    "metric_key": metric_key,
                    "mode_code": mode_code,
                    "mode_text": mode_text,
                }
            )
        running["west"].sort(key=lambda item: int(item.get("unit", 0) or 0))
        running["east"].sort(key=lambda item: int(item.get("unit", 0) or 0))
        return running

    @staticmethod
    def _extract_option_map_from_field(field_def: Dict[str, Any]) -> Dict[str, str]:
        if not isinstance(field_def, dict):
            return {}
        property_cfg = field_def.get("property", {})
        option_containers: List[Any] = []
        if isinstance(property_cfg, dict):
            option_containers.append(property_cfg.get("options"))
            type_cfg = property_cfg.get("type")
            if isinstance(type_cfg, dict):
                ui_property = type_cfg.get("ui_property")
                if isinstance(ui_property, dict):
                    option_containers.append(ui_property.get("options"))
        option_map: Dict[str, str] = {}
        for options in option_containers:
            if not isinstance(options, list):
                continue
            for option in options:
                if not isinstance(option, dict):
                    continue
                option_id = str(option.get("id", "") or "").strip()
                option_name = str(option.get("name", "") or "").strip()
                if option_id:
                    option_map[option_id] = option_name
        return option_map

    def _load_field_option_maps(
        self,
        *,
        client: FeishuBitableClient,
        table_id: str,
        target_fields: List[str],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Dict[str, str]]:
        field_names = [str(name or "").strip() for name in (target_fields or []) if str(name or "").strip()]
        if not field_names or not str(table_id or "").strip():
            return {}
        try:
            field_defs = client.list_fields(table_id=table_id, page_size=200)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][容量报表][耗水摘要] 字段定义读取失败 building=全局, error={exc}")
            return {}
        output: Dict[str, Dict[str, str]] = {}
        for field_def in field_defs:
            if not isinstance(field_def, dict):
                continue
            field_name = str(
                field_def.get("field_name")
                or field_def.get("name")
                or field_def.get("title")
                or ""
            ).strip()
            if not field_name or field_name not in field_names:
                continue
            output[field_name] = self._extract_option_map_from_field(field_def)
        return output

    @staticmethod
    def tracked_cells() -> List[str]:
        return list(_CAPACITY_TRACKED_CELLS)

    @staticmethod
    def _extract_fixed_cells_from_document(document: Dict[str, Any]) -> Dict[str, str]:
        output: Dict[str, str] = {}
        fixed_blocks = document.get("fixed_blocks", []) if isinstance(document, dict) else []
        if not isinstance(fixed_blocks, list):
            return output
        for block in fixed_blocks:
            if not isinstance(block, dict):
                continue
            for field in block.get("fields", []):
                if not isinstance(field, dict):
                    continue
                cell_name = _text(field.get("cell")).upper()
                if not cell_name:
                    continue
                output[cell_name] = _text(field.get("value"))
        return output

    @classmethod
    def extract_tracked_cells_from_review_document(cls, document: Dict[str, Any]) -> Dict[str, str]:
        fixed_cells = cls._extract_fixed_cells_from_document(document if isinstance(document, dict) else {})
        return {cell: _text(fixed_cells.get(cell)) for cell in _CAPACITY_TRACKED_CELLS}

    @staticmethod
    def capacity_input_signature(cells: Dict[str, Any] | None) -> str:
        payload = cells if isinstance(cells, dict) else {}
        return "|".join(f"{cell}={_text(payload.get(cell))}" for cell in _CAPACITY_TRACKED_CELLS)

    @staticmethod
    def _normalize_capacity_sync_payload(raw: Dict[str, Any] | None) -> Dict[str, Any]:
        payload = raw if isinstance(raw, dict) else {}
        status = _text(payload.get("status")).lower()
        if status not in {"ready", "pending", "pending_input", "missing_file", "failed"}:
            status = "failed"
        return {
            "status": status,
            "updated_at": _text(payload.get("updated_at")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": _text(payload.get("error")),
            "tracked_cells": list(_CAPACITY_TRACKED_CELLS),
            "input_signature": _text(payload.get("input_signature")),
        }

    @staticmethod
    def _build_capacity_sync_payload(
        *,
        status: str,
        error: str = "",
        input_signature: str = "",
    ) -> Dict[str, Any]:
        return HandoverCapacityReportService._normalize_capacity_sync_payload(
            {
                "status": status,
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "error": error,
                "input_signature": input_signature,
            }
        )

    @staticmethod
    def _month_window_for_duty_date(duty_date: str) -> tuple[datetime, datetime]:
        duty_day = parse_duty_date(duty_date)
        month_start_dt = datetime(duty_day.year, duty_day.month, 1, 0, 0, 0)
        if duty_day.month == 12:
            month_end_dt = datetime(duty_day.year + 1, 1, 1, 0, 0, 0)
        else:
            month_end_dt = datetime(duty_day.year, duty_day.month + 1, 1, 0, 0, 0)
        return month_start_dt, month_end_dt

    @staticmethod
    def _derive_tank_pair_from_f8(value: Any) -> tuple[str, str]:
        text = _text(value)
        if not text:
            return "", ""
        west_match = re.search(r"西区\s*([+-]?\d+(?:\.\d+)?)", text)
        east_match = re.search(r"东区\s*([+-]?\d+(?:\.\d+)?)", text)
        if west_match and east_match:
            return _text(west_match.group(1)), _text(east_match.group(1))
        numbers = re.findall(r"[+-]?\d+(?:\.\d+)?", text)
        if len(numbers) >= 2:
            return _text(numbers[0]), _text(numbers[1])
        return "", ""

    @staticmethod
    def _weather_keyword_from_html(html: str) -> str:
        text = str(html or "")
        match = re.search(r"天气的关键词是[“\"]([^”\"<]{1,20})", text)
        if match:
            source = _text(match.group(1))
        else:
            source = ""
        if not source:
            source = _text(text)
        for keyword in _WEATHER_PHENOMENON_PRIORITY:
            if keyword in source:
                return keyword
            if keyword in text:
                return keyword
        return ""

    @staticmethod
    def _weather_humidity_from_html(html: str) -> str:
        text = str(html or "")
        match = re.search(r"湿度\s*([0-9]{1,3}\s*%)", text)
        if match:
            return _text(match.group(1)).replace(" ", "")
        match = re.search(r"相对.*?湿度\s*([0-9]{1,3}\s*%)", text, flags=re.DOTALL)
        if match:
            return _text(match.group(1)).replace(" ", "")
        return ""

    def _legacy_fetch_weather_payload_for_duty_date(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        duty_date_text = _text(duty_date)
        if not duty_date_text:
            return {"text": "", "humidity": ""}
        cache_key = f"legacy_html:{duty_date_text}"
        global_key = ("legacy", cache_key)
        while True:
            cached = _read_short_cache(
                _WEATHER_CACHE,
                _WEATHER_CACHE_LOCK,
                global_key,
                ttl_sec=_CAPACITY_BATCH_CACHE_TTL_SEC,
            )
            if cached is not None:
                return {"text": _text(cached.get("text")), "humidity": _text(cached.get("humidity"))}
            leader, event = _claim_singleflight(_WEATHER_INFLIGHT, _WEATHER_CACHE_LOCK, global_key)
            if leader:
                break
            event.wait()
        try:
            duty_day = parse_duty_date(duty_date_text)
            date_token = duty_day.strftime("%Y%m%d")
            url = f"https://www.tianqi.com/tianqi/chongchuanqu/{date_token}.html"
            req = Request(url=url, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(req, timeout=8) as resp:  # noqa: S310
                content = resp.read().decode("utf-8", errors="ignore")
            payload = {
                "text": self._weather_keyword_from_html(content),
                "humidity": self._weather_humidity_from_html(content),
            }
            if payload["text"] or payload["humidity"]:
                emit_log(
                    "[交接班][容量报表][天气] 查询完成 "
                    f"duty_date={duty_date_text}, provider=legacy_html, weather={payload.get('text') or '-'}, humidity={payload.get('humidity') or '-'}"
                )
                _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, payload)
                return payload
            emit_log(
                "[交接班][容量报表][天气] 查询失败 "
                f"duty_date={duty_date_text}, provider=legacy_html, reason=页面未解析到天气现象/湿度"
            )
            _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, {"text": "", "humidity": ""})
            return {"text": "", "humidity": ""}
        except (ValueError, OSError, TimeoutError, URLError) as exc:
            emit_log(
                "[交接班][容量报表][天气] 查询失败 "
                f"duty_date={duty_date_text}, provider=legacy_html, error={exc}"
            )
            _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, {"text": "", "humidity": ""})
            return {"text": "", "humidity": ""}
        finally:
            _finish_singleflight(_WEATHER_INFLIGHT, _WEATHER_CACHE_LOCK, global_key, event)

    @staticmethod
    def _format_seniverse_humidity(value: Any) -> str:
        text = _text(value)
        if not text:
            return ""
        return text if text.endswith("%") else f"{text}%"

    def _build_seniverse_signed_query_params(self) -> Dict[str, str]:
        weather_cfg = self._weather_cfg()
        uid = _text(weather_cfg.get("seniverse_public_key"))
        key = _text(weather_cfg.get("seniverse_private_key"))
        if not uid or not key:
            raise ValueError("心知天气鉴权配置缺失: seniverse_public_key/seniverse_private_key")
        ts = str(int(datetime.now().timestamp()))
        ttl = str(_SENIVERSE_SIGN_TTL_SEC)
        signing_text = f"ts={ts}&ttl={ttl}&uid={uid}"
        digest = hmac.new(
            key.encode("utf-8"),
            signing_text.encode("utf-8"),
            hashlib.sha1,
        ).digest()
        signature = base64.b64encode(digest).decode("utf-8")
        return {
            "ts": ts,
            "ttl": ttl,
            "uid": uid,
            "sig": signature,
        }

    def _seniverse_candidate_locations(self) -> List[str]:
        weather_cfg = self._weather_cfg()
        candidates: List[str] = []
        raw_fallback_locations = weather_cfg.get("fallback_locations")
        for raw_value in [
            weather_cfg.get("location"),
            *(raw_fallback_locations if isinstance(raw_fallback_locations, list) else []),
        ]:
            text = self._normalize_seniverse_location(raw_value)
            if text and text not in candidates:
                candidates.append(text)
        if not candidates:
            candidates = ["31.98:120.89", "南通"]
        elif "南通" not in candidates:
            candidates.append("南通")
        return candidates

    @staticmethod
    def _normalize_seniverse_location(value: Any) -> str:
        text = _text(value)
        if not text:
            return ""
        normalized = text.replace(" ", "").casefold()
        if normalized in {"崇川区", "崇川"}:
            return "31.98:120.89"
        return text

    @staticmethod
    def _parse_seniverse_error_body(body_text: str) -> str:
        text = _text(body_text)
        if not text:
            return ""
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            return text
        if not isinstance(payload, dict):
            return text
        status = _text(payload.get("status"))
        status_code = _text(payload.get("status_code"))
        if status_code and status:
            return f"{status_code}: {status}"
        return status_code or status or text

    def _fetch_seniverse_weather_payload_for_duty_date(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        duty_date_text = _text(duty_date)
        if not duty_date_text:
            return {"text": "", "humidity": ""}
        candidate_locations = self._seniverse_candidate_locations()
        cache_key = f"seniverse:{duty_date_text}:{'|'.join(candidate_locations)}"
        weather_cfg = self._weather_cfg()
        global_key = (
            "seniverse",
            cache_key,
            _text(weather_cfg.get("seniverse_public_key")),
            _text(weather_cfg.get("language")) or "zh-Hans",
            _text(weather_cfg.get("unit")) or "c",
            str(id(urlopen)),
        )
        while True:
            cached = _read_short_cache(
                _WEATHER_CACHE,
                _WEATHER_CACHE_LOCK,
                global_key,
                ttl_sec=_CAPACITY_BATCH_CACHE_TTL_SEC,
            )
            if cached is not None:
                return {"text": _text(cached.get("text")), "humidity": _text(cached.get("humidity"))}
            leader, event = _claim_singleflight(_WEATHER_INFLIGHT, _WEATHER_CACHE_LOCK, global_key)
            if leader:
                break
            event.wait()
        try:
            duty_day = parse_duty_date(duty_date_text)
            today = self._today_local_date()
            day_offset = (duty_day - today).days
            if day_offset < 0:
                emit_log(
                    "[交接班][容量报表][天气] 查询失败 "
                    f"duty_date={duty_date_text}, provider=seniverse, reason=历史日期无可用天气数据"
                )
                _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, {"text": "", "humidity": ""})
                return {"text": "", "humidity": ""}
            timeout_sec = int(weather_cfg.get("timeout_sec") or 8)
            primary_location = _text(weather_cfg.get("location")) or "崇川区"
            attempted_errors: List[str] = []
            for request_location in candidate_locations:
                try:
                    query_params = {
                        **self._build_seniverse_signed_query_params(),
                        "location": request_location,
                        "language": _text(weather_cfg.get("language")) or "zh-Hans",
                        "unit": _text(weather_cfg.get("unit")) or "c",
                        "start": "0",
                        "days": str(max(1, min(day_offset + 1, _SENIVERSE_MAX_DAYS))),
                    }
                    url = f"{_SENIVERSE_DAILY_WEATHER_ENDPOINT}?{urlencode(query_params)}"
                    request = Request(url=url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
                    with urlopen(request, timeout=timeout_sec) as resp:  # noqa: S310
                        status_code = getattr(resp, "status", None) or getattr(resp, "code", None) or 200
                        if int(status_code) != 200:
                            raise OSError(f"HTTP {status_code}")
                        payload_json = json.loads(resp.read().decode("utf-8", errors="ignore"))
                    results = payload_json.get("results") if isinstance(payload_json, dict) else None
                    result = results[0] if isinstance(results, list) and results else {}
                    daily_rows = result.get("daily") if isinstance(result, dict) else None
                    target_row = next(
                        (
                            item for item in (daily_rows if isinstance(daily_rows, list) else [])
                            if isinstance(item, dict) and _text(item.get("date")) == duty_date_text
                        ),
                        None,
                    )
                    if not isinstance(target_row, dict):
                        attempted_errors.append(f"{request_location}: 未匹配到目标日期天气")
                        continue
                    payload = {
                        "text": _text(target_row.get("text_day")) or _text(target_row.get("text_night")),
                        "humidity": self._format_seniverse_humidity(target_row.get("humidity")),
                    }
                    if payload["text"] or payload["humidity"]:
                        emit_log(
                            "[交接班][容量报表][天气] 查询完成 "
                            f"duty_date={duty_date_text}, provider=seniverse, location={primary_location}, "
                            f"request_location={request_location}, weather={payload.get('text') or '-'}, humidity={payload.get('humidity') or '-'}"
                        )
                        _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, payload)
                        return payload
                    attempted_errors.append(f"{request_location}: 未解析到天气现象/湿度")
                except HTTPError as exc:
                    error_body = exc.read().decode("utf-8", errors="ignore")
                    attempted_errors.append(
                        f"{request_location}: HTTP {exc.code} {self._parse_seniverse_error_body(error_body) or exc.reason or ''}".strip()
                    )
                except (ValueError, OSError, TimeoutError, URLError, json.JSONDecodeError) as exc:
                    attempted_errors.append(f"{request_location}: {exc}")
            emit_log(
                "[交接班][容量报表][天气] 查询失败 "
                f"duty_date={duty_date_text}, provider=seniverse, attempted_locations={candidate_locations}, "
                f"error={' | '.join(attempted_errors) if attempted_errors else '未知错误'}"
            )
            _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, {"text": "", "humidity": ""})
            return {"text": "", "humidity": ""}
        except (ValueError, OSError, TimeoutError, URLError, json.JSONDecodeError) as exc:
            emit_log(
                "[交接班][容量报表][天气] 查询失败 "
                f"duty_date={duty_date_text}, provider=seniverse, error={exc}"
            )
            _store_short_cache(_WEATHER_CACHE, _WEATHER_CACHE_LOCK, global_key, {"text": "", "humidity": ""})
            return {"text": "", "humidity": ""}
        finally:
            _finish_singleflight(_WEATHER_INFLIGHT, _WEATHER_CACHE_LOCK, global_key, event)

    def _fetch_weather_payload_for_duty_date(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        duty_date_text = _text(duty_date)
        if not duty_date_text:
            return {"text": "", "humidity": ""}
        try:
            duty_day = parse_duty_date(duty_date_text)
        except ValueError as exc:
            emit_log(
                "[交接班][容量报表][天气] 查询失败 "
                f"duty_date={duty_date_text}, error={exc}"
            )
            return {"text": "", "humidity": ""}
        if duty_day < self._today_local_date():
            return self._legacy_fetch_weather_payload_for_duty_date(
                duty_date=duty_date_text,
                emit_log=emit_log,
            )
        return self._fetch_seniverse_weather_payload_for_duty_date(
            duty_date=duty_date_text,
            emit_log=emit_log,
        )

    def _fetch_weather_text_for_duty_date(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> str:
        return _text(self._fetch_weather_payload_for_duty_date(duty_date=duty_date, emit_log=emit_log).get("text"))

    def _new_capacity_water_client(self) -> FeishuBitableClient:
        global_feishu = require_feishu_auth_settings(self.config)
        app_token = str(_CAPACITY_WATER_SOURCE.get("app_token", "") or "").strip()
        table_id = str(_CAPACITY_WATER_SOURCE.get("table_id", "") or "").strip()
        if not app_token or not table_id:
            raise ValueError("容量报表耗水多维配置缺失: app_token/table_id")
        return FeishuBitableClient(
            app_id=str(global_feishu.get("app_id", "") or "").strip(),
            app_secret=str(global_feishu.get("app_secret", "") or "").strip(),
            app_token=app_token,
            calc_table_id=table_id,
            attachment_table_id=table_id,
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
        )

    @staticmethod
    def _build_capacity_water_filter_formula(*, start_date: str, end_date: str) -> str:
        fields_cfg = _CAPACITY_WATER_SOURCE.get("fields", {})
        if not isinstance(fields_cfg, dict):
            fields_cfg = {}
        date_field = str(fields_cfg.get("date", "执行日期") or "执行日期").strip()
        return (
            f'AND(CurrentValue.[{date_field}]>=TODATE({_formula_literal(start_date)}), '
            f'CurrentValue.[{date_field}]<TODATE({_formula_literal(end_date)}))'
        )

    @staticmethod
    def _matches_building(record_building: Any, target_building: str) -> bool:
        record_text = _field_text(record_building)
        target_text = _text(target_building)
        if not record_text or not target_text:
            return False
        if _normalize_building_text(record_text) == _normalize_building_text(target_text):
            return True
        return bool(
            _extract_building_code(record_text)
            and _extract_building_code(record_text) == _extract_building_code(target_text)
        )

    def _query_capacity_water_batch(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        duty_date_text = _text(duty_date)
        month_start_dt, month_end_dt = self._month_window_for_duty_date(duty_date_text)
        start_text = month_start_dt.strftime("%Y-%m-%d")
        end_text = month_end_dt.strftime("%Y-%m-%d")

        fields_cfg = _CAPACITY_WATER_SOURCE.get("fields", {})
        if not isinstance(fields_cfg, dict):
            fields_cfg = {}
        date_field = str(fields_cfg.get("date", "") or "").strip()
        building_field = str(fields_cfg.get("building", "") or "").strip()
        water_total_field = str(fields_cfg.get("water_total", "") or "").strip()
        table_id = str(_CAPACITY_WATER_SOURCE.get("table_id", "") or "").strip()
        app_token = str(_CAPACITY_WATER_SOURCE.get("app_token", "") or "").strip()
        cache_key = (app_token, table_id, date_field, building_field, water_total_field, start_text, end_text)
        ttl_sec = max(1, int(_CAPACITY_WATER_SOURCE.get("cache_ttl_sec", _CAPACITY_BATCH_CACHE_TTL_SEC) or _CAPACITY_BATCH_CACHE_TTL_SEC))
        while True:
            cached = _read_short_cache(
                _WATER_SUMMARY_CACHE,
                _WATER_SUMMARY_CACHE_LOCK,
                cache_key,
                ttl_sec=ttl_sec,
            )
            if cached is not None:
                emit_log(
                    "[交接班][容量报表][耗水摘要] 命中短缓存 "
                    f"window={start_text}~{end_text}, raw={cached.get('raw_count', 0)}, matched={cached.get('matched_count', 0)}"
                )
                return cached
            leader, event = _claim_singleflight(_WATER_SUMMARY_INFLIGHT, _WATER_SUMMARY_CACHE_LOCK, cache_key)
            if leader:
                break
            event.wait()

        result: Dict[str, Any] = {
            "start_date": start_text,
            "end_date": end_text,
            "by_building": {},
            "raw_count": 0,
            "matched_count": 0,
        }
        for building_name in _BUILDER_BY_BUILDING:
            result["by_building"][building_name] = {
                "month_total": "",
                "latest_daily_total": "",
                "matched_records": 0,
            }

        try:
            client = self._new_capacity_water_client()
            option_maps = self._load_field_option_maps(
                client=client,
                table_id=table_id,
                target_fields=[building_field],
                emit_log=emit_log,
            )
            filter_formula = self._build_capacity_water_filter_formula(start_date=start_text, end_date=end_text)
            records = client.list_records(
                table_id=table_id,
                page_size=max(1, int(_CAPACITY_WATER_SOURCE.get("page_size", 500) or 500)),
                max_records=max(1, int(_CAPACITY_WATER_SOURCE.get("max_records", 20000) or 20000)),
                filter_formula=filter_formula,
            )
        except Exception as exc:  # noqa: BLE001
            emit_log(
                "[交接班][容量报表][耗水摘要] 查询失败 "
                f"window={start_text}~{end_text}, duty_date={duty_date_text}, error={exc}"
            )
            _store_short_cache(_WATER_SUMMARY_CACHE, _WATER_SUMMARY_CACHE_LOCK, cache_key, result)
            _finish_singleflight(_WATER_SUMMARY_INFLIGHT, _WATER_SUMMARY_CACHE_LOCK, cache_key, event)
            return result

        result["raw_count"] = len(records)
        building_option_map = option_maps.get(building_field, {})
        month_totals: Dict[str, float] = {building_name: 0.0 for building_name in _BUILDER_BY_BUILDING}
        latest_values: Dict[str, float | None] = {building_name: None for building_name in _BUILDER_BY_BUILDING}
        latest_record_dates: Dict[str, datetime | None] = {building_name: None for building_name in _BUILDER_BY_BUILDING}
        for item in records:
            if not isinstance(item, dict):
                continue
            fields = item.get("fields", {})
            if not isinstance(fields, dict):
                continue
            record_dt = _parse_datetime(fields.get(date_field))
            if record_dt is None or record_dt < month_start_dt or record_dt >= month_end_dt:
                continue
            record_building = _field_text_with_option_map(fields.get(building_field), building_option_map)
            matched_building = ""
            for building_name in _BUILDER_BY_BUILDING:
                if self._matches_building(record_building, building_name):
                    matched_building = building_name
                    break
            if not matched_building:
                continue
            water_total = _to_float(fields.get(water_total_field))
            if water_total is None:
                continue
            bucket = result["by_building"][matched_building]
            bucket["matched_records"] = int(bucket.get("matched_records", 0) or 0) + 1
            result["matched_count"] = int(result.get("matched_count", 0) or 0) + 1
            month_totals[matched_building] += water_total
            latest_record_dt = latest_record_dates.get(matched_building)
            if latest_record_dt is None or record_dt >= latest_record_dt:
                latest_record_dates[matched_building] = record_dt
                latest_values[matched_building] = water_total

        for building_name, bucket in result["by_building"].items():
            bucket["month_total"] = format_number(month_totals.get(building_name, 0.0))
            bucket["latest_daily_total"] = format_number(latest_values.get(building_name))
            emit_log(
                "[交接班][容量报表][耗水摘要] 楼栋结果 "
                f"building={building_name}, O57={bucket.get('latest_daily_total', '') or '-'}, "
                f"AC25={bucket.get('month_total', '') or '-'}, matched={bucket.get('matched_records', 0)}"
            )
        emit_log(
            "[交接班][容量报表][耗水摘要] 查询完成 "
            f"window={start_text}~{end_text}, raw={result.get('raw_count', 0)}, matched={result.get('matched_count', 0)}"
        )
        _store_short_cache(_WATER_SUMMARY_CACHE, _WATER_SUMMARY_CACHE_LOCK, cache_key, result)
        _finish_singleflight(_WATER_SUMMARY_INFLIGHT, _WATER_SUMMARY_CACHE_LOCK, cache_key, event)
        return result

    def _query_capacity_water_summary(
        self,
        *,
        building: str,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        building_text = _text(building)
        batch = self._query_capacity_water_batch(duty_date=duty_date, emit_log=emit_log)
        by_building = batch.get("by_building", {}) if isinstance(batch.get("by_building", {}), dict) else {}
        summary = by_building.get(building_text, {}) if isinstance(by_building.get(building_text, {}), dict) else {}
        return {
            "month_total": _text(summary.get("month_total")),
            "latest_daily_total": _text(summary.get("latest_daily_total")),
        }

    def _new_total_electricity_client(self) -> FeishuBitableClient:
        global_feishu = require_feishu_auth_settings(self.config)
        app_token = str(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("app_token", "") or "").strip()
        table_id = str(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("table_id", "") or "").strip()
        if not app_token or not table_id:
            raise ValueError("容量报表总用电量多维配置缺失: app_token/table_id")
        return FeishuBitableClient(
            app_id=str(global_feishu.get("app_id", "") or "").strip(),
            app_secret=str(global_feishu.get("app_secret", "") or "").strip(),
            app_token=app_token,
            calc_table_id=table_id,
            attachment_table_id=table_id,
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
        )

    @staticmethod
    def _total_electricity_window(duty_date: str) -> tuple[datetime, datetime, datetime]:
        duty_day = parse_duty_date(duty_date)
        if duty_day.day == 1:
            month_end_dt = datetime(duty_day.year, duty_day.month, 1, 0, 0, 0)
            if duty_day.month == 1:
                month_start_dt = datetime(duty_day.year - 1, 12, 1, 0, 0, 0)
            else:
                month_start_dt = datetime(duty_day.year, duty_day.month - 1, 1, 0, 0, 0)
        else:
            month_start_dt = datetime(duty_day.year, duty_day.month, 1, 0, 0, 0)
            if duty_day.month == 12:
                month_end_dt = datetime(duty_day.year + 1, 1, 1, 0, 0, 0)
            else:
                month_end_dt = datetime(duty_day.year, duty_day.month + 1, 1, 0, 0, 0)
        prev_day = datetime(duty_day.year, duty_day.month, duty_day.day, 0, 0, 0) - timedelta(days=1)
        return month_start_dt, month_end_dt, prev_day

    @staticmethod
    def _build_total_electricity_filter_formula(*, start_date: str, end_date: str) -> str:
        fields_cfg = _CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("fields", {})
        if not isinstance(fields_cfg, dict):
            fields_cfg = {}
        date_field = str(fields_cfg.get("date", "日期") or "日期").strip()
        category_field = str(fields_cfg.get("category", "汇总分类") or "汇总分类").strip()
        category_value = str(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("category_value", "总用电量") or "总用电量").strip()
        return (
            f'AND(CurrentValue.[{date_field}]>=TODATE({_formula_literal(start_date)}), '
            f'CurrentValue.[{date_field}]<TODATE({_formula_literal(end_date)}), '
            f'CurrentValue.[{category_field}]={_formula_literal(category_value)})'
        )

    def _query_total_electricity_batch(
        self,
        *,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        duty_date_text = _text(duty_date)
        month_start_dt, month_end_dt, prev_day_dt = self._total_electricity_window(duty_date_text)
        start_text = month_start_dt.strftime("%Y-%m-%d")
        end_text = month_end_dt.strftime("%Y-%m-%d")
        prev_day_text = prev_day_dt.strftime("%Y-%m-%d")
        cache_key = (start_text, end_text, prev_day_text)
        ttl_sec = max(1, int(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("cache_ttl_sec", 30) or 30))
        while True:
            cached = _read_short_cache(
                _TOTAL_ELECTRICITY_CACHE,
                _TOTAL_ELECTRICITY_CACHE_LOCK,
                cache_key,
                ttl_sec=ttl_sec,
            )
            if cached is not None:
                emit_log(
                    "[交接班][容量报表][总用电量] 命中短缓存 "
                    f"window={start_text}~{end_text}, prev_day={prev_day_text}, raw={cached.get('raw_count', 0)}, matched={cached.get('matched_count', 0)}"
                )
                return cached
            leader, event = _claim_singleflight(_TOTAL_ELECTRICITY_INFLIGHT, _TOTAL_ELECTRICITY_CACHE_LOCK, cache_key)
            if leader:
                break
            event.wait()

        fields_cfg = _CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("fields", {})
        if not isinstance(fields_cfg, dict):
            fields_cfg = {}
        category_field = str(fields_cfg.get("category", "") or "").strip()
        building_field = str(fields_cfg.get("building", "") or "").strip()
        date_field = str(fields_cfg.get("date", "") or "").strip()
        value_field = str(fields_cfg.get("value", "") or "").strip()
        table_id = str(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("table_id", "") or "").strip()
        filter_formula = self._build_total_electricity_filter_formula(start_date=start_text, end_date=end_text)
        category_value = str(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("category_value", "总用电量") or "总用电量").strip()
        result: Dict[str, Any] = {
            "start_date": start_text,
            "end_date": end_text,
            "prev_day": prev_day_text,
            "by_building": {},
            "warnings": [],
            "raw_count": 0,
            "matched_count": 0,
        }
        for building_name in _BUILDER_BY_BUILDING:
            result["by_building"][building_name] = {
                "prev_day_value": "0",
                "month_total": "0",
                "prev_day_found": False,
                "prev_day_value_found": False,
                "month_records": 0,
                "month_value_records": 0,
            }
        try:
            client = self._new_total_electricity_client()
            option_maps = self._load_field_option_maps(
                client=client,
                table_id=table_id,
                target_fields=[category_field, building_field],
                emit_log=emit_log,
            )
            emit_log(
                "[交接班][容量报表][总用电量] 查询开始 "
                f"window={start_text}~{end_text}, prev_day={prev_day_text}, filter={filter_formula}"
            )
            records = client.list_records(
                table_id=table_id,
                page_size=max(1, int(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("page_size", 500) or 500)),
                max_records=int(_CAPACITY_TOTAL_ELECTRICITY_SOURCE.get("max_records", 0) or 0),
                filter_formula=filter_formula,
            )
        except Exception as exc:  # noqa: BLE001
            warning = f"总用电量查询失败，V57/Y57 已按 0 处理: {exc}"
            result["warnings"].append(warning)
            emit_log(
                "[交接班][容量报表][总用电量] 查询失败 "
                f"window={start_text}~{end_text}, prev_day={prev_day_text}, error={exc}"
            )
            _store_short_cache(_TOTAL_ELECTRICITY_CACHE, _TOTAL_ELECTRICITY_CACHE_LOCK, cache_key, result)
            _finish_singleflight(_TOTAL_ELECTRICITY_INFLIGHT, _TOTAL_ELECTRICITY_CACHE_LOCK, cache_key, event)
            return result

        result["raw_count"] = len(records)
        category_option_map = option_maps.get(category_field, {})
        building_option_map = option_maps.get(building_field, {})
        month_totals: Dict[str, float] = {building_name: 0.0 for building_name in _BUILDER_BY_BUILDING}
        prev_values: Dict[str, float | None] = {building_name: None for building_name in _BUILDER_BY_BUILDING}
        for item in records:
            if not isinstance(item, dict):
                continue
            fields = item.get("fields", {})
            if not isinstance(fields, dict):
                continue
            category_text = _field_text_with_option_map(fields.get(category_field), category_option_map)
            if category_text != category_value:
                continue
            record_date = _date_only(fields.get(date_field))
            if not record_date or record_date < start_text or record_date >= end_text:
                continue
            matched_building = ""
            record_building = _field_text_with_option_map(fields.get(building_field), building_option_map)
            for building_name in _BUILDER_BY_BUILDING:
                if self._matches_building(record_building, building_name):
                    matched_building = building_name
                    break
            if not matched_building:
                continue
            bucket = result["by_building"][matched_building]
            bucket["month_records"] = int(bucket.get("month_records", 0) or 0) + 1
            value = _to_float(fields.get(value_field))
            if record_date == prev_day_text:
                bucket["prev_day_found"] = True
                if value is not None:
                    bucket["prev_day_value_found"] = True
                    prev_values[matched_building] = value
            if value is None:
                continue
            bucket["month_value_records"] = int(bucket.get("month_value_records", 0) or 0) + 1
            month_totals[matched_building] += value
            result["matched_count"] = int(result.get("matched_count", 0) or 0) + 1

        for building_name, bucket in result["by_building"].items():
            prev_value = prev_values.get(building_name)
            bucket["prev_day_value"] = format_number(prev_value if prev_value is not None else 0)
            bucket["month_total"] = format_number(month_totals.get(building_name, 0.0))
            if not bool(bucket.get("prev_day_found")):
                result["warnings"].append(f"{building_name} V57 缺少 {prev_day_text} 总用电量记录，已按 0 写入")
            elif not bool(bucket.get("prev_day_value_found")):
                result["warnings"].append(f"{building_name} V57 {prev_day_text} 总用电量数值为空，已按 0 写入")
            if int(bucket.get("month_records", 0) or 0) <= 0:
                result["warnings"].append(f"{building_name} Y57 缺少 {start_text}~{end_text} 总用电量记录，已按 0 写入")
            elif int(bucket.get("month_value_records", 0) or 0) <= 0:
                result["warnings"].append(f"{building_name} Y57 {start_text}~{end_text} 总用电量数值为空，已按 0 写入")
            emit_log(
                "[交接班][容量报表][总用电量] 楼栋结果 "
                f"building={building_name}, V57={bucket.get('prev_day_value', '0')}, Y57={bucket.get('month_total', '0')}, "
                f"matched_month_records={int(bucket.get('month_records', 0) or 0)}, prev_day_found={bool(bucket.get('prev_day_found'))}"
            )
        emit_log(
            "[交接班][容量报表][总用电量] 查询完成 "
            f"window={start_text}~{end_text}, prev_day={prev_day_text}, raw={result.get('raw_count', 0)}, matched={result.get('matched_count', 0)}"
        )
        _store_short_cache(_TOTAL_ELECTRICITY_CACHE, _TOTAL_ELECTRICITY_CACHE_LOCK, cache_key, result)
        _finish_singleflight(_TOTAL_ELECTRICITY_INFLIGHT, _TOTAL_ELECTRICITY_CACHE_LOCK, cache_key, event)
        return result

    def query_total_electricity_summary(
        self,
        *,
        building: str,
        duty_date: str,
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        building_text = _text(building)
        batch = self._query_total_electricity_batch(duty_date=duty_date, emit_log=emit_log)
        by_building = batch.get("by_building", {}) if isinstance(batch.get("by_building", {}), dict) else {}
        summary = by_building.get(building_text, {}) if isinstance(by_building.get(building_text, {}), dict) else {}
        payload = {
            "V57": _text(summary.get("prev_day_value")) or "0",
            "Y57": _text(summary.get("month_total")) or "0",
            "start_date": _text(batch.get("start_date")),
            "end_date": _text(batch.get("end_date")),
            "prev_day": _text(batch.get("prev_day")),
            "warnings": [
                str(item or "").strip()
                for item in (batch.get("warnings", []) if isinstance(batch.get("warnings", []), list) else [])
                if str(item or "").strip() and str(item or "").strip().startswith(building_text)
            ],
        }
        payload["signature"] = "|".join(
            [
                f"start={payload['start_date']}",
                f"end={payload['end_date']}",
                f"prev={payload['prev_day']}",
                f"building={building_text}",
                f"V57={payload['V57']}",
                f"Y57={payload['Y57']}",
            ]
        )
        return payload

    def _build_capacity_overlay_values(
        self,
        *,
        building: str,
        duty_date: str,
        handover_cells: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, str]:
        handover = handover_cells if isinstance(handover_cells, dict) else {}
        water_summary = self._query_capacity_water_summary(
            building=building,
            duty_date=duty_date,
            emit_log=emit_log,
        )
        total_electricity = self.query_total_electricity_summary(
            building=building,
            duty_date=duty_date,
            emit_log=emit_log,
        )
        weather_payload = self._fetch_weather_payload_for_duty_date(duty_date=duty_date, emit_log=emit_log)
        weather_text = _text(weather_payload.get("text"))
        weather_humidity = _text(weather_payload.get("humidity"))
        west_tank, east_tank = self._derive_tank_pair_from_f8(handover.get("F8"))
        overlay = {
            "AC24": _text(handover.get("D8")),
            "U15": _text(handover.get("H6")),
            "AD22": west_tank,
            "AD23": east_tank,
            "V60": _text(handover.get("B6")),
            "O60": _text(handover.get("D6")),
            "S60": _text(handover.get("F6")),
            "AB56": _text(handover.get("B13")),
            "AC56": _text(handover.get("D13")),
            "L2": weather_text,
            "X2": weather_humidity,
            "AC25": _text(water_summary.get("month_total")),
            "O57": _text(water_summary.get("latest_daily_total")),
            "V57": _text(total_electricity.get("V57")) or "0",
            "Y57": _text(total_electricity.get("Y57")) or "0",
        }
        return {cell: value for cell, value in overlay.items() if value != ""}

    @staticmethod
    def _build_substation_110kv_values(shared_110kv: Dict[str, Any] | None) -> Dict[str, str]:
        block = shared_110kv if isinstance(shared_110kv, dict) else {}
        rows = block.get("rows", [])
        if not isinstance(rows, list):
            rows = []
        if not rows:
            return {}
        by_label = {
            _text(row.get("label")): row
            for row in rows
            if isinstance(row, dict) and _text(row.get("label"))
        }
        values: Dict[str, str] = {}
        for label, row_number in _SUBSTATION_110KV_TARGET_ROWS.items():
            source = by_label.get(label, {})
            for key, column in _SUBSTATION_110KV_TARGET_COLUMNS.items():
                values[f"{column}{row_number}"] = _text(source.get(key)) if isinstance(source, dict) else ""
        return values

    @staticmethod
    def _build_cooling_pump_pressure_values(cooling_pump_pressures: Dict[str, Any] | None) -> Dict[str, str]:
        payload = cooling_pump_pressures if isinstance(cooling_pump_pressures, dict) else {}
        rows = payload.get("rows", [])
        values: Dict[str, str] = {}
        for inlet_cell, outlet_cell in _COOLING_PUMP_PRESSURE_TARGETS.values():
            values[inlet_cell] = ""
            values[outlet_cell] = ""
        if not isinstance(rows, list):
            return values
        zone_positions: Dict[str, int] = {"west": 0, "east": 0}
        for row in rows:
            if not isinstance(row, dict):
                continue
            zone = _text(row.get("zone")).lower()
            if zone not in zone_positions:
                continue
            position = int(row.get("position", 0) or 0)
            if position <= 0:
                zone_positions[zone] += 1
                position = zone_positions[zone]
            if position not in {1, 2}:
                continue
            target = _COOLING_PUMP_PRESSURE_TARGETS.get((zone, position))
            if not target:
                continue
            inlet_cell, outlet_cell = target
            values[inlet_cell] = _text(row.get("inlet_pressure"))
            values[outlet_cell] = _text(row.get("outlet_pressure"))
        return values

    @staticmethod
    def _capacity_file_fingerprint(capacity_output_file: str) -> Dict[str, Any]:
        path = Path(_text(capacity_output_file))
        if not _text(capacity_output_file) or not path.exists() or not path.is_file():
            return {
                "path": _text(capacity_output_file),
                "exists": False,
                "size": 0,
                "mtime_ns": 0,
            }
        try:
            stat = path.stat()
            return {
                "path": str(path.resolve(strict=False)),
                "exists": True,
                "size": int(getattr(stat, "st_size", 0) or 0),
                "mtime_ns": int(getattr(stat, "st_mtime_ns", 0) or int(getattr(stat, "st_mtime", 0) or 0)),
            }
        except Exception:  # noqa: BLE001
            return {
                "path": str(path),
                "exists": False,
                "size": 0,
                "mtime_ns": 0,
            }

    def build_capacity_overlay_signature(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        handover_cells: Dict[str, Any],
        capacity_output_file: str,
        shared_110kv: Dict[str, Any] | None = None,
        cooling_pump_pressures: Dict[str, Any] | None = None,
        client_id: str = "",
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        handover = handover_cells if isinstance(handover_cells, dict) else {}
        input_signature = self.capacity_input_signature(handover)
        valid, error = self._validate_capacity_overlay_inputs(handover)
        overlay_values = self._build_capacity_overlay_values(
            building=building,
            duty_date=duty_date,
            handover_cells=handover,
            emit_log=emit_log,
        )
        shared_block = (
            shared_110kv
            if isinstance(shared_110kv, dict)
            else self._shared_substation_110kv_for_batch(
                duty_date=duty_date,
                duty_shift=duty_shift,
                client_id=client_id,
                emit_log=emit_log,
            )
        )
        overlay_values.update(self._build_substation_110kv_values(shared_block))
        if isinstance(cooling_pump_pressures, dict):
            overlay_values.update(self._build_cooling_pump_pressure_values(cooling_pump_pressures))
        normalized_values = {
            _text(cell).upper(): _text(value)
            for cell, value in overlay_values.items()
            if _text(cell)
        }
        file_fingerprint = self._capacity_file_fingerprint(capacity_output_file)
        payload = {
            "version": "capacity-overlay-signature-v2",
            "building": _text(building),
            "duty_date": _text(duty_date),
            "duty_shift": _text(duty_shift).lower(),
            "input_signature": input_signature,
            "file": file_fingerprint,
            "values": {cell: normalized_values[cell] for cell in sorted(normalized_values)},
        }
        signature = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        return {
            "signature": signature,
            "input_signature": input_signature,
            "valid": bool(valid),
            "error": error,
            "file": file_fingerprint,
            "values": payload["values"],
        }

    def _shared_substation_110kv_for_batch(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        client_id: str = "",
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        batch_key = ReviewSessionService.build_batch_key(_text(duty_date), _text(duty_shift).lower())
        if not batch_key:
            return {}
        try:
            state = self._review_session_service.get_substation_110kv_state(
                batch_key=batch_key,
                client_id=_text(client_id),
            )
            block = state.get("shared_blocks", {}).get("substation_110kv", {})
            return block if isinstance(block, dict) else {}
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][容量报表][110KV] 读取共享数据失败 batch={batch_key}, error={exc}")
            return {}

    def _cooling_pump_pressures_from_defaults(
        self,
        *,
        building: str,
        running_units: Dict[str, List[Dict[str, Any]]],
    ) -> Dict[str, Any]:
        try:
            raw_defaults = ReviewBuildingDocumentStore(config=self.config, building=building).get_default("cooling_pump_pressures")
        except Exception:  # noqa: BLE001
            raw_defaults = {}
        defaults = raw_defaults if isinstance(raw_defaults, dict) else {}
        rows: List[Dict[str, Any]] = []
        for zone in ("west", "east"):
            zone_label = "西区" if zone == "west" else "东区"
            for position, unit_info in enumerate(list((running_units or {}).get(zone, []))[:2], start=1):
                try:
                    unit = int(unit_info.get("unit", 0) or 0)
                except Exception:  # noqa: BLE001
                    unit = 0
                if unit <= 0:
                    continue
                key = f"{zone}:{unit}"
                default = defaults.get(key, {}) if isinstance(defaults.get(key, {}), dict) else {}
                rows.append(
                    {
                        "row_id": key,
                        "zone": zone,
                        "zone_label": zone_label,
                        "unit": unit,
                        "unit_label": f"{unit}#制冷单元",
                        "position": position,
                        "mode_text": _text(unit_info.get("mode_text")),
                        "inlet_pressure": _text(default.get("inlet_pressure")),
                        "outlet_pressure": _text(default.get("outlet_pressure")),
                        "cooling_tower_level": _text(default.get("cooling_tower_level")),
                    }
                )
        tanks: Dict[str, Dict[str, str]] = {}
        for zone in ("west", "east"):
            key = f"tank:{zone}"
            default = defaults.get(key, {}) if isinstance(defaults.get(key, {}), dict) else {}
            tanks[zone] = {
                "zone": zone,
                "zone_label": "西区" if zone == "west" else "东区",
                "temperature": _text(default.get("temperature")),
                "level": _text(default.get("level")),
            }
        return {"rows": rows, "tanks": tanks}

    @staticmethod
    def _format_with_unit(value: Any, unit: str) -> str:
        text = _text(value)
        if not text:
            return ""
        return text if unit and unit in text else f"{text}{unit}"

    @staticmethod
    def _extract_equipment_numbers(row: Any) -> List[int]:
        text = " ".join(
            _text(getattr(row, attr, ""))
            for attr in ("d_name", "c_text", "b_text")
        )
        numbers: List[int] = []
        for match in re.finditer(r"(?<!\d)([1-9])\s*[#号]", text):
            number = int(match.group(1))
            if number not in numbers:
                numbers.append(number)
        return sorted(numbers)

    def _secondary_pump_running_text(self, query: CapacitySourceQuery, *, zone: str) -> str:
        rows = query.rows_by_d_regexes(
            [re.escape(alias) for alias in _COOLING_SECONDARY_PUMP_ALIASES],
            zone=zone,
            allow_global=True,
        )
        running_rows = [
            row for row in rows
            if getattr(row, "value", None) is not None and float(getattr(row, "value", 0) or 0) > 10
        ]
        if not running_rows:
            return ""
        numbers: List[int] = []
        for row in running_rows:
            for number in self._extract_equipment_numbers(row):
                if number not in numbers:
                    numbers.append(number)
        if numbers:
            return f"{''.join(f'{number}#' for number in sorted(numbers))}二次泵运行正常"
        return f"{len(running_rows)}台二次泵运行正常"

    @staticmethod
    def _manual_cooling_rows(cooling_pump_pressures: Dict[str, Any] | None, *, zone: str) -> Dict[int, Dict[str, Any]]:
        payload = cooling_pump_pressures if isinstance(cooling_pump_pressures, dict) else {}
        rows = payload.get("rows", []) if isinstance(payload.get("rows", []), list) else []
        output: Dict[int, Dict[str, Any]] = {}
        for row in rows:
            if not isinstance(row, dict) or _text(row.get("zone")).lower() != zone:
                continue
            try:
                unit = int(row.get("unit", 0) or 0)
            except Exception:  # noqa: BLE001
                unit = 0
            if unit > 0:
                output[unit] = row
        return output

    @staticmethod
    def _manual_cooling_tank(cooling_pump_pressures: Dict[str, Any] | None, *, zone: str) -> Dict[str, Any]:
        payload = cooling_pump_pressures if isinstance(cooling_pump_pressures, dict) else {}
        tanks = payload.get("tanks", {}) if isinstance(payload.get("tanks", {}), dict) else {}
        tank = tanks.get(zone, {}) if isinstance(tanks.get(zone, {}), dict) else {}
        return tank

    def _cooling_zone_sentence(
        self,
        *,
        zone: str,
        running_units: Dict[str, List[Dict[str, Any]]],
        query: CapacitySourceQuery,
        cooling_pump_pressures: Dict[str, Any] | None = None,
    ) -> tuple[str, List[str]]:
        zone_name = "A区" if zone == "west" else "B区"
        active_units = list((running_units or {}).get(zone, []))
        running_count = len(active_units)
        backup_count = max(0, 3 - running_count)
        warnings: List[str] = []
        manual_rows = self._manual_cooling_rows(cooling_pump_pressures, zone=zone)
        manual_tank = self._manual_cooling_tank(cooling_pump_pressures, zone=zone)
        if running_count <= 0:
            warnings.append(f"冷冻站{zone_name}未识别到运行制冷单元")
        parts: List[str] = [
            f"冷冻站{zone_name}3套制冷单元{running_count}用{backup_count}备",
            "群控模式为开启状态",
            "备用机组与备用二次泵状态正常可用",
        ]
        for unit_info in active_units:
            try:
                unit = int(unit_info.get("unit", 0) or 0)
            except Exception:  # noqa: BLE001
                unit = 0
            if unit <= 0:
                continue
            mode_text = _text(unit_info.get("mode_text"))
            if mode_text:
                parts.append(f"{unit}#制冷单元{mode_text}模式运行正常")
            else:
                parts.append(f"{unit}#制冷单元运行正常")
            tower_level = _text(manual_rows.get(unit, {}).get("cooling_tower_level")) or query.first_text_by_d_aliases(
                _COOLING_TOWER_LEVEL_ALIASES,
                zone=zone,
                unit=unit,
                allow_global=False,
            )
            if tower_level:
                parts.append(f"{unit}#冷却塔液位{self._format_with_unit(tower_level, 'm')}正常")
            else:
                warnings.append(f"冷冻站{zone_name}{unit}#冷却塔液位未识别")

        secondary_text = self._secondary_pump_running_text(query, zone=zone)
        if secondary_text:
            parts.append(secondary_text)
        else:
            warnings.append(f"冷冻站{zone_name}二次泵运行信息未识别")

        tank_temp = _text(manual_tank.get("temperature")) or query.first_text_by_d_aliases(
            _COOLING_TANK_TEMP_ALIASES,
            zone=zone,
            allow_global=True,
        )
        tank_level = _text(manual_tank.get("level")) or query.first_text_by_d_aliases(
            _COOLING_TANK_LEVEL_ALIASES,
            zone=zone,
            allow_global=True,
        )
        tank_parts: List[str] = []
        if tank_temp:
            tank_parts.append(f"后备温度{self._format_with_unit(tank_temp, '℃')}正常")
        else:
            warnings.append(f"冷冻站{zone_name}蓄冷罐后备温度未识别")
        if tank_level:
            tank_parts.append(f"液位{self._format_with_unit(tank_level, 'm')}正常")
        else:
            warnings.append(f"冷冻站{zone_name}蓄冷罐液位未识别")
        if tank_parts:
            parts.append(f"蓄冷罐{'、'.join(tank_parts)}")
        return "，".join(parts).rstrip("，") + "；", warnings

    def build_capacity_cooling_summary(
        self,
        *,
        capacity_rows: List[Any],
        running_units: Dict[str, List[Dict[str, Any]]],
        cooling_pump_pressures: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        query = CapacitySourceQuery(capacity_rows if isinstance(capacity_rows, list) else [])
        lines: Dict[str, str] = {}
        warnings: List[str] = []
        for zone in ("west", "east"):
            line, zone_warnings = self._cooling_zone_sentence(
                zone=zone,
                running_units=running_units,
                query=query,
                cooling_pump_pressures=cooling_pump_pressures,
            )
            lines[zone] = line
            warnings.extend(zone_warnings)
        return {
            "version": "capacity-cooling-summary-v1",
            "lines": lines,
            "warnings": warnings,
        }

    def _validate_capacity_overlay_inputs(self, handover_cells: Dict[str, Any]) -> tuple[bool, str]:
        handover = handover_cells if isinstance(handover_cells, dict) else {}
        missing_cells = [cell for cell in _CAPACITY_SYNC_REQUIRED_CELLS if not _text(handover.get(cell))]
        if missing_cells:
            return False, f"容量报表待补写输入不完整: 缺少{','.join(missing_cells)}"
        west_tank, east_tank = self._derive_tank_pair_from_f8(handover.get("F8"))
        if not west_tank or not east_tank:
            return False, "容量报表待补写输入不完整: F8 未解析出西区/东区数字"
        return True, ""

    def sync_overlay_for_existing_report(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        handover_output_file: str,
        capacity_output_file: str,
        handover_sheet_name: str = "",
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        handover_cells = self._read_handover_cells(
            handover_output_file,
            list(_CAPACITY_TRACKED_CELLS),
            sheet_name=handover_sheet_name,
        )
        return self.sync_overlay_for_existing_report_from_cells(
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            handover_cells=handover_cells,
            capacity_output_file=capacity_output_file,
            emit_log=emit_log,
        )

    def sync_overlay_for_existing_report_from_cells(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        handover_cells: Dict[str, Any],
        capacity_output_file: str,
        shared_110kv: Dict[str, Any] | None = None,
        cooling_pump_pressures: Dict[str, Any] | None = None,
        client_id: str = "",
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        input_signature = self.capacity_input_signature(handover_cells)
        capacity_output_path = Path(_text(capacity_output_file))
        if not _text(capacity_output_file) or not capacity_output_path.exists():
            return self._build_capacity_sync_payload(
                status="missing_file",
                error="交接班容量报表文件不存在，请重新生成",
                input_signature=input_signature,
            )
        valid, error = self._validate_capacity_overlay_inputs(handover_cells)

        try:
            workbook = load_workbook_quietly(capacity_output_path)
            try:
                sheet_name = self._sheet_name(workbook, _text(self._template_cfg().get("sheet_name")))
                sheet = workbook[sheet_name]
                overlay_values = self._build_capacity_overlay_values(
                    building=building,
                    duty_date=duty_date,
                    handover_cells=handover_cells,
                    emit_log=emit_log,
                )
                shared_block = (
                    shared_110kv
                    if isinstance(shared_110kv, dict)
                    else self._shared_substation_110kv_for_batch(
                        duty_date=duty_date,
                        duty_shift=duty_shift,
                        client_id=client_id,
                        emit_log=emit_log,
                    )
                )
                overlay_values.update(self._build_substation_110kv_values(shared_block))
                if isinstance(cooling_pump_pressures, dict):
                    overlay_values.update(self._build_cooling_pump_pressure_values(cooling_pump_pressures))
                _write_cells_with_merged_support(sheet, overlay_values)
                atomic_save_workbook(workbook, capacity_output_path)
            finally:
                workbook.close()
        except Exception as exc:  # noqa: BLE001
            emit_log(
                "[交接班][容量报表][补写] 失败 "
                f"building={building}, duty={duty_date}/{duty_shift}, error={exc}"
            )
            return self._build_capacity_sync_payload(
                status="failed",
                error=f"容量报表补写失败: {exc}",
                input_signature=input_signature,
            )

        emit_log(
            "[交接班][容量报表][补写] 完成 "
            f"building={building}, duty={duty_date}/{duty_shift}, output={capacity_output_path}"
        )
        if not valid:
            return self._build_capacity_sync_payload(
                status="pending_input",
                error=error,
                input_signature=input_signature,
            )
        return self._build_capacity_sync_payload(
            status="ready",
            input_signature=input_signature,
        )

    def generate(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        handover_output_file: str,
        capacity_source_file: str,
        roster_assignment: Any | None,
        current_alarm_summary: Dict[str, Any] | None,
        previous_alarm_summary: Dict[str, Any] | None,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None,
        hvdc_source_d_name: Any = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        building_text = _text(building)
        duty_date_text = _text(duty_date)
        duty_shift_text = _text(duty_shift).lower()
        if not building_text or not duty_date_text or duty_shift_text not in {"day", "night"}:
            raise ValueError("生成交接班容量报表缺少楼栋或班次上下文")

        output_file = self._next_available_output_path(
            building=building_text,
            duty_date=duty_date_text,
            duty_shift=duty_shift_text,
        )
        output_file.parent.mkdir(parents=True, exist_ok=True)
        template_selection = self.resolve_template_selection(building=building_text)
        template_path = template_selection["path"]
        if not template_path.exists():
            raise FileNotFoundError(f"交接班容量报表模板不存在: {template_path}")

        handover_sheet_name = _text(
            (
                self.config.get("template", {})
                if isinstance(self.config.get("template", {}), dict)
                else {}
            ).get("sheet_name", "")
        )
        handover_cells = self._read_handover_cells(
            handover_output_file,
            [
                "B4",
                "F4",
                "B6",
                "D6",
                "F6",
                "H6",
                "D8",
                "F8",
                "B13",
                "D13",
                "C3",
                "G3",
                "B7",
                "D7",
                "B10",
                "D10",
                "B15",
                "D15",
                "F15",
            ],
            sheet_name=handover_sheet_name,
        )
        capacity_rows = self._load_capacity_rows(capacity_source_file)
        oil_current = self._extract_current_oil_display_values(
            building=building_text,
            rows=capacity_rows,
            emit_log=emit_log,
        )
        oil_previous, previous_oil_warning = self._load_previous_capacity_display_oil_values(
            building=building_text,
            duty_date=duty_date_text,
            duty_shift=duty_shift_text,
            current_display_values=oil_current,
            emit_log=emit_log,
        )
        warnings: List[str] = []
        if previous_oil_warning:
            warnings.append(previous_oil_warning)

        current_alarm = self._normalize_alarm_summary(current_alarm_summary)
        previous_alarm = self._normalize_alarm_summary(previous_alarm_summary)
        capacity_water_summary = self._query_capacity_water_summary(
            building=building_text,
            duty_date=duty_date_text,
            emit_log=emit_log,
        )
        weather_payload = self._fetch_weather_payload_for_duty_date(
            duty_date=duty_date_text,
            emit_log=emit_log,
        )
        weather_text = _text(weather_payload.get("text"))
        weather_humidity = _text(weather_payload.get("humidity"))
        builder = _BUILDER_BY_BUILDING.get(building_text)
        if builder is None:
            raise ValueError(f"不支持的容量报表楼栋: {building_text}")

        hvdc_debug = self._resolve_hvdc_text(
            resolved_values_by_id=resolved_values_by_id,
            metric_origin_context=metric_origin_context,
            hvdc_source_d_name=hvdc_source_d_name,
        )
        emit_log(
            "[交接班][容量报表][HVDC] "
            f"building={building_text}, raw_value={hvdc_debug.get('raw_value', '') or '-'}, "
            f"source_d_name={hvdc_debug.get('source_d_name', '') or '-'}, "
            f"position_code={hvdc_debug.get('position_code', '') or '-'}, "
            f"H17={hvdc_debug.get('formatted', '') or '-'}"
        )

        workbook = load_workbook_quietly(template_path)
        try:
            sheet_name = self._sheet_name(workbook, _text(self._template_cfg().get("sheet_name")))
            sheet = workbook[sheet_name]
            template_snapshot = build_capacity_template_snapshot(sheet, building_text)
            template_snapshot["template_family"] = _text(template_selection.get("template_family"))
            running_units = self._resolve_running_units(resolved_values_by_id)
            cooling_pump_pressures = self._cooling_pump_pressures_from_defaults(
                building=building_text,
                running_units=running_units,
            )
            capacity_cooling_summary = self.build_capacity_cooling_summary(
                capacity_rows=capacity_rows,
                running_units=running_units,
                cooling_pump_pressures=cooling_pump_pressures,
            )
            shared_110kv = self._shared_substation_110kv_for_batch(
                duty_date=duty_date_text,
                duty_shift=duty_shift_text,
                emit_log=emit_log,
            )
            context = {
                "building": building_text,
                "duty_date": duty_date_text,
                "duty_shift": duty_shift_text,
                "handover_cells": handover_cells,
                "roster": {
                    "current_team": _text(getattr(roster_assignment, "current_team", "")),
                    "next_team": _text(getattr(roster_assignment, "next_team", "")),
                },
                "current_alarm_summary": current_alarm,
                "previous_alarm_summary": previous_alarm,
                "oil_previous": oil_previous,
                "oil_current": oil_current,
                "capacity_water_summary": capacity_water_summary,
                "night_water_summary": capacity_water_summary,
                "weather_text": weather_text,
                "weather_humidity": weather_humidity,
                "hvdc_text": hvdc_debug.get("formatted", ""),
                "capacity_rows": capacity_rows,
                "running_units": running_units,
                "resolved_values_by_id": resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {},
                "template_snapshot": template_snapshot,
            }
            cell_values = builder(context)
            cell_values.update(_build_fixed_header_cells(building_text))
            cell_values.update(
                self._build_capacity_overlay_values(
                    building=building_text,
                    duty_date=duty_date_text,
                    handover_cells=handover_cells,
                    emit_log=emit_log,
                )
            )
            cell_values.update(self._build_substation_110kv_values(shared_110kv))
            cell_values.update(self._build_cooling_pump_pressure_values(cooling_pump_pressures))
            _write_cells_with_merged_support(sheet, cell_values)
            atomic_save_workbook(workbook, output_file)
        finally:
            workbook.close()

        emit_log(
            "[交接班][容量报表] 生成完成 "
            f"building={building_text}, duty_date={duty_date_text}, duty_shift={duty_shift_text}, output={output_file}"
        )
        input_signature = self.capacity_input_signature({cell: handover_cells.get(cell, "") for cell in _CAPACITY_TRACKED_CELLS})
        valid, validation_error = self._validate_capacity_overlay_inputs(handover_cells)
        if valid:
            capacity_sync = self._build_capacity_sync_payload(
                status="ready",
                input_signature=input_signature,
            )
        else:
            capacity_sync = self._build_capacity_sync_payload(
                status="pending_input",
                error=validation_error,
                input_signature=input_signature,
            )
        return {
            "status": "success",
            "output_file": str(output_file),
            "warnings": warnings,
            "error": "",
            "capacity_sync": capacity_sync,
            "running_units": running_units,
            "capacity_cooling_summary": capacity_cooling_summary,
        }


