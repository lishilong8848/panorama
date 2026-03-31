from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from pipeline_utils import get_app_dir
from handover_log_module.core.section_layout import parse_category_sections
from handover_log_module.repository.event_followup_cache_store import EventFollowupCacheStore
from handover_log_module.repository.event_sections_repository import (
    EventRow,
    EventQueryByBuilding,
    EventSectionQueryResult,
    EventSectionsRepository,
)


def _norm_header(value: Any) -> str:
    return str(value or "").replace(" ", "").strip().casefold()


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


@dataclass
class _HistoryRow:
    record_id: str
    payload: Dict[str, str]
    unresolved_snapshot: Dict[str, Any] | None


class EventCategoryPayloadBuilder:
    def __init__(
        self,
        handover_cfg: Dict[str, Any],
        *,
        repository: EventSectionsRepository | None = None,
        cache_store: EventFollowupCacheStore | None = None,
    ) -> None:
        self.handover_cfg = handover_cfg
        self.repo = repository or EventSectionsRepository(handover_cfg)
        self._cache_store = cache_store

    def _get_cache_store(self, cfg: Dict[str, Any]) -> EventFollowupCacheStore:
        if self._cache_store is not None:
            return self._cache_store
        cache_cfg = cfg.get("cache", {}) if isinstance(cfg.get("cache", {}), dict) else {}
        state_file = str(cache_cfg.get("state_file", "")).strip() or "handover_shared_cache.json"
        global_paths = self.handover_cfg.get("_global_paths", {})
        self._cache_store = EventFollowupCacheStore(
            cache_state_file=state_file,
            global_paths=global_paths if isinstance(global_paths, dict) else {},
        )
        return self._cache_store

    @staticmethod
    def _build_base_row_payload(
        *,
        row: EventRow,
        progress_text: str,
        work_window_text: str,
        follower_text: str,
    ) -> Dict[str, str]:
        return {
            "event_level": str(row.event_level or "").strip(),
            "event_time": _fmt_dt(row.event_time),
            "description": str(row.description or "").strip(),
            "work_window": str(work_window_text or "").strip() or "/",
            "progress": str(progress_text or "").strip(),
            "follower": str(follower_text or "").strip(),
        }

    @staticmethod
    def _make_pending_snapshot(*, row: EventRow) -> Dict[str, Any]:
        return {
            "record_id": row.record_id,
            "event_level": row.event_level,
            "event_time": _fmt_dt(row.event_time),
            "description": row.description,
            "building_text": row.building_text,
            "final_status_text": row.final_status_text,
            "excluded_checked": bool(row.excluded_checked),
            "to_maint": bool(row.to_maint),
            "maint_done_time": _fmt_dt(row.maint_done_time),
            "event_done_time": _fmt_dt(row.event_done_time),
        }

    @staticmethod
    def _row_from_snapshot(snapshot: Dict[str, Any]) -> EventRow:
        from datetime import datetime

        def _parse_dt(text: Any):
            raw = str(text or "").strip()
            if not raw:
                return None
            try:
                return datetime.strptime(raw, "%Y-%m-%d %H:%M:%S")
            except Exception:  # noqa: BLE001
                return None

        return EventRow(
            record_id=str(snapshot.get("record_id", "")).strip(),
            event_time=_parse_dt(snapshot.get("event_time")),
            event_level=str(snapshot.get("event_level", "")).strip(),
            description=str(snapshot.get("description", "")).strip(),
            building_text=str(snapshot.get("building_text", "")).strip(),
            final_status_text=str(snapshot.get("final_status_text", "")).strip(),
            excluded_checked=bool(snapshot.get("excluded_checked", False)),
            to_maint=bool(snapshot.get("to_maint", False)),
            maint_done_time=_parse_dt(snapshot.get("maint_done_time")),
            event_done_time=_parse_dt(snapshot.get("event_done_time")),
            raw_fields={},
        )

    def _resolve_template_path(self, source_path: str) -> Path:
        path = Path(str(source_path or "").strip())
        if path.is_absolute():
            return path
        return get_app_dir() / path

    def _resolve_section_column_mapping(
        self,
        *,
        cfg: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Dict[str, str]]:
        col_cfg = cfg.get("column_mapping", {}) if isinstance(cfg.get("column_mapping", {}), dict) else {}
        fallback_cols = col_cfg.get("fallback_cols", {}) if isinstance(col_cfg.get("fallback_cols", {}), dict) else {}
        header_alias = col_cfg.get("header_alias", {}) if isinstance(col_cfg.get("header_alias", {}), dict) else {}
        resolve_by_header = bool(col_cfg.get("resolve_by_header", True))
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        new_event_name = str(sections_cfg.get("new_event", "新事件处理")).strip() or "新事件处理"
        history_name = str(sections_cfg.get("history_followup", "历史事件跟进")).strip() or "历史事件跟进"

        semantic_keys = ["event_level", "event_time", "description", "work_window", "progress", "follower"]
        default_map = {
            key: str(fallback_cols.get(key, "")).strip().upper() for key in semantic_keys
        }
        for key in semantic_keys:
            if not default_map.get(key):
                default_map[key] = {
                    "event_level": "B",
                    "event_time": "C",
                    "description": "D",
                    "work_window": "E",
                    "progress": "F",
                    "follower": "G",
                }[key]
        output = {
            new_event_name: dict(default_map),
            history_name: dict(default_map),
        }
        if not resolve_by_header:
            return output

        template_cfg = self.handover_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return output
        source_path = self._resolve_template_path(str(template_cfg.get("source_path", "")).strip())
        sheet_name = str(template_cfg.get("sheet_name", "交接班日志")).strip() or "交接班日志"
        if not source_path.exists():
            return output

        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return output
            ws = wb[sheet_name]
            sections = parse_category_sections(ws)
            section_by_name = {sec.name: sec for sec in sections}
            for sec_name in (new_event_name, history_name):
                sec = section_by_name.get(sec_name)
                if sec is None:
                    continue
                header_row = sec.header_row
                normalized_header_map: Dict[str, str] = {}
                max_col = int(ws.max_column or 0)
                for col_idx in range(1, max_col + 1):
                    text = _norm_header(ws.cell(row=header_row, column=col_idx).value)
                    if not text:
                        continue
                    col_letter = get_column_letter(col_idx)
                    normalized_header_map[text] = col_letter
                for semantic_key in semantic_keys:
                    aliases_raw = header_alias.get(semantic_key, [])
                    aliases = aliases_raw if isinstance(aliases_raw, list) else []
                    found_col = ""
                    for alias in aliases:
                        norm_alias = _norm_header(alias)
                        if not norm_alias:
                            continue
                        if norm_alias in normalized_header_map:
                            found_col = normalized_header_map[norm_alias]
                            break
                    if found_col:
                        output[sec_name][semantic_key] = found_col
            emit_log(
                f"[交接班][事件分类] 列映射已解析: new={output.get(new_event_name)}, history={output.get(history_name)}"
            )
            return output
        finally:
            wb.close()

    @staticmethod
    def _to_cells(row_payload: Dict[str, str], col_map: Dict[str, str]) -> Dict[str, str]:
        cells: Dict[str, str] = {}
        for semantic_key, text in row_payload.items():
            col = str(col_map.get(semantic_key, "")).strip().upper()
            if not col:
                continue
            cells[col] = str(text or "")
        return cells

    @staticmethod
    def _get_new_event_progress_text(
        *,
        row: EventRow,
        shift_end,
        progress_cfg: Dict[str, Any],
    ) -> str:
        done_text = str(progress_cfg.get("done", "已完成")).strip() or "已完成"
        todo_text = str(progress_cfg.get("todo", "未完成")).strip() or "未完成"
        if row.event_done_time is None or row.event_done_time > shift_end:
            return todo_text
        return done_text

    def build(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        follower_text: str,
        is_current_duty_context: bool,
        preloaded_query_result_by_building: EventQueryByBuilding | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        try:
            if preloaded_query_result_by_building is not None and building in preloaded_query_result_by_building:
                query_result = preloaded_query_result_by_building[building]
                emit_log(
                    "[交接班][事件分类] 命中批量预取: "
                    f"building={building}, current={len(query_result.current_rows)}, "
                    f"outside={len(query_result.outside_shift_ongoing_rows)}, "
                    f"history={len(query_result.historical_open_rows)}"
                )
            else:
                query_result = self.repo.load_current_shift_events(
                    building=building,
                    duty_date=duty_date,
                    duty_shift=duty_shift,
                    emit_log=emit_log,
                )
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][事件分类] 读取失败，分类留空: {exc}")
            return {}

        if not isinstance(query_result, EventSectionQueryResult):
            raise TypeError(f"load_current_shift_events returned unsupported type: {type(query_result)!r}")

        cfg = query_result.cfg
        if not bool(cfg.get("enabled", True)):
            return {}

        progress_cfg = cfg.get("progress_text", {}) if isinstance(cfg.get("progress_text", {}), dict) else {}
        done_text = str(progress_cfg.get("done", "已完成")).strip() or "已完成"
        todo_text = str(progress_cfg.get("todo", "未完成")).strip() or "未完成"
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        new_event_name = str(sections_cfg.get("new_event", "新事件处理")).strip() or "新事件处理"
        history_name = str(sections_cfg.get("history_followup", "历史事件跟进")).strip() or "历史事件跟进"
        cache_cfg = cfg.get("cache", {}) if isinstance(cfg.get("cache", {}), dict) else {}

        new_rows_payload: List[Dict[str, str]] = []
        history_rows_map: Dict[str, _HistoryRow] = {}
        pending_next_by_id: Dict[str, Dict[str, Any]] = {}
        queried_ids: List[str] = []
        outside_ongoing_count = 0
        historical_open_count = 0
        cache_completed_count = 0
        cache_from_snapshot_count = 0

        for row in query_result.current_rows:
            progress_text = self._get_new_event_progress_text(
                row=row,
                shift_end=query_result.shift_end,
                progress_cfg=progress_cfg,
            )
            payload = self._build_base_row_payload(
                row=row,
                progress_text=progress_text,
                work_window_text="/",
                follower_text=follower_text,
            )
            new_rows_payload.append(payload)
            queried_ids.append(row.record_id)
            if is_current_duty_context and progress_text == todo_text:
                pending_next_by_id[row.record_id] = self._make_pending_snapshot(row=row)

        if is_current_duty_context:
            for row in query_result.outside_shift_ongoing_rows:
                progress_text = self.repo.get_progress_text(row, progress_cfg)
                if progress_text != todo_text:
                    continue
                outside_ongoing_count += 1
                payload = self._build_base_row_payload(
                    row=row,
                    progress_text=todo_text,
                    work_window_text="/",
                    follower_text=follower_text,
                )
                queried_ids.append(row.record_id)
                pending_next_by_id[row.record_id] = self._make_pending_snapshot(row=row)
                history_rows_map[row.record_id] = _HistoryRow(
                    record_id=row.record_id,
                    payload=payload,
                    unresolved_snapshot=self._make_pending_snapshot(row=row),
                )

            outside_ongoing_ids = set(history_rows_map.keys())
            if bool(cache_cfg.get("enabled", True)):
                cache_store = self._get_cache_store(cfg)
                cached_rows = cache_store.list_pending_for_building(building)
                cache_before_count = len(cached_rows)
                for snap in cached_rows:
                    record_id = str(snap.get("record_id", "")).strip()
                    if not record_id:
                        continue
                    if record_id in outside_ongoing_ids:
                        continue
                    queried_ids.append(record_id)
                    resolved = None
                    try:
                        resolved = self.repo.get_record_by_id(record_id=record_id)
                    except Exception as exc:  # noqa: BLE001
                        emit_log(f"[交接班][事件分类] 回查record失败 record_id={record_id}: {exc}")
                    if resolved is None:
                        cache_from_snapshot_count += 1
                        resolved = self._row_from_snapshot(snap)

                    payload = self._build_base_row_payload(
                        row=resolved,
                        progress_text=done_text,
                        work_window_text="/",
                        follower_text=follower_text,
                    )
                    history_rows_map[record_id] = _HistoryRow(
                        record_id=record_id,
                        payload=payload,
                        unresolved_snapshot=None,
                    )
                    pending_next_by_id.pop(record_id, None)
                    cache_completed_count += 1

                try:
                    unique_ids = []
                    seen = set()
                    for rid in queried_ids:
                        if rid in seen:
                            continue
                        seen.add(rid)
                        unique_ids.append(rid)
                    cache_store.update_building_pending(
                        building=building,
                        pending_rows=list(pending_next_by_id.values()),
                        max_pending=int(cache_cfg.get("max_pending", 20000) or 20000),
                        last_query_record_ids=unique_ids,
                        max_last_query_ids=int(cache_cfg.get("max_last_query_ids", 5000) or 5000),
                    )
                    emit_log(
                        "[交接班][事件分类] 缓存更新: "
                        f"before={cache_before_count}, after={len(pending_next_by_id)}, removed={cache_completed_count}"
                    )
                except Exception as exc:  # noqa: BLE001
                    emit_log(f"[交接班][事件分类] 缓存更新失败: {exc}")
        else:
            emit_log(
                "[交接班][事件分类] 历史班次生成: 跳过缓存读写 "
                f"building={building}, duty={duty_date}/{duty_shift}"
            )
            for row in query_result.historical_open_rows:
                historical_open_count += 1
                history_rows_map[row.record_id] = _HistoryRow(
                    record_id=row.record_id,
                    payload=self._build_base_row_payload(
                        row=row,
                        progress_text=todo_text,
                        work_window_text="/",
                        follower_text=follower_text,
                    ),
                    unresolved_snapshot=None,
                )

        col_map_by_section = self._resolve_section_column_mapping(cfg=cfg, emit_log=emit_log)
        new_map = col_map_by_section.get(new_event_name, {})
        history_map = col_map_by_section.get(history_name, {})
        new_rows_cells = [{"cells": self._to_cells(row, new_map)} for row in new_rows_payload]
        history_rows_cells = [
            {"cells": self._to_cells(history.payload, history_map)}
            for history in history_rows_map.values()
        ]
        todo_count = sum(1 for row in new_rows_payload if str(row.get("progress", "")).strip() == todo_text)
        done_count = max(0, len(new_rows_payload) - todo_count)
        emit_log(
            f"[交接班][事件分类] 新事件构建: count={len(new_rows_cells)}, todo={todo_count}, done={done_count}"
        )
        if is_current_duty_context:
            emit_log(
                "[交接班][事件分类] 历史构建: "
                f"from_outside={outside_ongoing_count}, "
                f"from_cache_completed={cache_completed_count}, "
                f"from_cache_snapshot={cache_from_snapshot_count}, "
                f"history={len(history_rows_cells)}, pending={len(pending_next_by_id)}"
            )
        else:
            emit_log(
                "[交接班][事件分类] 历史构建(历史班次): "
                f"from_historical_open={historical_open_count}, history={len(history_rows_cells)}"
            )
        return {
            new_event_name: new_rows_cells,
            history_name: history_rows_cells,
        }
