from __future__ import annotations

import copy
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import openpyxl

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.bitable_target_resolver import BitableTargetResolver
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from handover_log_module.core.models import MetricHit


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(base or {})
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key] = _deep_merge(out[key], value)
        else:
            out[key] = copy.deepcopy(value)
    return out


def _extract_numbers(text: Any) -> List[float]:
    raw = str(text or "").strip()
    if not raw:
        return []
    matches = re.findall(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?", raw)
    values: List[float] = []
    for match in matches:
        candidate = match.replace(",", "")
        try:
            values.append(float(candidate))
        except Exception:  # noqa: BLE001
            continue
    return values


class DayMetricBitableExportService:
    def __init__(self, handover_cfg: Dict[str, Any]) -> None:
        self.handover_cfg = handover_cfg

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "enabled": True,
            "only_day_shift": True,
            "source": {
                "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
                "table_id": "tblAHGF8mV6U9jid",
                "base_url": "",
                "wiki_url": "",
                "page_size": 500,
                "max_records": 5000,
                "delete_batch_size": 200,
                "create_batch_size": 200,
            },
            "fields": {
                "type": "类型",
                "building": "楼栋",
                "date": "日期",
                "value": "数值",
                "position_code": "位置/编号",
            },
            "missing_value_policy": "zero",
            "types": [
                {"name": "总负荷（KW）", "source": "cell", "cell": "D6"},
                {"name": "IT总负荷（KW）", "source": "cell", "cell": "F6"},
                {"name": "室外湿球最高温度（℃）", "source": "cell", "cell": "D7"},
                {"name": "冷水系统供水最高温度（℃）", "source": "metric", "metric_id": "chilled_supply_temp_max"},
                {"name": "蓄水池后备最短时间（H）", "source": "cell", "cell": "D8"},
                {"name": "蓄冷罐后备最短时间（min）", "source": "cell_min_pair", "cell": "F8"},
                {"name": "供油可用时长（H）", "source": "cell", "cell": "H6"},
                {"name": "冷通道最高温度（℃）", "source": "metric", "metric_id": "cold_temp_max"},
                {"name": "冷通道最高湿度（%）", "source": "metric", "metric_id": "cold_humi_max"},
                {"name": "变压器负载率（MAX）", "source": "cell_percent", "cell": "B10"},
                {"name": "UPS负载率（MAX）", "source": "cell_percent", "cell": "D10"},
                {"name": "HVDC负载率（MAX）", "source": "metric", "metric_id": "hvdc_load_max"},
            ],
        }

    def _normalize_cfg(self) -> Dict[str, Any]:
        raw = self.handover_cfg.get("day_metric_export", {})
        cfg = _deep_merge(self._defaults(), raw if isinstance(raw, dict) else {})

        source = cfg.get("source", {})
        fields = cfg.get("fields", {})
        types = cfg.get("types", [])
        if not isinstance(source, dict):
            source = {}
        if not isinstance(fields, dict):
            fields = {}
        if not isinstance(types, list):
            types = []

        source["app_token"] = str(source.get("app_token", "")).strip()
        source["table_id"] = str(source.get("table_id", "")).strip()
        source["base_url"] = str(source.get("base_url", "")).strip()
        source["wiki_url"] = str(source.get("wiki_url", "")).strip()
        source["page_size"] = max(1, int(source.get("page_size", 500) or 500))
        source["max_records"] = max(0, int(source.get("max_records", 5000) or 5000))
        source["delete_batch_size"] = max(1, int(source.get("delete_batch_size", 200) or 200))
        source["create_batch_size"] = max(1, int(source.get("create_batch_size", 200) or 200))
        cfg["source"] = source

        fields["type"] = str(fields.get("type", "类型")).strip() or "类型"
        fields["building"] = str(fields.get("building", "楼栋")).strip() or "楼栋"
        fields["date"] = str(fields.get("date", "日期")).strip() or "日期"
        fields["value"] = str(fields.get("value", "数值")).strip() or "数值"
        fields["position_code"] = str(fields.get("position_code", "位置/编号")).strip() or "位置/编号"
        cfg["fields"] = fields

        normalized_types: List[Dict[str, Any]] = []
        for raw_item in types:
            if not isinstance(raw_item, dict):
                continue
            item = {
                "name": str(raw_item.get("name", "")).strip(),
                "source": str(raw_item.get("source", "cell")).strip().lower() or "cell",
                "cell": str(raw_item.get("cell", "")).strip().upper(),
                "metric_id": str(raw_item.get("metric_id", "")).strip(),
            }
            if item["name"]:
                normalized_types.append(item)

        cfg["types"] = normalized_types
        cfg["enabled"] = bool(cfg.get("enabled", True))
        cfg["only_day_shift"] = bool(cfg.get("only_day_shift", True))
        cfg["missing_value_policy"] = str(cfg.get("missing_value_policy", "zero")).strip().lower() or "zero"
        return cfg

    def _sheet_name(self) -> str:
        template_cfg = self.handover_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return ""
        return str(template_cfg.get("sheet_name", "")).strip()

    def _resolve_target(self, cfg: Dict[str, Any]) -> Dict[str, str]:
        global_feishu = self.handover_cfg.get("_global_feishu", {})
        if not isinstance(global_feishu, dict):
            global_feishu = {}

        app_id = str(global_feishu.get("app_id", "")).strip()
        app_secret = str(global_feishu.get("app_secret", "")).strip()
        if not app_id or not app_secret:
            raise ValueError("飞书配置缺失: common.feishu_auth.app_id/app_secret")

        source = cfg.get("source", {}) if isinstance(cfg.get("source", {}), dict) else {}
        return BitableTargetResolver(
            app_id=app_id,
            app_secret=app_secret,
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
        ).resolve(source)

    def _new_client(
        self,
        cfg: Dict[str, Any],
        *,
        resolved_target: Dict[str, str] | None = None,
    ) -> FeishuBitableClient:
        global_feishu = self.handover_cfg.get("_global_feishu", {})
        if not isinstance(global_feishu, dict):
            global_feishu = {}

        app_id = str(global_feishu.get("app_id", "")).strip()
        app_secret = str(global_feishu.get("app_secret", "")).strip()
        if not app_id or not app_secret:
            raise ValueError("飞书配置缺失: common.feishu_auth.app_id/app_secret")

        target = resolved_target if isinstance(resolved_target, dict) else self._resolve_target(cfg)
        return FeishuBitableClient(
            app_id=app_id,
            app_secret=app_secret,
            app_token=target["app_token"],
            calc_table_id=target["table_id"],
            attachment_table_id=target["table_id"],
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda x: str(x or "").strip(),
            dimension_mapping={},
        )

    @staticmethod
    def _midnight_timestamp_ms(duty_date: str) -> int:
        dt = datetime.strptime(str(duty_date).strip(), "%Y-%m-%d")
        return int(dt.timestamp() * 1000)

    @staticmethod
    def _normalize_text_field(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, dict):
            for key in ("text", "name", "value"):
                text = str(value.get(key, "")).strip()
                if text:
                    return text
            return ""
        if isinstance(value, list):
            for item in value:
                text = DayMetricBitableExportService._normalize_text_field(item)
                if text:
                    return text
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_date_field_to_midnight_ms(value: Any) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            number = int(value)
            if number <= 0:
                return None
            if number < 10**11:
                number *= 1000
            dt = datetime.fromtimestamp(number / 1000)
            midnight = datetime(dt.year, dt.month, dt.day, 0, 0, 0)
            return int(midnight.timestamp() * 1000)
        text = str(value).strip()
        if not text:
            return None
        candidate = text[:10].replace("/", "-")
        try:
            dt = datetime.strptime(candidate, "%Y-%m-%d")
        except ValueError:
            return None
        return int(dt.timestamp() * 1000)

    @staticmethod
    def _normalize_cell_values(values: Dict[str, Any] | None) -> Dict[str, Any]:
        if not isinstance(values, dict):
            return {}
        out: Dict[str, Any] = {}
        for raw_key, raw_val in values.items():
            key = str(raw_key or "").strip().upper()
            if key:
                out[key] = raw_val
        return out

    def build_metric_context(self, resolved_values_by_id: Dict[str, Any] | None) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        out: Dict[str, Any] = {}
        for item in cfg.get("types", []):
            if not isinstance(item, dict):
                continue
            if str(item.get("source", "")).strip().lower() != "metric":
                continue
            metric_id = str(item.get("metric_id", "")).strip()
            if not metric_id:
                continue
            if metric_id in resolved:
                out[metric_id] = resolved.get(metric_id)
        return out

    @staticmethod
    def _serialize_hit_payload(metric_key: str, hit: MetricHit | Dict[str, Any] | Any) -> Dict[str, Any]:
        if isinstance(hit, dict):
            row_index = int(hit.get("row_index", 0) or 0)
            b_norm = str(hit.get("b_norm", "")).strip()
            c_norm = str(hit.get("c_norm", "")).strip()
            b_text = str(hit.get("b_text", "")).strip()
            c_text = str(hit.get("c_text", "")).strip()
        else:
            row_index = int(getattr(hit, "row_index", 0) or 0)
            b_norm = str(getattr(hit, "b_norm", "")).strip()
            c_norm = str(getattr(hit, "c_norm", "")).strip()
            b_text = str(getattr(hit, "b_text", "")).strip()
            c_text = str(getattr(hit, "c_text", "")).strip()
        return {
            "metric_key": str(metric_key or "").strip(),
            "row_index": row_index,
            "b_norm": b_norm,
            "c_norm": c_norm,
            "b_text": b_text,
            "c_text": c_text,
        }

    @classmethod
    def _normalize_origin_payload(cls, payload: Dict[str, Any] | None) -> Dict[str, Any]:
        data = payload if isinstance(payload, dict) else {}
        return {
            "metric_key": str(data.get("metric_key", "")).strip(),
            "row_index": int(data.get("row_index", 0) or 0),
            "b_norm": str(data.get("b_norm", "")).strip(),
            "c_norm": str(data.get("c_norm", "")).strip(),
            "b_text": str(data.get("b_text", "")).strip(),
            "c_text": str(data.get("c_text", "")).strip(),
        }

    @classmethod
    def _normalize_metric_origin_context(cls, metric_origin_context: Dict[str, Any] | None) -> Dict[str, Any]:
        raw = metric_origin_context if isinstance(metric_origin_context, dict) else {}
        by_metric_id_raw = raw.get("by_metric_id", {})
        by_target_cell_raw = raw.get("by_target_cell", {})
        by_metric_id: Dict[str, Dict[str, Any]] = {}
        by_target_cell: Dict[str, Dict[str, Any]] = {}
        if isinstance(by_metric_id_raw, dict):
            for metric_id, payload in by_metric_id_raw.items():
                metric_key = str(metric_id or "").strip()
                if metric_key:
                    by_metric_id[metric_key] = cls._normalize_origin_payload(payload if isinstance(payload, dict) else {})
        if isinstance(by_target_cell_raw, dict):
            for cell_name, payload in by_target_cell_raw.items():
                target_cell = str(cell_name or "").strip().upper()
                if target_cell:
                    by_target_cell[target_cell] = cls._normalize_origin_payload(payload if isinstance(payload, dict) else {})
        return {
            "by_metric_id": by_metric_id,
            "by_target_cell": by_target_cell,
        }

    def serialize_metric_origin_context(
        self,
        *,
        hits: Dict[str, MetricHit] | Dict[str, Any] | None,
        effective_config: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        by_metric_id: Dict[str, Dict[str, Any]] = {}
        hits_map = hits if isinstance(hits, dict) else {}
        for metric_key, hit in hits_map.items():
            metric_id = str(metric_key or "").strip()
            if not metric_id:
                continue
            by_metric_id[metric_id] = self._serialize_hit_payload(metric_id, hit)

        by_target_cell: Dict[str, Dict[str, Any]] = {}
        effective = effective_config if isinstance(effective_config, dict) else {}
        cell_mapping = effective.get("cell_mapping", {})
        if isinstance(cell_mapping, dict):
            for metric_key, target_cell in cell_mapping.items():
                metric_id = str(metric_key or "").strip()
                cell_name = str(target_cell or "").strip().upper()
                if not metric_id or not cell_name:
                    continue
                payload = by_metric_id.get(metric_id)
                if not isinstance(payload, dict):
                    continue
                by_target_cell[cell_name] = {
                    **payload,
                    "metric_key": metric_id,
                }

        return self._normalize_metric_origin_context(
            {
                "by_metric_id": by_metric_id,
                "by_target_cell": by_target_cell,
            }
        )

    @staticmethod
    def _compose_position_code(origin_payload: Dict[str, Any] | None) -> str:
        payload = origin_payload if isinstance(origin_payload, dict) else {}
        b_norm = str(payload.get("b_norm", "")).strip()
        c_norm = str(payload.get("c_norm", "")).strip()
        c_text = str(payload.get("c_text", "")).strip()
        if b_norm and c_norm:
            return f"{b_norm} {c_norm}"
        if c_norm:
            return c_norm
        if c_text:
            return c_text
        if b_norm:
            return b_norm
        return ""

    def _resolve_origin_payload(
        self,
        *,
        type_item: Dict[str, Any],
        metric_origin_context: Dict[str, Any],
    ) -> Dict[str, Any]:
        source_type = str(type_item.get("source", "")).strip().lower()
        if source_type == "metric":
            metric_id = str(type_item.get("metric_id", "")).strip()
            payload = metric_origin_context.get("by_metric_id", {}).get(metric_id, {})
            return payload if isinstance(payload, dict) else {}
        if source_type in {"cell", "cell_percent", "cell_min_pair"}:
            cell_name = str(type_item.get("cell", "")).strip().upper()
            payload = metric_origin_context.get("by_target_cell", {}).get(cell_name, {})
            return payload if isinstance(payload, dict) else {}
        return {}

    def build_deferred_state(
        self,
        *,
        duty_shift: str,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        if not cfg.get("enabled", True):
            return {
                "status": "skipped",
                "reason": "disabled",
                "uploaded_count": 0,
                "error": "",
                "uploaded_at": "",
                "uploaded_revision": 0,
                "metric_values_by_id": {},
                "metric_origin_context": {"by_metric_id": {}, "by_target_cell": {}},
            }
        return {
            "status": "pending_review",
            "reason": "await_all_confirmed",
            "uploaded_count": 0,
            "error": "",
            "uploaded_at": "",
            "uploaded_revision": 0,
            "metric_values_by_id": self.build_metric_context(resolved_values_by_id),
            "metric_origin_context": self._normalize_metric_origin_context(metric_origin_context),
        }

    def _load_workbook_cell_values(self, output_file: str, cfg: Dict[str, Any]) -> Dict[str, Any]:
        path = Path(str(output_file or "").strip())
        if not path.exists():
            raise FileNotFoundError(f"交接班成品文件不存在: {path}")

        needed_cells: List[str] = []
        for item in cfg.get("types", []):
            if not isinstance(item, dict):
                continue
            source_type = str(item.get("source", "")).strip().lower()
            if source_type == "metric":
                continue
            cell_name = str(item.get("cell", "")).strip().upper()
            if cell_name and cell_name not in needed_cells:
                needed_cells.append(cell_name)

        workbook = openpyxl.load_workbook(path, data_only=True)
        try:
            sheet_name = self._sheet_name()
            ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            values: Dict[str, Any] = {}
            for cell_name in needed_cells:
                values[cell_name] = ws[cell_name].value
            return self._normalize_cell_values(values)
        finally:
            workbook.close()

    @staticmethod
    def _resolve_numeric_from_type(
        *,
        type_item: Dict[str, Any],
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
    ) -> Optional[float]:
        source_type = str(type_item.get("source", "")).strip().lower()
        if source_type == "metric":
            metric_id = str(type_item.get("metric_id", "")).strip()
            raw = resolved_values_by_id.get(metric_id)
            numbers = _extract_numbers(raw)
            return numbers[0] if numbers else None

        cell_name = str(type_item.get("cell", "")).strip().upper()
        raw_cell = cell_values.get(cell_name, "")
        numbers = _extract_numbers(raw_cell)
        if not numbers:
            return None
        if source_type == "cell_min_pair":
            return min(numbers)
        return numbers[0]

    def _prepare_records(
        self,
        *,
        cfg: Dict[str, Any],
        building: str,
        duty_date: str,
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
        metric_origin_context: Dict[str, Any] | None = None,
    ) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        fields = cfg.get("fields", {})
        missing_policy = str(cfg.get("missing_value_policy", "zero")).strip().lower()
        duty_ms = self._midnight_timestamp_ms(duty_date)
        origin_context = self._normalize_metric_origin_context(metric_origin_context)
        records: List[Dict[str, Any]] = []
        preview: List[Dict[str, Any]] = []

        for item in cfg.get("types", []):
            if not isinstance(item, dict):
                continue

            numeric = self._resolve_numeric_from_type(
                type_item=item,
                cell_values=cell_values,
                resolved_values_by_id=resolved_values_by_id,
            )
            if numeric is None and missing_policy == "zero":
                numeric = 0.0
            if numeric is None:
                numeric = 0.0
            position_code_text = self._compose_position_code(
                self._resolve_origin_payload(
                    type_item=item,
                    metric_origin_context=origin_context,
                )
            )

            row_fields = {
                str(fields.get("type", "类型")).strip(): str(item.get("name", "")).strip(),
                str(fields.get("building", "楼栋")).strip(): str(building or "").strip(),
                str(fields.get("date", "日期")).strip(): duty_ms,
                str(fields.get("value", "数值")).strip(): numeric,
                str(fields.get("position_code", "位置/编号")).strip(): position_code_text,
            }
            records.append(row_fields)
            if len(preview) < 3:
                preview.append(row_fields)
        return records, preview

    def _matching_existing_records(
        self,
        *,
        existing_records: List[Dict[str, Any]],
        building: str,
        duty_date: str,
        cfg: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        fields = cfg.get("fields", {})
        type_field = str(fields.get("type", "类型")).strip()
        building_field = str(fields.get("building", "楼栋")).strip()
        date_field = str(fields.get("date", "日期")).strip()
        target_building = str(building or "").strip()
        target_date_ms = self._midnight_timestamp_ms(duty_date)
        allowed_types = {
            str(item.get("name", "")).strip()
            for item in cfg.get("types", [])
            if isinstance(item, dict) and str(item.get("name", "")).strip()
        }
        matched: List[Dict[str, Any]] = []
        for item in existing_records:
            if not isinstance(item, dict):
                continue
            payload_fields = item.get("fields", {})
            if not isinstance(payload_fields, dict):
                continue
            item_building = self._normalize_text_field(payload_fields.get(building_field))
            item_type = self._normalize_text_field(payload_fields.get(type_field))
            item_date_ms = self._normalize_date_field_to_midnight_ms(payload_fields.get(date_field))
            if item_building != target_building:
                continue
            if item_date_ms != target_date_ms:
                continue
            if item_type not in allowed_types:
                continue
            matched.append(copy.deepcopy(item))
        return matched

    def list_existing_records_for_unit(self, *, building: str, duty_date: str) -> List[Dict[str, Any]]:
        cfg = self._normalize_cfg()
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        client = self._new_client(cfg, resolved_target=resolved_target)
        table_id = str(resolved_target.get("table_id", "")).strip()
        existing_records = client.list_records(
            table_id=table_id,
            page_size=int(source.get("page_size", 500) or 500),
            max_records=int(source.get("max_records", 5000) or 5000),
        )
        return self._matching_existing_records(
            existing_records=existing_records,
            building=building,
            duty_date=duty_date,
            cfg=cfg,
        )

    def delete_existing_records_for_unit(
        self,
        *,
        building: str,
        duty_date: str,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        client = self._new_client(cfg, resolved_target=resolved_target)
        table_id = str(resolved_target.get("table_id", "")).strip()
        existing_records = client.list_records(
            table_id=table_id,
            page_size=int(source.get("page_size", 500) or 500),
            max_records=int(source.get("max_records", 5000) or 5000),
        )
        matched = self._matching_existing_records(
            existing_records=existing_records,
            building=building,
            duty_date=duty_date,
            cfg=cfg,
        )
        record_ids = [
            str(item.get("record_id", "")).strip()
            for item in matched
            if str(item.get("record_id", "")).strip()
        ]
        deleted_count = 0
        if record_ids:
            deleted_count = client.batch_delete_records(
                table_id=table_id,
                record_ids=record_ids,
                batch_size=int(source.get("delete_batch_size", 200) or 200),
            )
        emit_log(f"[交接班][白班多维] 删除旧记录 building={building}, duty_date={duty_date}, count={deleted_count}")
        return {
            "status": "ok",
            "deleted_count": deleted_count,
            "matched_records": matched,
        }

    def _run_with_values(
        self,
        *,
        cfg: Dict[str, Any],
        building: str,
        duty_date: str,
        duty_shift: str,
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
        metric_origin_context: Dict[str, Any] | None,
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        if not cfg.get("enabled", True):
            return {"status": "skipped", "reason": "disabled", "uploaded_count": 0}
        if cfg.get("only_day_shift", True) and str(duty_shift or "").strip().lower() != "day":
            emit_log("[交接班][白班多维] 跳过: 非白班")
            return {"status": "skipped", "reason": "non_day_shift", "uploaded_count": 0}

        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        table_id = str(resolved_target.get("table_id", "")).strip()
        batch_size = int(source.get("create_batch_size", 200) or 200)
        records, preview = self._prepare_records(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            cell_values=cell_values,
            resolved_values_by_id=resolved_values_by_id,
            metric_origin_context=metric_origin_context,
        )

        try:
            client = self._new_client(cfg, resolved_target=resolved_target)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][白班多维] 初始化失败: {exc}")
            return {"status": "failed", "uploaded_count": 0, "error": str(exc)}

        emit_log(f"[交接班][白班多维] 开始上传 building={building}, duty_date={duty_date}, records={len(records)}")
        try:
            client.batch_create_records(table_id=table_id, fields_list=records, batch_size=batch_size)
            emit_log(f"[交接班][白班多维] 上传完成 uploaded={len(records)}")
            return {
                "status": "ok",
                "uploaded_count": len(records),
                "records_preview": preview,
                "error": "",
            }
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][白班多维] 上传失败 error={exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "records_preview": preview,
                "error": str(exc),
            }

    def run(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        filled_cell_values: Dict[str, Any] | None,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        cell_values = self._normalize_cell_values(filled_cell_values)
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        return self._run_with_values(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            cell_values=cell_values,
            resolved_values_by_id=resolved,
            metric_origin_context=metric_origin_context,
            emit_log=emit_log,
        )

    def rewrite_from_output_file(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        output_file: str,
        metric_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        try:
            cell_values = self._load_workbook_cell_values(output_file, cfg)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][白班多维] 读取成品失败: {exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
                "error": str(exc),
            }

        if not cfg.get("enabled", True):
            return {
                "status": "skipped",
                "reason": "disabled",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
            }
        if cfg.get("only_day_shift", True) and str(duty_shift or "").strip().lower() != "day":
            emit_log("[交接班][白班多维] 跳过: 非白班")
            return {
                "status": "skipped",
                "reason": "non_day_shift",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
            }

        resolved = metric_values_by_id if isinstance(metric_values_by_id, dict) else {}
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        table_id = str(resolved_target.get("table_id", "")).strip()
        batch_size = int(source.get("create_batch_size", 200) or 200)
        records, preview = self._prepare_records(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            cell_values=cell_values,
            resolved_values_by_id=resolved,
            metric_origin_context=metric_origin_context,
        )

        try:
            client = self._new_client(cfg, resolved_target=resolved_target)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][白班多维] 初始化失败: {exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
                "error": str(exc),
            }

        emit_log(f"[交接班][白班多维] 开始重写 building={building}, duty_date={duty_date}, records={len(records)}")
        try:
            existing_records = client.list_records(
                table_id=table_id,
                page_size=int(source.get("page_size", 500) or 500),
                max_records=int(source.get("max_records", 5000) or 5000),
            )
            matched = self._matching_existing_records(
                existing_records=existing_records,
                building=building,
                duty_date=duty_date,
                cfg=cfg,
            )
            record_ids = [
                str(item.get("record_id", "")).strip()
                for item in matched
                if str(item.get("record_id", "")).strip()
            ]
            deleted_count = 0
            if record_ids:
                deleted_count = client.batch_delete_records(
                    table_id=table_id,
                    record_ids=record_ids,
                    batch_size=int(source.get("delete_batch_size", 200) or 200),
                )
            emit_log(
                f"[交接班][白班多维] 删除旧记录完成 building={building}, duty_date={duty_date}, count={deleted_count}"
            )
            client.batch_create_records(table_id=table_id, fields_list=records, batch_size=batch_size)
            emit_log(
                f"[交接班][白班多维] 重写完成 building={building}, duty_date={duty_date}, created={len(records)}"
            )
            return {
                "status": "ok",
                "uploaded_count": len(records),
                "created_records": len(records),
                "deleted_records": deleted_count,
                "records_preview": preview,
                "error": "",
            }
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][白班多维] 重写失败 error={exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
                "records_preview": preview,
                "error": str(exc),
            }

    def run_from_output_file(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        output_file: str,
        metric_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        return self.rewrite_from_output_file(
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            output_file=output_file,
            metric_values_by_id=metric_values_by_id,
            metric_origin_context=metric_origin_context,
            emit_log=emit_log,
        )
