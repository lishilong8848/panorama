from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from handover_log_module.core.footer_layout import find_footer_inventory_layout


@dataclass
class CategorySection:
    name: str
    title_row: int
    header_row: int
    template_data_row: int
    start_row: int
    end_row: int
    baseline_rows: int


@dataclass
class SectionCellSnapshot:
    style: Any
    value: Any


@dataclass
class SectionRowSnapshot:
    row_idx: int
    row_height: float | None
    cells: Dict[int, SectionCellSnapshot]
    merges: List[Tuple[int, int]]


@dataclass
class CategorySectionSnapshot:
    title_row: int
    name: str
    title: SectionRowSnapshot
    header: SectionRowSnapshot
    template: SectionRowSnapshot


@dataclass
class SectionLogicalColumn:
    key: str
    label: str
    source_cols: List[str]
    span: int


def _cell_text(ws: Worksheet, row: int, col: int = 1) -> str:
    value = ws.cell(row=row, column=col).value
    return str(value or "").strip()


def _row_merge_spans(ws: Worksheet, row_idx: int) -> List[Tuple[int, int]]:
    spans: List[Tuple[int, int]] = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == row_idx and merged.max_row == row_idx:
            spans.append((merged.min_col, merged.max_col))
    return spans


def _capture_row_snapshot(ws: Worksheet, row_idx: int, *, min_col: int = 1, max_col: int = 9) -> SectionRowSnapshot:
    cells: Dict[int, SectionCellSnapshot] = {}
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cells[col_idx] = SectionCellSnapshot(style=copy(cell._style), value=cell.value)

    return SectionRowSnapshot(
        row_idx=row_idx,
        row_height=ws.row_dimensions[row_idx].height,
        cells=cells,
        merges=_row_merge_spans(ws, row_idx),
    )


def parse_category_sections(
    ws: Worksheet,
    *,
    stop_title: str = "交接确认",
) -> List[CategorySection]:
    max_row = int(ws.max_row or 0)
    title_rows: List[int] = []

    for row in range(1, max_row):
        current = _cell_text(ws, row, 1)
        if not current:
            continue
        if _cell_text(ws, row + 1, 1) == "序号":
            title_rows.append(row)

    if not title_rows:
        return []

    stop_row = max_row + 1
    footer_layout = find_footer_inventory_layout(ws)
    if footer_layout is not None and footer_layout.title_row > title_rows[-1]:
        stop_row = footer_layout.title_row
    else:
        for row in range(title_rows[-1] + 1, max_row + 1):
            if _cell_text(ws, row, 1) == stop_title:
                stop_row = row
                break

    sections: List[CategorySection] = []
    for idx, title_row in enumerate(title_rows):
        name = _cell_text(ws, title_row, 1)
        next_title_row = title_rows[idx + 1] if idx + 1 < len(title_rows) else stop_row
        header_row = title_row + 1
        template_data_row = title_row + 2
        end_row = max(template_data_row, next_title_row - 1)
        baseline_rows = max(1, end_row - template_data_row + 1)
        sections.append(
            CategorySection(
                name=name,
                title_row=title_row,
                header_row=header_row,
                template_data_row=template_data_row,
                start_row=template_data_row,
                end_row=end_row,
                baseline_rows=baseline_rows,
            )
        )

    return sections


def capture_section_snapshots(ws: Worksheet, sections: List[CategorySection]) -> Dict[int, CategorySectionSnapshot]:
    snapshots: Dict[int, CategorySectionSnapshot] = {}
    for section in sections:
        snapshots[section.title_row] = CategorySectionSnapshot(
            title_row=section.title_row,
            name=section.name,
            title=_capture_row_snapshot(ws, section.title_row),
            header=_capture_row_snapshot(ws, section.header_row),
            template=_capture_row_snapshot(ws, section.template_data_row),
        )
    return snapshots


def build_section_logical_columns(
    ws: Worksheet,
    section: CategorySection,
    *,
    hidden_columns: List[str] | None = None,
    min_col: int = 2,
    max_col: int = 9,
) -> List[SectionLogicalColumn]:
    hidden = {str(value or "").strip().upper() for value in (hidden_columns or []) if str(value or "").strip()}
    merge_spans = _row_merge_spans(ws, section.template_data_row)
    follower_cols: set[int] = set()
    span_by_lead: Dict[int, Tuple[int, int]] = {}

    for min_span_col, max_span_col in merge_spans:
        span_by_lead[min_span_col] = (min_span_col, max_span_col)
        for col_idx in range(min_span_col + 1, max_span_col + 1):
            follower_cols.add(col_idx)

    columns: List[SectionLogicalColumn] = []
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx).upper()
        if col_letter in hidden:
            continue
        if col_idx in follower_cols:
            continue

        min_span_col, max_span_col = span_by_lead.get(col_idx, (col_idx, col_idx))
        source_cols = [get_column_letter(span_col).upper() for span_col in range(min_span_col, max_span_col + 1)]
        header_value = ws.cell(row=section.header_row, column=col_idx).value
        label = "" if header_value is None else str(header_value)
        columns.append(
            SectionLogicalColumn(
                key=col_letter,
                label=label,
                source_cols=source_cols,
                span=len(source_cols),
            )
        )

    if columns:
        return columns

    fallback_columns: List[SectionLogicalColumn] = []
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx).upper()
        if col_letter in hidden:
            continue
        header_value = ws.cell(row=section.header_row, column=col_idx).value
        label = "" if header_value is None else str(header_value)
        fallback_columns.append(
            SectionLogicalColumn(
                key=col_letter,
                label=label,
                source_cols=[col_letter],
                span=1,
            )
        )
    return fallback_columns
