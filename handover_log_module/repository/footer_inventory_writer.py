from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Tuple

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from handover_log_module.core.footer_layout import (
    FOOTER_GROUP_TITLE_TEXT,
    FOOTER_INVENTORY_COLUMNS,
    FooterInventoryLayout,
    find_footer_inventory_layout,
    trim_rows_below_footer,
)
from handover_log_module.repository.section_writer import _insert_rows_like_excel


@dataclass
class FooterRowSnapshot:
    row_idx: int
    row_height: float | None
    cells: Dict[int, Tuple[Any, Any]]
    merges: List[Tuple[int, int, int, int]]


def _capture_row_snapshot(ws: Worksheet, row_idx: int, *, min_col: int = 1, max_col: int = 9) -> FooterRowSnapshot:
    cells: Dict[int, Tuple[Any, Any]] = {}
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cells[col_idx] = (copy(cell._style), copy(cell.value))

    merges: List[Tuple[int, int, int, int]] = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == row_idx and merged.max_row == row_idx:
            merges.append((merged.min_col, merged.min_row, merged.max_col, merged.max_row))

    return FooterRowSnapshot(
        row_idx=row_idx,
        row_height=ws.row_dimensions[row_idx].height,
        cells=cells,
        merges=merges,
    )


def _intersects_row_range(merged, start_row: int, end_row: int) -> bool:
    return not (merged.max_row < start_row or merged.min_row > end_row)


def _clear_merges_in_row_range(ws: Worksheet, start_row: int, end_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if _intersects_row_range(merged, start_row, end_row):
            try:
                ws.unmerge_cells(str(merged))
            except Exception:  # noqa: BLE001
                try:
                    ws.merged_cells.ranges.remove(merged)
                except Exception:  # noqa: BLE001
                    pass


def _restore_row_snapshot(
    ws: Worksheet,
    row_idx: int,
    snapshot: FooterRowSnapshot,
    *,
    restore_values: bool,
) -> None:
    for col_idx, (style, value) in snapshot.cells.items():
        existing = ws._cells.get((row_idx, col_idx))  # noqa: SLF001
        if isinstance(existing, MergedCell):
            del ws._cells[(row_idx, col_idx)]  # noqa: SLF001
        cell = ws.cell(row=row_idx, column=col_idx)
        cell._style = copy(style)
        if restore_values:
            cell.value = copy(value)
    ws.row_dimensions[row_idx].height = snapshot.row_height


def _apply_row_merges(ws: Worksheet, row_idx: int, merges: List[Tuple[int, int, int, int]]) -> None:
    for min_col, _, max_col, _ in merges:
        ws.merge_cells(
            start_row=row_idx,
            start_column=min_col,
            end_row=row_idx,
            end_column=max_col,
        )


def _normalize_inventory_columns(columns: Any) -> List[Dict[str, Any]]:
    if isinstance(columns, list) and columns:
        normalized: List[Dict[str, Any]] = []
        for column in columns:
            if not isinstance(column, dict):
                continue
            key = str(column.get("key", "")).strip().upper()
            if not key:
                continue
            source_cols = column.get("source_cols", [])
            if not isinstance(source_cols, list) or not source_cols:
                source_cols = [key]
            normalized.append(
                {
                    "key": key,
                    "label": str(column.get("label", "") or key),
                    "source_cols": [str(item or "").strip().upper() for item in source_cols if str(item or "").strip()],
                    "span": int(column.get("span", len(source_cols)) or len(source_cols)),
                }
            )
        if normalized:
            return normalized
    return [dict(item) for item in FOOTER_INVENTORY_COLUMNS]


def _blank_inventory_row(columns: List[Dict[str, Any]]) -> Dict[str, str]:
    return {str(column["key"]).upper(): "" for column in columns}


def _normalize_inventory_rows(rows: Any, columns: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    normalized: List[Dict[str, str]] = []
    if not isinstance(rows, list):
        rows = []
    for row in rows:
        cells = row.get("cells", {}) if isinstance(row, dict) else {}
        if not isinstance(cells, dict):
            cells = {}
        normalized_row = _blank_inventory_row(columns)
        for column in columns:
            key = str(column["key"]).upper()
            normalized_row[key] = str(cells.get(key, "") or "")
        normalized.append(normalized_row)
    if normalized:
        return normalized
    return [_blank_inventory_row(columns)]


def _set_inventory_row_values(ws: Worksheet, row_idx: int, row_payload: Dict[str, str], columns: List[Dict[str, Any]]) -> None:
    for col_idx in range(2, 10):
        existing = ws._cells.get((row_idx, col_idx))  # noqa: SLF001
        if isinstance(existing, MergedCell):
            del ws._cells[(row_idx, col_idx)]  # noqa: SLF001
        cell = ws.cell(row=row_idx, column=col_idx)
        if not isinstance(cell, MergedCell):
            cell.value = ""
    for column in columns:
        key = str(column["key"]).upper()
        source_cols = column.get("source_cols", [])
        if not isinstance(source_cols, list) or not source_cols:
            continue
        lead_col = str(source_cols[0]).upper()
        value = str(row_payload.get(key, "") or "")
        existing_lead = ws._cells.get((row_idx, ws[f"{lead_col}1"].column))  # noqa: SLF001
        if isinstance(existing_lead, MergedCell):
            del ws._cells[(row_idx, ws[f"{lead_col}1"].column)]  # noqa: SLF001
        lead_cell = ws[f"{lead_col}{row_idx}"]
        if not isinstance(lead_cell, MergedCell):
            lead_cell.value = value
        for follower_col in source_cols[1:]:
            follower_col_text = str(follower_col).upper()
            existing_follower = ws._cells.get((row_idx, ws[f"{follower_col_text}1"].column))  # noqa: SLF001
            if isinstance(existing_follower, MergedCell):
                del ws._cells[(row_idx, ws[f"{follower_col_text}1"].column)]  # noqa: SLF001
            follower_cell = ws[f"{follower_col_text}{row_idx}"]
            if not isinstance(follower_cell, MergedCell):
                follower_cell.value = ""


def write_footer_inventory_table(
    *,
    ws: Worksheet,
    inventory_block: Dict[str, Any] | None,
    emit_log: Callable[[str], None] = print,
) -> None:
    if not isinstance(inventory_block, dict):
        return

    layout = find_footer_inventory_layout(ws)
    if layout is None:
        emit_log("[交接班][审核页][工具表写回] 跳过: 未找到交接确认区域")
        return

    title_snapshot = _capture_row_snapshot(ws, layout.title_row)
    header_snapshot = _capture_row_snapshot(ws, layout.header_row)
    template_snapshot = _capture_row_snapshot(ws, layout.data_start_row)
    signoff_snapshots: List[FooterRowSnapshot] = []
    if layout.signoff_start_row:
        for row_idx in range(layout.signoff_start_row, layout.last_row + 1):
            signoff_snapshots.append(_capture_row_snapshot(ws, row_idx))

    columns = _normalize_inventory_columns(inventory_block.get("columns", []))
    rows = _normalize_inventory_rows(inventory_block.get("rows", []), columns)
    current_rows = max(1, layout.data_end_row - layout.data_start_row + 1)
    target_rows = max(1, len(rows))

    if target_rows > current_rows:
        delta = target_rows - current_rows
        _insert_rows_like_excel(
            ws,
            insert_at=layout.data_end_row + 1,
            amount=delta,
            template_row=layout.data_start_row,
            emit_log=lambda message: emit_log(f"[交接班][审核页][工具表写回] {message}"),
        )
    elif target_rows < current_rows:
        ws.delete_rows(layout.data_start_row + target_rows, amount=current_rows - target_rows)

    layout = find_footer_inventory_layout(ws)
    if layout is None:
        emit_log("[交接班][审核页][工具表写回] 跳过: 结构调整后未找到交接确认区域")
        return

    new_data_end = layout.data_start_row + target_rows - 1
    signoff_start = new_data_end + 1 if signoff_snapshots else None
    signoff_end = signoff_start + len(signoff_snapshots) - 1 if signoff_start else new_data_end
    affected_end = max(layout.last_row, signoff_end)

    _clear_merges_in_row_range(ws, layout.title_row, affected_end)

    _restore_row_snapshot(ws, layout.title_row, title_snapshot, restore_values=True)
    _restore_row_snapshot(ws, layout.header_row, header_snapshot, restore_values=True)

    for row_idx in range(layout.data_start_row, new_data_end + 1):
        _restore_row_snapshot(ws, row_idx, template_snapshot, restore_values=False)

    if signoff_start is not None:
        for offset, snapshot in enumerate(signoff_snapshots):
            _restore_row_snapshot(ws, signoff_start + offset, snapshot, restore_values=True)

    for offset, row_payload in enumerate(rows):
        row_idx = layout.data_start_row + offset
        _set_inventory_row_values(ws, row_idx, row_payload, columns)

    ws.cell(row=layout.header_row, column=1).value = FOOTER_GROUP_TITLE_TEXT

    _apply_row_merges(ws, layout.title_row, title_snapshot.merges)
    ws.merge_cells(
        start_row=layout.header_row,
        start_column=1,
        end_row=new_data_end,
        end_column=1,
    )
    _apply_row_merges(ws, layout.header_row, header_snapshot.merges)
    for row_idx in range(layout.data_start_row, new_data_end + 1):
        _apply_row_merges(ws, row_idx, template_snapshot.merges)
    if signoff_start is not None:
        for offset, snapshot in enumerate(signoff_snapshots):
            _apply_row_merges(ws, signoff_start + offset, snapshot.merges)

    final_layout = find_footer_inventory_layout(ws)
    deleted_count = trim_rows_below_footer(ws, final_layout) if final_layout is not None else 0

    emit_log(
        "[交接班][审核页][工具表写回] "
        f"title_row={layout.title_row}, header_row={layout.header_row}, current_rows={current_rows}, "
        f"target_rows={target_rows}, signoff_start_row={signoff_start or '-'}, tail_trimmed={deleted_count}"
    )
