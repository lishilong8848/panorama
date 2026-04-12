from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict

import openpyxl

from app.shared.utils.atomic_file import atomic_write_file, validate_excel_workbook_file
from app.shared.utils.artifact_naming import (
    OUTPUT_TYPE_HANDOVER_LOG,
    build_output_base_path,
    with_index,
)
from app.shared.utils.file_utils import fallback_missing_windows_drive_path
from handover_log_module.core.footer_layout import find_footer_inventory_layout, trim_rows_below_footer
from handover_log_module.repository.excel_reader import load_workbook_quietly
from handover_log_module.core.footer_snapshot import (
    capture_footer_block_snapshot,
    restore_footer_block_snapshot,
)
from handover_log_module.core.section_layout import capture_section_snapshots, parse_category_sections
from handover_log_module.repository.section_writer import (
    resolve_section_target_row_count,
    write_category_sections,
)
from pipeline_utils import get_app_dir


def _safe_filename(building: str) -> str:
    text = str(building).strip()
    if not text:
        return "unknown_building"
    for ch in "\\/:*?\"<>|":
        text = text.replace(ch, "_")
    return text


def _resolve_template_path(raw_path: str) -> Path:
    path = Path(str(raw_path).strip())
    if path.is_absolute():
        return path
    project_root = Path(__file__).resolve().parents[2]
    return project_root / path


def build_output_filename(
    building: str,
    file_name_pattern: str,
    date_format: str,
    date_ref: datetime,
) -> str:
    building_text = _safe_filename(building)
    date_text = date_ref.strftime(date_format)
    return file_name_pattern.format(building=building_text, date=date_text)


def _is_file_in_use(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError):
        return True
    if isinstance(exc, OSError) and getattr(exc, "winerror", None) == 32:
        return True
    return False


def _normalize_payload_rows(value: Any) -> list[Any]:
    if not isinstance(value, list):
        return []
    return value


def _calculate_footer_row_shift(
    *,
    sections,
    category_payloads: Dict[str, Any] | None,
    empty_section_mode: str = "single_blank_row",
) -> int:
    payload_map = category_payloads if isinstance(category_payloads, dict) else {}
    total_delta = 0
    for section in sections:
        payload_count = len(_normalize_payload_rows(payload_map.get(section.name, [])))
        current_rows = max(section.end_row - section.template_data_row + 1, 1)
        target_rows = resolve_section_target_row_count(
            payload_count,
            empty_section_mode=empty_section_mode,
        )
        total_delta += target_rows - current_rows
    return total_delta


def _write_workbook(
    *,
    source_path: Path,
    out_path: Path,
    sheet_name: str,
    cell_values: Dict[str, str],
    category_payloads: Dict[str, Any] | None = None,
    emit_log: Callable[[str], None] = print,
) -> None:
    def _writer(temp_path: Path) -> None:
        shutil.copy2(source_path, temp_path)
        wb = load_workbook_quietly(temp_path)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"template sheet not found: {sheet_name}")

            ws = wb[sheet_name]
            for cell, text in cell_values.items():
                if not str(cell).strip():
                    continue
                if text is None:
                    continue
                value_text = str(text).strip()
                if not value_text:
                    continue
                ws[str(cell).strip()] = value_text

            if category_payloads is not None:
                sections = parse_category_sections(ws)
                snapshots = capture_section_snapshots(ws, sections)
                footer_layout = find_footer_inventory_layout(ws)
                footer_snapshot = (
                    capture_footer_block_snapshot(ws, footer_layout)
                    if footer_layout is not None
                    else None
                )
                footer_row_shift = _calculate_footer_row_shift(
                    sections=sections,
                    category_payloads=category_payloads,
                    empty_section_mode="single_blank_row",
                )
                emit_log(f"[交接班][分类解析] sections={len(sections)}")
                write_category_sections(
                    ws=ws,
                    sections=sections,
                    category_payloads=category_payloads,
                    snapshots=snapshots,
                    empty_section_mode="single_blank_row",
                    preserve_template_values=False,
                    emit_log=emit_log,
                )
                if footer_snapshot is not None:
                    restore_footer_block_snapshot(
                        ws,
                        footer_snapshot,
                        row_shift=footer_row_shift,
                        emit_log=emit_log,
                    )

            final_footer_layout = find_footer_inventory_layout(ws)
            if final_footer_layout is not None:
                deleted_count = trim_rows_below_footer(ws, final_footer_layout)
                emit_log(
                    f"[交接班][footer裁剪] last_row={final_footer_layout.last_row}, deleted={deleted_count}"
                )

            wb.save(temp_path)
        finally:
            wb.close()

    atomic_write_file(
        out_path,
        _writer,
        validator=validate_excel_workbook_file,
        temp_suffix=".tmp",
    )


def copy_template_and_fill(
    *,
    building: str,
    template_cfg: Dict[str, str],
    cell_values: Dict[str, str],
    date_ref: datetime,
    duty_date: str = "",
    duty_shift: str = "",
    category_payloads: Dict[str, Any] | None = None,
    emit_log: Callable[[str], None] = print,
) -> Path:
    source_path = _resolve_template_path(str(template_cfg.get("source_path", "")).strip())
    if not source_path.exists():
        raise FileNotFoundError(f"template not found: {source_path}")

    sheet_name = str(template_cfg.get("sheet_name", "")).strip()
    if not sheet_name:
        raise ValueError("template.sheet_name is required")

    output_dir = Path(str(template_cfg.get("output_dir", "")).strip())
    if not str(output_dir):
        raise ValueError("template.output_dir is required")
    if not output_dir.is_absolute():
        output_dir = get_app_dir() / output_dir
    output_dir = fallback_missing_windows_drive_path(
        output_dir,
        app_dir=get_app_dir(),
        emit_log=emit_log,
        label="交接班日志输出目录",
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    if str(duty_date or "").strip() and str(duty_shift or "").strip():
        base_path = build_output_base_path(
            output_root=output_dir,
            output_type=OUTPUT_TYPE_HANDOVER_LOG,
            building=building,
            suffix=".xlsx",
            duty_date=duty_date,
            duty_shift=duty_shift,
        )
    else:
        file_name_pattern = str(template_cfg.get("file_name_pattern", "{building}_{date}_交接班日志.xlsx")).strip()
        date_format = str(template_cfg.get("date_format", "%Y%m%d")).strip()
        base_name = build_output_filename(
            building=building,
            file_name_pattern=file_name_pattern,
            date_format=date_format,
            date_ref=date_ref,
        )
        base_path = output_dir / base_name

    last_err: BaseException | None = None
    for idx in range(1, 1000):
        out_path = with_index(base_path, idx)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        if out_path.exists():
            continue
        try:
            _write_workbook(
                source_path=source_path,
                out_path=out_path,
                sheet_name=sheet_name,
                cell_values=cell_values,
                category_payloads=category_payloads,
                emit_log=emit_log,
            )
            return out_path
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            try:
                if out_path.exists():
                    out_path.unlink()
                    emit_log(f"[交接班][模板填充] 失败已清理未完成输出文件: {out_path}")
            except Exception:  # noqa: BLE001
                pass
            if _is_file_in_use(exc):
                try:
                    if out_path.exists():
                        out_path.unlink()
                except Exception:  # noqa: BLE001
                    pass
                continue
            raise

    raise RuntimeError(f"failed to write output file after indexed retries: {base_path}") from last_err
