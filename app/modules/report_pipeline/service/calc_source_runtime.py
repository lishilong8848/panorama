from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

import openpyxl


Worksheet = openpyxl.worksheet.worksheet.Worksheet


def apply_building_source_overrides(
    *,
    building: str,
    source_map: Dict[str, Any],
    canonical_metric_name: Callable[[Any], str],
) -> Dict[str, Any]:
    """
    按楼栋应用指标来源兼容规则。
    当前规则：
    - A楼“光伏发电量”优先采用“总_电度”行。
    """
    building_text = str(building or "").strip()
    if building_text != "A楼":
        return source_map
    pv_total_key = canonical_metric_name("总_电度")
    pv_total_row = source_map.get(pv_total_key) or source_map.get("总_电度")
    if pv_total_row is not None:
        source_map["光伏发电量"] = pv_total_row
    return source_map


def locate_stat_columns(
    ws: Worksheet,
    *,
    header_row: int,
    data_start_col: int,
    max_label: str,
    min_label: str,
    avg_label: str,
    norm_text: Callable[[Any], str],
) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for col in range(data_start_col, ws.max_column + 1):
        label = norm_text(ws.cell(header_row, col).value)
        if label == max_label and "max" not in result:
            result["max"] = col
        if label == min_label and "min" not in result:
            result["min"] = col
        if label == avg_label and "avg" not in result:
            result["avg"] = col

    missing = [name for name in ("max", "min", "avg") if name not in result]
    if missing:
        raise ValueError(f"未能在第{header_row}行定位统计列: {missing}")
    return result


def extract_month(
    ws: Worksheet,
    max_col: int,
    *,
    header_row: int,
    data_start_col: int,
    month_pattern: re.Pattern[str] | None = None,
    now_factory: Callable[[], datetime] = datetime.now,
) -> str:
    pattern = month_pattern or re.compile(r"(\d{4})[-/年](\d{1,2})")
    for col in range(data_start_col, max_col):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        if isinstance(value, datetime):
            return value.strftime("%Y-%m")
        if isinstance(value, date):
            return value.strftime("%Y-%m")

        text = str(value).strip()
        match = pattern.search(text)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            return f"{year:04d}-{month:02d}"
    return now_factory().strftime("%Y-%m")


def extract_building(file_path: str) -> str:
    stem = Path(file_path).stem
    return stem.split("_")[0] if "_" in stem else stem


def extract_row_sources(
    ws: Worksheet,
    stat_cols: Dict[str, int],
    *,
    data_start_col: int,
    data_start_row: int,
    type_col: int,
    category_col: int,
    item_col: int,
    to_float: Callable[[Any], float | None],
    canonical_metric_name: Callable[[Any], str],
    row_source_factory: Callable[..., Any],
) -> Tuple[Dict[str, Any], str, str]:
    data_end_col = stat_cols["max"] - 1
    if data_end_col < data_start_col:
        raise ValueError("数据区定位异常：最大值列在 E 列之前")

    current_type = ""
    current_category = ""
    source_map: Dict[str, Any] = {}

    for row in range(data_start_row, ws.max_row + 1):
        type_cell = ws.cell(row, type_col).value
        category_cell = ws.cell(row, category_col).value
        item_cell = ws.cell(row, item_col).value

        if type_cell is not None and str(type_cell).strip():
            current_type = str(type_cell).strip()
        if category_cell is not None and str(category_cell).strip():
            current_category = str(category_cell).strip()

        if item_cell is None or not str(item_cell).strip():
            continue
        item_name = str(item_cell).strip()

        data_values: List[float] = []
        for col in range(data_start_col, data_end_col + 1):
            number = to_float(ws.cell(row, col).value)
            if number is not None:
                data_values.append(number)

        max_value = to_float(ws.cell(row, stat_cols["max"]).value)
        min_value = to_float(ws.cell(row, stat_cols["min"]).value)
        avg_value = to_float(ws.cell(row, stat_cols["avg"]).value)

        if max_value is None and data_values:
            max_value = max(data_values)
        if min_value is None and data_values:
            min_value = min(data_values)
        if avg_value is None and data_values:
            avg_value = sum(data_values) / len(data_values)

        row_source = row_source_factory(
            row_index=row,
            type_name=current_type,
            category_name=current_category,
            item_name=item_name,
            max_value=max_value,
            min_value=min_value,
            avg_value=avg_value,
            data_values=data_values,
        )
        source_map[canonical_metric_name(item_name)] = row_source

    return source_map, current_type, current_category


__all__ = [
    "locate_stat_columns",
    "extract_month",
    "extract_building",
    "extract_row_sources",
    "apply_building_source_overrides",
]
