from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from handover_log_module.core.footer_layout import FooterInventoryLayout


@dataclass
class FooterCellSnapshot:
    style: Any
    value: Any


@dataclass
class FooterRowSnapshot:
    row_idx: int
    row_height: float | None
    cells: Dict[int, FooterCellSnapshot]


@dataclass
class FooterBlockSnapshot:
    layout: FooterInventoryLayout
    rows: List[FooterRowSnapshot]
    merges: List[Tuple[int, int, int, int]]


def _capture_row_snapshot(
    ws: Worksheet,
    row_idx: int,
    *,
    min_col: int = 1,
    max_col: int = 9,
) -> FooterRowSnapshot:
    cells: Dict[int, FooterCellSnapshot] = {}
    for col_idx in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cells[col_idx] = FooterCellSnapshot(style=copy(cell._style), value=copy(cell.value))

    return FooterRowSnapshot(
        row_idx=row_idx,
        row_height=ws.row_dimensions[row_idx].height,
        cells=cells,
    )


def _intersects_row_range(
    merged_range,
    start_row: int,
    end_row: int,
) -> bool:
    return not (merged_range.max_row < start_row or merged_range.min_row > end_row)


def _clear_merges_in_row_range(
    ws: Worksheet,
    start_row: int,
    end_row: int,
) -> None:
    for merged in list(ws.merged_cells.ranges):
        if not _intersects_row_range(merged, start_row, end_row):
            continue
        try:
            ws.unmerge_cells(str(merged))
        except Exception:  # noqa: BLE001
            try:
                ws.merged_cells.ranges.remove(merged)
            except Exception:  # noqa: BLE001
                pass


def capture_footer_block_snapshot(
    ws: Worksheet,
    layout: FooterInventoryLayout,
) -> FooterBlockSnapshot:
    rows = [
        _capture_row_snapshot(ws, row_idx)
        for row_idx in range(layout.title_row, layout.last_row + 1)
    ]
    merges = [
        (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        for merged in ws.merged_cells.ranges
        if _intersects_row_range(merged, layout.title_row, layout.last_row)
    ]
    return FooterBlockSnapshot(layout=layout, rows=rows, merges=merges)


def restore_footer_block_snapshot(
    ws: Worksheet,
    snapshot: FooterBlockSnapshot,
    *,
    row_shift: int,
    emit_log: Callable[[str], None] = print,
) -> FooterInventoryLayout:
    target_title_row = snapshot.layout.title_row + row_shift
    target_last_row = snapshot.layout.last_row + row_shift

    _clear_merges_in_row_range(ws, target_title_row, target_last_row)

    for row_snapshot in snapshot.rows:
        target_row = row_snapshot.row_idx + row_shift
        for col_idx, cell_snapshot in row_snapshot.cells.items():
            cell = ws.cell(row=target_row, column=col_idx)
            cell._style = copy(cell_snapshot.style)
            cell.value = copy(cell_snapshot.value)
        ws.row_dimensions[target_row].height = row_snapshot.row_height

    for min_col, min_row, max_col, max_row in snapshot.merges:
        ws.merge_cells(
            start_row=min_row + row_shift,
            start_column=min_col,
            end_row=max_row + row_shift,
            end_column=max_col,
        )

    restored_layout = FooterInventoryLayout(
        title_row=snapshot.layout.title_row + row_shift,
        header_row=snapshot.layout.header_row + row_shift,
        data_start_row=snapshot.layout.data_start_row + row_shift,
        data_end_row=snapshot.layout.data_end_row + row_shift,
        signoff_start_row=(
            snapshot.layout.signoff_start_row + row_shift
            if snapshot.layout.signoff_start_row is not None
            else None
        ),
        last_row=snapshot.layout.last_row + row_shift,
    )
    emit_log(
        "[交接班][footer恢复] "
        f"title_row={restored_layout.title_row}, data_rows={restored_layout.data_start_row}-{restored_layout.data_end_row}, "
        f"signoff_start_row={restored_layout.signoff_start_row or '-'}"
    )
    return restored_layout
