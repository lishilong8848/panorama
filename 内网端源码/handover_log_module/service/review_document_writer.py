from __future__ import annotations

from typing import Any, Dict, List

import openpyxl

from app.shared.utils.atomic_file import atomic_save_workbook
from handover_log_module.core.footer_layout import FOOTER_DYNAMIC_FIXED_CELLS
from handover_log_module.core.section_layout import capture_section_snapshots, parse_category_sections
from handover_log_module.repository.excel_reader import load_workbook_quietly
from handover_log_module.repository.footer_inventory_writer import write_footer_inventory_table
from handover_log_module.repository.section_writer import write_category_sections


class ReviewDocumentWriter:
    EDITABLE_COLUMNS = tuple("BCDEFGHI")

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    def _sheet_name(self) -> str:
        template_cfg = self.config.get("template", {})
        return str(template_cfg.get("sheet_name", "")).strip()

    @staticmethod
    def _normalize_dirty_regions(dirty_regions: Dict[str, Any] | None) -> Dict[str, bool]:
        if not isinstance(dirty_regions, dict):
            return {
                "fixed_blocks": True,
                "sections": True,
                "footer_inventory": True,
            }
        normalized = {
            "fixed_blocks": bool(dirty_regions.get("fixed_blocks")),
            "sections": bool(dirty_regions.get("sections")),
            "footer_inventory": bool(dirty_regions.get("footer_inventory")),
        }
        if not any(normalized.values()):
            return normalized
        return normalized

    @staticmethod
    def _fixed_cells_from_document(document: Dict[str, Any]) -> Dict[str, str]:
        fixed_cells: Dict[str, str] = {}
        fixed_blocks = document.get("fixed_blocks", [])
        if isinstance(fixed_blocks, list):
            for block in fixed_blocks:
                if not isinstance(block, dict):
                    continue
                for field in block.get("fields", []):
                    if not isinstance(field, dict):
                        continue
                    cell_name = str(field.get("cell", "")).strip().upper()
                    if not cell_name or cell_name in FOOTER_DYNAMIC_FIXED_CELLS:
                        continue
                    fixed_cells[cell_name] = str(field.get("value", "") or "")
        if "A1" not in fixed_cells:
            fixed_cells["A1"] = str(document.get("title", "") or "")
        return fixed_cells

    def _normalize_section_columns(self, section: Dict[str, Any]) -> List[Dict[str, Any]]:
        columns = section.get("columns", [])
        if isinstance(columns, list) and columns:
            normalized: List[Dict[str, Any]] = []
            for idx, column in enumerate(columns):
                if not isinstance(column, dict):
                    continue
                key = str(column.get("key", "")).strip().upper()
                if not key:
                    continue
                raw_source_cols = column.get("source_cols", [])
                if not isinstance(raw_source_cols, list) or not raw_source_cols:
                    raw_source_cols = [key]
                source_cols = [str(value or "").strip().upper() for value in raw_source_cols if str(value or "").strip()]
                if not source_cols:
                    source_cols = [key]
                normalized.append(
                    {
                        "key": key,
                        "label": str(column.get("label", "") or key),
                        "source_cols": source_cols,
                        "span": int(column.get("span", len(source_cols)) or len(source_cols)),
                    }
                )
            if normalized:
                return normalized

        header = section.get("header", [])
        fallback_columns: List[Dict[str, Any]] = []
        for idx, col in enumerate(self.EDITABLE_COLUMNS):
            label = ""
            if isinstance(header, list) and idx < len(header):
                label = str(header[idx] or "")
            fallback_columns.append(
                {
                    "key": col,
                    "label": label or col,
                    "source_cols": [col],
                    "span": 1,
                }
            )
        return fallback_columns

    def _category_payloads_from_document(self, document: Dict[str, Any]) -> Dict[str, Any]:
        output: Dict[str, Any] = {}
        sections = document.get("sections", [])
        if not isinstance(sections, list):
            return output
        for section in sections:
            if not isinstance(section, dict):
                continue
            section_name = str(section.get("name", "")).strip()
            if not section_name:
                continue
            section_columns = self._normalize_section_columns(section)
            payload_rows: List[Dict[str, Any]] = []
            for row in section.get("rows", []):
                if not isinstance(row, dict):
                    continue
                source_cells = row.get("cells", {})
                if not isinstance(source_cells, dict):
                    continue
                normalized_cells = {col: "" for col in self.EDITABLE_COLUMNS}
                for column in section_columns:
                    key = str(column.get("key", "")).strip().upper()
                    source_cols = column.get("source_cols", [])
                    if not key or not isinstance(source_cols, list) or not source_cols:
                        continue
                    value = str(source_cells.get(key, "") or "")
                    lead_col = str(source_cols[0] or "").strip().upper()
                    if lead_col in normalized_cells:
                        normalized_cells[lead_col] = value
                    for follower_col in source_cols[1:]:
                        follower_col = str(follower_col or "").strip().upper()
                        if follower_col in normalized_cells:
                            normalized_cells[follower_col] = ""
                payload_rows.append({"cells": normalized_cells})
            output[section_name] = payload_rows
        return output

    @staticmethod
    def _inventory_footer_block(document: Dict[str, Any]) -> Dict[str, Any] | None:
        footer_blocks = document.get("footer_blocks", [])
        if not isinstance(footer_blocks, list):
            return None
        for block in footer_blocks:
            if isinstance(block, dict) and str(block.get("type", "")).strip() == "inventory_table":
                return block
        return None

    def write(
        self,
        *,
        output_file: str,
        document: Dict[str, Any],
        dirty_regions: Dict[str, Any] | None = None,
    ) -> None:
        normalized_dirty = self._normalize_dirty_regions(dirty_regions)
        workbook = load_workbook_quietly(output_file)
        try:
            sheet_name = self._sheet_name()
            ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

            if normalized_dirty.get("fixed_blocks"):
                fixed_cells = self._fixed_cells_from_document(document)
                for cell_name, value in fixed_cells.items():
                    ws[cell_name] = value

            if normalized_dirty.get("sections"):
                sections = parse_category_sections(ws)
                snapshots = capture_section_snapshots(ws, sections)
                write_category_sections(
                    ws=ws,
                    sections=sections,
                    category_payloads=self._category_payloads_from_document(document),
                    snapshots=snapshots,
                    empty_section_mode="single_blank_row",
                    preserve_template_values=False,
                )

            if normalized_dirty.get("footer_inventory"):
                write_footer_inventory_table(
                    ws=ws,
                    inventory_block=self._inventory_footer_block(document),
                )
            atomic_save_workbook(workbook, output_file, temp_suffix=".tmp")
        finally:
            workbook.close()
