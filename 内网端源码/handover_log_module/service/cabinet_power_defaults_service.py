from __future__ import annotations

import copy
from pathlib import Path
from typing import Any, Dict

import openpyxl

from app.shared.utils.atomic_file import atomic_save_workbook
from handover_log_module.repository.excel_reader import load_workbook_quietly


class CabinetPowerDefaultsService:
    DEFAULTS_KEY = "cabinet_power_defaults_by_building"
    CELLS = ("B13", "D13", "F13", "H13")

    @classmethod
    def normalize_cells(cls, cells: Any) -> Dict[str, str]:
        source = cells if isinstance(cells, dict) else {}
        return {cell: str(source.get(cell, "") or "").strip() for cell in cls.CELLS}

    @staticmethod
    def _review_ui_from_v3(cfg: Dict[str, Any], *, create: bool) -> Dict[str, Any]:
        features = cfg.setdefault("features", {}) if create else cfg.get("features", {})
        if not isinstance(features, dict):
            features = {}
            if create:
                cfg["features"] = features
        handover = features.setdefault("handover_log", {}) if create else features.get("handover_log", {})
        if not isinstance(handover, dict):
            handover = {}
            if create:
                features["handover_log"] = handover
        review_ui = handover.setdefault("review_ui", {}) if create else handover.get("review_ui", {})
        if not isinstance(review_ui, dict):
            review_ui = {}
            if create:
                handover["review_ui"] = review_ui
        return review_ui

    @staticmethod
    def _review_ui_from_runtime_root(cfg: Dict[str, Any], *, create: bool) -> Dict[str, Any]:
        handover = cfg.setdefault("handover_log", {}) if create else cfg.get("handover_log", {})
        if not isinstance(handover, dict):
            handover = {}
            if create:
                cfg["handover_log"] = handover
        review_ui = handover.setdefault("review_ui", {}) if create else handover.get("review_ui", {})
        if not isinstance(review_ui, dict):
            review_ui = {}
            if create:
                handover["review_ui"] = review_ui
        return review_ui

    @staticmethod
    def _review_ui_from_handover_cfg(cfg: Dict[str, Any], *, create: bool) -> Dict[str, Any]:
        review_ui = cfg.setdefault("review_ui", {}) if create else cfg.get("review_ui", {})
        if not isinstance(review_ui, dict):
            review_ui = {}
            if create:
                cfg["review_ui"] = review_ui
        return review_ui

    def _review_ui_cfg(self, cfg: Dict[str, Any], *, create: bool = False) -> Dict[str, Any]:
        if not isinstance(cfg, dict):
            return {}
        if isinstance(cfg.get("features"), dict):
            return self._review_ui_from_v3(cfg, create=create)
        if isinstance(cfg.get("handover_log"), dict) and isinstance(cfg.get("download"), dict):
            return self._review_ui_from_runtime_root(cfg, create=create)
        return self._review_ui_from_handover_cfg(cfg, create=create)

    def extract_cells_from_document(self, document: Dict[str, Any]) -> Dict[str, str]:
        output = self.normalize_cells({})
        fixed_blocks = document.get("fixed_blocks", []) if isinstance(document, dict) else []
        if not isinstance(fixed_blocks, list):
            return output
        for block in fixed_blocks:
            if not isinstance(block, dict):
                continue
            for field in block.get("fields", []):
                if not isinstance(field, dict):
                    continue
                cell = str(field.get("cell", "")).strip().upper()
                if cell not in self.CELLS:
                    continue
                output[cell] = str(field.get("value", "") or "").strip()
        return output

    def get_building_defaults(self, config: Dict[str, Any], building: str) -> Dict[str, str] | None:
        review_ui = self._review_ui_cfg(config, create=False)
        defaults_by_building = review_ui.get(self.DEFAULTS_KEY, {}) if isinstance(review_ui, dict) else {}
        if not isinstance(defaults_by_building, dict):
            return None
        building_name = str(building or "").strip()
        if not building_name:
            return None
        if building_name not in defaults_by_building:
            return None
        payload = defaults_by_building.get(building_name, {})
        if not isinstance(payload, dict):
            return None
        cells = payload.get("cells", {})
        if not isinstance(cells, dict):
            return None
        return self.normalize_cells(cells)

    def set_building_defaults(
        self,
        config: Dict[str, Any],
        building: str,
        cells: Dict[str, str],
    ) -> Dict[str, Any]:
        updated = copy.deepcopy(config if isinstance(config, dict) else {})
        review_ui = self._review_ui_cfg(updated, create=True)
        defaults_by_building = review_ui.get(self.DEFAULTS_KEY, {})
        if not isinstance(defaults_by_building, dict):
            defaults_by_building = {}
        defaults_by_building[str(building or "").strip()] = {"cells": self.normalize_cells(cells)}
        review_ui[self.DEFAULTS_KEY] = defaults_by_building
        return updated

    def apply_building_defaults_to_output(
        self,
        *,
        config: Dict[str, Any],
        building: str,
        output_file: str | Path,
        sheet_name: str,
        emit_log=print,
    ) -> int | None:
        cells = self.get_building_defaults(config, building)
        if cells is None:
            return None

        output_path = Path(str(output_file).strip())
        workbook = load_workbook_quietly(output_path)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"交接班模板sheet不存在: {sheet_name}")
            ws = workbook[sheet_name]
            for cell, value in cells.items():
                ws[cell] = value
            atomic_save_workbook(workbook, output_path, temp_suffix=".tmp")
        finally:
            workbook.close()
        emit_log(
            f"[交接班][机柜上下电默认] 已应用楼栋默认值 building={building}, fields={len(self.CELLS)}, output={output_path}"
        )
        return len(self.CELLS)
