from __future__ import annotations

import copy
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import openpyxl

from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.bitable_target_resolver import BitableTargetResolver
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from app.modules.report_pipeline.core.metrics_rules import DERIVED_CATEGORY_HINT
from handover_log_module.core.day_metric_direct_rules import DAY_METRIC_DIRECT_TYPE_DEFINITIONS
from handover_log_module.core.models import MetricHit


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(base or {})
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key] = _deep_merge(out[key], value)
        else:
            out[key] = copy.deepcopy(value)
    return out


def _extract_numbers(text: Any) -> List[float]:
    raw = str(text or "").strip()
    if not raw:
        return []
    matches = re.findall(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?", raw)
    values: List[float] = []
    for match in matches:
        candidate = match.replace(",", "")
        try:
            values.append(float(candidate))
        except Exception:  # noqa: BLE001
            continue
    return values


class DayMetricBitableExportService:
    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "source": {
                "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
                "table_id": "tblAHGF8mV6U9jid",
                "base_url": "",
                "wiki_url": "",
                "page_size": 500,
                "max_records": 5000,
                "delete_batch_size": 200,
                "create_batch_size": 200,
            },
            "fields": {
                "type": "类型",
                "building": "楼栋",
                "date": "日期",
                "value": "数值",
                "position_code": "位置/编号",
            },
            "missing_value_policy": "zero",
        }

    @staticmethod
    def _type_definitions() -> List[Dict[str, str]]:
        return [copy.deepcopy(item) for item in DAY_METRIC_DIRECT_TYPE_DEFINITIONS]

    def _runtime_day_metric_cfg(self) -> Dict[str, Any] | None:
        if "day_metric_upload" in self.config and isinstance(self.config.get("day_metric_upload"), dict):
            return self.config.get("day_metric_upload", {})
        features = self.config.get("features", {})
        if isinstance(features, dict) and "day_metric_upload" in features and isinstance(features.get("day_metric_upload"), dict):
            return features.get("day_metric_upload", {})
        if "handover_log" in self.config and isinstance(self.config.get("handover_log"), dict):
            handover = self.config.get("handover_log", {})
            if "day_metric_upload" in handover and isinstance(handover.get("day_metric_upload"), dict):
                return handover.get("day_metric_upload", {})
        return None

    def _global_feishu_cfg(self) -> Dict[str, Any]:
        return require_feishu_auth_settings(self.config)

    def _template_cfg(self) -> Dict[str, Any]:
        if "template" in self.config and isinstance(self.config.get("template"), dict):
            return self.config.get("template", {})
        handover = self.config.get("handover_log", {})
        if isinstance(handover, dict) and "template" in handover and isinstance(handover.get("template"), dict):
            return handover.get("template", {})
        features = self.config.get("features", {})
        if isinstance(features, dict):
            handover_feature = features.get("handover_log", {})
            if isinstance(handover_feature, dict) and "template" in handover_feature and isinstance(handover_feature.get("template"), dict):
                return handover_feature.get("template", {})
        return {}

    def _normalize_cfg(self) -> Dict[str, Any]:
        runtime_day_metric_cfg = self._runtime_day_metric_cfg()
        if isinstance(runtime_day_metric_cfg, dict):
            raw = runtime_day_metric_cfg.get("target", {})
            cfg = _deep_merge(self._defaults(), raw if isinstance(raw, dict) else {})
        else:
            raw = self.config.get("day_metric_export", {})
            cfg = _deep_merge(self._defaults(), raw if isinstance(raw, dict) else {})

        source = cfg.get("source", {})
        fields = cfg.get("fields", {})
        if not isinstance(source, dict):
            source = {}
        if not isinstance(fields, dict):
            fields = {}

        source["app_token"] = str(source.get("app_token", "")).strip()
        source["table_id"] = str(source.get("table_id", "")).strip()
        source["base_url"] = str(source.get("base_url", "")).strip()
        source["wiki_url"] = str(source.get("wiki_url", "")).strip()
        source["page_size"] = max(1, int(source.get("page_size", 500) or 500))
        source["max_records"] = max(0, int(source.get("max_records", 5000) or 5000))
        source["delete_batch_size"] = max(1, int(source.get("delete_batch_size", 200) or 200))
        source["create_batch_size"] = max(1, int(source.get("create_batch_size", 200) or 200))
        cfg["source"] = source

        fields["type"] = str(fields.get("type", "类型")).strip() or "类型"
        fields["building"] = str(fields.get("building", "楼栋")).strip() or "楼栋"
        fields["date"] = str(fields.get("date", "日期")).strip() or "日期"
        fields["value"] = str(fields.get("value", "数值")).strip() or "数值"
        fields["position_code"] = str(fields.get("position_code", "位置/编号")).strip() or "位置/编号"
        cfg["fields"] = fields

        cfg["missing_value_policy"] = str(cfg.get("missing_value_policy", "zero")).strip().lower() or "zero"
        return cfg

    def _sheet_name(self) -> str:
        template_cfg = self._template_cfg()
        return str(template_cfg.get("sheet_name", "")).strip()

    def _resolve_target(self, cfg: Dict[str, Any]) -> Dict[str, str]:
        global_feishu = self._global_feishu_cfg()

        app_id = str(global_feishu.get("app_id", "")).strip()
        app_secret = str(global_feishu.get("app_secret", "")).strip()
        if not app_id or not app_secret:
            raise ValueError("飞书配置缺失: common.feishu_auth.app_id/app_secret")

        source = cfg.get("source", {}) if isinstance(cfg.get("source", {}), dict) else {}
        return BitableTargetResolver(
            app_id=app_id,
            app_secret=app_secret,
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
        ).resolve(source)

    def build_target_descriptor(self, cfg: Dict[str, Any] | None = None, *, force_refresh: bool = False) -> Dict[str, str]:
        normalized = cfg if isinstance(cfg, dict) else self._normalize_cfg()
        source = normalized.get("source", {}) if isinstance(normalized.get("source", {}), dict) else {}
        global_feishu = self._global_feishu_cfg()
        resolver = BitableTargetResolver(
            app_id=str(global_feishu.get("app_id", "")).strip(),
            app_secret=str(global_feishu.get("app_secret", "")).strip(),
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
        )
        base_url = str(source.get("base_url", "") or "").strip()
        wiki_url = str(source.get("wiki_url", "") or "").strip()
        if base_url or wiki_url:
            resolved = resolver.resolve(source)
            display_url = wiki_url or base_url or str(resolved.get("bitable_url", "")).strip()
            return {
                "configured_app_token": str(source.get("app_token", "") or "").strip(),
                "operation_app_token": str(resolved.get("app_token", "") or "").strip(),
                "app_token": str(resolved.get("app_token", "") or "").strip(),
                "table_id": str(resolved.get("table_id", "") or "").strip(),
                "target_kind": "wiki_url" if wiki_url else "base_url",
                "resolved_from": str(resolved.get("resolved_from", "") or "").strip() or ("wiki_url" if wiki_url else "base_url"),
                "display_url": display_url,
                "bitable_url": display_url,
                "wiki_node_token": str(resolved.get("wiki_node_token", "") or "").strip(),
                "message": "",
                "resolved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        return dict(
            resolver.resolve_token_pair_preview(
                configured_app_token=str(source.get("app_token", "")).strip(),
                table_id=str(source.get("table_id", "")).strip(),
                force_refresh=force_refresh,
            )
        )

    def _new_client(
        self,
        cfg: Dict[str, Any],
        *,
        resolved_target: Dict[str, str] | None = None,
    ) -> FeishuBitableClient:
        global_feishu = self._global_feishu_cfg()

        app_id = str(global_feishu.get("app_id", "")).strip()
        app_secret = str(global_feishu.get("app_secret", "")).strip()
        if not app_id or not app_secret:
            raise ValueError("飞书配置缺失: common.feishu_auth.app_id/app_secret")

        target = resolved_target if isinstance(resolved_target, dict) else self._resolve_target(cfg)
        return FeishuBitableClient(
            app_id=app_id,
            app_secret=app_secret,
            app_token=target["app_token"],
            calc_table_id=target["table_id"],
            attachment_table_id=target["table_id"],
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda x: str(x or "").strip(),
            dimension_mapping={},
        )

    @staticmethod
    def _midnight_timestamp_ms(duty_date: str) -> int:
        dt = datetime.strptime(str(duty_date).strip(), "%Y-%m-%d")
        return int(dt.timestamp() * 1000)

    @staticmethod
    def _normalize_text_field(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, dict):
            for key in ("text", "name", "value"):
                text = str(value.get(key, "")).strip()
                if text:
                    return text
            return ""
        if isinstance(value, list):
            for item in value:
                text = DayMetricBitableExportService._normalize_text_field(item)
                if text:
                    return text
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_date_field_to_midnight_ms(value: Any) -> int | None:
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            number = int(value)
            if number <= 0:
                return None
            if number < 10**11:
                number *= 1000
            dt = datetime.fromtimestamp(number / 1000)
            midnight = datetime(dt.year, dt.month, dt.day, 0, 0, 0)
            return int(midnight.timestamp() * 1000)
        text = str(value).strip()
        if not text:
            return None
        candidate = text[:10].replace("/", "-")
        try:
            dt = datetime.strptime(candidate, "%Y-%m-%d")
        except ValueError:
            return None
        return int(dt.timestamp() * 1000)

    @staticmethod
    def _normalize_cell_values(values: Dict[str, Any] | None) -> Dict[str, Any]:
        if not isinstance(values, dict):
            return {}
        out: Dict[str, Any] = {}
        for raw_key, raw_val in values.items():
            key = str(raw_key or "").strip().upper()
            if key:
                out[key] = raw_val
        return out

    def build_metric_context(self, resolved_values_by_id: Dict[str, Any] | None) -> Dict[str, Any]:
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        out: Dict[str, Any] = {}
        for item in self._type_definitions():
            metric_id = str(item.get("metric_id", "")).strip()
            if not metric_id:
                continue
            if metric_id in resolved:
                out[metric_id] = resolved.get(metric_id)
        return out

    @staticmethod
    def _serialize_hit_payload(metric_key: str, hit: MetricHit | Dict[str, Any] | Any) -> Dict[str, Any]:
        if isinstance(hit, dict):
            row_index = int(hit.get("row_index", 0) or 0)
            d_name = str(hit.get("d_name", "")).strip()
            b_norm = str(hit.get("b_norm", "")).strip()
            c_norm = str(hit.get("c_norm", "")).strip()
            b_text = str(hit.get("b_text", "")).strip()
            c_text = str(hit.get("c_text", "")).strip()
        else:
            row_index = int(getattr(hit, "row_index", 0) or 0)
            d_name = str(getattr(hit, "d_name", "")).strip()
            b_norm = str(getattr(hit, "b_norm", "")).strip()
            c_norm = str(getattr(hit, "c_norm", "")).strip()
            b_text = str(getattr(hit, "b_text", "")).strip()
            c_text = str(getattr(hit, "c_text", "")).strip()
        return {
            "metric_key": str(metric_key or "").strip(),
            "row_index": row_index,
            "d_name": d_name,
            "b_norm": b_norm,
            "c_norm": c_norm,
            "b_text": b_text,
            "c_text": c_text,
        }

    @classmethod
    def _normalize_origin_payload(cls, payload: Dict[str, Any] | None) -> Dict[str, Any]:
        data = payload if isinstance(payload, dict) else {}
        return {
            "metric_key": str(data.get("metric_key", "")).strip(),
            "row_index": int(data.get("row_index", 0) or 0),
            "d_name": str(data.get("d_name", "")).strip(),
            "b_norm": str(data.get("b_norm", "")).strip(),
            "c_norm": str(data.get("c_norm", "")).strip(),
            "b_text": str(data.get("b_text", "")).strip(),
            "c_text": str(data.get("c_text", "")).strip(),
        }

    @classmethod
    def _normalize_metric_origin_context(cls, metric_origin_context: Dict[str, Any] | None) -> Dict[str, Any]:
        raw = metric_origin_context if isinstance(metric_origin_context, dict) else {}
        by_metric_id_raw = raw.get("by_metric_id", {})
        by_target_cell_raw = raw.get("by_target_cell", {})
        by_metric_id: Dict[str, Dict[str, Any]] = {}
        by_target_cell: Dict[str, Dict[str, Any]] = {}
        if isinstance(by_metric_id_raw, dict):
            for metric_id, payload in by_metric_id_raw.items():
                metric_key = str(metric_id or "").strip()
                if metric_key:
                    by_metric_id[metric_key] = cls._normalize_origin_payload(payload if isinstance(payload, dict) else {})
        if isinstance(by_target_cell_raw, dict):
            for cell_name, payload in by_target_cell_raw.items():
                target_cell = str(cell_name or "").strip().upper()
                if target_cell:
                    by_target_cell[target_cell] = cls._normalize_origin_payload(payload if isinstance(payload, dict) else {})
        return {
            "by_metric_id": by_metric_id,
            "by_target_cell": by_target_cell,
        }

    def serialize_metric_origin_context(
        self,
        *,
        hits: Dict[str, MetricHit] | Dict[str, Any] | None,
        effective_config: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        by_metric_id: Dict[str, Dict[str, Any]] = {}
        hits_map = hits if isinstance(hits, dict) else {}
        for metric_key, hit in hits_map.items():
            metric_id = str(metric_key or "").strip()
            if not metric_id:
                continue
            by_metric_id[metric_id] = self._serialize_hit_payload(metric_id, hit)

        by_target_cell: Dict[str, Dict[str, Any]] = {}
        effective = effective_config if isinstance(effective_config, dict) else {}
        cell_mapping = effective.get("cell_mapping", {})
        if isinstance(cell_mapping, dict):
            for metric_key, target_cell in cell_mapping.items():
                metric_id = str(metric_key or "").strip()
                cell_name = str(target_cell or "").strip().upper()
                if not metric_id or not cell_name:
                    continue
                payload = by_metric_id.get(metric_id)
                if not isinstance(payload, dict):
                    continue
                by_target_cell[cell_name] = {
                    **payload,
                    "metric_key": metric_id,
                }

        return self._normalize_metric_origin_context(
            {
                "by_metric_id": by_metric_id,
                "by_target_cell": by_target_cell,
            }
        )

    @staticmethod
    def _clean_position_text(value: Any) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        raw = re.sub(r"[-_/\\ ]*冷通道\s*$", "", raw)
        raw = re.sub(r"[-_/\\ ]+$", "", raw).strip()
        raw = re.sub(r"\s+", " ", raw)
        return raw

    @classmethod
    def _compose_position_code(
        cls,
        origin_payload: Dict[str, Any] | None,
        *,
        type_item: Dict[str, Any] | None = None,
    ) -> str:
        payload = origin_payload if isinstance(origin_payload, dict) else {}
        type_cfg = type_item if isinstance(type_item, dict) else {}
        type_name = str(type_cfg.get("name", "")).strip()
        metric_key = str(payload.get("metric_key") or type_cfg.get("metric_id") or "").strip()
        b_norm = cls._clean_position_text(payload.get("b_norm", ""))
        c_norm = cls._clean_position_text(payload.get("c_norm", ""))
        c_text = cls._clean_position_text(payload.get("c_text", ""))
        b_text = cls._clean_position_text(payload.get("b_text", ""))
        d_name = cls._clean_position_text(payload.get("d_name", ""))

        pue_hint = str(DERIVED_CATEGORY_HINT.get("PUE", "pue能耗数据计算")).strip() or "pue能耗数据计算"
        if type_name in {"总负荷（KW）", "IT总负荷（KW）"} or metric_key in {"city_power", "it_power"}:
            return pue_hint
        if type_name == "蓄水池后备最短时间（H）" or metric_key in {"water_pool_backup_time", "water_backup_shortest"}:
            for candidate in (d_name, c_text, c_norm, b_text, b_norm):
                if "负载率" in candidate:
                    return candidate
            return "负载率"
        if type_name == "供油可用时长（H）" or metric_key == "oil_backup_time":
            for candidate in (b_text, b_norm, c_text, c_norm, d_name):
                if "燃油系统" in candidate:
                    return candidate
            return "燃油系统"

        if b_norm and c_norm:
            return f"{b_norm} {c_norm}"
        if c_norm:
            return c_norm
        if c_text:
            return c_text
        if b_norm:
            return b_norm
        if b_text:
            return b_text
        if d_name:
            return d_name
        return ""

    def _resolve_origin_payload(
        self,
        *,
        type_item: Dict[str, Any],
        metric_origin_context: Dict[str, Any],
    ) -> Dict[str, Any]:
        source_type = str(type_item.get("source", "")).strip().lower()
        metric_id = str(type_item.get("metric_id", "")).strip()
        if metric_id:
            payload = metric_origin_context.get("by_metric_id", {}).get(metric_id, {})
            return payload if isinstance(payload, dict) else {}
        if source_type == "metric":
            payload = metric_origin_context.get("by_metric_id", {}).get(metric_id, {})
            return payload if isinstance(payload, dict) else {}
        if source_type in {"cell", "cell_percent", "cell_min_pair"}:
            cell_name = str(type_item.get("cell", "")).strip().upper()
            payload = metric_origin_context.get("by_target_cell", {}).get(cell_name, {})
            return payload if isinstance(payload, dict) else {}
        return {}

    def build_deferred_state(
        self,
        *,
        duty_shift: str,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        return {
            "status": "pending_review",
            "reason": "await_all_confirmed",
            "uploaded_count": 0,
            "error": "",
            "uploaded_at": "",
            "uploaded_revision": 0,
            "metric_values_by_id": self.build_metric_context(resolved_values_by_id),
            "metric_origin_context": self._normalize_metric_origin_context(metric_origin_context),
        }

    def _load_workbook_cell_values(self, output_file: str, cfg: Dict[str, Any]) -> Dict[str, Any]:
        path = Path(str(output_file or "").strip())
        if not path.exists():
            raise FileNotFoundError(f"交接班成品文件不存在: {path}")

        needed_cells: List[str] = []
        for item in cfg.get("types", []):
            if not isinstance(item, dict):
                continue
            source_type = str(item.get("source", "")).strip().lower()
            if source_type == "metric":
                continue
            cell_name = str(item.get("cell", "")).strip().upper()
            if cell_name and cell_name not in needed_cells:
                needed_cells.append(cell_name)

        workbook = openpyxl.load_workbook(path, data_only=True)
        try:
            sheet_name = self._sheet_name()
            ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            values: Dict[str, Any] = {}
            for cell_name in needed_cells:
                values[cell_name] = ws[cell_name].value
            return self._normalize_cell_values(values)
        finally:
            workbook.close()

    @staticmethod
    def _resolve_numeric_from_type(
        *,
        type_item: Dict[str, Any],
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
    ) -> Optional[float]:
        metric_id = str(type_item.get("metric_id", "")).strip()
        if metric_id:
            raw = resolved_values_by_id.get(metric_id)
            numbers = _extract_numbers(raw)
            return numbers[0] if numbers else None

        source_type = str(type_item.get("source", "")).strip().lower()
        if source_type == "metric":
            raw = resolved_values_by_id.get(metric_id)
            numbers = _extract_numbers(raw)
            return numbers[0] if numbers else None

        cell_name = str(type_item.get("cell", "")).strip().upper()
        raw_cell = cell_values.get(cell_name, "")
        numbers = _extract_numbers(raw_cell)
        if not numbers:
            return None
        if source_type == "cell_min_pair":
            return min(numbers)
        return numbers[0]

    def _prepare_records(
        self,
        *,
        cfg: Dict[str, Any],
        building: str,
        duty_date: str,
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
        metric_origin_context: Dict[str, Any] | None = None,
    ) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        fields = cfg.get("fields", {})
        missing_policy = str(cfg.get("missing_value_policy", "zero")).strip().lower()
        duty_ms = self._midnight_timestamp_ms(duty_date)
        origin_context = self._normalize_metric_origin_context(metric_origin_context)
        records: List[Dict[str, Any]] = []
        preview: List[Dict[str, Any]] = []

        for item in self._type_definitions():
            if not isinstance(item, dict):
                continue

            numeric = self._resolve_numeric_from_type(
                type_item=item,
                cell_values=cell_values,
                resolved_values_by_id=resolved_values_by_id,
            )
            if numeric is None and missing_policy == "zero":
                numeric = 0.0
            if numeric is None:
                numeric = 0.0
            position_code_text = self._compose_position_code(
                self._resolve_origin_payload(
                    type_item=item,
                    metric_origin_context=origin_context,
                ),
                type_item=item,
            )

            row_fields = {
                str(fields.get("type", "类型")).strip(): str(item.get("name", "")).strip(),
                str(fields.get("building", "楼栋")).strip(): str(building or "").strip(),
                str(fields.get("date", "日期")).strip(): duty_ms,
                str(fields.get("value", "数值")).strip(): numeric,
                str(fields.get("position_code", "位置/编号")).strip(): position_code_text,
            }
            records.append(row_fields)
            if len(preview) < 3:
                preview.append(row_fields)
        return records, preview

    def _matching_existing_records(
        self,
        *,
        existing_records: List[Dict[str, Any]],
        building: str,
        duty_date: str,
        cfg: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        fields = cfg.get("fields", {})
        type_field = str(fields.get("type", "类型")).strip()
        building_field = str(fields.get("building", "楼栋")).strip()
        date_field = str(fields.get("date", "日期")).strip()
        target_building = str(building or "").strip()
        target_date_ms = self._midnight_timestamp_ms(duty_date)
        allowed_types = {str(item.get("name", "")).strip() for item in self._type_definitions() if str(item.get("name", "")).strip()}
        matched: List[Dict[str, Any]] = []
        for item in existing_records:
            if not isinstance(item, dict):
                continue
            payload_fields = item.get("fields", {})
            if not isinstance(payload_fields, dict):
                continue
            item_building = self._normalize_text_field(payload_fields.get(building_field))
            item_type = self._normalize_text_field(payload_fields.get(type_field))
            item_date_ms = self._normalize_date_field_to_midnight_ms(payload_fields.get(date_field))
            if item_building != target_building:
                continue
            if item_date_ms != target_date_ms:
                continue
            if item_type not in allowed_types:
                continue
            matched.append(copy.deepcopy(item))
        return matched

    def list_existing_records_for_unit(self, *, building: str, duty_date: str) -> List[Dict[str, Any]]:
        cfg = self._normalize_cfg()
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        client = self._new_client(cfg, resolved_target=resolved_target)
        table_id = str(resolved_target.get("table_id", "")).strip()
        existing_records = client.list_records(
            table_id=table_id,
            page_size=int(source.get("page_size", 500) or 500),
            max_records=int(source.get("max_records", 5000) or 5000),
        )
        return self._matching_existing_records(
            existing_records=existing_records,
            building=building,
            duty_date=duty_date,
            cfg=cfg,
        )

    def delete_existing_records_for_unit(
        self,
        *,
        building: str,
        duty_date: str,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        client = self._new_client(cfg, resolved_target=resolved_target)
        table_id = str(resolved_target.get("table_id", "")).strip()
        existing_records = client.list_records(
            table_id=table_id,
            page_size=int(source.get("page_size", 500) or 500),
            max_records=int(source.get("max_records", 5000) or 5000),
        )
        matched = self._matching_existing_records(
            existing_records=existing_records,
            building=building,
            duty_date=duty_date,
            cfg=cfg,
        )
        record_ids = [
            str(item.get("record_id", "")).strip()
            for item in matched
            if str(item.get("record_id", "")).strip()
        ]
        deleted_count = 0
        if record_ids:
            deleted_count = client.batch_delete_records(
                table_id=table_id,
                record_ids=record_ids,
                batch_size=int(source.get("delete_batch_size", 200) or 200),
            )
        emit_log(f"[12项独立上传] 删除旧记录 building={building}, duty_date={duty_date}, count={deleted_count}")
        return {
            "status": "ok",
            "deleted_count": deleted_count,
            "matched_records": matched,
        }

    def _run_with_values(
        self,
        *,
        cfg: Dict[str, Any],
        building: str,
        duty_date: str,
        duty_shift: str,
        cell_values: Dict[str, Any],
        resolved_values_by_id: Dict[str, Any],
        metric_origin_context: Dict[str, Any] | None,
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        source = cfg.get("source", {})
        resolved_target = self._resolve_target(cfg)
        table_id = str(resolved_target.get("table_id", "")).strip()
        batch_size = int(source.get("create_batch_size", 200) or 200)
        records, preview = self._prepare_records(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            cell_values=cell_values,
            resolved_values_by_id=resolved_values_by_id,
            metric_origin_context=metric_origin_context,
        )

        try:
            client = self._new_client(cfg, resolved_target=resolved_target)
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[12项独立上传] 初始化失败: {exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
                "error": str(exc),
            }

        emit_log(f"[12项独立上传] 开始写入 building={building}, duty_date={duty_date}, records={len(records)}")
        try:
            existing_records = client.list_records(
                table_id=table_id,
                page_size=int(source.get("page_size", 500) or 500),
                max_records=int(source.get("max_records", 5000) or 5000),
            )
            matched = self._matching_existing_records(
                existing_records=existing_records,
                building=building,
                duty_date=duty_date,
                cfg=cfg,
            )
            record_ids = [
                str(item.get("record_id", "")).strip()
                for item in matched
                if str(item.get("record_id", "")).strip()
            ]
            deleted_count = 0
            if record_ids:
                deleted_count = client.batch_delete_records(
                    table_id=table_id,
                    record_ids=record_ids,
                    batch_size=int(source.get("delete_batch_size", 200) or 200),
                )
            emit_log(
                f"[12项独立上传] 删除旧记录完成 building={building}, duty_date={duty_date}, count={deleted_count}"
            )
            client.batch_create_records(table_id=table_id, fields_list=records, batch_size=batch_size)
            emit_log(
                f"[12项独立上传] 写入完成 building={building}, duty_date={duty_date}, created={len(records)}"
            )
            return {
                "status": "ok",
                "uploaded_count": len(records),
                "created_records": len(records),
                "deleted_records": deleted_count,
                "records_preview": preview,
                "error": "",
            }
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[12项独立上传] 写入失败 error={exc}")
            return {
                "status": "failed",
                "uploaded_count": 0,
                "created_records": 0,
                "deleted_records": 0,
                "records_preview": preview,
                "error": str(exc),
            }

    def run(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        filled_cell_values: Dict[str, Any] | None,
        resolved_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        cell_values = self._normalize_cell_values(filled_cell_values)
        resolved = resolved_values_by_id if isinstance(resolved_values_by_id, dict) else {}
        return self._run_with_values(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            cell_values=cell_values,
            resolved_values_by_id=resolved,
            metric_origin_context=metric_origin_context,
            emit_log=emit_log,
        )

    def rewrite_from_output_file(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        output_file: str,
        metric_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        resolved = metric_values_by_id if isinstance(metric_values_by_id, dict) else {}
        return self._run_with_values(
            cfg=cfg,
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            cell_values={},
            resolved_values_by_id=resolved,
            metric_origin_context=metric_origin_context,
            emit_log=emit_log,
        )

    def run_from_output_file(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        output_file: str,
        metric_values_by_id: Dict[str, Any] | None,
        metric_origin_context: Dict[str, Any] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        return self.rewrite_from_output_file(
            building=building,
            duty_date=duty_date,
            duty_shift=duty_shift,
            output_file=output_file,
            metric_values_by_id=metric_values_by_id,
            metric_origin_context=metric_origin_context,
            emit_log=emit_log,
        )
