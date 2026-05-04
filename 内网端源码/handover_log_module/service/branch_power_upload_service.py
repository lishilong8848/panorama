from __future__ import annotations

import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

import openpyxl

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from app.shared.utils.atomic_file import atomic_write_text
from pipeline_utils import get_app_dir


class BranchPowerUploadService:
    APP_TOKEN = "ASLxbfESPahdTKs0A9NccgbrnXc"
    TABLE_ID = "tblT5KbsxGCK1SwA"
    FIELD_BUILDING = "机楼"
    FIELD_ROOM = "包间"
    FIELD_ROW = "机列"
    FIELD_PDU = "PDU"

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    def _emit(self, emit_log: Callable[[str], None], text: str) -> None:
        try:
            emit_log(text)
        except Exception:  # noqa: BLE001
            pass

    @staticmethod
    def _elapsed_ms(start: float) -> int:
        return int((time.perf_counter() - start) * 1000)

    def _progress_logger(
        self,
        emit_log: Callable[[str], None],
        *,
        label: str,
        total: int,
        step: int = 1000,
    ) -> Callable[[int, int], None]:
        last_logged = {"done": 0}

        def _log(done: int, total_count: int) -> None:
            total_value = int(total_count or total or 0)
            done_value = int(done or 0)
            if done_value >= total_value or done_value - last_logged["done"] >= step:
                last_logged["done"] = done_value
                self._emit(emit_log, f"[支路功率上传] {label}进度 {done_value}/{total_value}")

        return _log

    def _client(self, emit_log: Callable[[str], None]) -> FeishuBitableClient:
        auth = require_feishu_auth_settings(self.config)
        return FeishuBitableClient(
            app_id=str(auth.get("app_id", "") or "").strip(),
            app_secret=str(auth.get("app_secret", "") or "").strip(),
            app_token=self.APP_TOKEN,
            calc_table_id=self.TABLE_ID,
            attachment_table_id=self.TABLE_ID,
            timeout=int(auth.get("timeout", 30) or 30),
            request_retry_count=int(auth.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(auth.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
            emit_log=emit_log,
        )

    def _runtime_state_root(self) -> Path:
        common = self.config.get("common", {}) if isinstance(self.config.get("common", {}), dict) else {}
        paths = common.get("paths", {}) if isinstance(common.get("paths", {}), dict) else {}
        root_text = str(paths.get("runtime_state_root", "") or "").strip()
        root = Path(root_text) if root_text else Path(".runtime")
        if not root.is_absolute():
            root = get_app_dir() / root
        root.mkdir(parents=True, exist_ok=True)
        return root

    def _state_path(self) -> Path:
        return self._runtime_state_root() / "branch_power_upload_state.json"

    def _load_state(self) -> Dict[str, Any]:
        path = self._state_path()
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            return {}
        return payload if isinstance(payload, dict) else {}

    def _save_state(self, payload: Dict[str, Any]) -> None:
        path = self._state_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(path, json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _parse_bucket_datetime(bucket_key: str) -> datetime:
        text = str(bucket_key or "").strip()
        for fmt in ("%Y-%m-%d %H", "%Y%m%d%H"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) >= 10:
            return datetime.strptime(digits[:10], "%Y%m%d%H")
        raise ValueError(f"支路功率小时桶无效: {bucket_key}")

    @classmethod
    def _resolve_upload_bucket_datetime(cls, *, bucket_key: str, source_units: List[Dict[str, Any]]) -> datetime:
        data_bucket_keys: List[str] = []
        for unit in source_units:
            if not isinstance(unit, dict):
                continue
            metadata = unit.get("metadata", {}) if isinstance(unit.get("metadata", {}), dict) else {}
            raw = str(
                unit.get("data_hour_bucket", "")
                or metadata.get("data_hour_bucket", "")
                or metadata.get("data_bucket_key", "")
                or ""
            ).strip()
            if not raw:
                continue
            parsed = cls._parse_bucket_datetime(raw)
            normalized = parsed.strftime("%Y-%m-%d %H")
            if normalized not in data_bucket_keys:
                data_bucket_keys.append(normalized)
        if len(data_bucket_keys) > 1:
            raise RuntimeError(f"支路功率源文件数据小时不一致: {', '.join(data_bucket_keys)}")
        if data_bucket_keys:
            return cls._parse_bucket_datetime(data_bucket_keys[0])
        return cls._parse_bucket_datetime(bucket_key)

    @staticmethod
    def _norm_text(value: Any) -> str:
        return str(value or "").strip()

    @classmethod
    def _record_key(cls, building: Any, room: Any, row: Any, pdu: Any) -> Tuple[str, str, str, str]:
        return (
            cls._norm_text(building),
            cls._norm_text(room),
            cls._norm_text(row),
            cls._norm_text(pdu),
        )

    @classmethod
    def _field_text(cls, value: Any) -> str:
        if isinstance(value, list):
            if not value:
                return ""
            return cls._field_text(value[0])
        if isinstance(value, dict):
            for key in ("text", "name", "value"):
                if key in value:
                    return cls._field_text(value.get(key))
            return ""
        return cls._norm_text(value)

    @classmethod
    def _choice_field_text(cls, value: Any, option_text_by_id: Dict[str, str]) -> str:
        if isinstance(value, list):
            if not value:
                return ""
            return cls._choice_field_text(value[0], option_text_by_id)
        if isinstance(value, dict):
            for key in ("text", "name", "value"):
                text = cls._norm_text(value.get(key))
                if text:
                    return option_text_by_id.get(text, text)
            for key in ("id", "option_id"):
                option_id = cls._norm_text(value.get(key))
                if option_id:
                    return option_text_by_id.get(option_id, option_id)
            return ""
        text = cls._norm_text(value)
        return option_text_by_id.get(text, text)

    def _building_option_text_by_id(self, client: FeishuBitableClient, emit_log: Callable[[str], None]) -> Dict[str, str]:
        try:
            fields = client.list_fields(self.TABLE_ID)
        except Exception as exc:  # noqa: BLE001
            self._emit(emit_log, f"[支路功率上传][告警] 读取机楼下拉选项失败，按原值匹配: {exc}")
            return {}
        for field in fields:
            if not isinstance(field, dict):
                continue
            field_name = self._norm_text(field.get("field_name") or field.get("name"))
            if field_name != self.FIELD_BUILDING:
                continue
            prop = field.get("property", {}) if isinstance(field.get("property", {}), dict) else {}
            options = prop.get("options", []) if isinstance(prop.get("options", []), list) else []
            mapping: Dict[str, str] = {}
            for option in options:
                if not isinstance(option, dict):
                    continue
                option_name = self._norm_text(option.get("name") or option.get("text") or option.get("value"))
                if not option_name:
                    continue
                for key in ("id", "option_id", "name", "text", "value"):
                    raw = self._norm_text(option.get(key))
                    if raw:
                        mapping[raw] = option_name
            return mapping
        return {}

    @staticmethod
    def _number_value(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        return float(text)

    def _load_rows(self, *, file_path: Path, building: str, hour_field: str) -> tuple[List[Dict[str, Any]], List[str]]:
        warnings: List[str] = []
        rows: List[Dict[str, Any]] = []
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook["支路功率"] if "支路功率" in workbook.sheetnames else workbook.active
            if hasattr(sheet, "reset_dimensions"):
                # FineReport exports sometimes declare the worksheet dimension as A1
                # even when the sheet XML contains thousands of rows.  Streaming
                # openpyxl trusts that stale dimension unless it is reset first.
                sheet.reset_dimensions()
            for index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                room = self._norm_text(row[0] if len(row) > 0 else "")
                machine_row = self._norm_text(row[1] if len(row) > 1 else "")
                pdu = self._norm_text(row[2] if len(row) > 2 else "")
                raw_value = row[3] if len(row) > 3 else None
                if not any([room, machine_row, pdu, self._norm_text(raw_value)]):
                    continue
                if not room or not machine_row or not pdu:
                    warnings.append(f"{building} {file_path.name} 第{index}行缺少包间/机列/PDU，已跳过")
                    continue
                try:
                    numeric_value = self._number_value(raw_value)
                except Exception:  # noqa: BLE001
                    warnings.append(f"{building} {file_path.name} 第{index}行数值不是数字: {raw_value}")
                    continue
                if numeric_value is None:
                    warnings.append(f"{building} {file_path.name} 第{index}行数值为空，已跳过")
                    continue
                rows.append(
                    {
                        self.FIELD_BUILDING: building,
                        self.FIELD_ROOM: room,
                        self.FIELD_ROW: machine_row,
                        self.FIELD_PDU: pdu,
                        hour_field: numeric_value,
                        "_source_row": index,
                    }
                )
        finally:
            workbook.close()
        return rows, warnings

    def _clear_table_if_needed(
        self,
        *,
        client: FeishuBitableClient,
        hour_field: str,
        bucket_dt: datetime,
        emit_log: Callable[[str], None],
    ) -> bool:
        if hour_field != "0:00":
            return False
        day_key = bucket_dt.strftime("%Y-%m-%d")
        state = self._load_state()
        cleared_days = state.get("cleared_days", {}) if isinstance(state.get("cleared_days", {}), dict) else {}
        if str(cleared_days.get(day_key, "") or "").strip():
            return False
        self._emit(emit_log, f"[支路功率上传] 0点数据上传前清空目标表 date={day_key}")
        deleted = client.clear_table(self.TABLE_ID)
        cleared_days[day_key] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        state["cleared_days"] = cleared_days
        self._save_state(state)
        self._emit(emit_log, f"[支路功率上传] 目标表清空完成 date={day_key}, deleted={deleted}")
        return True

    def _index_existing_records(
        self,
        records: List[Dict[str, Any]],
        *,
        building_option_text_by_id: Dict[str, str],
    ) -> Dict[Tuple[str, str, str, str], List[Dict[str, Any]]]:
        indexed: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = {}
        for record in records:
            if not isinstance(record, dict):
                continue
            fields = record.get("fields", {}) if isinstance(record.get("fields", {}), dict) else {}
            key = self._record_key(
                self._choice_field_text(fields.get(self.FIELD_BUILDING), building_option_text_by_id),
                self._field_text(fields.get(self.FIELD_ROOM)),
                self._field_text(fields.get(self.FIELD_ROW)),
                self._field_text(fields.get(self.FIELD_PDU)),
            )
            if not all(key):
                continue
            indexed.setdefault(key, []).append(record)
        return indexed

    def continue_from_source_files(
        self,
        *,
        bucket_key: str,
        source_units: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        total_start = time.perf_counter()
        units = [unit for unit in (source_units or []) if isinstance(unit, dict)]
        if not units:
            raise RuntimeError(f"没有可上传的支路功率共享文件 bucket={bucket_key}")
        storage_bucket_dt = self._parse_bucket_datetime(bucket_key)
        bucket_dt = self._resolve_upload_bucket_datetime(bucket_key=bucket_key, source_units=units)
        hour_field = f"{bucket_dt.hour}:00"
        self._emit(
            emit_log,
            f"[支路功率上传] 开始 storage_bucket={bucket_key}, data_bucket={bucket_dt.strftime('%Y-%m-%d %H')}, hour={hour_field}, files={len(units)}",
        )

        parsed_rows: List[Dict[str, Any]] = []
        warnings: List[str] = []
        for unit in units:
            building = self._norm_text(unit.get("building"))
            file_path = Path(str(unit.get("file_path", "") or unit.get("source_file", "") or "").strip())
            if not building or not file_path.exists():
                warnings.append(f"共享文件不可用 building={building or '-'}, path={file_path}")
                continue
            read_start = time.perf_counter()
            rows, row_warnings = self._load_rows(file_path=file_path, building=building, hour_field=hour_field)
            parsed_rows.extend(rows)
            warnings.extend(row_warnings)
            self._emit(
                emit_log,
                f"[支路功率上传] 读取完成 building={building}, file={file_path}, rows={len(rows)}, elapsed_ms={self._elapsed_ms(read_start)}",
            )

        if not parsed_rows:
            for warning in warnings:
                self._emit(emit_log, f"[支路功率上传][告警] {warning}")
            self._emit(emit_log, f"[支路功率上传] 失败 bucket={bucket_key}, error=共享文件没有可上传数据")
            raise RuntimeError(f"支路功率共享文件没有可上传数据 bucket={bucket_key}")

        unique_rows: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
        duplicate_source_rows = 0
        for row in parsed_rows:
            key = self._record_key(
                row.get(self.FIELD_BUILDING),
                row.get(self.FIELD_ROOM),
                row.get(self.FIELD_ROW),
                row.get(self.FIELD_PDU),
            )
            if key in unique_rows:
                duplicate_source_rows += 1
                warnings.append(f"源文件存在重复支路 key={key}，已使用最后一条数值")
            unique_rows[key] = row
        upload_rows = list(unique_rows.values())
        self._emit(
            emit_log,
            f"[支路功率上传] 解析汇总 parsed={len(parsed_rows)}, unique={len(upload_rows)}, source_duplicates={duplicate_source_rows}, warnings={len(warnings)}",
        )

        client_start = time.perf_counter()
        self._emit(emit_log, f"[支路功率上传] 飞书客户端初始化开始 app_token={self.APP_TOKEN}, table_id={self.TABLE_ID}")
        client = self._client(emit_log)
        self._emit(emit_log, f"[支路功率上传] 飞书客户端初始化完成 elapsed_ms={self._elapsed_ms(client_start)}")
        cleared = self._clear_table_if_needed(
            client=client,
            hour_field=hour_field,
            bucket_dt=bucket_dt,
            emit_log=emit_log,
        )
        if cleared:
            existing_records = []
            self._emit(emit_log, "[支路功率上传] 已清空目标表，跳过读取现有记录")
        else:
            list_start = time.perf_counter()
            self._emit(emit_log, f"[支路功率上传] 读取目标表现有记录开始 table_id={self.TABLE_ID}")
            existing_records = client.list_records(self.TABLE_ID, page_size=500, max_records=0)
            self._emit(
                emit_log,
                f"[支路功率上传] 读取目标表现有记录完成 records={len(existing_records)}, elapsed_ms={self._elapsed_ms(list_start)}",
            )
        field_start = time.perf_counter()
        building_option_text_by_id = self._building_option_text_by_id(client, emit_log)
        self._emit(
            emit_log,
            f"[支路功率上传] 读取机楼选项完成 options={len(building_option_text_by_id)}, elapsed_ms={self._elapsed_ms(field_start)}",
        )
        index_start = time.perf_counter()
        existing_index = self._index_existing_records(
            existing_records,
            building_option_text_by_id=building_option_text_by_id,
        )
        self._emit(
            emit_log,
            f"[支路功率上传] 目标表索引构建完成 keys={len(existing_index)}, elapsed_ms={self._elapsed_ms(index_start)}",
        )

        create_fields: List[Dict[str, Any]] = []
        update_records: List[Dict[str, Any]] = []
        duplicate_count = 0
        for row in upload_rows:
            key = self._record_key(
                row.get(self.FIELD_BUILDING),
                row.get(self.FIELD_ROOM),
                row.get(self.FIELD_ROW),
                row.get(self.FIELD_PDU),
            )
            matches = existing_index.get(key, [])
            fields = {
                self.FIELD_BUILDING: row[self.FIELD_BUILDING],
                self.FIELD_ROOM: row[self.FIELD_ROOM],
                self.FIELD_ROW: row[self.FIELD_ROW],
                self.FIELD_PDU: row[self.FIELD_PDU],
                hour_field: row[hour_field],
            }
            if matches:
                if len(matches) > 1:
                    duplicate_count += len(matches) - 1
                    warnings.append(f"目标表存在重复记录 key={key}，已更新第一条")
                update_records.append({"record_id": str(matches[0].get("record_id", "") or "").strip(), "fields": fields})
            else:
                create_fields.append(fields)
        self._emit(
            emit_log,
            f"[支路功率上传] 写入计划 ready update={len(update_records)}, create={len(create_fields)}, target_duplicates={duplicate_count}",
        )

        created = 0
        updated = 0
        if update_records:
            update_start = time.perf_counter()
            self._emit(emit_log, f"[支路功率上传] 批量更新开始 count={len(update_records)}, hour={hour_field}")
            client.batch_update_records(
                self.TABLE_ID,
                update_records,
                progress_callback=self._progress_logger(emit_log, label="批量更新", total=len(update_records)),
            )
            updated = len(update_records)
            self._emit(emit_log, f"[支路功率上传] 批量更新完成 count={updated}, elapsed_ms={self._elapsed_ms(update_start)}")
        if create_fields:
            create_start = time.perf_counter()
            self._emit(emit_log, f"[支路功率上传] 批量创建开始 count={len(create_fields)}, hour={hour_field}")
            client.batch_create_records(
                self.TABLE_ID,
                create_fields,
                progress_callback=self._progress_logger(emit_log, label="批量创建", total=len(create_fields)),
            )
            created = len(create_fields)
            self._emit(emit_log, f"[支路功率上传] 批量创建完成 count={created}, elapsed_ms={self._elapsed_ms(create_start)}")
        self._emit(
            emit_log,
            f"[支路功率上传] 完成 storage_bucket={bucket_key}, data_bucket={bucket_dt.strftime('%Y-%m-%d %H')}, hour={hour_field}, parsed={len(parsed_rows)}, unique={len(upload_rows)}, updated={updated}, created={created}, target_duplicates={duplicate_count}, source_duplicates={duplicate_source_rows}, elapsed_ms={self._elapsed_ms(total_start)}",
        )
        for warning in warnings:
            self._emit(emit_log, f"[支路功率上传][告警] {warning}")
        return {
            "ok": True,
            "bucket_key": bucket_key,
            "storage_bucket_key": storage_bucket_dt.strftime("%Y-%m-%d %H"),
            "data_bucket_key": bucket_dt.strftime("%Y-%m-%d %H"),
            "hour_field": hour_field,
            "parsed": len(parsed_rows),
            "unique": len(upload_rows),
            "updated": updated,
            "created": created,
            "duplicate_count": duplicate_count,
            "source_duplicate_count": duplicate_source_rows,
            "warnings": warnings,
        }
