from __future__ import annotations

import copy
import math
import time
from collections import OrderedDict
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Dict, List, Tuple

from openpyxl import load_workbook

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.internal_bridge_http.service.client import InternalBridgeHttpClient


SOURCE_FAMILY = "air_conditioner_temperature_humidity_family"
DEFAULT_BUILDINGS = ["A楼", "B楼", "C楼", "D楼", "E楼"]
DEFAULT_TARGET = {
    "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
    "table_id": "tblfnTbEWK9607zV",
    "page_size": 500,
    "max_records": 0,
    "delete_batch_size": 200,
    "create_batch_size": 200,
    "max_upload_records": 10000,
    "fields": {
        "building": "楼栋",
        "position": "位置",
        "temperature": "温度",
        "humidity": "湿度",
        "running_status": "运行状态",
    },
}
DEFAULT_CONFIG = {
    "enabled": True,
    "source_family": SOURCE_FAMILY,
    "buildings": DEFAULT_BUILDINGS,
    "trigger_missing_download": True,
    "wait_source_timeout_sec": 600,
    "wait_source_poll_sec": 5,
    "target": DEFAULT_TARGET,
}
EXPECTED_TARGET_FIELD_TYPES = {
    "building": 3,
    "position": 1,
    "temperature": 2,
    "humidity": 2,
    "running_status": 3,
}


def _dict(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _list(value: Any) -> List[Any]:
    return value if isinstance(value, list) else []


def _deep_merge(raw: Any, defaults: Any) -> Any:
    if isinstance(defaults, dict):
        source = raw if isinstance(raw, dict) else {}
        output: Dict[str, Any] = {}
        for key, default_value in defaults.items():
            output[key] = _deep_merge(source.get(key), default_value)
        for key, value in source.items():
            if key not in output:
                output[key] = copy.deepcopy(value)
        return output
    if isinstance(defaults, list):
        return copy.deepcopy(raw) if isinstance(raw, list) else copy.deepcopy(defaults)
    return copy.deepcopy(defaults if raw is None else raw)


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    for suffix in ("℃", "%", "°C", "C"):
        if text.endswith(suffix):
            text = text[: -len(suffix)].strip()
            break
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _running_status(value: Any) -> str:
    number = _number(value)
    if number == 1:
        return "开启"
    if number == 0:
        return "关闭"
    return ""


class TemperatureHumidityUploadService:
    def __init__(
        self,
        runtime_config: Dict[str, Any],
        *,
        internal_client: InternalBridgeHttpClient | None = None,
        bitable_client: FeishuBitableClient | None = None,
    ) -> None:
        self.runtime_config = runtime_config if isinstance(runtime_config, dict) else {}
        self._internal_client = internal_client
        self._bitable_client = bitable_client

    @staticmethod
    def normalize_source_date(value: Any | None = None) -> str:
        text = str(value or "").strip()
        if not text:
            return datetime.now().strftime("%Y-%m-%d")
        for fmt in ("%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError("源文件日期必须为 YYYY-MM-DD")

    def _config(self) -> Dict[str, Any]:
        raw = _dict(self.runtime_config.get("temperature_humidity_upload"))
        return _deep_merge(raw, DEFAULT_CONFIG)

    @staticmethod
    def _target(cfg: Dict[str, Any]) -> Dict[str, Any]:
        return _deep_merge(_dict(cfg.get("target")), DEFAULT_TARGET)

    @staticmethod
    def _buildings(cfg: Dict[str, Any]) -> List[str]:
        output: List[str] = []
        for item in _list(cfg.get("buildings")):
            building = _text(item)
            if building and building not in output:
                output.append(building)
        return output or list(DEFAULT_BUILDINGS)

    @staticmethod
    def _target_fields(target: Dict[str, Any]) -> Dict[str, str]:
        raw = _dict(target.get("fields"))
        output = {
            key: _text(raw.get(key, default_name))
            for key, default_name in DEFAULT_TARGET["fields"].items()
        }
        missing = [key for key, value in output.items() if not value]
        if missing:
            raise ValueError(f"温湿度上传目标字段配置不完整: {','.join(missing)}")
        return output

    def _make_internal_client(self) -> InternalBridgeHttpClient:
        if self._internal_client is not None:
            return self._internal_client
        client = InternalBridgeHttpClient.from_runtime_config(self.runtime_config)
        if client is None:
            raise RuntimeError("内网端 HTTP 桥接 base_url 未配置，无法读取空调温湿度源文件")
        return client

    def _make_bitable_client(self, target: Dict[str, Any]) -> FeishuBitableClient:
        if self._bitable_client is not None:
            return self._bitable_client
        feishu = _dict(self.runtime_config.get("feishu"))
        app_token = _text(target.get("app_token"))
        table_id = _text(target.get("table_id"))
        if not app_token or not table_id:
            raise ValueError("温湿度上传目标 app_token/table_id 未配置")
        return FeishuBitableClient(
            app_id=_text(feishu.get("app_id")),
            app_secret=_text(feishu.get("app_secret")),
            app_token=app_token,
            calc_table_id=table_id,
            attachment_table_id=table_id,
            timeout=int(feishu.get("timeout", 30) or 30),
            request_retry_count=int(feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=lambda **_: 0,
            canonical_metric_name_fn=lambda item: _text(item),
            dimension_mapping={},
        )

    @staticmethod
    def _entry_sort_key(item: Dict[str, Any]) -> Tuple[str, str, str]:
        return (
            _text(item.get("downloaded_at")),
            _text(item.get("updated_at")),
            _text(item.get("entry_id")),
        )

    def _external_shared_root(self) -> str:
        shared_bridge = _dict(self.runtime_config.get("shared_bridge"))
        if not shared_bridge:
            shared_bridge = _dict(_dict(self.runtime_config.get("common")).get("shared_bridge"))
        return _text(
            shared_bridge.get("external_root_dir")
            or shared_bridge.get("root_dir")
        )

    @staticmethod
    def _safe_relative_path(value: Any) -> PurePosixPath | None:
        text = _text(value).replace("\\", "/").strip("/")
        if not text:
            return None
        candidate = PurePosixPath(text)
        if candidate.is_absolute() or ".." in candidate.parts or ":" in text:
            return None
        return candidate

    def _entry_file(self, item: Dict[str, Any]) -> Path | None:
        candidate: Path | None = None
        relative_path = self._safe_relative_path(item.get("relative_path"))
        shared_root = self._external_shared_root()
        if relative_path is not None and shared_root:
            candidate = Path(shared_root).joinpath(*relative_path.parts)
        if candidate is None:
            file_path = _text(item.get("file_path") or item.get("resolved_file_path"))
            if not file_path:
                return None
            candidate = Path(file_path)
        if candidate.suffix.lower() not in {".xlsx", ".xlsm"}:
            return None
        try:
            if not candidate.is_file() or candidate.stat().st_size <= 0:
                return None
        except OSError:
            return None
        return candidate

    @staticmethod
    def _entry_matches_source(
        item: Dict[str, Any],
        *,
        source_family: str,
        source_date: str,
        building: str,
    ) -> bool:
        metadata = _dict(item.get("metadata"))
        item_family = _text(item.get("source_family") or metadata.get("family"))
        item_building = _text(item.get("building") or metadata.get("building"))
        bucket_kind = _text(item.get("bucket_kind")).lower()
        item_dates = {
            _text(item.get("bucket_key")),
            _text(item.get("duty_date")),
            _text(metadata.get("business_date")),
            _text(metadata.get("data_day")),
        }
        return (
            item_family == source_family
            and item_building == building
            and bucket_kind in {"daily", "day"}
            and source_date in item_dates
        )

    def _query_source_files(
        self,
        internal: InternalBridgeHttpClient,
        *,
        source_family: str,
        source_date: str,
        buildings: List[str],
    ) -> Tuple[Dict[str, Path], List[str], List[str]]:
        queries = [
            {
                "source_family": source_family,
                "bucket_or_date": source_date,
                "building": building,
                "bucket_kind": "daily",
                "status": "ready",
                "limit": 20,
            }
            for building in buildings
        ]
        results = internal.source_index_batch(queries, default_limit=20)
        selected: Dict[str, Path] = {}
        errors: List[str] = []
        for index, building in enumerate(buildings):
            result = results[index] if index < len(results) and isinstance(results[index], dict) else {}
            if not bool(result.get("ok", False)):
                errors.append(f"{building}: {_text(result.get('error')) or 'source-index 查询失败'}")
                continue
            entries = result.get("entries", []) if isinstance(result.get("entries", []), list) else []
            candidates: List[Tuple[Tuple[str, str, str], Path]] = []
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                if not self._entry_matches_source(
                    entry,
                    source_family=source_family,
                    source_date=source_date,
                    building=building,
                ):
                    continue
                if _text(entry.get("status")).lower() != "ready":
                    continue
                if entry.get("file_verified") is False:
                    continue
                path = self._entry_file(entry)
                if path is not None:
                    candidates.append((self._entry_sort_key(entry), path))
            if candidates:
                candidates.sort(key=lambda item: item[0], reverse=True)
                selected[building] = candidates[0][1]
        missing = [building for building in buildings if building not in selected]
        return selected, missing, errors

    def _wait_for_source_files(
        self,
        internal: InternalBridgeHttpClient,
        *,
        cfg: Dict[str, Any],
        source_date: str,
        buildings: List[str],
        log: Callable[[str], None],
    ) -> Dict[str, Path]:
        source_family = _text(cfg.get("source_family")) or SOURCE_FAMILY
        selected, missing, errors = self._query_source_files(
            internal,
            source_family=source_family,
            source_date=source_date,
            buildings=buildings,
        )
        reported_errors: set[str] = set()
        for error in errors:
            if error not in reported_errors:
                reported_errors.add(error)
                log(f"[空调温湿度上传] source-index 查询异常: {error}")
        if not missing:
            return selected

        timeout_sec = max(0.0, float(cfg.get("wait_source_timeout_sec", 600) or 0))
        poll_sec = max(1.0, float(cfg.get("wait_source_poll_sec", 5) or 5))
        is_current_date = source_date == datetime.now().strftime("%Y-%m-%d")
        if bool(cfg.get("trigger_missing_download", True)) and is_current_date:
            try:
                refresh = internal.refresh_latest_source_cache(
                    source_family=source_family,
                    buildings=missing,
                    target_bucket_key=source_date,
                )
                log(
                    "[空调温湿度上传] 已请求内网端补采 "
                    f"date={source_date}, buildings={','.join(missing)}, "
                    f"accepted={int(refresh.get('accepted_count', 0) or 0)}"
                )
            except Exception as exc:  # noqa: BLE001
                log(f"[空调温湿度上传] 内网端补采请求失败，继续等待索引恢复: {exc}")
        elif not is_current_date:
            timeout_sec = min(timeout_sec, 30.0)
            log(
                "[空调温湿度上传] 历史日期只等待已有文件索引恢复，不触发当前时间窗口补采 "
                f"date={source_date}, missing={','.join(missing)}"
            )

        deadline = time.monotonic() + timeout_sec
        while missing and time.monotonic() < deadline:
            time.sleep(min(poll_sec, max(0.1, deadline - time.monotonic())))
            selected, missing, errors = self._query_source_files(
                internal,
                source_family=source_family,
                source_date=source_date,
                buildings=buildings,
            )
            for error in errors:
                if error not in reported_errors:
                    reported_errors.add(error)
                    log(f"[空调温湿度上传] source-index 查询异常: {error}")
        if missing:
            raise RuntimeError(
                f"{source_date} 空调温湿度源文件缺失或外网端不可访问: {','.join(missing)}"
            )
        return selected

    @staticmethod
    def parse_workbook(file_path: Path | str, *, building: str) -> Dict[str, Any]:
        source_path = Path(file_path)
        workbook = load_workbook(source_path, data_only=True, read_only=True)
        try:
            if not workbook.sheetnames:
                raise ValueError(f"{building} 空调温湿度源文件没有工作表")
            sheet = workbook[workbook.sheetnames[0]]
            reset_dimensions = getattr(sheet, "reset_dimensions", None)
            if callable(reset_dimensions):
                reset_dimensions()

            temperature_rows: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
            status_rows: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
            current_temperature_position = ""
            for row_index, values in enumerate(
                sheet.iter_rows(min_col=3, max_col=5, values_only=True),
                start=1,
            ):
                column_c = _text(values[0] if len(values) > 0 else None)
                column_d = _text(values[1] if len(values) > 1 else None)
                column_e = values[2] if len(values) > 2 else None
                if not column_d:
                    continue
                measurement_key = ""
                if "温度" in column_d:
                    measurement_key = "temperature"
                elif "湿度" in column_d:
                    measurement_key = "humidity"
                if measurement_key:
                    if column_c:
                        current_temperature_position = column_c
                    if not current_temperature_position:
                        raise ValueError(
                            f"{building} 空调温湿度源文件第{row_index}行缺少位置，无法关联合并单元格"
                        )
                    number = _number(column_e)
                    if number is None:
                        raise ValueError(
                            f"{building} 空调温湿度源文件第{row_index}行{column_d}不是有效数字"
                        )
                    item = temperature_rows.setdefault(
                        current_temperature_position,
                        {"building": building, "position": current_temperature_position},
                    )
                    item[measurement_key] = number
                    continue

                # 运行状态行结束当前温湿度合并单元格分组，避免后续缺位置行误沿用旧位置。
                current_temperature_position = ""
                status = _running_status(column_e)
                if not status:
                    raw_status = _text(column_e) or "空"
                    raise ValueError(
                        f"{building} 空调温湿度源文件第{row_index}行运行状态必须为0或1，"
                        f"实际为“{raw_status}”"
                    )
                status_rows[column_d] = {
                    "building": building,
                    "position": column_d,
                    "running_status": status,
                }

            incomplete = [
                position
                for position, item in temperature_rows.items()
                if "temperature" not in item or "humidity" not in item
            ]
            if incomplete:
                preview = ",".join(incomplete[:5])
                raise ValueError(
                    f"{building} 空调温湿度源文件存在未成对温湿度位置: {preview}"
                )
            rows = [*temperature_rows.values(), *status_rows.values()]
            if not rows:
                raise ValueError(f"{building} 空调温湿度源文件未解析到可上传数据")
            return {
                "building": building,
                "file_path": str(source_path),
                "sheet_name": sheet.title,
                "rows": rows,
                "temperature_location_count": len(temperature_rows),
                "status_location_count": len(status_rows),
                "unparsed_status_count": 0,
            }
        finally:
            workbook.close()

    @staticmethod
    def _validate_target_fields(
        client: FeishuBitableClient,
        *,
        table_id: str,
        field_names: Dict[str, str],
    ) -> None:
        definitions = client.list_fields(table_id, page_size=200)
        by_name = {
            _text(item.get("field_name")): item
            for item in definitions
            if isinstance(item, dict) and _text(item.get("field_name"))
        }
        errors: List[str] = []
        for key, expected_type in EXPECTED_TARGET_FIELD_TYPES.items():
            field_name = field_names[key]
            definition = by_name.get(field_name)
            if definition is None:
                errors.append(f"缺少字段“{field_name}”")
                continue
            try:
                actual_type = int(definition.get("type", -1))
            except (TypeError, ValueError):
                actual_type = -1
            if actual_type != expected_type:
                errors.append(f"字段“{field_name}”类型应为{expected_type}，实际为{actual_type}")
        if errors:
            raise ValueError("温湿度目标多维表字段校验失败: " + "；".join(errors))

    @staticmethod
    def _build_fields(rows: List[Dict[str, Any]], field_names: Dict[str, str]) -> List[Dict[str, Any]]:
        output: List[Dict[str, Any]] = []
        for item in rows:
            fields: Dict[str, Any] = {
                field_names["building"]: item["building"],
                field_names["position"]: item["position"],
            }
            if "temperature" in item:
                fields[field_names["temperature"]] = float(item["temperature"])
            if "humidity" in item:
                fields[field_names["humidity"]] = float(item["humidity"])
            if _text(item.get("running_status")):
                fields[field_names["running_status"]] = _text(item.get("running_status"))
            output.append(fields)
        return output

    @staticmethod
    def _record_ids(records: List[Dict[str, Any]]) -> List[str]:
        return [
            _text(item.get("record_id"))
            for item in records
            if isinstance(item, dict) and _text(item.get("record_id"))
        ]

    @staticmethod
    def _record_field_snapshot(
        records: List[Dict[str, Any]],
        *,
        allowed_field_names: List[str],
    ) -> List[Dict[str, Any]]:
        allowed = {name for name in allowed_field_names if _text(name)}
        snapshot: List[Dict[str, Any]] = []
        for item in records:
            if not isinstance(item, dict):
                continue
            raw_fields = item.get("fields")
            if not isinstance(raw_fields, dict):
                continue
            fields = {
                str(key): copy.deepcopy(value)
                for key, value in raw_fields.items()
                if str(key) in allowed
            }
            if fields:
                snapshot.append(fields)
        return snapshot

    @staticmethod
    def _restore_table_snapshot(
        client: FeishuBitableClient,
        *,
        table_id: str,
        snapshot_fields: List[Dict[str, Any]],
        page_size: int,
        max_records: int,
        delete_batch_size: int,
        create_batch_size: int,
        field_names: List[str],
    ) -> Dict[str, Any]:
        current_records = client.list_records(
            table_id,
            page_size=page_size,
            max_records=max_records,
            field_names=field_names,
        )
        current_ids = TemperatureHumidityUploadService._record_ids(current_records)
        removed_count = 0
        if current_ids:
            removed_count = client.batch_delete_records(
                table_id,
                current_ids,
                batch_size=delete_batch_size,
            )
        restored_count = 0
        if snapshot_fields:
            client.batch_create_records(
                table_id,
                snapshot_fields,
                batch_size=create_batch_size,
            )
            restored_count = len(snapshot_fields)
        return {
            "removed_partial_count": removed_count,
            "restored_count": restored_count,
        }

    def run(
        self,
        *,
        source_date: Any | None = None,
        emit_log: Callable[[str], None] | None = None,
    ) -> Dict[str, Any]:
        log = emit_log if callable(emit_log) else (lambda _message: None)
        cfg = self._config()
        if not bool(cfg.get("enabled", True)):
            return {"status": "skipped", "reason": "disabled"}

        date_text = self.normalize_source_date(source_date)
        buildings = self._buildings(cfg)
        target = self._target(cfg)
        app_token = _text(target.get("app_token"))
        table_id = _text(target.get("table_id"))
        if not app_token or not table_id:
            raise ValueError("温湿度上传目标 app_token/table_id 未配置")

        internal = self._make_internal_client()
        log(
            "[空调温湿度上传] 开始读取内网端 source-index "
            f"date={date_text}, family={_text(cfg.get('source_family')) or SOURCE_FAMILY}, "
            f"buildings={','.join(buildings)}"
        )
        source_files = self._wait_for_source_files(
            internal,
            cfg=cfg,
            source_date=date_text,
            buildings=buildings,
            log=log,
        )

        parsed_results: List[Dict[str, Any]] = []
        all_rows: List[Dict[str, Any]] = []
        for building in buildings:
            parsed = self.parse_workbook(source_files[building], building=building)
            parsed_results.append(parsed)
            all_rows.extend(parsed["rows"])
            log(
                "[空调温湿度上传] 源文件解析完成 "
                f"building={building}, temperature={parsed['temperature_location_count']}, "
                f"status={parsed['status_location_count']}, unparsed_status={parsed['unparsed_status_count']}"
            )

        max_upload_records = max(1, int(target.get("max_upload_records", 10000) or 10000))
        if len(all_rows) > max_upload_records:
            raise ValueError(
                f"温湿度待上传记录数 {len(all_rows)} 超过安全上限 {max_upload_records}"
            )

        bitable = self._make_bitable_client(target)
        field_names = self._target_fields(target)
        self._validate_target_fields(bitable, table_id=table_id, field_names=field_names)
        old_records = bitable.list_records(
            table_id,
            page_size=max(1, int(target.get("page_size", 500) or 500)),
            max_records=max(0, int(target.get("max_records", 0) or 0)),
            field_names=list(field_names.values()),
        )
        old_record_ids = self._record_ids(old_records)
        old_snapshot_fields = self._record_field_snapshot(
            old_records,
            allowed_field_names=list(field_names.values()),
        )

        fields_list = self._build_fields(all_rows, field_names)
        page_size = max(1, int(target.get("page_size", 500) or 500))
        max_records = max(0, int(target.get("max_records", 0) or 0))
        delete_batch_size = max(1, int(target.get("delete_batch_size", 200) or 200))
        create_batch_size = max(1, int(target.get("create_batch_size", 200) or 200))
        deleted_count = 0
        try:
            if old_record_ids:
                deleted_count = bitable.batch_delete_records(
                    table_id,
                    old_record_ids,
                    batch_size=delete_batch_size,
                )
                if deleted_count != len(old_record_ids):
                    raise RuntimeError(
                        "目标表清空结果异常: "
                        f"计划删除={len(old_record_ids)}, 实际删除={deleted_count}"
                    )
            bitable.batch_create_records(
                table_id,
                fields_list,
                batch_size=create_batch_size,
            )
        except Exception as exc:  # noqa: BLE001
            rollback_text = ""
            try:
                rollback = self._restore_table_snapshot(
                    bitable,
                    table_id=table_id,
                    snapshot_fields=old_snapshot_fields,
                    page_size=page_size,
                    max_records=max_records,
                    delete_batch_size=delete_batch_size,
                    create_batch_size=create_batch_size,
                    field_names=list(field_names.values()),
                )
                rollback_text = (
                    "；旧数据已回滚"
                    f"(恢复={rollback['restored_count']}, 清理半成品={rollback['removed_partial_count']})"
                )
                log(f"[空调温湿度上传] 新旧数据替换失败，旧数据已回滚: {exc}")
            except Exception as rollback_exc:  # noqa: BLE001
                rollback_text = f"；旧数据回滚失败: {rollback_exc}"
                log(
                    "[空调温湿度上传] 新旧数据替换失败且旧数据回滚失败: "
                    f"replace_error={exc}, rollback_error={rollback_exc}"
                )
            raise RuntimeError(f"温湿度目标表清空并上传失败: {exc}{rollback_text}") from exc

        temperature_count = sum(int(item["temperature_location_count"]) for item in parsed_results)
        status_count = sum(int(item["status_location_count"]) for item in parsed_results)
        unparsed_status_count = sum(int(item["unparsed_status_count"]) for item in parsed_results)
        log(
            "[空调温湿度上传] 上传完成 "
            f"date={date_text}, uploaded={len(fields_list)}, deleted={deleted_count}, "
            f"temperature={temperature_count}, status={status_count}"
        )
        result: Dict[str, Any] = {
            "status": "success",
            "source_date": date_text,
            "app_token": app_token,
            "table_id": table_id,
            "uploaded_count": len(fields_list),
            "deleted_count": deleted_count,
            "temperature_location_count": temperature_count,
            "status_location_count": status_count,
            "unparsed_status_count": unparsed_status_count,
            "buildings": buildings,
            "source_files": {
                item["building"]: item["file_path"]
                for item in parsed_results
            },
        }
        return result
