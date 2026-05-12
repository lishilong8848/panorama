from __future__ import annotations

from dataclasses import dataclass
import re
from typing import Dict, List

from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class FooterInventoryLayout:
    title_row: int
    header_row: int
    data_start_row: int
    data_end_row: int
    signoff_start_row: int | None
    last_row: int


FOOTER_TITLE_TEXT = "交接确认"
FOOTER_GROUP_TITLE_TEXT = "工具及物品交接清点"
FOOTER_SIGNOFF_MARKER = "双方班组签字确认清楚知道以上信息"
FOOTER_DYNAMIC_FIXED_CELLS = {"H52", "H53", "H54", "H55"}
FOOTER_INVENTORY_COLUMNS: List[Dict[str, object]] = [
    {"key": "B", "label": "交接工具名称", "source_cols": ["B"], "span": 1},
    {"key": "C", "label": "存放位置", "source_cols": ["C", "D"], "span": 2},
    {"key": "E", "label": "数量", "source_cols": ["E"], "span": 1},
    {"key": "F", "label": "是否存在损坏", "source_cols": ["F"], "span": 1},
    {"key": "G", "label": "其他补充说明", "source_cols": ["G"], "span": 1},
    {"key": "H", "label": "清点确认人（接班）", "source_cols": ["H"], "span": 1},
]


def first_person_text(raw_people: object, split_regex: str = r"[、,/，；;\s]+") -> str:
    people = str(raw_people or "").strip()
    if not people:
        return ""
    try:
        parts = re.split(split_regex, people)
    except Exception:  # noqa: BLE001
        parts = re.split(r"[、,/，；;\s]+", people)
    for part in parts:
        text = str(part or "").strip()
        if text:
            return text
    return ""


def cell_text(ws: Worksheet, row_idx: int, col_idx: int = 1) -> str:
    value = ws.cell(row=row_idx, column=col_idx).value
    return str(value or "").strip()


def row_texts(ws: Worksheet, row_idx: int, *, max_col: int = 9) -> List[str]:
    return [cell_text(ws, row_idx, col_idx) for col_idx in range(1, max_col + 1)]


def _row_has_text(texts: List[str], expected: str) -> bool:
    needle = str(expected or "").strip()
    if not needle:
        return False
    return any(needle in str(text or "").strip() for text in texts)


def _is_signoff_row(texts: List[str], *, signoff_marker: str) -> bool:
    if _row_has_text(texts, signoff_marker):
        return True
    return _row_has_text(texts, "交班值班长签字") and _row_has_text(texts, "接班值班长签字")


def _row_has_inventory_content(ws: Worksheet, row_idx: int) -> bool:
    for col_idx in (2, 3, 5, 6, 7, 8):
        if cell_text(ws, row_idx, col_idx):
            return True
    return False


def _remove_merge_range(ws: Worksheet, merged) -> None:
    try:
        ws.unmerge_cells(str(merged))
    except Exception:  # noqa: BLE001
        try:
            ws.merged_cells.ranges.remove(merged)
        except Exception:  # noqa: BLE001
            pass


def find_footer_inventory_layout(
    ws: Worksheet,
    *,
    title_text: str = FOOTER_TITLE_TEXT,
    signoff_marker: str = FOOTER_SIGNOFF_MARKER,
) -> FooterInventoryLayout | None:
    max_row = int(ws.max_row or 0)
    title_row = None
    for row_idx in range(1, max_row + 1):
        if cell_text(ws, row_idx, 1) == title_text:
            title_row = row_idx
            break
    if title_row is None:
        return None

    signoff_start_row = None
    last_row = title_row
    group_row = None

    for row_idx in range(title_row, max_row + 1):
        texts = row_texts(ws, row_idx)
        if any(texts):
            last_row = row_idx
        if group_row is None and row_idx > title_row and _row_has_text(texts, FOOTER_GROUP_TITLE_TEXT):
            group_row = row_idx
        if signoff_start_row is None and row_idx > title_row and _is_signoff_row(texts, signoff_marker=signoff_marker):
            signoff_start_row = row_idx

    header_row = group_row if group_row is not None else title_row + 1
    scan_end = (signoff_start_row - 1) if signoff_start_row else last_row
    data_start_row = None
    data_end_row = None
    for row_idx in range(header_row + 1, scan_end + 1):
        texts = row_texts(ws, row_idx)
        if _is_signoff_row(texts, signoff_marker=signoff_marker):
            break
        if _row_has_text(texts, FOOTER_GROUP_TITLE_TEXT):
            continue
        if not _row_has_inventory_content(ws, row_idx):
            continue
        if data_start_row is None:
            data_start_row = row_idx
        data_end_row = row_idx

    if data_start_row is None:
        data_start_row = header_row + 1
    if data_end_row is None:
        data_end_row = data_start_row

    if signoff_start_row is not None:
        last_row = max(last_row, signoff_start_row)
        for row_idx in range(signoff_start_row + 1, max_row + 1):
            texts = row_texts(ws, row_idx)
            if any(texts):
                last_row = row_idx
                continue
            break

    return FooterInventoryLayout(
        title_row=title_row,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        signoff_start_row=signoff_start_row,
        last_row=max(last_row, data_end_row),
    )


def trim_rows_below_footer(
    ws: Worksheet,
    layout: FooterInventoryLayout,
) -> int:
    max_row = int(ws.max_row or 0)
    cut_from = int(layout.last_row) + 1
    if cut_from > max_row:
        return 0

    delete_count = max_row - int(layout.last_row)
    if delete_count <= 0:
        return 0

    for merged in list(ws.merged_cells.ranges):
        if merged.max_row >= cut_from:
            _remove_merge_range(ws, merged)

    ws.delete_rows(cut_from, amount=delete_count)

    for key in list(ws._cells.keys()):  # noqa: SLF001
        row_idx, _ = key
        if row_idx > int(layout.last_row):
            del ws._cells[key]  # noqa: SLF001

    for row_idx in list(ws.row_dimensions.keys()):
        if row_idx > int(layout.last_row):
            del ws.row_dimensions[row_idx]

    for merged in list(ws.merged_cells.ranges):
        if merged.max_row > int(layout.last_row):
            _remove_merge_range(ws, merged)

    return delete_count
