from __future__ import annotations

from copy import copy
import re
from typing import Any, Callable, Dict, List, Tuple

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from handover_log_module.core.section_layout import (
    CategorySection,
    CategorySectionSnapshot,
    parse_category_sections,
)


def _extract_cells_map(row_payload: Any) -> Dict[str, Any]:
    if isinstance(row_payload, dict) and isinstance(row_payload.get("cells"), dict):
        source = row_payload.get("cells", {})
    elif isinstance(row_payload, dict):
        source = row_payload
    else:
        source = {}

    out: Dict[str, Any] = {}
    for raw_key, raw_value in source.items():
        key = str(raw_key or "").strip().upper()
        if not re.fullmatch(r"[B-I]", key):
            continue
        out[key] = "" if raw_value is None else raw_value
    return out


def _normalize_payload_rows(value: Any) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [
        row
        for row in (_extract_cells_map(item) for item in value)
        if any(_is_meaningful_payload_value(cell_value) for cell_value in row.values())
    ]


def _is_meaningful_payload_value(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(text and text != "/")


def _cleanup_row_merged_cells(ws: Worksheet, row_idx: int) -> None:
    for key, cell in list(ws._cells.items()):  # noqa: SLF001
        row, _ = key
        if row == row_idx and isinstance(cell, MergedCell):
            try:
                del ws._cells[key]  # noqa: SLF001
            except Exception:  # noqa: BLE001
                continue


def _clear_row_merges(
    ws: Worksheet,
    row_idx: int,
    *,
    emit_log: Callable[[str], None],
) -> None:
    to_remove = [rng for rng in ws.merged_cells.ranges if rng.min_row == row_idx and rng.max_row == row_idx]
    for merged in to_remove:
        merged_ref = str(merged)
        try:
            ws.unmerge_cells(merged_ref)
        except KeyError as exc:
            try:
                ws.merged_cells.ranges.remove(merged)
            except Exception:  # noqa: BLE001
                pass
            emit_log(
                f"[交接班][分类写入] 合并恢复容错: row={row_idx}, range={merged_ref}, err={exc}"
            )
        except Exception as exc:  # noqa: BLE001
            emit_log(
                f"[交接班][分类写入] 合并恢复容错: row={row_idx}, range={merged_ref}, err={exc}"
            )
    _cleanup_row_merged_cells(ws, row_idx)


def _unmerge_all_safely(ws: Worksheet, emit_log: Callable[[str], None]) -> None:
    for merged in list(ws.merged_cells.ranges):
        merged_ref = str(merged)
        try:
            ws.unmerge_cells(merged_ref)
        except KeyError as exc:
            try:
                ws.merged_cells.ranges.remove(merged)
            except Exception:  # noqa: BLE001
                pass
            emit_log(f"[交接班][插行] 合并清理容错: range={merged_ref}, err={exc}")
        except Exception as exc:  # noqa: BLE001
            emit_log(f"[交接班][插行] 合并清理容错: range={merged_ref}, err={exc}")


def _ranges_overlap(a: Tuple[int, int, int, int], b: Tuple[int, int, int, int]) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = a
    b_min_col, b_min_row, b_max_col, b_max_row = b
    row_overlap = not (a_max_row < b_min_row or b_max_row < a_min_row)
    col_overlap = not (a_max_col < b_min_col or b_max_col < a_min_col)
    return row_overlap and col_overlap


def _filter_non_overlapping_merges(
    merged_defs: List[Tuple[int, int, int, int]],
) -> Tuple[List[Tuple[int, int, int, int]], int]:
    ordered = sorted(set(merged_defs), key=lambda x: (x[1], x[0], x[3], x[2]))
    accepted: List[Tuple[int, int, int, int]] = []
    skipped = 0
    for item in ordered:
        if item[0] > item[2] or item[1] > item[3]:
            skipped += 1
            continue
        if any(_ranges_overlap(item, exist) for exist in accepted):
            skipped += 1
            continue
        accepted.append(item)
    return accepted, skipped


def _apply_merge_defs(
    ws: Worksheet,
    merged_defs: List[Tuple[int, int, int, int]],
    *,
    emit_log: Callable[[str], None],
) -> Tuple[int, int]:
    accepted, skipped = _filter_non_overlapping_merges(merged_defs)
    _unmerge_all_safely(ws, emit_log)
    for min_col, min_row, max_col, max_row in accepted:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )
    return len(accepted), skipped


def _insert_rows_like_excel(
    ws: Worksheet,
    *,
    insert_at: int,
    amount: int,
    template_row: int,
    emit_log: Callable[[str], None],
) -> None:
    if amount <= 0:
        return

    original_merged = list(ws.merged_cells.ranges)
    old_count = len(original_merged)

    # Clear merges before rows move. If stale merged ranges are unmerged after
    # insert_rows, openpyxl may delete values that have already shifted into the
    # old follower cells.
    _unmerge_all_safely(ws, emit_log)
    ws.insert_rows(insert_at, amount=amount)

    src_row = template_row if template_row < insert_at else template_row + amount
    max_col = ws.max_column
    for row_idx in range(insert_at, insert_at + amount):
        for col in range(1, max_col + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=row_idx, column=col)
            dst._style = copy(src._style)
        if src_row in ws.row_dimensions:
            ws.row_dimensions[row_idx].height = ws.row_dimensions[src_row].height

    merged_defs: List[Tuple[int, int, int, int]] = []
    template_merges: List[Tuple[int, int, int, int]] = []
    for merged in original_merged:
        min_col, min_row, max_col_merged, max_row_merged = (
            merged.min_col,
            merged.min_row,
            merged.max_col,
            merged.max_row,
        )
        if min_row >= insert_at:
            min_row += amount
            max_row_merged += amount
        merged_defs.append((min_col, min_row, max_col_merged, max_row_merged))
        if merged.min_row == template_row and merged.max_row == template_row:
            template_merges.append((merged.min_col, merged.min_row, merged.max_col, merged.max_row))

    for i in range(amount):
        row_idx = insert_at + i
        for min_col, _, max_col_merged, _ in template_merges:
            merged_defs.append((min_col, row_idx, max_col_merged, row_idx))

    applied_count, skipped = _apply_merge_defs(ws, merged_defs, emit_log=emit_log)
    emit_log(
        f"[交接班][插行] insert_at={insert_at}, amount={amount}, template_row={template_row}"
    )
    emit_log(
        f"[交接班][插行] 合并重建: old={old_count}, new={applied_count}, skipped_conflict={skipped}"
    )


def _delete_rows_like_excel(
    ws: Worksheet,
    *,
    delete_at: int,
    amount: int,
    emit_log: Callable[[str], None],
) -> None:
    if amount <= 0:
        return

    original_merged = list(ws.merged_cells.ranges)
    old_count = len(original_merged)
    delete_end = int(delete_at) + int(amount) - 1

    _unmerge_all_safely(ws, emit_log)
    ws.delete_rows(delete_at, amount=amount)

    merged_defs: List[Tuple[int, int, int, int]] = []
    for merged in original_merged:
        min_col, min_row, max_col_merged, max_row_merged = (
            merged.min_col,
            merged.min_row,
            merged.max_col,
            merged.max_row,
        )
        if max_row_merged < delete_at:
            pass
        elif min_row > delete_end:
            min_row -= amount
            max_row_merged -= amount
        elif min_row < delete_at and max_row_merged > delete_end:
            max_row_merged -= amount
        elif min_row < delete_at <= max_row_merged <= delete_end:
            max_row_merged = delete_at - 1
        elif delete_at <= min_row <= delete_end < max_row_merged:
            min_row = delete_at
            max_row_merged -= amount
        else:
            continue
        if min_row > max_row_merged:
            continue
        merged_defs.append((min_col, min_row, max_col_merged, max_row_merged))

    applied_count, skipped = _apply_merge_defs(ws, merged_defs, emit_log=emit_log)
    emit_log(f"[交接班][删行] delete_at={delete_at}, amount={amount}")
    emit_log(
        f"[交接班][删行] 合并重建: old={old_count}, new={applied_count}, skipped_conflict={skipped}"
    )


def _apply_row_snapshot(
    ws: Worksheet,
    *,
    target_row: int,
    snapshot,
    restore_values: bool,
    emit_log: Callable[[str], None],
) -> None:
    _clear_row_merges(ws, target_row, emit_log=emit_log)

    for col_idx, cell_snapshot in snapshot.cells.items():
        target = ws.cell(row=target_row, column=col_idx)
        target._style = copy(cell_snapshot.style)
        if restore_values:
            target.value = copy(cell_snapshot.value)

    ws.row_dimensions[target_row].height = snapshot.row_height

    existing = {str(item) for item in ws.merged_cells.ranges}
    for min_col, max_col in snapshot.merges:
        merged_ref = f"{get_column_letter(min_col)}{target_row}:{get_column_letter(max_col)}{target_row}"
        if merged_ref in existing:
            continue
        ws.merge_cells(merged_ref)
        existing.add(merged_ref)


def _blank_data_row(ws: Worksheet, row_idx: int) -> None:
    for col in "ABCDEFGHI":
        cell = ws[f"{col}{row_idx}"]
        if isinstance(cell, MergedCell):
            continue
        cell.value = ""


def _fill_empty_section_row(ws: Worksheet, row_idx: int) -> None:
    for col in "ABCDEFGHI":
        cell = ws[f"{col}{row_idx}"]
        if isinstance(cell, MergedCell):
            continue
        cell.value = "/"


def resolve_section_target_row_count(
    payload_count: int,
    *,
    empty_section_mode: str = "single_blank_row",
) -> int:
    if payload_count > 0:
        return payload_count
    if empty_section_mode == "single_blank_row":
        return 1
    raise ValueError(f"unsupported empty_section_mode: {empty_section_mode}")


def write_category_sections(
    *,
    ws: Worksheet,
    sections: List[CategorySection],
    category_payloads: Dict[str, Any] | None,
    snapshots: Dict[int, CategorySectionSnapshot] | None = None,
    empty_section_mode: str = "single_blank_row",
    preserve_template_values: bool = False,
    emit_log: Callable[[str], None] = print,
) -> None:
    payload_map = category_payloads if isinstance(category_payloads, dict) else {}
    normalized_payloads: Dict[str, List[Dict[str, Any]]] = {
        str(key): _normalize_payload_rows(value) for key, value in payload_map.items()
    }
    snapshots = snapshots if isinstance(snapshots, dict) else {}
    snapshot_by_name: Dict[str, CategorySectionSnapshot] = {item.name: item for item in snapshots.values()}
    target_rows_by_name: Dict[str, int] = {}

    for section in reversed(sections):
        payload_rows = normalized_payloads.get(section.name, [])
        payload_count = len(payload_rows)
        current_n = max(section.end_row - section.template_data_row + 1, 1)
        target_n = resolve_section_target_row_count(
            payload_count,
            empty_section_mode=empty_section_mode,
        )
        target_rows_by_name[section.name] = target_n
        section_changed = target_n != current_n or payload_count > 0

        op = "keep"
        if target_n > current_n:
            delta = target_n - current_n
            emit_log(
                f"[交接班][插行] 分类={section.name}, insert_at={section.end_row + 1}, amount={delta}, "
                f"template_row={section.template_data_row}"
            )
            _insert_rows_like_excel(
                ws,
                insert_at=section.end_row + 1,
                amount=delta,
                template_row=section.template_data_row,
                emit_log=emit_log,
            )
            op = f"insert+{delta}"
        elif target_n < current_n:
            delta = current_n - target_n
            _delete_rows_like_excel(
                ws,
                delete_at=section.template_data_row + target_n,
                amount=delta,
                emit_log=emit_log,
            )
            op = f"delete-{delta}"

        snapshot = snapshots.get(section.title_row)
        if snapshot is not None:
            _apply_row_snapshot(
                ws,
                target_row=section.title_row,
                snapshot=snapshot.title,
                restore_values=True,
                emit_log=emit_log,
            )
            _apply_row_snapshot(
                ws,
                target_row=section.header_row,
                snapshot=snapshot.header,
                restore_values=True,
                emit_log=emit_log,
            )

        for row_idx in range(section.template_data_row, section.template_data_row + target_n):
            if snapshot is not None:
                _apply_row_snapshot(
                    ws,
                    target_row=row_idx,
                    snapshot=snapshot.template,
                    restore_values=False,
                    emit_log=emit_log,
                )
            if not preserve_template_values or row_idx < section.template_data_row + payload_count:
                _blank_data_row(ws, row_idx)

        for idx, row in enumerate(payload_rows):
            row_idx = section.template_data_row + idx
            a_cell = ws.cell(row=row_idx, column=1)
            if not isinstance(a_cell, MergedCell):
                a_cell.value = idx + 1
            for col in "BCDEFGHI":
                target = ws[f"{col}{row_idx}"]
                if isinstance(target, MergedCell):
                    continue
                target.value = row.get(col, "")
        if payload_count == 0:
            _fill_empty_section_row(ws, section.template_data_row)

        title_cell = ws.cell(row=section.title_row, column=1)
        if not isinstance(title_cell, MergedCell):
            title_cell.value = section.name

        title_merge_desc = "-"
        if snapshot is not None and snapshot.title.merges:
            title_merge_desc = ",".join(
                f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
                for min_col, max_col in snapshot.title.merges
            )
            emit_log(
                f"[交接班][分类快照] 分类={section.name}, baseline_rows={section.baseline_rows}, "
                f"title_merge={title_merge_desc}"
            )
            emit_log(f"[交接班][分类恢复] 分类={section.name}, 标题/表头/数据样式已恢复")

        emit_log(
            f"[交接班][分类写入] 分类={section.name}, current={current_n}, target={target_n}, "
            f"payload={payload_count}, changed={int(section_changed)}, op={op}"
        )

    # 所有增删行完成后，基于最终行号再次恢复快照，避免上方插行污染下方分类样式/合并。
    if snapshot_by_name:
        final_sections = parse_category_sections(ws)
        for section in final_sections:
            snapshot = snapshot_by_name.get(section.name)
            if snapshot is None:
                continue

            _apply_row_snapshot(
                ws,
                target_row=section.title_row,
                snapshot=snapshot.title,
                restore_values=True,
                emit_log=emit_log,
            )
            _apply_row_snapshot(
                ws,
                target_row=section.header_row,
                snapshot=snapshot.header,
                restore_values=True,
                emit_log=emit_log,
            )

            final_target_n = max(
                target_rows_by_name.get(section.name, 1),
                1,
            )
            max_row_allowed = max(section.end_row, section.template_data_row)
            for row_idx in range(section.template_data_row, section.template_data_row + final_target_n):
                if row_idx > max_row_allowed:
                    emit_log(
                        f"[交接班][分类写入] 分类={section.name} 最终恢复越界跳过: "
                        f"row={row_idx}, max_row={max_row_allowed}"
                    )
                    continue
                _apply_row_snapshot(
                    ws,
                    target_row=row_idx,
                    snapshot=snapshot.template,
                    restore_values=False,
                    emit_log=emit_log,
                )

            title_cell = ws.cell(row=section.title_row, column=1)
            if not isinstance(title_cell, MergedCell):
                title_cell.value = section.name
