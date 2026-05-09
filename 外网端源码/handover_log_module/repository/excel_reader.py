from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

from handover_log_module.core.models import RawRow
from handover_log_module.core.normalizers import format_extracted_value, normalize_b, normalize_c, to_float


def _pick_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str | None,
    sheet_index: int | None,
) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"数据表缺少工作表: {sheet_name}")
        return wb[sheet_name]
    idx = int(sheet_index or 0)
    if idx < 0 or idx >= len(wb.sheetnames):
        raise ValueError(f"sheet_index越界: {idx}")
    return wb[wb.sheetnames[idx]]


def load_workbook_quietly(path: str | Path, **kwargs: Any) -> openpyxl.Workbook:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module=r"openpyxl\.styles\.stylesheet",
        )
        warnings.filterwarnings(
            "ignore",
            message="Unknown extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        warnings.filterwarnings(
            "ignore",
            message="Conditional Formatting extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        return openpyxl.load_workbook(path, **kwargs)


def load_rows(
    data_file: str,
    parsing_cfg: Dict[str, Any],
    normalize_cfg: Dict[str, Any],
) -> List[RawRow]:
    path = Path(data_file)
    if not path.exists():
        raise FileNotFoundError(f"数据表不存在: {path}")

    start_row = int(parsing_cfg.get("start_row", 4))
    col_b = int(parsing_cfg.get("col_b", 2))
    col_c = int(parsing_cfg.get("col_c", 3))
    col_d = int(parsing_cfg.get("col_d", 4))
    col_e = int(parsing_cfg.get("col_e", 5))
    forward_fill_b = bool(parsing_cfg.get("forward_fill_b", True))
    forward_fill_c = bool(parsing_cfg.get("forward_fill_c", True))

    b_regex = str(normalize_cfg.get("b_extract_regex", "")).strip()
    c_regex = str(normalize_cfg.get("c_extract_regex", "")).strip()

    wb = load_workbook_quietly(path, data_only=True)

    try:
        sheet_name = str(parsing_cfg.get("sheet_name", "")).strip() or None
        sheet_index = parsing_cfg.get("sheet_index", 0)
        ws = _pick_sheet(wb, sheet_name=sheet_name, sheet_index=sheet_index)

        rows: List[RawRow] = []
        last_b = ""
        last_c = ""

        for row_idx in range(start_row, ws.max_row + 1):
            b_raw = ws.cell(row_idx, col_b).value
            c_raw = ws.cell(row_idx, col_c).value
            d_raw = ws.cell(row_idx, col_d).value
            e_raw = ws.cell(row_idx, col_e).value
            e_value = to_float(e_raw)
            e_display = format_extracted_value(e_raw)

            b_text = str(b_raw).strip() if b_raw is not None else ""
            c_text = str(c_raw).strip() if c_raw is not None else ""
            d_name = str(d_raw).strip() if d_raw is not None else ""
            if not d_name:
                continue

            if b_text:
                last_b = b_text
            elif forward_fill_b:
                b_text = last_b

            if c_text:
                last_c = c_text
            elif forward_fill_c:
                c_text = last_c

            rows.append(
                RawRow(
                    row_index=row_idx,
                    b_text=b_text,
                    c_text=c_text,
                    d_name=d_name,
                    e_raw=e_display,
                    value=e_value,
                    b_norm=normalize_b(b_text, b_regex),
                    c_norm=normalize_c(c_text, c_regex),
                )
            )
        return rows
    finally:
        wb.close()
