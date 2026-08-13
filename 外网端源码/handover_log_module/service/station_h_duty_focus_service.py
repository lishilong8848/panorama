from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
import threading
import uuid
import hashlib
import json
from functools import lru_cache
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image

from app.shared.utils.atomic_file import atomic_write_bytes, atomic_write_text, validate_image_file
from app.shared.utils.runtime_temp_workspace import resolve_runtime_state_root
from handover_log_module.service.capacity_report_image_delivery_service import (
    _find_soffice_executable,
    _subprocess_creation_flags,
)
from handover_log_module.service.chiller_mode_focus_state_service import (
    ChillerModeFocusStateService,
)
from handover_log_module.service.review_session_service import ReviewSessionService
from handover_log_module.service.station_h_review_selection_service import (
    split_station_h_people,
    station_h_build_batch_key,
)
from handover_log_module.service.station_h_signature_service import (
    StationHSignatureError,
    StationHSignatureService,
)
from pipeline_utils import get_app_dir, get_bundle_dir


STATION_H_DUTY_FOCUS_TEMPLATE_NAME = "值班关注点模板.xlsx"
STATION_H_DUTY_FOCUS_BUILDINGS = ("A楼", "B楼", "C楼", "D楼", "E楼")
STATION_H_DUTY_FOCUS_MODE_VALUES = ("制冷", "预冷", "板换", "△")
STATION_H_DUTY_FOCUS_CHECK_ROWS = tuple(range(11, 40))
_SIGNATURE_MERGED_RANGES = {"handover": "E3:F3", "takeover": "H3:I3"}
_PDF_CONVERT_LOCK = threading.Lock()
_PRINT_DOCUMENT_LOCK = threading.Lock()
_DUTY_FOCUS_IMAGE_MAX_WIDTH = 2000
_DUTY_FOCUS_IMAGE_MAX_HEIGHT = 3200


class StationHDutyFocusError(RuntimeError):
    pass


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _bounded_text(value: Any, limit: int = 500) -> str:
    return _text(value)[: max(0, int(limit or 0))]


def _mode_text(value: Any) -> str:
    text = _text(value)
    aliases = {
        "1": "制冷",
        "2": "预冷",
        "3": "板换",
        "4": "△",
        "停机": "△",
        "未启动": "△",
        "▲": "△",
    }
    normalized = aliases.get(text, text)
    return normalized if normalized in STATION_H_DUTY_FOCUS_MODE_VALUES else "△"


def _safe_filename(value: Any) -> str:
    text = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_.-]+", "_", _text(value))
    return text.strip("._") or "值班关注点"


def _temperature(value: Any) -> str:
    text = _text(value).replace("℃", "").strip()
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return ""
    rendered = f"{number:.1f}".rstrip("0").rstrip(".")
    return f"{rendered}℃"


def _runtime_root(config: Dict[str, Any]) -> Path:
    payload = config if isinstance(config, dict) else {}
    paths = payload.get("paths", {}) if isinstance(payload.get("paths", {}), dict) else {}
    global_paths = payload.get("_global_paths", {}) if isinstance(payload.get("_global_paths", {}), dict) else {}
    runtime_state_root = _text(paths.get("runtime_state_root")) or _text(global_paths.get("runtime_state_root"))
    return resolve_runtime_state_root(
        runtime_config={"paths": {"runtime_state_root": runtime_state_root}} if runtime_state_root else {},
        app_dir=get_app_dir(),
    )


@lru_cache(maxsize=4)
def _cached_check_labels(template_path_text: str, modified_ns: int) -> Dict[str, str]:
    del modified_ns
    fallback = {str(row): f"确认项 {row}" for row in STATION_H_DUTY_FOCUS_CHECK_ROWS}
    workbook = load_workbook(Path(template_path_text), read_only=False, data_only=False)
    try:
        worksheet = workbook.worksheets[0]
        labels: Dict[str, str] = {}
        for row in STATION_H_DUTY_FOCUS_CHECK_ROWS:
            system = _text(worksheet.cell(row, 2).value)
            content = _text(worksheet.cell(row, 3).value)
            labels[str(row)] = f"{system} · {content}" if system else content or fallback[str(row)]
        return labels
    finally:
        workbook.close()


class StationHDutyFocusService:
    def __init__(
        self,
        config: Dict[str, Any] | None = None,
        *,
        review_service: ReviewSessionService | None = None,
        signature_service: StationHSignatureService | None = None,
        chiller_focus_service: ChillerModeFocusStateService | None = None,
        template_path: str | Path | None = None,
        emit_log: Callable[[str], None] | None = None,
    ) -> None:
        self.config = config if isinstance(config, dict) else {}
        self.review_service = review_service or ReviewSessionService(self.config)
        self.signature_service = signature_service or StationHSignatureService(self.config, emit_log=emit_log)
        self.chiller_focus_service = chiller_focus_service or ChillerModeFocusStateService(
            self.config,
            emit_log=emit_log,
        )
        self._template_path = Path(template_path) if template_path else None
        self._emit_log = emit_log

    def template_path(self) -> Path:
        configured = self.config.get("station_h_duty_focus", {}) if isinstance(self.config, dict) else {}
        configured_path = _text(configured.get("template_path")) if isinstance(configured, dict) else ""
        candidates: List[Path] = []
        if self._template_path is not None:
            candidates.append(self._template_path)
        if configured_path:
            path = Path(configured_path)
            candidates.append(path if path.is_absolute() else get_app_dir() / path)
        candidates.extend(
            [
                get_app_dir() / STATION_H_DUTY_FOCUS_TEMPLATE_NAME,
                get_bundle_dir() / STATION_H_DUTY_FOCUS_TEMPLATE_NAME,
            ]
        )
        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                return candidate
        raise StationHDutyFocusError(f"值班关注点模板不存在: {STATION_H_DUTY_FOCUS_TEMPLATE_NAME}")

    @staticmethod
    def _session_map(sessions: Iterable[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        return {
            _text(item.get("building")): item
            for item in sessions
            if isinstance(item, dict) and _text(item.get("building"))
        }

    @staticmethod
    def _modes_from_session(session: Dict[str, Any] | None) -> Dict[str, str]:
        if not isinstance(session, dict) or not session:
            return {str(unit): "" for unit in range(1, 7)}
        modes = {str(unit): "△" for unit in range(1, 7)}
        running = session.get("capacity_running_units", {})
        if not isinstance(running, dict):
            return modes
        for zone in ("west", "east"):
            rows = running.get(zone, [])
            if not isinstance(rows, list):
                continue
            for item in rows:
                if not isinstance(item, dict):
                    continue
                try:
                    unit = int(item.get("unit", 0) or 0)
                except Exception:
                    unit = 0
                if 1 <= unit <= 6:
                    modes[str(unit)] = _mode_text(item.get("mode_text") or item.get("mode_code"))
        return modes

    def _check_labels(self) -> Dict[str, str]:
        fallback = {str(row): f"确认项 {row}" for row in STATION_H_DUTY_FOCUS_CHECK_ROWS}
        try:
            path = self.template_path().resolve()
            return dict(_cached_check_labels(str(path), path.stat().st_mtime_ns))
        except Exception:
            return fallback

    def _outdoor_temperature_text(self, batch_key: str) -> str:
        try:
            state = self.review_service.get_outdoor_temperature_state(batch_key=batch_key)
            blocks = state.get("shared_blocks", {}) if isinstance(state, dict) else {}
            block = blocks.get("outdoor_temperature", {}) if isinstance(blocks, dict) else {}
            cells = block.get("cells", {}) if isinstance(block, dict) else {}
            dry = _temperature(cells.get("B7")) if isinstance(cells, dict) else ""
            wet = _temperature(cells.get("D7")) if isinstance(cells, dict) else ""
            return f"{dry}/{wet}" if dry and wet else ""
        except Exception:
            return ""

    def build_auto_focus(self, *, duty_date: str, duty_shift: str, selection: Dict[str, Any]) -> Dict[str, Any]:
        batch_key = station_h_build_batch_key(duty_date, duty_shift)
        try:
            chiller_state = self.chiller_focus_service.ensure_batch_state(batch_key)
        except Exception as exc:  # noqa: BLE001
            chiller_state = {"buildings": {}, "refresh_error": str(exc)}
        chiller_buildings = (
            chiller_state.get("buildings", {})
            if isinstance(chiller_state.get("buildings", {}), dict)
            else {}
        )
        complete_chiller_buildings = {
            building
            for building, item in chiller_buildings.items()
            if isinstance(item, dict)
            and isinstance(item.get("modes", {}), dict)
            and set(item.get("modes", {})) == {str(unit) for unit in range(1, 7)}
        }
        current_sessions: Dict[str, Dict[str, Any]] = {}
        if complete_chiller_buildings != set(STATION_H_DUTY_FOCUS_BUILDINGS):
            try:
                current_sessions = self._session_map(self.review_service.list_batch_sessions(batch_key))
            except Exception:
                current_sessions = {}
        rows: List[Dict[str, Any]] = []
        missing_buildings: List[str] = []
        for building in STATION_H_DUTY_FOCUS_BUILDINGS:
            current_session = current_sessions.get(building)
            chiller_building = chiller_buildings.get(building, {})
            chiller_building = chiller_building if isinstance(chiller_building, dict) else {}
            chiller_modes = chiller_building.get("modes", {})
            chiller_modes = chiller_modes if isinstance(chiller_modes, dict) else {}
            if set(chiller_modes) == {str(unit) for unit in range(1, 7)}:
                current_modes = {
                    str(unit): _mode_text(chiller_modes.get(str(unit)))
                    for unit in range(1, 7)
                }
                change_note = _bounded_text(chiller_building.get("change_note"), 200) or "无"
            else:
                current_modes = self._modes_from_session(current_session)
                change_note = "无"
            if not chiller_modes and not current_session:
                missing_buildings.append(building)
            rows.append(
                {
                    "building": building,
                    "modes": current_modes,
                    "change_note": change_note,
                }
            )
        checks = {str(row): "√" for row in STATION_H_DUTY_FOCUS_CHECK_ROWS}
        checks["19"] = self._outdoor_temperature_text(batch_key)
        current_people = split_station_h_people(selection.get("current_people", selection.get("current_people_text", "")))
        next_people = split_station_h_people(selection.get("next_people", selection.get("next_people_text", "")))
        return {
            "date_text": duty_date,
            "shift": duty_shift,
            "rows": rows,
            "checks": checks,
            "signatures": {
                "handover": {"selection_id": "", "table_id": "", "record_id": "", "name": current_people[0] if current_people else "", "match_source": "auto"},
                "takeover": {"selection_id": "", "table_id": "", "record_id": "", "name": next_people[0] if next_people else "", "match_source": "auto"},
            },
            "source": "auto",
            "auto_source": {
                "current_batch": batch_key,
                "missing_buildings": missing_buildings,
                "mode_revision": _text(chiller_state.get("mode_revision")),
                "mode_observed_at": _text(chiller_state.get("observed_at")),
                "mode_source": _text(chiller_state.get("source")),
                "mode_error": _text(chiller_state.get("refresh_error")),
                "mode_buildings": sorted(complete_chiller_buildings),
            },
        }

    @staticmethod
    def _saved_focus_complete(raw: Any) -> bool:
        if not isinstance(raw, dict):
            return False
        rows = raw.get("rows", [])
        if not isinstance(rows, list):
            return False
        by_building = {
            _text(item.get("building")): item
            for item in rows
            if isinstance(item, dict) and _text(item.get("building"))
        }
        for building in STATION_H_DUTY_FOCUS_BUILDINGS:
            item = by_building.get(building, {})
            modes = item.get("modes", {}) if isinstance(item, dict) else {}
            if not isinstance(modes, dict) or any(str(unit) not in modes for unit in range(1, 7)):
                return False
        checks = raw.get("checks", {})
        return isinstance(checks, dict) and all(str(row) in checks for row in STATION_H_DUTY_FOCUS_CHECK_ROWS)

    @staticmethod
    def _saved_focus_fallback(*, duty_date: str, duty_shift: str, selection: Dict[str, Any]) -> Dict[str, Any]:
        current_people = split_station_h_people(selection.get("current_people", selection.get("current_people_text", "")))
        next_people = split_station_h_people(selection.get("next_people", selection.get("next_people_text", "")))
        return {
            "date_text": duty_date,
            "shift": duty_shift,
            "rows": [
                {
                    "building": building,
                    "modes": {str(unit): "△" for unit in range(1, 7)},
                    "change_note": "无",
                }
                for building in STATION_H_DUTY_FOCUS_BUILDINGS
            ],
            "checks": {str(row): "" for row in STATION_H_DUTY_FOCUS_CHECK_ROWS},
            "signatures": {
                "handover": {"name": current_people[0] if current_people else "", "match_source": "auto"},
                "takeover": {"name": next_people[0] if next_people else "", "match_source": "auto"},
            },
            "source": "manual",
            "auto_source": {
                "current_batch": station_h_build_batch_key(duty_date, duty_shift),
                "missing_buildings": [],
                "mode_revision": "",
                "mode_observed_at": "",
                "mode_source": "",
                "mode_error": "",
                "mode_buildings": [],
            },
        }

    @staticmethod
    def _normalize_signature(raw: Any, fallback: Dict[str, Any]) -> Dict[str, Any]:
        payload = raw if isinstance(raw, dict) else {}
        selection_id = _bounded_text(payload.get("selection_id"), 256)
        table_id = _bounded_text(payload.get("table_id"), 128)
        record_id = _bounded_text(payload.get("record_id"), 128)
        if selection_id and ":" in selection_id and (not table_id or not record_id):
            table_id, record_id = selection_id.split(":", 1)
        if table_id and record_id:
            selection_id = f"{table_id}:{record_id}"
        raw_match_source = _text(payload.get("match_source")).lower()
        fallback_match_source = _text(fallback.get("match_source")).lower()
        match_source = raw_match_source if raw_match_source in {"auto", "manual"} else fallback_match_source
        if match_source not in {"auto", "manual"}:
            match_source = "manual" if selection_id else "auto"
        return {
            "selection_id": selection_id,
            "table_id": table_id,
            "record_id": record_id,
            "name": _bounded_text(payload.get("name"), 100) or _bounded_text(fallback.get("name"), 100),
            "signature_revision": (
                _bounded_text(payload.get("signature_revision"), 128)
                or _bounded_text(fallback.get("signature_revision"), 128)
            ),
            "match_source": match_source,
        }

    @classmethod
    def normalize_focus(cls, raw: Any, *, fallback: Dict[str, Any]) -> Dict[str, Any]:
        payload = raw if isinstance(raw, dict) else {}
        fallback_rows = {
            _text(item.get("building")): item
            for item in fallback.get("rows", [])
            if isinstance(item, dict)
        }
        raw_rows = {
            _text(item.get("building")): item
            for item in payload.get("rows", [])
            if isinstance(item, dict)
        }
        rows: List[Dict[str, Any]] = []
        for building in STATION_H_DUTY_FOCUS_BUILDINGS:
            base = fallback_rows.get(building, {})
            incoming = raw_rows.get(building, {})
            base_modes = base.get("modes", {}) if isinstance(base.get("modes", {}), dict) else {}
            incoming_modes = incoming.get("modes", {}) if isinstance(incoming.get("modes", {}), dict) else {}
            rows.append(
                {
                    "building": building,
                    "modes": {
                        str(unit): _mode_text(incoming_modes.get(str(unit), base_modes.get(str(unit), "△")))
                        for unit in range(1, 7)
                    },
                    "change_note": _bounded_text(incoming.get("change_note"), 200)
                    if "change_note" in incoming
                    else _bounded_text(base.get("change_note"), 200),
                }
            )
        base_checks = fallback.get("checks", {}) if isinstance(fallback.get("checks", {}), dict) else {}
        raw_checks = payload.get("checks", {}) if isinstance(payload.get("checks", {}), dict) else {}
        checks = {
            str(row): _bounded_text(raw_checks.get(str(row), base_checks.get(str(row), "")), 500)
            for row in STATION_H_DUTY_FOCUS_CHECK_ROWS
        }
        if not checks["19"]:
            checks["19"] = _bounded_text(base_checks.get("19"), 500)
        base_signatures = fallback.get("signatures", {}) if isinstance(fallback.get("signatures", {}), dict) else {}
        raw_signatures = payload.get("signatures", {}) if isinstance(payload.get("signatures", {}), dict) else {}
        base_auto_source = (
            dict(fallback.get("auto_source", {}))
            if isinstance(fallback.get("auto_source", {}), dict)
            else {}
        )
        raw_auto_source = payload.get("auto_source", {}) if isinstance(payload.get("auto_source", {}), dict) else {}
        for key in (
            "current_batch",
            "mode_revision",
            "mode_observed_at",
            "mode_source",
            "mode_error",
        ):
            if key in raw_auto_source:
                base_auto_source[key] = _bounded_text(raw_auto_source.get(key), 256)
        return {
            "date_text": _text(payload.get("date_text")) or _text(fallback.get("date_text")),
            "shift": _text(payload.get("shift")).lower() if _text(payload.get("shift")).lower() in {"day", "night"} else _text(fallback.get("shift")),
            "rows": rows,
            "checks": checks,
            "signatures": {
                slot: cls._normalize_signature(raw_signatures.get(slot), base_signatures.get(slot, {}))
                for slot in ("handover", "takeover")
            },
            "source": "manual" if isinstance(raw, dict) and raw else _text(fallback.get("source")) or "auto",
            "auto_source": base_auto_source,
        }

    @classmethod
    def normalize_submitted_focus(
        cls,
        raw: Any,
        *,
        duty_date: str,
        duty_shift: str,
        selection: Dict[str, Any],
    ) -> Dict[str, Any]:
        fallback = cls._saved_focus_fallback(
            duty_date=duty_date,
            duty_shift=duty_shift,
            selection=selection,
        )
        normalized = cls.normalize_focus(raw, fallback=fallback)
        normalized["date_text"] = duty_date
        normalized["shift"] = duty_shift
        return normalized

    def build_status(self, *, duty_date: str, duty_shift: str, selection: Dict[str, Any]) -> Dict[str, Any]:
        saved_focus = selection.get("duty_focus") if isinstance(selection, dict) else None
        fallback = self.build_auto_focus(duty_date=duty_date, duty_shift=duty_shift, selection=selection)
        focus = self.normalize_focus(saved_focus, fallback=fallback)
        saved_auto_source = saved_focus.get("auto_source", {}) if isinstance(saved_focus, dict) else {}
        saved_revision = _text(saved_auto_source.get("mode_revision")) if isinstance(saved_auto_source, dict) else ""
        current_revision = _text(fallback.get("auto_source", {}).get("mode_revision"))
        if self._saved_focus_complete(saved_focus) and current_revision and saved_revision != current_revision:
            automatic_mode_buildings = {
                _text(building)
                for building in fallback.get("auto_source", {}).get("mode_buildings", [])
                if _text(building)
            }
            automatic_rows = {
                _text(item.get("building")): item
                for item in fallback.get("rows", [])
                if isinstance(item, dict)
            }
            for item in focus.get("rows", []):
                if not isinstance(item, dict):
                    continue
                if _text(item.get("building")) not in automatic_mode_buildings:
                    continue
                automatic = automatic_rows.get(_text(item.get("building")), {})
                if automatic:
                    item["modes"] = dict(automatic.get("modes", {}))
                    item["change_note"] = _text(automatic.get("change_note")) or "无"
        focus["auto_source"] = dict(fallback.get("auto_source", {}))
        current_people = split_station_h_people(selection.get("current_people", selection.get("current_people_text", "")))
        next_people = split_station_h_people(selection.get("next_people", selection.get("next_people_text", "")))
        expected_names = {
            "handover": current_people[0] if current_people else "",
            "takeover": next_people[0] if next_people else "",
        }
        required_names: List[str] = []
        required_selection_ids: List[str] = []
        for slot in ("handover", "takeover"):
            selected = focus["signatures"][slot]
            selection_id = _text(selected.get("selection_id"))
            if _text(selected.get("match_source")).lower() == "manual" and selection_id:
                required_selection_ids.append(selection_id)
            elif expected_names[slot]:
                required_names.append(expected_names[slot])
        directory = self.signature_service.cached_directory()
        try:
            ensure_directory = getattr(self.signature_service, "ensure_directory", None)
            directory = (
                ensure_directory(
                    required_names=required_names,
                    required_selection_ids=required_selection_ids,
                )
                if callable(ensure_directory)
                else self.signature_service.refresh_directory()
            )
        except Exception as exc:  # noqa: BLE001
            directory = dict(directory)
            directory["error"] = f"签名目录自动刷新失败: {exc}"
        people = directory.get("people", []) if isinstance(directory.get("people", []), list) else []
        available_by_id = {
            _text(person.get("selection_id")): person
            for person in people
            if isinstance(person, dict) and bool(person.get("available"))
        }
        for slot in ("handover", "takeover"):
            selected = focus["signatures"][slot]
            selection_id = _text(selected.get("selection_id"))
            match_source = _text(selected.get("match_source")).lower()
            expected_name = expected_names[slot]
            auto_match = (
                match_source == "auto"
                or not selection_id
                or (match_source not in {"auto", "manual"} and _text(selected.get("name")) == expected_name)
            )
            matched = None
            if auto_match:
                matched = self.signature_service.match_person(people, expected_name)
            elif selection_id:
                matched = available_by_id.get(selection_id)
            if matched:
                focus["signatures"][slot] = {
                    "selection_id": _text(matched.get("selection_id")),
                    "table_id": _text(matched.get("table_id")),
                    "record_id": _text(matched.get("record_id")),
                    "name": _text(matched.get("name")),
                    "signature_revision": _text(matched.get("signature_revision")),
                    "match_source": "auto" if auto_match else "manual",
                }
            elif auto_match:
                focus["signatures"][slot] = {
                    "selection_id": "",
                    "table_id": "",
                    "record_id": "",
                    "name": expected_name,
                    "signature_revision": "",
                    "match_source": "auto",
                }
        missing_signature_slots: List[str] = []
        for slot, label in (("handover", "交班确认人"), ("takeover", "接班确认人")):
            selection_id = _text(focus["signatures"][slot].get("selection_id"))
            if not selection_id or selection_id not in available_by_id:
                missing_signature_slots.append(label)
        check_labels = self._check_labels()
        image_status = self.image_status(
            duty_date=duty_date,
            duty_shift=duty_shift,
            focus=focus,
        )
        return {
            **focus,
            "check_items": [
                {"row": row, "cell": f"H{row}", "label": check_labels.get(str(row), f"确认项 {row}")}
                for row in STATION_H_DUTY_FOCUS_CHECK_ROWS
            ],
            "mode_options": list(STATION_H_DUTY_FOCUS_MODE_VALUES),
            "signature_directory": directory,
            "print_ready": not missing_signature_slots,
            "print_block_reason": "、".join(missing_signature_slots) + "缺少可用签名" if missing_signature_slots else "",
            "image_status": image_status,
        }

    def output_root(self) -> Path:
        root = _runtime_root(self.config) / "handover" / "station_h_duty_focus"
        root.mkdir(parents=True, exist_ok=True)
        return root

    @classmethod
    def focus_content_hash(
        cls,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
    ) -> str:
        normalized = cls.normalize_focus(
            focus,
            fallback={
                "date_text": duty_date,
                "shift": duty_shift,
                "rows": [],
                "checks": {},
                "signatures": {},
                "source": "manual",
            },
        )
        signatures = normalized.get("signatures", {}) if isinstance(normalized.get("signatures"), dict) else {}
        payload = {
            "date_text": duty_date,
            "shift": duty_shift,
            "rows": normalized.get("rows", []),
            "checks": normalized.get("checks", {}),
            "signatures": {
                slot: {
                    "table_id": _text((signatures.get(slot) or {}).get("table_id")),
                    "record_id": _text((signatures.get(slot) or {}).get("record_id")),
                    "name": _text((signatures.get(slot) or {}).get("name")),
                    "signature_revision": _text((signatures.get(slot) or {}).get("signature_revision")),
                }
                for slot in ("handover", "takeover")
            },
        }
        encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()

    def _image_artifacts(self, *, duty_date: str, duty_shift: str) -> tuple[Path, Path]:
        output_dir = self.output_root() / f"{duty_date}_{duty_shift}"
        output_dir.mkdir(parents=True, exist_ok=True)
        stem = f"值班关注点_{_safe_filename(duty_date)}_{duty_shift}"
        return output_dir / f"{stem}.png", output_dir / f"{stem}.image.json"

    @staticmethod
    def _read_image_metadata(path: Path) -> Dict[str, Any]:
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        return dict(payload) if isinstance(payload, dict) else {}

    def image_status(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
    ) -> Dict[str, Any]:
        image_path, metadata_path = self._image_artifacts(
            duty_date=duty_date,
            duty_shift=duty_shift,
        )
        content_hash = self.focus_content_hash(
            duty_date=duty_date,
            duty_shift=duty_shift,
            focus=focus,
        )
        metadata = self._read_image_metadata(metadata_path)
        available = bool(image_path.exists() and image_path.is_file() and image_path.stat().st_size > 0)
        current = bool(available and _text(metadata.get("content_hash")) == content_hash)
        return {
            "status": "current" if current else ("stale" if available else "missing"),
            "available": available,
            "current": current,
            "content_hash": content_hash,
            "generated_at": _text(metadata.get("generated_at")),
            "image_sha256": _text(metadata.get("image_sha256")),
        }

    def current_image_path(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
    ) -> Path:
        image_path, _metadata_path = self._image_artifacts(
            duty_date=duty_date,
            duty_shift=duty_shift,
        )
        status = self.image_status(duty_date=duty_date, duty_shift=duty_shift, focus=focus)
        if not bool(status.get("current", False)):
            raise StationHDutyFocusError("值班关注点图片尚未生成或已过期，请重新生成")
        try:
            validate_image_file(image_path)
        except Exception as exc:
            raise StationHDutyFocusError(f"值班关注点图片文件无效: {exc}") from exc
        return image_path

    @staticmethod
    def _center_cell(cell: Any) -> None:
        alignment = copy(cell.alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        alignment.wrap_text = True
        cell.alignment = alignment

    @staticmethod
    def _use_text_font(cell: Any) -> None:
        # The empty input cells in the source template use Wingdings 2 so a
        # handwritten check mark renders correctly. Program-filled text must
        # switch back to a Chinese text font or LibreOffice prints it as
        # symbols/boxes.
        font = copy(cell.font)
        font.name = "宋体"
        font.charset = 134
        cell.font = font

    @staticmethod
    def _use_mode_font(cell: Any) -> None:
        font = copy(cell.font)
        font.name = "宋体"
        font.charset = 134
        font.sz = 11
        font.bold = False
        font.italic = False
        font.underline = None
        font.strike = False
        font.color = "000000"
        cell.font = font

    @staticmethod
    def _column_pixels(worksheet: Any, start_col: int, end_col: int) -> int:
        pixels = 0
        for col_idx in range(start_col, end_col + 1):
            letter = get_column_letter(col_idx)
            width = float(worksheet.column_dimensions[letter].width or 8.43)
            pixels += max(1, int(width * 7 + 5))
        return pixels

    @classmethod
    def _add_signature_image(cls, worksheet: Any, *, png: bytes, merged_range: str) -> None:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(merged_range)
        with Image.open(BytesIO(png)) as source:
            width, height = source.size
        if width <= 0 or height <= 0:
            raise StationHDutyFocusError("签名图片尺寸无效")
        box_width = max(20, cls._column_pixels(worksheet, min_col, max_col) - 8)
        row_height_pt = float(worksheet.row_dimensions[min_row].height or 30)
        box_height = max(18, int(row_height_pt * 96 / 72) - 4)
        scale = min(box_width / width, box_height / height)
        target_width = max(1, int(width * scale))
        target_height = max(1, int(height * scale))
        x_offset = max(0, int((box_width - target_width) / 2) + 4)
        y_offset = max(0, int((box_height - target_height) / 2) + 2)
        image = OpenpyxlImage(BytesIO(png))
        image.width = target_width
        image.height = target_height
        image.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=min_col - 1,
                colOff=pixels_to_EMU(x_offset),
                row=min_row - 1,
                rowOff=pixels_to_EMU(y_offset),
            ),
            ext=XDRPositiveSize2D(pixels_to_EMU(target_width), pixels_to_EMU(target_height)),
        )
        worksheet.add_image(image)

    @staticmethod
    def _shift_cell_text(shift: str) -> str:
        return "√白班\n□夜班" if shift == "day" else "□白班\n√夜班"

    def _resolve_signatures(self, focus: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
        results: Dict[str, Dict[str, Any]] = {}
        for slot, label in (("handover", "交班确认人"), ("takeover", "接班确认人")):
            signature = focus.get("signatures", {}).get(slot, {})
            table_id = _text(signature.get("table_id")) if isinstance(signature, dict) else ""
            record_id = _text(signature.get("record_id")) if isinstance(signature, dict) else ""
            if not table_id or not record_id:
                raise StationHDutyFocusError(f"{label}缺少可用签名，暂不能打印")
            try:
                results[slot] = self.signature_service.resolve_signature_png(table_id=table_id, record_id=record_id)
            except StationHSignatureError as exc:
                raise StationHDutyFocusError(f"{label}签名不可用: {exc}") from exc
        return results

    def build_workbook(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
        require_signatures: bool = True,
    ) -> Path:
        normalized = self.normalize_focus(focus, fallback={
            "date_text": duty_date,
            "shift": duty_shift,
            "rows": [],
            "checks": {},
            "signatures": {},
            "source": "manual",
        })
        signatures = self._resolve_signatures(normalized) if require_signatures else {}
        output_dir = self.output_root() / f"{duty_date}_{duty_shift}"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"值班关注点_{_safe_filename(duty_date)}_{duty_shift}.xlsx"
        temp_path = output_dir / f".{output_path.stem}.{os.getpid()}.{threading.get_ident()}.tmp.xlsx"
        try:
            workbook = load_workbook(self.template_path(), read_only=False, data_only=False)
            try:
                worksheet = workbook.worksheets[0]
                worksheet["B3"] = normalized.get("date_text") or duty_date
                worksheet["C3"] = self._shift_cell_text(normalized.get("shift") or duty_shift)
                self._use_text_font(worksheet["B3"])
                row_by_building = {building: row for row, building in enumerate(STATION_H_DUTY_FOCUS_BUILDINGS, start=5)}
                for item in normalized.get("rows", []):
                    if not isinstance(item, dict):
                        continue
                    building = _text(item.get("building"))
                    target_row = row_by_building.get(building)
                    if not target_row:
                        continue
                    modes = item.get("modes", {}) if isinstance(item.get("modes", {}), dict) else {}
                    for unit in range(1, 7):
                        worksheet.cell(target_row, unit + 1).value = _mode_text(modes.get(str(unit)))
                        self._use_mode_font(worksheet.cell(target_row, unit + 1))
                        self._center_cell(worksheet.cell(target_row, unit + 1))
                    worksheet.cell(target_row, 8).value = _text(item.get("change_note"))
                    self._center_cell(worksheet.cell(target_row, 8))
                checks = normalized.get("checks", {}) if isinstance(normalized.get("checks", {}), dict) else {}
                for row in STATION_H_DUTY_FOCUS_CHECK_ROWS:
                    worksheet.cell(row, 8).value = _text(checks.get(str(row)))
                    if row == 19:
                        self._use_text_font(worksheet.cell(row, 8))
                    self._center_cell(worksheet.cell(row, 8))
                for coordinate in ("B3", "C3"):
                    self._center_cell(worksheet[coordinate])
                if signatures:
                    for slot, merged_range in _SIGNATURE_MERGED_RANGES.items():
                        self._add_signature_image(worksheet, png=bytes(signatures[slot]["png"]), merged_range=merged_range)
                worksheet.print_area = "A1:I41"
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 1
                worksheet.page_setup.orientation = "portrait"
                workbook.save(temp_path)
            finally:
                workbook.close()
            os.replace(temp_path, output_path)
        finally:
            temp_path.unlink(missing_ok=True)
        return output_path

    def build_print_pdf(self, *, workbook_path: Path) -> Path:
        soffice = _find_soffice_executable()
        if soffice is None:
            raise StationHDutyFocusError("当前运行环境未安装 LibreOffice，无法生成原样打印文件")
        output_dir = workbook_path.parent / "print"
        output_dir.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(prefix="station_h_focus_", dir=str(output_dir)) as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            print_workbook = temp_dir / workbook_path.name
            workbook = load_workbook(workbook_path, read_only=False, data_only=False)
            try:
                for index, worksheet in enumerate(workbook.worksheets):
                    worksheet.sheet_state = "visible" if index == 0 else "hidden"
                workbook.save(print_workbook)
            finally:
                workbook.close()
            # LibreOffice on Windows may outlive the launcher briefly. A stable
            # profile outside the temporary conversion folder avoids a startup
            # race where the profile disappears before soffice.bin is ready.
            profile_dir = self.output_root() / f"libreoffice_profile_{os.getpid()}"
            profile_dir.mkdir(parents=True, exist_ok=True)
            command = [
                str(soffice),
                "--headless",
                "--nologo",
                "--nofirststartwizard",
                "--nodefault",
                "--nolockcheck",
                f"-env:UserInstallation={profile_dir.as_uri()}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(temp_dir),
                str(print_workbook),
            ]
            with _PDF_CONVERT_LOCK:
                result = subprocess.run(
                    command,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=180,
                    creationflags=_subprocess_creation_flags(),
                )
            generated = temp_dir / f"{print_workbook.stem}.pdf"
            if result.returncode != 0:
                detail = _text(result.stderr or result.stdout)
                raise StationHDutyFocusError(
                    f"值班关注点打印文件生成失败: code={result.returncode}, detail={detail[-500:] or '-'}"
                )
            if not generated.exists() or generated.stat().st_size <= 0:
                raise StationHDutyFocusError("值班关注点打印文件生成后为空")
            target = output_dir / f"{workbook_path.stem}_{uuid.uuid4().hex}.pdf"
            temp_target = target.with_suffix(".tmp.pdf")
            shutil.copy2(generated, temp_target)
            os.replace(temp_target, target)
            return target

    @staticmethod
    def _render_pdf_to_png_bytes(pdf_path: Path) -> bytes:
        try:
            import pypdfium2
        except Exception as exc:  # noqa: BLE001
            raise StationHDutyFocusError(f"缺少 pypdfium2 依赖，无法生成值班关注点图片: {exc}") from exc

        document = pypdfium2.PdfDocument(str(pdf_path))
        pages: List[Image.Image] = []
        try:
            page_count = len(document)
            if page_count <= 0:
                raise StationHDutyFocusError("值班关注点打印文件没有可渲染页面")
            for index in range(page_count):
                page = document[index]
                try:
                    bitmap = page.render(scale=2.0)
                    pages.append(bitmap.to_pil().convert("RGB"))
                finally:
                    try:
                        page.close()
                    except Exception:
                        pass
        finally:
            try:
                document.close()
            except Exception:
                pass
        if not pages:
            raise StationHDutyFocusError("值班关注点图片渲染结果为空")

        gap = 12 if len(pages) > 1 else 0
        width = max(page.width for page in pages)
        height = sum(page.height for page in pages) + gap * max(0, len(pages) - 1)
        combined = Image.new("RGB", (width, height), "white")
        y_offset = 0
        for page_image in pages:
            x_offset = max(0, int((width - page_image.width) / 2))
            combined.paste(page_image, (x_offset, y_offset))
            y_offset += page_image.height + gap

        scale = min(
            1.0,
            _DUTY_FOCUS_IMAGE_MAX_WIDTH / max(1, combined.width),
            _DUTY_FOCUS_IMAGE_MAX_HEIGHT / max(1, combined.height),
        )
        if scale < 1.0:
            combined = combined.resize(
                (max(1, int(combined.width * scale)), max(1, int(combined.height * scale))),
                Image.Resampling.LANCZOS,
            )
        buffer = BytesIO()
        combined.save(buffer, format="PNG", optimize=True)
        content = buffer.getvalue()
        if not content:
            raise StationHDutyFocusError("值班关注点图片生成后为空")
        return content

    def build_image_document(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
        force: bool = False,
    ) -> Dict[str, Any]:
        with _PRINT_DOCUMENT_LOCK:
            before = self.image_status(duty_date=duty_date, duty_shift=duty_shift, focus=focus)
            image_path, metadata_path = self._image_artifacts(
                duty_date=duty_date,
                duty_shift=duty_shift,
            )
            if not force and bool(before.get("current", False)):
                try:
                    validate_image_file(image_path)
                    return {**before, "path": image_path, "generated": False}
                except Exception:
                    pass

            workbook_path = self.build_workbook(
                duty_date=duty_date,
                duty_shift=duty_shift,
                focus=focus,
                require_signatures=True,
            )
            pdf_path: Path | None = None
            try:
                pdf_path = self.build_print_pdf(workbook_path=workbook_path)
                image_bytes = self._render_pdf_to_png_bytes(pdf_path)
                atomic_write_bytes(
                    image_path,
                    image_bytes,
                    validator=validate_image_file,
                    temp_suffix=".tmp",
                )
                metadata = {
                    "content_hash": self.focus_content_hash(
                        duty_date=duty_date,
                        duty_shift=duty_shift,
                        focus=focus,
                    ),
                    "image_sha256": hashlib.sha256(image_bytes).hexdigest(),
                    "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "file_name": image_path.name,
                }
                atomic_write_text(
                    metadata_path,
                    json.dumps(metadata, ensure_ascii=False, indent=2),
                    temp_suffix=".tmp",
                )
                return {
                    "status": "current",
                    "available": True,
                    "current": True,
                    "content_hash": metadata["content_hash"],
                    "generated_at": metadata["generated_at"],
                    "image_sha256": metadata["image_sha256"],
                    "path": image_path,
                    "generated": True,
                }
            finally:
                workbook_path.unlink(missing_ok=True)
                if pdf_path is not None:
                    pdf_path.unlink(missing_ok=True)

    def build_print_document(
        self,
        *,
        duty_date: str,
        duty_shift: str,
        focus: Dict[str, Any],
    ) -> Path:
        with _PRINT_DOCUMENT_LOCK:
            workbook_path = self.build_workbook(
                duty_date=duty_date,
                duty_shift=duty_shift,
                focus=focus,
                require_signatures=True,
            )
            try:
                return self.build_print_pdf(workbook_path=workbook_path)
            finally:
                workbook_path.unlink(missing_ok=True)
