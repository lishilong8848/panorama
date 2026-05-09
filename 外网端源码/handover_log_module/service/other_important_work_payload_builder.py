from __future__ import annotations

from pathlib import Path
from typing import Any, Callable, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from pipeline_utils import get_app_dir
from handover_log_module.repository.excel_reader import load_workbook_quietly
from handover_log_module.core.section_layout import parse_category_sections
from handover_log_module.core.specialty_normalizer import normalize_specialty_text, pick_engineer_supervisor
from handover_log_module.repository.other_important_work_repository import (
    OtherImportantWorkRepository,
    OtherImportantWorkRow,
    OtherImportantWorkRowsByBuilding,
)
from handover_log_module.repository.maintenance_management_repository import (
    MaintenanceManagementRow,
    MaintenanceRowsByBuilding,
)
from handover_log_module.repository.shift_roster_repository import ShiftRosterRepository


def _norm_header(value: Any) -> str:
    return str(value or "").replace(" ", "").strip().casefold()


def _dedupe_text(value: Any) -> str:
    return "".join(str(value or "").split()).casefold()


def _is_equalizing_charge_notice(value: Any) -> bool:
    return "均充" in str(value or "").strip()


class OtherImportantWorkPayloadBuilder:
    def __init__(
        self,
        handover_cfg: Dict[str, Any],
        *,
        repository: OtherImportantWorkRepository | None = None,
        shift_roster_repo: ShiftRosterRepository | None = None,
    ) -> None:
        self.handover_cfg = handover_cfg
        self.repo = repository or OtherImportantWorkRepository(handover_cfg)
        self.shift_roster_repo = shift_roster_repo or ShiftRosterRepository(handover_cfg)

    def _resolve_template_path(self, source_path: str) -> Path:
        path = Path(str(source_path or "").strip())
        if path.is_absolute():
            return path
        return get_app_dir() / path

    def _resolve_section_column_mapping(
        self,
        *,
        cfg: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Dict[str, str]]:
        col_cfg = cfg.get("column_mapping", {}) if isinstance(cfg.get("column_mapping", {}), dict) else {}
        fallback_cols = col_cfg.get("fallback_cols", {}) if isinstance(col_cfg.get("fallback_cols", {}), dict) else {}
        header_alias = col_cfg.get("header_alias", {}) if isinstance(col_cfg.get("header_alias", {}), dict) else {}
        resolve_by_header = bool(col_cfg.get("resolve_by_header", True))
        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("other_important_work", "其他重要工作记录")).strip() or "其他重要工作记录"
        semantic_keys = ["description", "completion", "executor"]
        defaults = {
            "description": "B",
            "completion": "F",
            "executor": "H",
        }
        output = {
            section_name: {
                key: str(fallback_cols.get(key, "")).strip().upper() or defaults[key]
                for key in semantic_keys
            }
        }
        if not resolve_by_header:
            return output

        template_cfg = self.handover_cfg.get("template", {})
        if not isinstance(template_cfg, dict):
            return output
        source_path = self._resolve_template_path(str(template_cfg.get("source_path", "")).strip())
        sheet_name = str(template_cfg.get("sheet_name", "交接班日志")).strip() or "交接班日志"
        if not source_path.exists():
            return output

        wb = load_workbook_quietly(source_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return output
            ws = wb[sheet_name]
            sections = parse_category_sections(ws)
            section = next((item for item in sections if item.name == section_name), None)
            if section is None:
                return output
            header_row = section.header_row
            normalized_header_map: Dict[str, str] = {}
            max_col = int(ws.max_column or 0)
            for col_idx in range(1, max_col + 1):
                text = _norm_header(ws.cell(row=header_row, column=col_idx).value)
                if not text:
                    continue
                normalized_header_map[text] = get_column_letter(col_idx)
            for semantic_key in semantic_keys:
                aliases_raw = header_alias.get(semantic_key, [])
                aliases = aliases_raw if isinstance(aliases_raw, list) else []
                found_col = ""
                for alias in aliases:
                    normalized = _norm_header(alias)
                    if normalized and normalized in normalized_header_map:
                        found_col = normalized_header_map[normalized]
                        break
                if found_col:
                    output[section_name][semantic_key] = found_col
            emit_log(f"[交接班][其他重要工作] 列映射已解析: {output.get(section_name)}")
            return output
        finally:
            wb.close()

    @staticmethod
    def _to_cells(row_payload: Dict[str, str], col_map: Dict[str, str]) -> Dict[str, str]:
        cells: Dict[str, str] = {}
        for semantic_key, text in row_payload.items():
            col = str(col_map.get(semantic_key, "")).strip().upper()
            if not col:
                continue
            cells[col] = str(text or "")
        return cells

    @staticmethod
    def _pick_supervisor(
        engineers: List[Dict[str, str]],
        *,
        building: str,
        specialty_text: str,
    ) -> str:
        return pick_engineer_supervisor(
            engineers,
            building=building,
            specialty_text=specialty_text,
        )

    def _resolve_executor(
        self,
        *,
        row: OtherImportantWorkRow,
        building: str,
        engineers: List[Dict[str, str]],
        emit_log: Callable[[str], None],
    ) -> str:
        supervisor = self._pick_supervisor(
            engineers,
            building=building,
            specialty_text=row.specialty_text,
        )
        if supervisor:
            emit_log(
                "[交接班][其他重要工作] 执行人匹配: "
                f"source={row.source_key}, record_id={row.record_id}, building={building}, "
                f"specialty={row.specialty_text or '-'}, supervisor={supervisor}"
            )
            return supervisor

        emit_log(
            "[交接班][其他重要工作] 执行人未匹配: "
            f"source={row.source_key}, record_id={row.record_id}, building={building}, "
            f"specialty={row.specialty_text or '-'}"
        )
        return "/"

    def _resolve_executor_for_specialty(
        self,
        *,
        source: str,
        record_id: str,
        specialty_text: str,
        building: str,
        engineers: List[Dict[str, str]],
        emit_log: Callable[[str], None],
    ) -> str:
        supervisor = self._pick_supervisor(
            engineers,
            building=building,
            specialty_text=specialty_text,
        )
        if supervisor:
            emit_log(
                "[交接班][其他重要工作] 执行人匹配: "
                f"source={source}, record_id={record_id}, building={building}, "
                f"specialty={specialty_text or '-'}, supervisor={supervisor}"
            )
            return supervisor

        emit_log(
            "[交接班][其他重要工作] 执行人未匹配: "
            f"source={source}, record_id={record_id}, building={building}, "
            f"specialty={specialty_text or '-'}"
        )
        return "/"

    def _maintenance_equalizing_charge_rows(
        self,
        *,
        building: str,
        preloaded_rows_by_building: MaintenanceRowsByBuilding | None,
    ) -> List[MaintenanceManagementRow]:
        if preloaded_rows_by_building is None:
            return []
        return [
            row
            for row in list(preloaded_rows_by_building.get(building, []))
            if isinstance(row, MaintenanceManagementRow) and _is_equalizing_charge_notice(row.item_text)
        ]

    def _maintenance_completion_text(self) -> str:
        cfg = self.handover_cfg.get("maintenance_management_section", {})
        fixed_values = cfg.get("fixed_values", {}) if isinstance(cfg, dict) else {}
        return str(fixed_values.get("completion", "已完成")).strip() or "已完成"

    def build(
        self,
        *,
        building: str,
        duty_date: str,
        duty_shift: str,
        preloaded_rows_by_building: OtherImportantWorkRowsByBuilding | None = None,
        preloaded_maintenance_rows_by_building: MaintenanceRowsByBuilding | None = None,
        preloaded_engineers: List[Dict[str, str]] | None = None,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self.repo.get_config()
        if not bool(cfg.get("enabled", True)):
            return {}

        if preloaded_rows_by_building is None:
            try:
                rows, cfg = self.repo.list_current_shift_rows(
                    building=building,
                    duty_date=duty_date,
                    duty_shift=duty_shift,
                    emit_log=emit_log,
                )
            except Exception as exc:  # noqa: BLE001
                emit_log(f"[交接班][其他重要工作] 读取失败，按空分类继续: {exc}")
                return {}
        else:
            rows = list(preloaded_rows_by_building.get(building, []))

        if preloaded_engineers is not None:
            engineers = list(preloaded_engineers)
            emit_log(f"[交接班][其他重要工作] 命中批量预取工程师目录: count={len(engineers)}")
        else:
            try:
                engineers = self.shift_roster_repo.list_engineer_directory(
                    duty_date=duty_date,
                    duty_shift=duty_shift,
                    emit_log=emit_log,
                )
            except Exception as exc:  # noqa: BLE001
                emit_log(f"[交接班][其他重要工作] 工程师目录读取失败，执行人按 / 继续: {exc}")
                engineers = []

        sections_cfg = cfg.get("sections", {}) if isinstance(cfg.get("sections", {}), dict) else {}
        section_name = str(sections_cfg.get("other_important_work", "其他重要工作记录")).strip() or "其他重要工作记录"
        col_map = self._resolve_section_column_mapping(cfg=cfg, emit_log=emit_log).get(section_name, {})

        payload_rows: List[Dict[str, str]] = []
        matched_supervisor = 0
        unmatched_supervisor = 0
        deduped_count = 0
        maintenance_equalizing_charge_count = 0
        seen_descriptions: set[str] = set()
        for row in rows:
            description_text = str(row.description_text or "").strip()
            if not description_text:
                continue
            dedupe_key = _dedupe_text(description_text)
            if dedupe_key and dedupe_key in seen_descriptions:
                deduped_count += 1
                continue
            if dedupe_key:
                seen_descriptions.add(dedupe_key)
            executor = self._resolve_executor(row=row, building=building, engineers=engineers, emit_log=emit_log)
            if executor == "/":
                unmatched_supervisor += 1
            else:
                matched_supervisor += 1
            payload_rows.append(
                {
                    "description": description_text,
                    "completion": str(row.completion_text or "").strip() or "/",
                    "executor": executor,
                }
            )

        maintenance_completion = self._maintenance_completion_text()
        for row in self._maintenance_equalizing_charge_rows(
            building=building,
            preloaded_rows_by_building=preloaded_maintenance_rows_by_building,
        ):
            description_text = str(row.item_text or "").strip()
            if not description_text:
                continue
            dedupe_key = _dedupe_text(description_text)
            if dedupe_key and dedupe_key in seen_descriptions:
                deduped_count += 1
                continue
            if dedupe_key:
                seen_descriptions.add(dedupe_key)
            executor = self._resolve_executor_for_specialty(
                source="maintenance_equalizing_charge",
                record_id=row.record_id,
                specialty_text=row.specialty_text,
                building=building,
                engineers=engineers,
                emit_log=emit_log,
            )
            if executor == "/":
                unmatched_supervisor += 1
            else:
                matched_supervisor += 1
            maintenance_equalizing_charge_count += 1
            payload_rows.append(
                {
                    "description": description_text,
                    "completion": maintenance_completion,
                    "executor": executor,
                }
            )

        output_rows = [{"cells": self._to_cells(row_payload, col_map)} for row_payload in payload_rows]
        emit_log(
            f"[交接班][其他重要工作] 构建完成: building={building}, count={len(output_rows)}, "
            f"deduped={deduped_count}, maintenance_equalizing_charge={maintenance_equalizing_charge_count}, "
            f"matched_supervisor={matched_supervisor}, "
            f"unmatched_supervisor={unmatched_supervisor}"
        )
        return {section_name: output_rows}
