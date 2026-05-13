from __future__ import annotations

import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List

import openpyxl

from app.config.config_adapter import normalize_role_mode
from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from handover_log_module.service.power_alert_sync_service import PowerAlertSyncService


@dataclass(frozen=True)
class _MetricRow:
    source_row: int
    room: str
    machine_row: str
    point: str
    value: Any


@dataclass(frozen=True)
class _SourceBundle:
    building: str
    power_file: Path
    current_file: Path
    switch_file: Path
    metadata: Dict[str, Any]


class BranchPowerUploadService:
    """Collect branch power/current/switch values locally and upload once daily."""

    APP_TOKEN = "ASLxbfESPahdTKs0A9NccgbrnXc"
    TABLE_ID = "tblT5KbsxGCK1SwA"

    FIELD_BUILDING = "机楼"
    FIELD_ROOM = "包间"
    FIELD_ROW = "机列"
    FIELD_BRANCH_ID = "支路编号"
    FIELD_PDU_ID = "PDU编号"

    FAMILY_POWER = "branch_power_family"
    FAMILY_CURRENT = "branch_current_family"
    FAMILY_SWITCH = "branch_switch_family"

    SWITCH_UNKNOWN_VALUE = "未知"
    SWITCH_ALLOWED_VALUES = {"分闸", "合闸", SWITCH_UNKNOWN_VALUE}

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    def _emit(self, emit_log: Callable[[str], None], text: str) -> None:
        try:
            emit_log(text)
        except Exception:  # noqa: BLE001
            pass

    def _is_external_role(self) -> bool:
        deployment = self.config.get("deployment", {}) if isinstance(self.config.get("deployment", {}), dict) else {}
        return normalize_role_mode(deployment.get("role_mode")) == "external"

    @staticmethod
    def _elapsed_ms(start: float) -> int:
        return int((time.perf_counter() - start) * 1000)

    def _progress_logger(
        self,
        emit_log: Callable[[str], None],
        *,
        label: str,
        total: int,
        step: int = 1000,
    ) -> Callable[[int, int], None]:
        last_logged = {"done": 0}

        def _log(done: int, total_count: int) -> None:
            total_value = int(total_count or total or 0)
            done_value = int(done or 0)
            if done_value >= total_value or done_value - last_logged["done"] >= step:
                last_logged["done"] = done_value
                self._emit(emit_log, f"[支路功率上传] {label}进度 {done_value}/{total_value}")

        return _log

    def _client(self, emit_log: Callable[[str], None]) -> FeishuBitableClient:
        auth = require_feishu_auth_settings(self.config)
        return FeishuBitableClient(
            app_id=str(auth.get("app_id", "") or "").strip(),
            app_secret=str(auth.get("app_secret", "") or "").strip(),
            app_token=self.APP_TOKEN,
            calc_table_id=self.TABLE_ID,
            attachment_table_id=self.TABLE_ID,
            timeout=int(auth.get("timeout", 30) or 30),
            request_retry_count=int(auth.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(auth.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
            emit_log=emit_log,
        )

    @staticmethod
    def _norm_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @classmethod
    def _hour_fields(cls, hour: int) -> tuple[str, str, str]:
        return (f"功率-{hour}:00", f"电流-{hour}:00", f"开关状态-{hour}:00")

    @classmethod
    def _upload_fields(cls) -> List[str]:
        output = [
            cls.FIELD_BUILDING,
            cls.FIELD_ROOM,
            cls.FIELD_ROW,
            cls.FIELD_BRANCH_ID,
            cls.FIELD_PDU_ID,
        ]
        for hour in range(24):
            output.extend(cls._hour_fields(hour))
        return output

    @staticmethod
    def _parse_header_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value.replace(minute=0, second=0, microsecond=0)
        text = str(value or "").strip().replace("/", "-")
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H"):
            try:
                return datetime.strptime(text, fmt).replace(minute=0, second=0, microsecond=0)
            except ValueError:
                continue
        return None

    @classmethod
    def _detect_header_bucket_keys(cls, file_path: Path) -> List[str]:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except OSError:
            return []
        try:
            sheet = workbook.active
            if hasattr(sheet, "reset_dimensions"):
                sheet.reset_dimensions()
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            detected: List[str] = []
            for value in header:
                parsed = cls._parse_header_datetime(value)
                if parsed is None:
                    continue
                normalized = parsed.strftime("%Y-%m-%d %H")
                if normalized not in detected:
                    detected.append(normalized)
            return detected
        finally:
            workbook.close()

    def _resolve_hour_columns(
        self,
        sheet: Any,
        *,
        data_bucket_dts: List[datetime],
        file_path: Path,
    ) -> Dict[str, int]:
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        target_by_key = {
            item.strftime("%Y-%m-%d %H"): item.replace(minute=0, second=0, microsecond=0)
            for item in data_bucket_dts
        }
        parsed_header: Dict[datetime, int] = {}
        for index, value in enumerate(header, start=1):
            parsed = self._parse_header_datetime(value)
            if parsed is not None and parsed not in parsed_header:
                parsed_header[parsed] = index
        columns: Dict[str, int] = {}
        for key, target in target_by_key.items():
            column = parsed_header.get(target)
            if column:
                columns[key] = column
        return columns

    def _row_has_data(self, row: tuple[Any, ...], value_index: int) -> bool:
        room = self._norm_text(row[0] if len(row) > 0 else "")
        machine_row = self._norm_text(row[1] if len(row) > 1 else "")
        point = self._norm_text(row[2] if len(row) > 2 else "")
        raw_value = row[value_index] if len(row) > value_index else None
        return any([room, machine_row, point, self._norm_text(raw_value)])

    def _load_metric_rows_by_hour(
        self,
        *,
        file_path: Path,
        data_bucket_dts: List[datetime],
        metric_label: str,
    ) -> Dict[str, List[_MetricRow]]:
        rows_by_bucket = {item.strftime("%Y-%m-%d %H"): [] for item in data_bucket_dts}
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if hasattr(sheet, "reset_dimensions"):
                sheet.reset_dimensions()
            hour_columns = self._resolve_hour_columns(
                sheet,
                data_bucket_dts=data_bucket_dts,
                file_path=file_path,
            )
            max_col = max(hour_columns.values()) if hour_columns else 0
            data_rows = sheet.iter_rows(min_row=4, max_col=max_col, values_only=True)
            for index, raw_row in enumerate(data_rows, start=4):
                row = tuple(raw_row or ())
                room = self._norm_text(row[0] if len(row) > 0 else "")
                machine_row = self._norm_text(row[1] if len(row) > 1 else "")
                point = self._norm_text(row[2] if len(row) > 2 else "")
                for bucket_key, hour_col in hour_columns.items():
                    value_index = hour_col - 1
                    if not self._row_has_data(row, value_index):
                        continue
                    value = row[value_index] if len(row) > value_index else None
                    if not room or not machine_row or not point:
                        raise RuntimeError(f"{file_path.name} 第{index}行缺少包间/机列/测点")
                    if value is None or self._norm_text(value) == "":
                        continue
                    rows_by_bucket[bucket_key].append(
                        _MetricRow(
                            source_row=index,
                            room=room,
                            machine_row=machine_row,
                            point=point,
                            value=value,
                        )
                    )
        finally:
            workbook.close()
        return rows_by_bucket

    @staticmethod
    def _number_value(value: Any, *, label: str, source_row: int) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value or "").strip().replace(",", "")
        if not text:
            raise RuntimeError(f"第{source_row}行{label}为空")
        try:
            return float(text)
        except ValueError as exc:
            raise RuntimeError(f"第{source_row}行{label}不是数字: {value}") from exc

    @staticmethod
    def _extract_pdu_id(raw_point: str, *, source_row: int) -> str:
        def normalize_number(text: str) -> str:
            normalized = str(text or "").lstrip("0")
            return normalized or "0"

        text = str(raw_point or "")
        match = re.match(r"\s*(\d+)", text)
        if match:
            return normalize_number(match.group(1))
        match = re.search(r"支路\s*(\d+)\s*路", text)
        if match:
            return normalize_number(match.group(1))
        match = re.search(r"支路\s*(\d+)", text)
        if not match:
            raise RuntimeError(f"第{source_row}行无法从功率测点提取PDU编号: {raw_point}")
        return normalize_number(match.group(1))

    @staticmethod
    def _extract_branch_id(raw_point: str, *, source_row: int) -> str:
        text = str(raw_point or "").strip()
        if not text:
            raise RuntimeError(f"第{source_row}行支路开关测点为空")
        return text.split("_", 1)[0].strip()

    @classmethod
    def _normalize_switch_value(cls, value: Any) -> str:
        if isinstance(value, (int, float)) and float(value) == 0:
            return cls.SWITCH_UNKNOWN_VALUE
        text = cls._norm_text(value)
        if text:
            try:
                if float(text.replace(",", "")) == 0:
                    return cls.SWITCH_UNKNOWN_VALUE
            except ValueError:
                pass
        return text

    def _combine_rows(
        self,
        *,
        building: str,
        power_rows: List[_MetricRow],
        current_rows: List[_MetricRow],
        switch_rows: List[_MetricRow],
        hour: int,
    ) -> List[Dict[str, Any]]:
        if not power_rows or not current_rows or not switch_rows:
            raise RuntimeError(
                f"{building} 三类支路源文件数据为空: 功率={len(power_rows)}, 电流={len(current_rows)}, 开关={len(switch_rows)}"
            )
        if len(power_rows) != len(current_rows) or len(power_rows) != len(switch_rows):
            raise RuntimeError(
                f"{building} 三类支路源文件行数不一致: 功率={len(power_rows)}, 电流={len(current_rows)}, 开关={len(switch_rows)}"
            )
        power_field, current_field, switch_field = self._hour_fields(hour)
        output: List[Dict[str, Any]] = []
        seen_keys: set[tuple[str, str, str, str, str]] = set()
        for index, (power, current, switch) in enumerate(zip(power_rows, current_rows, switch_rows), start=1):
            if power.room != current.room or power.room != switch.room:
                raise RuntimeError(
                    f"{building} 第{index}条包间不一致: 功率={power.room}, 电流={current.room}, 开关={switch.room}"
                )
            if power.machine_row != current.machine_row or power.machine_row != switch.machine_row:
                raise RuntimeError(
                    f"{building} 第{index}条机列不一致: 功率={power.machine_row}, 电流={current.machine_row}, 开关={switch.machine_row}"
                )
            switch_value = self._normalize_switch_value(switch.value)
            if switch_value not in self.SWITCH_ALLOWED_VALUES:
                raise RuntimeError(f"{building} 第{switch.source_row}行开关状态无效: {switch.value}")
            power_pdu_id = self._extract_pdu_id(power.point, source_row=power.source_row)
            current_pdu_id = self._extract_pdu_id(current.point, source_row=current.source_row)
            if power_pdu_id != current_pdu_id:
                raise RuntimeError(
                    f"{building} 第{index}条功率/电流PDU编号不一致: "
                    f"功率第{power.source_row}行={power_pdu_id}, 电流第{current.source_row}行={current_pdu_id}"
                )
            branch_id = self._extract_branch_id(switch.point, source_row=switch.source_row)
            row_key = (building, power.room, power.machine_row, branch_id, power_pdu_id)
            if row_key in seen_keys:
                raise RuntimeError(f"{building} 支路源文件存在重复支路记录: {'/'.join(row_key)}")
            seen_keys.add(row_key)
            output.append(
                {
                    self.FIELD_BUILDING: building,
                    self.FIELD_ROOM: power.room,
                    self.FIELD_ROW: power.machine_row,
                    self.FIELD_BRANCH_ID: branch_id,
                    self.FIELD_PDU_ID: power_pdu_id,
                    power_field: self._number_value(power.value, label="功率", source_row=power.source_row),
                    current_field: self._number_value(current.value, label="电流", source_row=current.source_row),
                    switch_field: switch_value,
                }
            )
        return output

    def _choose_range_source_file(
        self,
        candidates: List[Path],
        *,
        target_bucket_keys: List[str],
        building: str,
        label: str,
    ) -> Path:
        unique: List[Path] = []
        seen: set[str] = set()
        for path in candidates:
            if not isinstance(path, Path):
                continue
            key = str(path)
            if key in seen:
                continue
            seen.add(key)
            unique.append(path)
        if not unique:
            raise RuntimeError(f"{building} 缺少{label}")

        target_set = set(target_bucket_keys)
        scored: List[tuple[int, int, Path, List[str]]] = []
        for path in unique:
            detected = self._detect_header_bucket_keys(path)
            coverage = len(target_set.intersection(detected))
            covers_all = 1 if target_set and target_set.issubset(set(detected)) else 0
            scored.append((covers_all, coverage, path, detected))
        scored.sort(key=lambda item: (item[0], item[1]), reverse=True)
        best = scored[0]
        if best[1] <= 0:
            raise RuntimeError(f"{building} {label}未覆盖目标小时: {','.join(target_bucket_keys)}")
        return best[2]

    def _normalize_range_source_bundles(
        self,
        source_units: List[Dict[str, Any]],
        *,
        data_bucket_dts: List[datetime],
    ) -> List[_SourceBundle]:
        target_bucket_keys = [item.strftime("%Y-%m-%d %H") for item in data_bucket_dts]
        grouped: Dict[str, Dict[str, Any]] = {}
        for unit in source_units:
            if not isinstance(unit, dict):
                continue
            building = self._norm_text(unit.get("building"))
            if not building:
                continue
            bucket = grouped.setdefault(
                building,
                {"building": building, "metadata": {}, "power_files": [], "current_files": [], "switch_files": []},
            )
            metadata = unit.get("metadata", {}) if isinstance(unit.get("metadata", {}), dict) else {}
            if metadata and not bucket.get("metadata"):
                bucket["metadata"] = metadata
            source_files = unit.get("source_files", {}) if isinstance(unit.get("source_files", {}), dict) else {}
            for key, list_key in (
                ("power_file", "power_files"),
                ("current_file", "current_files"),
                ("switch_file", "switch_files"),
            ):
                raw = unit.get(key) or source_files.get(key)
                if raw:
                    bucket[list_key].append(Path(str(raw).strip()))
            raw_source = unit.get("source_file") or unit.get("file_path")
            if raw_source:
                source_family = self._norm_text(unit.get("source_family") or metadata.get("family"))
                list_key = "power_files"
                if source_family == self.FAMILY_CURRENT:
                    list_key = "current_files"
                elif source_family == self.FAMILY_SWITCH:
                    list_key = "switch_files"
                bucket[list_key].append(Path(str(raw_source).strip()))

        bundles: List[_SourceBundle] = []
        for payload in grouped.values():
            building = self._norm_text(payload.get("building"))
            power_file = self._choose_range_source_file(
                payload.get("power_files", []) if isinstance(payload.get("power_files", []), list) else [],
                target_bucket_keys=target_bucket_keys,
                building=building,
                label="支路功率源文件",
            )
            current_candidates = payload.get("current_files", []) if isinstance(payload.get("current_files", []), list) else []
            switch_candidates = payload.get("switch_files", []) if isinstance(payload.get("switch_files", []), list) else []
            current_file = self._choose_range_source_file(
                current_candidates,
                target_bucket_keys=target_bucket_keys,
                building=building,
                label="支路电流源文件",
            )
            switch_file = self._choose_range_source_file(
                switch_candidates,
                target_bucket_keys=target_bucket_keys,
                building=building,
                label="支路开关源文件",
            )
            bundles.append(
                _SourceBundle(
                    building=building,
                    power_file=power_file,
                    current_file=current_file,
                    switch_file=switch_file,
                    metadata=payload.get("metadata", {}) if isinstance(payload.get("metadata", {}), dict) else {},
                )
            )
        return bundles

    def _target_fields_from_bitable(
        self,
        client: FeishuBitableClient,
        emit_log: Callable[[str], None],
    ) -> List[str]:
        fields = client.list_fields(self.TABLE_ID)
        names = [self._norm_text(field.get("field_name") or field.get("name")) for field in fields if isinstance(field, dict)]
        required = self._upload_fields()
        missing = [field for field in required if field not in names]
        if missing:
            raise RuntimeError(f"支路功率目标多维表缺少字段: {', '.join(missing[:20])}")
        self._emit(emit_log, f"[支路功率上传] 目标表字段校验完成 fields={len(names)}, upload_fields={len(required)}")
        return required

    def _parse_bundle_rows_by_hour(
        self,
        *,
        bundle: _SourceBundle,
        data_bucket_dts: List[datetime],
        emit_log: Callable[[str], None],
    ) -> tuple[Dict[str, List[Dict[str, Any]]], Dict[str, str]]:
        building = bundle.building
        read_start = time.perf_counter()
        power_rows_by_hour = self._load_metric_rows_by_hour(
            file_path=bundle.power_file,
            data_bucket_dts=data_bucket_dts,
            metric_label="功率",
        )
        current_rows_by_hour = self._load_metric_rows_by_hour(
            file_path=bundle.current_file,
            data_bucket_dts=data_bucket_dts,
            metric_label="电流",
        )
        switch_rows_by_hour = self._load_metric_rows_by_hour(
            file_path=bundle.switch_file,
            data_bucket_dts=data_bucket_dts,
            metric_label="开关状态",
        )
        output: Dict[str, List[Dict[str, Any]]] = {}
        failures: Dict[str, str] = {}
        summary: List[str] = []
        for data_bucket_dt in data_bucket_dts:
            data_bucket_key = data_bucket_dt.strftime("%Y-%m-%d %H")
            power_rows = power_rows_by_hour.get(data_bucket_key, [])
            current_rows = current_rows_by_hour.get(data_bucket_key, [])
            switch_rows = switch_rows_by_hour.get(data_bucket_key, [])
            try:
                output[data_bucket_key] = self._combine_rows(
                    building=building,
                    power_rows=power_rows,
                    current_rows=current_rows,
                    switch_rows=switch_rows,
                    hour=data_bucket_dt.hour,
                )
            except Exception as exc:  # noqa: BLE001
                failures[data_bucket_key] = str(exc)
            summary.append(
                f"{data_bucket_key}:功率={len(power_rows)},电流={len(current_rows)},开关={len(switch_rows)}"
            )
        self._emit(
            emit_log,
            f"[支路功率上传] 范围三源文件读取完成 building={building}, "
            f"hours={';'.join(summary)}, elapsed_ms={self._elapsed_ms(read_start)}",
        )
        return output, failures

    @staticmethod
    def _parse_business_date(value: Any) -> datetime:
        text = str(value or "").strip().replace("/", "-")
        for fmt in ("%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        raise ValueError(f"业务日期无效: {value}")

    def _validate_daily_direct_records(
        self,
        *,
        records: List[Dict[str, Any]],
        upload_fields: List[str],
        expected_buildings: List[str],
        business_date: str,
    ) -> Dict[str, Any]:
        key_fields = [
            self.FIELD_BUILDING,
            self.FIELD_ROOM,
            self.FIELD_ROW,
            self.FIELD_BRANCH_ID,
            self.FIELD_PDU_ID,
        ]
        missing: List[str] = []
        row_count_by_building: Dict[str, int] = {}
        missing_hour_fields = [
            field
            for hour in range(24)
            for field in self._hour_fields(hour)
            if field not in upload_fields
        ]
        if missing_hour_fields:
            missing.append(f"目标多维表缺少小时字段: {','.join(missing_hour_fields[:20])}")
        for row in records:
            key_parts = [self._norm_text(row.get(field)) for field in key_fields]
            building = key_parts[0]
            if building:
                row_count_by_building[building] = row_count_by_building.get(building, 0) + 1
            key_label = "/".join(part or "-" for part in key_parts)
            for field, value in zip(key_fields, key_parts):
                if not value:
                    missing.append(f"{key_label}/{field}为空")
                    if len(missing) >= 100:
                        break
            if len(missing) >= 100:
                break
        for building in expected_buildings:
            if row_count_by_building.get(building, 0) <= 0:
                missing.append(f"{building}/无整日解析记录")
        return {
            "ok": not missing,
            "business_date": business_date,
            "missing_count": len(missing),
            "missing": missing[:100],
            "row_count": len(records),
            "building_rows": row_count_by_building,
        }

    def upload_day_from_source_files(
        self,
        *,
        business_date: str,
        source_units: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        total_start = time.perf_counter()
        day_dt = self._parse_business_date(business_date)
        resolved_business_date = day_dt.strftime("%Y-%m-%d")
        data_bucket_dts = [day_dt.replace(hour=hour, minute=0, second=0, microsecond=0) for hour in range(24)]
        units = [unit for unit in (source_units or []) if isinstance(unit, dict)]
        if not units:
            raise RuntimeError(f"没有可上传的支路整日源文件 business_date={resolved_business_date}")
        bundles = self._normalize_range_source_bundles(units, data_bucket_dts=data_bucket_dts)
        if not bundles:
            raise RuntimeError(f"没有可上传的支路整日源文件 bundle business_date={resolved_business_date}")
        self._emit(
            emit_log,
            f"[支路功率上传] 整日直传开始 business_date={resolved_business_date}, buildings={len(bundles)}, "
            f"hours=00:00-23:00",
        )

        records_by_key: Dict[tuple[str, str, str, str, str], Dict[str, Any]] = {}
        parse_failures: List[Dict[str, Any]] = []
        parsed_hour_count = 0
        for bundle in bundles:
            rows_by_bucket, failures = self._parse_bundle_rows_by_hour(
                bundle=bundle,
                data_bucket_dts=data_bucket_dts,
                emit_log=emit_log,
            )
            for bucket_key, error_text in failures.items():
                parse_failures.append(
                    {
                        "business_date": resolved_business_date,
                        "bucket_key": bucket_key,
                        "building": bundle.building,
                        "error": str(error_text or "").strip(),
                    }
                )
            for data_bucket_dt in data_bucket_dts:
                bucket_key = data_bucket_dt.strftime("%Y-%m-%d %H")
                rows = rows_by_bucket.get(bucket_key, [])
                if failures.get(bucket_key):
                    continue
                if not rows:
                    parse_failures.append(
                        {
                            "business_date": resolved_business_date,
                            "bucket_key": bucket_key,
                            "building": bundle.building,
                            "error": "目标小时无有效数据",
                        }
                    )
                    continue
                parsed_hour_count += 1
                for row in rows:
                    key = self._row_key(row)
                    if not all(key):
                        parse_failures.append(
                            {
                                "business_date": resolved_business_date,
                                "bucket_key": bucket_key,
                                "building": bundle.building,
                                "error": "支路记录关键字段为空",
                            }
                        )
                        continue
                    record = records_by_key.setdefault(
                        key,
                        {
                            self.FIELD_BUILDING: key[0],
                            self.FIELD_ROOM: key[1],
                            self.FIELD_ROW: key[2],
                            self.FIELD_BRANCH_ID: key[3],
                            self.FIELD_PDU_ID: key[4],
                        },
                    )
                    for field in self._hour_fields(data_bucket_dt.hour):
                        value = row.get(field)
                        if value is not None and self._norm_text(value) != "":
                            record[field] = value

        if parse_failures:
            detail = "; ".join(
                f"{item.get('bucket_key', '-')}/{item.get('building', '-')}: {item.get('error', '-')}"
                for item in parse_failures[:30]
            )
            suffix = f" 等{len(parse_failures)}项" if len(parse_failures) > 30 else ""
            raise RuntimeError(f"支路整日源文件解析存在失败，未清空多维表: {detail}{suffix}")
        records = sorted(
            records_by_key.values(),
            key=lambda row: (
                self._norm_text(row.get(self.FIELD_BUILDING)),
                self._norm_text(row.get(self.FIELD_ROOM)),
                self._norm_text(row.get(self.FIELD_ROW)),
                self._norm_text(row.get(self.FIELD_BRANCH_ID)),
                self._norm_text(row.get(self.FIELD_PDU_ID)),
            ),
        )
        if not records:
            raise RuntimeError(f"支路整日源文件没有解析出可上传记录 business_date={resolved_business_date}")

        client = self._client(emit_log)
        upload_fields = self._target_fields_from_bitable(client, emit_log)
        validation = self._validate_daily_direct_records(
            records=records,
            upload_fields=upload_fields,
            expected_buildings=[bundle.building for bundle in bundles],
            business_date=resolved_business_date,
        )
        if not validation.get("ok"):
            self._emit(
                emit_log,
                f"[支路功率上传] 整日直传校验失败 business_date={resolved_business_date}, "
                f"row_count={validation.get('row_count', 0)}, missing_count={validation.get('missing_count', 0)}, "
                f"missing_sample={validation.get('missing', [])[:20]}",
            )
            raise RuntimeError(f"支路整日解析结果字段未齐全，未清空多维表: {validation.get('missing', [])[:20]}")

        upload_records: List[Dict[str, Any]] = []
        for row in records:
            payload = {
                field: row.get(field)
                for field in upload_fields
                if row.get(field) is not None and self._norm_text(row.get(field)) != ""
            }
            if all(self._norm_text(payload.get(field)) for field in [
                self.FIELD_BUILDING,
                self.FIELD_ROOM,
                self.FIELD_ROW,
                self.FIELD_BRANCH_ID,
                self.FIELD_PDU_ID,
            ]):
                upload_records.append(payload)
        if not upload_records:
            raise RuntimeError(f"支路整日解析结果为空，无法上传 business_date={resolved_business_date}")

        clear_start = time.perf_counter()
        self._emit(emit_log, f"[支路功率上传] 整日直传清空目标多维表开始 table_id={self.TABLE_ID}")
        deleted = client.clear_table(
            self.TABLE_ID,
            progress_callback=self._progress_logger(emit_log, label="整日直传清空目标表", total=1),
        )
        self._emit(
            emit_log,
            f"[支路功率上传] 整日直传清空目标多维表完成 deleted={deleted}, elapsed_ms={self._elapsed_ms(clear_start)}",
        )

        create_start = time.perf_counter()
        self._emit(emit_log, f"[支路功率上传] 整日直传批量上传开始 records={len(upload_records)}")
        client.batch_create_records(
            self.TABLE_ID,
            upload_records,
            progress_callback=self._progress_logger(emit_log, label="整日直传上传", total=len(upload_records)),
        )
        elapsed_ms = self._elapsed_ms(total_start)
        self._emit(
            emit_log,
            f"[支路功率上传] 整日直传批量上传完成 business_date={resolved_business_date}, "
            f"records={len(upload_records)}, parsed_hours={parsed_hour_count}, elapsed_ms={elapsed_ms}",
        )
        power_alert_sync = PowerAlertSyncService(self.config).sync(
            report_date=resolved_business_date,
            emit_log=emit_log,
        )
        return {
            "ok": True,
            "status": "success",
            "mode": "daily_direct_upload",
            "business_date": resolved_business_date,
            "parsed_hours": parsed_hour_count,
            "buildings": [bundle.building for bundle in bundles],
            "records": len(upload_records),
            "deleted": deleted,
            "elapsed_ms": elapsed_ms,
            "power_alert_sync": power_alert_sync,
        }

    @classmethod
    def _row_key(cls, row: Dict[str, Any]) -> tuple[str, str, str, str, str]:
        return (
            cls._norm_text(row.get(cls.FIELD_BUILDING)),
            cls._norm_text(row.get(cls.FIELD_ROOM)),
            cls._norm_text(row.get(cls.FIELD_ROW)),
            cls._norm_text(row.get(cls.FIELD_BRANCH_ID)),
            cls._norm_text(row.get(cls.FIELD_PDU_ID)),
        )
