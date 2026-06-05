from __future__ import annotations

import copy
import re
import time
import warnings
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from app.shared.utils.atomic_file import atomic_save_workbook
from app.shared.utils.file_utils import fallback_missing_windows_drive_path
from pipeline_utils import get_app_dir


_ALL_BUILDINGS = ["A楼", "B楼", "C楼", "D楼", "E楼"]
_BUILDING_CODES = {"A楼": "A", "B楼": "B", "C楼": "C", "D楼": "D", "E楼": "E"}
_SUMMARY_SHEET_NAME = "汇总信息表"
_TOP_N = 5
_DEFAULT_TEMPLATE_SOURCE_PATH = "阿里月报高功率TOP5报表模板.xlsx"
_LEGACY_TEMPLATE_SOURCE_PATHS = {"", "TOP5功率报表空模板.xlsx"}
_HEADERS = [
    "序号",
    "楼栋",
    "变压器编号",
    "变压器功率（KW）",
    "HVDC编号",
    "HVDC功率（KW）",
    "机列编号",
    "机列负载（KW）",
    "UPS编号",
    "UPS功率（KW）",
]
_HEADER_FILL = PatternFill("solid", fgColor="FFFF00")
_DATA_FILL = PatternFill("solid", fgColor="C6E0B4")
_TOP5_RED_FILL = PatternFill("solid", fgColor="FFFF0000")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="宋体", size=20, bold=True)
_HEADER_FONT = Font(name="宋体", size=11, bold=True)
_DATA_FONT = Font(name="宋体", size=11)
_MONTHLY_TRANSFORMER_PATTERN = re.compile(r"([A-E]-\d{3}-[AB]变压器容量)$", re.IGNORECASE)
_MONTHLY_HVDC_PATTERN = re.compile(r"(?:([A-E])-)?(\d{3}-HVDC-\d+)$", re.IGNORECASE)
_MONTHLY_UPS_PATTERN = re.compile(r"(?:([A-E])-)?(\d{3}-UPS-\d+(?:_UPS)?)$", re.IGNORECASE)
_MONTHLY_CABINET_PATTERN = re.compile(r"([A-E])-(\d{3})-([A-Z])(\d{2})$", re.IGNORECASE)


@dataclass(frozen=True)
class PowerRecord:
    building: str
    identifier: str
    power_kw: float
    source_file: str = ""


@dataclass(frozen=True)
class RowLineAggregate:
    building: str
    room: str
    column: str
    identifier: str
    cabinet_count: int
    cabinet_ids: List[str]
    power_kw: float
    source_file: str = ""


@dataclass(frozen=True)
class MonthlyTop5Groups:
    transformers: List[PowerRecord]
    hvdcs: List[PowerRecord]
    row_lines: List[PowerRecord]
    upss: List[PowerRecord]
    all_transformers: List[PowerRecord]
    all_hvdcs: List[PowerRecord]
    all_row_lines: List[PowerRecord]
    all_upss: List[PowerRecord]
    row_line_aggregates: List[RowLineAggregate]


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    merged = copy.deepcopy(base or {})
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = copy.deepcopy(value)
    return merged


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_building(value: Any) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    if text in _ALL_BUILDINGS:
        return text
    if len(text) == 1 and text in {"A", "B", "C", "D", "E"}:
        return f"{text}楼"
    match = re.search(r"([A-E])\s*楼", text, re.IGNORECASE)
    if match:
        return f"{match.group(1).upper()}楼"
    return str(value or "").strip()


def _building_code(building: str) -> str:
    normalized = _normalize_building(building)
    return _BUILDING_CODES.get(normalized, normalized[:1].upper())


def _dedupe_records(records: Iterable[PowerRecord]) -> List[PowerRecord]:
    best: Dict[str, PowerRecord] = {}
    for record in records:
        key = record.identifier
        current = best.get(key)
        if current is None or record.power_kw > current.power_kw:
            best[key] = record
    return list(best.values())


def _sorted_records(records: Iterable[PowerRecord]) -> List[PowerRecord]:
    return sorted(
        _dedupe_records(records),
        key=lambda item: (-item.power_kw, item.identifier),
    )


def _top_records(records: Iterable[PowerRecord], *, building: str, group_name: str) -> List[PowerRecord]:
    sorted_records = _sorted_records(records)
    if len(sorted_records) < _TOP_N:
        raise RuntimeError(f"{building}{group_name}有效数据不足{_TOP_N}条，当前{len(sorted_records)}条")
    return sorted_records[:_TOP_N]


def _format_power(value: float) -> float:
    return round(float(value), 2)


def _monthly_entry_date(entry: Dict[str, Any]) -> str:
    metadata = entry.get("metadata", {}) if isinstance(entry.get("metadata", {}), dict) else {}
    for key in ("upload_date", "duty_date", "bucket_key"):
        text = str(metadata.get(key, "") or entry.get(key, "") or "").strip()
        if re.fullmatch(r"20\d{2}-\d{2}-\d{2}", text):
            return text
    return ""


def _load_rows(path: Path) -> List[List[Any]]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _load_workbook(path: Path):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        return openpyxl.load_workbook(path, data_only=False)


def _clone_image(image: Any) -> OpenpyxlImage | None:
    try:
        raw = image._data()
    except Exception:  # noqa: BLE001
        return None
    cloned = OpenpyxlImage(BytesIO(raw))
    try:
        cloned.width = image.width
        cloned.height = image.height
    except Exception:  # noqa: BLE001
        pass
    try:
        cloned.anchor = copy.copy(image.anchor)
    except Exception:  # noqa: BLE001
        return None
    return cloned


def _monthly_transformer_label(text: str) -> str:
    match = _MONTHLY_TRANSFORMER_PATTERN.search(str(text or "").strip())
    return match.group(1).upper() if match else ""


def _monthly_hvdc_label(text: str, *, building: str) -> str:
    match = _MONTHLY_HVDC_PATTERN.search(str(text or "").strip())
    if not match:
        return ""
    if match.group(1):
        return f"{match.group(1).upper()}-{match.group(2).upper()}"
    return match.group(2).upper()


def _monthly_ups_label(text: str, *, building: str) -> str:
    match = _MONTHLY_UPS_PATTERN.search(str(text or "").strip())
    if not match:
        return ""
    prefix = match.group(1).upper() if match.group(1) else _building_code(building)
    suffix = match.group(2).upper()
    return suffix if suffix.startswith(f"{prefix}-") else f"{prefix}-{suffix}"


def _monthly_cabinet_parts(text: str) -> tuple[str, str, str, str] | None:
    match = _MONTHLY_CABINET_PATTERN.fullmatch(str(text or "").strip().upper())
    if not match:
        return None
    return match.group(1), match.group(2), match.group(3), match.group(4)


class Top5PowerReportService:
    def __init__(self, runtime_config: Dict[str, Any]) -> None:
        self.runtime_config = runtime_config if isinstance(runtime_config, dict) else {}

    @staticmethod
    def all_buildings() -> List[str]:
        return list(_ALL_BUILDINGS)

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "enabled": True,
            "template": {
                "source_path": _DEFAULT_TEMPLATE_SOURCE_PATH,
                "output_dir": r"D:\QLDownload\TOP5功率文件生成",
                "file_name_pattern": "TOP5功率文件_{timestamp}.xlsx",
            },
            "over_power_attachment": {
                "enabled": True,
                "app_token": "MliKbC3fXa8PXrsndKscmxjdn1g",
                "table_id": "tblkh6YCMYtS8nHa",
                "view_id": "vewrHJHl3v",
                "output_dir": r"D:\QLDownload\月度超功率附件",
                "zip_file_name_pattern": "月度超功率附件_{year}{month}_{timestamp}.zip",
            },
            "report_upload": {
                "enabled": True,
                "app_token": "MliKbC3fXa8PXrsndKscmxjdn1g",
                "table_id": "tblkh6YCMYtS8nHa",
                "sub_category": "高功率TOP5",
                "fields": {
                    "sub_category": "子分类",
                    "year": "年度",
                    "month": "月份",
                    "attachment": "上传文件",
                    "link": "链接",
                },
            },
        }

    def _normalize_cfg(self) -> Dict[str, Any]:
        handover_cfg = self.runtime_config.get("handover_log", {})
        if not isinstance(handover_cfg, dict):
            handover_cfg = {}
        raw_cfg = handover_cfg.get("top5_power_report", {})
        cfg = _deep_merge(self._defaults(), raw_cfg if isinstance(raw_cfg, dict) else {})
        cfg["enabled"] = bool(cfg.get("enabled", True))
        template = cfg.get("template", {}) if isinstance(cfg.get("template", {}), dict) else {}
        source_path = str(template.get("source_path", "") or "").strip()
        if source_path in _LEGACY_TEMPLATE_SOURCE_PATHS:
            source_path = _DEFAULT_TEMPLATE_SOURCE_PATH
        template["source_path"] = source_path
        template["output_dir"] = (
            str(template.get("output_dir", "") or "").strip() or r"D:\QLDownload\TOP5功率文件生成"
        )
        template["file_name_pattern"] = (
            str(template.get("file_name_pattern", "") or "").strip() or "TOP5功率文件_{timestamp}.xlsx"
        )
        cfg["template"] = template
        return cfg

    def is_enabled(self) -> bool:
        return bool(self._normalize_cfg().get("enabled", True))

    def _app_dir(self) -> Path:
        return get_app_dir()

    def _resolve_path(self, value: str) -> Path:
        path = Path(str(value or "").strip())
        if path.is_absolute():
            return fallback_missing_windows_drive_path(path, app_dir=self._app_dir())
        return self._app_dir() / path

    def resolve_template_path(self) -> Path:
        return self._resolve_path(self._normalize_cfg()["template"]["source_path"])

    def resolve_output_dir(self) -> Path:
        output_dir = self._resolve_path(self._normalize_cfg()["template"]["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    def _build_output_path(self) -> Path:
        cfg = self._normalize_cfg()["template"]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = str(cfg["file_name_pattern"]).format(
            timestamp=timestamp,
            date=datetime.now().strftime("%Y%m%d"),
        )
        if not file_name.lower().endswith(".xlsx"):
            file_name = f"{file_name}.xlsx"
        candidate = self.resolve_output_dir() / file_name
        index = 2
        while candidate.exists():
            candidate = candidate.with_name(f"{candidate.stem}_{index}{candidate.suffix}")
            index += 1
        return candidate

    @staticmethod
    def extract_monthly_top5_groups(path: Path | str, *, building: str) -> MonthlyTop5Groups:
        source_path = Path(path)
        rows = _load_rows(source_path)
        normalized_building = _normalize_building(building)
        transformers: List[PowerRecord] = []
        hvdcs: List[PowerRecord] = []
        upss: List[PowerRecord] = []
        row_line_groups: Dict[str, Dict[str, Any]] = {}

        for row in rows:
            device_text = _cell_text(row[1] if len(row) > 1 else "")
            actual_load = _to_float(row[4] if len(row) > 4 else None)

            transformer_id = _monthly_transformer_label(device_text)
            if transformer_id and actual_load is not None:
                transformers.append(PowerRecord(normalized_building, transformer_id, actual_load, str(source_path)))
                continue

            hvdc_id = _monthly_hvdc_label(device_text, building=normalized_building)
            if hvdc_id and actual_load is not None:
                hvdcs.append(PowerRecord(normalized_building, hvdc_id, actual_load, str(source_path)))
                continue

            ups_id = _monthly_ups_label(device_text, building=normalized_building)
            if ups_id and actual_load is not None:
                upss.append(PowerRecord(normalized_building, ups_id, actual_load, str(source_path)))
                continue

            cabinet_parts = _monthly_cabinet_parts(device_text)
            if cabinet_parts is None or actual_load is None:
                continue
            _building, room, column, _cabinet_no = cabinet_parts
            identifier = f"{_building_code(normalized_building)}-{room}-{column}列功率和"
            group = row_line_groups.setdefault(
                identifier,
                {
                    "building": normalized_building,
                    "room": room,
                    "column": column,
                    "identifier": identifier,
                    "cabinet_ids": [],
                    "power_kw": 0.0,
                },
            )
            group["cabinet_ids"].append(device_text.strip().upper())
            group["power_kw"] += float(actual_load)

        row_line_aggregates = [
            RowLineAggregate(
                building=str(item["building"]),
                room=str(item["room"]),
                column=str(item["column"]),
                identifier=str(item["identifier"]),
                cabinet_count=len(item["cabinet_ids"]),
                cabinet_ids=list(item["cabinet_ids"]),
                power_kw=float(item["power_kw"]),
                source_file=str(source_path),
            )
            for item in row_line_groups.values()
        ]
        row_lines = [
            PowerRecord(
                normalized_building,
                item.identifier,
                float(item.power_kw),
                str(source_path),
            )
            for item in row_line_aggregates
        ]
        return MonthlyTop5Groups(
            transformers=_top_records(transformers, building=normalized_building, group_name="变压器"),
            hvdcs=_top_records(hvdcs, building=normalized_building, group_name="HVDC"),
            row_lines=_top_records(row_lines, building=normalized_building, group_name="机列"),
            upss=_top_records(upss, building=normalized_building, group_name="UPS"),
            all_transformers=_sorted_records(transformers),
            all_hvdcs=_sorted_records(hvdcs),
            all_row_lines=_sorted_records(row_lines),
            all_upss=_sorted_records(upss),
            row_line_aggregates=sorted(
                row_line_aggregates,
                key=lambda item: (-item.power_kw, item.identifier),
            ),
        )

    @staticmethod
    def _entries_by_building(entries: List[Dict[str, Any]], *, label: str) -> Dict[str, Dict[str, Any]]:
        result: Dict[str, Dict[str, Any]] = {}
        for entry in entries if isinstance(entries, list) else []:
            if not isinstance(entry, dict):
                continue
            building = _normalize_building(entry.get("building"))
            file_path = str(entry.get("file_path", "") or "").strip()
            if not building or not file_path:
                continue
            result[building] = {**entry, "building": building, "file_path": file_path}
        missing = [building for building in _ALL_BUILDINGS if building not in result]
        if missing:
            raise RuntimeError(f"缺少{label}最新源文件: {', '.join(missing)}")
        return result

    @staticmethod
    def _style_cell(cell, *, is_header: bool = False, is_data_fill: bool = False, fill: PatternFill | None = None) -> None:
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
        cell.font = _HEADER_FONT if is_header else _DATA_FONT
        if fill is not None:
            cell.fill = fill
        elif is_header:
            cell.fill = _HEADER_FILL
        elif is_data_fill:
            cell.fill = _DATA_FILL

    def _create_summary_workbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = _SUMMARY_SHEET_NAME
        self._format_summary_sheet(sheet)
        return workbook

    def _load_workbook_for_output(self):
        template_path = self.resolve_template_path()
        if template_path.exists():
            workbook = openpyxl.load_workbook(template_path)
            if _SUMMARY_SHEET_NAME not in workbook.sheetnames:
                workbook.active.title = _SUMMARY_SHEET_NAME
            return workbook
        return self._create_summary_workbook()

    def _format_summary_sheet(self, sheet) -> None:
        if "A1:J1" not in [str(range_ref) for range_ref in sheet.merged_cells.ranges]:
            sheet.merge_cells("A1:J1")
        title_cell = sheet["A1"]
        title_cell.value = "高功率设备排名（TOP5）"
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER
        for column, header in enumerate(_HEADERS, start=1):
            cell = sheet.cell(row=2, column=column, value=header)
            self._style_cell(cell, is_header=True)
        widths = {
            "A": 8,
            "B": 10,
            "C": 24,
            "D": 18,
            "E": 24,
            "F": 18,
            "G": 28,
            "H": 18,
            "I": 24,
            "J": 18,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 22

    def _prepare_summary_sheet(self, workbook):
        sheet = workbook[_SUMMARY_SHEET_NAME] if _SUMMARY_SHEET_NAME in workbook.sheetnames else workbook.active
        sheet.title = _SUMMARY_SHEET_NAME
        for worksheet in list(workbook.worksheets):
            if worksheet.title != _SUMMARY_SHEET_NAME:
                workbook.remove(worksheet)
        self._format_summary_sheet(sheet)
        for row_index in range(3, max(sheet.max_row, 27) + 1):
            for column_index in range(1, 11):
                sheet.cell(row=row_index, column=column_index).value = None
        return sheet

    def _write_summary_rows(
        self,
        sheet,
        *,
        monthly_by_building: Dict[str, MonthlyTop5Groups],
    ) -> None:
        row_index = 3
        sequence = 1
        for building in _ALL_BUILDINGS:
            groups = monthly_by_building[building]
            for index in range(_TOP_N):
                values = [
                    sequence,
                    _building_code(building),
                    groups.transformers[index].identifier,
                    _format_power(groups.transformers[index].power_kw),
                    groups.hvdcs[index].identifier,
                    _format_power(groups.hvdcs[index].power_kw),
                    groups.row_lines[index].identifier,
                    _format_power(groups.row_lines[index].power_kw),
                    groups.upss[index].identifier,
                    _format_power(groups.upss[index].power_kw),
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_index, column=column_index, value=value)
                    self._style_cell(cell, is_data_fill=column_index >= 3)
                    if column_index in {4, 6, 8, 10}:
                        cell.number_format = "0.00"
                row_index += 1
                sequence += 1

    def _append_source_sheet(self, workbook, *, sheet_name: str, source_path: Path) -> int:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        target_sheet = workbook.create_sheet(sheet_name)
        source_workbook = _load_workbook(source_path)
        try:
            source_sheet = source_workbook.active
            if hasattr(source_sheet, "reset_dimensions"):
                source_sheet.reset_dimensions()
            for row in source_sheet.iter_rows():
                for source_cell in row:
                    target_cell = target_sheet.cell(
                        row=source_cell.row,
                        column=source_cell.column,
                        value=source_cell.value,
                    )
                    if source_cell.has_style:
                        target_cell._style = copy.copy(source_cell._style)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                    if source_cell.font is not None:
                        target_cell.font = copy.copy(source_cell.font)
                    if source_cell.fill is not None:
                        target_cell.fill = copy.copy(source_cell.fill)
                    if source_cell.border is not None:
                        target_cell.border = copy.copy(source_cell.border)
                    if source_cell.alignment is not None:
                        target_cell.alignment = copy.copy(source_cell.alignment)
                    if source_cell.protection is not None:
                        target_cell.protection = copy.copy(source_cell.protection)
                    if source_cell.hyperlink:
                        target_cell._hyperlink = copy.copy(source_cell.hyperlink)
                    if source_cell.comment:
                        target_cell.comment = copy.copy(source_cell.comment)
            for merged_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_range))
            for key, dimension in source_sheet.column_dimensions.items():
                target_dimension = target_sheet.column_dimensions[key]
                target_dimension.width = dimension.width
                target_dimension.hidden = dimension.hidden
                target_dimension.bestFit = dimension.bestFit
                target_dimension.outlineLevel = dimension.outlineLevel
                target_dimension.collapsed = dimension.collapsed
            for key, dimension in source_sheet.row_dimensions.items():
                target_dimension = target_sheet.row_dimensions[key]
                target_dimension.height = dimension.height
                target_dimension.hidden = dimension.hidden
                target_dimension.outlineLevel = dimension.outlineLevel
                target_dimension.collapsed = dimension.collapsed
            target_sheet.freeze_panes = source_sheet.freeze_panes
            target_sheet.sheet_state = source_sheet.sheet_state
            if getattr(source_sheet.auto_filter, "ref", None):
                target_sheet.auto_filter.ref = source_sheet.auto_filter.ref
            for image in list(getattr(source_sheet, "_images", []) or []):
                cloned_image = _clone_image(image)
                if cloned_image is not None:
                    target_sheet.add_image(cloned_image)
            return int(source_sheet.max_row or 0)
        finally:
            source_workbook.close()

    def _append_building_detail_sheet(
        self,
        workbook,
        *,
        sheet_name: str,
        groups: MonthlyTop5Groups,
    ) -> int:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        sheet = workbook.create_sheet(sheet_name)
        blocks = [
            (1, groups.all_transformers),
            (4, groups.all_hvdcs),
            (7, groups.all_upss),
            (10, groups.all_row_lines),
        ]
        for start_column, records in blocks:
            for offset, header in enumerate(("地点", "值")):
                cell = sheet.cell(row=1, column=start_column + offset, value=header)
                self._style_cell(cell, is_header=True)
            for index, record in enumerate(records, start=2):
                name_cell = sheet.cell(row=index, column=start_column, value=record.identifier)
                value_cell = sheet.cell(row=index, column=start_column + 1, value=_format_power(record.power_kw))
                highlight_fill = _TOP5_RED_FILL if index <= 6 else None
                self._style_cell(name_cell, fill=highlight_fill)
                self._style_cell(value_cell, fill=highlight_fill)
                value_cell.number_format = "0.00"
        sheet.freeze_panes = "A2"
        for column in ("A", "D", "G", "J"):
            sheet.column_dimensions[column].width = 24
        for column in ("B", "E", "H", "K"):
            sheet.column_dimensions[column].width = 12
        for column in ("C", "F", "I"):
            sheet.column_dimensions[column].width = 4
        return max(
            len(groups.all_transformers),
            len(groups.all_hvdcs),
            len(groups.all_upss),
            len(groups.all_row_lines),
        )

    def _append_sheet1(self, workbook, *, records: List[PowerRecord]) -> int:
        if "Sheet1" in workbook.sheetnames:
            workbook.remove(workbook["Sheet1"])
        sheet = workbook.create_sheet("Sheet1")
        for row_index, record in enumerate(records, start=1):
            sheet.cell(row=row_index, column=1, value=record.identifier)
            value_cell = sheet.cell(row=row_index, column=2, value=_format_power(record.power_kw))
            value_cell.number_format = "0.00"
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 12
        return len(records)

    def run(
        self,
        *,
        monthly_entries: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        if not self.is_enabled():
            raise RuntimeError("TOP5功率文件生成已禁用")

        started_at = datetime.now()
        emit_log("[TOP5功率文件生成] 开始读取共享源文件")
        monthly_entry_map = self._entries_by_building(monthly_entries, label="TOP5月报")

        monthly_by_building: Dict[str, MonthlyTop5Groups] = {}
        for building in _ALL_BUILDINGS:
            monthly_path = Path(monthly_entry_map[building]["file_path"])
            emit_log(f"[TOP5功率文件生成] 解析{building}: monthly={monthly_path.name}")
            monthly_by_building[building] = self.extract_monthly_top5_groups(monthly_path, building=building)

        output_path = self._build_output_path()
        workbook = self._load_workbook_for_output()
        try:
            summary_sheet = self._prepare_summary_sheet(workbook)
            self._write_summary_rows(
                summary_sheet,
                monthly_by_building=monthly_by_building,
            )
            source_sheet_rows: Dict[str, int] = {}
            for building in _ALL_BUILDINGS:
                code = _building_code(building)
                row_count = self._append_building_detail_sheet(
                    workbook,
                    sheet_name=code,
                    groups=monthly_by_building[building],
                )
                source_sheet_rows[code] = row_count
            for building in _ALL_BUILDINGS:
                row_count = self._append_source_sheet(
                    workbook,
                    sheet_name=f"{building}容量",
                    source_path=Path(monthly_entry_map[building]["file_path"]),
                )
                source_sheet_rows[f"{building}容量"] = row_count
            source_sheet_rows["Sheet1"] = self._append_sheet1(
                workbook,
                records=monthly_by_building["E楼"].all_row_lines,
            )
            atomic_save_workbook(workbook, output_path)
        finally:
            workbook.close()

        finished_at = datetime.now()
        source_files = {
            "monthly_report": {
                building: {
                    "building": building,
                    "file_path": monthly_entry_map[building]["file_path"],
                    "file_name": Path(monthly_entry_map[building]["file_path"]).name,
                    "bucket_key": str(monthly_entry_map[building].get("bucket_key", "") or "").strip(),
                    "duty_date": _monthly_entry_date(monthly_entry_map[building]),
                }
                for building in _ALL_BUILDINGS
            },
        }
        result = {
            "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
            "status": "ok",
            "report_type": "top5_power_report",
            "output_file": str(output_path),
            "file_name": output_path.name,
            "output_dir": str(output_path.parent),
            "summary_row_count": len(_ALL_BUILDINGS) * _TOP_N,
            "source_sheet_count": len(source_sheet_rows),
            "source_sheet_rows": source_sheet_rows,
            "source_files": source_files,
        }
        emit_log(
            "[TOP5功率文件生成] 文件生成完成: "
            f"output={output_path}, summary_rows={result['summary_row_count']}, "
            f"source_sheets={result['source_sheet_count']}"
        )
        return result


def _upload_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value).strip()
    if isinstance(value, dict):
        for key in ("text", "name", "value", "label"):
            text = _upload_text(value.get(key))
            if text:
                return text
        return ""
    if isinstance(value, list):
        parts = [_upload_text(item) for item in value]
        return " ".join(part for part in parts if part).strip()
    return str(value).strip()


class Top5PowerReportBitableUploadService:
    """Upload the generated TOP5 workbook to the monthly report attachment table."""

    def __init__(self, runtime_config: Dict[str, Any]) -> None:
        self.runtime_config = runtime_config if isinstance(runtime_config, dict) else {}

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return Top5PowerReportService._defaults()["report_upload"]

    def _normalize_cfg(self) -> Dict[str, Any]:
        handover_cfg = self.runtime_config.get("handover_log", {})
        if not isinstance(handover_cfg, dict):
            handover_cfg = {}
        top5_cfg = handover_cfg.get("top5_power_report", {})
        if not isinstance(top5_cfg, dict):
            top5_cfg = {}
        raw_cfg = top5_cfg.get("report_upload", {})
        cfg = _deep_merge(self._defaults(), raw_cfg if isinstance(raw_cfg, dict) else {})

        # Keep compatibility with the previously-added over_power_attachment
        # config so existing deployments do not need another manual table setup.
        over_power_cfg = top5_cfg.get("over_power_attachment", {})
        if isinstance(over_power_cfg, dict):
            for key in ("app_token", "table_id"):
                if not str(cfg.get(key, "") or "").strip() and str(over_power_cfg.get(key, "") or "").strip():
                    cfg[key] = str(over_power_cfg.get(key, "") or "").strip()

        cfg["enabled"] = bool(cfg.get("enabled", True))
        for key in ("app_token", "table_id", "sub_category"):
            cfg[key] = str(cfg.get(key, "") or "").strip()
        fields = cfg.get("fields", {}) if isinstance(cfg.get("fields", {}), dict) else {}
        defaults = self._defaults()["fields"]
        cfg["fields"] = {
            key: str(fields.get(key, "") or defaults.get(key, "")).strip()
            for key in defaults
        }
        return cfg

    def _client(self, cfg: Dict[str, Any], emit_log: Callable[[str], None]) -> FeishuBitableClient:
        auth = require_feishu_auth_settings(self.runtime_config)
        return FeishuBitableClient(
            app_id=str(auth.get("app_id", "") or "").strip(),
            app_secret=str(auth.get("app_secret", "") or "").strip(),
            app_token=str(cfg.get("app_token", "") or "").strip(),
            calc_table_id=str(cfg.get("table_id", "") or "").strip(),
            attachment_table_id=str(cfg.get("table_id", "") or "").strip(),
            timeout=int(auth.get("timeout", 30) or 30),
            request_retry_count=int(auth.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(auth.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
            emit_log=emit_log,
        )

    @staticmethod
    def _validate_year_month(year: Any, month: Any) -> tuple[str, str]:
        year_text = str(year or "").strip()
        if not re.fullmatch(r"20\d{2}", year_text):
            raise ValueError("TOP5上传年度必须为四位年份")
        try:
            month_number = int(month)
        except Exception as exc:  # noqa: BLE001
            raise ValueError("TOP5上传月份必须为 1-12 的数字") from exc
        if month_number < 1 or month_number > 12:
            raise ValueError("TOP5上传月份必须在 1-12 之间")
        return year_text, f"{month_number:02d}"

    @staticmethod
    def _extract_attachment_link(record: Dict[str, Any], attachment_field: str) -> str:
        fields = record.get("fields", {}) if isinstance(record.get("fields", {}), dict) else {}
        attachments = fields.get(attachment_field)
        if not isinstance(attachments, list):
            return ""
        for item in attachments:
            if not isinstance(item, dict):
                continue
            for key in ("url", "tmp_url", "download_url"):
                text = str(item.get(key, "") or "").strip()
                if text:
                    return text
        return ""

    @staticmethod
    def _matching_record_ids(
        records: List[Dict[str, Any]],
        *,
        fields: Dict[str, str],
        sub_category: str,
        year: str,
        month: str,
    ) -> List[str]:
        output: List[str] = []
        for record in records if isinstance(records, list) else []:
            if not isinstance(record, dict):
                continue
            record_fields = record.get("fields", {}) if isinstance(record.get("fields", {}), dict) else {}
            record_category = _upload_text(record_fields.get(fields["sub_category"]))
            record_year = _upload_text(record_fields.get(fields["year"]))
            record_month = _upload_text(record_fields.get(fields["month"]))
            if record_category != sub_category:
                continue
            if record_year != year:
                continue
            if record_month.zfill(2) != month:
                continue
            record_id = str(record.get("record_id", "") or "").strip()
            if record_id:
                output.append(record_id)
        return output

    def upload_report(
        self,
        *,
        file_path: str | Path,
        year: Any,
        month: Any,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        cfg = self._normalize_cfg()
        if not cfg["enabled"]:
            return {"status": "skipped", "reason": "disabled"}
        missing = [key for key in ("app_token", "table_id", "sub_category") if not str(cfg.get(key, "") or "").strip()]
        if missing:
            raise RuntimeError(f"TOP5上传多维配置缺失: {', '.join(missing)}")

        output_path = Path(file_path)
        if not output_path.exists() or not output_path.is_file():
            raise FileNotFoundError(str(output_path))

        target_year, target_month = self._validate_year_month(year, month)
        fields = cfg["fields"]
        client = self._client(cfg, emit_log)
        table_id = str(cfg["table_id"])
        sub_category = str(cfg["sub_category"])

        emit_log(f"[TOP5功率文件生成] 开始上传多维附件: year={target_year}, month={target_month}, file={output_path.name}")
        existing_records = client.list_records(
            table_id=table_id,
            field_names=[fields["sub_category"], fields["year"], fields["month"]],
        )
        delete_ids = self._matching_record_ids(
            existing_records,
            fields=fields,
            sub_category=sub_category,
            year=target_year,
            month=target_month,
        )
        file_token = client.upload_attachment(str(output_path))
        create_fields = {
            fields["sub_category"]: sub_category,
            fields["year"]: target_year,
            fields["month"]: target_month,
            fields["attachment"]: [{"file_token": file_token}],
        }
        responses = client.batch_create_records(table_id=table_id, fields_list=[create_fields], batch_size=1)
        record_id = ""
        if responses:
            data = responses[0].get("data") if isinstance(responses[0], dict) else {}
            records = data.get("records") if isinstance(data, dict) else []
            if isinstance(records, list) and records:
                record_id = str(records[0].get("record_id", "") or "").strip()
        link = ""
        if record_id:
            for attempt in range(1, 4):
                record = client.get_record_by_id(table_id=table_id, record_id=record_id)
                link = self._extract_attachment_link(record, fields["attachment"])
                if link:
                    break
                if attempt < 3:
                    time.sleep(attempt)
            if link and fields.get("link"):
                try:
                    client.update_record(table_id=table_id, record_id=record_id, fields={fields["link"]: link})
                except Exception:
                    try:
                        client.batch_delete_records(table_id=table_id, record_ids=[record_id], batch_size=1)
                    except Exception:  # noqa: BLE001
                        pass
                    raise
        if not link:
            if record_id:
                try:
                    client.batch_delete_records(table_id=table_id, record_ids=[record_id], batch_size=1)
                except Exception:  # noqa: BLE001
                    pass
            raise RuntimeError("TOP5多维附件已创建，但未读取到附件链接，未更新“链接”字段")
        deleted = 0
        if delete_ids:
            deleted = client.batch_delete_records(table_id=table_id, record_ids=delete_ids, batch_size=200)
            emit_log(f"[TOP5功率文件生成] 已删除旧多维记录: year={target_year}, month={target_month}, count={deleted}")
        emit_log(
            "[TOP5功率文件生成] 多维附件上传完成: "
            f"year={target_year}, month={target_month}, record_id={record_id or '-'}, link={'yes' if link else 'no'}"
        )
        return {
            "status": "ok",
            "app_token": str(cfg["app_token"]),
            "table_id": table_id,
            "sub_category": sub_category,
            "year": target_year,
            "month": target_month,
            "record_id": record_id,
            "deleted_count": int(deleted or 0),
            "file_token": file_token,
            "link": link,
        }
