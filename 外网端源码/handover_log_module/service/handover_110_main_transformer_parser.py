from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl.utils import get_column_letter

from handover_log_module.repository.excel_reader import load_workbook_quietly


TRANSFORMER_COLUMNS = (9, 10, 11, 12)  # I:L
GIS_TRANSFORMER_COLUMNS = (7, 8, 9, 10)  # G:J
TRANSFORMER_NAMES = ("1号主变", "2号主变", "3号主变", "4号主变")
LINE_NAMES = ("阿开线", "阿开线", "阿家线", "阿家线")


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(",", "").replace("%", "")
    if not text or text in {"/", "-", "N/A"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _format_percent(value: Any) -> str:
    number = _number(value)
    if number is None:
        text = _text(value)
        return text if text.endswith("%") else ""
    if abs(number) <= 1:
        number *= 100
    return f"{number:.2f}%"


def _pressure_threshold(label: Any) -> float:
    text = str(label or "").strip()
    if "断路器" in text:
        return 0.55
    if "PT" in text.upper() or "压变" in text:
        return 0.45
    return 0.35


def _gis_statuses(worksheet: Any) -> List[str]:
    statuses: List[str] = []
    for col in GIS_TRANSFORMER_COLUMNS:
        ok = True
        seen = False
        for row_idx in range(132, 138):
            label = worksheet.cell(row=row_idx, column=6).value
            value = _number(worksheet.cell(row=row_idx, column=col).value)
            if value is None:
                continue
            seen = True
            if value < _pressure_threshold(label):
                ok = False
                break
        statuses.append("正常" if seen and ok else "异常")
    return statuses


def _label_text(value: Any) -> str:
    return str(value or "").strip().replace(" ", "").replace("\n", "")


def _find_metric_row(worksheet: Any, keyword: str, *, fallback: int) -> int:
    target = _label_text(keyword)
    for row_idx in range(28, 46):
        label = _label_text(worksheet.cell(row=row_idx, column=8).value)
        if target and target in label:
            return row_idx
    return fallback


def _missing_fields(
    *,
    worksheet: Any,
    column: int,
    row_map: Dict[str, int],
    current_a: float | None,
    max_load_mw: float | None,
    load_rate: str,
) -> List[str]:
    checks = [
        ("电流", current_a, row_map.get("current")),
        ("本班最大负载", max_load_mw, row_map.get("max_load")),
        ("负载率", load_rate, row_map.get("load_rate")),
    ]
    missing: List[str] = []
    col_letter = get_column_letter(column)
    for label, parsed, row_idx in checks:
        if parsed not in (None, ""):
            continue
        raw = worksheet.cell(row=int(row_idx or 0), column=column).value if row_idx else None
        missing.append(f"{label}({col_letter}{row_idx}={_text(raw) or '空'})")
    return missing


def normalize_110_main_transformer_rows(rows: List[Dict[str, Any]] | None) -> List[Dict[str, Any]]:
    source = rows if isinstance(rows, list) else []
    by_name = {
        str(item.get("transformer_name", "") or "").strip(): item
        for item in source
        if isinstance(item, dict)
    }
    normalized: List[Dict[str, Any]] = []
    for index, transformer_name in enumerate(TRANSFORMER_NAMES):
        item = by_name.get(transformer_name, {})
        load_kw_raw = item.get("load_kw")
        current_raw = item.get("current_a")
        load_kw = _number(load_kw_raw)
        current_a = _number(current_raw)
        normalized.append(
            {
                "transformer_name": transformer_name,
                "line_name": str(item.get("line_name", "") or LINE_NAMES[index]).strip() or LINE_NAMES[index],
                "oil_temp": _text(item.get("oil_temp")),
                "tap_position": _text(item.get("tap_position")),
                "load_kw": round(load_kw) if load_kw is not None else "",
                "current_a": round(current_a, 2) if current_a is not None else "",
                "load_rate": _format_percent(item.get("load_rate")),
                "gis_status": str(item.get("gis_status", "") or "").strip(),
                "missing_fields": (
                    list(item.get("missing_fields", []))
                    if isinstance(item.get("missing_fields", []), list)
                    else []
                ),
            }
        )
    return normalized


def missing_110_main_transformer_fields(rows: List[Dict[str, Any]] | None) -> List[str]:
    missing: List[str] = []
    for row in normalize_110_main_transformer_rows(rows):
        name = str(row.get("transformer_name", "") or "").strip()
        checks = [
            ("电流", row.get("current_a")),
            ("本班最大负载", row.get("load_kw")),
            ("负载率", row.get("load_rate")),
        ]
        for label, value in checks:
            if value in (None, ""):
                missing.append(f"{name}{label}")
    return missing


def parse_110_main_transformer_rows(source_file: str | Path, *, strict: bool = True) -> List[Dict[str, Any]]:
    path = Path(source_file)
    if not path.exists():
        raise FileNotFoundError(f"110站上传文件不存在: {path}")
    workbook = load_workbook_quietly(path, data_only=True)
    try:
        if not workbook.worksheets:
            raise ValueError("110站文件缺少工作表")
        worksheet = workbook.worksheets[0]
        gis_statuses = _gis_statuses(worksheet)
        row_map = {
            "current": _find_metric_row(worksheet, "输出电流", fallback=31),
            "max_load": _find_metric_row(worksheet, "本班最大负载", fallback=34),
            "load_rate": _find_metric_row(worksheet, "负载率", fallback=35),
            "oil_temp": _find_metric_row(worksheet, "油温", fallback=39),
            "tap_position": _find_metric_row(worksheet, "档位", fallback=41),
        }
        rows: List[Dict[str, Any]] = []
        for index, column in enumerate(TRANSFORMER_COLUMNS):
            current_a = _number(worksheet.cell(row=row_map["current"], column=column).value)
            max_load_mw = _number(worksheet.cell(row=row_map["max_load"], column=column).value)
            load_rate = _format_percent(worksheet.cell(row=row_map["load_rate"], column=column).value)
            oil_temp = _text(worksheet.cell(row=row_map["oil_temp"], column=column).value)
            tap_position = _text(worksheet.cell(row=row_map["tap_position"], column=column).value)
            missing = _missing_fields(
                worksheet=worksheet,
                column=column,
                row_map=row_map,
                current_a=current_a,
                max_load_mw=max_load_mw,
                load_rate=load_rate,
            )
            if strict and missing:
                raise ValueError(f"{TRANSFORMER_NAMES[index]} 主变关键数据缺失: {', '.join(missing)}")
            rows.append(
                {
                    "transformer_name": TRANSFORMER_NAMES[index],
                    "line_name": LINE_NAMES[index],
                    "oil_temp": oil_temp,
                    "tap_position": tap_position,
                    "load_kw": round(max_load_mw * 1000) if max_load_mw is not None else "",
                    "current_a": round(current_a, 2) if current_a is not None else "",
                    "load_rate": load_rate,
                    "gis_status": gis_statuses[index],
                    "missing_fields": missing,
                }
            )
        return normalize_110_main_transformer_rows(rows)
    finally:
        workbook.close()
