from __future__ import annotations

from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from handover_log_module.core.footer_layout import (
    FOOTER_DYNAMIC_FIXED_CELLS,
    FOOTER_GROUP_TITLE_TEXT,
    FOOTER_INVENTORY_COLUMNS,
    first_person_text,
    find_footer_inventory_layout,
)
from handover_log_module.core.fixed_cell_overrides import forced_fixed_cell_value
from handover_log_module.core.section_layout import build_section_logical_columns, parse_category_sections
from handover_log_module.repository.excel_reader import load_workbook_quietly


class ReviewDocumentParser:
    BLOCK_TITLE_MAP = {
        "header_basic": "基础信息",
        "metrics_summary": "指标与摘要",
        "cabinet_power_info": "机柜上下电信息",
    }
    FIELD_LABEL_MAP = {
        "A1": "标题",
        "B2": "日期",
        "F2": "班次",
        "C3": "当前班组人员",
        "G3": "下一班组人员",
        "B4": "白班长白岗",
        "F4": "夜班长白岗",
        "B6": "PUE",
        "D6": "总负荷",
        "F6": "IT总负荷",
        "H6": "供油可用时长",
        "B7": "室外温度",
        "D7": "室外湿球最高温度",
        "F7": "冷机模式汇总",
        "H7": "冷水供水汇总",
        "B8": "市政补水压力",
        "D8": "蓄水池后备时间",
        "F8": "蓄冷罐后备时间",
        "B9": "冷通道最高温度",
        "D9": "冷通道最高湿度",
        "F9": "冷通道最低温度",
        "H9": "冷通道最低湿度",
        "B10": "变压器负载率",
        "D10": "UPS负载率",
        "F10": "电池放电后备时间",
        "B13": "机房总规划机柜数（个）",
        "D13": "实际上电机柜数（个）",
        "F13": "本班组上电机柜数（个）",
        "H13": "本班组下电机柜数（个）",
        "B15": "告警总数",
        "D15": "未恢复告警数",
        "F15": "告警描述",
        "H52": "清点确认人1",
        "H53": "清点确认人2",
        "H54": "清点确认人3",
        "H55": "清点确认人4",
    }

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config if isinstance(config, dict) else {}

    def _sheet_name(self) -> str:
        template_cfg = self.config.get("template", {})
        return str(template_cfg.get("sheet_name", "")).strip()

    def _review_ui_cfg(self) -> Dict[str, Any]:
        review_ui = self.config.get("review_ui", {})
        return review_ui if isinstance(review_ui, dict) else {}

    @staticmethod
    def _normalize_fixed_cell_entry(raw_entry: Any) -> Dict[str, str] | str | None:
        if isinstance(raw_entry, str):
            cell_name = str(raw_entry or "").strip().upper()
            return cell_name or None
        if not isinstance(raw_entry, dict):
            return None
        label_cell = str(
            raw_entry.get("label_cell", raw_entry.get("LABEL_CELL", "")) or ""
        ).strip().upper()
        value_cell = str(
            raw_entry.get("value_cell", raw_entry.get("VALUE_CELL", "")) or ""
        ).strip().upper()
        if not value_cell:
            return None
        if label_cell:
            return {"label_cell": label_cell, "value_cell": value_cell}
        return value_cell

    @staticmethod
    def _fixed_cell_key(entry: Dict[str, str] | str) -> str:
        if isinstance(entry, dict):
            return str(entry.get("value_cell", "")).strip().upper()
        return str(entry or "").strip().upper()

    def _read_cell_text(self, ws, cell_name: str) -> str:
        value = ws[cell_name].value
        return "" if value is None else str(value)

    def _build_fixed_field(self, ws, entry: Dict[str, str] | str) -> Dict[str, Any]:
        if isinstance(entry, dict):
            value_cell = str(entry.get("value_cell", "")).strip().upper()
            label_cell = str(entry.get("label_cell", "")).strip().upper()
            label_text = self._read_cell_text(ws, label_cell) if label_cell else ""
            if not label_text:
                label_text = self.FIELD_LABEL_MAP.get(value_cell, value_cell)
            forced_value = forced_fixed_cell_value(value_cell)
            return {
                "cell": value_cell,
                "label": label_text,
                "value": forced_value if forced_value is not None else self._read_cell_text(ws, value_cell),
            }
        cell_name = str(entry or "").strip().upper()
        forced_value = forced_fixed_cell_value(cell_name)
        return {
            "cell": cell_name,
            "label": self.FIELD_LABEL_MAP.get(cell_name, cell_name),
            "value": forced_value if forced_value is not None else self._read_cell_text(ws, cell_name),
        }

    def _section_hidden_columns(self) -> List[str]:
        hidden_columns = self._review_ui_cfg().get("section_hidden_columns", [])
        if not isinstance(hidden_columns, list):
            return []
        return [str(value or "").strip().upper() for value in hidden_columns if str(value or "").strip()]

    def _fixed_blocks(self, ws) -> List[Dict[str, Any]]:
        fixed_cells = self._review_ui_cfg().get("fixed_cells", {})
        if not isinstance(fixed_cells, dict):
            fixed_cells = {}

        blocks: List[Dict[str, Any]] = []
        for block_id, cells in fixed_cells.items():
            if not isinstance(cells, list):
                continue
            fields: List[Dict[str, Any]] = []
            seen_cells: set[str] = set()
            for raw_cell in cells:
                normalized_entry = self._normalize_fixed_cell_entry(raw_cell)
                if normalized_entry is None:
                    continue
                cell_name = self._fixed_cell_key(normalized_entry)
                if not cell_name or cell_name in FOOTER_DYNAMIC_FIXED_CELLS or cell_name in seen_cells:
                    continue
                fields.append(self._build_fixed_field(ws, normalized_entry))
                seen_cells.add(cell_name)
            if fields:
                block_key = str(block_id or "").strip() or "block"
                blocks.append(
                    {
                        "id": block_key,
                        "title": self.BLOCK_TITLE_MAP.get(block_key, block_key),
                        "fields": fields,
                    }
                )
        return blocks

    def _sections(self, ws) -> List[Dict[str, Any]]:
        sections = parse_category_sections(ws)
        hidden_columns = self._section_hidden_columns()
        output: List[Dict[str, Any]] = []
        for section in sections:
            logical_columns = build_section_logical_columns(ws, section, hidden_columns=hidden_columns)
            columns = [
                {
                    "key": column.key,
                    "label": column.label,
                    "source_cols": list(column.source_cols),
                    "span": int(column.span),
                }
                for column in logical_columns
            ]
            rows: List[Dict[str, Any]] = []
            for row_idx in range(section.template_data_row, section.end_row + 1):
                cells = {}
                for column in columns:
                    lead_col = str(column.get("key", "")).strip().upper()
                    if not lead_col:
                        continue
                    value = ws[f"{lead_col}{row_idx}"].value
                    cells[lead_col] = "" if value is None else str(value)
                rows.append(
                    {
                        "row_id": f"{section.name}:{row_idx}",
                        "cells": cells,
                        "is_placeholder_row": not any(str(value).strip() for value in cells.values()),
                    }
                )
            output.append(
                {
                    "name": section.name,
                    "columns": columns,
                    "header": [column["label"] for column in columns],
                    "rows": rows,
                }
            )
        return output

    @staticmethod
    def _row_merge_spans(ws, row_idx: int) -> Dict[int, int]:
        spans: Dict[int, int] = {}
        for merged in ws.merged_cells.ranges:
            if merged.min_row == row_idx and merged.max_row == row_idx:
                spans[int(merged.min_col)] = int(merged.max_col - merged.min_col + 1)
        return spans

    def _inventory_footer_block(self, ws, layout) -> Dict[str, Any]:
        receiver_text = first_person_text(self._read_cell_text(ws, "G3"))
        rows: List[Dict[str, Any]] = []
        for row_idx in range(layout.data_start_row, layout.data_end_row + 1):
            cells = {}
            has_content = False
            for column in FOOTER_INVENTORY_COLUMNS:
                key = str(column["key"])
                value = ws[f"{key}{row_idx}"].value
                text = "" if value is None else str(value)
                if key.upper() == "H":
                    text = first_person_text(text)
                cells[key] = text
                if text.strip():
                    has_content = True
            has_inventory_content = any(
                str(cells.get(str(column["key"]), "") or "").strip()
                for column in FOOTER_INVENTORY_COLUMNS
                if str(column["key"]).upper() != "H"
            )
            if receiver_text and has_inventory_content and not str(cells.get("H", "") or "").strip():
                cells["H"] = receiver_text
                has_content = True
            rows.append(
                {
                    "row_id": f"inventory:{row_idx}",
                    "cells": cells,
                    "is_placeholder_row": not has_content,
                }
            )

        if not rows:
            rows = [
                {
                    "row_id": "inventory:placeholder",
                    "cells": {str(column["key"]): "" for column in FOOTER_INVENTORY_COLUMNS},
                    "is_placeholder_row": True,
                }
            ]

        return {
            "id": "handover_inventory_table",
            "type": "inventory_table",
            "title": "交接确认",
            "group_title": FOOTER_GROUP_TITLE_TEXT,
            "columns": [dict(item) for item in FOOTER_INVENTORY_COLUMNS],
            "rows": rows,
        }

    def _readonly_footer_block(self, ws, layout) -> Dict[str, Any] | None:
        if not layout.signoff_start_row:
            return None

        rows: List[Dict[str, Any]] = []
        for row_idx in range(layout.signoff_start_row, layout.last_row + 1):
            merge_spans = self._row_merge_spans(ws, row_idx)
            follower_cols = set()
            for start_col, span in merge_spans.items():
                for col_idx in range(start_col + 1, start_col + span):
                    follower_cols.add(col_idx)

            row_cells: List[Dict[str, Any]] = []
            row_has_content = False
            for col_idx in range(1, 10):
                if col_idx in follower_cols:
                    continue
                value = ws.cell(row=row_idx, column=col_idx).value
                text = "" if value is None else str(value)
                colspan = int(merge_spans.get(col_idx, 1) or 1)
                if text.strip():
                    row_has_content = True
                row_cells.append(
                    {
                        "column": get_column_letter(col_idx).upper(),
                        "value": text,
                        "colspan": colspan,
                        "cell_key": f"signoff:{row_idx}:{col_idx}",
                    }
                )
            if row_has_content or row_cells:
                rows.append({"row_key": f"signoff:{row_idx}", "cells": row_cells})

        if not rows:
            return None

        return {
            "id": "handover_signoff_block",
            "type": "readonly_grid",
            "title": "交接确认签字区",
            "rows": rows,
        }

    def _footer_blocks(self, ws) -> List[Dict[str, Any]]:
        layout = find_footer_inventory_layout(ws)
        if layout is None:
            return []

        blocks: List[Dict[str, Any]] = [self._inventory_footer_block(ws, layout)]
        readonly_block = self._readonly_footer_block(ws, layout)
        if readonly_block:
            blocks.append(readonly_block)
        return blocks

    def parse(self, output_file: str) -> Dict[str, Any]:
        workbook = load_workbook_quietly(output_file)
        try:
            sheet_name = self._sheet_name()
            ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            title_value = ws["A1"].value
            return {
                "title": "" if title_value is None else str(title_value),
                "fixed_blocks": self._fixed_blocks(ws),
                "sections": self._sections(ws),
                "footer_blocks": self._footer_blocks(ws),
            }
        finally:
            workbook.close()
