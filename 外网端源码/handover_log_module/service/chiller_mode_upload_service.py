from __future__ import annotations

import copy
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List

from openpyxl import load_workbook

from app.config.config_compat_cleanup import sanitize_chiller_mode_upload_config
from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.bitable_target_resolver import BitableTargetResolver
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from handover_log_module.service.hvac_bitable_sync_service import HvacBitableSyncService


class ChillerModeUploadService:
    """Upload latest chiller mode switch parameter source files to Feishu bitable."""

    SOURCE_FAMILY = "chiller_mode_switch_family"

    def __init__(self, runtime_config: Dict[str, Any]) -> None:
        self.runtime_config = runtime_config if isinstance(runtime_config, dict) else {}

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "enabled": True,
            "scheduler": {
                "enabled": True,
                "auto_start_in_gui": False,
                "interval_minutes": 10,
                "check_interval_sec": 30,
                "retry_failed_on_next_tick": True,
                "state_file": "chiller_mode_upload_scheduler_state.json",
            },
            "target": {
                "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
                "table_id": "tblkvVCNRbtMmjQg",
                "page_size": 500,
                "max_records": 5000,
                "delete_batch_size": 500,
                "create_batch_size": 100,
                "create_timeout_sec": 20,
                "create_retry_count": 1,
                "replace_existing": True,
            },
            "fields": {
                "building": "楼栋",
                "controller": "所属控制器",
                "point": "采集点",
                "value": "数据",
                "chiller_mode": "冷机模式",
            },
            "mode_value_map": {
                "1": "制冷",
                "2": "预冷",
                "3": "板换",
                "4": "停机",
            },
            "hvac_bitable_sync": {
                "enabled": True,
                "required": False,
                "dry_run": False,
                "send_mode_switch_alerts": False,
                "page_size": 500,
                "batch_size": 200,
                "base_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
                "source": {
                    "table_id": "tblkvVCNRbtMmjQg",
                    "all_view_id": "vewtnp2Ay9",
                    "running_view_id": "vewSen1ncq",
                },
                "target": {
                    "table_id": "tblxOyKdyyiMTdhR",
                    "view_id": "vewyJLUSVm",
                },
                "weather": {
                    "enabled": True,
                    "latitude": 31.94,
                    "longitude": 120.98,
                    "timezone": "Asia/Shanghai",
                    "summary_hours": 8,
                    "past_days": 2,
                    "forecast_days": 2,
                    "temperature_trend_threshold_c": 0.5,
                    "precipitation_threshold_mm": 0.1,
                    "precipitation_probability_threshold": 50,
                    "timeout_seconds": 15,
                    "warnings": {
                        "enabled": False,
                        "provider": "cma",
                        "station_id": "58259",
                    },
                },
                "notifications": {
                    "mode_switch_alerts": {
                        "enabled": False,
                        "chat_id": "oc_9961bb057de8bd715447559c5e63c4f2",
                        "identity": "bot",
                        "max_items": 10,
                    },
                },
            },
        }

    @classmethod
    def _deep_merge(cls, base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
        merged = copy.deepcopy(base or {})
        for key, value in (override or {}).items():
            if isinstance(value, dict) and isinstance(merged.get(key), dict):
                merged[key] = cls._deep_merge(merged[key], value)
            else:
                merged[key] = copy.deepcopy(value)
        return merged

    def _normalize_cfg(self) -> Dict[str, Any]:
        raw = sanitize_chiller_mode_upload_config(self.runtime_config.get("chiller_mode_upload", {}))
        cfg = self._deep_merge(self._defaults(), raw)

        scheduler = cfg.get("scheduler", {}) if isinstance(cfg.get("scheduler", {}), dict) else {}
        target = cfg.get("target", {}) if isinstance(cfg.get("target", {}), dict) else {}
        fields = cfg.get("fields", {}) if isinstance(cfg.get("fields", {}), dict) else {}
        mode_map = cfg.get("mode_value_map", {}) if isinstance(cfg.get("mode_value_map", {}), dict) else {}

        cfg["enabled"] = bool(cfg.get("enabled", True))

        scheduler["enabled"] = bool(scheduler.get("enabled", True))
        scheduler["auto_start_in_gui"] = bool(scheduler.get("auto_start_in_gui", False))
        scheduler["interval_minutes"] = max(1, int(scheduler.get("interval_minutes", 10) or 10))
        scheduler["check_interval_sec"] = max(1, int(scheduler.get("check_interval_sec", 30) or 30))
        scheduler["retry_failed_on_next_tick"] = bool(scheduler.get("retry_failed_on_next_tick", True))
        scheduler["state_file"] = str(scheduler.get("state_file", "") or "").strip() or "chiller_mode_upload_scheduler_state.json"
        cfg["scheduler"] = scheduler

        target["app_token"] = str(target.get("app_token", "") or "").strip()
        target["table_id"] = str(target.get("table_id", "") or "").strip()
        target["page_size"] = max(1, int(target.get("page_size", 500) or 500))
        target["max_records"] = max(1, int(target.get("max_records", 5000) or 5000))
        target["delete_batch_size"] = max(1, int(target.get("delete_batch_size", 500) or 500))
        target["create_batch_size"] = max(1, min(100, int(target.get("create_batch_size", 100) or 100)))
        target["create_timeout_sec"] = max(5, int(target.get("create_timeout_sec", 20) or 20))
        target["create_retry_count"] = max(0, int(target.get("create_retry_count", 1) or 1))
        target["replace_existing"] = bool(target.get("replace_existing", True))
        cfg["target"] = target

        for key, default in self._defaults()["fields"].items():
            fields[key] = str(fields.get(key, default) or "").strip() or default
        cfg["fields"] = fields

        normalized_map: Dict[str, str] = {}
        defaults = self._defaults()["mode_value_map"]
        source_map = mode_map if mode_map else defaults
        for key, value in source_map.items():
            code = self._normalize_mode_code(key)
            text = str(value or "").strip()
            if code and text:
                normalized_map[code] = text
        for key, value in defaults.items():
            normalized_map.setdefault(str(key), str(value))
        cfg["mode_value_map"] = normalized_map

        hvac_sync = cfg.get("hvac_bitable_sync", {}) if isinstance(cfg.get("hvac_bitable_sync", {}), dict) else {}
        hvac_default = self._defaults()["hvac_bitable_sync"]
        cfg["hvac_bitable_sync"] = self._deep_merge(hvac_default, hvac_sync)
        return cfg

    def _new_target_resolver(self) -> BitableTargetResolver:
        global_feishu = self.runtime_config.get("feishu", {}) if isinstance(self.runtime_config.get("feishu", {}), dict) else {}
        return BitableTargetResolver(
            app_id=str(global_feishu.get("app_id", "") or "").strip(),
            app_secret=str(global_feishu.get("app_secret", "") or "").strip(),
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
        )

    def build_target_descriptor(self, cfg: Dict[str, Any] | None = None, *, force_refresh: bool = False) -> Dict[str, str]:
        normalized = cfg if isinstance(cfg, dict) else self._normalize_cfg()
        target = normalized.get("target", {}) if isinstance(normalized.get("target", {}), dict) else {}
        return dict(
            self._new_target_resolver().resolve_token_pair_preview(
                configured_app_token=str(target.get("app_token", "") or "").strip(),
                table_id=str(target.get("table_id", "") or "").strip(),
                force_refresh=force_refresh,
            )
        )

    def _new_client(self, cfg: Dict[str, Any], *, target_descriptor: Dict[str, Any] | None = None) -> FeishuBitableClient:
        global_feishu = require_feishu_auth_settings(self.runtime_config)
        resolved_target = dict(target_descriptor or self.build_target_descriptor(cfg, force_refresh=True))
        operation_app_token = str(resolved_target.get("operation_app_token", "") or "").strip()
        table_id = str(resolved_target.get("table_id", "") or "").strip()
        if str(resolved_target.get("target_kind", "")).strip() not in {"base_token_pair", "wiki_token_pair"}:
            raise ValueError(str(resolved_target.get("message", "")).strip() or "制冷模式目标多维表不可用")
        if not operation_app_token or not table_id:
            raise ValueError("制冷模式目标多维表缺少 operation_app_token/table_id")
        return FeishuBitableClient(
            app_id=str(global_feishu.get("app_id", "") or "").strip(),
            app_secret=str(global_feishu.get("app_secret", "") or "").strip(),
            app_token=operation_app_token,
            calc_table_id=table_id,
            attachment_table_id=table_id,
            timeout=int(global_feishu.get("timeout", 30) or 30),
            request_retry_count=int(global_feishu.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(global_feishu.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=lambda **_: 0,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
        )

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_number(value: Any) -> int | float | None:
        if isinstance(value, bool) or value is None:
            return None
        if isinstance(value, (int, float)):
            number = float(value)
            return int(number) if number.is_integer() else number
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(",", "")
        if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
            return None
        number = float(text)
        return int(number) if number.is_integer() else number

    @classmethod
    def _normalize_mode_code(cls, value: Any) -> str:
        number = cls._normalize_number(value)
        if number is not None:
            value = number
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            as_float = float(text)
        except ValueError:
            return text
        if as_float.is_integer():
            return str(int(as_float))
        return str(as_float)

    @staticmethod
    def _is_header_row(controller: str, point: str, value: Any) -> bool:
        text = f"{controller} {point} {value}".strip()
        return bool(text) and ("所属控制器" in text or "采集点" in text) and "数据" in text

    @staticmethod
    def _is_chiller_mode_point(point: str) -> bool:
        normalized = re.sub(r"\s+", "", str(point or "").strip())
        return bool(normalized and "冷机" in normalized and "模式" in normalized)

    def _parse_source_file(
        self,
        *,
        building: str,
        file_path: str,
        cfg: Dict[str, Any],
        emit_log: Callable[[str], None],
    ) -> Dict[str, Any]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"源文件不存在: {file_path}")

        fields = cfg.get("fields", {}) if isinstance(cfg.get("fields", {}), dict) else {}
        mode_map = cfg.get("mode_value_map", {}) if isinstance(cfg.get("mode_value_map", {}), dict) else {}
        workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            worksheet = workbook.worksheets[0]
            rows: List[Dict[str, Any]] = []
            skipped_rows = 0
            controller_last = ""
            mode_count = 0

            for row_index in range(1, worksheet.max_row + 1):
                controller_raw = self._text(worksheet.cell(row=row_index, column=3).value)
                if controller_raw:
                    controller_last = controller_raw
                controller = controller_raw or controller_last
                point = self._text(worksheet.cell(row=row_index, column=4).value)
                value_raw = worksheet.cell(row=row_index, column=5).value
                if not controller and not point and value_raw in (None, ""):
                    continue
                if self._is_header_row(controller, point, value_raw):
                    continue
                if not point or value_raw in (None, ""):
                    skipped_rows += 1
                    continue
                value_number = self._normalize_number(value_raw)
                if value_number is None:
                    skipped_rows += 1
                    continue
                upload_fields = {
                    str(fields.get("building", "楼栋")).strip(): building,
                    str(fields.get("controller", "所属控制器")).strip(): controller,
                    str(fields.get("point", "采集点")).strip(): point,
                    str(fields.get("value", "数据")).strip(): value_number,
                }
                if self._is_chiller_mode_point(point):
                    mode_code = self._normalize_mode_code(value_number)
                    mode_text = str(mode_map.get(mode_code, "") or "").strip()
                    if mode_text:
                        upload_fields[str(fields.get("chiller_mode", "冷机模式")).strip()] = mode_text
                        mode_count += 1
                rows.append(upload_fields)

            emit_log(
                f"[制冷模式参数上传][{building}] 源文件解析完成: file={path.name}, "
                f"rows={len(rows)}, mode_rows={mode_count}, skipped={skipped_rows}"
            )
            return {
                "building": building,
                "file_path": str(path),
                "rows": rows,
                "mode_count": mode_count,
                "skipped_rows": skipped_rows,
            }
        finally:
            try:
                workbook.close()
            except Exception:  # noqa: BLE001
                pass

    def continue_from_source_units(
        self,
        *,
        source_units: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
        cfg: Dict[str, Any] | None = None,
        target_descriptor: Dict[str, Any] | None = None,
    ) -> Dict[str, Any]:
        normalized_cfg = cfg if isinstance(cfg, dict) else self._normalize_cfg()
        run_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        deleted_count = 0
        created_count = 0
        failed_files: List[Dict[str, Any]] = []
        parsed_files: List[Dict[str, Any]] = []
        prepared_rows: List[Dict[str, Any]] = []

        if not normalized_cfg.get("enabled", True):
            emit_log("[制冷模式参数上传] 功能未启用，跳过本次处理")
            return {
                "status": "skipped",
                "run_at": run_at,
                "parsed_files": [],
                "failed_files": [],
                "deleted_count": 0,
                "created_count": 0,
                "target": dict(target_descriptor or {}),
            }

        emit_log(
            "[制冷模式参数上传] 共享源文件处理开始: "
            + "; ".join(
                f"{str(item.get('building', '') or '').strip() or '-'}={Path(str(item.get('file_path', '') or '')).name or '-'}"
                for item in source_units
                if isinstance(item, dict)
            )
        )
        resolved_target = dict(target_descriptor or self.build_target_descriptor(normalized_cfg, force_refresh=True))
        if str(resolved_target.get("target_kind", "")).strip() not in {"base_token_pair", "wiki_token_pair"}:
            message = str(resolved_target.get("message", "") or "").strip() or "制冷模式目标多维表不可用"
            emit_log(f"[制冷模式参数上传] 目标多维表不可用: {message}")
            return {
                "status": "failed",
                "run_at": run_at,
                "parsed_files": [],
                "failed_files": [{"building": "-", "file_path": "", "error": message}],
                "deleted_count": 0,
                "created_count": 0,
                "target": resolved_target,
            }

        seen: set[tuple[str, str]] = set()
        valid_source_units: List[Dict[str, str]] = []
        for item in source_units if isinstance(source_units, list) else []:
            if not isinstance(item, dict):
                continue
            building = str(item.get("building", "") or "").strip()
            file_path = str(item.get("file_path", "") or item.get("source_file", "") or "").strip()
            if not building or not file_path:
                failed_files.append(
                    {
                        "building": building or "-",
                        "file_path": file_path,
                        "error": "共享索引缺少楼栋或源文件路径",
                    }
                )
                continue
            dedupe_key = (building, file_path)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            valid_source_units.append({"building": building, "file_path": file_path})

        if not valid_source_units and not failed_files:
            failed_files.append({"building": "-", "file_path": "", "error": "未提供有效制冷模式参数源文件"})

        for item in valid_source_units:
            building = item["building"]
            file_path = item["file_path"]
            try:
                parsed = self._parse_source_file(
                    building=building,
                    file_path=file_path,
                    cfg=normalized_cfg,
                    emit_log=emit_log,
                )
                rows = parsed.get("rows", []) if isinstance(parsed.get("rows", []), list) else []
                if not rows:
                    raise ValueError("源文件未解析出可上传数据")
                if int(parsed.get("mode_count", 0) or 0) <= 0:
                    emit_log(f"[制冷模式参数上传][{building}] 未解析到冷机模式行，本次仍上传基础采集点数据")
                parsed_files.append(
                    {
                        "building": building,
                        "file_path": file_path,
                        "row_count": len(rows),
                        "mode_count": int(parsed.get("mode_count", 0) or 0),
                        "skipped_rows": int(parsed.get("skipped_rows", 0) or 0),
                    }
                )
                prepared_rows.extend(rows)
            except Exception as exc:  # noqa: BLE001
                failed_files.append({"building": building or "-", "file_path": file_path, "error": str(exc)})
                emit_log(f"[制冷模式参数上传][{building or '-'}] 解析失败: {exc}")

        if failed_files:
            emit_log(f"[制冷模式参数上传] 存在解析失败，未清空目标多维表: failed={len(failed_files)}")
            return {
                "status": "failed",
                "run_at": run_at,
                "parsed_files": parsed_files,
                "failed_files": failed_files,
                "deleted_count": 0,
                "created_count": 0,
                "target": resolved_target,
            }

        if not prepared_rows:
            emit_log("[制冷模式参数上传] 未解析出任何可上传数据，保留目标多维表现有数据")
            return {
                "status": "skipped",
                "run_at": run_at,
                "parsed_files": parsed_files,
                "failed_files": [],
                "deleted_count": 0,
                "created_count": 0,
                "target": resolved_target,
            }

        target = normalized_cfg.get("target", {}) if isinstance(normalized_cfg.get("target", {}), dict) else {}
        max_records = max(1, int(target.get("max_records", 5000) or 5000))
        if len(prepared_rows) > max_records:
            message = f"解析记录数超过上限: records={len(prepared_rows)}, max_records={max_records}"
            emit_log(f"[制冷模式参数上传] {message}，保留目标多维表现有数据")
            return {
                "status": "failed",
                "run_at": run_at,
                "parsed_files": parsed_files,
                "failed_files": [{"building": "-", "file_path": "", "error": message}],
                "deleted_count": 0,
                "created_count": 0,
                "target": resolved_target,
            }

        client = self._new_client(normalized_cfg, target_descriptor=resolved_target)
        table_id = str(resolved_target.get("table_id", "") or "").strip()
        emit_log(
            f"[制冷模式参数上传] 全部解析成功，开始清表重传: files={len(parsed_files)}, records={len(prepared_rows)}, table_id={table_id}"
        )
        deleted_count = client.clear_table(
            table_id=table_id,
            list_page_size=int(target.get("page_size", 500) or 500),
            delete_batch_size=int(target.get("delete_batch_size", 500) or 500),
            list_field_names=[str(normalized_cfg.get("fields", {}).get("building", "楼栋") or "楼栋").strip()],
        )
        emit_log(f"[制冷模式参数上传] 目标表清空完成: deleted={deleted_count}")
        batch_size = max(1, int(target.get("create_batch_size", 100) or 100))
        total_batches = (len(prepared_rows) + batch_size - 1) // batch_size
        created_count = 0
        original_timeout = int(getattr(client, "timeout", 30) or 30)
        original_retry_count = int(getattr(client, "request_retry_count", 3) or 3)
        client.timeout = min(original_timeout, int(target.get("create_timeout_sec", 20) or 20))
        client.request_retry_count = min(original_retry_count, int(target.get("create_retry_count", 1) or 1))
        try:
            for batch_index, start in enumerate(range(0, len(prepared_rows), batch_size), start=1):
                chunk = prepared_rows[start : start + batch_size]
                emit_log(
                    f"[制冷模式参数上传] 批量写入开始: batch={batch_index}/{total_batches}, "
                    f"records={len(chunk)}, uploaded={created_count}/{len(prepared_rows)}, "
                    f"timeout={client.timeout}s, retries={client.request_retry_count}"
                )
                batch_started = time.perf_counter()
                try:
                    client.batch_create_records(
                        table_id=table_id,
                        fields_list=chunk,
                        batch_size=len(chunk),
                    )
                except Exception as exc:  # noqa: BLE001
                    emit_log(
                        f"[制冷模式参数上传][失败] 批量写入失败: batch={batch_index}/{total_batches}, "
                        f"uploaded={created_count}/{len(prepared_rows)}, error={exc}"
                    )
                    raise
                created_count += len(chunk)
                emit_log(
                    f"[制冷模式参数上传] 批量写入进度: uploaded={created_count}/{len(prepared_rows)}, "
                    f"batch={batch_index}/{total_batches}, elapsed_ms={int((time.perf_counter() - batch_started) * 1000)}"
                )
        finally:
            client.timeout = original_timeout
            client.request_retry_count = original_retry_count
        emit_log(f"[制冷模式参数上传] 批量写入完成: created={created_count}")
        hvac_sync_result = HvacBitableSyncService(self.runtime_config).safe_sync_after_chiller_upload(
            emit_log=emit_log,
            chiller_cfg=normalized_cfg,
        )
        return {
            "status": "ok",
            "run_at": run_at,
            "parsed_files": parsed_files,
            "failed_files": [],
            "deleted_count": deleted_count,
            "created_count": created_count,
            "target": resolved_target,
            "hvac_bitable_sync": hvac_sync_result,
        }
