from __future__ import annotations

import calendar
import copy
import re
import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.feishu.service.bitable_client_runtime import FeishuBitableClient
from app.modules.feishu.service.feishu_auth_resolver import require_feishu_auth_settings
from app.modules.report_pipeline.core.metrics_math import date_text_to_timestamp_ms
from app.shared.utils.atomic_file import atomic_save_workbook
from app.shared.utils.file_utils import fallback_missing_windows_drive_path
from pipeline_utils import get_app_dir


_DEFAULT_APP_TOKEN = "ASLxbfESPahdTKs0A9NccgbrnXc"
_DATA_CENTER_NAME = "EA118"
_TABLE_DEFAULTS = {
    "branch": {"table_id": "tblvyOhLPCIH52gB", "view_id": "vewcScal5R", "name": "单支路超6.25KW功率"},
    "cabinet": {"table_id": "tbloY4JDihNu0aJ7", "view_id": "vewjNHFVZK", "name": "机柜超18KW统计"},
    "line_head": {"table_id": "tblqxn3BajmZHxb6", "view_id": "vew07sLZm1", "name": "列头柜超107.5功率统计"},
    "row_line": {"table_id": "tblCcFzmKKz50tYT", "view_id": "vewQialYnD", "name": "机列超215KW功率统计"},
}
_BUILDING_ORDER = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
_TITLE_FILL = PatternFill("solid", fgColor="FFFFFF00")
_HEADER_FILL = PatternFill("solid", fgColor="F4B183")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="宋体", size=20, bold=True)
_HEADER_FONT = Font(name="宋体", size=11, bold=True)
_BODY_FONT = Font(name="宋体", size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    merged = copy.deepcopy(base or {})
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = copy.deepcopy(value)
    return merged


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value).strip()
    if isinstance(value, dict):
        for key in ("text", "name", "value", "label", "url"):
            text = _text(value.get(key))
            if text:
                return text
        return ""
    if isinstance(value, list):
        parts = [_text(item) for item in value]
        return " ".join(part for part in parts if part).strip()
    return str(value).strip()


def _number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "").strip()
    if not text:
        return None
    # If a cell contains a later correction note such as "729 (728)", use the
    # first displayed number, which is the current valid value.
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _int_value(value: Any) -> int:
    number = _number(value)
    if number is None:
        return 0
    return int(round(number))


def _format_kw(value: float | None, digits: int = 3) -> str:
    if value is None:
        return ""
    rounded = round(float(value), digits)
    text = f"{rounded:.{digits}f}".rstrip("0").rstrip(".")
    return f"{text}kw"


def _format_number(value: float | None, digits: int = 3) -> str:
    if value is None:
        return ""
    return f"{round(float(value), digits):.{digits}f}".rstrip("0").rstrip(".")


def _building_code(value: Any) -> str:
    text = _text(value).upper()
    match = re.search(r"([A-E])", text)
    return match.group(1) if match else text[:1]


def _room_sort_key(room: str) -> Tuple[int, str]:
    match = re.search(r"(\d{3})", room)
    return (int(match.group(1)) if match else 9999, room)


def _sort_key(building: str, room: str, identifier: str) -> Tuple[int, Tuple[int, str], str]:
    return (_BUILDING_ORDER.get(_building_code(building), 99), _room_sort_key(room), str(identifier or ""))


def _parse_record_date(value: Any) -> datetime | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if number > 10_000_000_000:
            return datetime.fromtimestamp(number / 1000)
        if number > 1_000_000_000:
            return datetime.fromtimestamp(number)
    text = _text(value)
    if not text:
        return None
    match = re.search(r"(20\d{2})[/-](\d{1,2})[/-](\d{1,2})", text)
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def _month_label(year: str, month: int) -> str:
    last_day = calendar.monthrange(int(year), int(month))[1]
    return f"{int(year)}/{int(month)}/1-{int(year)}/{int(month)}/{last_day}"


def _normalize_cabinet_no(value: Any) -> str:
    text = _text(value).upper()
    match = re.search(r"([A-Z])列([A-Z]\d{2})", text)
    if match:
        return match.group(2)
    match = re.search(r"\b([A-Z]\d{2})\b", text)
    return match.group(1) if match else text


@dataclass
class _DailyMetric:
    max_value: float | None = None
    run_count: int = 0
    duration_hours: int = 0
    opposite_value: float | None = None

    def merge_once(self, *, max_value: float | None, run_count: int, duration_hours: int, opposite_value: float | None = None) -> None:
        if max_value is not None and (self.max_value is None or max_value > self.max_value):
            self.max_value = max_value
        if opposite_value is not None and (self.opposite_value is None or opposite_value > self.opposite_value):
            self.opposite_value = opposite_value
        self.run_count = max(self.run_count, int(run_count or 0))
        self.duration_hours = max(self.duration_hours, int(duration_hours or 0))


@dataclass
class _BaseMonthlyItem:
    building: str
    room: str
    identifier: str
    daily: Dict[str, _DailyMetric] = field(default_factory=dict)
    remark: str = ""

    @property
    def max_value(self) -> float | None:
        values = [item.max_value for item in self.daily.values() if item.max_value is not None]
        return max(values) if values else None

    @property
    def run_count(self) -> int:
        return sum(item.run_count for item in self.daily.values())

    @property
    def duration_hours(self) -> int:
        return sum(item.duration_hours for item in self.daily.values())


@dataclass
class _LineHeadItem(_BaseMonthlyItem):
    opposite_identifier: str = "/"

    @property
    def opposite_max_value(self) -> float | None:
        values = [item.opposite_value for item in self.daily.values() if item.opposite_value is not None]
        return max(values) if values else None


@dataclass
class _CabinetPduItem:
    pdu: str
    max_current: float | None = None

    def merge_current(self, value: float | None) -> None:
        if value is not None and (self.max_current is None or value > self.max_current):
            self.max_current = value


@dataclass
class _CabinetItem(_BaseMonthlyItem):
    display_cabinet: str = ""
    imbalance: str = "均匀"
    pdus: Dict[str, _CabinetPduItem] = field(default_factory=dict)


@dataclass
class _BranchItem(_BaseMonthlyItem):
    pdu: str = ""
    opposite_pdu: str = ""

    @property
    def opposite_max_value(self) -> float | None:
        values = [item.opposite_value for item in self.daily.values() if item.opposite_value is not None]
        return max(values) if values else None


class MonthlyPowerAlertReportService:
    """Generate the EA118 monthly overpower warning workbook from four Bitable tables."""

    def __init__(self, runtime_config: Dict[str, Any]) -> None:
        self.runtime_config = runtime_config if isinstance(runtime_config, dict) else {}

    @staticmethod
    def _defaults() -> Dict[str, Any]:
        return {
            "enabled": True,
            "output_dir": r"D:\QLDownload\月度超功率统计表",
            "file_name_pattern": "EA118机柜超功耗告警统计{month}月份_{timestamp}.xlsx",
            "page_size": 500,
            "max_records": 0,
        }

    def _normalize_cfg(self) -> Dict[str, Any]:
        handover = self.runtime_config.get("handover_log", {}) if isinstance(self.runtime_config.get("handover_log", {}), dict) else {}
        top5 = handover.get("top5_power_report", {}) if isinstance(handover.get("top5_power_report", {}), dict) else {}
        report_cfg = top5.get("monthly_power_alert_report", {}) if isinstance(top5.get("monthly_power_alert_report", {}), dict) else {}
        cfg = _deep_merge(self._defaults(), report_cfg)
        cfg["enabled"] = bool(cfg.get("enabled", True))
        cfg["output_dir"] = str(cfg.get("output_dir", "") or "").strip() or self._defaults()["output_dir"]
        cfg["file_name_pattern"] = str(cfg.get("file_name_pattern", "") or "").strip() or self._defaults()["file_name_pattern"]
        cfg["page_size"] = max(1, min(500, int(cfg.get("page_size", 500) or 500)))
        cfg["max_records"] = max(0, int(cfg.get("max_records", 0) or 0))
        return cfg

    def _power_cfg(self) -> Dict[str, Any]:
        features = self.runtime_config.get("features", {}) if isinstance(self.runtime_config.get("features", {}), dict) else {}
        branch_power = features.get("branch_power_upload", {}) if isinstance(features.get("branch_power_upload", {}), dict) else {}
        raw = branch_power.get("power_alert_sync", {}) if isinstance(branch_power.get("power_alert_sync", {}), dict) else {}
        cfg = _deep_merge(
            {
                "app_token": _DEFAULT_APP_TOKEN,
                "data_center_name": _DATA_CENTER_NAME,
                "tables": _TABLE_DEFAULTS,
            },
            raw,
        )
        cfg["app_token"] = str(cfg.get("app_token", "") or _DEFAULT_APP_TOKEN).strip()
        cfg["data_center_name"] = str(cfg.get("data_center_name", "") or _DATA_CENTER_NAME).strip() or _DATA_CENTER_NAME
        tables = cfg.get("tables", {}) if isinstance(cfg.get("tables", {}), dict) else {}
        cfg["tables"] = {
            key: _deep_merge(default, tables.get(key, {}) if isinstance(tables.get(key, {}), dict) else {})
            for key, default in _TABLE_DEFAULTS.items()
        }
        return cfg

    def _client(self, power_cfg: Dict[str, Any], emit_log: Callable[[str], None]) -> FeishuBitableClient:
        auth = require_feishu_auth_settings(self.runtime_config)
        return FeishuBitableClient(
            app_id=str(auth.get("app_id", "") or "").strip(),
            app_secret=str(auth.get("app_secret", "") or "").strip(),
            app_token=str(power_cfg.get("app_token", "") or _DEFAULT_APP_TOKEN).strip(),
            calc_table_id="",
            attachment_table_id="",
            timeout=int(auth.get("timeout", 30) or 30),
            request_retry_count=int(auth.get("request_retry_count", 3) or 3),
            request_retry_interval_sec=float(auth.get("request_retry_interval_sec", 2) or 2),
            date_text_to_timestamp_ms_fn=date_text_to_timestamp_ms,
            canonical_metric_name_fn=lambda value: str(value or "").strip(),
            dimension_mapping={},
            emit_log=emit_log,
        )

    def _resolve_output_dir(self, cfg: Dict[str, Any]) -> Path:
        path = Path(str(cfg.get("output_dir", "") or "").strip())
        if path.is_absolute():
            output_dir = fallback_missing_windows_drive_path(path, app_dir=get_app_dir())
        else:
            output_dir = get_app_dir() / path
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir

    def _build_output_path(self, cfg: Dict[str, Any], *, year: str, month: int) -> Path:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        file_name = str(cfg.get("file_name_pattern", "") or self._defaults()["file_name_pattern"]).format(
            year=year,
            month=str(month),
            month2=f"{month:02d}",
            timestamp=timestamp,
        )
        if not file_name.lower().endswith(".xlsx"):
            file_name += ".xlsx"
        candidate = self._resolve_output_dir(cfg) / file_name
        index = 2
        while candidate.exists():
            candidate = candidate.with_name(f"{candidate.stem}_{index}{candidate.suffix}")
            index += 1
        return candidate

    @staticmethod
    def validate_year_month(year: Any, month: Any) -> Tuple[str, int]:
        year_text = str(year or "").strip()
        if not re.fullmatch(r"20\d{2}", year_text):
            raise ValueError("月度超功率统计表年份必须为四位年份")
        try:
            month_number = int(month)
        except Exception as exc:  # noqa: BLE001
            raise ValueError("月度超功率统计表月份必须为 1-12 的数字") from exc
        if not 1 <= month_number <= 12:
            raise ValueError("月度超功率统计表月份必须在 1-12 之间")
        return year_text, month_number

    def _list_table_records(
        self,
        client: FeishuBitableClient,
        table_cfg: Dict[str, Any],
        report_cfg: Dict[str, Any],
        *,
        emit_log: Callable[[str], None],
    ) -> List[Dict[str, Any]]:
        table_id = str(table_cfg.get("table_id", "") or "").strip()
        if not table_id:
            raise RuntimeError(f"月度超功率统计表配置缺少 table_id: {table_cfg.get('name') or '-'}")
        records = client.list_records(
            table_id=table_id,
            view_id=str(table_cfg.get("view_id", "") or "").strip(),
            page_size=int(report_cfg.get("page_size", 500) or 500),
            max_records=int(report_cfg.get("max_records", 0) or 0),
        )
        emit_log(f"[月度超功率统计表] 多维读取完成: table={table_cfg.get('name') or table_id}, records={len(records)}")
        return records

    @staticmethod
    def _iter_month_records(records: Iterable[Dict[str, Any]], *, year: str, month: int) -> Iterable[Tuple[str, Dict[str, Any]]]:
        for record in records:
            fields = record.get("fields", {}) if isinstance(record.get("fields", {}), dict) else {}
            parsed = _parse_record_date(fields.get("数据时间"))
            if parsed is None or parsed.year != int(year) or parsed.month != int(month):
                continue
            yield parsed.strftime("%Y-%m-%d"), fields

    @staticmethod
    def _merge_daily(item: _BaseMonthlyItem, date_text: str, *, max_value: float | None, run_count: int, duration_hours: int, opposite_value: float | None = None) -> None:
        daily = item.daily.setdefault(date_text, _DailyMetric())
        daily.merge_once(max_value=max_value, run_count=run_count, duration_hours=duration_hours, opposite_value=opposite_value)

    def _aggregate_row_line(self, records: List[Dict[str, Any]], *, year: str, month: int) -> List[_BaseMonthlyItem]:
        items: Dict[str, _BaseMonthlyItem] = {}
        for date_text, fields in self._iter_month_records(records, year=year, month=month):
            building = _building_code(fields.get("楼栋"))
            room = _text(fields.get("房间"))
            identifier = _text(fields.get("机列"))
            if not building or not room or not identifier:
                continue
            key = f"{building}|{room}|{identifier}"
            item = items.setdefault(key, _BaseMonthlyItem(building=building, room=room, identifier=identifier, remark=_text(fields.get("备注"))))
            self._merge_daily(
                item,
                date_text,
                max_value=_number(fields.get("功率")),
                run_count=_int_value(fields.get("次数")),
                duration_hours=_int_value(fields.get("时长")),
            )
        return sorted(items.values(), key=lambda item: _sort_key(item.building, item.room, item.identifier))

    def _aggregate_line_head(self, records: List[Dict[str, Any]], *, year: str, month: int) -> List[_LineHeadItem]:
        items: Dict[str, _LineHeadItem] = {}
        for date_text, fields in self._iter_month_records(records, year=year, month=month):
            building = _building_code(fields.get("楼栋"))
            room = _text(fields.get("房间"))
            identifier = _text(fields.get("机列"))
            if not building or not room or not identifier:
                continue
            key = f"{building}|{room}|{identifier}"
            item = items.setdefault(
                key,
                _LineHeadItem(
                    building=building,
                    room=room,
                    identifier=identifier,
                    opposite_identifier=_text(fields.get("对侧机列")) or "/",
                    remark=_text(fields.get("备注")),
                ),
            )
            if _text(fields.get("对侧机列")) and _text(fields.get("对侧机列")) != "/":
                item.opposite_identifier = _text(fields.get("对侧机列"))
            self._merge_daily(
                item,
                date_text,
                max_value=_number(fields.get("功率")),
                run_count=_int_value(fields.get("次数")),
                duration_hours=_int_value(fields.get("时长")),
                opposite_value=_number(fields.get("对侧机列最大功率")),
            )
        return sorted(items.values(), key=lambda item: _sort_key(item.building, item.room, item.identifier))

    def _aggregate_cabinet(self, records: List[Dict[str, Any]], *, year: str, month: int) -> List[_CabinetItem]:
        items: Dict[str, _CabinetItem] = {}
        for date_text, fields in self._iter_month_records(records, year=year, month=month):
            building = _building_code(fields.get("楼栋"))
            room = _text(fields.get("房间"))
            cabinet_no = _normalize_cabinet_no(fields.get("机柜号"))
            if not building or not room or not cabinet_no:
                continue
            key = f"{building}|{room}|{cabinet_no}"
            raw_cabinet = _text(fields.get("机柜号"))
            item = items.setdefault(
                key,
                _CabinetItem(
                    building=building,
                    room=room,
                    identifier=cabinet_no,
                    display_cabinet=raw_cabinet or cabinet_no,
                    imbalance=_text(fields.get("是否负载不均匀")) or "均匀",
                    remark=_text(fields.get("备注")),
                ),
            )
            self._merge_daily(
                item,
                date_text,
                max_value=_number(fields.get("机柜功率")),
                run_count=_int_value(fields.get("次数")),
                duration_hours=_int_value(fields.get("时长")),
            )
            pdu = _text(fields.get("PDU编号"))
            if pdu:
                item.pdus.setdefault(pdu, _CabinetPduItem(pdu=pdu)).merge_current(_number(fields.get("电流值")))
        return sorted(items.values(), key=lambda item: _sort_key(item.building, item.room, item.identifier))

    def _aggregate_branch(self, records: List[Dict[str, Any]], *, year: str, month: int) -> List[_BranchItem]:
        items: Dict[str, _BranchItem] = {}
        for date_text, fields in self._iter_month_records(records, year=year, month=month):
            building = _building_code(fields.get("楼栋"))
            room = _text(fields.get("房间"))
            branch_code = _text(fields.get("支路编号")) or _text(fields.get("支路号"))
            pdu = _text(fields.get("PDU编号"))
            if not building or not room or not branch_code:
                continue
            key = f"{building}|{room}|{branch_code}|{pdu}"
            item = items.setdefault(
                key,
                _BranchItem(
                    building=building,
                    room=room,
                    identifier=branch_code,
                    pdu=pdu,
                    opposite_pdu=_text(fields.get("对侧PDU编号")),
                    remark=_text(fields.get("备注")),
                ),
            )
            if _text(fields.get("对侧PDU编号")):
                item.opposite_pdu = _text(fields.get("对侧PDU编号"))
            self._merge_daily(
                item,
                date_text,
                max_value=_number(fields.get("支路功率")),
                run_count=0,
                duration_hours=_int_value(fields.get("时长")),
                opposite_value=_number(fields.get("对侧支路功率")),
            )
        return sorted(items.values(), key=lambda item: _sort_key(item.building, item.room, item.identifier))

    @staticmethod
    def _style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.alignment = _CENTER
                cell.border = _BORDER
                cell.font = _BODY_FONT

    @staticmethod
    def _setup_main_sheet(ws, *, title: str, headers: List[str], widths: List[float]) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1, title)
        ws.cell(1, 1).fill = _TITLE_FILL
        ws.cell(1, 1).font = _TITLE_FONT
        ws.cell(1, 1).alignment = _CENTER
        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col, header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER
            width = widths[col - 1] if col - 1 < len(widths) else 14
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

    @staticmethod
    def _write_rows(ws, rows: List[List[Any]], *, start_row: int = 3) -> None:
        for row_index, row_values in enumerate(rows, start_row):
            for col_index, value in enumerate(row_values, 1):
                ws.cell(row_index, col_index, value)
        if rows:
            MonthlyPowerAlertReportService._style_range(ws, start_row, start_row + len(rows) - 1, 1, len(rows[0]))

    @staticmethod
    def _merge_same_group(ws, start_row: int, end_row: int, columns: Iterable[int]) -> None:
        if end_row <= start_row:
            return
        for col in columns:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            ws.cell(start_row, col).alignment = _CENTER

    def _write_row_line_sheet(self, workbook, items: List[_BaseMonthlyItem], *, period_label: str) -> None:
        ws = workbook.create_sheet("机列超215功率统计")
        headers = ["序号", "数据时间", "机房", "楼栋", "房间", "机列", "功率", "次数", "时长", "备注"]
        self._setup_main_sheet(ws, title="EA118机列超功率告警统计表", headers=headers, widths=[8, 22, 8, 8, 16, 16, 12, 10, 10, 18])
        rows = [
            [
                index,
                period_label,
                _DATA_CENTER_NAME,
                item.building,
                item.room,
                item.identifier,
                _format_kw(item.max_value, 3),
                item.run_count,
                f"{item.duration_hours}h",
                item.remark,
            ]
            for index, item in enumerate(items, 1)
        ]
        self._write_rows(ws, rows)

    def _write_line_head_sheet(self, workbook, items: List[_LineHeadItem], *, period_label: str) -> None:
        ws = workbook.create_sheet("列头柜超107.5功率统计")
        headers = ["序号", "数据时间", "机房", "楼栋", "房间", "机列", "功率", "对侧机列", "对侧机列最大功率", "次数", "时长", "备注"]
        self._setup_main_sheet(ws, title="EA118列头柜超功率告警统计表", headers=headers, widths=[8, 22, 8, 8, 16, 18, 12, 22, 18, 10, 10, 18])
        rows = [
            [
                index,
                period_label,
                _DATA_CENTER_NAME,
                item.building,
                item.room,
                item.identifier,
                _format_kw(item.max_value, 3),
                item.opposite_identifier or "/",
                _format_kw(item.opposite_max_value, 3) if item.opposite_max_value is not None else "/",
                item.run_count,
                f"{item.duration_hours}h",
                item.remark,
            ]
            for index, item in enumerate(items, 1)
        ]
        self._write_rows(ws, rows)

    def _write_cabinet_sheet(self, workbook, items: List[_CabinetItem], *, period_label: str) -> None:
        ws = workbook.create_sheet("机柜超18KW统计")
        headers = ["序号", "数据时间", "机房", "楼栋", "房间", "机柜号", "机柜功率", "PDU编号", "电流值", "是否负载不均匀", "次数", "时长", "备注"]
        self._setup_main_sheet(ws, title="EA118机柜超功率告警统计表", headers=headers, widths=[8, 22, 8, 8, 14, 12, 12, 12, 10, 18, 10, 10, 18])
        current_row = 3
        serial = 1
        for item in items:
            pdus = sorted(item.pdus.values(), key=lambda row: row.pdu) or [_CabinetPduItem(pdu="/")]
            start_row = current_row
            for pdu_item in pdus:
                values = [
                    serial,
                    period_label,
                    _DATA_CENTER_NAME,
                    item.building,
                    item.room,
                    item.identifier,
                    _format_kw(item.max_value, 2),
                    pdu_item.pdu,
                    _format_number(pdu_item.max_current, 3),
                    item.imbalance or "均匀",
                    item.run_count,
                    f"{item.duration_hours}h",
                    item.remark,
                ]
                for col_index, value in enumerate(values, 1):
                    ws.cell(current_row, col_index, value)
                current_row += 1
            self._style_range(ws, start_row, current_row - 1, 1, len(headers))
            self._merge_same_group(ws, start_row, current_row - 1, columns=[1, 2, 3, 4, 5, 6, 7, 10, 11, 12, 13])
            serial += 1

    def _write_branch_sheet(self, workbook, items: List[_BranchItem], *, period_label: str) -> None:
        ws = workbook.create_sheet("单支路超6.25KW功率")
        headers = ["序号", "数据时间", "机房", "楼栋", "房间", "支路号", "PDU编号", "支路功率", "对侧PDU编号", "支路功率", "时长", "备注"]
        self._setup_main_sheet(ws, title="EA118单支路超功率告警统计表", headers=headers, widths=[8, 22, 8, 8, 16, 20, 14, 12, 22, 12, 10, 18])
        rows = [
            [
                index,
                period_label,
                _DATA_CENTER_NAME,
                item.building,
                item.room,
                item.identifier,
                item.pdu,
                _format_number(item.max_value, 3),
                item.opposite_pdu or "",
                _format_number(item.opposite_max_value, 3),
                f"{item.duration_hours}h",
                item.remark,
            ]
            for index, item in enumerate(items, 1)
        ]
        self._write_rows(ws, rows)

    @staticmethod
    def _appendix_label(item: _BaseMonthlyItem) -> str:
        if isinstance(item, _CabinetItem):
            return f"{item.building}/{item.room}/{item.display_cabinet or item.identifier}"
        return f"{item.building}/{item.room}/{item.identifier}"

    def _write_appendix_sheet(self, workbook, *, sheet_name: str, items: List[_BaseMonthlyItem], year: str, month: int) -> None:
        ws = workbook.create_sheet(sheet_name)
        last_day = calendar.monthrange(int(year), int(month))[1]
        ws.merge_cells("A1:A3")
        ws.merge_cells("B1:B3")
        ws.cell(1, 1, "序号")
        ws.cell(1, 2, "时间")
        for col, item in enumerate(items, 3):
            ws.cell(1, col, self._appendix_label(item))
            ws.cell(2, col, "日最大功率")
            ws.cell(3, col, "KW")
            ws.column_dimensions[get_column_letter(col)].width = max(14, min(42, len(self._appendix_label(item)) * 1.6))
        for row in range(1, 4):
            for col in range(1, max(2, len(items) + 2) + 1):
                cell = ws.cell(row, col)
                cell.alignment = _CENTER
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL if row == 1 else PatternFill("solid", fgColor="FCE4D6")
                cell.border = _BORDER
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        for day in range(1, last_day + 1):
            row = day + 3
            date_text = f"{int(year)}/{int(month)}/{day}"
            date_key = f"{int(year):04d}-{int(month):02d}-{day:02d}"
            ws.cell(row, 1, day)
            ws.cell(row, 2, date_text)
            for col, item in enumerate(items, 3):
                metric = item.daily.get(date_key)
                ws.cell(row, col, _format_number(metric.max_value if metric else None, 3))
            self._style_range(ws, row, row, 1, max(2, len(items) + 2))
        ws.freeze_panes = "C4"

    def _write_note_sheet(self, workbook) -> None:
        ws = workbook.create_sheet("说明")
        rows = [
            ("填写说明", ""),
            ("隐藏部分", "为历史超功率告警（本月未触发超功率），勿删除！"),
            ("", ""),
            ("生成说明", "本表由四类超功率多维表日记录聚合生成。"),
        ]
        for row_index, values in enumerate(rows, 1):
            for col_index, value in enumerate(values, 1):
                ws.cell(row_index, col_index, value)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 62
        self._style_range(ws, 1, len(rows), 1, 2)
        for row in (1, 2, 4):
            ws.cell(row, 1).font = _HEADER_FONT

    def _build_workbook(
        self,
        *,
        year: str,
        month: int,
        row_line_items: List[_BaseMonthlyItem],
        line_head_items: List[_LineHeadItem],
        cabinet_items: List[_CabinetItem],
        branch_items: List[_BranchItem],
    ):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        period_label = _month_label(year, month)
        self._write_row_line_sheet(workbook, row_line_items, period_label=period_label)
        self._write_line_head_sheet(workbook, line_head_items, period_label=period_label)
        self._write_cabinet_sheet(workbook, cabinet_items, period_label=period_label)
        self._write_branch_sheet(workbook, branch_items, period_label=period_label)
        self._write_note_sheet(workbook)
        self._write_appendix_sheet(workbook, sheet_name="机列超功率附表", items=row_line_items, year=year, month=month)
        self._write_appendix_sheet(workbook, sheet_name="列头柜超功率附表", items=line_head_items, year=year, month=month)
        self._write_appendix_sheet(workbook, sheet_name="机柜超功率附表", items=cabinet_items, year=year, month=month)
        return workbook

    def run(
        self,
        *,
        year: Any,
        month: Any,
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        report_cfg = self._normalize_cfg()
        if not report_cfg["enabled"]:
            return {"status": "skipped", "reason": "disabled"}
        year_text, month_number = self.validate_year_month(year, month)
        started_at = datetime.now()
        power_cfg = self._power_cfg()
        client = self._client(power_cfg, emit_log)
        table_cfgs = power_cfg["tables"]

        emit_log(f"[月度超功率统计表] 开始读取四张超功率多维表: year={year_text}, month={month_number:02d}")
        table_records = {
            key: self._list_table_records(client, table_cfgs[key], report_cfg, emit_log=emit_log)
            for key in ("row_line", "line_head", "cabinet", "branch")
        }
        row_line_items = self._aggregate_row_line(table_records["row_line"], year=year_text, month=month_number)
        line_head_items = self._aggregate_line_head(table_records["line_head"], year=year_text, month=month_number)
        cabinet_items = self._aggregate_cabinet(table_records["cabinet"], year=year_text, month=month_number)
        branch_items = self._aggregate_branch(table_records["branch"], year=year_text, month=month_number)
        emit_log(
            "[月度超功率统计表] 聚合完成: "
            f"row_line={len(row_line_items)}, line_head={len(line_head_items)}, cabinet={len(cabinet_items)}, branch={len(branch_items)}"
        )

        output_path = self._build_output_path(report_cfg, year=year_text, month=month_number)
        workbook = self._build_workbook(
            year=year_text,
            month=month_number,
            row_line_items=row_line_items,
            line_head_items=line_head_items,
            cabinet_items=cabinet_items,
            branch_items=branch_items,
        )
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning)
                atomic_save_workbook(workbook, output_path)
        finally:
            workbook.close()

        finished_at = datetime.now()
        result = {
            "status": "ok",
            "report_type": "monthly_power_alert_report",
            "year": year_text,
            "month": f"{month_number:02d}",
            "started_at": started_at.strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": finished_at.strftime("%Y-%m-%d %H:%M:%S"),
            "output_file": str(output_path),
            "file_name": output_path.name,
            "output_dir": str(output_path.parent),
            "row_counts": {
                "row_line": len(row_line_items),
                "line_head": len(line_head_items),
                "cabinet": len(cabinet_items),
                "branch": len(branch_items),
            },
            "source_record_counts": {key: len(value) for key, value in table_records.items()},
        }
        emit_log(f"[月度超功率统计表] 文件生成完成: output={output_path}")
        return result
