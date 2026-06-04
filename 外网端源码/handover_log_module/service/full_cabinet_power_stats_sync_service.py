from __future__ import annotations

import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List

import openpyxl

from handover_log_module.service.power_alert_sync_service import PowerAlertSyncService, _PowerAlertTable


@dataclass(frozen=True)
class _CabinetMetricRow:
    building: str
    room_code: str
    room_short: str
    cabinet_col: str
    cabinet_no: str
    powers: List[float]


@dataclass(frozen=True)
class _LineHeadMetricRow:
    building: str
    room_code: str
    room_short: str
    line_raw: str
    line: Dict[str, Any]
    powers: List[float]


@dataclass(frozen=True)
class _RowLineMetricRow:
    building: str
    room_code: str
    room_short: str
    row_col: str
    powers: List[float]


class FullCabinetPowerStatsSyncService(PowerAlertSyncService):
    HEADER_SCAN_ROWS = 4
    HOURS = list(range(24))

    def _iter_target_tables(self, cfg: Dict[str, Any]) -> tuple[List[_PowerAlertTable], List[str]]:
        target_tables, missing = self._resolve_target_tables(cfg)
        selected = [table for table in target_tables if table.key in {"cabinet", "line_head", "row_line"}]
        return selected, missing

    @staticmethod
    def _extract_room_code(room_text: str, fallback_text: str = "") -> str:
        for raw in (str(room_text or "").strip(), str(fallback_text or "").strip()):
            match = re.search(r"([A-Z]-\d{3})", raw, re.IGNORECASE)
            if match:
                return match.group(1).upper()
        return ""

    @classmethod
    def _parse_header_datetime(cls, value: Any) -> datetime | None:
        return BranchCompatibleHeaderMixin.parse_header_datetime(value)

    def _detect_header_row(self, sheet: Any) -> tuple[int, tuple[Any, ...]]:
        best_row_index = 1
        best_header: tuple[Any, ...] = ()
        best_count = -1
        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=self.HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            values = tuple(row or ())
            parsed_count = sum(1 for value in values if self._parse_header_datetime(value) is not None)
            if parsed_count > best_count:
                best_count = parsed_count
                best_row_index = row_index
                best_header = values
        if best_count <= 0:
            raise RuntimeError("楼栋全机柜功率源文件未识别到小时表头")
        return best_row_index, best_header

    def _resolve_hour_columns(self, sheet: Any, *, business_date: str) -> tuple[int, Dict[int, int]]:
        header_row_index, header = self._detect_header_row(sheet)
        target_hours = {
            datetime.strptime(f"{business_date} {hour:02d}:00:00", "%Y-%m-%d %H:%M:%S"): hour
            for hour in self.HOURS
        }
        hour_columns: Dict[int, int] = {}
        for index, value in enumerate(header, start=1):
            parsed = self._parse_header_datetime(value)
            if parsed is None:
                continue
            normalized = parsed.replace(minute=0, second=0, microsecond=0)
            hour = target_hours.get(normalized)
            if hour is not None and hour not in hour_columns:
                hour_columns[hour] = index
        missing_hours = [str(hour).zfill(2) for hour in self.HOURS if hour not in hour_columns]
        if missing_hours:
            raise RuntimeError(f"楼栋全机柜功率源文件缺少小时列: {','.join(missing_hours[:12])}")
        return header_row_index, hour_columns

    @staticmethod
    def _number_or_zero(value: Any) -> float:
        return PowerAlertSyncService._number_or_zero(value)

    def _parse_metric_file(
        self,
        *,
        file_path: Path,
        building: str,
        business_date: str,
    ) -> Dict[str, List[Any]]:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            if hasattr(sheet, "reset_dimensions"):
                sheet.reset_dimensions()
            header_row_index, hour_columns = self._resolve_hour_columns(sheet, business_date=business_date)
            cabinet_rows: List[_CabinetMetricRow] = []
            line_head_rows: List[_LineHeadMetricRow] = []
            row_line_rows: List[_RowLineMetricRow] = []
            current_room_text = ""
            current_group_text = ""
            data_start_row = max(4, header_row_index + 2)
            for row_index, raw_row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                row = tuple(raw_row or ())
                room_text = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                group_text = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                item_text = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                if room_text:
                    current_room_text = room_text
                if group_text:
                    current_group_text = group_text
                if not current_group_text or not item_text:
                    continue
                powers = [
                    self._number_or_zero(row[column_index - 1] if len(row) >= column_index else None)
                    for _hour, column_index in sorted(hour_columns.items())
                ]
                room_code = self._extract_room_code(current_room_text, current_group_text)
                if not room_code:
                    raise RuntimeError(f"{file_path.name} 第{row_index}行无法识别房间编码")
                cabinet_match = re.fullmatch(r"(?P<room>\d+)包间(?P<col>[A-Z])列功率和", current_group_text)
                item_cabinet_match = re.fullmatch(r"(?P<cabinet>[A-Z]\d{2})机柜功率和", item_text)
                if cabinet_match and item_cabinet_match:
                    cabinet_rows.append(
                        _CabinetMetricRow(
                            building=building,
                            room_code=room_code,
                            room_short=room_code,
                            cabinet_col=item_cabinet_match.group("cabinet")[0].upper(),
                            cabinet_no=item_cabinet_match.group("cabinet")[1:].zfill(2),
                            powers=powers,
                        )
                    )
                    continue
                parsed_line = self._parse_line(current_group_text)
                if parsed_line and item_text == "总_负载功率_KW":
                    line_head_rows.append(
                        _LineHeadMetricRow(
                            building=building,
                            room_code=room_code,
                            room_short=parsed_line.get("room_short", room_code),
                            line_raw=current_group_text,
                            line=parsed_line,
                            powers=powers,
                        )
                    )
                    continue
                row_line_group = re.fullmatch(r"(?P<room>[A-Z]-\d{3})列头柜功率和", current_group_text)
                row_line_item = re.fullmatch(r"(?P<col>[A-Z])列功率和", item_text)
                if row_line_group and row_line_item:
                    row_line_rows.append(
                        _RowLineMetricRow(
                            building=building,
                            room_code=room_code,
                            room_short=row_line_group.group("room"),
                            row_col=row_line_item.group("col").upper(),
                            powers=powers,
                        )
                    )
            return {
                "cabinet": cabinet_rows,
                "line_head": line_head_rows,
                "row_line": row_line_rows,
            }
        finally:
            workbook.close()

    def _build_detail_index(self, detail_records: List[Dict[str, Any]]) -> Dict[tuple[str, str, str], List[Dict[str, Any]]]:
        output: Dict[tuple[str, str, str], List[Dict[str, Any]]] = {}
        for item in detail_records if isinstance(detail_records, list) else []:
            if not isinstance(item, dict):
                continue
            building = self._text(item.get("机楼"))
            room = self._extract_room_code(self._text(item.get("包间")))
            pdu_code = self._text(item.get("支路编号"))
            if not building or not room or not pdu_code:
                continue
            pdu_info = self._parse_pdu(pdu_code)
            if not pdu_info:
                continue
            cabinet_id = f"{str(pdu_info.get('col', '')).upper()}{str(pdu_info.get('num_pad2', '')).zfill(2)}"
            hours = []
            currents = []
            for hour in self.HOURS:
                hours.append(self._number_or_zero(item.get(f"功率-{hour}:00")))
                currents.append(self._number_or_zero(item.get(f"电流-{hour}:00")))
            output.setdefault((building, room, cabinet_id), []).append(
                {
                    "building": building,
                    "room": room,
                    "pdu_code": pdu_code,
                    "pdu_info": pdu_info,
                    "powers": hours,
                    "currents": currents,
                }
            )
        for rows in output.values():
            rows.sort(
                key=lambda row: (
                    str(row.get("pdu_info", {}).get("side", "")),
                    int(row.get("pdu_info", {}).get("feed", 0) or 0),
                    str(row.get("pdu_code", "")),
                )
            )
        return output

    def _generate_cabinet_rows(
        self,
        rows: List[_CabinetMetricRow],
        *,
        detail_index: Dict[tuple[str, str, str], List[Dict[str, Any]]],
        threshold: float,
        report_date: str,
        data_center_name: str,
    ) -> List[Dict[str, Any]]:
        output: List[Dict[str, Any]] = []
        for row in rows:
            stats = self._threshold_stats(row.powers, threshold)
            if not int(stats["over_count"] or 0):
                continue
            cabinet_id = f"{row.cabinet_col}{row.cabinet_no}"
            details = detail_index.get((row.building, row.room_code, cabinet_id), [])
            max_hour = int(stats["max_hour"])
            if not details:
                output.append(
                    {
                        "序号": str(len(output) + 1),
                        "数据时间": report_date,
                        "机房": data_center_name,
                        "楼栋": row.building,
                        "房间": row.room_code,
                        "机柜号": cabinet_id,
                        "机柜功率": f"{self._fmt_trim(stats['max_value'], 2)}kw",
                        "PDU编号": None,
                        "电流值": None,
                        "是否负载不均匀": "均匀",
                        "次数": stats["runs"],
                        "时长": f"{stats['over_count']}h",
                        "备注": None,
                    }
                )
                continue
            for item in details:
                output.append(
                    {
                        "序号": str(len(output) + 1),
                        "数据时间": report_date,
                        "机房": data_center_name,
                        "楼栋": row.building,
                        "房间": row.room_code,
                        "机柜号": cabinet_id,
                        "机柜功率": f"{self._fmt_trim(stats['max_value'], 2)}kw",
                        "PDU编号": item.get("pdu_code"),
                        "电流值": float(self._fmt_trim(item.get("currents", [0])[max_hour], 3) or 0),
                        "是否负载不均匀": "均匀",
                        "次数": stats["runs"],
                        "时长": f"{stats['over_count']}h",
                        "备注": None,
                    }
                )
        return output

    def _generate_line_head_rows(
        self,
        rows: List[_LineHeadMetricRow],
        *,
        threshold: float,
        report_date: str,
        data_center_name: str,
    ) -> List[Dict[str, Any]]:
        group_stats = {row.line_raw: {"group": [row], "totals": row.powers} for row in rows}
        output: List[Dict[str, Any]] = []
        for row in rows:
            stats = self._threshold_stats(row.powers, threshold)
            if not int(stats["over_count"] or 0):
                continue
            opposite = self._find_opposite_line_group(row.line_raw, group_stats)
            opposite_max = self._max_of(opposite["totals"]) if opposite else None
            output.append(
                {
                    "序号": len(output) + 1,
                    "数据时间": report_date,
                    "机房": data_center_name,
                    "楼栋": row.building,
                    "房间": f"{row.room_short}.{data_center_name}",
                    "机列": self._line_display(row.line),
                    "功率": f"{self._fmt_trim(stats['max_value'], 3)}kw",
                    "对侧机列": self._line_display(opposite["group"][0].line) if opposite else None,
                    "对侧机列最大功率": f"{self._fmt_trim(opposite_max, 3)}kw" if opposite else None,
                    "次数": stats["runs"],
                    "时长": f"{stats['over_count']}h",
                    "备注": None,
                }
            )
        return output

    def _generate_row_line_rows(
        self,
        rows: List[_RowLineMetricRow],
        *,
        threshold: float,
        report_date: str,
        data_center_name: str,
    ) -> List[Dict[str, Any]]:
        output: List[Dict[str, Any]] = []
        for row in rows:
            stats = self._threshold_stats(row.powers, threshold)
            if not int(stats["over_count"] or 0):
                continue
            output.append(
                {
                    "序号": len(output) + 1,
                    "数据时间": report_date,
                    "机房": data_center_name,
                    "楼栋": row.building,
                    "房间": f"{row.room_short}.{data_center_name}",
                    "机列": f"{row.row_col}列",
                    "功率": f"{self._fmt_trim(stats['max_value'], 3)}KW",
                    "次数": stats["runs"],
                    "时长": f"{stats['over_count']}h",
                    "备注": None,
                }
            )
        return output

    def sync_from_source_units(
        self,
        *,
        report_date: str,
        source_units: List[Dict[str, Any]],
        detail_records: List[Dict[str, Any]],
        emit_log: Callable[[str], None] = print,
    ) -> Dict[str, Any]:
        started = time.perf_counter()
        cfg = self._cfg()
        if not self._bool(cfg.get("enabled"), True):
            return {"ok": True, "status": "skipped", "reason": "disabled"}

        required = self._bool(cfg.get("required"), False)
        dry_run = self._bool(cfg.get("dry_run"), False)
        page_size = max(1, self._as_int(cfg.get("page_size"), 500))
        batch_size = max(1, min(500, self._as_int(cfg.get("batch_size"), 200)))
        data_center_name = self._text(cfg.get("data_center_name")) or self._text(cfg.get("dataCenterName")) or "EA118"
        report_date_slash = self._normalize_date(report_date)
        target_tables, missing = self._iter_target_tables(cfg)
        if missing or not target_tables:
            message = "；".join(missing or ["未配置任何动环功率统计目标表"])
            if required:
                raise RuntimeError(f"楼栋全机柜功率统计同步配置缺失: {message}")
            self._emit(emit_log, f"[楼栋全机柜功率统计同步] 已跳过: {message}")
            return {
                "ok": True,
                "status": "skipped",
                "reason": "missing_target_table_config",
                "missing": missing,
            }

        detail_index = self._build_detail_index(detail_records)
        building_file_map: Dict[str, Path] = {}
        for item in source_units if isinstance(source_units, list) else []:
            if not isinstance(item, dict):
                continue
            building = self._text(item.get("building"))
            source_files = item.get("source_files", {}) if isinstance(item.get("source_files", {}), dict) else {}
            file_path = self._text(item.get("full_cabinet_power_file") or source_files.get("full_cabinet_power_file"))
            if building and file_path and building not in building_file_map:
                building_file_map[building] = Path(file_path)
        if not building_file_map:
            raise RuntimeError("缺少楼栋全机柜功率整日源文件")

        cabinet_rows: List[_CabinetMetricRow] = []
        line_head_rows: List[_LineHeadMetricRow] = []
        row_line_rows: List[_RowLineMetricRow] = []
        for building, file_path in building_file_map.items():
            parsed = self._parse_metric_file(file_path=file_path, building=building, business_date=report_date)
            cabinet_rows.extend(parsed.get("cabinet", []))
            line_head_rows.extend(parsed.get("line_head", []))
            row_line_rows.extend(parsed.get("row_line", []))
        if not cabinet_rows and not line_head_rows and not row_line_rows:
            raise RuntimeError("楼栋全机柜功率源文件未解析出任何可用统计数据")

        generated = {
            "cabinet": self._generate_cabinet_rows(
                cabinet_rows,
                detail_index=detail_index,
                threshold=next((table.threshold for table in target_tables if table.key == "cabinet"), 18.0),
                report_date=report_date_slash,
                data_center_name=data_center_name,
            ),
            "line_head": self._generate_line_head_rows(
                line_head_rows,
                threshold=next((table.threshold for table in target_tables if table.key == "line_head"), 107.5),
                report_date=report_date_slash,
                data_center_name=data_center_name,
            ),
            "row_line": self._generate_row_line_rows(
                row_line_rows,
                threshold=next((table.threshold for table in target_tables if table.key == "row_line"), 215.0),
                report_date=report_date_slash,
                data_center_name=data_center_name,
            ),
        }

        clients: Dict[str, Any] = {}

        def _client_for(table: _PowerAlertTable):
            if table.app_token not in clients:
                clients[table.app_token] = self._new_client(
                    app_token=table.app_token,
                    table_id=table.table_id,
                    emit_log=emit_log,
                )
            return clients[table.app_token]

        results: Dict[str, Any] = {}
        for table in target_tables:
            results[table.key] = self._replace_target_rows(
                client=_client_for(table),
                table=table,
                rows=generated.get(table.key, []),
                report_date=report_date_slash,
                dry_run=dry_run,
                page_size=page_size,
                batch_size=batch_size,
                emit_log=emit_log,
            )
        elapsed_ms = int((time.perf_counter() - started) * 1000)
        return {
            "ok": True,
            "status": "success",
            "report_date": report_date_slash,
            "dry_run": dry_run,
            "source_rows": {
                "cabinet": len(cabinet_rows),
                "line_head": len(line_head_rows),
                "row_line": len(row_line_rows),
            },
            "targets": results,
            "elapsed_ms": elapsed_ms,
        }


class BranchCompatibleHeaderMixin:
    @staticmethod
    def parse_header_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value.replace(minute=0, second=0, microsecond=0)
        text = str(value or "").strip().replace("/", "-")
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d %H"):
            try:
                return datetime.strptime(text, fmt).replace(minute=0, second=0, microsecond=0)
            except ValueError:
                continue
        return None
