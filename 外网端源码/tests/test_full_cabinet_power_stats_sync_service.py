from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from handover_log_module.service.branch_power_upload_service import BranchPowerUploadService
from handover_log_module.service.full_cabinet_power_stats_sync_service import (
    FullCabinetPowerStatsSyncService,
)
from handover_log_module.service.power_alert_sync_service import _PowerAlertTable, _SourceRow


def _write_full_cabinet_source(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([None] * 28)
    header = [None, None, None, None]
    start = datetime(2026, 5, 31, 0, 0, 0)
    header.extend([start + timedelta(hours=hour) for hour in range(24)])
    sheet.append(header)
    sheet.append([None] * len(header))

    cabinet_values = [20.5] + [1.0] * 23
    line_head_values = [130.0] + [10.0] * 23
    row_line_values = [220.0] + [10.0] * 23

    sheet.append([None, "南通阿里保税A区E楼/E楼/二层/包间M1 E-202", "202包间A列功率和", "A01机柜功率和", *cabinet_values])
    sheet.append([None, None, "E-202-A列A路-DC019", "总_负载功率_KW", *line_head_values])
    sheet.append([None, None, "E-202列头柜功率和", "A列功率和", *row_line_values])
    workbook.save(path)


class FullCabinetPowerStatsSyncServiceTests(unittest.TestCase):
    def test_parse_metric_file_detects_header_below_first_row(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "full_cabinet.xlsx"
            _write_full_cabinet_source(source_path)
            service = FullCabinetPowerStatsSyncService({})

            parsed = service._parse_metric_file(
                file_path=source_path,
                building="E楼",
                business_date="2026-05-31",
            )

            self.assertEqual(len(parsed["cabinet"]), 1)
            self.assertEqual(len(parsed["line_head"]), 1)
            self.assertEqual(len(parsed["row_line"]), 1)
            self.assertEqual(parsed["cabinet"][0].room_code, "E-202")
            self.assertEqual(parsed["cabinet"][0].cabinet_col, "A")
            self.assertEqual(parsed["cabinet"][0].cabinet_no, "01")
            self.assertEqual(parsed["line_head"][0].line_raw, "E-202-A列A路-DC019")
            self.assertEqual(parsed["row_line"][0].row_col, "A")

    def test_parse_metric_file_accepts_hyphenated_dc_line_head(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "full_cabinet.xlsx"
            _write_full_cabinet_source(source_path)
            workbook = openpyxl.load_workbook(source_path)
            try:
                sheet = workbook.active
                sheet["C5"] = "A-202-A列-DC-001"
                workbook.save(source_path)
            finally:
                workbook.close()
            service = FullCabinetPowerStatsSyncService({})

            parsed = service._parse_metric_file(
                file_path=source_path,
                building="A楼",
                business_date="2026-05-31",
            )

            self.assertEqual(len(parsed["line_head"]), 1)
            self.assertEqual(parsed["line_head"][0].line["type"], "DC")
            self.assertEqual(parsed["line_head"][0].line["num"], "001")

    def test_parse_metric_file_accepts_c_building_row_line_total_format(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "full_cabinet.xlsx"
            _write_full_cabinet_source(source_path)
            workbook = openpyxl.load_workbook(source_path)
            try:
                sheet = workbook.active
                sheet["B6"] = "南通阿里保税A区C楼/C楼/二层/包间M1 C-202"
                sheet["C6"] = "202包间A列功率和"
                sheet["D6"] = "202包间A列列头柜总功率和"
                workbook.save(source_path)
            finally:
                workbook.close()
            service = FullCabinetPowerStatsSyncService({})

            parsed = service._parse_metric_file(
                file_path=source_path,
                building="C楼",
                business_date="2026-05-31",
            )

            self.assertEqual(len(parsed["row_line"]), 1)
            self.assertEqual(parsed["row_line"][0].room_short, "C-202")
            self.assertEqual(parsed["row_line"][0].row_col, "A")

    def test_generate_line_head_rows_uses_placeholder_when_opposite_missing(self) -> None:
        service = FullCabinetPowerStatsSyncService({})
        line_head_row = service._parse_metric_file(
            file_path=self._temp_full_cabinet_source(),
            building="E楼",
            business_date="2026-05-31",
        )["line_head"][0]

        rows = service._generate_line_head_rows(
            [line_head_row],
            threshold=107.5,
            report_date="2026/05/31",
            data_center_name="EA118",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["对侧机列"], "/")
        self.assertEqual(rows[0]["对侧机列最大功率"], "/")

    def test_generate_line_head_rows_backfills_opposite_from_old_detail_rows(self) -> None:
        service = FullCabinetPowerStatsSyncService({})
        line_head_row = service._parse_metric_file(
            file_path=self._temp_full_cabinet_source(),
            building="E楼",
            business_date="2026-05-31",
        )["line_head"][0]
        legacy_stats = service._build_legacy_line_group_stats(
            [
                {
                    "机楼": "E楼",
                    "包间": "E-202包间",
                    "机列": "E-202-A列-AC019",
                    "支路编号": "A01-A1",
                    "PDU编号": "1",
                    "功率-0:00": 45.0,
                    "功率-1:00": 99.0,
                }
            ]
        )

        rows = service._generate_line_head_rows(
            [line_head_row],
            threshold=107.5,
            report_date="2026/05/31",
            data_center_name="EA118",
            legacy_line_group_stats=legacy_stats,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["对侧机列"], "A列-AC019")
        self.assertEqual(rows[0]["对侧机列最大功率"], "99kw")

    def test_threshold_stats_keeps_cross_day_continuity_in_sqlite(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FullCabinetPowerStatsSyncService({"paths": {"runtime_state_root": temp_dir}})
            object_key = service._power_alert_object_key("C楼", "C-301", "H列-DC001")

            previous = service._threshold_stats(
                [120.0] * 24,
                107.5,
                table_key="line_head",
                object_key=object_key,
                report_date="2026/05/03",
                source_hint=object_key,
            )
            current_values = (
                [120.0] * 11
                + [10.0] * 4
                + [120.0] * 2
                + [10.0] * 2
                + [120.0]
                + [10.0]
                + [120.0] * 3
            )
            current = service._threshold_stats(
                current_values,
                107.5,
                table_key="line_head",
                object_key=object_key,
                report_date="2026/05/04",
                source_hint=object_key,
            )

            self.assertEqual(previous["over_count"], 24)
            self.assertEqual(previous["runs"], 1)
            self.assertTrue(previous["end_over"])
            self.assertEqual(current["over_count"], 17)
            self.assertEqual(current["runs"], 3)
            self.assertTrue(current["previous_end_over"])

    def test_threshold_stats_ignores_boundary_point_after_24_hours(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FullCabinetPowerStatsSyncService({"paths": {"runtime_state_root": temp_dir}})

            stats = service._threshold_stats(
                [120.0] * 25,
                107.5,
                previous_end_over=False,
            )

            self.assertEqual(stats["over_count"], 24)
            self.assertEqual(stats["runs"], 1)
            self.assertNotIn(24, stats["over_hours"])

    def test_branch_rows_write_fixed_data_center_name_to_machine_room(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FullCabinetPowerStatsSyncService({"paths": {"runtime_state_root": temp_dir}})
            row = _SourceRow(
                building="A楼",
                room="A-301包间",
                room_short="A-301",
                line_raw="A-301-C列-DC010",
                line=service._parse_line("A-301-C列-DC010"),
                pdu="C02-A2",
                pdu_info=service._parse_pdu("C02-A2"),
                branch_no="20",
                powers=[8.28] * 24,
            )

            rows = service._generate_branch_rows(
                [row],
                threshold=6.25,
                report_date="2026/05/31",
                data_center_name="EA118",
            )

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["机房"], "EA118")
            self.assertNotEqual(rows[0]["机房"], "A-301-C列-DC010")
            self.assertEqual(rows[0]["支路编号"], "C列-DC010 #20")

    def test_branch_rows_only_use_same_feed_opposite_pdu(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FullCabinetPowerStatsSyncService({"paths": {"runtime_state_root": temp_dir}})

            def make_row(pdu: str, branch_no: str, power: float = 7.0) -> _SourceRow:
                return _SourceRow(
                    building="B楼",
                    room="B-302包间",
                    room_short="B-302",
                    line_raw="B-302-H列-DC012",
                    line=service._parse_line("B-302-H列-DC012"),
                    pdu=pdu,
                    pdu_info=service._parse_pdu(pdu),
                    branch_no=branch_no,
                    powers=[power] * 24,
                )

            rows = service._generate_branch_rows(
                [
                    make_row("H01-B1", "5"),
                    make_row("H01-A1", "5"),
                    make_row("H01-A2", "38"),
                    make_row("H01-B2", "19", 2.4),
                    make_row("H01-B1", "1"),
                ],
                threshold=6.25,
                report_date="2026/05/31",
                data_center_name="EA118",
            )

            by_key = {(row["PDU编号"], row["支路号"]): row for row in rows}
            self.assertEqual(by_key[("H01-B1", "5")]["对侧PDU编号"], "H01-A1")
            self.assertEqual(by_key[("H01-A1", "5")]["对侧PDU编号"], "H01-B1")
            self.assertEqual(by_key[("H01-A2", "38")]["对侧PDU编号"], "H01-B2")
            self.assertEqual(by_key[("H01-A2", "38")]["对侧支路功率"], "2.4")
            self.assertEqual(by_key[("H01-B1", "1")]["对侧PDU编号"], "H01-A1")
            self.assertEqual(by_key[("H01-B1", "1")]["对侧支路功率"], "7")

    def test_branch_rows_pick_nonzero_opposite_pdu_without_same_branch_no(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = FullCabinetPowerStatsSyncService({"paths": {"runtime_state_root": temp_dir}})

            def make_row(pdu: str, branch_no: str, powers: list[float]) -> _SourceRow:
                return _SourceRow(
                    building="B楼",
                    room="B-402包间",
                    room_short="B-402",
                    line_raw="B-402-J列-DC010",
                    line=service._parse_line("B-402-J列-DC010"),
                    pdu=pdu,
                    pdu_info=service._parse_pdu(pdu),
                    branch_no=branch_no,
                    powers=powers,
                )

            source_powers = [0.0] * 24
            source_powers[3] = 6.6
            opposite_missing_branch_powers = [0.0] * 24
            opposite_real_branch_powers = [0.0] * 24
            opposite_real_branch_powers[3] = 0.873

            rows = service._generate_branch_rows(
                [
                    make_row("J01-B1", "1", source_powers),
                    make_row("J01-A1", "1", opposite_missing_branch_powers),
                    make_row("J01-A1", "31", opposite_real_branch_powers),
                ],
                threshold=6.25,
                report_date="2026/05/31",
                data_center_name="EA118",
            )

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["PDU编号"], "J01-B1")
            self.assertEqual(rows[0]["支路号"], "1")
            self.assertEqual(rows[0]["对侧PDU编号"], "J01-A1")
            self.assertEqual(rows[0]["对侧支路功率"], "0.873")

    def test_generate_cabinet_rows_backfills_pdu_and_current_from_old_detail_rows(self) -> None:
        service = FullCabinetPowerStatsSyncService({})
        cabinet_row = service._parse_metric_file(
            file_path=self._temp_full_cabinet_source(),
            building="E楼",
            business_date="2026-05-31",
        )["cabinet"][0]
        detail_records = [
            {
                "机楼": "E楼",
                "包间": "E-202",
                "机列": "E-202-A列A路-DC019",
                "支路编号": "A01-A1",
                "PDU编号": "1",
                "功率-0:00": 20.5,
                "电流-0:00": 5.2,
            }
        ]
        detail_index = service._build_detail_index(detail_records)

        rows = service._generate_cabinet_rows(
            [cabinet_row],
            detail_index=detail_index,
            threshold=18.0,
            report_date="2026/05/31",
            data_center_name="EA118",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["房间"], "E-202包间")
        self.assertEqual(rows[0]["机柜号"], "A列A01")
        self.assertEqual(rows[0]["PDU编号"], "A01-A1")
        self.assertEqual(rows[0]["电流值"], 5.2)
        self.assertEqual(rows[0]["次数"], 1)
        self.assertEqual(rows[0]["时长"], "1h")

    def test_generate_cabinet_rows_keeps_multiple_pdu_records_for_same_cabinet(self) -> None:
        service = FullCabinetPowerStatsSyncService({})
        cabinet_row = service._parse_metric_file(
            file_path=self._temp_full_cabinet_source(),
            building="E楼",
            business_date="2026-05-31",
        )["cabinet"][0]
        detail_records = [
            {
                "机楼": "E楼",
                "包间": "E-202包间",
                "机列": "E-202-A列A路-DC019",
                "支路编号": pdu,
                "PDU编号": branch,
                "功率-0:00": 20.5,
                "电流-0:00": current,
            }
            for pdu, branch, current in (
                ("A01-A1", "1", 5.2),
                ("A01-A2", "2", 5.3),
                ("A01-B1", "1", 4.8),
                ("A01-B2", "2", 4.9),
            )
        ]
        detail_index = service._build_detail_index(detail_records)

        rows = service._generate_cabinet_rows(
            [cabinet_row],
            detail_index=detail_index,
            threshold=18.0,
            report_date="2026/05/31",
            data_center_name="EA118",
        )

        self.assertEqual(len(rows), 4)
        self.assertEqual({row["机柜号"] for row in rows}, {"A列A01"})
        self.assertEqual([row["PDU编号"] for row in rows], ["A01-A1", "A01-A2", "A01-B1", "A01-B2"])
        self.assertEqual([row["电流值"] for row in rows], [5.2, 5.3, 4.8, 4.9])

    def test_generate_cabinet_rows_keeps_record_when_old_detail_rows_missing(self) -> None:
        service = FullCabinetPowerStatsSyncService({})
        cabinet_row = service._parse_metric_file(
            file_path=self._temp_full_cabinet_source(),
            building="E楼",
            business_date="2026-05-31",
        )["cabinet"][0]

        rows = service._generate_cabinet_rows(
            [cabinet_row],
            detail_index={},
            threshold=18.0,
            report_date="2026/05/31",
            data_center_name="EA118",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["房间"], "E-202包间")
        self.assertEqual(rows[0]["机柜号"], "A列A01")
        self.assertIsNone(rows[0]["PDU编号"])
        self.assertIsNone(rows[0]["电流值"])
        self.assertEqual(rows[0]["次数"], 1)
        self.assertEqual(rows[0]["时长"], "1h")

    def test_branch_source_header_detection_scans_multiple_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "branch_like.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append([None] * 8)
            sheet.append([None, None, None, datetime(2026, 5, 31, 0, 0, 0), datetime(2026, 5, 31, 1, 0, 0)])
            sheet.append([None] * 8)
            workbook.save(source_path)

            bucket_keys = BranchPowerUploadService._detect_header_bucket_keys(source_path)

            self.assertEqual(bucket_keys[:2], ["2026-05-31 00", "2026-05-31 01"])

    def test_replace_target_rows_does_not_delete_old_records_when_create_fails(self) -> None:
        class FakeClient:
            def __init__(self) -> None:
                self.deleted: list[str] = []
                self.create_called = False

            def list_records(self, **_kwargs):
                return [{"record_id": "old_1", "fields": {"数据时间": "2026/05/31"}}]

            def batch_create_records(self, **_kwargs):
                self.create_called = True
                raise RuntimeError("create failed")

            def batch_delete_records(self, **_kwargs):
                self.deleted.extend(_kwargs.get("record_ids", []))
                return len(self.deleted)

        class SafeReplaceService(FullCabinetPowerStatsSyncService):
            def _field_meta_map(self, client, table):  # noqa: ANN001
                return {name: {"name": name, "type": 1, "property": {}} for name in self.TARGET_FIELDS[table.key]}

        service = SafeReplaceService({})
        table = _PowerAlertTable(
            key="cabinet",
            name="机柜超18KW统计",
            table_id="tbl_test",
            view_id="",
            threshold=18,
        )
        client = FakeClient()

        with self.assertRaisesRegex(RuntimeError, "create failed"):
            service._replace_target_rows(
                client=client,
                table=table,
                rows=[
                    {
                        "序号": "1",
                        "数据时间": "2026/05/31",
                        "机房": "EA118",
                        "楼栋": "E楼",
                        "房间": "E-202",
                        "机柜号": "A01",
                        "机柜功率": "20kw",
                    }
                ],
                report_date="2026/05/31",
                dry_run=False,
                page_size=500,
                batch_size=200,
                emit_log=lambda _message: None,
            )

        self.assertTrue(client.create_called)
        self.assertEqual(client.deleted, [])

    def _temp_full_cabinet_source(self) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        source_path = Path(temp_dir.name) / "full_cabinet.xlsx"
        _write_full_cabinet_source(source_path)
        return source_path


if __name__ == "__main__":
    unittest.main()
