from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from handover_log_module.service.handover_110_station_upload_service import (
    Handover110StationUploadService,
)


ROW_LABELS = ("阿开", "阿家", "1号主变", "2号主变", "3号主变", "4号主变")


def _write_substation_workbook(path: Path, *, shift: str, target_sheet_index: int) -> None:
    workbook = Workbook()
    workbook.active.title = "源数据"
    while len(workbook.worksheets) <= target_sheet_index:
        workbook.create_sheet(f"占位{len(workbook.worksheets)}")
    worksheet = workbook.worksheets[target_sheet_index]
    worksheet.title = "白班110" if shift == "day" else "夜班110"
    label_column = 4 if shift == "day" else 5
    value_columns = (5, 6, 7, 8, 9) if shift == "day" else (6, 7, 8, 9, 10)
    for row_index, label in enumerate(ROW_LABELS, start=3):
        worksheet.cell(row=row_index, column=label_column, value=label)
        for offset, column in enumerate(value_columns, start=1):
            worksheet.cell(row=row_index, column=column, value=row_index * 10 + offset)
    workbook.save(path)
    workbook.close()


def _complete_rows():
    return [
        {
            "row_id": row_id,
            "label": label,
            "group": "incoming" if row_id.startswith("incoming") else "transformer",
            "line_voltage": "110",
            "current": "10",
            "power_kw": "1000",
            "power_factor": "0.99",
            "load_rate": "20%",
        }
        for row_id, label in (
            ("incoming_akai", "阿开"),
            ("incoming_ajia", "阿家"),
            ("transformer_1", "1#主变"),
            ("transformer_2", "2#主变"),
            ("transformer_3", "3#主变"),
            ("transformer_4", "4#主变"),
        )
    ]


def test_parse_workbook_finds_day_layout_after_sheet_order_changes(tmp_path):
    path = tmp_path / "moved_day.xlsx"
    _write_substation_workbook(path, shift="day", target_sheet_index=3)
    service = object.__new__(Handover110StationUploadService)

    parsed = service._parse_workbook(path, duty_shift="day")

    assert parsed["substation_sheet"]["title"] == "白班110"
    assert parsed["substation_sheet"]["sheet_index"] == 4
    assert parsed["substation_sheet"]["detected_shift"] == "day"
    assert [row["row_id"] for row in parsed["parsed_110kv_rows"]] == [
        "incoming_akai",
        "incoming_ajia",
        "transformer_1",
        "transformer_2",
        "transformer_3",
        "transformer_4",
    ]


def test_parse_workbook_reports_shift_mismatch_instead_of_missing_all_rows(tmp_path):
    path = tmp_path / "night.xlsx"
    _write_substation_workbook(path, shift="night", target_sheet_index=2)
    service = object.__new__(Handover110StationUploadService)

    with pytest.raises(ValueError, match="识别为夜班版式，与当前白班批次不一致"):
        service._parse_workbook(path, duty_shift="day")


def test_manual_complete_rows_recover_failed_upload_state(tmp_path):
    source_path = tmp_path / "110.xlsx"
    source_path.write_bytes(b"placeholder")
    service = object.__new__(Handover110StationUploadService)
    service.config = {}
    saved = {}

    class ReviewService:
        @staticmethod
        def build_batch_key(duty_date, duty_shift):
            return f"{duty_date}|{duty_shift}"

        @staticmethod
        def get_cloud_batch(_batch_key):
            return {}

    service._review_service = ReviewService()
    service._load_state = lambda _batch_key: {
        "batch_key": "2026-07-13|day",
        "duty_date": "2026-07-13",
        "duty_shift": "day",
        "status": "failed",
        "error": "旧解析错误",
        "stored_path": str(source_path),
        "cloud_sync": {"status": "failed"},
    }

    def save_state(_batch_key, payload):
        saved.update(payload)
        return dict(payload)

    service._save_state = save_state
    service._save_substation_rows = lambda **_kwargs: {}

    result = service.update_parsed_rows(
        duty_date="2026-07-13",
        duty_shift="day",
        rows=_complete_rows(),
        emit_log=lambda _message: None,
    )

    assert result["upload"]["status"] == "success"
    assert result["upload"]["error"] == ""
    assert result["upload"]["cloud_sync"]["status"] == "pending"
    assert saved["parsed_110kv_rows"][2]["label"] == "1#主变"


def test_manual_rows_do_not_accept_blank_skeleton():
    rows = _complete_rows()
    rows[2].update({key: "" for key in ("line_voltage", "current", "power_kw", "power_factor", "load_rate")})

    with pytest.raises(ValueError, match="整行无数据: 1#主变"):
        Handover110StationUploadService._normalize_edited_substation_rows(rows)


def test_manual_rows_preserve_numeric_zero_values():
    rows = _complete_rows()
    rows[0]["power_kw"] = 0

    normalized = Handover110StationUploadService._normalize_edited_substation_rows(rows)

    assert normalized[0]["power_kw"] == "0"


def test_existing_successful_same_file_cloud_sync_is_idempotent(tmp_path):
    source_path = tmp_path / "110.xlsx"
    source_path.write_bytes(b"content")
    service = object.__new__(Handover110StationUploadService)
    service.config = {}
    called = []
    state = {
        "status": "success",
        "stored_path": str(source_path),
        "cloud_sync": {
            "status": "success",
            "source_file": str(source_path),
            "spreadsheet_token": "sheet_token",
        },
    }

    class ReviewService:
        @staticmethod
        def get_cloud_batch(_batch_key):
            return {"spreadsheet_token": "sheet_token"}

    service._review_service = ReviewService()
    service._load_state = lambda _batch_key: state
    service._sync_state_to_cloud = lambda **_kwargs: called.append(True)

    result = service.sync_existing_upload_to_cloud(
        batch_key="2026-07-13|day",
        emit_log=lambda _message: None,
    )

    assert result["status"] == "skipped"
    assert result["reason"] == "already_synced_same_source"
    assert called == []
