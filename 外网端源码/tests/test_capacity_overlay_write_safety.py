from __future__ import annotations

from openpyxl import Workbook, load_workbook

from handover_log_module.service.handover_capacity_report_service import HandoverCapacityReportService


def test_incomplete_overlay_does_not_modify_existing_report(tmp_path) -> None:
    path = tmp_path / "capacity.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "original"
    workbook.save(path)
    workbook.close()

    result = HandoverCapacityReportService({}).sync_overlay_for_existing_report_from_cells(
        building="A楼",
        duty_date="2026-08-13",
        duty_shift="day",
        handover_cells={},
        capacity_output_file=str(path),
        emit_log=lambda _text: None,
    )

    assert result["status"] == "pending_input"
    checked = load_workbook(path)
    try:
        assert checked.active["A1"].value == "original"
    finally:
        checked.close()
