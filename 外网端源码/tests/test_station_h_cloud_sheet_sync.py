from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from handover_log_module.service.handover_cloud_sheet_sync_service import HandoverCloudSheetSyncService


def _write_station_h_template(path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "H楼"
    worksheet["H50"] = "template-boundary"
    worksheet["H12"].font = Font(name="宋体", size=10, bold=True, color="000000")
    worksheet["H12"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["H12"].border = Border(right=Side(style="thin", color="1F2329"))
    workbook.save(path)
    workbook.close()


class _FakeSheetsClient:
    def __init__(self, target_sheet):
        self.target_sheet = dict(target_sheet) if target_sheet else {}
        self.value_batches = []
        self.style_batches = []
        self.image_writes = []

    def query_sheets(self, _spreadsheet_token, *, sheet_cache=None, force_refresh=False):
        del sheet_cache, force_refresh
        return [dict(self.target_sheet)] if self.target_sheet else []

    def dedupe_named_sheets(self, _spreadsheet_token, _sheet_title, *, sheet_cache=None):
        del sheet_cache
        return dict(self.target_sheet)

    def batch_update_values(self, _spreadsheet_token, value_ranges):
        self.value_batches.append(list(value_ranges))
        return {}

    def batch_update_styles(self, _spreadsheet_token, style_ranges):
        self.style_batches.append(list(style_ranges))
        return {}

    def write_cell_image(self, spreadsheet_token, *, range_name, image_path, name=""):
        self.image_writes.append(
            {
                "spreadsheet_token": spreadsheet_token,
                "range_name": range_name,
                "image_path": str(image_path),
                "name": name,
            }
        )
        return {}

    def batch_unmerge_cells(self, *_args, **_kwargs):
        raise AssertionError("H楼值同步不应取消合并单元格")

    def batch_merge_cells(self, *_args, **_kwargs):
        raise AssertionError("H楼值同步不应重建合并单元格")

    def batch_clear_values(self, *_args, **_kwargs):
        raise AssertionError("H楼值同步不应清空整张表")

    def delete_dimension(self, *_args, **_kwargs):
        raise AssertionError("H楼值同步不应删除行列")

    def add_dimension(self, *_args, **_kwargs):
        raise AssertionError("H楼值同步不应增补行列")


def _service_with_client(tmp_path, client):
    template_path = tmp_path / "H楼交接班日志空模板.xlsx"
    _write_station_h_template(template_path)
    service = HandoverCloudSheetSyncService(
        {"cloud_sheet_sync": {"station_h_template_path": str(template_path)}}
    )
    service._build_client = lambda: client
    return service


def test_station_h_sync_updates_only_allowed_values_and_preserves_cloud_layout(tmp_path):
    target_sheet = {
        "sheet_id": "h_sheet",
        "title": "H楼",
        "index": 5,
        "row_count": 200,
        "column_count": 20,
        "merges": [
            {
                "start_row_index": 1,
                "end_row_index": 2,
                "start_column_index": 1,
                "end_column_index": 4,
            }
        ],
    }
    client = _FakeSheetsClient(target_sheet)
    service = _service_with_client(tmp_path, client)

    result = service.sync_station_h_sheet(
        batch_meta={"batch_key": "2026-08-03|day", "spreadsheet_token": "sheet_token"},
        cell_values={"B2": "2026-08-03", "H12": 7.5, "H7": "不得覆盖的静态值"},
        emit_log=lambda _message: None,
    )

    assert result["status"] == "success"
    assert result["rebuild_mode"] == "values_only_preserve_layout"
    assert result["synced_row_count"] == 200
    assert result["synced_column_count"] == 20
    assert len(client.value_batches) == 1
    assert len(client.style_batches) == 1

    value_ranges = client.value_batches[0]
    assert len(value_ranges) == len(service.STATION_H_ALLOWED_CELLS)
    values_by_range = {item["range"]: item["values"][0][0] for item in value_ranges}
    assert values_by_range["h_sheet!B2:B2"] == "2026-08-03"
    assert values_by_range["h_sheet!H12:H12"] == 7.5
    assert "h_sheet!H7:H7" not in values_by_range
    assert "h_sheet!H8:H8" not in values_by_range
    assert "h_sheet!H9:H9" not in values_by_range

    style_ranges = client.style_batches[0]
    h12_styles = [
        item["style"]
        for item in style_ranges
        if any("H12" in range_name for range_name in item.get("ranges", []))
    ]
    assert any(
        style.get("hAlign") == 1
        and style.get("vAlign") == 1
        and style.get("font", {}).get("bold") is True
        and style.get("font", {}).get("fontSize") == "10pt/1.5"
        for style in h12_styles
    )
    assert any(
        style.get("borderType") == "RIGHT_BORDER"
        and style.get("borderColor") == "#1F2329"
        for style in h12_styles
    )


def test_station_h_sync_rejects_sheet_without_template_merge_structure(tmp_path):
    client = _FakeSheetsClient(
        {
            "sheet_id": "h_sheet",
            "title": "H楼",
            "index": 5,
            "row_count": 200,
            "column_count": 20,
            "merges": [],
        }
    )
    service = _service_with_client(tmp_path, client)

    result = service.sync_station_h_sheet(
        batch_meta={"batch_key": "2026-08-03|day", "spreadsheet_token": "sheet_token"},
        cell_values={"B2": "2026-08-03"},
        emit_log=lambda _message: None,
    )

    assert result["status"] == "failed"
    assert "未检测到模板合并结构" in result["error"]
    assert client.value_batches == []
    assert client.style_batches == []


def test_station_h_sync_writes_duty_focus_image_only_to_merged_j2(tmp_path):
    target_sheet = {
        "sheet_id": "h_sheet",
        "title": "H楼",
        "index": 5,
        "row_count": 200,
        "column_count": 20,
        "merges": [
            {
                "start_row_index": 1,
                "end_row_index": 20,
                "start_column_index": 9,
                "end_column_index": 16,
            }
        ],
    }
    image_path = tmp_path / "值班关注点.png"
    image_path.write_bytes(b"\x89PNG\r\n\x1a\nimage")
    client = _FakeSheetsClient(target_sheet)
    service = _service_with_client(tmp_path, client)

    result = service.sync_station_h_sheet(
        batch_meta={"batch_key": "2026-08-03|day", "spreadsheet_token": "sheet_token"},
        cell_values={"B2": "2026-08-03"},
        duty_focus_image_path=image_path,
        emit_log=lambda _message: None,
    )

    assert result["status"] == "success"
    assert result["duty_focus_image_synced"] is True
    assert result["duty_focus_image_cell"] == "J2"
    assert client.image_writes == [
        {
            "spreadsheet_token": "sheet_token",
            "range_name": "h_sheet!J2:J2",
            "image_path": str(image_path),
            "name": image_path.name,
        }
    ]


def test_station_h_sync_rejects_image_when_j2_is_not_merge_anchor(tmp_path):
    target_sheet = {
        "sheet_id": "h_sheet",
        "title": "H楼",
        "index": 5,
        "row_count": 200,
        "column_count": 20,
        "merges": [
            {
                "start_row_index": 1,
                "end_row_index": 20,
                "start_column_index": 8,
                "end_column_index": 16,
            }
        ],
    }
    image_path = tmp_path / "值班关注点.png"
    image_path.write_bytes(b"\x89PNG\r\n\x1a\nimage")
    client = _FakeSheetsClient(target_sheet)
    service = _service_with_client(tmp_path, client)

    result = service.sync_station_h_sheet(
        batch_meta={"batch_key": "2026-08-03|day", "spreadsheet_token": "sheet_token"},
        cell_values={"B2": "2026-08-03"},
        duty_focus_image_path=image_path,
        emit_log=lambda _message: None,
    )

    assert result["status"] == "failed"
    assert "J2 不是合并区域左上角" in result["error"]
    assert client.image_writes == []
