from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from handover_log_module.repository.review_building_document_store import ReviewBuildingDocumentStore
from handover_log_module.service.handover_xlsx_write_queue_service import HandoverXlsxWriteQueueService
from handover_log_module.service.review_document_writer import ReviewDocumentWriter


def _store(tmp_path: Path) -> ReviewBuildingDocumentStore:
    return ReviewBuildingDocumentStore(
        config={"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}},
        building="A楼",
    )


def _session(tmp_path: Path, *, revision: int = 1) -> dict:
    output_file = tmp_path / "handover.xlsx"
    output_file.write_bytes(b"placeholder")
    return {
        "session_id": "A楼|2026-07-09|day",
        "building": "A楼",
        "duty_date": "2026-07-09",
        "duty_shift": "day",
        "batch_key": "2026-07-09|day",
        "revision": revision,
        "output_file": str(output_file),
    }


def test_save_document_keeps_previous_unsynced_dirty_regions(tmp_path: Path) -> None:
    store = _store(tmp_path)
    session = _session(tmp_path)
    store.upsert_imported_document(session=session, document={"fixed_blocks": []}, imported_from_excel=True)

    first, _ = store.save_document(
        session=session,
        document={"fixed_blocks": [{"fields": [{"cell": "B4", "value": "长白岗：甲"}]}]},
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    assert first["dirty_regions"] == {"fixed_blocks": True}

    second, _ = store.save_document(
        session=session,
        document={
            "fixed_blocks": [{"fields": [{"cell": "B4", "value": "长白岗：甲"}]}],
            "sections": [{"name": "一、本班完成工作"}],
        },
        base_revision=2,
        dirty_regions={"sections": True},
    )

    assert second["dirty_regions"]["fixed_blocks"] is True
    assert second["dirty_regions"]["sections"] is True


def test_save_document_does_not_keep_previous_dirty_regions_after_sync(tmp_path: Path) -> None:
    store = _store(tmp_path)
    session = _session(tmp_path)
    store.upsert_imported_document(session=session, document={"fixed_blocks": []}, imported_from_excel=True)
    store.save_document(
        session=session,
        document={"fixed_blocks": [{"fields": [{"cell": "B4", "value": "长白岗：甲"}]}]},
        base_revision=1,
        dirty_regions={"fixed_blocks": True},
    )
    store.update_sync_state(
        session_id=session["session_id"],
        status="synced",
        synced_revision=2,
        pending_revision=0,
        error="",
    )

    second, _ = store.save_document(
        session=session,
        document={
            "fixed_blocks": [{"fields": [{"cell": "B4", "value": "长白岗：甲"}]}],
            "sections": [{"name": "一、本班完成工作"}],
        },
        base_revision=2,
        dirty_regions={"sections": True},
    )

    assert second["dirty_regions"] == {"sections": True}


def test_writer_syncs_fixed_cells_when_dirty_region_was_lost(tmp_path: Path) -> None:
    output_file = tmp_path / "handover.xlsx"
    workbook = Workbook()
    try:
        workbook.active["B4"] = "长白岗：旧值"
        workbook.save(output_file)
    finally:
        workbook.close()

    writer = ReviewDocumentWriter({"template": {"sheet_name": "Sheet"}})
    writer.write(
        output_file=str(output_file),
        document={
            "fixed_blocks": [
                {
                    "fields": [
                        {"cell": "B4", "value": "长白岗：新值"},
                    ],
                },
            ],
        },
        dirty_regions={"fixed_blocks": False, "sections": False, "footer_inventory": False},
    )

    reloaded = load_workbook(output_file, data_only=False)
    try:
        assert reloaded.active["B4"].value == "长白岗：新值"
    finally:
        reloaded.close()


def test_download_full_sync_overrides_pending_incremental_sync(tmp_path: Path) -> None:
    config = {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}}
    queue = HandoverXlsxWriteQueueService(config)
    queue._start_worker = lambda building: None  # type: ignore[method-assign]
    session = _session(tmp_path, revision=3)

    queue.enqueue_review_excel_sync(
        session,
        target_revision=3,
        force_all_regions=False,
    )
    queue.enqueue_review_excel_sync(
        session,
        target_revision=3,
        force_all_regions=True,
    )

    store = ReviewBuildingDocumentStore(config=config, building="A楼")
    with store.connect(read_only=True) as conn:
        rows = conn.execute(
            """
            SELECT * FROM xlsx_write_jobs
             WHERE task_type='review_excel_sync'
               AND dedupe_key=?
               AND status='pending'
            """,
            (session["session_id"],),
        ).fetchall()

    assert len(rows) == 1
    assert json.loads(rows[0]["payload_json"])["force_all_regions"] is True
