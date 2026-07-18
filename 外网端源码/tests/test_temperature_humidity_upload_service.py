from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.modules.temperature_humidity_upload.service.temperature_humidity_upload_service import (
    TemperatureHumidityUploadService,
)


def _write_source_file(
    path: Path,
    *,
    temperature: float,
    humidity: float,
    invalid_status: bool = False,
    temperature_label: str = "温度",
    humidity_label: str = "湿度",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.merge_cells("C3:C4")
    sheet["C3"] = "B-144变电所P1_TH-01"
    sheet["D3"] = temperature_label
    sheet["E3"] = temperature
    sheet["D4"] = humidity_label
    sheet["E4"] = humidity
    sheet["D6"] = "B-111-CRAH-01_运行状态"
    sheet["E6"] = 1
    sheet["D7"] = "B-111-CRAH-02_运行状态"
    sheet["E7"] = 0
    if invalid_status:
        sheet["D8"] = "B-111-CRAH-03_运行状态"
        sheet["E8"] = 2
    workbook.save(path)
    workbook.close()


class FakeInternalClient:
    def __init__(
        self,
        files: dict[str, Path],
        *,
        entry_overrides: dict[str, dict] | None = None,
    ):
        self.files = files
        self.entry_overrides = entry_overrides or {}
        self.refresh_calls = []

    def source_index_batch(self, queries, *, default_limit=50):
        output = []
        for index, query in enumerate(queries):
            building = query["building"]
            path = self.files.get(building)
            entries = []
            if path is not None:
                entry = {
                    "entry_id": f"entry-{building}",
                    "source_family": query["source_family"],
                    "building": building,
                    "bucket_kind": "daily",
                    "bucket_key": query["bucket_or_date"],
                    "duty_date": query["bucket_or_date"],
                    "status": "ready",
                    "file_verified": True,
                    "file_path": str(path),
                    "downloaded_at": "2026-07-18 09:20:00",
                }
                entry.update(self.entry_overrides.get(building, {}))
                entries.append(entry)
            output.append({"index": index, "ok": True, "entries": entries})
        return output

    def refresh_latest_source_cache(self, *, source_family, buildings, target_bucket_key=""):
        self.refresh_calls.append((source_family, list(buildings), target_bucket_key))
        return {"ok": True, "accepted_count": len(buildings)}


class FakeBitableClient:
    def __init__(self, *, fail_create: bool = False):
        self.fail_create = fail_create
        self.create_failed = False
        self.created = []
        self.deleted = []
        self.events = []
        self.next_record_number = 1
        self.records = {
            "rec-old-1": {
                "楼栋": "A楼",
                "位置": "旧温湿度位置",
                "温度": 19.5,
                "湿度": 45.0,
            }
        }

    def list_fields(self, table_id, page_size=500):
        return [
            {"field_name": "楼栋", "type": 3},
            {"field_name": "位置", "type": 1},
            {"field_name": "温度", "type": 2},
            {"field_name": "湿度", "type": 2},
            {"field_name": "运行状态", "type": 3},
        ]

    def list_records(self, table_id, page_size=500, max_records=0, *, view_id="", filter_formula="", field_names=None):
        return [
            {"record_id": record_id, "fields": dict(fields)}
            for record_id, fields in self.records.items()
        ]

    def batch_create_records(self, table_id, fields_list, batch_size=200, progress_callback=None):
        self.events.append("create")
        if self.fail_create and not self.create_failed:
            self.create_failed = True
            raise RuntimeError("create failed")
        for fields in fields_list:
            record_id = f"rec-new-{self.next_record_number}"
            self.next_record_number += 1
            self.records[record_id] = dict(fields)
            self.created.append(dict(fields))
        return [{"code": 0}]

    def batch_delete_records(self, table_id, record_ids, batch_size=500, progress_callback=None):
        self.events.append("delete")
        deleted_count = 0
        for record_id in record_ids:
            if record_id in self.records:
                self.records.pop(record_id)
                deleted_count += 1
            self.deleted.append(record_id)
        return deleted_count


def _runtime_config(buildings: list[str], *, shared_root: Path | None = None) -> dict:
    config = {
        "temperature_humidity_upload": {
            "enabled": True,
            "buildings": buildings,
            "trigger_missing_download": True,
            "wait_source_timeout_sec": 0,
            "target": {
                "app_token": "ASLxbfESPahdTKs0A9NccgbrnXc",
                "table_id": "tblfnTbEWK9607zV",
                "fields": {
                    "building": "楼栋",
                    "position": "位置",
                    "temperature": "温度",
                    "humidity": "湿度",
                    "running_status": "运行状态",
                },
            },
        }
    }
    if shared_root is not None:
        config["shared_bridge"] = {
            "root_dir": str(shared_root),
            "external_root_dir": str(shared_root),
        }
    return config


def test_parse_workbook_merges_temperature_humidity_and_maps_binary_status(tmp_path):
    source = tmp_path / "source.xlsx"
    _write_source_file(source, temperature=27.41, humidity=52.79)

    result = TemperatureHumidityUploadService.parse_workbook(source, building="A楼")

    assert result["temperature_location_count"] == 1
    assert result["status_location_count"] == 2
    assert result["unparsed_status_count"] == 0
    assert result["rows"][0] == {
        "building": "A楼",
        "position": "B-144变电所P1_TH-01",
        "temperature": 27.41,
        "humidity": 52.79,
    }
    assert result["rows"][1]["running_status"] == "开启"
    assert result["rows"][2]["running_status"] == "关闭"


def test_parse_workbook_accepts_labels_containing_temperature_or_humidity(tmp_path):
    source = tmp_path / "source-with-prefixed-labels.xlsx"
    _write_source_file(
        source,
        temperature=28.88,
        humidity=61.25,
        temperature_label="室内温度值",
        humidity_label="相对湿度",
    )

    result = TemperatureHumidityUploadService.parse_workbook(source, building="C楼")

    assert result["temperature_location_count"] == 1
    assert result["rows"][0] == {
        "building": "C楼",
        "position": "B-144变电所P1_TH-01",
        "temperature": 28.88,
        "humidity": 61.25,
    }


def test_parse_workbook_rejects_non_binary_running_status(tmp_path):
    source = tmp_path / "invalid-status.xlsx"
    _write_source_file(
        source,
        temperature=27.41,
        humidity=52.79,
        invalid_status=True,
    )

    with pytest.raises(ValueError, match="运行状态必须为0或1"):
        TemperatureHumidityUploadService.parse_workbook(source, building="A楼")


def test_parse_workbook_does_not_reuse_temperature_position_across_status_row(tmp_path):
    source = tmp_path / "broken-temperature-group.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["C3"] = "B-144变电所P1_TH-01"
    sheet["D3"] = "温度"
    sheet["E3"] = 27.41
    sheet["D4"] = "B-111-CRAH-01_运行状态"
    sheet["E4"] = 1
    sheet["D5"] = "湿度"
    sheet["E5"] = 52.79
    workbook.save(source)
    workbook.close()

    with pytest.raises(ValueError, match="缺少位置"):
        TemperatureHumidityUploadService.parse_workbook(source, building="A楼")


def test_run_validates_then_clears_old_records_before_uploading_snapshot(tmp_path):
    files = {}
    for index, building in enumerate(("A楼", "B楼"), start=1):
        path = tmp_path / f"{building}.xlsx"
        _write_source_file(path, temperature=20 + index, humidity=50 + index)
        files[building] = path
    internal = FakeInternalClient(files)
    bitable = FakeBitableClient()

    result = TemperatureHumidityUploadService(
        _runtime_config(["A楼", "B楼"]),
        internal_client=internal,
        bitable_client=bitable,
    ).run(source_date="2026-07-18")

    assert result["status"] == "success"
    assert result["uploaded_count"] == 6
    assert result["temperature_location_count"] == 2
    assert result["status_location_count"] == 4
    assert result["unparsed_status_count"] == 0
    assert bitable.events == ["delete", "create"]
    assert bitable.deleted == ["rec-old-1"]
    temperature_row = next(row for row in bitable.created if row.get("温度") == 21.0)
    assert temperature_row == {
        "楼栋": "A楼",
        "位置": "B-144变电所P1_TH-01",
        "温度": 21.0,
        "湿度": 51.0,
    }
    opened = next(row for row in bitable.created if row.get("位置") == "B-111-CRAH-01_运行状态")
    assert opened["运行状态"] == "开启"


def test_create_failure_restores_old_record_snapshot(tmp_path):
    source = tmp_path / "A楼.xlsx"
    _write_source_file(source, temperature=27.0, humidity=50.0)
    bitable = FakeBitableClient(fail_create=True)

    with pytest.raises(RuntimeError, match="create failed"):
        TemperatureHumidityUploadService(
            _runtime_config(["A楼"]),
            internal_client=FakeInternalClient({"A楼": source}),
            bitable_client=bitable,
        ).run(source_date="2026-07-18")

    assert bitable.events == ["delete", "create", "create"]
    assert bitable.deleted == ["rec-old-1"]
    assert list(bitable.records.values()) == [
        {
            "楼栋": "A楼",
            "位置": "旧温湿度位置",
            "温度": 19.5,
            "湿度": 45.0,
        }
    ]


def test_missing_source_requests_internal_refresh_without_deleting_old_records():
    internal = FakeInternalClient({})
    bitable = FakeBitableClient()
    source_date = datetime.now().strftime("%Y-%m-%d")

    with pytest.raises(RuntimeError, match="B楼"):
        TemperatureHumidityUploadService(
            _runtime_config(["A楼", "B楼"]),
            internal_client=internal,
            bitable_client=bitable,
        ).run(source_date=source_date)

    assert internal.refresh_calls == [
        ("air_conditioner_temperature_humidity_family", ["A楼", "B楼"], source_date)
    ]
    assert bitable.events == []
    assert bitable.deleted == []


@pytest.mark.parametrize(
    "entry_override",
    [
        {"source_family": "branch_power_family"},
        {"bucket_key": "2026-07-17", "duty_date": "2026-07-17"},
        {"bucket_kind": "latest"},
    ],
)
def test_wrong_source_index_entry_is_rejected_without_deleting_old_records(
    tmp_path,
    entry_override,
):
    source = tmp_path / "A楼.xlsx"
    _write_source_file(source, temperature=27.0, humidity=50.0)
    bitable = FakeBitableClient()
    source_date = datetime.now().strftime("%Y-%m-%d")
    internal = FakeInternalClient(
        {"A楼": source},
        entry_overrides={"A楼": entry_override},
    )

    with pytest.raises(RuntimeError, match="A楼"):
        TemperatureHumidityUploadService(
            _runtime_config(["A楼"]),
            internal_client=internal,
            bitable_client=bitable,
        ).run(source_date=source_date)

    assert internal.refresh_calls == [
        ("air_conditioner_temperature_humidity_family", ["A楼"], source_date)
    ]
    assert bitable.events == []


def test_internal_absolute_path_is_remapped_through_external_shared_root(tmp_path):
    shared_root = tmp_path / "external-share"
    relative_path = Path(
        "空调温湿度源文件",
        "202607",
        "20260718--整日",
        "20260718--整日--空调温湿度源文件--A楼.xlsx",
    )
    mapped_source = shared_root / relative_path
    mapped_source.parent.mkdir(parents=True, exist_ok=True)
    _write_source_file(mapped_source, temperature=27.0, humidity=50.0)
    internal = FakeInternalClient(
        {"A楼": Path(r"D:\share\internal-only\A楼.xlsx")},
        entry_overrides={"A楼": {"relative_path": relative_path.as_posix()}},
    )
    bitable = FakeBitableClient()

    result = TemperatureHumidityUploadService(
        _runtime_config(["A楼"], shared_root=shared_root),
        internal_client=internal,
        bitable_client=bitable,
    ).run(source_date="2026-07-18")

    assert result["status"] == "success"
    assert result["source_files"]["A楼"] == str(mapped_source)
    assert result["uploaded_count"] == 3
