from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl

from handover_log_module.service.top5_power_report_service import (
    Top5PowerReportBitableUploadService,
    Top5PowerReportService,
)


BUILDINGS = ["A楼", "B楼", "C楼", "D楼", "E楼"]


def _code(building: str) -> str:
    return building[:1]


def _write_monthly_source(path: Path, building: str) -> None:
    code = _code(building)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([None] * 8)
    sheet.append(["变压器", "变压器总容量", "设备容量(KVA)", "安全容量(KVA)", "实际负载KW", "负载率", "安全使用率"])
    for index in range(5):
        room = 120 + index
        side = "A" if index % 2 == 0 else "B"
        sheet.append([None, f"{code}-{room}-{side}变压器容量", "2500", "1317", 300 + index * 10, 0.1, 0.2])

    sheet.append([None] * 4)
    sheet.append(["HVDC容量", "HVDC总容量", "设备容量(KW)", "安全容量(KW)", "实际负载KW", "负载率", "安全使用率"])
    for index in range(5):
        room = 218 + index
        sheet.append([None, f"{code}-{room}-HVDC-{112 + index}", "324", "145", 100 + index * 5, 0.3, 0.7])

    sheet.append([None] * 4)
    sheet.append(["UPS容量", "UPS使用总容量", "设计容量(KVA)", "安全容量(KVA)", "实际负载KW", "负载率", "安全使用率"])
    for index in range(5):
        room = 120 + index
        sheet.append([None, f"{code}-{room}-UPS-{101 + index}_UPS", "300", "210", 20 + index * 2, 0.1, 0.2])

    sheet.append([None] * 4)
    sheet.append(["区域", "设备编号", "上电情况", "设计容量（KW）", "实际负载（KW）", "使用率"])
    cabinet_rows = [
        ("M1包间", f"{code}-202-A01", "闭合", "13", 11.25, 0.86),
        (None, f"{code}-202-A02", "闭合", "13", 12.35, 0.95),
        (None, f"{code}-202-B01", "闭合", "13", 8.10, 0.62),
        (None, f"{code}-202-B02", "闭合", "13", 7.90, 0.61),
        ("M2包间", f"{code}-201-A01", "闭合", "13", 14.60, 1.12),
        (None, f"{code}-201-A02", "闭合", "13", 13.40, 1.03),
        (None, f"{code}-201-C01", "闭合", "13", 9.20, 0.71),
        (None, f"{code}-302-A01", "闭合", "18", 5.10, 0.28),
    ]
    for row in cabinet_rows:
        sheet.append(list(row))
    workbook.save(path)


def _entry(building: str, path: Path) -> dict:
    return {"building": building, "file_path": str(path), "bucket_key": "2026-05-01"}


class Top5PowerReportServiceTest(unittest.TestCase):
    def test_extract_monthly_top5_groups_normalizes_top5(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source = temp_path / "monthly_A.xlsx"
            _write_monthly_source(source, "A楼")

            groups = Top5PowerReportService.extract_monthly_top5_groups(source, building="A楼")

            self.assertEqual(groups.transformers[0].identifier, "A-124-A变压器容量")
            self.assertEqual(groups.transformers[0].power_kw, 340)
            self.assertEqual(groups.hvdcs[0].identifier, "A-222-HVDC-116")
            self.assertEqual(groups.upss[0].identifier, "A-124-UPS-105_UPS")
            self.assertEqual(groups.row_lines[0].identifier, "A-201-A列")
            self.assertAlmostEqual(groups.row_lines[0].power_kw, 28.0)
            self.assertTrue(any(item.identifier == "A-202-A列" for item in groups.row_line_aggregates))

    def test_run_writes_summary_and_source_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            monthly_entries = []
            for building in BUILDINGS:
                monthly_path = temp_path / f"monthly_{_code(building)}.xlsx"
                _write_monthly_source(monthly_path, building)
                monthly_entries.append(_entry(building, monthly_path))

            service = Top5PowerReportService(
                {
                    "handover_log": {
                        "top5_power_report": {
                            "template": {
                                "source_path": str(temp_path / "missing_template.xlsx"),
                                "output_dir": str(temp_path / "out"),
                            }
                        }
                    }
                }
            )

            result = service.run(monthly_entries=monthly_entries, emit_log=lambda _: None)

            output_path = Path(result["output_file"])
            self.assertTrue(output_path.exists())
            workbook = openpyxl.load_workbook(output_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames[0], "汇总信息表")
                self.assertEqual(
                    workbook.sheetnames,
                    ["汇总信息表", "A", "B", "C", "D", "E", "A楼容量", "B楼容量", "C楼容量", "D楼容量", "E楼容量", "Sheet1"],
                )
                summary = workbook["汇总信息表"]
                self.assertEqual(summary["G2"].value, "机列编号")
                self.assertEqual(summary["H2"].value, "机列负载（KW）")
                self.assertEqual(summary["A3"].value, 1)
                self.assertEqual(summary["B3"].value, "A")
                self.assertTrue(summary["G3"].value.endswith("列"))
                self.assertEqual(summary["D3"].number_format, "0.00")
                self.assertEqual(summary["F3"].number_format, "0.00")
                self.assertEqual(summary["H3"].number_format, "0.00")
                self.assertEqual(summary["J3"].number_format, "0.00")
                detail = workbook["A"]
                self.assertEqual(detail["A1"].value, "地点")
                self.assertEqual(detail["J1"].value, "地点")
                self.assertTrue(str(detail["J2"].value).endswith("列"))
                self.assertEqual(detail["K2"].number_format, "0.00")
                self.assertIn("A楼容量", workbook.sheetnames)
                self.assertIn("E楼容量", workbook.sheetnames)
                self.assertIn("Sheet1", workbook.sheetnames)
            finally:
                workbook.close()

    def test_run_fails_when_any_building_monthly_source_missing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            monthly_entries = []
            for building in BUILDINGS[:-1]:
                monthly_path = temp_path / f"monthly_{_code(building)}.xlsx"
                _write_monthly_source(monthly_path, building)
                monthly_entries.append(_entry(building, monthly_path))

            service = Top5PowerReportService(
                {
                    "handover_log": {
                        "top5_power_report": {
                            "template": {
                                "source_path": str(temp_path / "missing_template.xlsx"),
                                "output_dir": str(temp_path / "out"),
                            }
                        }
                    }
                }
            )

            with self.assertRaisesRegex(RuntimeError, "缺少TOP5月报最新源文件: E楼"):
                service.run(monthly_entries=monthly_entries, emit_log=lambda _: None)

class _FakeBitableClient:
    def __init__(self) -> None:
        self.deleted_ids: list[str] = []
        self.created_fields: list[dict] = []
        self.updated: list[tuple[str, dict]] = []

    def list_records(self, table_id: str, field_names: list[str] | None = None) -> list[dict]:
        return [
            {
                "record_id": "old_top5",
                "fields": {"子分类": "高功率TOP5", "年度": "2026", "月份": "04"},
            },
            {
                "record_id": "other_category",
                "fields": {"子分类": "机柜超功耗", "年度": "2026", "月份": "04"},
            },
        ]

    def batch_delete_records(self, table_id: str, record_ids: list[str], batch_size: int = 500, progress_callback=None) -> int:
        self.deleted_ids.extend(record_ids)
        return len(record_ids)

    def upload_attachment(self, file_path: str) -> str:
        return "file_token_1"

    def batch_create_records(self, table_id: str, fields_list: list[dict], batch_size: int = 200, progress_callback=None) -> list[dict]:
        self.created_fields.extend(fields_list)
        return [{"data": {"records": [{"record_id": "new_top5"}]}}]

    def get_record_by_id(self, table_id: str, record_id: str) -> dict:
        return {"record_id": record_id, "fields": {"上传文件": [{"url": "https://example.test/top5.xlsx"}]}}

    def update_record(self, table_id: str, record_id: str, fields: dict) -> dict:
        self.updated.append((record_id, fields))
        return {"code": 0}


class _UploadServiceForTest(Top5PowerReportBitableUploadService):
    def __init__(self, runtime_config: dict[str, Any], client: _FakeBitableClient) -> None:
        super().__init__(runtime_config)
        self.client = client

    def _client(self, cfg: dict[str, Any], emit_log):  # noqa: ANN001
        return self.client


class Top5PowerReportBitableUploadServiceTest(unittest.TestCase):
    def test_upload_report_replaces_same_month_top5_record_and_keeps_other_category(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "top5.xlsx"
            workbook = openpyxl.Workbook()
            workbook.save(path)
            client = _FakeBitableClient()
            service = _UploadServiceForTest({}, client)

            result = service.upload_report(file_path=path, year="2026", month=4, emit_log=lambda _: None)

            self.assertEqual(result["status"], "ok")
            self.assertEqual(result["year"], "2026")
            self.assertEqual(result["month"], "04")
            self.assertEqual(client.deleted_ids, ["old_top5"])
            self.assertEqual(client.created_fields[0]["子分类"], "高功率TOP5")
            self.assertEqual(client.created_fields[0]["年度"], "2026")
            self.assertEqual(client.created_fields[0]["月份"], "04")
            self.assertEqual(client.created_fields[0]["上传文件"], [{"file_token": "file_token_1"}])
            self.assertEqual(client.updated, [("new_top5", {"链接": "https://example.test/top5.xlsx"})])


class Top5PowerReportWorkerHandlerTest(unittest.TestCase):
    def test_worker_uses_top5_monthly_report_family(self) -> None:
        from app.worker import task_handlers

        class FakeBridgeRuntime:
            instances: list["FakeBridgeRuntime"] = []

            def __init__(self, *args, **kwargs) -> None:  # noqa: ANN002, ANN003
                self.monthly_args: dict[str, Any] | None = None
                FakeBridgeRuntime.instances.append(self)

            def get_monthly_by_date_cache_entries(self, *args, **kwargs):  # noqa: ANN002, ANN003
                raise AssertionError("TOP5 worker must use top5_monthly_report_family")

            def get_top5_monthly_by_date_cache_entries(self, *args, **kwargs):  # noqa: ANN002, ANN003
                raise AssertionError("TOP5 worker must not query month-range source-index")

            def refresh_top5_monthly_latest_cache_entries(self, *, buildings, emit_log, cancel_check=None):  # noqa: ANN001
                self.monthly_args = {"buildings": buildings}
                return [
                    {
                        "building": "A楼",
                        "file_path": r"D:\share\monthly_A.xlsx",
                        "metadata": {"upload_date": "2026-05-20"},
                    }
                ]

            def stop(self) -> None:
                return None

        class FakeTop5Service:
            monthly_entries: list[dict[str, Any]] = []

            def __init__(self, config: dict[str, Any]) -> None:
                self.config = config

            def all_buildings(self) -> list[str]:
                return ["A楼"]

            def run(self, *, monthly_entries, emit_log):  # noqa: ANN001
                FakeTop5Service.monthly_entries = list(monthly_entries)
                return {"output_file": r"D:\tmp\top5.xlsx", "file_name": "top5.xlsx"}

        class FakeUploadService:
            @staticmethod
            def _validate_year_month(year, month):  # noqa: ANN001
                return str(year), f"{int(month):02d}"

            def __init__(self, config: dict[str, Any]) -> None:
                self.config = config

            def upload_report(self, *, file_path, year, month, emit_log):  # noqa: ANN001
                return {"status": "ok", "file_path": file_path, "year": str(year), "month": f"{int(month):02d}"}

        with (
            patch.object(task_handlers.shared_bridge_runtime_module, "SharedBridgeRuntimeService", FakeBridgeRuntime),
            patch.object(task_handlers, "Top5PowerReportService", FakeTop5Service),
            patch.object(task_handlers, "Top5PowerReportBitableUploadService", FakeUploadService),
        ):
            result = task_handlers.handle_top5_power_report(
                {},
                {"year": "2026", "month": 5},
                emit_log=lambda _message: None,
            )

        self.assertEqual(result["bitable_upload"]["status"], "ok")
        self.assertEqual(FakeBridgeRuntime.instances[0].monthly_args["buildings"], ["A楼"])
        self.assertEqual(FakeTop5Service.monthly_entries[0]["building"], "A楼")


class Top5PowerReportBridgeRuntimeTest(unittest.TestCase):
    def test_latest_selection_prefers_newer_failed_over_old_ready(self) -> None:
        from app.modules.shared_bridge.service import shared_bridge_runtime_service as bridge_module

        runtime = bridge_module.SharedBridgeRuntimeService.__new__(bridge_module.SharedBridgeRuntimeService)
        runtime._http_source_cache_buildings = lambda buildings=None: ["A楼"]

        def fake_entries(*, source_family, buildings, bucket_key, status="ready", limit_per_building):  # noqa: ANN001
            self.assertEqual(status, "all")
            return [
                {
                    "building": "A楼",
                    "status": "ready",
                    "bucket_key": "2026-06-04 08",
                    "file_path": r"D:\share\old.xlsx",
                    "relative_path": "old.xlsx",
                    "downloaded_at": "2026-06-04 08:10:00",
                    "updated_at": "2026-06-04 08:10:00",
                },
                {
                    "building": "A楼",
                    "status": "failed",
                    "bucket_key": "2026-06-04 09",
                    "metadata": {"error": "下载失败"},
                    "downloaded_at": "2026-06-04 09:10:00",
                    "updated_at": "2026-06-04 09:10:00",
                },
            ]

        runtime._http_source_index_entries = fake_entries
        selection = runtime._http_latest_source_cache_selection(
            source_family=bridge_module.FAMILY_TOP5_MONTHLY_REPORT,
            buildings=["A楼"],
            max_selection_age_hours=999999,
        )

        self.assertFalse(selection["can_proceed"])
        self.assertEqual(selection["failed_buildings"], ["A楼"])
        self.assertEqual(selection["selected_entries"], [])
        self.assertEqual(selection["buildings"][0]["status"], "failed")
        self.assertEqual(selection["buildings"][0]["last_error"], "下载失败")

    def test_refresh_latest_waits_for_entries_downloaded_after_request(self) -> None:
        from app.modules.shared_bridge.service import shared_bridge_runtime_service as bridge_module

        runtime = bridge_module.SharedBridgeRuntimeService.__new__(bridge_module.SharedBridgeRuntimeService)
        runtime._internal_bridge_http_client = object()
        runtime._http_bridge_should_try = lambda: True
        runtime.request_latest_source_cache_refresh = lambda *, source_family, buildings: {
            "ok": True,
            "accepted_count": len(buildings),
            "results": [],
        }
        calls = {"count": 0}

        def fake_entries(*, source_family, buildings, bucket_key, status="ready", limit_per_building):  # noqa: ANN001
            calls["count"] += 1
            if calls["count"] == 1:
                return [
                    {
                        "building": "A楼",
                        "status": "ready",
                        "file_path": r"D:\share\old.xlsx",
                        "downloaded_at": "2000-01-01 00:00:00",
                        "updated_at": "2000-01-01 00:00:00",
                    }
                ]
            return [
                {
                    "building": "A楼",
                    "status": "ready",
                    "file_path": r"D:\share\new.xlsx",
                    "downloaded_at": "2099-01-01 00:00:00",
                    "updated_at": "2099-01-01 00:00:00",
                }
            ]

        runtime._http_source_index_entries = fake_entries
        with patch.object(bridge_module.time, "sleep", lambda _seconds: None):
            entries = runtime.refresh_top5_monthly_latest_cache_entries(
                buildings=["A楼"],
                timeout_sec=30,
                poll_interval_sec=1,
            )

        self.assertEqual(calls["count"], 2)
        self.assertEqual(entries[0]["file_path"], r"D:\share\new.xlsx")


if __name__ == "__main__":
    unittest.main()
