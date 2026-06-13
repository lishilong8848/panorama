from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from handover_log_module.service.handover_110_main_transformer_bitable_sync_service import (
    Handover110MainTransformerBitableSyncService,
)


def _write_sample_110_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EA118中天变夜班"
    for index, column in enumerate((9, 10, 11, 12), start=1):
        ws.cell(row=29, column=column, value=f"#{index}主变10{index}\n（50MW）")
    values = {
        31: [817.78, 363.17, 376.94, 738.64],
        34: [14.77, 6.45, 6.67, 13.27],
        35: [0.2954, 0.1303, 0.1347, 0.2654],
        39: [53, 46, 46, 55],
        41: [3, 3, 4, 4],
    }
    for row_idx, row_values in values.items():
        for column, value in zip((9, 10, 11, 12), row_values):
            ws.cell(row=row_idx, column=column, value=value)

    labels = ["主变断路器（A/B/C）", "主变避雷器（A/B/C）", "避雷器过渡仓（A/B/C）", "主变正母侧", "主变副母侧", "主变分支"]
    pressure_rows = [
        [0.60, 0.62, 0.64, 0.61],
        [0.50, 0.50, 0.50, 0.50],
        [0.52, 0.52, 0.52, 0.52],
        [0.50, 0.52, 0.52, 0.50],
        [0.50, 0.52, 0.52, 0.50],
        [0.50, 0.52, 0.52, 0.50],
    ]
    for row_idx, label, row_values in zip(range(132, 138), labels, pressure_rows):
        ws.cell(row=row_idx, column=6, value=label)
        for column, value in zip((7, 8, 9, 10), row_values):
            ws.cell(row=row_idx, column=column, value=value)
    wb.save(path)
    wb.close()


def _write_shifted_day_110_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EA118中天变白班"
    for index, column in enumerate((9, 10, 11, 12), start=1):
        ws.cell(row=30, column=column, value=f"#{index}主变10{index}\n（50MW）")
    labels = {
        31: "运行情况",
        32: "输出电流（A）",
        35: "本班最大负载（MW）",
        36: "负载率",
        40: "油温（℃）",
        42: "档位",
    }
    for row_idx, label in labels.items():
        ws.cell(row=row_idx, column=8, value=label)
    values = {
        32: [858.82, 365.22, 370.2, 738.64],
        35: [15.5, 6.45, 6.85, 13.27],
        36: [0.31, 0.13030303030303, 0.138383838383838, 0.2654],
        40: [51, 45, 45, 53],
        42: [3, 3, 4, 4],
    }
    for row_idx, row_values in values.items():
        for column, value in zip((9, 10, 11, 12), row_values):
            ws.cell(row=row_idx, column=column, value=value)

    for row_idx in range(132, 138):
        ws.cell(row=row_idx, column=6, value="主变避雷器（A/B/C）")
        for column in (7, 8, 9, 10):
            ws.cell(row=row_idx, column=column, value=0.50)
    wb.save(path)
    wb.close()


def test_parse_workbook_builds_four_transformer_rows(tmp_path):
    path = tmp_path / "110.xlsx"
    _write_sample_110_workbook(path)

    rows = Handover110MainTransformerBitableSyncService.parse_workbook(path)

    assert [row["transformer_name"] for row in rows] == ["1号主变", "2号主变", "3号主变", "4号主变"]
    assert [row["line_name"] for row in rows] == ["阿开线", "阿开线", "阿家线", "阿家线"]
    assert rows[0]["load_kw"] == 14770
    assert rows[3]["current_a"] == 738.64
    assert rows[0]["load_rate"] == "29.54%"
    assert all(row["gis_status"] == "正常" for row in rows)


def test_parse_workbook_uses_labels_for_shifted_day_rows(tmp_path):
    path = tmp_path / "110_day.xlsx"
    _write_shifted_day_110_workbook(path)

    rows = Handover110MainTransformerBitableSyncService.parse_workbook(path)

    assert rows[0]["current_a"] == 858.82
    assert rows[0]["load_kw"] == 15500
    assert rows[0]["load_rate"] == "31.00%"
    assert rows[0]["oil_temp"] == "51"
    assert rows[0]["tap_position"] == "3"


def test_payload_and_existing_record_match_use_selected_duty_context():
    service = Handover110MainTransformerBitableSyncService({})
    rows = [
        {
            "transformer_name": "1号主变",
            "line_name": "阿开线",
            "oil_temp": "53",
            "tap_position": "3",
            "load_kw": 14770,
            "current_a": 817.78,
            "load_rate": "29.54%",
            "gis_status": "正常",
        }
    ]

    payloads = service._payload_fields(duty_date="2026-05-11", duty_shift="night", rows=rows)

    assert payloads[0]["班次"] == "夜班"
    assert payloads[0]["主变名称"] == "1号主变"
    assert payloads[0]["所属线路"] == "阿开线"
    assert payloads[0]["负载（KW）"] == 14770

    class FakeClient:
        def list_records(self, **_kwargs):
            return [
                {
                    "record_id": "rec_keep",
                    "fields": {
                        "日期": payloads[0]["日期"],
                        "班次": "夜班",
                        "主变名称": "1号主变",
                    },
                },
                {
                    "record_id": "rec_other_shift",
                    "fields": {
                        "日期": payloads[0]["日期"],
                        "班次": "白班",
                        "主变名称": "1号主变",
                    },
                },
            ]

    matched = service._existing_record_ids(
        client=FakeClient(),
        table_id="tbl8Ni54taYGeWAa",
        duty_date="2026-05-11",
        duty_shift="night",
        cfg=service._normalize_cfg(),
    )

    assert matched == ["rec_keep"]


def test_sync_from_upload_state_replaces_same_duty_records_without_network(tmp_path):
    path = tmp_path / "110.xlsx"
    _write_sample_110_workbook(path)

    class FakeClient:
        def __init__(self):
            self.deleted = []
            self.created = []
            self.calls = []

        def list_fields(self, table_id):
            assert table_id == "tbl8Ni54taYGeWAa"
            return [{"field_name": name} for name in Handover110MainTransformerBitableSyncService.REQUIRED_FIELDS]

        def list_records(self, **_kwargs):
            date_ms = Handover110MainTransformerBitableSyncService._date_ms("2026-05-11")
            return [
                {"record_id": "rec_old_1", "fields": {"日期": date_ms, "班次": "夜班", "主变名称": "1号主变"}},
                {"record_id": "rec_old_2", "fields": {"日期": date_ms, "班次": "夜班", "主变名称": "2号主变"}},
            ]

        def batch_delete_records(self, *, table_id, record_ids, batch_size):
            assert table_id == "tbl8Ni54taYGeWAa"
            self.calls.append("delete")
            self.deleted.extend(record_ids)
            return len(record_ids)

        def batch_create_records(self, *, table_id, fields_list, batch_size):
            assert table_id == "tbl8Ni54taYGeWAa"
            self.calls.append("create")
            self.created.extend(fields_list)
            return [{"data": {"records": [{"record_id": f"rec_new_{index}"} for index, _ in enumerate(fields_list, 1)]}}]

    fake_client = FakeClient()

    class Service(Handover110MainTransformerBitableSyncService):
        def _new_client(self, cfg):
            return fake_client

    service = Service({})
    result = service.sync_from_upload_state(
        duty_date="2026-05-11",
        duty_shift="night",
        upload_state={"status": "success", "stored_path": str(path)},
        emit_log=lambda _message: None,
    )

    assert result["status"] == "success"
    assert fake_client.calls == ["create", "delete"]
    assert fake_client.deleted == ["rec_old_1", "rec_old_2"]
    assert len(fake_client.created) == 4
    assert fake_client.created[0]["主变名称"] == "1号主变"
    assert fake_client.created[0]["负载（KW）"] == 14770
    assert fake_client.created[2]["所属线路"] == "阿家线"


def test_sync_from_upload_state_prefers_saved_transformer_rows(tmp_path):
    missing_source = tmp_path / "missing.xlsx"

    class FakeClient:
        def __init__(self):
            self.created = []

        def list_fields(self, table_id):
            assert table_id == "tbl8Ni54taYGeWAa"
            return [{"field_name": name} for name in Handover110MainTransformerBitableSyncService.REQUIRED_FIELDS]

        def list_records(self, **_kwargs):
            return []

        def batch_create_records(self, *, table_id, fields_list, batch_size):
            assert table_id == "tbl8Ni54taYGeWAa"
            self.created.extend(fields_list)
            return [{"data": {"records": [{"record_id": f"rec_new_{index}"} for index, _ in enumerate(fields_list, 1)]}}]

        def batch_delete_records(self, *, table_id, record_ids, batch_size):
            return len(record_ids)

    fake_client = FakeClient()

    class Service(Handover110MainTransformerBitableSyncService):
        def _new_client(self, cfg):
            return fake_client

    saved_rows = [
        {
            "transformer_name": "1号主变",
            "line_name": "阿开线",
            "oil_temp": "51",
            "tap_position": "3",
            "load_kw": 15500,
            "current_a": 858.82,
            "load_rate": "31.00%",
            "gis_status": "正常",
        },
        {
            "transformer_name": "2号主变",
            "line_name": "阿开线",
            "oil_temp": "45",
            "tap_position": "3",
            "load_kw": 6450,
            "current_a": 365.22,
            "load_rate": "13.03%",
            "gis_status": "正常",
        },
        {
            "transformer_name": "3号主变",
            "line_name": "阿家线",
            "oil_temp": "45",
            "tap_position": "4",
            "load_kw": 6850,
            "current_a": 370.2,
            "load_rate": "13.84%",
            "gis_status": "正常",
        },
        {
            "transformer_name": "4号主变",
            "line_name": "阿家线",
            "oil_temp": "53",
            "tap_position": "4",
            "load_kw": 13270,
            "current_a": 738.64,
            "load_rate": "26.54%",
            "gis_status": "正常",
        },
    ]

    service = Service({})
    result = service.sync_from_upload_state(
        duty_date="2026-05-11",
        duty_shift="day",
        upload_state={
            "status": "success",
            "stored_path": str(missing_source),
            "parsed_main_transformer_rows": saved_rows,
        },
        emit_log=lambda _message: None,
    )

    assert result["status"] == "success"
    assert len(fake_client.created) == 4
    assert fake_client.created[0]["负载（KW）"] == 15500
    assert fake_client.created[0]["电流（A）"] == 858.82
    assert fake_client.created[0]["班次"] == "白班"
