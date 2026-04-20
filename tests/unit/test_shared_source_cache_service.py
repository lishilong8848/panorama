from __future__ import annotations

import concurrent.futures
import json
import os
import shutil
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

import app.modules.shared_bridge.service.shared_source_cache_service as cache_module
from app.modules.shared_bridge.service.shared_bridge_store import SharedBridgeStore
from app.modules.shared_bridge.service.shared_source_cache_service import (
    ALARM_EVENT_BITABLE_TARGET_FIELDS,
    FAMILY_ALARM_EVENT,
    FAMILY_HANDOVER_CAPACITY_REPORT,
    FAMILY_HANDOVER_LOG,
    FAMILY_MONTHLY_REPORT,
    SharedSourceCacheService,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEMP_ROOT = PROJECT_ROOT / '.tmp_runtime_tests' / 'shared_source_cache_service'


@pytest.fixture
def work_dir() -> Path:
    root = TEMP_ROOT / uuid.uuid4().hex
    root.mkdir(parents=True, exist_ok=True)
    try:
        yield root
    finally:
        shutil.rmtree(root, ignore_errors=True)


def _build_runtime_config(
    *,
    role_mode: str,
    shared_root: Path | None = None,
    legacy_root: str = '',
    internal_root: str = '',
    external_root: str = '',
) -> dict:
    shared_bridge = {
        'enabled': True,
        'root_dir': legacy_root or (str(shared_root) if shared_root is not None else ''),
    }
    if internal_root:
        shared_bridge['internal_root_dir'] = internal_root
    if external_root:
        shared_bridge['external_root_dir'] = external_root
    return {
        'deployment': {'role_mode': role_mode},
        'shared_bridge': shared_bridge,
        'internal_source_cache': {'enabled': True},
        'feishu': {
            'app_id': 'test-app-id',
            'app_secret': 'test-app-secret',
            'timeout': 30,
            'request_retry_count': 1,
            'request_retry_interval_sec': 0,
        },
        'alarm_export': {
            'feishu': {
                'app_token': 'test-app-token',
                'table_id': 'test-table-id',
                'page_size': 200,
                'delete_batch_size': 100,
                'create_batch_size': 50,
            },
            'shared_source_upload': {
                'replace_existing_on_full': True,
            },
        },
    }


def _write_alarm_json(
    path: Path,
    *,
    rows: list[dict[str, object]] | None = None,
    query_start: str = '2026-02-01 00:00:00',
    query_end: str = '2026-04-01 12:05:00',
    building: str = 'A楼',
    bucket_kind: str = 'latest',
    bucket_key: str = '2026-04-01 08',
) -> None:
    payload = {
        'schema_version': 1,
        'source_family': FAMILY_ALARM_EVENT,
        'building': building,
        'bucket_kind': bucket_kind,
        'bucket_key': bucket_key,
        'generated_at': '2026-04-01 12:05:00',
        'query_start': query_start,
        'query_end': query_end,
        'row_count': len(rows or []),
        'count_summary': {},
        'rows': rows or [],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def _write_minimal_handover_workbook(
    path: Path,
    *,
    b_text: str = 'A-245-TEST-001',
    c_text: str = '测试系统',
    d_name: str = '测试点位',
    e_value: object = 1,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = '交接班日志'
    worksheet['B4'] = b_text
    worksheet['C4'] = c_text
    worksheet['D4'] = d_name
    worksheet['E4'] = e_value
    workbook.save(path)


def _add_alarm_cache_entry(
    *,
    store: SharedBridgeStore,
    shared_root: Path,
    building: str,
    bucket_kind: str,
    bucket_key: str,
    downloaded_at: str,
    status: str = 'ready',
    rows: list[dict[str, object]] | None = None,
    suffix: str = 'json',
) -> dict:
    folder_name = f"{building}-{bucket_kind}-{bucket_key}".replace(":", "-").replace(" ", "_")
    target_path = shared_root / '告警信息源文件' / 'test' / f'{folder_name}.{suffix}'
    if suffix.lower() == 'json':
        _write_alarm_json(
            target_path,
            building=building,
            bucket_kind=bucket_kind,
            bucket_key=bucket_key,
            rows=rows or [],
        )
    else:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_bytes(b'test')
    relative_path = str(target_path.relative_to(shared_root)).replace('\\', '/')
    store.upsert_source_cache_entry(
        source_family=FAMILY_ALARM_EVENT,
        building=building,
        bucket_kind=bucket_kind,
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at=downloaded_at,
        relative_path=relative_path,
        status=status,
        file_hash=f'hash-{building}-{bucket_kind}-{bucket_key}',
        size_bytes=int(target_path.stat().st_size) if target_path.exists() else 0,
        metadata={},
    )
    return {
        'file_path': str(target_path),
        'relative_path': relative_path,
    }


def test_health_snapshot_contains_building_level_current_hour_statuses(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-03-29 10'
    service._current_hour_bucket = bucket_key

    ready_file = shared_root / '交接班日志源文件' / '202603' / '20260329--10' / '20260329--10--交接班日志源文件--A楼.xlsx'
    _write_minimal_handover_workbook(ready_file)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-03-29 10:05:00',
        relative_path=str(ready_file.relative_to(shared_root)).replace('\\', '/'),
        status='ready',
        file_hash='hash-a',
        size_bytes=7,
    )
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='B楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-03-29 10:06:00',
        relative_path='source_cache/_failed/handover/latest/2026-03-29 10/B楼.failed',
        status='failed',
        file_hash='',
        size_bytes=0,
        metadata={'error': '下载失败'},
    )

    snapshot = service.get_health_snapshot()
    buildings = {item['building']: item for item in snapshot[FAMILY_HANDOVER_LOG]['buildings']}

    assert buildings['A楼']['status'] == 'ready'
    assert buildings['A楼']['ready'] is True
    assert buildings['A楼']['downloaded_at'] == '2026-03-29 10:05:00'
    assert buildings['A楼']['resolved_file_path'] == str(ready_file)
    assert buildings['B楼']['status'] == 'failed'
    assert buildings['B楼']['last_error'] == '下载失败'
    assert buildings['C楼']['status'] == 'waiting'
    assert snapshot[FAMILY_HANDOVER_LOG]['ready_count'] == 1
    assert snapshot[FAMILY_HANDOVER_LOG]['failed_buildings'] == ['B楼']


def test_external_source_cache_fast_overview_reads_index_without_file_validation(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-04-20 14'
    relative_path = '交接班日志源文件/202604/20260420--14/20260420--14--交接班日志源文件--A楼.xlsx'
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-04-20 14:05:00',
        relative_path=relative_path,
        status='ready',
        file_hash='hash-a',
        size_bytes=123,
    )

    overview = service.get_external_source_cache_overview_fast()
    families = {item['key']: item for item in overview['families']}
    handover_rows = {item['building']: item for item in families[FAMILY_HANDOVER_LOG]['buildings']}

    assert handover_rows['A楼']['status_key'] == 'ready'
    assert handover_rows['A楼']['relative_path'] == relative_path
    assert handover_rows['A楼']['resolved_file_path'].endswith('20260420--14--交接班日志源文件--A楼.xlsx')


def test_health_snapshot_does_not_mark_missing_ready_file_as_ready(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-03-29 11'
    service._current_hour_bucket = bucket_key

    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-03-29 11:03:00',
        relative_path='全景平台月报源文件/202603/20260329--11/缺失文件.xlsx',
        status='ready',
        file_hash='hash-missing',
        size_bytes=18,
    )

    snapshot = service.get_health_snapshot()
    buildings = {item['building']: item for item in snapshot[FAMILY_MONTHLY_REPORT]['buildings']}
    rows = store.list_source_cache_entries(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        limit=1,
    )

    assert buildings['A楼']['status'] == 'failed'
    assert buildings['A楼']['ready'] is False
    assert buildings['A楼']['last_error'] == '共享文件缺失或不可访问'
    assert snapshot[FAMILY_MONTHLY_REPORT]['ready_count'] == 0
    assert rows[0]['status'] == 'failed'
    assert rows[0]['metadata']['error'] == '共享文件缺失或不可访问'


def test_store_entry_marks_entry_refreshing_before_replacing_ready_file(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    source_path = work_dir / 'source.xlsx'
    source_path.write_bytes(b'test')
    target_path = service._source_target_path(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-10 09',
        duty_date='2026-04-10',
        duty_shift='day',
        source_path=source_path,
    )
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(b'old-ready')
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-10 09',
        duty_date='2026-04-10',
        duty_shift='day',
        downloaded_at='2026-04-10 09:00:00',
        relative_path=str(target_path.relative_to(shared_root)).replace('\\', '/'),
        status='ready',
        file_hash='old-hash',
        size_bytes=9,
        metadata={'naming_version': 2},
    )

    observed_statuses: list[str] = []

    def _fake_cache_file(*, source_path: Path, target_path: Path) -> dict[str, object]:
        rows = store.list_source_cache_entries(
            source_family=FAMILY_HANDOVER_LOG,
            building='A楼',
            bucket_kind='latest',
            bucket_key='2026-04-10 09',
            limit=5,
        )
        observed_statuses.extend(
            str(item.get('status', '') or '').strip().lower()
            for item in rows
            if isinstance(item, dict)
        )
        target_path.write_bytes(b'new-ready')
        return {
            'file_hash': 'new-hash',
            'size_bytes': int(target_path.stat().st_size),
            'target_path': target_path,
            'relative_path': target_path.relative_to(shared_root).as_posix(),
        }

    monkeypatch.setattr(service, '_cache_file', _fake_cache_file)

    result = service._store_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-10 09',
        duty_date='2026-04-10',
        duty_shift='day',
        source_path=source_path,
        status='ready',
        metadata={'family': FAMILY_HANDOVER_LOG, 'building': 'A楼'},
    )

    assert 'refreshing' in observed_statuses
    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-10 09',
        limit=1,
    )
    assert rows[0]['status'] == 'ready'
    assert rows[0]['file_hash'] == 'new-hash'
    assert result['file_path'] == str(target_path)


def test_get_health_snapshot_repairs_stale_refreshing_entry(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-04-10 09'
    service._current_hour_bucket = bucket_key
    stale_refreshing_at = (datetime.now() - timedelta(minutes=20)).strftime('%Y-%m-%d %H:%M:%S')
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-04-10 09:00:00',
        relative_path='交接班日志源文件/202604/20260410--09/A楼.xlsx',
        status='refreshing',
        file_hash='',
        size_bytes=0,
        metadata={'refreshing_at': stale_refreshing_at},
    )

    snapshot = service.get_health_snapshot()
    buildings = {item['building']: item for item in snapshot[FAMILY_HANDOVER_LOG]['buildings']}
    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        limit=1,
    )

    assert buildings['A楼']['status'] == 'failed'
    assert buildings['A楼']['last_error'] == '共享文件刷新超时'
    assert rows[0]['status'] == 'failed'
    assert rows[0]['metadata']['error'] == '共享文件刷新超时'


def test_internal_light_health_snapshot_skips_latest_selection(work_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    service._current_hour_bucket = '2026-04-01 10'

    def _explode(*_args, **_kwargs):
        raise AssertionError('internal_light snapshot should not compute latest selection')

    monkeypatch.setattr(service, 'get_latest_ready_selection', _explode)

    snapshot = service.get_health_snapshot(mode='internal_light')

    assert snapshot[FAMILY_HANDOVER_LOG]['latest_selection'] == {}
    assert snapshot[FAMILY_MONTHLY_REPORT]['latest_selection'] == {}
    assert snapshot[FAMILY_ALARM_EVENT]['latest_selection'] == {}


def test_internal_light_health_snapshot_uses_cached_building_rows_only(work_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-04-01 10'
    service._current_hour_bucket = bucket_key
    with service._lock:
        service._ensure_light_family_cache_unlocked(
            source_family=FAMILY_HANDOVER_LOG,
            bucket_key=bucket_key,
            buildings=['A楼', 'B楼', 'C楼', 'D楼', 'E楼'],
        )
        service._set_light_building_status_unlocked(
            source_family=FAMILY_HANDOVER_LOG,
            building='A楼',
            bucket_key=bucket_key,
            payload={
                'status': 'ready',
                'ready': True,
                'downloaded_at': '2026-04-01 10:05:00',
                'relative_path': 'handover/A.xlsx',
                'resolved_file_path': str(shared_root / 'handover/A.xlsx'),
            },
        )

    def _explode(*_args, **_kwargs):
        raise AssertionError('internal_light snapshot should not use heavy family health builder')

    monkeypatch.setattr(service, '_build_family_health_snapshot', _explode)

    snapshot = service.get_health_snapshot(mode='internal_light')
    buildings = {item['building']: item for item in snapshot[FAMILY_HANDOVER_LOG]['buildings']}

    assert buildings['A楼']['status'] == 'ready'
    assert buildings['A楼']['ready'] is True
    assert buildings['B楼']['status'] == 'waiting'


def test_external_full_health_snapshot_uses_cached_copy_until_marked_dirty(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    logs: list[str] = []
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda text, *_args, **_kwargs: logs.append(str(text)),
    )
    call_counter = {'count': 0}

    def _fake_build_family_health_snapshot(*, source_family: str, current_bucket: str, include_latest_selection: bool):  # noqa: ANN001
        call_counter['count'] += 1
        return {
            'ready_count': 0,
            'failed_buildings': [],
            'blocked_buildings': [],
            'last_success_at': '',
            'current_bucket': current_bucket,
            'buildings': [],
            'latest_selection': {} if not include_latest_selection else {'source_family': source_family},
        }

    monkeypatch.setattr(service, '_build_family_health_snapshot', _fake_build_family_health_snapshot)

    first = service.get_health_snapshot(mode='external_full')
    second = service.get_health_snapshot(mode='external_full')

    assert first[FAMILY_HANDOVER_LOG]['latest_selection'] == {'source_family': FAMILY_HANDOVER_LOG}
    assert second[FAMILY_HANDOVER_LOG]['latest_selection'] == {'source_family': FAMILY_HANDOVER_LOG}
    assert call_counter['count'] == 3

    service._mark_external_full_snapshot_dirty()
    service.get_health_snapshot(mode='external_full')
    assert call_counter['count'] == 6


def test_alarm_event_recent_bucket_uses_current_hour_bucket(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert service.current_alarm_bucket(cache_module.datetime(2026, 4, 1, 7, 30, 0)) == '2026-04-01 07'
    assert service.current_alarm_bucket(cache_module.datetime(2026, 4, 1, 8, 15, 0)) == '2026-04-01 08'
    assert service.current_alarm_bucket(cache_module.datetime(2026, 4, 1, 16, 20, 0)) == '2026-04-01 16'


def test_fill_alarm_event_latest_exports_workbook_and_indexes_entry(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()

    class _FakePool:
        @staticmethod
        def submit_building_alarm_job(building: str, runner):  # noqa: ANN001, ARG004
            future = concurrent.futures.Future()
            future.set_result(
                {
                    'query_start': '2026-02-01 00:00:00',
                    'query_end': '2026-04-01 08:05:00',
                    'rows': [
                        {
                            'level': '次要',
                            'content': '风机状态: 告警',
                            'position': 'E楼/三层',
                            'object': 'E-311-CRAH-10',
                            'event_time': '2026-04-01 08:01:00',
                            'accept_time': '2026-04-01 08:02:00',
                            'is_accept': '已处理',
                            'accept_by': '系统管理员',
                            'accept_content': '检修导致',
                            'recover_time': '--',
                            'is_recover': '未恢复',
                            'event_snapshot': '关闭',
                            'event_type': '不正常值',
                            'confirm_type': '真实告警',
                            'event_suggest': '',
                            'confirm_time': '2026-04-01 08:03:00',
                            'confirm_by': '系统管理员',
                            'confirm_description': '检修导致',
                            'real_value': '0',
                            'alarm_threshold': '0',
                        }
                    ],
                }
            )
            return future

    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        download_browser_pool=_FakePool(),
        emit_log=lambda *_args, **_kwargs: None,
    )

    entry = service.fill_alarm_event_latest(
        building='A楼',
        bucket_key='2026-04-01 08',
        emit_log=lambda *_args, **_kwargs: None,
    )

    output_path = Path(entry['file_path'])
    assert output_path.exists()
    payload = json.loads(output_path.read_text(encoding='utf-8'))
    assert payload['schema_version'] == 1
    assert payload['building'] == 'A楼'
    assert payload['bucket_key'] == '2026-04-01 08'
    assert payload['row_count'] == 1
    assert payload['rows'][0]['level'] == '次要'
    rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='ready',
        limit=1,
    )
    assert len(rows) == 1
    assert rows[0]['metadata']['row_count'] == 1
    assert rows[0]['metadata']['query_start'] == '2026-02-01 00:00:00'


def test_current_hour_refresh_also_refreshes_recent_alarm_bucket(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    calls: list[tuple[str, str, bool]] = []

    def _capture_refresh(*, source_family: str, bucket_key: str, fill_func, force_retry_failed: bool = False):  # noqa: ANN001, ARG001
        calls.append((source_family, bucket_key, force_retry_failed))
        service._family_status.setdefault(source_family, {})
        service._family_status[source_family]['failed_buildings'] = []
        service._family_status[source_family]['blocked_buildings'] = []

    monkeypatch.setattr(service, '_refresh_family_bucket', _capture_refresh)
    monkeypatch.setattr(service, 'current_hour_bucket', lambda when=None: '2026-04-01 10')
    monkeypatch.setattr(service, 'current_alarm_bucket', lambda when=None: '2026-04-01 08')

    service._run_current_hour_refresh_impl()

    assert calls == [
        (FAMILY_HANDOVER_LOG, '2026-04-01 10', True),
        (FAMILY_HANDOVER_CAPACITY_REPORT, '2026-04-01 10', True),
        (FAMILY_MONTHLY_REPORT, '2026-04-01 10', True),
        (FAMILY_ALARM_EVENT, '2026-04-01 08', True),
    ]


def test_alarm_temp_root_uses_filesystem_safe_manual_bucket_segment(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    path = service._alarm_temp_root(
        bucket_key='2026-04-01 17:26:16',
        building='A楼',
        bucket_kind='manual',
    )
    relative_path = path.relative_to(service._tmp_root)

    assert '20260401--172616--manual' in str(path)
    assert ':' not in str(relative_path)


def test_run_manual_alarm_refresh_impl_returns_parallel_summary(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()

    class _FakePool:
        @staticmethod
        def get_building_pause_info(building: str) -> dict[str, object]:
            if building == 'C楼':
                return {
                    'suspended': True,
                    'suspend_reason': '页面异常',
                }
            return {'suspended': False}

    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        download_browser_pool=_FakePool(),
        emit_log=lambda *_args, **_kwargs: None,
    )
    service.get_enabled_buildings = lambda: ['A楼', 'B楼', 'C楼']  # type: ignore[method-assign]

    def _fake_fill(*, building: str, bucket_key: str, emit_log):  # noqa: ANN001, ARG001
        if building == 'B楼':
            raise RuntimeError('B楼 页面异常')
        return {
            'building': building,
            'bucket_key': bucket_key,
            'downloaded_at': '2026-04-01 17:30:00',
            'relative_path': f'告警信息源文件/{building}.json',
            'file_path': str(shared_root / '告警信息源文件' / f'{building}.json'),
        }

    monkeypatch.setattr(service, 'fill_alarm_event_manual', _fake_fill)
    monkeypatch.setattr(service, '_alarm_manual_bucket', lambda when=None: '2026-04-01 17')

    result = service._run_manual_alarm_refresh_impl()

    assert result['running_buildings'] == ['A楼', 'B楼']
    assert sorted(result['completed_buildings']) == ['A楼']
    assert result['failed_buildings'] == ['B楼']
    assert result['blocked_buildings'] == ['C楼']


def test_run_current_hour_refresh_impl_tracks_running_and_completed_buildings(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    def _fake_refresh_family_bucket(*, source_family: str, bucket_key: str, fill_func, force_retry_failed: bool = False):  # noqa: ANN001, ARG001
        if source_family == FAMILY_HANDOVER_LOG:
            return {
                'ready_count': 2,
                'failed_buildings': ['B楼'],
                'blocked_buildings': [],
                'running_buildings': ['A楼', 'B楼'],
                'completed_buildings': ['A楼'],
                'current_bucket': bucket_key,
            }
        if source_family == FAMILY_MONTHLY_REPORT:
            return {
                'ready_count': 1,
                'failed_buildings': [],
                'blocked_buildings': ['D楼'],
                'running_buildings': ['C楼', 'D楼'],
                'completed_buildings': ['C楼'],
                'current_bucket': bucket_key,
            }
        return {
            'ready_count': 1,
            'failed_buildings': [],
            'blocked_buildings': [],
            'running_buildings': ['E楼'],
            'completed_buildings': ['E楼'],
            'current_bucket': bucket_key,
        }

    monkeypatch.setattr(service, '_refresh_family_bucket', _fake_refresh_family_bucket)
    monkeypatch.setattr(service, 'current_hour_bucket', lambda when=None: '2026-04-01 10')
    monkeypatch.setattr(service, 'current_alarm_bucket', lambda when=None: '2026-04-01 08')

    service._run_current_hour_refresh_impl()

    assert service._current_hour_refresh['running_buildings'] == [
        'A楼/handover_log_family',
        'B楼/handover_log_family',
        'E楼/handover_capacity_report_family',
        'C楼/monthly_report_family',
        'D楼/monthly_report_family',
        'E楼/alarm_event_family',
    ]
    assert service._current_hour_refresh['completed_buildings'] == [
        'A楼/handover_log_family',
        'E楼/handover_capacity_report_family',
        'C楼/monthly_report_family',
        'E楼/alarm_event_family',
    ]
    assert service._current_hour_refresh['failed_buildings'] == ['B楼/handover_log_family']
    assert service._current_hour_refresh['blocked_buildings'] == ['D楼/monthly_report_family']


def test_refresh_family_bucket_calls_fill_with_keyword_arguments(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    calls: list[tuple[str, str, bool]] = []

    def _fake_fill(*, building: str, bucket_key: str, emit_log):  # noqa: ANN001
        calls.append((building, bucket_key, callable(emit_log)))

    service.get_enabled_buildings = lambda: ['A楼']  # type: ignore[method-assign]
    service._refresh_family_bucket(
        source_family=FAMILY_HANDOVER_LOG,
        bucket_key='2026-03-29 23',
        fill_func=_fake_fill,
    )

    assert calls == [('A楼', '2026-03-29 23', True)]


def test_refresh_family_bucket_records_failed_entry_metadata(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    def _fake_fill(*, building: str, bucket_key: str, emit_log):  # noqa: ANN001, ARG001
        raise RuntimeError(f'{building} 下载异常')

    service.get_enabled_buildings = lambda: ['A楼']  # type: ignore[method-assign]
    service._refresh_family_bucket(
        source_family=FAMILY_HANDOVER_LOG,
        bucket_key='2026-03-29 23',
        fill_func=_fake_fill,
    )

    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-03-29 23',
        status='failed',
        limit=1,
    )

    assert len(rows) == 1
    assert rows[0]['relative_path'].startswith('source_cache/_failed/')
    assert rows[0]['metadata']['error'] == 'A楼 下载异常'


def test_get_monthly_by_date_entries_ignores_missing_indexed_files(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='date',
        bucket_key='2026-03-29',
        duty_date='2026-03-29',
        duty_shift='',
        downloaded_at='2026-03-29 23:59:00',
        relative_path='全景平台月报源文件/202603/20260329--月报/20260329--月报--A楼.xlsx',
        status='ready',
        file_hash='missing-hash',
        size_bytes=100,
    )

    entries = service.get_monthly_by_date_entries(selected_dates=['2026-03-29'], buildings=['A楼'])

    assert entries == []


def test_get_handover_by_date_entries_reuses_latest_matching_date_shift_entry(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    logs: list[str] = []
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda text, *_args, **_kwargs: logs.append(str(text)),
    )

    latest_file = shared_root / '交接班日志源文件' / '202603' / '20260331--21' / '20260331--21--交接班日志源文件--A楼.xlsx'
    _write_minimal_handover_workbook(latest_file)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-03-31 21',
        duty_date='2026-03-30',
        duty_shift='day',
        downloaded_at='2026-03-31 21:05:00',
        relative_path=latest_file.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-handover-a',
        size_bytes=10,
    )

    entries = service.get_handover_by_date_entries(duty_date='2026-03-30', duty_shift='day', buildings=['A楼'])

    assert len(entries) == 1
    assert entries[0]['building'] == 'A楼'
    assert entries[0]['bucket_kind'] == 'latest'
    assert entries[0]['file_path'] == str(latest_file)


def test_get_handover_by_date_entries_reuses_latest_entry_with_legacy_none_duty_context(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    latest_file = shared_root / '交接班日志源文件' / '202604' / '20260401--09' / '20260401--09--交接班日志源文件--A楼.xlsx'
    _write_minimal_handover_workbook(latest_file)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 09',
        duty_date='None',
        duty_shift='none',
        downloaded_at='2026-04-01 09:25:00',
        relative_path=latest_file.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-legacy-none',
        size_bytes=len(b'handover-legacy-none'),
    )

    entries = service.get_handover_by_date_entries(duty_date='2026-04-01', duty_shift='day', buildings=['A楼'])

    assert len(entries) == 1
    assert entries[0]['building'] == 'A楼'
    assert entries[0]['bucket_kind'] == 'latest'
    assert entries[0]['duty_date'] == '2026-04-01'
    assert entries[0]['duty_shift'] == 'day'
    assert entries[0]['file_path'] == str(latest_file)


def test_fill_handover_latest_infers_duty_context_when_downloader_returns_none(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    downloaded_file = work_dir / 'downloaded' / 'A楼.xlsx'
    _write_minimal_handover_workbook(downloaded_file)

    class _FakeDownloadService:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

        def run(self, **_kwargs):
            return {
                'duty_date': None,
                'duty_shift': None,
                'success_files': [
                    {
                        'building': 'A楼',
                        'file_path': str(downloaded_file),
                    }
                ],
            }

    monkeypatch.setattr(cache_module, 'HandoverDownloadService', _FakeDownloadService)

    entry = service.fill_handover_latest(building='A楼', bucket_key='2026-04-01 09', emit_log=lambda *_args, **_kwargs: None)

    assert entry['duty_date'] == '2026-04-01'
    assert entry['duty_shift'] == 'day'
    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 09',
        status='ready',
        limit=1,
    )
    assert len(rows) == 1
    assert rows[0]['duty_date'] == '2026-04-01'
    assert rows[0]['duty_shift'] == 'day'


def test_get_monthly_by_date_entries_ignores_inaccessible_indexed_files(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    actual_file = shared_root / '全景平台月报源文件' / '202603' / '20260329--月报' / '20260329--月报--A楼.xlsx'
    actual_file.parent.mkdir(parents=True, exist_ok=True)
    actual_file.write_bytes(b'monthly-a')
    relative_path = actual_file.relative_to(shared_root).as_posix()
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='date',
        bucket_key='2026-03-29',
        duty_date='2026-03-29',
        duty_shift='',
        downloaded_at='2026-03-29 23:59:00',
        relative_path=relative_path,
        status='ready',
        file_hash='hash-a',
        size_bytes=9,
    )

    monkeypatch.setattr(cache_module, 'is_accessible_cached_file_path', lambda _path: False)

    entries = service.get_monthly_by_date_entries(selected_dates=['2026-03-29'], buildings=['A楼'])

    assert entries == []


def test_external_health_snapshot_resolves_file_path_from_external_root_dir(work_dir: Path) -> None:
    internal_root = work_dir / 'internal-share'
    external_root = work_dir / 'external-share'
    store = SharedBridgeStore(external_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(
            role_mode='external',
            legacy_root=str(work_dir / 'legacy-share'),
            internal_root=str(internal_root),
            external_root=str(external_root),
        ),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-03-30 08'
    service._current_hour_bucket = bucket_key

    actual_file = external_root / '全景平台月报源文件' / '202603' / '20260330--08' / '20260330--08--全景平台月报源文件-A楼.xlsx'
    actual_file.parent.mkdir(parents=True, exist_ok=True)
    actual_file.write_bytes(b'external-health')
    relative_path = actual_file.relative_to(external_root).as_posix()
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-03-30 08:11:00',
        relative_path=relative_path,
        status='ready',
        file_hash='hash-health',
        size_bytes=len(b'external-health'),
    )

    snapshot = service.get_health_snapshot()
    buildings = {item['building']: item for item in snapshot[FAMILY_MONTHLY_REPORT]['buildings']}

    assert buildings['A楼']['status'] == 'ready'
    assert buildings['A楼']['resolved_file_path'] == str(actual_file)


def test_external_health_snapshot_resolves_failed_file_path_from_external_root_dir(work_dir: Path) -> None:
    internal_root = work_dir / 'internal-share'
    external_root = work_dir / 'external-share'
    store = SharedBridgeStore(external_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(
            role_mode='external',
            legacy_root=str(work_dir / 'legacy-share'),
            internal_root=str(internal_root),
            external_root=str(external_root),
        ),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    bucket_key = '2026-03-30 08'
    service._current_hour_bucket = bucket_key

    relative_path = '全景平台月报源文件/202603/20260330--08/20260330--08--全景平台月报源文件-A楼.xlsx'
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key=bucket_key,
        duty_date='',
        duty_shift='',
        downloaded_at='2026-03-30 08:11:00',
        relative_path=relative_path,
        status='failed',
        file_hash='',
        size_bytes=0,
        metadata={'error': '共享目录不可访问'},
    )

    snapshot = service.get_health_snapshot()
    buildings = {item['building']: item for item in snapshot[FAMILY_MONTHLY_REPORT]['buildings']}

    assert buildings['A楼']['status'] == 'failed'
    assert buildings['A楼']['resolved_file_path'] == str(external_root / relative_path)
    assert buildings['A楼']['last_error'] == '共享目录不可访问'


def test_get_latest_ready_selection_allows_fallback_within_three_buckets(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    logs: list[str] = []
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda text, *_args, **_kwargs: logs.append(str(text)),
    )

    latest_a = shared_root / '全景平台月报源文件' / '202603' / '20260330--08' / '20260330--08--全景平台月报源文件--A楼.xlsx'
    latest_b = shared_root / '全景平台月报源文件' / '202603' / '20260330--07' / '20260330--07--全景平台月报源文件--B楼.xlsx'
    latest_a.parent.mkdir(parents=True, exist_ok=True)
    latest_b.parent.mkdir(parents=True, exist_ok=True)
    latest_a.write_bytes(b'a')
    latest_b.write_bytes(b'b')
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-03-30 08',
        duty_date='2026-03-30',
        duty_shift='',
        downloaded_at='2026-03-30 08:01:00',
        relative_path=latest_a.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-a',
        size_bytes=1,
    )
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-03-30 07',
        duty_date='2026-03-30',
        duty_shift='',
        downloaded_at='2026-03-30 07:30:00',
        relative_path=latest_b.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-b',
        size_bytes=1,
    )
    monkeypatch.setattr(
        cache_module,
        '_now_dt',
        lambda: cache_module.datetime(2026, 3, 30, 8, 30, 0),
    )

    selection = service.get_latest_ready_selection(
        source_family=FAMILY_MONTHLY_REPORT,
        buildings=['A楼', 'B楼'],
        max_version_gap=3,
    )

    assert selection['can_proceed'] is True
    assert selection['best_bucket_key'] == '2026-03-30 08'
    assert selection['fallback_buildings'] == ['B楼']
    assert selection['missing_buildings'] == []
    assert selection['stale_buildings'] == []
    building_rows = {item['building']: item for item in selection['buildings']}
    assert building_rows['A楼']['status'] == 'ready'
    assert building_rows['A楼']['using_fallback'] is False
    assert building_rows['B楼']['status'] == 'ready'
    assert building_rows['B楼']['using_fallback'] is True
    assert building_rows['B楼']['version_gap'] == 1


def test_get_latest_ready_selection_blocks_stale_building_over_three_buckets(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    latest_a = shared_root / '交接班日志源文件' / '202603' / '20260330--08' / '20260330--08--交接班日志源文件--A楼.xlsx'
    stale_b = shared_root / '交接班日志源文件' / '202603' / '20260330--04' / '20260330--04--交接班日志源文件--B楼.xlsx'
    _write_minimal_handover_workbook(latest_a)
    _write_minimal_handover_workbook(stale_b)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-03-30 08',
        duty_date='2026-03-30',
        duty_shift='day',
        downloaded_at='2026-03-30 08:05:00',
        relative_path=latest_a.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-a',
        size_bytes=1,
    )
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-03-30 04',
        duty_date='2026-03-30',
        duty_shift='night',
        downloaded_at='2026-03-30 04:05:00',
        relative_path=stale_b.relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-b',
        size_bytes=1,
    )

    selection = service.get_latest_ready_selection(
        source_family=FAMILY_HANDOVER_LOG,
        buildings=['A楼', 'B楼'],
        max_version_gap=3,
    )

    assert selection['can_proceed'] is False
    assert selection['best_bucket_key'] == '2026-03-30 08'
    assert selection['selected_entries'][0]['building'] == 'A楼'
    assert selection['stale_buildings'] == ['B楼']
    building_rows = {item['building']: item for item in selection['buildings']}
    assert building_rows['B楼']['status'] == 'stale'
    assert building_rows['B楼']['version_gap'] == 4


def test_get_latest_ready_selection_blocks_best_bucket_older_than_three_hours(
    monkeypatch: pytest.MonkeyPatch,
    work_dir: Path,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    latest_a = shared_root / '全景平台月报源文件' / '202603' / '20260330--08' / '20260330--08--全景平台月报源文件--A楼.xlsx'
    latest_b = shared_root / '全景平台月报源文件' / '202603' / '20260330--08' / '20260330--08--全景平台月报源文件--B楼.xlsx'
    latest_a.parent.mkdir(parents=True, exist_ok=True)
    latest_b.parent.mkdir(parents=True, exist_ok=True)
    latest_a.write_bytes(b'a')
    latest_b.write_bytes(b'b')
    for building, target in (('A楼', latest_a), ('B楼', latest_b)):
        store.upsert_source_cache_entry(
            source_family=FAMILY_MONTHLY_REPORT,
            building=building,
            bucket_kind='latest',
            bucket_key='2026-03-30 08',
            duty_date='2026-03-30',
            duty_shift='',
            downloaded_at='2026-03-30 08:05:00',
            relative_path=target.relative_to(shared_root).as_posix(),
            status='ready',
            file_hash=f'hash-{building}',
            size_bytes=1,
        )

    monkeypatch.setattr(cache_module, '_now_dt', lambda: cache_module.datetime(2026, 3, 30, 12, 30, 0))

    selection = service.get_latest_ready_selection(
        source_family=FAMILY_MONTHLY_REPORT,
        buildings=['A楼', 'B楼'],
        max_version_gap=3,
        max_selection_age_hours=3.0,
    )

    assert selection['best_bucket_key'] == '2026-03-30 08'
    assert selection['best_bucket_age_hours'] == 4.5
    assert selection['is_best_bucket_too_old'] is True
    assert selection['can_proceed'] is False
    assert selection['stale_buildings'] == []
    assert selection['missing_buildings'] == []


def test_health_snapshot_marks_suspended_building_as_blocked(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()

    class _FakePool:
        @staticmethod
        def get_building_pause_info(building: str) -> dict:
            if building == 'A楼':
                return {
                    'building': 'A楼',
                    'suspended': True,
                    'suspend_reason': 'A楼 登录失败: 页面无响应，请检查楼栋页面服务或网络',
                    'failure_kind': 'login_failed',
                    'recovery_attempts': 3,
                    'last_failure_at': '2026-03-31 22:10:00',
                    'next_probe_at': '2026-03-31 22:11:00',
                    'pending_issue_summary': 'A楼 登录失败: 页面无响应，请检查楼栋页面服务或网络',
                    'login_state': 'failed',
                }
            return {'building': building, 'suspended': False}

    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        download_browser_pool=_FakePool(),
        emit_log=lambda *_args, **_kwargs: None,
    )
    service._current_hour_bucket = '2026-03-31 22'

    snapshot = service.get_health_snapshot()
    building = next(item for item in snapshot[FAMILY_HANDOVER_LOG]['buildings'] if item['building'] == 'A楼')

    assert building['status'] == 'waiting'
    assert building['blocked'] is True
    assert 'A楼 登录失败' in building['blocked_reason']
    assert building['next_probe_at'] == '2026-03-31 22:11:00'
    assert snapshot[FAMILY_HANDOVER_LOG]['blocked_buildings'] == ['A楼']


def test_refresh_family_bucket_skips_suspended_building_without_failed_entry(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()

    class _FakePool:
        @staticmethod
        def get_building_pause_info(building: str) -> dict:
            if building == 'A楼':
                return {
                    'building': 'A楼',
                    'suspended': True,
                    'suspend_reason': 'A楼 页面无响应: 页面无响应，请检查楼栋页面服务或网络',
                    'failure_kind': 'page_unreachable',
                    'recovery_attempts': 3,
                    'last_failure_at': '2026-03-31 22:10:00',
                    'next_probe_at': '2026-03-31 22:11:00',
                    'pending_issue_summary': 'A楼 页面无响应: 页面无响应，请检查楼栋页面服务或网络',
                    'login_state': 'failed',
                }
            return {'building': building, 'suspended': False}

    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        download_browser_pool=_FakePool(),
        emit_log=lambda *_args, **_kwargs: None,
    )

    calls: list[str] = []

    def _fake_fill(*, building: str, bucket_key: str, emit_log):  # noqa: ANN001, ARG001
        calls.append(f'{building}:{bucket_key}')

    service.get_enabled_buildings = lambda: ['A楼']  # type: ignore[method-assign]
    service._refresh_family_bucket(
        source_family=FAMILY_HANDOVER_LOG,
        bucket_key='2026-03-31 22',
        fill_func=_fake_fill,
    )

    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-03-31 22',
        limit=10,
    )

    assert calls == []
    assert rows == []
    assert service._family_status[FAMILY_HANDOVER_LOG]['blocked_buildings'] == ['A楼']


def test_fill_alarm_event_manual_indexes_latest_hour_bucket_and_path(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()

    class _FakePool:
        @staticmethod
        def submit_building_alarm_job(building: str, runner):  # noqa: ANN001, ARG004
            future = concurrent.futures.Future()
            future.set_result(
                {
                    'query_start': '2026-02-01 00:00:00',
                    'query_end': '2026-04-01 12:05:00',
                    'rows': [],
                }
            )
            return future

    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        download_browser_pool=_FakePool(),
        emit_log=lambda *_args, **_kwargs: None,
    )

    entry = service.fill_alarm_event_manual(
        building='A楼',
        bucket_key='2026-04-01 12',
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert entry['bucket_kind'] == 'latest'
    assert '20260401--12/20260401--12--告警信息源文件--A楼.json' in entry['relative_path']
    rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 12',
        status='ready',
        limit=1,
    )
    assert len(rows) == 1
    assert rows[0]['metadata']['manual'] is True

    second_entry = service.fill_alarm_event_manual(
        building='A楼',
        bucket_key='2026-04-01 12',
        emit_log=lambda *_args, **_kwargs: None,
    )
    assert second_entry['relative_path'] == entry['relative_path']
    rows_after_overwrite = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 12',
        status='ready',
        limit=10,
    )
    assert len(rows_after_overwrite) == 1


def test_delete_manual_alarm_files_only_deletes_manual_entries(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    manual_source = work_dir / 'manual.json'
    scheduled_source = work_dir / 'scheduled.json'
    _write_alarm_json(manual_source, bucket_kind='manual', bucket_key='2026-04-01 manual-1200')
    _write_alarm_json(scheduled_source, bucket_kind='latest', bucket_key='2026-04-01 08')

    manual_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='manual',
        bucket_key='2026-04-01 manual-1200',
        duty_date='',
        duty_shift='',
        source_path=manual_source,
        status='ready',
        metadata={'manual': True},
    )
    scheduled_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=scheduled_source,
        status='ready',
        metadata={},
    )

    manual_path = Path(manual_entry['file_path'])
    scheduled_path = Path(scheduled_entry['file_path'])
    assert manual_path.exists()
    assert scheduled_path.exists()

    result = service.delete_manual_alarm_files()

    assert result['accepted'] is True
    assert result['deleted_count'] == 1
    assert not manual_path.exists()
    assert scheduled_path.exists()
    assert store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='manual',
        bucket_key='2026-04-01 manual-1200',
        limit=10,
    ) == []
    remaining = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='ready',
        limit=1,
    )
    assert len(remaining) == 1


def test_cleanup_expired_entries_removes_db_row_before_file_delete(work_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    source_file = work_dir / 'old.xlsx'
    workbook = openpyxl.Workbook()
    workbook.save(source_file)
    entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=source_file,
        status='ready',
        metadata={},
    )
    store.upsert_source_cache_entry(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        downloaded_at='2026-01-01 00:00:00',
        relative_path=Path(entry['file_path']).relative_to(shared_root).as_posix(),
        status='ready',
        file_hash='hash-old',
        size_bytes=1,
        metadata={},
    )
    monkeypatch.setattr(service, '_safe_delete_cached_file', lambda **_kwargs: False)

    result = service.cleanup_expired_entries(limit=20)
    rows = store.list_source_cache_entries(
        source_family=FAMILY_MONTHLY_REPORT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        limit=5,
    )

    assert result['deleted_entries'] == 1
    assert rows == []


def test_cleanup_expired_entries_removes_old_orphan_cached_file(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    orphan_file = shared_root / '全景平台月报源文件' / '202501' / '20250101--08' / '20250101--08--全景平台月报源文件--A楼.xlsx'
    orphan_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    workbook.save(orphan_file)
    old_timestamp = (datetime.now() - timedelta(days=SharedSourceCacheService.ORPHAN_FILE_RETENTION_DAYS + 2)).timestamp()
    os.utime(orphan_file, (old_timestamp, old_timestamp))

    result = service.cleanup_expired_entries(limit=20)

    assert result['deleted_orphan_files'] == 1
    assert not orphan_file.exists()


def test_internal_light_snapshot_includes_manual_alarm_refresh_summary(work_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='internal', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    def _fake_fill_alarm_event_manual(*, building: str, bucket_key: str, emit_log):  # noqa: ANN001
        row_count_map = {'A楼': 12, 'B楼': 8}
        count = row_count_map.get(building, 0)
        return {
            'building': building,
            'bucket_kind': 'latest',
            'bucket_key': bucket_key,
            'downloaded_at': '2026-04-03 00:20:18',
            'relative_path': f'告警信息源文件/test/{building}.json',
            'file_path': str(shared_root / '告警信息源文件' / 'test' / f'{building}.json'),
            'metadata': {
                'row_count': count,
                'query_start': '2026-02-02 00:20:18',
                'query_end': '2026-04-03 00:20:18',
            },
        }

    monkeypatch.setattr(service, 'get_enabled_buildings', lambda: ['A楼', 'B楼'])
    monkeypatch.setattr(service, 'fill_alarm_event_manual', _fake_fill_alarm_event_manual)

    result = service._run_manual_alarm_refresh_impl()  # noqa: SLF001
    snapshot = service.get_health_snapshot(mode='internal_light')
    manual_refresh = snapshot[FAMILY_ALARM_EVENT]['manual_refresh']

    assert result['accepted'] is True
    assert result['total_row_count'] == 20
    assert result['building_row_counts'] == {'A楼': 12, 'B楼': 8}
    assert manual_refresh['running'] is False
    assert manual_refresh['total_row_count'] == 20
    assert manual_refresh['building_row_counts'] == {'A楼': 12, 'B楼': 8}
    assert manual_refresh['successful_buildings'] == ['A楼', 'B楼']
    assert manual_refresh['query_start'] == '2026-02-02 00:20:18'
    assert manual_refresh['query_end'] == '2026-04-03 00:20:18'


def test_external_consume_ready_alarm_event_entries_is_retired_and_keeps_files_unchanged(work_dir: Path) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    scheduled_08 = work_dir / 'scheduled_08.json'
    scheduled_16 = work_dir / 'scheduled_16.json'
    unscheduled_10 = work_dir / 'scheduled_10.json'
    manual_source = work_dir / 'manual.json'
    _write_alarm_json(scheduled_08, bucket_key='2026-04-01 08')
    _write_alarm_json(scheduled_16, building='B楼', bucket_key='2026-04-01 16')
    _write_alarm_json(unscheduled_10, building='C楼', bucket_key='2026-04-01 10')
    _write_alarm_json(manual_source, building='D楼', bucket_kind='manual', bucket_key='2026-04-01 manual-1200')

    entry_08 = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=scheduled_08,
        status='ready',
        metadata={},
    )
    entry_16 = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 16',
        duty_date='',
        duty_shift='',
        source_path=scheduled_16,
        status='ready',
        metadata={},
    )
    unscheduled_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='C楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 10',
        duty_date='',
        duty_shift='',
        source_path=unscheduled_10,
        status='ready',
        metadata={},
    )
    manual_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='D楼',
        bucket_kind='manual',
        bucket_key='2026-04-01 manual-1200',
        duty_date='',
        duty_shift='',
        source_path=manual_source,
        status='ready',
        metadata={'manual': True},
    )

    result = service.consume_ready_alarm_event_entries()

    assert result['accepted'] is False
    assert result['reason'] == 'retired'
    assert result['consumed_count'] == 0
    assert Path(entry_08['file_path']).exists()
    assert Path(entry_16['file_path']).exists()
    assert Path(unscheduled_entry['file_path']).exists()
    assert Path(manual_entry['file_path']).exists()
    ready_08 = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='ready',
        limit=1,
    )
    ready_16 = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 16',
        status='ready',
        limit=1,
    )
    assert len(ready_08) == 1
    assert len(ready_16) == 1
    remaining_unscheduled = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='C楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 10',
        status='ready',
        limit=1,
    )
    remaining_manual = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='D楼',
        bucket_kind='manual',
        bucket_key='2026-04-01 manual-1200',
        status='ready',
        limit=1,
    )
    assert len(remaining_unscheduled) == 1
    assert len(remaining_manual) == 1


def test_external_upload_alarm_entries_full_keeps_files_ready_after_success(work_dir: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    class _FakeBitableClient:
        instances = []

        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.clear_calls = []
            self.create_calls = []
            _FakeBitableClient.instances.append(self)

        def clear_table(self, table_id: str, list_page_size: int = 500, delete_batch_size: int = 500) -> int:
            self.clear_calls.append(
                {
                    'table_id': table_id,
                    'list_page_size': list_page_size,
                    'delete_batch_size': delete_batch_size,
                }
            )
            return 3

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append(
                {
                    'table_id': table_id,
                    'batch_size': batch_size,
                    'fields_list': fields_list,
                }
            )
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    source = work_dir / 'alarm_a.json'
    _write_alarm_json(
        source,
        building='A楼',
        bucket_key='2026-04-01 08',
        rows=[
            {
                'level': '次要',
                'content': '风机状态: 告警',
                'position': 'E楼/三层/空调区2',
                'object': 'E-311-CRAH-10',
                'event_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'accept_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'is_accept': '已处理',
                'accept_by': '系统管理员',
                'accept_content': '测试受理',
                'recover_time': '',
                'is_recover': '未恢复',
                'event_snapshot': '12.5',
                'event_type': '不正常值',
                'confirm_type': '真实告警',
                'event_suggest': '测试建议',
                'confirm_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'confirm_by': '系统管理员',
                'confirm_description': '测试确认',
                'alarm_threshold': '',
            }
        ],
    )
    entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=source,
        status='ready',
        metadata={},
    )
    assert not source.exists()
    cached_path = Path(entry['file_path'])
    assert cached_path.exists()

    result = service.upload_alarm_event_entries_full_to_bitable()

    assert result['accepted'] is True
    assert result['uploaded_record_count'] == 1
    assert result['consumed_count'] == 0
    assert result['consumed_buildings'] == []
    assert cached_path.exists()
    ready_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='ready',
        limit=1,
    )
    assert len(ready_rows) == 1
    assert ready_rows[0]['metadata']['last_uploaded_mode'] == 'full'
    assert ready_rows[0]['metadata']['last_uploaded_scope'] == 'all'

    assert len(_FakeBitableClient.instances) == 1
    fake_client = _FakeBitableClient.instances[0]
    assert len(fake_client.clear_calls) == 1
    assert len(fake_client.create_calls) == 1
    fields = fake_client.create_calls[0]['fields_list'][0]
    assert fields['楼栋'] == 'A楼'
    assert fields['告警内容'] == '风机状态: 告警'
    assert isinstance(fields['产生时间'], int)
    assert isinstance(fields['受理时间'], int)
    assert isinstance(fields['确认时间'], int)
    assert isinstance(fields['触发值'], (int, float))
    assert '恢复时间' not in fields
    assert set(fields.keys()).issubset(set(ALARM_EVENT_BITABLE_TARGET_FIELDS.values()))


def test_external_upload_alarm_entries_full_logs_progress_every_100_records(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows = [
        {
            'level': '次要',
            'content': f'记录-{idx}',
            'position': 'A楼/位置',
            'object': f'A-{idx}',
            'event_time': now_text,
            'accept_time': now_text,
            'is_accept': '已处理',
            'accept_by': '系统管理员',
            'accept_content': '受理',
            'recover_time': '',
            'is_recover': '未恢复',
            'event_snapshot': str(idx),
            'event_type': '不正常值',
            'confirm_type': '真实告警',
            'event_suggest': '建议',
            'confirm_time': now_text,
            'confirm_by': '系统管理员',
            'confirm_description': '确认',
            'alarm_threshold': '',
        }
        for idx in range(250)
    ]

    class _FakeBitableClient:
        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.create_calls = []

        def clear_table(self, table_id: str, list_page_size: int = 500, delete_batch_size: int = 500) -> int:
            return 0

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append({'table_id': table_id, 'batch_size': batch_size, 'size': len(fields_list)})
            for uploaded in (60, 120, 180, 240, len(fields_list)):
                if callable(progress_callback):
                    progress_callback(uploaded, len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    source = work_dir / 'alarm_progress.json'
    _write_alarm_json(
        source,
        building='A楼',
        bucket_key='2026-04-01 08',
        rows=rows,
    )
    service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=source,
        status='ready',
        metadata={},
    )

    logs: list[str] = []
    result = service.upload_alarm_event_entries_full_to_bitable(emit_log=logs.append)

    assert result['accepted'] is True
    assert any('uploaded=100/250' in line for line in logs)
    assert any('uploaded=200/250' in line for line in logs)
    assert not any('uploaded=300/250' in line for line in logs)


def test_external_upload_alarm_entries_full_rejects_precheck_failure_without_remote_mutation(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    class _FakeBitableClient:
        instances = []

        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.clear_calls = []
            self.create_calls = []
            _FakeBitableClient.instances.append(self)

        def clear_table(self, table_id: str, list_page_size: int = 500, delete_batch_size: int = 500) -> int:
            self.clear_calls.append({'table_id': table_id})
            return 0

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append({'table_id': table_id, 'fields_list': fields_list})
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    valid_source = work_dir / 'alarm_valid.json'
    invalid_source = work_dir / 'alarm_invalid.json'
    now_text = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _write_alarm_json(
        valid_source,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        rows=[
            {
                'level': '次要',
                'content': '有效记录',
                'position': 'A楼/位置',
                'object': 'A-OBJ',
                'event_time': now_text,
                'accept_time': now_text,
                'is_accept': '已处理',
                'accept_by': '系统管理员',
                'accept_content': '受理',
                'recover_time': '',
                'is_recover': '未恢复',
                'event_snapshot': '1',
                'event_type': '不正常值',
                'confirm_type': '真实告警',
                'event_suggest': '建议',
                'confirm_time': now_text,
                'confirm_by': '系统管理员',
                'confirm_description': '确认',
                'alarm_threshold': '',
            }
        ],
    )
    _write_alarm_json(
        invalid_source,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        rows=[],
    )
    invalid_payload = json.loads(invalid_source.read_text(encoding='utf-8'))
    invalid_payload['schema_version'] = 2
    invalid_source.write_text(json.dumps(invalid_payload, ensure_ascii=False, indent=2), encoding='utf-8')
    valid_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=valid_source,
        status='ready',
        metadata={},
    )
    invalid_entry = service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=invalid_source,
        status='ready',
        metadata={},
    )

    result = service.upload_alarm_event_entries_full_to_bitable()

    assert result['accepted'] is False
    assert result['reason'] == 'precheck_failed'
    assert result['uploaded_record_count'] == 0
    assert result['consumed_count'] == 0
    assert result['failed_entries'] == ['B楼']
    assert _FakeBitableClient.instances == []
    assert Path(valid_entry['file_path']).exists()
    assert Path(invalid_entry['file_path']).exists()
    ready_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='ready',
        limit=1,
    )
    failed_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        status='failed',
        limit=1,
    )
    assert len(ready_rows) == 1
    assert len(failed_rows) == 1


def test_external_upload_alarm_entries_is_single_flight_and_reports_running_state(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    source = work_dir / 'alarm_single_flight.json'
    _write_alarm_json(
        source,
        building='A楼',
        bucket_key='2026-04-01 08',
        rows=[
            {
                'level': '次要',
                'content': '风机状态: 告警',
                'position': 'A楼/三层/空调区2',
                'object': 'A-311-CRAH-10',
                'event_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'accept_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'is_accept': '已处理',
                'accept_by': '系统管理员',
                'accept_content': '测试受理',
                'recover_time': '',
                'is_recover': '未恢复',
                'event_snapshot': '12.5',
                'event_type': '不正常值',
                'confirm_type': '真实告警',
                'event_suggest': '测试建议',
                'confirm_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'confirm_by': '系统管理员',
                'confirm_description': '测试确认',
                'alarm_threshold': '',
            }
        ],
    )
    service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 08',
        duty_date='',
        duty_shift='',
        source_path=source,
        status='ready',
        metadata={},
    )

    started = threading.Event()
    release = threading.Event()

    class _BlockingBitableClient:
        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            pass

        def clear_table(self, table_id: str, list_page_size: int = 500, delete_batch_size: int = 500) -> int:
            started.set()
            assert release.wait(timeout=5)
            return 0

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _BlockingBitableClient)

    full_result: dict[str, object] = {}

    def _run_full_upload() -> None:
        full_result.update(service.upload_alarm_event_entries_full_to_bitable())

    worker = threading.Thread(target=_run_full_upload, daemon=True)
    worker.start()

    assert started.wait(timeout=2)
    snapshot = service.get_health_snapshot(mode='external_full')
    upload_state = snapshot[FAMILY_ALARM_EVENT]['external_upload']
    assert upload_state['running'] is True
    assert upload_state['current_mode'] == 'full'
    assert upload_state['current_scope'] == 'all'

    second_result = service.upload_alarm_event_entries_single_building_to_bitable(building='A楼')
    assert second_result['accepted'] is False
    assert second_result['reason'] == 'already_running'
    assert second_result['running'] is True

    release.set()
    worker.join(timeout=5)
    assert full_result['accepted'] is True

    snapshot_after = service.get_health_snapshot(mode='external_full')
    upload_state_after = snapshot_after[FAMILY_ALARM_EVENT]['external_upload']
    assert upload_state_after['running'] is False
    assert upload_state_after['last_mode'] == 'full'
    assert upload_state_after['last_scope'] == 'all'


def test_external_alarm_selection_prefers_today_latest_else_yesterday_fallback(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    fixed_now = datetime(2026, 4, 3, 10, 0, 0)
    monkeypatch.setattr(cache_module, '_now_dt', lambda: fixed_now)

    now_text = fixed_now.strftime('%Y-%m-%d %H:%M:%S')
    rows = [
        {
            'level': '次要',
            'content': '测试记录',
            'position': '测试位置',
            'object': 'TEST-OBJ',
            'event_time': now_text,
            'accept_time': now_text,
            'is_accept': '已处理',
            'accept_by': '系统管理员',
            'accept_content': '受理',
            'recover_time': '',
            'is_recover': '未恢复',
            'event_snapshot': '1',
            'event_type': '不正常值',
            'confirm_type': '真实告警',
            'event_suggest': '建议',
            'confirm_time': now_text,
            'confirm_by': '系统管理员',
            'confirm_description': '确认',
            'alarm_threshold': '',
        }
    ]
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='A楼',
        bucket_kind='latest',
        bucket_key='2026-04-03 08',
        downloaded_at='2026-04-03 08:00:00',
        rows=rows,
    )
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='A楼',
        bucket_kind='manual',
        bucket_key='2026-04-03 09:30:00',
        downloaded_at='2026-04-03 09:30:00',
        rows=rows,
    )
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-02 16',
        downloaded_at='2026-04-02 16:05:00',
        rows=rows,
    )
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='C楼',
        bucket_kind='latest',
        bucket_key='2026-04-02 16',
        downloaded_at='2026-04-02 16:10:00',
        rows=rows,
    )
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='C楼',
        bucket_kind='latest',
        bucket_key='2026-04-03 09',
        downloaded_at='2026-04-03 09:10:00',
        status='consumed',
        rows=rows,
    )
    _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='E楼',
        bucket_kind='latest',
        bucket_key='2026-04-03 07',
        downloaded_at='2026-04-03 07:05:00',
        rows=rows,
    )

    snapshot = service.get_health_snapshot(mode='external_full')
    alarm_family = snapshot[FAMILY_ALARM_EVENT]
    buildings = {item['building']: item for item in alarm_family['buildings']}
    selected_entries = service._select_alarm_ready_entries_for_external_upload()  # noqa: SLF001

    assert alarm_family['selection_policy'] == 'today_latest_else_yesterday_fallback'
    assert alarm_family['selection_reference_date'] == '2026-04-03'
    assert alarm_family['used_previous_day_fallback'] == ['B楼', 'C楼']
    assert sorted(alarm_family['missing_today_buildings']) == ['B楼', 'C楼', 'D楼']
    assert alarm_family['missing_both_days_buildings'] == ['D楼']

    assert buildings['A楼']['source_kind'] == 'manual'
    assert buildings['A楼']['selection_scope'] == 'today'
    assert buildings['A楼']['status'] == 'ready'
    assert buildings['B楼']['source_kind'] == 'latest'
    assert buildings['B楼']['selection_scope'] == 'yesterday_fallback'
    assert buildings['B楼']['status'] == 'ready'
    assert buildings['C楼']['selection_scope'] == 'yesterday_fallback'
    assert buildings['C楼']['status'] == 'ready'
    assert buildings['D楼']['selection_scope'] == 'missing'
    assert buildings['D楼']['status'] == 'waiting'
    assert buildings['E楼']['selection_scope'] == 'today'
    assert buildings['E楼']['status'] == 'ready'

    selected_by_building = {item['building']: item for item in selected_entries}
    assert set(selected_by_building.keys()) == {'A楼', 'B楼', 'C楼', 'E楼'}
    assert selected_by_building['A楼']['bucket_kind'] == 'manual'
    assert selected_by_building['B楼']['bucket_key'] == '2026-04-02 16'
    assert selected_by_building['C楼']['bucket_key'] == '2026-04-02 16'
    assert selected_by_building['E楼']['bucket_key'] == '2026-04-03 07'


def test_external_upload_alarm_entries_single_building_keeps_only_rows_within_60_days(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    logs: list[str] = []
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda text, *_args, **_kwargs: logs.append(str(text)),
    )

    class _FakeBitableClient:
        instances = []

        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.clear_calls = []
            self.create_calls = []
            self.list_calls = []
            self.delete_calls = []
            _FakeBitableClient.instances.append(self)

        def clear_table(self, table_id: str, list_page_size: int = 500, delete_batch_size: int = 500) -> int:
            self.clear_calls.append({'table_id': table_id, 'list_page_size': list_page_size, 'delete_batch_size': delete_batch_size})
            return 0

        def list_records(self, table_id: str, page_size: int = 500, max_records: int = 0, *, view_id: str = "", filter_formula: str = "") -> list[dict]:
            self.list_calls.append({'table_id': table_id, 'page_size': page_size})
            return [
                {'record_id': 'record-b-1', 'fields': {'楼栋': 'B楼', '告警内容': '旧记录'}},
                {'record_id': 'record-a-1', 'fields': {'楼栋': 'A楼', '告警内容': '其他楼栋记录'}},
            ]

        def batch_delete_records(
            self,
            table_id: str,
            record_ids: list[str],
            batch_size: int = 500,
            progress_callback=None,
        ) -> int:
            self.delete_calls.append({'table_id': table_id, 'record_ids': list(record_ids), 'batch_size': batch_size})
            if callable(progress_callback):
                progress_callback(0, len(record_ids))
                progress_callback(len(record_ids), len(record_ids))
            return len(record_ids)

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append({'table_id': table_id, 'batch_size': batch_size, 'fields_list': fields_list})
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['级别', '内容', '位置', '对象', '告警时间', '接警时间', '处理状态', '处理人', '处理内容', '恢复时间', '恢复状态', '告警快照', '事件类型', '确认类型', '建议', '确认时间', '确认人', '确认说明', '实时值', '阈值'])
    recent_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    old_time = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d %H:%M:%S')
    sheet.append(['次要', '近60天记录', 'B楼/位置', 'B-OBJ', recent_time, recent_time, '已处理', '值班员', '受理', '--', '未恢复', '关闭', '不正常值', '真实告警', '建议', recent_time, '值班员', '确认', '2', '0'])
    sheet.append(['次要', '超窗记录', 'B楼/位置', 'B-OBJ', old_time, old_time, '已处理', '值班员', '受理', '--', '未恢复', '关闭', '不正常值', '真实告警', '建议', old_time, '值班员', '确认', '3', '0'])
    source = work_dir / 'alarm_b.xlsx'
    workbook.save(source)
    service._store_entry(  # noqa: SLF001
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key='2026-04-01 16',
        duty_date='',
        duty_shift='',
        source_path=source,
        status='ready',
        metadata={},
    )

    result = service.upload_alarm_event_entries_single_building_to_bitable(building='B楼')

    assert result['accepted'] is True
    assert result['uploaded_record_count'] == 1
    assert result['consumed_count'] == 0
    assert result['consumed_buildings'] == []
    assert len(_FakeBitableClient.instances) == 1
    fake_client = _FakeBitableClient.instances[0]
    assert fake_client.clear_calls == []
    assert len(fake_client.list_calls) == 1
    assert len(fake_client.delete_calls) == 1
    assert fake_client.delete_calls[0]['record_ids'] == ['record-b-1']
    assert len(fake_client.create_calls) == 1
    uploaded_rows = fake_client.create_calls[0]['fields_list']
    assert len(uploaded_rows) == 1
    assert uploaded_rows[0]['告警内容'] == '近60天记录'
    assert any('外网告警开始删除旧记录' in text for text in logs)
    assert any('外网告警删除旧记录进度' in text and 'deleted=1/1' in text for text in logs)
    assert any('外网告警删除旧记录完成' in text and 'deleted=1' in text for text in logs)


def test_external_upload_alarm_entries_single_building_allows_yesterday_fallback(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    fixed_now = datetime(2026, 4, 3, 10, 0, 0)
    monkeypatch.setattr(cache_module, '_now_dt', lambda: fixed_now)

    class _FakeBitableClient:
        instances = []

        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.list_calls = []
            self.delete_calls = []
            self.create_calls = []
            _FakeBitableClient.instances.append(self)

        def list_records(self, table_id: str, page_size: int = 500, max_records: int = 0, *, view_id: str = "", filter_formula: str = "") -> list[dict]:
            self.list_calls.append({'table_id': table_id, 'page_size': page_size})
            return []

        def batch_delete_records(self, table_id: str, record_ids: list[str], batch_size: int = 500) -> int:
            self.delete_calls.append({'table_id': table_id, 'record_ids': list(record_ids), 'batch_size': batch_size})
            return 0

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append({'table_id': table_id, 'batch_size': batch_size, 'fields_list': fields_list})
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    rows = [
        {
            'level': '次要',
            'content': '昨天最新文件',
            'position': 'B楼/位置',
            'object': 'B-OBJ',
            'event_time': fixed_now.strftime('%Y-%m-%d %H:%M:%S'),
            'accept_time': fixed_now.strftime('%Y-%m-%d %H:%M:%S'),
            'is_accept': '已处理',
            'accept_by': '系统管理员',
            'accept_content': '受理',
            'recover_time': '',
            'is_recover': '未恢复',
            'event_snapshot': '2',
            'event_type': '不正常值',
            'confirm_type': '真实告警',
            'event_suggest': '建议',
            'confirm_time': fixed_now.strftime('%Y-%m-%d %H:%M:%S'),
            'confirm_by': '系统管理员',
            'confirm_description': '确认',
            'alarm_threshold': '',
        }
    ]
    entry = _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='B楼',
        bucket_kind='manual',
        bucket_key='2026-04-02 18:00:00',
        downloaded_at='2026-04-02 18:00:00',
        rows=rows,
    )

    result = service.upload_alarm_event_entries_single_building_to_bitable(building='B楼')

    assert result['accepted'] is True
    assert result['uploaded_record_count'] == 1
    assert result['consumed_count'] == 0
    assert result['consumed_buildings'] == []
    assert result['used_previous_day_fallback'] == ['B楼']
    assert result['selection_reference_date'] == '2026-04-03'
    assert Path(entry['file_path']).exists()
    ready_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='manual',
        bucket_key='2026-04-02 18:00:00',
        status='ready',
        limit=1,
    )
    assert len(ready_rows) == 1
    assert ready_rows[0]['metadata']['last_uploaded_mode'] == 'single_building'
    assert ready_rows[0]['metadata']['last_uploaded_scope'] == 'B楼'


def test_external_upload_alarm_entries_single_building_keeps_selected_and_older_ready_files(
    work_dir: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    shared_root = work_dir / 'shared'
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode='external', shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    class _FakeBitableClient:
        instances = []

        def __init__(self, *args, **kwargs):  # noqa: ANN002, ANN003
            self.list_calls = []
            self.delete_calls = []
            self.create_calls = []
            _FakeBitableClient.instances.append(self)

        def list_records(self, table_id: str, page_size: int = 500, max_records: int = 0, *, view_id: str = "", filter_formula: str = "") -> list[dict]:
            self.list_calls.append({'table_id': table_id, 'page_size': page_size})
            return []

        def batch_delete_records(self, table_id: str, record_ids: list[str], batch_size: int = 500) -> int:
            self.delete_calls.append({'table_id': table_id, 'record_ids': list(record_ids), 'batch_size': batch_size})
            return 0

        def batch_create_records(
            self,
            table_id: str,
            fields_list: list[dict],
            batch_size: int = 200,
            progress_callback=None,
        ) -> list[dict]:
            self.create_calls.append({'table_id': table_id, 'batch_size': batch_size, 'fields_list': fields_list})
            if callable(progress_callback):
                progress_callback(len(fields_list), len(fields_list))
            return []

    monkeypatch.setattr(cache_module, 'FeishuBitableClient', _FakeBitableClient)

    now = datetime.now()
    now_text = now.strftime('%Y-%m-%d %H:%M:%S')
    manual_bucket_key = now.strftime('%Y-%m-%d %H:%M:%S')
    yesterday_bucket_key = (now - timedelta(days=1)).strftime('%Y-%m-%d 16')
    yesterday_downloaded_at = (now - timedelta(days=1)).strftime('%Y-%m-%d 16:00:00')
    rows = [
        {
            'level': '次要',
            'content': '较新手动文件',
            'position': 'B楼/位置',
            'object': 'B-OBJ',
            'event_time': now_text,
            'accept_time': now_text,
            'is_accept': '已处理',
            'accept_by': '系统管理员',
            'accept_content': '受理',
            'recover_time': '',
            'is_recover': '未恢复',
            'event_snapshot': '1',
            'event_type': '不正常值',
            'confirm_type': '真实告警',
            'event_suggest': '建议',
            'confirm_time': now_text,
            'confirm_by': '系统管理员',
            'confirm_description': '确认',
            'alarm_threshold': '',
        }
    ]
    newer_entry = _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='B楼',
        bucket_kind='manual',
        bucket_key=manual_bucket_key,
        downloaded_at=now_text,
        rows=rows,
    )
    older_entry = _add_alarm_cache_entry(
        store=store,
        shared_root=shared_root,
        building='B楼',
        bucket_kind='latest',
        bucket_key=yesterday_bucket_key,
        downloaded_at=yesterday_downloaded_at,
        rows=rows,
    )

    result = service.upload_alarm_event_entries_single_building_to_bitable(building='B楼')

    assert result['accepted'] is True
    assert result['consumed_count'] == 0
    assert Path(newer_entry['file_path']).exists()
    assert Path(older_entry['file_path']).exists()
    newer_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='manual',
        bucket_key=manual_bucket_key,
        status='ready',
        limit=1,
    )
    older_rows = store.list_source_cache_entries(
        source_family=FAMILY_ALARM_EVENT,
        building='B楼',
        bucket_kind='latest',
        bucket_key=yesterday_bucket_key,
        status='ready',
        limit=1,
    )
    assert len(newer_rows) == 1
    assert len(older_rows) == 1
    assert newer_rows[0]['metadata']['last_uploaded_mode'] == 'single_building'
    assert newer_rows[0]['metadata']['last_uploaded_scope'] == 'B楼'
    assert 'consumed_reason' not in older_rows[0]['metadata']

