from __future__ import annotations

from pathlib import Path

import openpyxl

from handover_log_module.service.handover_summary_message_service import HandoverSummaryMessageService


def _write_handover_file(path: Path) -> None:
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "交接班日志"
    ws["A1"] = "EA118机房A栋数据中心交接班日志"
    ws["C3"] = "张三 李四"
    ws["G3"] = "王五 赵六"
    ws["B6"] = "1.45"
    ws["D6"] = "2432"
    ws["F6"] = "1674.72"
    ws["B13"] = "1272000"
    ws["D13"] = "330000"
    workbook.save(path)
    workbook.close()


def test_handover_summary_message_uses_capacity_cooling_summary(tmp_path: Path, monkeypatch) -> None:
    output_file = tmp_path / "handover.xlsx"
    _write_handover_file(output_file)
    service = HandoverSummaryMessageService({"template": {"sheet_name": "交接班日志"}})
    monkeypatch.setattr(service, "_lookup_contact_phones", lambda names, *, emit_log: {})

    message = service.build_for_session(
        {
            "building": "A楼",
            "duty_date": "2026-04-29",
            "duty_shift": "day",
            "output_file": str(output_file),
            "capacity_cooling_summary": {
                "lines": {
                    "west": "冷冻站A区3套制冷单元2用1备，2#制冷单元板换模式运行正常；",
                    "east": "冷冻站B区3套制冷单元2用1备，4#制冷单元板换模式运行正常；",
                }
            },
        },
        emit_log=lambda _msg: None,
    )

    assert "2、冷冻站A区3套制冷单元2用1备，2#制冷单元板换模式运行正常；" in message
    assert "3、冷冻站B区3套制冷单元2用1备，4#制冷单元板换模式运行正常；" in message
    assert "制冷单元按当前运行方式运行" not in message


def test_handover_summary_message_uses_manual_cooling_level_values(tmp_path: Path, monkeypatch) -> None:
    output_file = tmp_path / "handover.xlsx"
    _write_handover_file(output_file)
    service = HandoverSummaryMessageService({"template": {"sheet_name": "交接班日志"}})
    monkeypatch.setattr(service, "_lookup_contact_phones", lambda names, *, emit_log: {})

    message = service.build_for_session(
        {
            "building": "E楼",
            "duty_date": "2026-04-29",
            "duty_shift": "day",
            "output_file": str(output_file),
            "capacity_running_units": {
                "west": [{"unit": 2, "mode_text": "板换"}, {"unit": 3, "mode_text": "板换"}],
                "east": [{"unit": 4, "mode_text": "板换"}, {"unit": 5, "mode_text": "板换"}],
            },
            "capacity_cooling_summary": {
                "lines": {
                    "west": "冷冻站A区3套制冷单元2用1备，1#2#二次泵运行正常；",
                    "east": "冷冻站B区3套制冷单元2用1备，5#6#二次泵运行正常；",
                }
            },
            "cooling_pump_pressures": {
                "rows": [
                    {"zone": "west", "unit": 2, "mode_text": "板换", "cooling_tower_level": "0.43"},
                    {"zone": "west", "unit": 3, "mode_text": "板换", "cooling_tower_level": "0.4"},
                    {"zone": "east", "unit": 4, "mode_text": "板换", "cooling_tower_level": "0.36"},
                    {"zone": "east", "unit": 5, "mode_text": "板换", "cooling_tower_level": "0.35"},
                ],
                "tanks": {
                    "west": {"temperature": "16.9", "level": "27.32"},
                    "east": {"temperature": "16.9", "level": "27.3"},
                },
            },
        },
        emit_log=lambda _msg: None,
    )

    assert "2、冷冻站A区3套制冷单元2用1备" in message
    assert "2#冷却塔液位0.43m正常" in message
    assert "3#冷却塔液位0.4m正常" in message
    assert "蓄冷罐后备温度16.9℃正常、液位27.32m正常" in message
    assert "3、冷冻站B区3套制冷单元2用1备" in message
    assert "4#冷却塔液位0.36m正常" in message
    assert "5#冷却塔液位0.35m正常" in message
    assert "蓄冷罐后备温度16.9℃正常、液位27.3m正常" in message
