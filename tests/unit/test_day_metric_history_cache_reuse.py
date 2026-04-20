from __future__ import annotations

import shutil
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

from app.modules.shared_bridge.service.shared_bridge_store import SharedBridgeStore
from app.modules.shared_bridge.service.shared_source_cache_service import (
    FAMILY_HANDOVER_LOG,
    SharedSourceCacheService,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TEMP_ROOT = PROJECT_ROOT / ".tmp_runtime_tests" / "day_metric_history_cache_reuse"


@pytest.fixture
def work_dir() -> Path:
    root = TEMP_ROOT / uuid.uuid4().hex
    root.mkdir(parents=True, exist_ok=True)
    try:
        yield root
    finally:
        shutil.rmtree(root, ignore_errors=True)


def _build_runtime_config(*, role_mode: str, shared_root: Path) -> dict:
    return {
        "deployment": {"role_mode": role_mode},
        "shared_bridge": {
            "enabled": True,
            "root_dir": str(shared_root),
        },
        "internal_source_cache": {"enabled": True},
    }


def _write_handover_source_xlsx(path: Path, *, d_value: str = "市电总功率", e_value: float = 123.4) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["D4"] = d_value
    sheet["E4"] = e_value
    workbook.save(path)


def _write_empty_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "empty"
    workbook.save(path)


def _register_shared_handover_entry(
    *,
    store: SharedBridgeStore,
    shared_root: Path,
    building: str,
    duty_date: str,
    downloaded_at: str,
    e_value: float = 123.4,
) -> Path:
    source_path = (
        shared_root
        / "交接班日志源文件"
        / "202604"
        / f"{duty_date.replace('-', '')}--16"
        / f"{duty_date.replace('-', '')}--16--交接班日志源文件--{building}.xlsx"
    )
    _write_handover_source_xlsx(source_path, e_value=e_value)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building=building,
        bucket_kind="latest",
        bucket_key=f"{duty_date} 16",
        duty_date=duty_date,
        duty_shift="day",
        downloaded_at=downloaded_at,
        relative_path=source_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": building, "duty_date": duty_date, "duty_shift": "day"},
    )
    return source_path


def test_fill_day_metric_history_reuses_shared_handover_sources(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    _register_shared_handover_entry(
        store=store,
        shared_root=shared_root,
        building="A楼",
        duty_date="2026-04-05",
        downloaded_at="2026-04-05 12:00:00",
    )

    entries = service.fill_day_metric_history(
        selected_dates=["2026-04-05"],
        building_scope="single",
        building="A楼",
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert len(entries) == 1
    output_path = Path(entries[0]["file_path"])
    assert output_path.exists()
    assert entries[0]["bucket_kind"] == "date"
    assert entries[0]["bucket_key"] == "2026-04-05"
    assert entries[0]["metadata"]["alias_only"] is True
    assert entries[0]["metadata"]["canonical_relative_path"] == entries[0]["relative_path"]
    assert "20260405--16" in entries[0]["relative_path"]
    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-05",
        duty_date="2026-04-05",
        duty_shift="all",
        status="ready",
    )
    assert len(rows) == 1
    assert rows[0]["metadata"]["alias_only"] is True


def test_fill_day_metric_history_rejects_external_writer(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="external", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    _register_shared_handover_entry(
        store=store,
        shared_root=shared_root,
        building="A楼",
        duty_date="2026-04-05",
        downloaded_at="2026-04-05 12:00:00",
    )

    with pytest.raises(RuntimeError, match="仅允许内网端"):
        service.fill_day_metric_history(
            selected_dates=["2026-04-05"],
            building_scope="single",
            building="A楼",
            emit_log=lambda *_args, **_kwargs: None,
        )

    rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-05",
        duty_date="2026-04-05",
        duty_shift="all",
        status="ready",
    )
    assert rows == []


def test_fill_day_metric_history_raises_when_no_shared_handover_source(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    with pytest.raises(RuntimeError, match="缺少可复用的交接班源文件"):
        service.fill_day_metric_history(
            selected_dates=["2026-04-05"],
            building_scope="single",
            building="A楼",
            emit_log=lambda *_args, **_kwargs: None,
        )


def test_fill_day_metric_history_skips_empty_ready_entry_and_reuses_valid_latest(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    empty_path = shared_root / "source_cache" / "handover_log" / "date" / "202604" / "empty-A.xlsx"
    _write_empty_xlsx(empty_path)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-05",
        duty_date="2026-04-05",
        duty_shift="all",
        downloaded_at="2026-04-05 18:00:00",
        relative_path=empty_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-05", "duty_shift": "all"},
    )

    latest_valid = _register_shared_handover_entry(
        store=store,
        shared_root=shared_root,
        building="A楼",
        duty_date="2026-04-05",
        downloaded_at="2026-04-05 16:00:00",
    )

    entries = service.fill_day_metric_history(
        selected_dates=["2026-04-05"],
        building_scope="single",
        building="A楼",
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert len(entries) == 1
    output_path = Path(entries[0]["file_path"])
    assert output_path.exists()
    workbook = openpyxl.load_workbook(output_path, data_only=True)
    try:
        assert workbook.active["D4"].value == "市电总功率"
        assert workbook.active["E4"].value == 123.4
    finally:
        workbook.close()
    assert output_path.read_bytes() == latest_valid.read_bytes()
    assert entries[0]["bucket_kind"] == "date"
    assert entries[0]["metadata"]["alias_only"] is True


def test_fill_day_metric_history_ignores_unindexed_history_file_when_alias_exists(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    cached_path = shared_root / "source_cache" / "handover_log" / "date" / "202604" / "cached-A.xlsx"
    _write_handover_source_xlsx(cached_path, e_value=111.1)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-11",
        duty_date="2026-04-11",
        duty_shift="all",
        downloaded_at="2026-04-11 20:00:00",
        relative_path=cached_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-11", "duty_shift": "all"},
    )

    history_path = (
        shared_root
        / "交接班日志源文件"
        / "202604"
        / "20260411--11"
        / "20260411--11--交接班日志源文件--A楼.xlsx"
    )
    _write_handover_source_xlsx(history_path, e_value=222.2)

    entries = service.fill_day_metric_history(
        selected_dates=["2026-04-11"],
        building_scope="single",
        building="A楼",
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert len(entries) == 1
    assert entries[0]["metadata"]["resolution_source"] == "cache_date"
    workbook = openpyxl.load_workbook(Path(entries[0]["file_path"]), data_only=True)
    try:
        assert workbook.active["E4"].value == 111.1
    finally:
        workbook.close()


def test_fill_day_metric_history_uses_indexed_cache_without_directory_lookup(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    cached_path = shared_root / "source_cache" / "handover_log" / "date" / "202604" / "cached-A.xlsx"
    _write_handover_source_xlsx(cached_path, e_value=345.6)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-12",
        duty_date="2026-04-12",
        duty_shift="all",
        downloaded_at="2026-04-12 20:00:00",
        relative_path=cached_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-12", "duty_shift": "all"},
    )

    entries = service.fill_day_metric_history(
        selected_dates=["2026-04-12"],
        building_scope="single",
        building="A楼",
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert len(entries) == 1
    assert entries[0]["metadata"]["resolution_source"] == "cache_date"
    workbook = openpyxl.load_workbook(Path(entries[0]["file_path"]), data_only=True)
    try:
        assert workbook.active["E4"].value == 345.6
    finally:
        workbook.close()


def test_repair_day_metric_ready_entries_downgrades_empty_ready_rows(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    empty_path = shared_root / "source_cache" / "handover_log" / "date" / "202604" / "empty-A.xlsx"
    _write_empty_xlsx(empty_path)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-05",
        duty_date="2026-04-05",
        duty_shift="all",
        downloaded_at="2026-04-05 18:00:00",
        relative_path=empty_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-05", "duty_shift": "all"},
    )
    _register_shared_handover_entry(
        store=store,
        shared_root=shared_root,
        building="A楼",
        duty_date="2026-04-05",
        downloaded_at="2026-04-05 16:00:00",
    )

    summary = service.repair_day_metric_ready_entries(
        selected_dates=["2026-04-05"],
        buildings=["A楼"],
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert summary["scanned"] >= 2
    assert summary["downgraded"] >= 1
    failed_rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-05",
        duty_date="2026-04-05",
        duty_shift="all",
        status="failed",
    )
    assert len(failed_rows) == 1


def test_sweep_invalid_ready_entries_only_scans_latest_and_recent_window(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    now = datetime.now()
    latest_date = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    recent_date = (now - timedelta(days=2)).strftime("%Y-%m-%d")
    old_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    latest_bucket = f"{latest_date} 16"

    latest_empty = shared_root / "source_cache" / "handover_log" / "latest" / "20260411-16" / "A-empty.xlsx"
    _write_empty_xlsx(latest_empty)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key=latest_bucket,
        duty_date=latest_date,
        duty_shift="day",
        downloaded_at=f"{latest_date} 16:00:00",
        relative_path=latest_empty.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": latest_date, "duty_shift": "day"},
    )

    recent_valid = _register_shared_handover_entry(
        store=store,
        shared_root=shared_root,
        building="B楼",
        duty_date=recent_date,
        downloaded_at=f"{recent_date} 12:00:00",
    )
    assert recent_valid.exists()

    old_empty = shared_root / "source_cache" / "handover_log" / "date" / "202603" / "old-empty.xlsx"
    _write_empty_xlsx(old_empty)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="C楼",
        bucket_kind="date",
        bucket_key=old_date,
        duty_date=old_date,
        duty_shift="all",
        downloaded_at=f"{old_date} 09:00:00",
        relative_path=old_empty.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "C楼", "duty_date": old_date, "duty_shift": "all"},
    )

    summary = service.sweep_invalid_ready_entries(
        lookback_days=7,
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert summary["scanned"] >= 2
    assert summary["downgraded"] >= 1
    failed_latest = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key=latest_bucket,
        status="failed",
    )
    assert len(failed_latest) == 1
    assert failed_latest[0]["metadata"]["validated_by"] == "background_sweep"
    old_ready = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="C楼",
        bucket_kind="date",
        bucket_key=old_date,
        duty_date=old_date,
        duty_shift="all",
        status="ready",
    )
    assert len(old_ready) == 1


def test_background_sweep_candidates_only_include_recent_three_hours(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    now = datetime.now()
    recent_time = (now - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
    old_time = (now - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M:%S")
    recent_bucket = (now - timedelta(hours=1)).strftime("%Y-%m-%d %H")
    old_bucket = (now - timedelta(hours=5)).strftime("%Y-%m-%d %H")

    recent_file = shared_root / "recent.xlsx"
    old_file = shared_root / "old.xlsx"
    _write_handover_source_xlsx(recent_file)
    _write_handover_source_xlsx(old_file)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key=recent_bucket,
        duty_date=now.strftime("%Y-%m-%d"),
        duty_shift="day",
        downloaded_at=recent_time,
        relative_path=recent_file.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼"},
    )
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="B楼",
        bucket_kind="latest",
        bucket_key=old_bucket,
        duty_date=(now - timedelta(days=1)).strftime("%Y-%m-%d"),
        duty_shift="day",
        downloaded_at=old_time,
        relative_path=old_file.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "B楼"},
    )
    with store.connect() as conn:
        conn.execute(
            """
            UPDATE source_cache_entries
            SET downloaded_at=?, updated_at=?, created_at=?
            WHERE source_family=? AND building=?
            """,
            (old_time, old_time, old_time, FAMILY_HANDOVER_LOG, "B楼"),
        )

    candidates = service.list_background_sweep_candidates(recent_hours=3)

    buildings = {item["building"] for item in candidates}
    assert buildings == {"A楼"}


def test_sweep_invalid_ready_entries_pauses_and_returns_next_index(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )
    now = datetime.now()
    recent_time = now.strftime("%Y-%m-%d %H:%M:%S")
    recent_bucket = now.strftime("%Y-%m-%d %H")
    source_a = shared_root / "A.xlsx"
    source_b = shared_root / "B.xlsx"
    _write_handover_source_xlsx(source_a)
    _write_handover_source_xlsx(source_b)
    for building, path in (("A楼", source_a), ("B楼", source_b)):
        store.upsert_source_cache_entry(
            source_family=FAMILY_HANDOVER_LOG,
            building=building,
            bucket_kind="latest",
            bucket_key=recent_bucket,
            duty_date=now.strftime("%Y-%m-%d"),
            duty_shift="day",
            downloaded_at=recent_time,
            relative_path=path.relative_to(shared_root).as_posix(),
            status="ready",
            metadata={"family": FAMILY_HANDOVER_LOG, "building": building},
        )
    candidates = service.list_background_sweep_candidates(recent_hours=3)
    calls = {"count": 0}

    def _pause_after_first() -> bool:
        calls["count"] += 1
        return calls["count"] > 1

    summary = service.sweep_invalid_ready_entries(
        recent_hours=3,
        emit_log=lambda *_args, **_kwargs: None,
        candidates=candidates,
        should_pause=_pause_after_first,
    )

    assert summary["paused"] is True
    assert summary["status"] == "deferred"
    assert summary["next_index"] == 1
    assert summary["scanned"] == 1


def test_sweep_invalid_ready_entries_downgrades_missing_ready_entry_to_failed(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key="2026-04-11 16",
        duty_date="2026-04-11",
        duty_shift="day",
        downloaded_at="2026-04-12 09:00:00",
        relative_path="source_cache/handover_log/latest/20260411-16/missing-A.xlsx",
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-11", "duty_shift": "day"},
    )

    summary = service.sweep_invalid_ready_entries(
        lookback_days=7,
        emit_log=lambda *_args, **_kwargs: None,
    )

    assert summary["scanned"] >= 1
    assert summary["downgraded"] >= 1
    assert store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key="2026-04-11 16",
        status="ready",
    ) == []
    failed_rows = store.list_source_cache_entries(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="latest",
        bucket_key="2026-04-11 16",
        status="failed",
    )
    assert len(failed_rows) == 1


def test_fill_day_metric_history_does_not_scan_real_history_folder_when_store_missing_valid_entry(work_dir: Path) -> None:
    shared_root = work_dir / "shared"
    store = SharedBridgeStore(shared_root)
    store.ensure_ready()
    service = SharedSourceCacheService(
        runtime_config=_build_runtime_config(role_mode="internal", shared_root=shared_root),
        store=store,
        emit_log=lambda *_args, **_kwargs: None,
    )

    empty_path = shared_root / "source_cache" / "handover_log" / "date" / "202604" / "empty-A.xlsx"
    _write_empty_xlsx(empty_path)
    store.upsert_source_cache_entry(
        source_family=FAMILY_HANDOVER_LOG,
        building="A楼",
        bucket_kind="date",
        bucket_key="2026-04-11",
        duty_date="2026-04-11",
        duty_shift="all",
        downloaded_at="2026-04-11 23:00:00",
        relative_path=empty_path.relative_to(shared_root).as_posix(),
        status="ready",
        metadata={"family": FAMILY_HANDOVER_LOG, "building": "A楼", "duty_date": "2026-04-11", "duty_shift": "all"},
    )

    history_path = (
        shared_root
        / "交接班日志源文件"
        / "202604"
        / "20260411--11"
        / "20260411--11--交接班日志源文件--A楼.xlsx"
    )
    _write_handover_source_xlsx(history_path)

    with pytest.raises(RuntimeError, match="缺少可复用的交接班源文件"):
        service.fill_day_metric_history(
            selected_dates=["2026-04-11"],
            building_scope="single",
            building="A楼",
            emit_log=lambda *_args, **_kwargs: None,
        )
