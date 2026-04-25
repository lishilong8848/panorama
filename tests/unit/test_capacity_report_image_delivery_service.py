from __future__ import annotations

from copy import copy
import os
from pathlib import Path

import pytest
from openpyxl import Workbook
from PIL import Image

from handover_log_module.service import capacity_report_image_delivery_service as module


def _write_capacity_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本班组"
    sheet.print_area = "A1:D4"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "A楼交接班容量报表"
    font = copy(sheet["A1"].font)
    font.bold = True
    font.sz = 14
    sheet["A1"].font = font
    sheet["A2"] = "项目"
    sheet["B2"] = "数值"
    sheet["A3"] = "机柜"
    sheet["B3"] = 12
    sheet["C3"] = "备注"
    sheet["D3"] = "正常"
    workbook.save(path)
    workbook.close()


def _disable_excel_copy_picture(monkeypatch) -> None:
    monkeypatch.setattr(
        module.CapacityReportImageRenderer,
        "_render_with_excel_copy_picture",
        lambda self, *, source_path, output_path, emit_log=None: False,
    )


def _fake_excel_copy_picture(monkeypatch) -> None:
    def _fake_copy(self, *, source_path: Path, output_path: Path, emit_log=None) -> bool:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        image = Image.new("RGB", (240, 120), "white")
        image.save(output_path, format="PNG")
        if emit_log:
            emit_log(f"fake excel copy: {source_path}")
        return True

    monkeypatch.setattr(module.CapacityReportImageRenderer, "_render_with_excel_copy_picture", _fake_copy)


def test_capacity_report_renderer_outputs_valid_png(tmp_path: Path, monkeypatch) -> None:
    _disable_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    renderer = module.CapacityReportImageRenderer(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    image_path = renderer.render_to_image(
        source_file=str(source),
        building="A楼",
        duty_date="2026-04-24",
        duty_shift="day",
        session_id="A楼|2026-04-24|day",
    )

    assert image_path.exists()
    with Image.open(image_path) as image:
        assert image.format == "PNG"
        assert image.width > 100
        assert image.height > 60


def test_capacity_report_renderer_prefers_excel_copy_picture(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    renderer = module.CapacityReportImageRenderer(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )
    calls = []

    def _fake_excel_copy(*, source_path: Path, output_path: Path, emit_log=None) -> bool:
        calls.append(source_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        image = Image.new("RGB", (220, 120), "white")
        image.save(output_path, format="PNG")
        return True

    monkeypatch.setattr(renderer, "_render_with_excel_copy_picture", _fake_excel_copy)

    image_path = renderer.render_to_image(
        source_file=str(source),
        building="A楼",
        duty_date="2026-04-24",
        duty_shift="day",
        session_id="A楼|2026-04-24|day",
    )

    assert calls == [source]
    with Image.open(image_path) as image:
        assert image.size == (220, 120)


def test_capacity_report_renderer_calculates_formula_without_excel_dependency(tmp_path: Path, monkeypatch) -> None:
    _disable_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity_formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本班组"
    sheet.print_area = "A1:C3"
    sheet["A1"] = "公式项"
    sheet["B1"] = "=1+1"
    sheet["A2"] = 10
    sheet["A3"] = 20
    sheet["B2"] = "=SUM(A2:A3)"
    sheet["C2"] = "=B2/100"
    sheet["C2"].number_format = "0.00%"
    workbook.save(source)
    workbook.close()
    renderer = module.CapacityReportImageRenderer(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )
    formula_workbook = module.load_workbook(source, data_only=False)
    try:
        values = renderer._formula_display_values(formula_workbook["本班组"])
    finally:
        formula_workbook.close()

    image_path = renderer.render_to_image(
        source_file=str(source),
        building="A楼",
        duty_date="2026-04-24",
        duty_shift="day",
        session_id="A楼|2026-04-24|day",
    )

    assert values["B1"] == "2"
    assert values["B2"] == "30"
    assert values["C2"] == "30.00%"
    with Image.open(image_path) as image:
        assert image.format == "PNG"
        assert image.width > 100


def test_capacity_image_delivery_uploads_once_and_sends_to_all_recipients(tmp_path: Path, monkeypatch) -> None:
    _fake_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    sent_text = []
    sent_image = []
    uploaded = []
    updated_states = []
    logs = []

    class _FakeClient:
        def upload_image(self, image_path: str):
            uploaded.append(image_path)
            return {"image_key": "img-key"}

        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            sent_text.append({"receive_id": receive_id, "receive_id_type": receive_id_type, "text": text})
            return {"message_id": f"text-{len(sent_text)}"}

        def send_image_message(self, *, receive_id: str, receive_id_type: str, image_key: str):
            sent_image.append({"receive_id": receive_id, "receive_id_type": receive_id_type, "image_key": image_key})
            return {"message_id": f"image-{len(sent_image)}"}

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1", "note": "甲"}, {"open_id": "ou_2", "note": "乙"}]

        def _build_feishu_client(self):
            return _FakeClient()

        @staticmethod
        def _resolve_effective_receive_id_type(_recipient_id, _configured_receive_id_type="open_id"):
            return "open_id"

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": "甲"}, {"open_id": "ou_2", "note": "乙"}],
                "raw_count": 3,
                "enabled_count": 2,
                "disabled_count": 1,
                "invalid_count": 0,
                "open_ids": ["ou_1", "ou_2"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "摘要"

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            updated_states.append({"session_id": session_id, "delivery": dict(capacity_image_delivery)})
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            updated_states.append({"session_id": session_id, "review_delivery": dict(review_link_delivery)})
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready"},
        },
        building="A楼",
        emit_log=logs.append,
    )

    assert result["status"] == "success"
    assert len(uploaded) == 1
    assert Path(uploaded[0]).exists()
    assert [item["receive_id"] for item in sent_text] == ["ou_1", "ou_2"]
    assert [item["text"] for item in sent_text] == ["摘要", "摘要"]
    assert sent_image == [
        {"receive_id": "ou_1", "receive_id_type": "open_id", "image_key": "img-key"},
        {"receive_id": "ou_2", "receive_id_type": "open_id", "image_key": "img-key"},
    ]
    assert updated_states[0]["delivery"]["status"] == "sending"
    capacity_updates = [item["delivery"] for item in updated_states if "delivery" in item]
    review_updates = [item["review_delivery"] for item in updated_states if "review_delivery" in item]
    assert capacity_updates[-1]["status"] == "success"
    assert review_updates[-1]["status"] == "success"
    assert result["successful_recipients"] == ["ou_1", "ou_2"]
    assert result["failed_recipients"] == []
    assert any("开始生成容量表图片" in line for line in logs)
    assert any("开始上传飞书图片" in line for line in logs)
    assert any("本次将发送审核文本内容如下" in line and "摘要" in line for line in logs)
    assert any("准备发送审核文本" in line and "open_id=ou_1" in line for line in logs)
    assert any("容量图片发送成功" in line and "open_id=ou_2" in line for line in logs)


def test_capacity_image_delivery_reuses_image_when_capacity_signature_unchanged(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    excel_calls = []
    uploaded = []

    def _fake_copy(self, *, source_path: Path, output_path: Path, emit_log=None) -> bool:
        excel_calls.append(str(source_path))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (240, 120), "white").save(output_path, format="PNG")
        return True

    monkeypatch.setattr(module.CapacityReportImageRenderer, "_render_with_excel_copy_picture", _fake_copy)

    class _FakeClient:
        def upload_image(self, image_path: str):
            uploaded.append(image_path)
            return {"image_key": "img-key"}

        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            return {"message_id": f"text-{receive_id}"}

        def send_image_message(self, *, receive_id: str, receive_id_type: str, image_key: str):
            return {"message_id": f"image-{receive_id}"}

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1", "note": "甲"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": "甲"}],
                "raw_count": 1,
                "enabled_count": 1,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _resolve_effective_receive_id_type(_recipient_id, _configured_receive_id_type="open_id"):
            return "open_id"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            return _FakeClient()

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )
    session = {
        "session_id": "session-a",
        "building": "A楼",
        "duty_date": "2026-04-24",
        "duty_shift": "day",
        "revision": 1,
        "capacity_output_file": str(source),
        "capacity_sync": {"status": "ready", "input_signature": "cap-v1", "updated_at": "2026-04-24 08:00:00"},
    }

    first = service.send_for_session(session, building="A楼", emit_log=lambda _line: None)
    second = service.send_for_session({**session, "revision": 2}, building="A楼", emit_log=lambda _line: None)

    assert first["status"] == "success"
    assert second["status"] == "success"
    assert len(uploaded) == 2
    assert excel_calls == [str(source)]
    assert first["capacity_image_delivery"]["cache_hit"] is False
    assert second["capacity_image_delivery"]["cache_hit"] is True
    assert first["capacity_image_delivery"]["image_signature"] == second["capacity_image_delivery"]["image_signature"]


def test_capacity_image_delivery_rerenders_when_capacity_signature_or_file_changes(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    excel_calls = []

    def _fake_copy(self, *, source_path: Path, output_path: Path, emit_log=None) -> bool:
        excel_calls.append(str(source_path))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (240, 120), "white").save(output_path, format="PNG")
        return True

    monkeypatch.setattr(module.CapacityReportImageRenderer, "_render_with_excel_copy_picture", _fake_copy)

    class _FakeClient:
        def upload_image(self, _image_path: str):
            return {"image_key": "img-key"}

        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            return {"message_id": f"text-{receive_id}"}

        def send_image_message(self, *, receive_id: str, receive_id_type: str, image_key: str):
            return {"message_id": f"image-{receive_id}"}

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": ""}],
                "raw_count": 1,
                "enabled_count": 1,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _resolve_effective_receive_id_type(_recipient_id, _configured_receive_id_type="open_id"):
            return "open_id"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            return _FakeClient()

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )
    base_session = {
        "session_id": "session-a",
        "building": "A楼",
        "duty_date": "2026-04-24",
        "duty_shift": "day",
        "capacity_output_file": str(source),
        "capacity_sync": {"status": "ready", "input_signature": "cap-v1", "updated_at": "2026-04-24 08:00:00"},
    }

    first = service.send_for_session(base_session, building="A楼", emit_log=lambda _line: None)
    changed_capacity = service.send_for_session(
        {
            **base_session,
            "capacity_sync": {"status": "ready", "input_signature": "cap-v2", "updated_at": "2026-04-24 08:05:00"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )
    next_mtime = source.stat().st_mtime + 10
    os.utime(source, (next_mtime, next_mtime))
    changed_file = service.send_for_session(
        {
            **base_session,
            "capacity_sync": {"status": "ready", "input_signature": "cap-v1", "updated_at": "2026-04-24 08:00:00"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )
    missing_signature = service.send_for_session(
        {**base_session, "capacity_sync": {"status": "ready"}},
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert first["capacity_image_delivery"]["cache_hit"] is False
    assert changed_capacity["capacity_image_delivery"]["cache_hit"] is False
    assert changed_file["capacity_image_delivery"]["cache_hit"] is False
    assert missing_signature["capacity_image_delivery"]["cache_hit"] is False
    assert len(excel_calls) == 4


def test_capacity_image_delivery_lock_timeout_fails_without_upload_or_send(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)

    class _NeverAcquireLock:
        def acquire(self, timeout=None):
            return False

        def release(self):
            raise AssertionError("release should not be called when acquire failed")

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": ""}],
                "raw_count": 1,
                "enabled_count": 1,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            raise AssertionError("Excel截图锁超时时不应构建飞书客户端")

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "_CAPACITY_IMAGE_EXCEL_COPY_LOCK", _NeverAcquireLock())
    monkeypatch.setattr(module, "_CAPACITY_IMAGE_EXCEL_LOCK_TIMEOUT_SEC", 0.01)
    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready", "input_signature": "cap-v1", "updated_at": "2026-04-24 08:00:00"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert result["status"] == "failed"
    assert "容量图片截图繁忙，请稍后重试" in result["error"]
    assert result["failed_recipients"] == []


def test_capacity_image_delivery_rejects_duplicate_running_session(tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}]

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def get_session_by_id(self, _session_id: str):
            return {
                "session_id": "session-a",
                "building": "A楼",
                "capacity_output_file": str(source),
                "capacity_sync": {"status": "ready"},
                "capacity_image_delivery": {"status": "sending"},
            }

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService({"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}})

    with pytest.raises(ValueError, match="正在发送中"):
        service.begin_delivery(
            {
                "session_id": "session-a",
                "building": "A楼",
                "capacity_output_file": str(source),
                "capacity_sync": {"status": "ready"},
            },
            building="A楼",
        )


def test_capacity_image_delivery_marks_failed_when_upload_fails(tmp_path: Path, monkeypatch) -> None:
    _fake_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    updated_states = []

    class _FakeClient:
        def upload_image(self, _image_path: str):
            raise RuntimeError("upload failed")

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": ""}],
                "raw_count": 1,
                "enabled_count": 1,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "摘要"

        def _build_feishu_client(self):
            return _FakeClient()

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            updated_states.append({"session_id": session_id, "delivery": dict(capacity_image_delivery)})
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            updated_states.append({"session_id": session_id, "review_delivery": dict(review_link_delivery)})
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert result["ok"] is False
    assert result["status"] == "failed"
    assert "upload failed" in result["error"]
    assert updated_states[0]["delivery"]["status"] == "sending"
    capacity_updates = [item["delivery"] for item in updated_states if "delivery" in item]
    assert capacity_updates[-1]["status"] == "failed"
    assert "upload failed" in capacity_updates[-1]["error"]


def test_capacity_image_delivery_fails_when_excel_screenshot_fails_without_upload_or_send(tmp_path: Path, monkeypatch) -> None:
    _disable_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)
    updated_states = []

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": ""}],
                "raw_count": 1,
                "enabled_count": 1,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            raise AssertionError("Excel截图失败时不应构建飞书客户端")

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            updated_states.append({"session_id": session_id, "delivery": dict(capacity_image_delivery)})
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            updated_states.append({"session_id": session_id, "review_delivery": dict(review_link_delivery)})
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert result["ok"] is False
    assert result["status"] == "failed"
    assert "Excel截图失败" in result["error"]
    capacity_updates = [item["delivery"] for item in updated_states if "delivery" in item]
    assert capacity_updates[-1]["status"] == "failed"


def test_capacity_image_delivery_text_failure_returns_failed_recipient_step(tmp_path: Path, monkeypatch) -> None:
    _fake_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)

    class _FakeClient:
        def upload_image(self, _image_path: str):
            return {"image_key": "img-key"}

        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            if receive_id == "ou_1":
                raise RuntimeError("text failed")
            return {"message_id": f"text-{receive_id}"}

        def send_image_message(self, *, receive_id: str, receive_id_type: str, image_key: str):
            return {"message_id": f"image-{receive_id}"}

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}, {"open_id": "ou_2"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": "甲"}, {"open_id": "ou_2", "note": "乙"}],
                "raw_count": 2,
                "enabled_count": 2,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1", "ou_2"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _resolve_effective_receive_id_type(_recipient_id, _configured_receive_id_type="open_id"):
            return "open_id"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            return _FakeClient()

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert result["status"] == "failed"
    assert result["successful_recipients"] == ["ou_2"]
    assert result["failed_recipients"] == [{"open_id": "ou_1", "note": "甲", "step": "text", "error": "text failed"}]


def test_capacity_image_delivery_image_failure_returns_failed_recipient_step(tmp_path: Path, monkeypatch) -> None:
    _fake_excel_copy_picture(monkeypatch)
    source = tmp_path / "capacity.xlsx"
    _write_capacity_workbook(source)

    class _FakeClient:
        def upload_image(self, _image_path: str):
            return {"image_key": "img-key"}

        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            return {"message_id": f"text-{receive_id}"}

        def send_image_message(self, *, receive_id: str, receive_id_type: str, image_key: str):
            if receive_id == "ou_1":
                raise RuntimeError("image failed")
            return {"message_id": f"image-{receive_id}"}

    class _FakeLinkService:
        def __init__(self, *_args, **_kwargs):
            pass

        def _recipients_for_building(self, _building):
            return [{"open_id": "ou_1"}, {"open_id": "ou_2"}]

        @staticmethod
        def _recipient_snapshot_for_building(_building):
            return {
                "recipients": [{"open_id": "ou_1", "note": "甲"}, {"open_id": "ou_2", "note": "乙"}],
                "raw_count": 2,
                "enabled_count": 2,
                "disabled_count": 0,
                "invalid_count": 0,
                "open_ids": ["ou_1", "ou_2"],
            }

        @staticmethod
        def _review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _manual_test_review_url_for_building(_snapshot, _building):
            return "http://example.com/review/a"

        @staticmethod
        def _resolve_effective_receive_id_type(_recipient_id, _configured_receive_id_type="open_id"):
            return "open_id"

        class _summary_message_service:
            @staticmethod
            def build_for_session(_session, emit_log):
                return "交接班日志全文"

        def _build_feishu_client(self):
            return _FakeClient()

    class _FakeReviewSessionService:
        def __init__(self, *_args, **_kwargs):
            pass

        def update_capacity_image_delivery(self, *, session_id: str, capacity_image_delivery):
            return {"session_id": session_id, "capacity_image_delivery": dict(capacity_image_delivery)}

        def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
            return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}

    monkeypatch.setattr(module, "ReviewLinkDeliveryService", _FakeLinkService)
    monkeypatch.setattr(module, "ReviewSessionService", _FakeReviewSessionService)
    service = module.CapacityReportImageDeliveryService(
        {"_global_paths": {"runtime_state_root": str(tmp_path / "runtime")}, "capacity_report": {"template": {"sheet_name": "本班组"}}}
    )

    result = service.send_for_session(
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-24",
            "duty_shift": "day",
            "capacity_output_file": str(source),
            "capacity_sync": {"status": "ready"},
        },
        building="A楼",
        emit_log=lambda _line: None,
    )

    assert result["status"] == "failed"
    assert result["successful_recipients"] == ["ou_2"]
    assert result["failed_recipients"] == [{"open_id": "ou_1", "note": "甲", "step": "image", "error": "image failed"}]
