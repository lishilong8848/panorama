from __future__ import annotations

import pytest
from openpyxl import Workbook

from handover_log_module.core.models import RawRow
from handover_log_module.service import review_link_delivery_service as review_module
from handover_log_module.service import handover_summary_message_service as summary_module


class _FakeReviewSessionService:
    sessions = []
    updated_states = {}

    def __init__(self, _handover_cfg):
        pass

    def list_batch_sessions(self, _batch_key):
        return [dict(item) for item in self.sessions]

    def list_sessions(self):
        return [dict(item) for item in self.sessions]

    def update_review_link_delivery(self, *, session_id: str, review_link_delivery):
        self.updated_states[session_id] = dict(review_link_delivery)
        return {"session_id": session_id, "review_link_delivery": dict(review_link_delivery)}


def _make_service(monkeypatch, *, recipients_by_building, review_links, review_base_url_effective=""):
    _FakeReviewSessionService.updated_states = {}
    _FakeReviewSessionService.sessions = [
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-10",
            "duty_shift": "night",
            "review_link_delivery": {},
        }
    ]
    monkeypatch.setattr(review_module, "ReviewSessionService", _FakeReviewSessionService)
    monkeypatch.setattr(
        review_module,
        "materialize_review_access_snapshot",
        lambda _cfg: {
            "review_links": list(review_links),
            "review_base_url_effective": review_base_url_effective,
        },
    )
    monkeypatch.setattr(
        review_module,
        "load_review_access_state",
        lambda _cfg: {},
    )
    return review_module.ReviewLinkDeliveryService(
        {
            "review_ui": {
                "review_link_recipients_by_building": recipients_by_building,
            },
            "_global_feishu": {
                "app_id": "app-id",
                "app_secret": "app-secret",
            },
        }
    )


def test_send_for_batch_manual_unconfigured_raises(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )
    logs = []

    with pytest.raises(RuntimeError, match="当前楼未配置审核链接接收人"):
        service.send_for_batch(
            batch_key="2026-04-10|night",
            building="A楼",
            source="manual",
            emit_log=logs.append,
        )

    assert _FakeReviewSessionService.updated_states["session-a"]["status"] == "unconfigured"
    assert any("跳过发送" in line and "未配置审核链接接收人" in line for line in logs)
    assert any("批次完成" in line for line in logs)


def test_send_for_batch_manual_pending_access_raises(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
    )
    logs = []

    with pytest.raises(RuntimeError, match="审核访问地址尚未就绪"):
        service.send_for_batch(
            batch_key="2026-04-10|night",
            building="A楼",
            source="manual",
            emit_log=logs.append,
        )

    assert _FakeReviewSessionService.updated_states["session-a"]["status"] == "pending_access"
    assert any("跳过发送" in line and "审核访问地址尚未就绪" in line for line in logs)
    assert any("批次完成" in line for line in logs)


def test_send_for_batch_manual_success_uses_open_id(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "本人"}]},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )
    logs = []
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_for_batch(
        batch_key="2026-04-10|night",
        building="A楼",
        source="manual",
        emit_log=logs.append,
    )

    assert result["results"][0]["delivery"]["status"] == "success"
    assert calls == [
        {
            "receive_id": "ou_abc",
            "receive_id_type": "open_id",
            "text": (
                "这是一条交接班审核访问链接，请在办公电脑的浏览器中打开。\n"
                "楼栋：A楼\n"
                "日期：2026-04-10\n"
                "班次：夜班\n"
                "审核链接：http://example.com/review/A"
            ),
        }
    ]
    assert _FakeReviewSessionService.updated_states["session-a"]["status"] == "success"
    assert any("发送成功" in line and "receive_id_type=open_id" in line for line in logs)
    assert any("批次完成" in line for line in logs)


def test_send_for_batch_appends_handover_summary(monkeypatch, tmp_path):
    output_file = tmp_path / "handover.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "交接班日志"
    ws["A1"] = "EA118-E栋世纪互联 夜班"
    ws["C3"] = "张三、李四"
    ws["G3"] = "王五"
    ws["B6"] = 1.43
    ws["D6"] = 2480
    ws["F6"] = 1734.06
    ws["B13"] = 1272
    ws["D13"] = 330
    ws["A20"] = "日常工作"
    ws["A21"] = "序号"
    ws["B21"] = "工作内容"
    ws["C21"] = "完成情况"
    ws["B22"] = "值班巡检"
    ws["B23"] = "HVDC电池均充维护"
    ws["C23"] = "未完成"
    ws["A25"] = "新事件处理"
    ws["A26"] = "序号"
    ws["B27"] = "不应出现在摘要中"
    workbook.save(output_file)
    workbook.close()

    capacity_file = tmp_path / "capacity.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws["D23"] = "1号制冷单元→预冷"
    ws["D33"] = "2号制冷单元→预冷"
    ws["Q23"] = "4号制冷单元→制冷"
    ws["Q33"] = "5号制冷单元→板换"
    ws["AC27"] = 27.3
    ws["AC28"] = 27.9
    workbook.save(capacity_file)
    workbook.close()

    service = _make_service(
        monkeypatch,
        recipients_by_building={"E楼": [{"open_id": "ou_e", "note": "E楼"}]},
        review_links=[{"building": "E楼", "url": "http://example.com/review/E"}],
    )
    _FakeReviewSessionService.sessions = [
        {
            "session_id": "session-e",
            "building": "E楼",
            "duty_date": "2026-04-24",
            "duty_shift": "night",
            "output_file": str(output_file),
            "capacity_output_file": str(capacity_file),
            "review_link_delivery": {},
        }
    ]

    monkeypatch.setattr(
        summary_module,
        "require_feishu_auth_settings",
        lambda _cfg, config_path=None: {"app_id": "app-id", "app_secret": "secret"},
    )

    phone_by_name = {"张三": "111", "李四": "222", "王五": "333"}

    class _FakeBitableClient:
        def __init__(self, *args, **kwargs):
            pass

        def list_records(self, *, filter_formula: str, **kwargs):
            match = None
            for name in phone_by_name:
                if f'"{name}"' in filter_formula:
                    match = name
                    break
            phone = phone_by_name.get(match or "")
            return [{"fields": {"联系方式": phone}}] if phone else []

    monkeypatch.setattr(summary_module, "FeishuBitableClient", _FakeBitableClient)

    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append({"receive_id": receive_id, "receive_id_type": receive_id_type, "text": text})
            return {"message_id": "msg-e"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_for_batch(
        batch_key="2026-04-24|night",
        building="E楼",
        source="manual",
        emit_log=lambda _msg: None,
    )

    assert result["results"][0]["delivery"]["status"] == "success"
    message = calls[0]["text"]
    assert "审核链接：http://example.com/review/E" in message
    assert "【EA118-E栋世纪互联 夜班】" not in message
    assert "【交班人员】张三、李四" not in message
    assert "【交接内容】" not in message
    assert "E-144、E-120变电所" not in message
    assert "【本班完成工作】" in message
    assert "值班巡检" in message
    assert "HVDC电池均充维护，未完成" in message
    assert "不应出现在摘要中" not in message
    assert "【重点关注项】" in message
    assert "E楼IT负载功率:1734.06KW" in message


def test_summary_chiller_zone_line_uses_source_level_and_tank_values():
    service = summary_module.HandoverSummaryMessageService({})
    source = {
        "building": "A楼",
        "running_units": {"west": [{"unit": 1, "mode_text": "预冷"}, {"unit": 2, "mode_text": "预冷"}]},
        "rows": [
            RawRow(1, "西区", "西区一号冷机", "冷却塔液位", 0.34, 0.34),
            RawRow(2, "西区", "西区二号冷机", "冷却塔液位", "0.36m", 0.36),
            RawRow(3, "西区蓄冷罐", "西区蓄冷罐", "蓄冷罐温度", 17.3, 17.3),
            RawRow(4, "西区蓄冷罐", "西区蓄冷罐", "蓄冷罐液位", 27.3, 27.3),
        ],
    }

    line = service._build_chiller_zone_line("A", "west", source["running_units"], source, {})

    assert "冷冻站A区3套制冷单元2用1备" in line
    assert "1#制冷单元预冷模式运行正常，1#冷却塔液位0.34m正常" in line
    assert "2#制冷单元预冷模式运行正常，2#冷却塔液位0.36m正常" in line
    assert "1#、2#二次泵运行正常" in line
    assert "蓄冷罐后备温度17.3℃正常、液位27.3m正常" in line


def test_send_for_session_summary_failure_does_not_block_message(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "本人"}]},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )
    calls = []

    class _BrokenSummary:
        def build_review_link_summary_for_session(self, *_args, **_kwargs):
            raise RuntimeError("summary broken")

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(text)
            return {"message_id": "msg-1"}

    service._summary_message_service = _BrokenSummary()
    service._build_feishu_client = lambda: _FakeClient()
    logs = []

    result = service.send_for_session(
        dict(_FakeReviewSessionService.sessions[0]),
        source="manual",
        emit_log=logs.append,
    )

    assert result["status"] == "success"
    assert calls == [
        (
            "这是一条交接班审核访问链接，请在办公电脑的浏览器中打开。\n"
            "楼栋：A楼\n"
            "日期：2026-04-10\n"
            "班次：夜班\n"
            "审核链接：http://example.com/review/A"
        )
    ]
    assert any("摘要" in line and "不阻断发送" in line for line in logs)


def test_send_for_batch_old_recipients_without_enabled_still_send(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_legacy", "note": "旧配置"}]},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-legacy"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_for_batch(
        batch_key="2026-04-10|night",
        building="A楼",
        source="manual",
        emit_log=lambda _msg: None,
    )

    assert result["results"][0]["delivery"]["status"] == "success"
    assert calls[0]["receive_id"] == "ou_legacy"


def test_send_for_batch_skips_disabled_recipients(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={
            "A楼": [
                {"open_id": "ou_enabled", "note": "启用", "enabled": True},
                {"open_id": "ou_disabled", "note": "停用", "enabled": False},
            ]
        },
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(receive_id)
            return {"message_id": "msg-mixed"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_for_batch(
        batch_key="2026-04-10|night",
        building="A楼",
        source="manual",
        emit_log=lambda _msg: None,
    )

    assert result["results"][0]["delivery"]["status"] == "success"
    assert calls == ["ou_enabled"]


def test_recipient_normalization_skips_string_false_enabled(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={
            "A楼": [
                {"open_id": "ou_enabled", "note": "启用", "enabled": "true"},
                {"open_id": "ou_disabled_bool", "note": "停用1", "enabled": False},
                {"open_id": "ou_disabled_text", "note": "停用2", "enabled": "false"},
                {"open_id": "ou_disabled_cn", "note": "停用3", "enabled": "停用"},
            ]
        },
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    snapshot = service._recipient_snapshot_for_building("A楼")

    assert snapshot["open_ids"] == ["ou_enabled"]
    assert snapshot["disabled_open_ids"] == ["ou_disabled_bool", "ou_disabled_text", "ou_disabled_cn"]
    assert snapshot["enabled_count"] == 1
    assert snapshot["disabled_count"] == 3


def test_send_for_session_mixed_recipient_failure_is_failed(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={
            "A楼": [
                {"open_id": "ou_ok", "note": "成功"},
                {"open_id": "ou_fail", "note": "失败"},
            ]
        },
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            if receive_id == "ou_fail":
                raise RuntimeError("send failed")
            return {"message_id": "msg-ok"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_for_session(
        dict(_FakeReviewSessionService.sessions[0]),
        source="manual",
        emit_log=lambda _msg: None,
    )

    assert result["status"] == "failed"
    assert result["error"] == "发送失败，详见收件人明细"
    assert result["successful_recipients"] == ["ou_ok"]
    assert result["failed_recipients"] == [
        {"open_id": "ou_fail", "note": "失败", "step": "text", "error": "send failed"}
    ]
    assert "partial_failed" not in {result["status"], _FakeReviewSessionService.updated_states["session-a"]["status"]}


def test_send_manual_test_mixed_recipient_failure_is_failed(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={
            "A楼": [
                {"open_id": "ou_ok", "note": "成功"},
                {"open_id": "ou_fail", "note": "失败"},
            ]
        },
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            if receive_id == "ou_fail":
                raise RuntimeError("test failed")
            return {"message_id": "msg-ok"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_manual_test(
        building="A楼",
        batch_key="2026-04-10|night",
        emit_log=lambda _msg: None,
    )

    assert result["status"] == "failed"
    assert result["error"] == "发送失败，详见收件人明细"
    assert result["successful_recipients"] == ["ou_ok"]
    assert result["failed_recipients"] == [
        {"open_id": "ou_fail", "note": "失败", "step": "text", "error": "test failed"}
    ]


def test_send_manual_test_all_recipients_disabled_raises(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_disabled", "note": "停用", "enabled": False}]},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    with pytest.raises(ValueError, match="当前楼审核链接接收人均未启用"):
        service.send_manual_test(
            building="A楼",
            batch_key="2026-04-10|night",
            emit_log=lambda _msg: None,
        )


def test_dispatch_pending_review_links_sends_pending_link_only(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "本人"}]},
        review_links=[],
        review_base_url_effective="http://192.168.224.157:18765",
    )
    _FakeReviewSessionService.sessions = [
        {
            "session_id": "session-a",
            "building": "A楼",
            "duty_date": "2026-04-10",
            "duty_shift": "night",
            "review_link_delivery": {"status": "pending_access", "auto_attempted": False},
        }
    ]
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(text)
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    results = service.dispatch_pending_review_links(emit_log=lambda _msg: None)

    assert results[0]["delivery"]["status"] == "success"
    assert calls == [
        (
            "这是一条交接班审核访问链接，请在办公电脑的浏览器中打开。\n"
            "楼栋：A楼\n"
            "日期：2026-04-10\n"
            "班次：夜班\n"
            "审核链接：http://192.168.224.157:18765/handover/review/a"
        )
    ]


def test_send_for_batch_auto_uses_effective_base_url_fallback(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "本人"}]},
        review_links=[],
        review_base_url_effective="http://192.168.224.157:18765",
    )
    logs = []
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    session = dict(_FakeReviewSessionService.sessions[0])
    result = service.send_for_session(
        session,
        source="auto",
        force=False,
        emit_log=logs.append,
    )

    assert result["status"] == "success"
    assert result["url"] == "http://192.168.224.157:18765/handover/review/a"
    assert calls == [
        {
            "receive_id": "ou_abc",
            "receive_id_type": "open_id",
            "text": (
                "这是一条交接班审核访问链接，请在办公电脑的浏览器中打开。\n"
                "楼栋：A楼\n"
                "日期：2026-04-10\n"
                "班次：夜班\n"
                "审核链接：http://192.168.224.157:18765/handover/review/a"
            ),
        }
    ]


def test_validate_manual_send_preflight_building_unconfigured(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    with pytest.raises(ValueError, match="当前楼未配置审核链接接收人"):
        service.validate_manual_send_preflight(
            batch_key="2026-04-10|night",
            building="A楼",
        )


def test_validate_manual_send_preflight_building_pending_access(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
    )

    result = service.validate_manual_send_preflight(
        batch_key="2026-04-10|night",
        building="A楼",
    )
    assert result["building"] == "A楼"
    assert result["session_count"] == 1


def test_validate_manual_send_preflight_building_ok(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[{"building": "A楼", "url": "http://example.com/review/A"}],
    )

    result = service.validate_manual_send_preflight(
        batch_key="2026-04-10|night",
        building="A楼",
    )
    assert result["session_count"] == 1
    assert result["building"] == "A楼"


def test_send_manual_test_without_review_url_still_sends(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
    )
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_manual_test(
        building="A楼",
        batch_key="2026-04-10|night",
        emit_log=lambda _msg: None,
    )

    assert result["status"] == "success"
    assert result["review_url"] == ""
    assert calls == [
        {
            "receive_id": "ou_abc",
            "receive_id_type": "open_id",
            "text": (
                "这是一条交接班审核链接测试消息，请在办公电脑的浏览器中打开。\n"
                "楼栋：A楼\n"
                "日期：2026-04-10\n"
                "班次：夜班\n"
                "审核链接：当前尚未生成，本次仅测试发送通道"
            ),
        }
    ]


def test_send_manual_test_uses_fallback_building_review_url(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
        review_base_url_effective="http://192.168.224.157:18765",
    )
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_manual_test(
        building="A楼",
        batch_key="2026-04-12|day",
        emit_log=lambda _msg: None,
    )

    assert result["status"] == "success"
    assert result["review_url"] == "http://192.168.224.157:18765/handover/review/a"
    assert calls == [
        {
            "receive_id": "ou_abc",
            "receive_id_type": "open_id",
            "text": (
                "这是一条交接班审核链接测试消息，请在办公电脑的浏览器中打开。\n"
                "楼栋：A楼\n"
                "日期：2026-04-12\n"
                "班次：白班\n"
                "审核链接：http://192.168.224.157:18765/handover/review/a"
            ),
        }
    ]


def test_send_manual_test_uses_configured_public_base_url_when_snapshot_missing(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
    )
    service.handover_cfg.setdefault("review_ui", {})["public_base_url"] = "http://192.168.224.157:18765"
    calls = []

    class _FakeClient:
        def send_text_message(self, *, receive_id: str, receive_id_type: str, text: str):
            calls.append(
                {
                    "receive_id": receive_id,
                    "receive_id_type": receive_id_type,
                    "text": text,
                }
            )
            return {"message_id": "msg-1"}

    service._build_feishu_client = lambda: _FakeClient()

    result = service.send_manual_test(
        building="A楼",
        batch_key="2026-04-12|day",
        emit_log=lambda _msg: None,
    )

    assert result["status"] == "success"
    assert result["review_url"] == "http://192.168.224.157:18765/handover/review/a"
    assert calls == [
        {
            "receive_id": "ou_abc",
            "receive_id_type": "open_id",
            "text": (
                "这是一条交接班审核链接测试消息，请在办公电脑的浏览器中打开。\n"
                "楼栋：A楼\n"
                "日期：2026-04-12\n"
                "班次：白班\n"
                "审核链接：http://192.168.224.157:18765/handover/review/a"
            ),
        }
    ]


def test_build_feishu_client_uses_resolved_auth(monkeypatch):
    service = _make_service(
        monkeypatch,
        recipients_by_building={"A楼": [{"open_id": "ou_abc", "note": "测试"}]},
        review_links=[],
    )
    monkeypatch.setattr(
        review_module,
        "require_feishu_auth_settings",
        lambda _cfg, config_path=None: {
            "app_id": "resolved-app",
            "app_secret": "resolved-secret",
            "timeout": 31,
            "request_retry_count": 4,
            "request_retry_interval_sec": 5,
        },
    )

    client = service._build_feishu_client()

    assert client.app_id == "resolved-app"
    assert client.app_secret == "resolved-secret"
    assert client.timeout == 31
