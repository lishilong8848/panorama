import json
from datetime import date
from io import BytesIO
from urllib.error import HTTPError
from urllib.parse import parse_qs, urlparse

import openpyxl

from handover_log_module.core.models import RawRow
from handover_log_module.service.capacity_report_common import (
    _hvdc_search_tokens,
    _tr_replacement_search_tokens,
    build_capacity_cells_with_config,
    build_capacity_template_snapshot,
)
from handover_log_module.service import handover_capacity_report_service as capacity_module
from handover_log_module.service.handover_capacity_report_service import (
    HandoverCapacityReportService,
    _build_fixed_header_cells,
)


def test_build_fixed_header_cells_for_a_building() -> None:
    cells = _build_fixed_header_cells("A楼")

    assert cells["A1"] == "世纪互联南通数据中心A栋FM运维交接班重要事项"
    assert cells["E5"] == "A楼"
    assert cells["G16"] == "A楼"
    assert cells["G17"] == "A楼"
    assert cells["G18"] == "A楼"
    assert cells["S15"] == "A楼"
    assert cells["S16"] == "A楼"
    assert cells["S17"] == "A楼"
    assert cells["S18"] == "A楼"
    assert cells["A20"] == "A楼"
    assert cells["A65"] == "A楼容量一览表"
    assert cells["O55"] == "A楼能耗一览"


def test_build_capacity_cells_include_direct_source_values_and_zone_capacity_formula() -> None:
    rows = [
        RawRow(1, "", "西区 101 1号冷机", "冷机_冷冻水出水温度", "8", 8.0),
        RawRow(2, "", "西区 102 2号冷机", "冷机_冷冻水出水温度", "6", 6.0),
        RawRow(3, "", "西区", "板换冷却水进水温度", "10", 10.0),
        RawRow(4, "", "西区 101 1号冷机", "板交冷冻供水温度", "14", 14.0),
        RawRow(11, "", "西区 101 1号冷机", "板交冷冻回水温度", "9", 9.0),
        RawRow(5, "", "西区", "一次总流量", "100", 100.0),
        RawRow(6, "", "东区", "板换冷却水进水温度", "12", 12.0),
        RawRow(7, "", "东区 104 4号冷机", "板交冷冻供水温度", "18", 18.0),
        RawRow(12, "", "东区 104 4号冷机", "板交冷冻回水温度", "10", 10.0),
        RawRow(8, "", "东区", "一次总流量", "120", 120.0),
        RawRow(9, "", "", "蓄水池总储水量", "88.6", 88.6),
        RawRow(10, "", "", "油量", "55.2", 55.2),
    ]
    context = {
        "building": "A楼",
        "duty_shift": "day",
        "capacity_rows": rows,
        "running_units": {
            "west": [
                {"unit": 1, "mode_text": "预冷"},
                {"unit": 2, "mode_text": "制冷"},
                {"unit": 3, "mode_text": "板换"},
            ],
            "east": [
                {"unit": 4, "mode_text": "板换"},
            ],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert "AC29" not in values
    assert values["U16"] == "55.2"
    assert values["D22"] == "581.5"
    assert values["Q22"] == "1116.48"


def test_capacity_overlay_values_include_ac24_and_track_d8(monkeypatch) -> None:
    service = HandoverCapacityReportService({})
    weather_calls = {"count": 0}
    monkeypatch.setattr(
        service,
        "_query_capacity_water_summary",
        lambda **kwargs: {"month_total": "123.4", "latest_daily_total": "5.6"},
    )
    monkeypatch.setattr(
        service,
        "query_total_electricity_summary",
        lambda **kwargs: {"V57": "7", "Y57": "8"},
    )

    def _fake_weather(**kwargs):
        weather_calls["count"] += 1
        return {"text": "多云", "humidity": "96%"}

    monkeypatch.setattr(
        service,
        "_fetch_weather_payload_for_duty_date",
        _fake_weather,
    )

    values = service._build_capacity_overlay_values(
        building="A楼",
        duty_date="2026-04-11",
        handover_cells={
            "D8": "9.8",
            "H6": "12",
            "F8": "西区30/东区40",
            "B6": "1.23",
            "D6": "2000",
            "F6": "1200",
            "B13": "1000",
            "D13": "800",
        },
        emit_log=lambda _msg: None,
    )

    assert values["AC24"] == "9.8"
    assert values["X2"] == "96%"
    assert values["V57"] == "7"
    assert values["Y57"] == "8"
    assert weather_calls["count"] == 1
    assert "D8" in service.tracked_cells()


def test_capacity_water_summary_queries_once_for_batch(monkeypatch) -> None:
    with capacity_module._WATER_SUMMARY_CACHE_LOCK:
        capacity_module._WATER_SUMMARY_CACHE.clear()
        capacity_module._WATER_SUMMARY_INFLIGHT.clear()

    service = HandoverCapacityReportService({})

    class _FakeClient:
        def __init__(self) -> None:
            self.calls = 0

        def list_fields(self, **_kwargs):
            return []

        def list_records(self, **kwargs):
            self.calls += 1
            assert "filter_formula" in kwargs
            return [
                {"fields": {"执行日期": "2026-04-28", "楼栋": "A楼", "当日耗水总量（修正）": 10}},
                {"fields": {"执行日期": "2026-04-29", "楼栋": "A楼", "当日耗水总量（修正）": 12}},
                {"fields": {"执行日期": "2026-04-29", "楼栋": "B", "当日耗水总量（修正）": 5}},
            ]

    fake_client = _FakeClient()
    monkeypatch.setattr(service, "_new_capacity_water_client", lambda: fake_client)

    a_summary = service._query_capacity_water_summary(
        building="A楼",
        duty_date="2026-04-29",
        emit_log=lambda _msg: None,
    )
    b_summary = service._query_capacity_water_summary(
        building="B楼",
        duty_date="2026-04-29",
        emit_log=lambda _msg: None,
    )

    assert fake_client.calls == 1
    assert a_summary == {"month_total": "22", "latest_daily_total": "12"}
    assert b_summary == {"month_total": "5", "latest_daily_total": "5"}


def test_build_capacity_cooling_summary_uses_running_units_and_source_rows() -> None:
    service = HandoverCapacityReportService({})
    rows = [
        RawRow(1, "", "西区 102 2号冷机", "冷却塔液位", "0.43", 0.43),
        RawRow(2, "", "西区 103 3号冷机", "冷却塔液位", "0.4", 0.4),
        RawRow(3, "", "西区", "1#二次泵频率反馈", "35", 35.0),
        RawRow(4, "", "西区", "2#二次泵频率反馈", "36", 36.0),
        RawRow(5, "", "西区", "蓄冷罐后备温度", "16.9", 16.9),
        RawRow(6, "", "西区", "蓄冷罐液位", "27.32", 27.32),
    ]

    summary = service.build_capacity_cooling_summary(
        capacity_rows=rows,
        running_units={
            "west": [{"unit": 2, "mode_text": "板换"}, {"unit": 3, "mode_text": "板换"}],
            "east": [],
        },
    )

    west_line = summary["lines"]["west"]
    assert "冷冻站A区3套制冷单元2用1备" in west_line
    assert "2#制冷单元板换模式运行正常" in west_line
    assert "2#冷却塔液位0.43m正常" in west_line
    assert "1#2#二次泵运行正常" in west_line
    assert "蓄冷罐后备温度16.9℃正常、液位27.32m正常" in west_line


def test_weather_payload_uses_seniverse_for_today_and_caches(monkeypatch) -> None:
    service = HandoverCapacityReportService(
        {
            "capacity_report": {
                "weather": {
                    "seniverse_public_key": "test-public",
                    "seniverse_private_key": "test-private",
                    "location": "崇川区",
                    "language": "zh-Hans",
                    "unit": "c",
                    "timeout_sec": 8,
                }
            }
        }
    )
    monkeypatch.setattr(service, "_today_local_date", lambda: date(2026, 4, 17))
    calls = {"count": 0}

    class _FakeResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self) -> bytes:
            return json.dumps(
                {
                    "results": [
                        {
                            "daily": [
                                {
                                    "date": "2026-04-17",
                                    "text_day": "多云",
                                    "text_night": "阴",
                                    "humidity": "96",
                                }
                            ]
                        }
                    ]
                }
            ).encode("utf-8")

    def _fake_urlopen(request, timeout=0):  # noqa: ANN001
        calls["count"] += 1
        parsed = urlparse(request.full_url)
        query = parse_qs(parsed.query)
        assert query["uid"] == ["test-public"]
        assert query["location"] == ["31.98:120.89"]
        assert query["language"] == ["zh-Hans"]
        assert query["unit"] == ["c"]
        assert "sig" in query and query["sig"][0]
        assert timeout == 8
        return _FakeResponse()

    monkeypatch.setattr("handover_log_module.service.handover_capacity_report_service.urlopen", _fake_urlopen)

    first = service._fetch_weather_payload_for_duty_date(
        duty_date="2026-04-17",
        emit_log=lambda _msg: None,
    )
    second = service._fetch_weather_payload_for_duty_date(
        duty_date="2026-04-17",
        emit_log=lambda _msg: None,
    )

    assert first == {"text": "多云", "humidity": "96%"}
    assert second == first
    assert calls["count"] == 1


def test_weather_payload_falls_back_to_nantong_when_chongchuan_is_forbidden(monkeypatch) -> None:
    service = HandoverCapacityReportService(
        {
            "capacity_report": {
                "weather": {
                    "seniverse_public_key": "test-public",
                    "seniverse_private_key": "test-private",
                    "location": "崇川区",
                    "fallback_locations": ["南通"],
                    "language": "zh-Hans",
                    "unit": "c",
                    "timeout_sec": 8,
                }
            }
        }
    )
    monkeypatch.setattr(service, "_today_local_date", lambda: date(2026, 4, 17))
    attempted_locations = []

    class _FakeResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self) -> bytes:
            return json.dumps(
                {
                    "results": [
                        {
                            "daily": [
                                {
                                    "date": "2026-04-17",
                                    "text_day": "阴",
                                    "text_night": "阴",
                                    "humidity": "72",
                                }
                            ]
                        }
                    ]
                }
            ).encode("utf-8")

    def _fake_urlopen(request, timeout=0):  # noqa: ANN001
        parsed = urlparse(request.full_url)
        query = parse_qs(parsed.query)
        location = query["location"][0]
        attempted_locations.append(location)
        assert timeout == 8
        if location == "31.98:120.89":
            raise HTTPError(
                request.full_url,
                403,
                "Forbidden",
                hdrs=None,
                fp=BytesIO(b"{\"status\":\"You don't have access to data of this city.\",\"status_code\":\"AP010006\"}"),
            )
        return _FakeResponse()

    monkeypatch.setattr("handover_log_module.service.handover_capacity_report_service.urlopen", _fake_urlopen)

    payload = service._fetch_weather_payload_for_duty_date(
        duty_date="2026-04-17",
        emit_log=lambda _msg: None,
    )

    assert attempted_locations == ["31.98:120.89", "南通"]
    assert payload == {"text": "阴", "humidity": "72%"}


def test_weather_payload_uses_legacy_html_for_historical_dates(monkeypatch) -> None:
    service = HandoverCapacityReportService({})
    monkeypatch.setattr(service, "_today_local_date", lambda: date(2026, 4, 17))
    legacy_calls = {"count": 0}

    def _fake_legacy_fetch(**kwargs):  # noqa: ANN003
        legacy_calls["count"] += 1
        assert kwargs["duty_date"] == "2026-04-16"
        return {"text": "小雨", "humidity": "88%"}

    def _unexpected_seniverse_fetch(**kwargs):  # noqa: ANN003
        raise AssertionError("historical duty_date should not call Seniverse")

    monkeypatch.setattr(service, "_legacy_fetch_weather_payload_for_duty_date", _fake_legacy_fetch)
    monkeypatch.setattr(service, "_fetch_seniverse_weather_payload_for_duty_date", _unexpected_seniverse_fetch)

    payload = service._fetch_weather_payload_for_duty_date(
        duty_date="2026-04-16",
        emit_log=lambda _msg: None,
    )

    assert payload == {"text": "小雨", "humidity": "88%"}
    assert legacy_calls["count"] == 1


def test_ab_building_current_oil_values_are_scaled() -> None:
    service = HandoverCapacityReportService({})
    rows = [
        RawRow(1, "", "燃油自控系统", "1#油罐容积", "10", 10.0),
        RawRow(2, "", "燃油自控系统", "2#油罐容积", "20", 20.0),
    ]

    values = service._extract_current_oil_display_values(
        building="A楼",
        rows=rows,
        emit_log=lambda _msg: None,
    )

    assert values["first"] == "11904.76"
    assert values["second"] == "23809.52"


def test_d_building_current_oil_values_use_specific_tank_volume_aliases() -> None:
    service = HandoverCapacityReportService({})
    rows = [
        RawRow(1, "", "燃油自控系统", "1#油罐容积", "10", 10.0),
        RawRow(2, "", "燃油自控系统", "2#油罐容积", "20", 20.0),
        RawRow(3, "", "燃油自控系统", "1#油罐体积", "31", 31.0),
        RawRow(4, "", "燃油自控系统", "2#油罐体积", "41", 41.0),
    ]

    values = service._extract_current_oil_display_values(
        building="D楼",
        rows=rows,
        emit_log=lambda _msg: None,
    )

    assert values["first"] == "31"
    assert values["second"] == "41"


def test_previous_capacity_values_read_from_previous_capacity_file(tmp_path) -> None:
    service = HandoverCapacityReportService({"capacity_report": {"template": {"sheet_name": "Sheet"}}})
    previous_file = tmp_path / "previous_capacity.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet["U13"] = "123"
    sheet["X13"] = "456"
    workbook.save(previous_file)
    workbook.close()

    class _FakeReviewService:
        def get_latest_session_for_context(self, **kwargs):
            assert kwargs["building"] == "A楼"
            assert kwargs["duty_date"] == "2026-04-10"
            assert kwargs["duty_shift"] == "night"
            return {"capacity_output_file": str(previous_file), "revision": 2, "updated_at": "2026-04-11 09:00:00"}

    service._review_session_service = _FakeReviewService()

    values, warning = service._load_previous_capacity_display_oil_values(
        building="A楼",
        duty_date="2026-04-11",
        duty_shift="day",
        current_display_values={"first": "11", "second": "22"},
        emit_log=lambda _msg: None,
    )

    assert warning == ""
    assert values == {"first": "123", "second": "456"}


def test_previous_capacity_values_fallback_to_current_when_previous_missing() -> None:
    service = HandoverCapacityReportService({})

    class _FakeReviewService:
        def get_latest_session_for_context(self, **kwargs):
            return None

    service._review_session_service = _FakeReviewService()

    values, warning = service._load_previous_capacity_display_oil_values(
        building="B楼",
        duty_date="2026-04-11",
        duty_shift="day",
        current_display_values={"first": "33", "second": "44"},
        emit_log=lambda _msg: None,
    )

    assert warning == "上一班容量文件未命中，已回退当前班次值"
    assert values == {"first": "33", "second": "44"}


def test_capacity_template_snapshot_collects_other_building_entries_from_template_anchors() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row, value in [
        (67, "-245-TR101"),
        (72, "-245-TR102"),
        (77, "-218-TR201"),
        (82, "-218-TR202"),
        (87, "-345-TR101"),
        (94, "-345-TR102"),
        (101, "-317-TR201"),
        (108, "-317-TR202"),
        (115, "-445-TR101"),
        (120, "-445-TR102"),
        (125, "-417-TR201"),
        (130, "-417-TR202"),
        (135, "-144-TR101"),
        (140, "-144-TR102"),
        (145, "-120-TR201"),
        (150, "-120-TR202"),
    ]:
        sheet[f"B{row}"] = value
        sheet.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)

    for row, value in [
        (67, "-245-UPS-101"),
        (77, "-218-UPS-101"),
        (87, "-345-UPS-101"),
        (101, "-317-UPS-101"),
        (115, "-445-UPS-101"),
        (125, "-417-UPS-101"),
        (135, "-144-UPS-101"),
        (140, "-144-UPS-201"),
        (145, "-120-UPS-101"),
        (150, "-120-UPS-201"),
    ]:
        sheet[f"I{row}"] = value
        sheet.merge_cells(start_row=row, start_column=9, end_row=row + 1, end_column=9)

    sheet["O67"] = "-245-HVDC-111"
    sheet["O134"] = "-417-HVDC-251"

    snapshot = build_capacity_template_snapshot(sheet, "A楼")

    assert [entry["row"] for entry in snapshot["tr_entries"]] == [
        67, 72, 77, 82, 87, 94, 101, 108, 115, 120, 125, 130, 135, 140, 145, 150,
    ]
    assert [entry["row"] for entry in snapshot["ups_entries"]] == [
        67, 77, 87, 101, 115, 125, 135, 140, 145, 150,
    ]
    assert [entry["row"] for entry in snapshot["hvdc_entries"]] == [67, 134]


def test_capacity_template_snapshot_collects_e_building_entries_from_template_anchors() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row, value in [
        (67, "-245-TR-101"),
        (77, "-245-TR-102"),
        (87, "-218-TR-201"),
        (97, "-218-TR-202"),
        (107, "-345-TR-101"),
        (117, "-345-TR-102"),
        (127, "-317-TR-201"),
        (137, "-317-TR-202"),
        (147, "-445-TR-101"),
        (157, "-445-TR-102"),
        (167, "-417-TR-201"),
        (177, "-417-TR-202"),
        (187, "-144-TR-101"),
        (192, "-144-TR-102"),
        (197, "-120-TR-201"),
        (202, "-120-TR-202"),
    ]:
        sheet[f"B{row}"] = value
        sheet.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=2)

    for row, value in [
        (67, "-245-UPS-101"),
        (87, "-218-UPS-101"),
        (107, "-345-UPS-101"),
        (127, "-317-UPS-101"),
        (147, "-445-UPS-101"),
        (167, "-417-UPS-101"),
        (187, "-144-UPS-101"),
        (192, "-144-UPS-201"),
        (197, "-120-UPS-101"),
        (202, "-120-UPS-201"),
    ]:
        sheet[f"I{row}"] = value
        sheet.merge_cells(start_row=row, start_column=9, end_row=row + 2, end_column=9)

    sheet["O67"] = "-245-HVDC-102"
    sheet["O186"] = "-417-HVDC-292"

    snapshot = build_capacity_template_snapshot(sheet, "E楼")

    assert [entry["row"] for entry in snapshot["tr_entries"]] == [
        67, 77, 87, 97, 107, 117, 127, 137, 147, 157, 167, 177, 187, 192, 197, 202,
    ]
    assert [entry["row"] for entry in snapshot["ups_entries"]] == [
        67, 87, 107, 127, 147, 167, 187, 192, 197, 202,
    ]
    assert [entry["row"] for entry in snapshot["hvdc_entries"]] == [67, 186]


def test_tr_replacement_search_tokens_support_unhyphenated_suffixes() -> None:
    assert _tr_replacement_search_tokens("A-245-TR102") == ["A-245-TRB201", "A-245-TRB-201"]
    assert _tr_replacement_search_tokens("A-245-TR202") == ["A-245-TRB201", "A-245-TRB-201"]
    assert _tr_replacement_search_tokens("A-245-TR201") == ["A-245-TRB101", "A-245-TRB-101"]
    assert _tr_replacement_search_tokens("A-245-TR-102") == ["A-245-TRB201", "A-245-TRB-201"]


def test_hvdc_search_tokens_for_other_buildings_prioritize_last_digit_two() -> None:
    assert _hvdc_search_tokens("A-245-HVDC-111", template_family="other_buildings") == [
        "A-245-HVDC-112",
        "A-245-HVDC-111",
    ]
    assert _hvdc_search_tokens("E-245-HVDC-102", template_family="e_building") == [
        "E-245-HVDC-102",
    ]


def test_build_capacity_cells_fill_e_column_from_trb_201_source_row() -> None:
    context = {
        "capacity_rows": [
            RawRow(1, "A-245-TRB-201", "", "", "45.6", 45.6),
        ],
        "running_units": {},
        "template_snapshot": {
            "tr_entries": [
                {"row": 67, "identifier": "A-245-TR102", "search_tokens": ["A-245-TR102"]},
            ],
            "ups_entries": [],
            "hvdc_entries": [],
            "rpp_entries": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["E67"] == "45.6"


def test_build_capacity_cells_fill_e_column_from_trb_source_c_column_for_d_building() -> None:
    context = {
        "capacity_rows": [
            RawRow(1, "", "D-218-TRB-101", "", "101.1", 101.1),
            RawRow(2, "", "D-218-TRB-201", "", "201.2", 201.2),
        ],
        "running_units": {},
        "template_snapshot": {
            "tr_entries": [
                {"row": 77, "identifier": "D-218-TR201", "search_tokens": ["D-218-TR201"]},
                {"row": 82, "identifier": "D-218-TR202", "search_tokens": ["D-218-TR202"]},
            ],
            "ups_entries": [],
            "hvdc_entries": [],
            "rpp_entries": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["E77"] == "101.1"
    assert values["E82"] == "201.2"


def test_build_capacity_cells_prioritize_exact_oil_tonnage_alias_for_u16() -> None:
    context = {
        "building": "D楼",
        "capacity_rows": [
            RawRow(1, "", "燃油自控系统", "油量", "legacy", None),
            RawRow(2, "", "燃油自控系统", "燃油总吨数", "66.6", 66.6),
        ],
        "running_units": {},
    }

    values = build_capacity_cells_with_config(context)

    assert values["U16"] == "66.6"


def test_d_building_capacity_cells_keep_formula_cells_and_fill_required_oil_and_tr_values() -> None:
    context = {
        "building": "D楼",
        "duty_shift": "day",
        "oil_current": {"first": "31", "second": "41"},
        "capacity_rows": [
            RawRow(1, "", "", "蓄水池总储水量", "88.6", 88.6),
            RawRow(2, "", "燃油自控系统", "燃油总吨数", "66.6", 66.6),
            RawRow(3, "", "D-218-TRB-101", "", "101.1", 101.1),
            RawRow(4, "", "D-218-TRB-201", "", "201.2", 201.2),
        ],
        "running_units": {},
        "template_snapshot": {
            "tr_entries": [
                {"row": 77, "identifier": "D-218-TR201", "search_tokens": ["D-218-TR201"]},
                {"row": 82, "identifier": "D-218-TR202", "search_tokens": ["D-218-TR202"]},
            ],
            "ups_entries": [],
            "hvdc_entries": [],
            "rpp_entries": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["U13"] == "31"
    assert values["X13"] == "41"
    assert values["U16"] == "66.6"
    assert values["E77"] == "101.1"
    assert values["E82"] == "201.2"
    assert "AC29" not in values


def test_build_capacity_cells_hvdc_for_other_buildings_prefers_last_digit_two_source_row() -> None:
    context = {
        "capacity_rows": [
            RawRow(1, "", "A-245-HVDC-112", "电池组电压", "91", 91.0),
            RawRow(2, "", "A-245-HVDC-112", "直流电压", "92", 92.0),
            RawRow(3, "", "A-245-HVDC-112", "直流总功率", "93", 93.0),
            RawRow(4, "", "A-245-HVDC-111", "电池组电压", "11", 11.0),
        ],
        "running_units": {},
        "template_snapshot": {
            "hvdc_entries": [
                {
                    "row": 67,
                    "identifier": "A-245-HVDC-111",
                    "search_tokens": ["A-245-HVDC-112", "A-245-HVDC-111"],
                },
            ],
            "tr_entries": [],
            "ups_entries": [],
            "rpp_entries": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["P67"] == "91"
    assert values["Q67"] == "92"
    assert values["R67"] == "93"


def test_capacity_template_selection_uses_dual_defaults_for_legacy_source_path() -> None:
    service = HandoverCapacityReportService(
        {"capacity_report": {"template": {"source_path": "交接班容量报表空模板.xlsx"}}}
    )

    a_selection = service.resolve_template_selection(building="A楼")
    e_selection = service.resolve_template_selection(building="E楼")

    assert a_selection["path"].name == "其他楼交接班容量报表空模板.xlsx"
    assert a_selection["template_family"] == "other_buildings"
    assert e_selection["path"].name == "E楼交接班容量报表空模板.xlsx"
    assert e_selection["template_family"] == "e_building"


def test_capacity_template_selection_respects_explicit_override(tmp_path) -> None:
    custom_template = tmp_path / "custom_capacity_template.xlsx"
    custom_template.write_bytes(b"demo")
    service = HandoverCapacityReportService(
        {"capacity_report": {"template": {"source_path": str(custom_template)}}}
    )

    selection = service.resolve_template_selection(building="A楼")

    assert selection["path"] == custom_template
    assert selection["template_family"] == "other_buildings"


def test_d_building_cooling_tower_out_temp_alias_fills_f30() -> None:
    rows = [
        RawRow(1, "", "西区 101 1号冷机", "冷却塔出水温度", "27.5", 27.5),
    ]
    context = {
        "building": "D楼",
        "duty_shift": "day",
        "capacity_rows": rows,
        "running_units": {
            "west": [
                {"unit": 1, "mode_text": "制冷"},
            ],
            "east": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["F30"] == "27.5"


def test_e_building_plate_mode_skips_chiller_fields_but_keeps_fan_and_pump_values() -> None:
    rows = [
        RawRow(1, "", "西区 101 1号冷机", "冷机_电流百分比", "45", 45.0),
        RawRow(2, "", "西区 101 1号冷机", "冷机_冷冻水出水温度", "7.2", 7.2),
        RawRow(3, "", "西区 101 1号冷机", "冷机_冷却水进水温度", "27.1", 27.1),
        RawRow(4, "", "西区 101 1号冷机", "冷凝器压力", "100", 100.0),
        RawRow(5, "", "西区 101 1号冷机", "蒸发器小温差", "2.1", 2.1),
        RawRow(6, "", "西区 101 1号冷机", "冷凝器小温差", "1.8", 1.8),
        RawRow(7, "", "西区 101 1号冷机", "冷塔1#风机频率反馈", "33", 33.0),
        RawRow(8, "", "西区 101 1号冷机", "冷塔2#风机频率反馈", "34", 34.0),
        RawRow(9, "", "西区 101 1号冷机", "冷却泵频率反馈", "35", 35.0),
        RawRow(10, "", "西区 101 1号冷机", "一次冷冻泵频率反馈", "36", 36.0),
        RawRow(11, "", "西区", "板换冷却水进水温度", "10", 10.0),
    ]
    context = {
        "building": "E楼",
        "duty_shift": "day",
        "capacity_rows": rows,
        "running_units": {
            "west": [{"unit": 1, "mode_text": "板换"}],
            "east": [],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert "D25" not in values
    assert "D27" not in values
    assert "D28" not in values
    assert "D29" not in values
    assert "D30" not in values
    assert "D31" not in values
    assert values["F27"] == "33"
    assert values["G27"] == "34"
    assert values["I26"] == "35"
    assert values["J26"] == "36"


def test_hvdc_missing_r_fills_zero_without_touching_u() -> None:
    context = {
        "capacity_rows": [],
        "running_units": {},
        "template_snapshot": {
            "hvdc_entries": [
                {"row": 67, "identifier": "E-317-HVDC-252", "search_tokens": ["E-317-HVDC-252"]},
            ],
        },
    }

    values = build_capacity_cells_with_config(context)

    assert values["R67"] == "0"
    assert "U67" not in values


def test_aircon_matrix_mapping_for_other_buildings_uses_other_template_targets() -> None:
    rows = [
        RawRow(1, "A楼/四层/空调区1 A-412", "A-412-CRAHB-A_电量仪", "总_有功功率", "2.23", 2.23),
        RawRow(2, "A楼/四层/空调区2 A-411", "A-411-CRAHB-A_电量仪", "总_有功功率", "0.8", 0.8),
        RawRow(3, "A楼/四层/空调区3 A-441", "A-441-CRAHB-A_电量仪", "总_有功功率", "0", 0.0),
        RawRow(4, "A楼/四层/空调区4 A-440", "A-440-CRAHB-A_电量仪", "总_有功功率", "1.36", 1.36),
    ]
    context = {
        "capacity_rows": rows,
        "running_units": {},
        "template_snapshot": {"building_code": "A", "template_family": "other_buildings"},
    }

    values = build_capacity_cells_with_config(context)

    assert values["AE128"] == "2.23"
    assert values["AE133"] == "0.8"
    assert values["AE118"] == "0"
    assert values["AE123"] == "1.36"


def test_aircon_matrix_mapping_fills_last_two_targets() -> None:
    rows = [
        RawRow(1, "南通阿里保税A区E楼/E楼/四层/空调区1 E-412", "E-412-CRAHB-A_电量仪", "总_有功功率", "2.23", 2.23),
        RawRow(2, "南通阿里保税A区E楼/E楼/四层/空调区2 E-411", "E-411-CRAHB-A_电量仪", "总_有功功率", "0.8", 0.8),
        RawRow(3, "南通阿里保税A区E楼/E楼/四层/空调区3 E-441", "E-441-CRAHB-A_电量仪", "总_有功功率", "0", 0.0),
        RawRow(4, "南通阿里保税A区E楼/E楼/四层/空调区4 E-440", "E-440-CRAHB-A_电量仪", "总_有功功率", "1.36", 1.36),
    ]
    context = {
        "capacity_rows": rows,
        "running_units": {},
        "template_snapshot": {"building_code": "E", "template_family": "e_building"},
    }

    values = build_capacity_cells_with_config(context)

    assert values["AE172"] == "2.23"
    assert values["AE182"] == "0.8"
    assert values["AE152"] == "0"
    assert values["AE162"] == "1.36"


def test_aircon_matrix_mapping_skips_blank_first_hit_and_keeps_later_value_for_e_building() -> None:
    rows = [
        RawRow(1, "南通阿里保税A区E楼/E楼/三层/空调区3 E-341", "E-341-CRAHB-A_电量仪", "总_有功功率", "", None),
        RawRow(2, "南通阿里保税A区E楼/E楼/三层/空调区3 E-341", "E-341-CRAHB-A_电量仪", "总_有功功率", "1.12", 1.12),
        RawRow(3, "南通阿里保税A区E楼/E楼/三层/空调区2 E-311", "E-311-CRAHB-A_电量仪", "总_有功功率", "", None),
        RawRow(4, "南通阿里保税A区E楼/E楼/三层/空调区2 E-311", "E-311-CRAHB-A_电量仪", "总_有功功率", "1.42", 1.42),
        RawRow(5, "南通阿里保税A区E楼/E楼/四层/空调区3 E-441", "E-441-CRAHB-A_电量仪", "总_有功功率", "", None),
        RawRow(6, "南通阿里保税A区E楼/E楼/四层/空调区3 E-441", "E-441-CRAHB-A_电量仪", "总_有功功率", "1.52", 1.52),
    ]
    context = {
        "capacity_rows": rows,
        "running_units": {},
        "template_snapshot": {"building_code": "E", "template_family": "e_building"},
    }

    values = build_capacity_cells_with_config(context)

    assert values["AE112"] == "1.12"
    assert values["AE142"] == "1.42"
    assert values["AE152"] == "1.52"
