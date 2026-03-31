from dataclasses import dataclass

from openpyxl import Workbook

from app.modules.report_pipeline.service.calc_source_runtime import (
    apply_building_source_overrides,
    extract_building,
    extract_month,
    extract_row_sources,
    locate_stat_columns,
)


@dataclass
class _Row:
    row_index: int
    type_name: str
    category_name: str
    item_name: str
    max_value: float | None
    min_value: float | None
    avg_value: float | None
    data_values: list[float]


HEADER_ROW = 2
DATA_START_ROW = 4
TYPE_COL = 2
CATEGORY_COL = 3
ITEM_COL = 4
DATA_START_COL = 5


def _norm_text(v):
    return str(v or "").strip().lower()


def _to_float(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return float(v)
    except Exception:
        return None


def test_extract_building_from_filename():
    assert extract_building(r"D:\\QLDownload\\C楼_20260308_120000.xlsx") == "C楼"


def test_locate_and_extract_sources():
    wb = Workbook()
    ws = wb.active

    # header
    ws.cell(HEADER_ROW, DATA_START_COL, "2026-03-01 00:00:00")
    ws.cell(HEADER_ROW, DATA_START_COL + 1, "2026-03-01 01:00:00")
    ws.cell(HEADER_ROW, DATA_START_COL + 2, "最大值")
    ws.cell(HEADER_ROW, DATA_START_COL + 3, "最小值")
    ws.cell(HEADER_ROW, DATA_START_COL + 4, "平均值")

    # data row
    ws.cell(DATA_START_ROW, TYPE_COL, "总览指标")
    ws.cell(DATA_START_ROW, CATEGORY_COL, "分析指标")
    ws.cell(DATA_START_ROW, ITEM_COL, "IT负载率")
    ws.cell(DATA_START_ROW, DATA_START_COL, 50)
    ws.cell(DATA_START_ROW, DATA_START_COL + 1, 70)

    stat_cols = locate_stat_columns(
        ws,
        header_row=HEADER_ROW,
        data_start_col=DATA_START_COL,
        max_label=_norm_text("最大值"),
        min_label=_norm_text("最小值"),
        avg_label=_norm_text("平均值"),
        norm_text=_norm_text,
    )

    assert stat_cols["max"] == DATA_START_COL + 2
    assert stat_cols["min"] == DATA_START_COL + 3
    assert stat_cols["avg"] == DATA_START_COL + 4

    source_map, default_type, default_category = extract_row_sources(
        ws,
        stat_cols,
        data_start_col=DATA_START_COL,
        data_start_row=DATA_START_ROW,
        type_col=TYPE_COL,
        category_col=CATEGORY_COL,
        item_col=ITEM_COL,
        to_float=_to_float,
        canonical_metric_name=lambda x: str(x),
        row_source_factory=_Row,
    )

    assert default_type == "总览指标"
    assert default_category == "分析指标"
    row = source_map["IT负载率"]
    assert row.max_value == 70
    assert row.min_value == 50
    assert row.avg_value == 60


def test_extract_month_from_header_datetime_text():
    wb = Workbook()
    ws = wb.active
    ws.cell(HEADER_ROW, DATA_START_COL, "2026-02-01 00:00:00")

    month = extract_month(
        ws,
        max_col=DATA_START_COL + 2,
        header_row=HEADER_ROW,
        data_start_col=DATA_START_COL,
    )
    assert month == "2026-02"


def test_apply_building_source_overrides_for_a_building():
    marker = object()
    source_map = {"总_电度": marker}
    out = apply_building_source_overrides(
        building="A楼",
        source_map=source_map,
        canonical_metric_name=lambda x: str(x),
    )
    assert out["光伏发电量"] is marker


def test_apply_building_source_overrides_does_not_apply_for_other_building():
    marker = object()
    source_map = {"总_电度": marker}
    out = apply_building_source_overrides(
        building="B楼",
        source_map=source_map,
        canonical_metric_name=lambda x: str(x),
    )
    assert "光伏发电量" not in out
