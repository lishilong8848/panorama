from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict

import openpyxl

from app.modules.sheet_import.repository.workbook_repository import (
    build_raw_header_name_by_column,
    extract_rows_with_row_index,
    extract_sheet_images_by_anchor,
    safe_file_token,
)


@dataclass
class _RowPayload:
    row_index: int
    fields: Dict[str, Any]


@dataclass
class _ImagePlacement:
    row_index: int
    column_index: int
    image_index: int
    file_name: str
    mime_type: str
    content: bytes


def test_build_raw_header_name_by_column_and_rows() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="名称")
    ws.cell(row=1, column=3, value="值")
    ws.cell(row=2, column=2, value="A")
    ws.cell(row=2, column=3, value=1)
    ws.cell(row=3, column=2, value=None)
    ws.cell(row=3, column=3, value=None)

    headers = build_raw_header_name_by_column(ws, 1)
    assert headers[2] == "名称"
    assert headers[3] == "值"

    rows = extract_rows_with_row_index(
        ws=ws,
        header_row=1,
        row_payload_factory=lambda row_index, fields: _RowPayload(row_index=row_index, fields=fields),
    )
    assert len(rows) == 1
    assert rows[0].row_index == 2
    assert rows[0].fields["名称"] == "A"
    wb.close()


def test_safe_file_token() -> None:
    assert safe_file_token("A楼/Sheet 1") == "A_Sheet_1"


def test_extract_sheet_images_by_anchor_empty_when_no_images() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    placements = extract_sheet_images_by_anchor(
        ws=ws,
        header_row=1,
        image_placement_factory=lambda **kwargs: _ImagePlacement(**kwargs),
    )
    assert placements == []
    wb.close()
